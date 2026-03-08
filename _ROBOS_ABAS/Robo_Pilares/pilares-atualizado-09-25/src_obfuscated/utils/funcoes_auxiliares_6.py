
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import win32com.client
import win32gui
import pythoncom
import time
import shutil

# IMPORTAR HELPER FROZEN GLOBAL - garante que paths estão configurados
try:
    from _frozen_helper import ensure_paths
    ensure_paths()
except ImportError:
    try:
        from src._frozen_helper import ensure_paths
        ensure_paths()
    except ImportError:
        # Fallback manual
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
            if script_dir not in sys.path:
                sys.path.insert(0, script_dir)
            src_dir = os.path.join(script_dir, 'src')
            if os.path.exists(src_dir) and src_dir not in sys.path:
                sys.path.insert(0, src_dir)

# Importar funcoes_auxiliares_3 com múltiplos fallbacks
try:
    from .funcoes_auxiliares_3 import AC_Module
except ImportError:
    try:
        from funcoes_auxiliares_3 import AC_Module
    except ImportError:
        try:
            from src.utils.funcoes_auxiliares_3 import AC_Module
        except ImportError:
            try:
                from utils.funcoes_auxiliares_3 import AC_Module
            except ImportError:
                # Fallback se AC_Module não estiver disponível
                AC_Module = None
                print("⚠️ AC_Module não disponível - algumas funcionalidades podem não funcionar")

from natsort import natsorted  # Importar para ordenação natural dos arquivos

class PainelControleDesenhoPMixin:
    """
    Mixin para adicionar funcionalidades do Painel de Controle de Desenho
    """
    
    def __init__(self):
        self.ac = None
        self.processing_lock = False  # Flag para evitar execução simultânea
    
    def _get_project_root(self):
        """Retorna o diretório raiz do projeto"""
        # Usar o mesmo path resolver que está funcionando no conector
        try:
            from ..utils.robust_path_resolver import robust_path_resolver
            return robust_path_resolver.get_project_root()
        except ImportError:
            try:
                from utils.robust_path_resolver import robust_path_resolver
                return robust_path_resolver.get_project_root()
            except ImportError:
                try:
                    from src.utils.robust_path_resolver import robust_path_resolver
                    return robust_path_resolver.get_project_root()
                except ImportError:
                    # Fallback final - calcular manualmente
                    if getattr(sys, 'frozen', False):
                        return os.path.dirname(sys.executable)
                    else:
                        return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    
    def _get_item_selecionado_safe(self):
        """
        Obtém o item selecionado de forma segura, verificando se o conector existe e não é None.
        
        Returns:
            tuple: (numero, dados) ou (None, None) se houver erro
        """
        # Verificar se o conector está disponível e não é None
        if not hasattr(self, 'conector_painel') or self.conector_painel is None:
            error_msg = "Conector não disponível. Certifique-se de que a interface principal está aberta e funcionando corretamente."
            print(f">>> ERRO: {error_msg}")
            try:
                messagebox.showerror("Erro", error_msg)
            except Exception:
                pass
            return None, None
        
        try:
            # CORREÇÃO: Chamar o método do conector, não a si mesmo
            numero, dados = self.conector_painel.get_item_selecionado()
            return numero, dados
        except AttributeError as e:
            error_msg = f"Erro ao obter item selecionado: {str(e)}. O conector pode não estar inicializado corretamente."
            print(f">>> ERRO: {error_msg}")
            try:
                messagebox.showerror("Erro", error_msg)
            except Exception:
                pass
            return None, None
        except Exception as e:
            error_msg = f"Erro inesperado ao obter item selecionado: {str(e)}"
            print(f">>> ERRO: {error_msg}")
            try:
                messagebox.showerror("Erro", error_msg)
            except Exception:
                pass
            return None, None
    
    def _verificar_se_item_especial(self, dados_item):
        """
        Verifica se um item é especial baseado nos dados do pilar
        
        Args:
            dados_item: Dicionário com dados do item/pilar
        
        Returns:
            bool: True se for especial, False se for comum
        """
        try:
            pilar_especial_ativo = False
            tipo_pilar_especial = ""
            
            if isinstance(dados_item, dict):
                # Verificar diretamente no dicionário
                pilar_especial_ativo = dados_item.get('ativar_pilar_especial', False)
                tipo_pilar_especial = dados_item.get('tipo_pilar_especial', '')
                
                # Verificar em pilar_especial (aninhado)
                pilar_especial_data = dados_item.get('pilar_especial', {})
                if isinstance(pilar_especial_data, dict):
                    if not pilar_especial_ativo:
                        pilar_especial_ativo = pilar_especial_data.get('ativar_pilar_especial', False)
                    if not tipo_pilar_especial:
                        tipo_pilar_especial = pilar_especial_data.get('tipo_pilar_especial', '')
            
            # É especial se: ativo E tipo válido (L, T, U)
            return pilar_especial_ativo and tipo_pilar_especial in ['L', 'T', 'U']
            
        except Exception as e:
            print(f"Erro ao verificar se item é especial: {e}")
            return False
    
    def _obter_parafusos_especiais_para_transferencia(self, tipo_parafuso):
        """
        Obtém os parafusos especiais para transferência ao AutoCAD
        
        Parâmetros:
        - tipo_parafuso: 'a' para parafusos A, 'e' para parafusos E
        
        Lógica:
        - Pilar 1: usa parafusos A (todos os campos)
        - Pilar 2: usa parafusos E (desconsidera primeiro campo, usa a partir do segundo)
        """
        try:
            parafusos = []
            
            if tipo_parafuso == 'a':
                # Parafusos A: todos os campos (para Pilar 1)
                for i in range(1, 10):  # par_a_1 a par_a_9
                    campo_nome = f"par_a_{i}"
                    if hasattr(self, campo_nome):
                        campo = getattr(self, campo_nome)
                        valor = campo.get()
                        if valor and valor.strip():
                            try:
                                parafusos.append(float(valor))
                            except ValueError:
                                parafusos.append(0)
                        else:
                            parafusos.append(0)
                    else:
                        parafusos.append(0)
                        
            elif tipo_parafuso == 'e':
                # Parafusos E: desconsidera primeiro campo, usa a partir do segundo (para Pilar 2)
                for i in range(2, 10):  # par_e_2 a par_e_9 (pula par_e_1)
                    campo_nome = f"par_e_{i}"
                    if hasattr(self, campo_nome):
                        campo = getattr(self, campo_nome)
                        valor = campo.get()
                        if valor and valor.strip():
                            try:
                                parafusos.append(float(valor))
                            except ValueError:
                                parafusos.append(0)
                        else:
                            parafusos.append(0)
                    else:
                        parafusos.append(0)
                        
            print(f"[TRANSFERENCIA] Parafusos {tipo_parafuso.upper()} coletados: {parafusos}")
            return parafusos
            
        except Exception as e:
            print(f"[ERRO] Falha ao coletar parafusos {tipo_parafuso.upper()}: {str(e)}")
            return [0] * 9 if tipo_parafuso == 'a' else [0] * 8
    
    def criar_painel_controle_desenho(self, parent_frame):
        """
        Cria o painel de controle de desenho com 4 categorias de botões e scroll
        """
        # Frame principal com scroll - mais estreito
        scroll_frame = ttk.Frame(parent_frame)
        scroll_frame.pack(fill="y", expand=False, padx=5, pady=5)
        scroll_frame.configure(width=120)  # Largura fixa menor
        
        # Canvas para scroll - largura fixa com estilo moderno
        canvas = tk.Canvas(scroll_frame, highlightthickness=0, width=100, 
                          bg="#F5F5F5", relief="flat", bd=0)
        canvas.pack(side="left", fill="y", expand=False)
        
        # Scrollbar sempre visível
        scrollbar = ttk.Scrollbar(scroll_frame, orient="vertical", command=canvas.yview)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Frame interno que será scrollable
        painel_frame = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=painel_frame, anchor="nw")
        
        # Configurar o grid para expansão
        painel_frame.grid_columnconfigure(0, weight=1)
        
        # Configurar estilos modernos com cores por categoria
        style = ttk.Style()
        style.configure("Mini.TButton", padding=(1, 1))  # Padding mínimo em Y
        
        # Estilos modernos para LabelFrames
        style.configure("Cima.TLabelframe", 
                       background="#E3F2FD", 
                       borderwidth=2, 
                       relief="solid",
                       foreground="#1976D2")
        style.configure("Cima.TLabelframe.Label", 
                       background="#E3F2FD", 
                       foreground="#1976D2",
                       font=("Arial", 9, "bold"))
        
        style.configure("Abcd.TLabelframe", 
                       background="#E8F5E8", 
                       borderwidth=2, 
                       relief="solid",
                       foreground="#388E3C")
        style.configure("Abcd.TLabelframe.Label", 
                       background="#E8F5E8", 
                       foreground="#388E3C",
                       font=("Arial", 9, "bold"))
        
        style.configure("Grades.TLabelframe", 
                       background="#FFF3E0", 
                       borderwidth=2, 
                       relief="solid",
                       foreground="#F57C00")
        style.configure("Grades.TLabelframe.Label", 
                       background="#FFF3E0", 
                       foreground="#F57C00",
                       font=("Arial", 9, "bold"))
        
        style.configure("Utilitarios.TLabelframe", 
                       background="#F3E5F5", 
                       borderwidth=2, 
                       relief="solid",
                       foreground="#7B1FA2")
        style.configure("Utilitarios.TLabelframe.Label", 
                       background="#F3E5F5", 
                       foreground="#7B1FA2",
                       font=("Arial", 9, "bold"))
        
        # Criar as 4 seções com canvas como referência
        self._criar_secao_visao_cima(painel_frame, 0, canvas)
        self._criar_secao_visao_abcd(painel_frame, 1, canvas) 
        self._criar_secao_grades(painel_frame, 2, canvas)
        self._criar_secao_utilitarios(painel_frame, 3, canvas)
        
        # Configurar scroll
        def update_scrollregion(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        painel_frame.bind("<Configure>", update_scrollregion)
        
        # Scroll com mouse
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind("<MouseWheel>", on_mousewheel)
        painel_frame.bind("<MouseWheel>", on_mousewheel)
        
        # Armazenar referências para que não sejam perdidas
        self.painel_canvas = canvas
        self.painel_scrollbar = scrollbar
        
        return scroll_frame
    
    def _verificar_modo_offline(self):
        """Verifica se está em modo offline - compatível com frozen (Nuitka/PyInstaller)"""
        try:
            # Tentar múltiplos imports para compatibilidade com frozen
            obter_gerenciador_creditos = None
            try:
                from core.credit_system import obter_gerenciador_creditos
            except ImportError:
                try:
                    from src.core.credit_system import obter_gerenciador_creditos
                except ImportError:
                    # Último fallback: tentar importar o módulo completo
                    try:
                        import core.credit_system as credit_system
                        obter_gerenciador_creditos = credit_system.obter_gerenciador_creditos
                    except ImportError:
                        try:
                            import src.core.credit_system as credit_system
                            obter_gerenciador_creditos = credit_system.obter_gerenciador_creditos
                        except ImportError:
                            return False
            
            if obter_gerenciador_creditos:
                credit_manager = obter_gerenciador_creditos()
                if credit_manager and hasattr(credit_manager, 'modo_offline'):
                    return credit_manager.modo_offline
        except Exception:
            # Em caso de qualquer erro, retornar False (não está em modo offline)
            pass
        return False
    
    def _criar_botao_compacto(self, parent, texto, comando, posicao, canvas=None):
        """Cria um botão compacto com grid layout e quebra de texto"""
        # Determinar cor baseada no estilo do parent
        parent_style = str(parent.cget('style')) if hasattr(parent, 'cget') else ""
        
        # Cores modernas por categoria
        cores = {
            "Cima.TLabelframe": {"bg": "#1976D2", "fg": "white", "activebg": "#1565C0", "activefg": "white"},
            "Abcd.TLabelframe": {"bg": "#388E3C", "fg": "white", "activebg": "#2E7D32", "activefg": "white"},
            "Grades.TLabelframe": {"bg": "#F57C00", "fg": "white", "activebg": "#EF6C00", "activefg": "white"},
            "Utilitarios.TLabelframe": {"bg": "#7B1FA2", "fg": "white", "activebg": "#6A1B9A", "activefg": "white"}
        }
        
        cor_atual = cores.get(parent_style, {"bg": "#607D8B", "fg": "white", "activebg": "#455A64", "activefg": "white"})
        
        # Verificar se é um botão de desenho que deve ser bloqueado em modo offline
        texto_lower = texto.lower()
        botoes_desenho = ["desenhar", "pilar selecionado", "pavimento selecionado", "item completo", "pavimento completo"]
        eh_botao_desenho = any(palavra in texto_lower for palavra in botoes_desenho)
        
        # Criar wrapper para verificar modo offline antes de executar
        if eh_botao_desenho:
            def comando_wrapper():
                if self._verificar_modo_offline():
                    messagebox.showwarning(
                        "Modo Offline", 
                        "Você está em modo offline.\n\n"
                        "A geração de scripts de desenho está desabilitada.\n\n"
                        "Faça login para habilitar as funcionalidades de desenho."
                    )
                    return
                comando()
            comando_final = comando_wrapper
        else:
            comando_final = comando
        
        # Botão tk.Button moderno com cores
        btn = tk.Button(parent, text=texto, command=comando_final, width=10, 
                       wraplength=80, justify='center',
                       bg=cor_atual["bg"], fg=cor_atual["fg"],
                       activebackground=cor_atual["activebg"], activeforeground=cor_atual["activefg"],
                       font=("Arial", 8, "bold"),
                       relief="flat", bd=0,
                       cursor="hand2")  # Cursor de mão
        
        # Desabilitar botão se estiver em modo offline e for botão de desenho
        if eh_botao_desenho and self._verificar_modo_offline():
            btn.config(state=tk.DISABLED, bg="#9E9E9E", fg="#616161")
        
        btn.grid(row=posicao, column=0, padx=2, pady=2, sticky="ew")  # Padding aumentado
        
        # Efeito hover moderno (apenas se não estiver desabilitado)
        def on_enter(event):
            if btn.cget("state") != tk.DISABLED:
                btn.configure(bg=cor_atual["activebg"])
        def on_leave(event):
            if btn.cget("state") != tk.DISABLED:
                btn.configure(bg=cor_atual["bg"])
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        # Bind scroll no botão individual
        if canvas:
            def on_botao_mousewheel(event, canv=canvas):
                canv.yview_scroll(int(-1*(event.delta/120)), "units")
                return "break"  # Impede propagação do evento
            
            btn.bind("<MouseWheel>", on_botao_mousewheel)
    
    def _criar_secao_visao_cima(self, parent, row, canvas):
        """Cria a seção Visão de CIMA"""
        frame = ttk.LabelFrame(parent, text="Visão CIMA", style="Cima.TLabelframe")
        frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        frame.grid_columnconfigure(0, weight=1)
        
        # Botões da seção
        botoes = [
            # ("Desenhar Item", self._desenhar_item_cima),  # Botões ocultados para caso deseje retornar, funções mantidas
            # ("Desenhar Pavimento", self._desenhar_pavimento_cima),  # Botões ocultados para caso deseje retornar, funções mantidas
            ("Desenhar Visão/Corte Pilar Selecionado", self._dcad_item_cima),
            ("Desenhar Visão/Corte de todo o Pavimento Selecionado", self._dcad_pavimento_cima),
            ("Configurações", self._config_desenho_cima),
            ("Configurar Ordenador", self._config_ordenamento_cima),
            # ("Ordenador CIMA", self._ordenador_cima)  # Botões ocultados para caso deseje retornar, funções mantidas
        ]
        
        for i, (texto, comando) in enumerate(botoes):
            self._criar_botao_compacto(frame, texto, comando, i, canvas)
    
    def _criar_secao_visao_abcd(self, parent, row, canvas):
        """Cria a seção Visão lados ABCD"""
        frame = ttk.LabelFrame(parent, text="Visão ABCD", style="Abcd.TLabelframe")
        frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        frame.grid_columnconfigure(0, weight=1)
        
        # Botões da seção
        botoes = [
            # ("Desenhar Item", self._desenhar_item_abcd),  # Botões ocultados para caso deseje retornar, funções mantidas
            # ("Desenhar Pavimento", self._desenhar_pavimento_abcd),  # Botões ocultados para caso deseje retornar, funções mantidas
            ("Desenhar Paineis ABCD do Pilar Selecionado", self._dcad_item_abcd),
            ("Desenhar Paineis ABCD de todo o Pavimento Selecionado", self._dcad_pavimento_abcd),
            ("Configurações", self._config_desenho_abcd),
            ("Configurar Ordenador", self._config_ordenamento_abcd),
            # ("Ordenador ABCD", self._ordenador_abcd)  # Botões ocultados para caso deseje retornar, funções mantidas
        ]
        
        for i, (texto, comando) in enumerate(botoes):
            self._criar_botao_compacto(frame, texto, comando, i, canvas)
    
    def _criar_secao_grades(self, parent, row, canvas):
        """Cria a seção Grades"""
        frame = ttk.LabelFrame(parent, text="Grades", style="Grades.TLabelframe")
        frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        frame.grid_columnconfigure(0, weight=1)
        
        # Botões da seção
        botoes = [
            # ("Desenhar Item", self._desenhar_item_grades),  # Botões ocultados para caso deseje retornar, funções mantidas
            # ("Desenhar Pavimento", self._desenhar_pavimento_grades),  # Botões ocultados para caso deseje retornar, funções mantidas
            ("Desenhar Grades do Pilar Selecionado", self._dcad_item_grades),
            ("Desenhar Todas Grades do Pavimento Selecionado", self._dcad_pavimento_grades),
            ("Configurações", self._config_desenho_grades),
            ("Configurar Ordenador", self._config_ordenamento_grades),
            # ("Ordenador GRADES", self._ordenador_grades)  # Botões ocultados para caso deseje retornar, funções mantidas
        ]
        
        for i, (texto, comando) in enumerate(botoes):
            self._criar_botao_compacto(frame, texto, comando, i, canvas)
    
    def _criar_secao_utilitarios(self, parent, row, canvas):
        """Cria a seção de utilitários"""
        frame = ttk.LabelFrame(parent, text="Utilitários", style="Utilitarios.TLabelframe")
        frame.grid(row=row, column=0, sticky="ew", padx=5, pady=5)
        frame.grid_columnconfigure(0, weight=1)
        
        # Botões da seção
        botoes = [
            ("Desenhar 1 Item Completo", self._desenhar_1_item_completo),
            ("Desenhar 1 Pavimento Completo", self._desenhar_1_pavimento_completo),
            ("Criar Comando LISP", self._criar_comando_lisp),
        ]
        
        for i, (texto, comando) in enumerate(botoes):
            self._criar_botao_compacto(frame, texto, comando, i, canvas)
    
    # Métodos dos botões - vinculados às funções do conector
    
    def _desenhar_item_cima(self):
        """Desenha item individual para visão de cima"""
        if hasattr(self, 'conector_painel'):
            # Usar apenas a execução direta (sem Excel temporário)
            self._dcad_item_cima()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _desenhar_pavimento_cima(self):
        """Desenha pavimento completo para visão de cima - APENAS COMBINA SCRIPTS EXISTENTES"""
        # Proteção contra execução simultânea
        if self.processing_lock:
            print(">>> Processamento já em andamento. Aguarde...")
            return
        
        self.processing_lock = True
        
        try:
            if not hasattr(self, 'conector_painel'):
                messagebox.showwarning("Aviso", "Conector não disponível")
                return
                
            # Obter informações do pavimento selecionado (usar item selecionado como referência)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                messagebox.showerror("Erro", "Nenhum item selecionado. Selecione um item na lista para usar como referência do pavimento.")
                return
            
            pavimento = dados.get("pavimento", "")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se a pasta do pavimento já existe
            robo_dir = self._get_project_root()
            
            # Construir nome da pasta baseado no pavimento
            nome_pasta = f"{pavimento.replace(' ', '_')}_CIMA"
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            
            print(f">>> Verificando pasta do pavimento: {pasta_pavimento}")
            
            # Verificar se a pasta do pavimento existe
            if not os.path.exists(pasta_pavimento):
                print(f">>> Pasta do pavimento não encontrada. Use o botão 'D.Cad Pavimento' para gerar os scripts primeiro.")
                messagebox.showwarning("Aviso", "Pasta do pavimento não encontrada.\n\nUse o botão 'D.Cad Pavimento' para gerar os scripts primeiro.")
                return
            
            # Verificar se os scripts individuais existem
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            scripts_existem = True
            
            for numero, dados in itens:
                nome = dados.get("nome", numero)
                # Verificar primeiro script combinado, depois normal
                script_combinado_path = os.path.join(pasta_pavimento, f"{nome}_COMBINADO_CIMA.scr")
                script_normal_path = os.path.join(pasta_pavimento, f"{nome}_CIMA.scr")
                
                if os.path.exists(script_combinado_path):
                    script_path = script_combinado_path
                else:
                    script_path = script_normal_path
                    
                if not os.path.exists(script_path):
                    scripts_existem = False
                    break
            
            if not scripts_existem:
                print(f">>> Scripts individuais não encontrados. Use o botão 'D.Cad Pavimento' para gerar os scripts primeiro.")
                messagebox.showwarning("Aviso", "Scripts individuais não encontrados.\n\nUse o botão 'D.Cad Pavimento' para gerar os scripts primeiro.")
                return
            
            print(f">>> Scripts individuais encontrados. Executando apenas o combinador...")
            # Se os scripts existem, executar apenas o combinador
            self.conector_painel.combinar_codigos_cima()
            
        finally:
            # Sempre liberar o lock ao final
            self.processing_lock = False
    
    def _config_desenho_cima(self):
        """Configurações de desenho para visão de cima"""
        print(f"\n>>> EXECUTANDO CONFIGURAÇÕES CIMA")
        print(f"[DEBUG CONFIG] Abrindo configurações CIMA (layers/blocks)")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_robo_cima()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _config_ordenamento_cima(self):
        """Configurações de ordenamento para visão de cima"""
        print(f"\n>>> EXECUTANDO CONFIGURAR ORDENADOR CIMA")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_configuracao_cima()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _ordenador_cima(self):
        """Executa diretamente o ordenador de coordenadas CIMA"""
        if hasattr(self, 'conector_painel'):
            try:
                import sys
                import os
                
                # Adicionar o diretório dos ordenadores ao path - usar nova estrutura
                try:
                    from ..interfaces import interface_ordenador_cima
                    ordenador_dir = os.path.join(self._get_project_root(), "src", "interfaces")
                except ImportError:
                    ordenador_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "interfaces")
                sys.path.append(ordenador_dir)
                
                # Importar e executar a interface do ordenador CIMA
                try:
                    from interface_ordenador_cima import InterfaceOrdenadorCima
                except ImportError:
                    # Fallback para subprocess se não conseguir importar
                    import subprocess
                    ordenador_path = os.path.join(ordenador_dir, "interface_ordenador_cima.py")
                    if os.path.exists(ordenador_path):
                        subprocess.Popen([sys.executable, ordenador_path])
                        return
                    else:
                        raise ImportError("Não foi possível encontrar interface_ordenador_cima")
                
                # Criar e iniciar a interface
                app = InterfaceOrdenadorCima()
                app.iniciar()
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir ordenador CIMA:\n{str(e)}")
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _desenhar_item_abcd(self):
        """Desenha item individual para visão lados ABCD"""
        if hasattr(self, 'conector_painel'):
            self.conector_painel.executar_abcd_excel()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _desenhar_pavimento_abcd(self):
        """Desenha pavimento completo para visão lados ABCD"""
        if hasattr(self, 'conector_painel'):
            self.conector_painel.executar_todos_abcd_excel()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _config_desenho_abcd(self):
        """Configurações de desenho para visão lados ABCD"""
        print(f"\n>>> EXECUTANDO CONFIGURAÇÕES ABCD")
        print(f"[DEBUG CONFIG] Abrindo configurações ABCD (layers/blocks)")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_robo_abcd()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _config_ordenamento_abcd(self):
        """Configurações de ordenamento para visão lados ABCD"""
        print(f"\n>>> EXECUTANDO CONFIGURAR ORDENADOR ABCD")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_configuracao_abcd()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _ordenador_abcd(self):
        """Executa diretamente o ordenador de coordenadas ABCD"""
        if hasattr(self, 'conector_painel'):
            try:
                import sys
                import os
                
                # Adicionar o diretório dos ordenadores ao path
                ordenador_dir = os.path.join(self._get_project_root(), "src", "interfaces")
                sys.path.append(ordenador_dir)
                
                # Importar e executar a interface do ordenador ABCD
                try:
                    from interface_ordenador_abcd import InterfaceOrdenador
                except ImportError:
                    # Fallback para subprocess se não conseguir importar
                    import subprocess
                    ordenador_path = os.path.join(ordenador_dir, "interface_ordenador_abcd.py")
                    if os.path.exists(ordenador_path):
                        subprocess.Popen([sys.executable, ordenador_path])
                        return
                    else:
                        raise ImportError("Não foi possível encontrar interface_ordenador_abcd")
                
                # Criar e iniciar a interface
                app = InterfaceOrdenador()
                app.iniciar()
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir ordenador ABCD:\n{str(e)}")
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _desenhar_item_grades(self):
        """Desenha item individual para grades"""
        if hasattr(self, 'conector_painel'):
            self.conector_painel.executar_grade_excel()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _desenhar_pavimento_grades(self):
        """Desenha pavimento completo para grades"""
        if hasattr(self, 'conector_painel'):
            self.conector_painel.executar_todos_grade_excel()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _config_desenho_grades(self):
        """Configurações de desenho para grades"""
        print(f"\n>>> EXECUTANDO CONFIGURAÇÕES GRADES")
        print(f"[DEBUG CONFIG] Abrindo configurações GRADES (layers/blocks)")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_robo_grades()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _config_ordenamento_grades(self):
        """Configurações de ordenamento para grades"""
        print(f"\n>>> EXECUTANDO CONFIGURAR ORDENADOR GRADES")
        if hasattr(self, 'conector_painel'):
            self.conector_painel.abrir_configuracao_grades()
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")
    
    def _ordenador_grades(self):
        """Executa diretamente o ordenador de coordenadas GRADES"""
        if hasattr(self, 'conector_painel'):
            try:
                import sys
                import os
                
                # Adicionar o diretório dos ordenadores ao path
                ordenador_dir = os.path.join(self._get_project_root(), "src", "interfaces")
                sys.path.append(ordenador_dir)
                
                # Importar e executar a interface do ordenador GRADES
                try:
                    from interface_ordenador_grades import InterfaceOrdenadorGrade
                except ImportError:
                    # Fallback para subprocess se não conseguir importar
                    import subprocess
                    ordenador_path = os.path.join(ordenador_dir, "interface_ordenador_grades.py")
                    if os.path.exists(ordenador_path):
                        subprocess.Popen([sys.executable, ordenador_path])
                        return
                    else:
                        raise ImportError("Não foi possível encontrar interface_ordenador_grades")
                
                # Criar e iniciar a interface
                app = InterfaceOrdenadorGrade()
                app.iniciar()
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao abrir ordenador GRADES:\n{str(e)}")
        else:
            messagebox.showwarning("Aviso", "Conector não disponível")

    def _dcad_item_cima(self):
        """Executa o desenho de um item no CAD para visão CIMA"""
        try:
            print(f"\n>>> EXECUTANDO D.CAD ITEM CIMA")
            print(f"[DEBUG] _dcad_item_cima CHAMADA - self: {self}")
            print(f"[DEBUG] _dcad_item_cima CHAMADA - conector_painel: {hasattr(self, 'conector_painel')}")
            
            # Obter informações do item selecionado (com verificação segura)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            # Verificar se o sistema de créditos está disponível
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="CIMA"
                )
                
                # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                    return
                
                print(_get_obf_str("credit"))
                
                # Atualizar interface imediatamente
                self.atualizar_creditos_interface()

            # Verificar se a pasta do pavimento já existe e excluí-la antes de executar
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_CIMA")
            
            if os.path.exists(pasta_pavimento):
                print(f">>> Pasta do pavimento já existe: {pasta_pavimento}")
                print(f">>> Excluindo pasta antes de executar...")
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                    print(f">>> ✅ Pasta excluída com sucesso")
                except Exception as e:
                    print(f">>> ⚠️ Erro ao excluir pasta: {str(e)}")
                    # Continuar mesmo se não conseguir excluir
            else:
                print(f">>> Pasta do pavimento não existe ainda: {pasta_pavimento}")

            # Usar a nova função que não depende do Excel
            try:
                # Importar a nova função com múltiplos fallbacks para ambiente frozen
                import sys
                
                # Tentar múltiplos caminhos
                interfaces_dir = None
                
                # 1. Tentar caminho relativo
                try:
                    current_file = os.path.abspath(__file__)
                    utils_dir = os.path.dirname(current_file)
                    interfaces_dir = os.path.join(utils_dir, '..', 'interfaces')
                    interfaces_dir = os.path.abspath(interfaces_dir)
                    if not os.path.exists(interfaces_dir):
                        interfaces_dir = None
                except Exception:
                    pass
                
                # 2. Se frozen, tentar caminho do executável
                if not interfaces_dir and getattr(sys, 'frozen', False):
                    try:
                        exe_dir = os.path.dirname(sys.executable)
                        interfaces_dir = os.path.join(exe_dir, 'src', 'interfaces')
                        if not os.path.exists(interfaces_dir):
                            # Tentar diretório do executável diretamente
                            interfaces_dir = os.path.join(exe_dir, 'interfaces')
                    except Exception:
                        pass
                
                # 3. Tentar importação direta (módulo já no path)
                try:
                    from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                except ImportError:
                    try:
                        from interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                    except ImportError:
                        try:
                            from src.interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                        except ImportError:
                            # Último fallback: adicionar ao path e tentar novamente
                            if interfaces_dir and interfaces_dir not in sys.path:
                                sys.path.insert(0, interfaces_dir)
                            from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                
                # TRANSFERÊNCIA DIRETA: Passar a interface principal para o robô
                print(f"\n>>> 🎯 TRANSFERÊNCIA DIRETA - SEM MAPEAMENTO")
                print(f">>> Enviando interface principal diretamente para o robô")
                
                # Passar a interface principal diretamente
                # Preparar dados do pilar especial com parafusos especiais dentro
                pilar_especial_data = dados.get('pilar_especial', {}).copy() if dados.get('pilar_especial', {}) else {}
                # Adicionar parafusos especiais dentro de pilar_especial
                pilar_especial_data['parafusos_especiais'] = {
                    'parafusos_a': self._obter_parafusos_especiais_para_transferencia('a'),
                    'parafusos_e': self._obter_parafusos_especiais_para_transferencia('e')
                }
                
                dados_pilar = {
                    'interface_principal': self,  # Interface principal completa (self é a própria interface)
                    'pavimento': pavimento,
                    'nome': nome,
                    'comprimento': dados.get('comprimento', ''),
                    'largura': dados.get('largura', ''),
                    'parafuso_p1_p2': int(float(dados.get('parafusos', {}).get('par_1_2', 0) or 0)),
                    'parafuso_p2_p3': int(float(dados.get('parafusos', {}).get('par_2_3', 0) or 0)),
                    'parafuso_p3_p4': int(float(dados.get('parafusos', {}).get('par_3_4', 0) or 0)),
                    'parafuso_p4_p5': int(float(dados.get('parafusos', {}).get('par_4_5', 0) or 0)),
                    'parafuso_p5_p6': int(float(dados.get('parafusos', {}).get('par_5_6', 0) or 0)),
                    'parafuso_p6_p7': int(float(dados.get('parafusos', {}).get('par_6_7', 0) or 0)),
                    'parafuso_p7_p8': int(float(dados.get('parafusos', {}).get('par_7_8', 0) or 0)),
                    'parafuso_p8_p9': int(float(dados.get('parafusos', {}).get('par_8_9', 0) or 0)),
                    # CORREÇÃO: Adicionar dados das grades grupo 1 (Grade A) - CORRIGIDO
                    'grades_grupo1': dados.get('grades', {}),  # Os dados da Grade A estão em 'grades'
                    'detalhes_grades': dados.get('detalhes_grades', {}),
                    # CORREÇÃO: Adicionar dados das grades grupo 2 (Grade B)
                    'grades_grupo2': dados.get('grades_grupo2', {}),
                    'detalhes_grades_grupo2': dados.get('detalhes_grades_grupo2', {}),
                    # CORREÇÃO: Passar o objeto pilar_especial completo (contém as grades A, B, E, F)
                    'pilar_especial': pilar_especial_data,
                    # ADIÇÃO: Dados do pilar especial (mantidos para compatibilidade)
                    'ativar_pilar_especial': pilar_especial_data.get('ativar_pilar_especial', False),
                    'tipo_pilar_especial': pilar_especial_data.get('tipo_pilar_especial', ''),
                    'comp_1': pilar_especial_data.get('comprimentos', {}).get('comp_1', ''),
                    'comp_2': pilar_especial_data.get('comprimentos', {}).get('comp_2', ''),
                    'comp_3': pilar_especial_data.get('comprimentos', {}).get('comp_3', ''),
                    'larg_1': pilar_especial_data.get('larguras', {}).get('larg_1', ''),
                    'larg_2': pilar_especial_data.get('larguras', {}).get('larg_2', ''),
                    'larg_3': pilar_especial_data.get('larguras', {}).get('larg_3', ''),
                    # NOVA FUNCIONALIDADE: Parafusos especiais para transferência (também no nível raiz para compatibilidade)
                    'parafusos_especiais': pilar_especial_data.get('parafusos_especiais', {})
                }
                
                print(f">>> Executando nova função sem Excel...")
                print(f">>> Dados preparados: {dados_pilar}")
                
                # DEBUG: Print dados completos
                print(f"\n>>> 🎯 DADOS COMPLETOS DO ITEM:")
                print(f">>> dados completos: {dados}")
                print(f">>> grades (Grupo 1): {dados.get('grades', 'NÃO ENCONTRADO')}")
                print(f">>> grades_grupo2: {dados.get('grades_grupo2', 'NÃO ENCONTRADO')}")
                print(f">>> detalhes_grades: {dados.get('detalhes_grades', 'NÃO ENCONTRADO')}")
                print(f">>> detalhes_grades_grupo2: {dados.get('detalhes_grades_grupo2', 'NÃO ENCONTRADO')}")
                
                # DEBUG: Print grades being sent
                print(f"\n>>> 🎯 GRADES ENVIADAS:")
                print(f">>> Grade Grupo 1 (Grade A): {dados_pilar.get('grades_grupo1', 'NÃO ENCONTRADO')}")
                print(f">>> Grade Grupo 2 (Grade B): {dados_pilar.get('grades_grupo2', 'NÃO ENCONTRADO')}")
                print(f">>> Detalhes Grade Grupo 1: {dados_pilar.get('detalhes_grades', 'NÃO ENCONTRADO')}")
                print(f">>> Detalhes Grade Grupo 2: {dados_pilar.get('detalhes_grades_grupo2', 'NÃO ENCONTRADO')}")
                print(f">>> 🎯 FIM GRADES ENVIADAS\n")
                
                # DEBUG: Print dados do pilar especial
                print(f"\n>>> 🎯 DADOS PILAR ESPECIAL:")
                print(f">>> ativar_pilar_especial: {dados_pilar.get('ativar_pilar_especial', 'NÃO ENCONTRADO')}")
                print(f">>> tipo_pilar_especial: {dados_pilar.get('tipo_pilar_especial', 'NÃO ENCONTRADO')}")
                print(f">>> comp_1: {dados_pilar.get('comp_1', 'NÃO ENCONTRADO')}")
                print(f">>> comp_2: {dados_pilar.get('comp_2', 'NÃO ENCONTRADO')}")
                print(f">>> comp_3: {dados_pilar.get('comp_3', 'NÃO ENCONTRADO')}")
                print(f">>> larg_1: {dados_pilar.get('larg_1', 'NÃO ENCONTRADO')}")
                print(f">>> larg_2: {dados_pilar.get('larg_2', 'NÃO ENCONTRADO')}")
                print(f">>> larg_3: {dados_pilar.get('larg_3', 'NÃO ENCONTRADO')}")
                
                # DEBUG: Print objeto pilar_especial completo
                pilar_especial_completo = dados_pilar.get('pilar_especial', {})
                print(f"\n>>> 🎯 OBJETO PILAR_ESPECIAL COMPLETO:")
                print(f">>> pilar_especial: {pilar_especial_completo}")
                if pilar_especial_completo:
                    grades_especiais = pilar_especial_completo.get('grades', {})
                    detalhes_especiais = pilar_especial_completo.get('detalhes_grades_especiais', {})
                    print(f">>> grades (contém A, B, E, F): {grades_especiais}")
                    print(f">>> detalhes_grades_especiais: {detalhes_especiais}")
                    print(f">>> Número de grades especiais: {len(grades_especiais)}")
                    print(f">>> Número de detalhes especiais: {len(detalhes_especiais)}")
                else:
                    print(f">>> ❌ OBJETO PILAR_ESPECIAL VAZIO!")
                print(f">>> 🎯 FIM DADOS PILAR ESPECIAL\n")
                
                # ===== NOVA FUNCIONALIDADE: PREENCHIMENTO AUTOMÁTICO DE GRADES PARA PILARES ESPECIAIS TIPO "L" =====
                # NOTA: A lógica de detecção do pilar_atual foi movida para a função preencher_campos_diretamente_e_gerar_scripts
                # onde o pilar_atual está disponível durante a geração dos scripts múltiplos
                
                # Chamar a nova função
                preencher_campos_diretamente_e_gerar_scripts(dados_pilar)
                
                print(f">>> ✅ Nova função executada com sucesso")
                
            except Exception as e:
                print(f">>> ERRO na nova função: {str(e)}")
                import traceback
                traceback.print_exc()
                
                # Não usar fallback - apenas a execução direta
                print(f">>> ❌ Erro na execução direta - não há fallback")
                
                # Mostrar erro de forma mais detalhada para debugging
                error_msg = f"Erro na geração do script:\n{str(e)}\n\nVerifique os logs para mais detalhes."
                try:
                    messagebox.showerror("Erro", error_msg)
                except Exception as msg_error:
                    print(f"[AVISO] Não foi possível mostrar messagebox: {msg_error}")
                
                # NÃO fechar o programa - apenas retornar
                return

            # Obter o caminho do script gerado
            robo_dir = self._get_project_root()
            
            # Obter informações do item selecionado
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Construir o caminho do script (usar nome normalizado do pavimento)
            import unicodedata
            def _normalizar_nome_pasta_local(texto):
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(ch for ch in texto_normalizado if unicodedata.category(ch) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                return texto_limpo.strip('_')

            pav_norm = _normalizar_nome_pasta_local(pavimento)
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pav_norm}_CIMA")
            
            # Primeiro, tentar encontrar o script combinado (para pilares especiais)
            script_combinado_path = os.path.join(pasta_pavimento, f"{nome}_COMBINADO_CIMA.scr")
            script_normal_path = os.path.join(pasta_pavimento, f"{nome}_CIMA.scr")
            
            # Verificar qual script usar
            if os.path.exists(script_combinado_path):
                script_path = script_combinado_path
                print(f">>> Usando script combinado (pilar especial)")
            else:
                script_path = script_normal_path
                print(f">>> Usando script normal")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Caminho do script: {script_path}")

            if not os.path.exists(script_path):
                print(f">>> ERRO: Script não encontrado!")
                print(f">>> Verificando se a pasta existe: {os.path.exists(pasta_pavimento)}")
                
                if os.path.exists(pasta_pavimento):
                    print(f">>> Conteúdo da pasta {pasta_pavimento}:")
                    try:
                        for item in os.listdir(pasta_pavimento):
                            print(f">>>   - {item}")
                    except Exception as e:
                        print(f">>> Erro ao listar pasta: {e}")
                
                messagebox.showerror("Erro", f"Script não encontrado em:\n{script_path}\n\nVerifique se o item foi processado corretamente.")
                return

            print(f">>> Script encontrado com sucesso!")

            # Ler e filtrar o script
            script_content = self._ler_e_filtrar_script(script_path)
            if not script_content:
                print(f">>> ERRO: Falha ao ler/filtrar script")
                return

            print(f">>> Script processado com sucesso, executando no AutoCAD...")

            # Registrar item desenhado nesta sessão para evitar duplicidades no pavimento
            try:
                base = os.path.basename(script_path)
                if base.endswith('_COMBINADO_CIMA.scr'):
                    base_core = base[:-len('_COMBINADO_CIMA.scr')]
                elif base.endswith('_CIMA.scr'):
                    base_core = base[:-len('_CIMA.scr')]
                else:
                    base_core = os.path.splitext(base)[0]
                if not hasattr(self, '_itens_desenhados_sessao'):
                    self._itens_desenhados_sessao = set()
                self._itens_desenhados_sessao.add(base_core)
                print(f">>> [SESSAO] Registrado como desenhado: {base_core}")
            except Exception as _:
                pass

            # ========================================================
            # 🎯 ATUALIZAR SCRIPT_PAZ.scr E ADICIONAR COMANDOS SCRIPT
            # ========================================================
            
            # Verificar se é pilar especial (tem script combinado)
            pilar_especial = os.path.exists(script_combinado_path)
            
            if pilar_especial:
                # PILAR ESPECIAL: usar apenas o script combinado (já contém ambos CIMA_1 e CIMA_2)
                print(f">>> [PAZ] Pilar especial detectado - usando script combinado")
                
                # Para pilares especiais, o script combinado já contém tudo
                # Os scripts individuais foram removidos após a combinação
                if os.path.exists(script_combinado_path):
                    # Atualizar PAZ com o script combinado
                    self._atualizar_script_paz(script_combinado_path)
                    
                    print(f">>> [PAZ] ✅ Script PAZ atualizado com script combinado (contém CIMA_1 e CIMA_2)")
                else:
                    print(f">>> [PAZ] ⚠️ Script combinado não encontrado, pulando atualização do PAZ")
            else:
                # PILAR COMUM: 1 script só
                print(f">>> [PAZ] Pilar comum detectado - processando 1 script")
                
                # Atualizar PAZ com o único script
                self._atualizar_script_paz(script_path)
                
                print(f">>> [PAZ] ✅ Script PAZ atualizado com script único")

            # Perguntar se deseja executar no AutoCAD
            opcao = self._perguntar_execucao_autocad("desenho visão cima")
            
            if opcao is None:  # Cancelado
                return
            
            if opcao == 1:
                # Opção 1: Desenhar no CAD diretamente
                self._executar_no_autocad(script_content, pavimento, "CIMA")
                print(f"✅ Desenho visão cima executado com sucesso!")
            elif opcao == 2:
                # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                print(f"✅ Scripts gerados (não executados no AutoCAD)")
            elif opcao == 3:
                # Opção 3: Gerar DFX (em desenvolvimento)
                messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                return

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD item CIMA: {str(e)}")
            import traceback
            error_traceback = traceback.format_exc()
            print(f">>> TRACEBACK COMPLETO:\n{error_traceback}")
            
            # Mostrar erro sem fechar o programa
            error_msg = f"Erro ao executar D.CAD item CIMA:\n{str(e)}\n\nVerifique os logs para mais detalhes."
            try:
                messagebox.showerror("Erro", error_msg)
            except Exception as msg_error:
                print(f"[AVISO] Não foi possível mostrar messagebox: {msg_error}")
            
            # NÃO levantar exceção novamente - apenas registrar o erro
            # Isso evita que o programa feche inesperadamente

    def _dcad_pavimento_cima(self):
        """Executa o desenho do pavimento no CAD para visão CIMA"""
        # Proteção contra execução simultânea mais robusta
        if hasattr(self, 'dcad_lock') and self.dcad_lock:
            print(">>> Processamento D.CAD já em andamento. Aguarde...")
            return
        
        # Usar lock específico para D.CAD
        self.dcad_lock = True
        self.processing_lock = True
        
        try:
            print(f"\n>>> EXECUTANDO D.CAD PAVIMENTO CIMA")
            # Usar datetime ao invés de time.time() para evitar conflitos
            from datetime import datetime
            print(f">>> Timestamp: {datetime.now().timestamp()}")
            print(f">>> PID: {os.getpid()}")

            # Reiniciar controle de itens desenhados na sessão a cada execução completa de pavimento
            self._itens_desenhados_sessao = set()
            print(f">>> [SESSAO] Controle de itens desenhados reiniciado para este pavimento")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                messagebox.showwarning("Aviso", "Conector não disponível")
                return

            # Obter informações do pavimento selecionado (usar item selecionado como referência)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                messagebox.showerror("Erro", "Nenhum item selecionado. Selecione um item na lista para usar como referência do pavimento.")
                return
            
            # VALIDAÇÃO: Obter pavimento do item E do filtro para garantir consistência
            pavimento_item = dados.get("pavimento", "")
            pavimento_filtro = self.conector_painel.get_pavimento_selecionado()
            
            # Usar o filtro se disponível e válido, senão usar o do item
            if pavimento_filtro and pavimento_filtro != "Todos":
                pavimento = pavimento_filtro
                print(f">>> ✅ Pavimento obtido do FILTRO: '{pavimento}'")
            elif pavimento_item and pavimento_item != "Todos":
                pavimento = pavimento_item
                print(f">>> ✅ Pavimento obtido do ITEM selecionado: '{pavimento}'")
            else:
                pavimento = pavimento_item if pavimento_item else ""
                print(f">>> ⚠️ Pavimento obtido do item (fallback): '{pavimento}'")
            
            print(f">>> [VALIDAÇÃO] Pavimento final que será usado: '{pavimento}'")
            print(f">>> [VALIDAÇÃO] Pavimento do item: '{pavimento_item}'")
            print(f">>> [VALIDAÇÃO] Pavimento do filtro: '{pavimento_filtro}'")
            
            # Verificar se o pavimento não é "Todos" ou vazio
            if not pavimento or pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Pavimento não funciona quando 'Todos' está selecionado.")
                return

            # Obter todos os itens do pavimento primeiro (para contar créditos)
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            print(f">>> Total de itens encontrados para pavimento '{pavimento}': {len(itens)}")
            
            # LOG DETALHADO: Listar todos os itens encontrados
            print(f">>> [DEBUG] Itens encontrados:")
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                pavimento_item = dados_item.get("pavimento", "N/A")
                print(f">>>   [{idx}] Número: {numero_item}, Nome: {nome_item}, Pavimento: {pavimento_item}")
            
            if not itens:
                messagebox.showerror("Erro", f"Nenhum item encontrado para o pavimento {pavimento}")
                return
            
            # VALIDAÇÃO ADICIONAL: Verificar se todos os itens são do pavimento correto e remover duplicatas
            itens_validos = []
            itens_numeros_ja_vistos = set()  # Para evitar duplicatas por número
            itens_nomes_ja_vistos = set()    # Para evitar duplicatas por nome também
            
            for numero_item, dados_item in itens:
                pavimento_item = dados_item.get("pavimento", "")
                nome_item = dados_item.get("nome", numero_item)
                
                # Verificar se pertence ao pavimento correto
                if pavimento_item != pavimento:
                    print(f">>> ⚠️ ATENÇÃO: Item {numero_item} pertence ao pavimento '{pavimento_item}', não '{pavimento}' - REMOVIDO")
                    continue
                
                # Verificar duplicata por número E por nome (alguns podem ter números diferentes mas mesmo nome)
                numero_str = str(numero_item)
                nome_str = str(nome_item)
                chave_duplicata = (numero_str, nome_str)
                
                # Verificar se já foi adicionado (por número ou por nome similar)
                ja_adicionado = (
                    numero_str in itens_numeros_ja_vistos or
                    nome_str in itens_nomes_ja_vistos or
                    (numero_str in [str(n) for n in itens_numeros_ja_vistos] if itens_numeros_ja_vistos else False)
                )
                
                if ja_adicionado:
                    print(f">>> ⚠️ ATENÇÃO: Item {numero_item} (nome: {nome_str}) está DUPLICADO na lista - REMOVIDO")
                    continue
                
                itens_validos.append((numero_item, dados_item))
                itens_numeros_ja_vistos.add(numero_str)
                itens_nomes_ja_vistos.add(nome_str)
            
            if len(itens_validos) != len(itens):
                print(f">>> ⚠️ CORREÇÃO: {len(itens)} itens retornados, mas apenas {len(itens_validos)} são válidos e únicos")
                print(f">>> Usando apenas os {len(itens_validos)} itens válidos sem duplicatas")
            else:
                print(f">>> ✅ Todos os {len(itens_validos)} itens são válidos e únicos")
            
            itens = itens_validos

            # ============================================================
            # VERIFICAÇÃO E DÉBITO DE CRÉDITOS
            # ============================================================
            quantidade_itens = len(itens)
            print(f">>> [DEBUG] Quantidade final de itens após validação: {quantidade_itens}")
            credit_manager = getattr(self, 'credit_manager', None)
            
            if credit_manager is None:
                print(">>> Sistema de créditos não disponível, executando sem verificação de créditos")
            else:
                print(">>> Sistema de créditos disponível, calculando créditos necessários...")
                
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                itens_dados = []
                for numero_item, dados_item in itens:
                    itens_dados.append(dados_item)
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_comuns = quantidade_itens - quantidade_especiais
                
                # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    quantidade_itens,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área total do pavimento
                area_total = credit_manager.calcular_area_pavimento(itens_dados)
                
                # Obter obra (do primeiro item)
                obra = itens_dados[0].get('obra', '') if itens_dados else ''
                
                print(f">>> Quantidade de itens no pavimento: {quantidade_itens} ({quantidade_comuns} comuns, {quantidade_especiais} especiais)")
                print(_get_obf_str("credit"))
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento,
                    tipo="",  # Vazio para pavimento completo
                    area_m2=area_total,
                    parte_desenho=f"PAVIMENTO COMPLETO {quantidade_itens} ITENS"
                )
                
                # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                    return
                
                print(_get_obf_str("credit"))
                
                # Atualizar interface imediatamente
                self.atualizar_creditos_interface()

            # ============================================================
            # LIMPEZA DA PASTA DO PAVIMENTO - DEVE SER PRIMEIRO!
            # ============================================================
            robo_dir = self._get_project_root()
            
            # VALIDAÇÃO CRÍTICA: Construir nome da pasta APENAS baseado no PAVIMENTO (nunca no nome do item!)
            # IMPORTANTE: Normalizar o nome do pavimento para evitar problemas de encoding e múltiplas pastas
            import unicodedata
            
            def normalizar_nome_pasta(texto):
                """
                Normaliza o nome para criar pasta consistente, removendo acentos e caracteres especiais
                IMPORTANTE: Esta função garante que pastas com encoding diferente sejam tratadas igualmente
                Exemplo: 'Térreo' e 'TÃ©rreo' ambos se tornam 'Terreo'
                """
                if not texto:
                    return ""
                # Converter para string e remover espaços extras
                texto = str(texto).strip()
                if not texto:
                    return ""
                # Normalizar unicode (NFD) e remover diacríticos (acentos)
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    # Se falhar na normalização, apenas remover caracteres especiais
                    texto_sem_acentos = texto
                # Substituir espaços por underscore e remover caracteres especiais
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                # Remover caracteres não alfanuméricos exceto underscore
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                # Remover underscores múltiplos
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                # Remover underscores no início e fim
                texto_limpo = texto_limpo.strip('_')
                return texto_limpo
            
            pavimento_normalizado = normalizar_nome_pasta(pavimento)
            nome_pasta = f"{pavimento_normalizado}_CIMA"
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")
            
            print(f">>> [NORMALIZAÇÃO] Pavimento original: '{pavimento}'")
            print(f">>> [NORMALIZAÇÃO] Pavimento normalizado: '{pavimento_normalizado}'")

            print(f"\n>>> ============================================================")
            print(f">>> LIMPEZA COMPLETA E FORÇADA DA PASTA DO PAVIMENTO")
            print(f">>> ============================================================")
            print(f">>> ✅ VALIDAÇÃO: Usando PAVIMENTO para nome da pasta (NÃO o nome do item)")
            print(f">>>   Pavimento selecionado: '{pavimento_normalizado}'")
            print(f">>>   Nome da pasta gerado: '{nome_pasta}'")
            print(f">>>   Caminho completo: {pasta_pavimento}")
            print(f">>>   Diretório Combinados: {combinados_dir}")
            
            # VERIFICAÇÃO: Listar todas as pastas existentes para comparação (debug)
            pastas_scripts_dir = os.path.join(robo_dir, "output", "scripts")
            if os.path.exists(pastas_scripts_dir):
                pastas_existentes = [p for p in os.listdir(pastas_scripts_dir) if os.path.isdir(os.path.join(pastas_scripts_dir, p))]
                print(f">>>   [DEBUG] Pastas existentes em output/scripts: {len(pastas_existentes)}")
                if pastas_existentes:
                    # Procurar pastas que possam corresponder ao pavimento (mesmo com encoding diferente)
                    pastas_similares = []
                    pavimento_normalizado_lower = pavimento_normalizado.lower()
                    for pasta in pastas_existentes:
                        pasta_normalizada = normalizar_nome_pasta(pasta.replace('_CIMA', '').replace('_ABCD', '').replace('_GRADES', ''))
                        if pasta_normalizada.lower() == pavimento_normalizado_lower:
                            pastas_similares.append(pasta)
                    
                    if pastas_similares:
                        print(f">>>   ⚠️ ATENÇÃO: Encontradas {len(pastas_similares)} pasta(s) similar(es) ao pavimento:")
                        for pasta_sim in pastas_similares:
                            print(f">>>     - {pasta_sim}")
                            if pasta_sim != nome_pasta:
                                print(f">>>       ⚠️ Esta pasta TAMBÉM SERÁ LIMPA para evitar duplicatas")
                    
                    # DEFINIR LISTA DE PASTAS A LIMPAR: A pasta correta + todas as similares
                    pastas_para_limpar = [nome_pasta]
                    if pastas_similares:
                        for pasta_sim in pastas_similares:
                            if pasta_sim != nome_pasta:
                                pastas_para_limpar.append(pasta_sim)
                    
                    if nome_pasta in pastas_existentes:
                        print(f">>>   ✅ Pasta CORRETA '{nome_pasta}' encontrada - será limpa")
                        if len(pastas_para_limpar) > 1:
                            print(f">>>   🗑️ Total de {len(pastas_para_limpar)} pasta(s) a serem limpas (incluindo similares)")
                    else:
                        print(f">>>   ℹ️ Pasta '{nome_pasta}' não existe ainda - será criada")
                        if len(pastas_para_limpar) > 1:
                            print(f">>>   🗑️ {len(pastas_para_limpar) - 1} pasta(s) similar(es) será(ão) limpa(s) antes")
            
            # ============================================================
            # LIMPEZA COMPLETA E FORÇADA DA PASTA ANTES DE QUALQUER PROCESSAMENTO
            # ============================================================
            # ESTRATÉGIA: EXCLUIR A PASTA INTEIRA E RECRIAR VAZIA
            # Isso garante que não haja resquícios de execuções anteriores
            import shutil
            
            # PRIMEIRO: Limpar pasta Combinados se existir (antes de excluir a pasta principal)
            if os.path.exists(combinados_dir):
                print(f">>> 🗑️ LIMPANDO PASTA COMBINADOS PRIMEIRO...")
                print(f">>>   Pasta Combinados: {combinados_dir}")
                try:
                    # Listar arquivos antes de excluir
                    arquivos_combinados_antes = []
                    for item in os.listdir(combinados_dir):
                        if item.endswith('.scr'):
                            arquivos_combinados_antes.append(item)
                    
                    if arquivos_combinados_antes:
                        print(f">>>   ⚠️ Encontrados {len(arquivos_combinados_antes)} arquivo(s) SCR em Combinados")
                        print(f">>>   Arquivos: {arquivos_combinados_antes[:10]}...")  # Mostrar primeiros 10
                    
                    # Excluir pasta Combinados completamente
                    shutil.rmtree(combinados_dir, ignore_errors=False)
                    print(f">>>   ✅ Pasta Combinados excluída completamente")
                    time.sleep(0.2)  # Aguardar um pouco
                except Exception as e:
                    print(f">>>   ⚠️ Erro ao excluir pasta Combinados: {e}")
                    # Tentar com ignore_errors
                    try:
                        shutil.rmtree(combinados_dir, ignore_errors=True)
                        time.sleep(0.3)
                        print(f">>>   ✅ Pasta Combinados excluída com ignore_errors=True")
                    except Exception as e2:
                        print(f">>>   ❌ Erro definitivo ao excluir pasta Combinados: {e2}")
                        # Última tentativa: remover arquivos individualmente
                        try:
                            for item in os.listdir(combinados_dir):
                                item_path = os.path.join(combinados_dir, item)
                                try:
                                    if os.path.isdir(item_path):
                                        shutil.rmtree(item_path, ignore_errors=True)
                                    else:
                                        os.remove(item_path)
                                except:
                                    pass
                            print(f">>>   ⚠️ Arquivos Combinados removidos manualmente")
                        except:
                            print(f">>>   ⚠️ Não foi possível limpar pasta Combinados completamente")
            else:
                print(f">>> ℹ️ Pasta Combinados não existe (será criada após combinar scripts)")
            
            # AGORA: Limpar pasta principal do pavimento E todas as similares (se existirem)
            # Primeiro, limpar todas as pastas similares que não são a correta
            pastas_scripts_dir = os.path.join(robo_dir, "output", "scripts")
            if os.path.exists(pastas_scripts_dir):
                # Buscar pastas similares novamente para garantir que temos a lista atualizada
                pastas_existentes_atual = [p for p in os.listdir(pastas_scripts_dir) 
                                         if os.path.isdir(os.path.join(pastas_scripts_dir, p))]
                pastas_similares_para_limpar = []
                pavimento_normalizado_lower = pavimento_normalizado.lower()
                
                for pasta in pastas_existentes_atual:
                    # Remover sufixos (_CIMA, _ABCD, etc) e normalizar
                    pasta_sem_sufixo = pasta.replace('_CIMA', '').replace('_ABCD', '').replace('_GRADES', '')
                    pasta_normalizada = normalizar_nome_pasta(pasta_sem_sufixo)
                    
                    # Se normalizada é igual ao pavimento E tem sufixo _CIMA E não é a pasta correta
                    if (pasta_normalizada.lower() == pavimento_normalizado_lower and 
                        pasta.endswith('_CIMA') and 
                        pasta != nome_pasta):
                        pastas_similares_para_limpar.append(pasta)
                
                # LIMPAR PASTAS SIMILARES PRIMEIRO (se existirem)
                if pastas_similares_para_limpar:
                    print(f"\n>>> 🗑️ LIMPANDO {len(pastas_similares_para_limpar)} PASTA(S) SIMILAR(ES) PRIMEIRO...")
                    for pasta_sim in pastas_similares_para_limpar:
                        caminho_pasta_sim = os.path.join(pastas_scripts_dir, pasta_sim)
                        if os.path.exists(caminho_pasta_sim):
                            try:
                                # Contar arquivos antes de excluir
                                arquivos_sim = []
                                for item in os.listdir(caminho_pasta_sim):
                                    if item.endswith('.scr'):
                                        arquivos_sim.append(item)
                                
                                if arquivos_sim:
                                    print(f">>>   ⚠️ Encontrados {len(arquivos_sim)} arquivo(s) SCR em '{pasta_sim}'")
                                
                                shutil.rmtree(caminho_pasta_sim, ignore_errors=False)
                                print(f">>>   ✅ Pasta similar '{pasta_sim}' excluída completamente")
                                time.sleep(0.2)
                            except Exception as e:
                                print(f">>>   ⚠️ Erro ao excluir pasta similar '{pasta_sim}': {e}")
                                try:
                                    shutil.rmtree(caminho_pasta_sim, ignore_errors=True)
                                    print(f">>>   ✅ Pasta similar '{pasta_sim}' excluída com ignore_errors")
                                except:
                                    print(f">>>   ❌ Não foi possível excluir pasta similar '{pasta_sim}'")
            
            # AGORA: Limpar pasta principal do pavimento (a correta)
            if os.path.exists(pasta_pavimento):
                print(f"\n>>> 🗑️ PASTA PRINCIPAL JÁ EXISTE - EXCLUINDO COMPLETAMENTE...")
                print(f">>>   Pasta: {pasta_pavimento}")
                
                # Contar arquivos antes de excluir (para log)
                try:
                    arquivos_antes_limpeza = []
                    for item in os.listdir(pasta_pavimento):
                        item_path = os.path.join(pasta_pavimento, item)
                        if os.path.isfile(item_path) and item.endswith('.scr'):
                            arquivos_antes_limpeza.append(item)
                        elif os.path.isdir(item_path):
                            # Contar arquivos em subpastas também
                            for subitem in os.listdir(item_path):
                                if subitem.endswith('.scr'):
                                    arquivos_antes_limpeza.append(f"{item}/{subitem}")
                    
                    if arquivos_antes_limpeza:
                        print(f">>>   ⚠️ Encontrados {len(arquivos_antes_limpeza)} arquivo(s) SCR para remover")
                        print(f">>>   Arquivos: {arquivos_antes_limpeza[:10]}...")  # Mostrar primeiros 10
                except Exception as e:
                    print(f">>>   ⚠️ Erro ao listar arquivos antes da limpeza: {e}")
                
                # EXCLUIR A PASTA COMPLETAMENTE
                try:
                    shutil.rmtree(pasta_pavimento, ignore_errors=False)
                    print(f">>>   ✅ Pasta excluída completamente")
                    # Aguardar um pouco para o sistema liberar os arquivos
                    time.sleep(0.3)
                except Exception as e:
                    print(f">>>   ❌ Erro ao excluir pasta: {e}")
                    # Tentar novamente com ignore_errors
                    try:
                        shutil.rmtree(pasta_pavimento, ignore_errors=True)
                        time.sleep(0.5)
                        print(f">>>   ✅ Pasta excluída com ignore_errors=True")
                    except Exception as e2:
                        print(f">>>   ❌ Erro definitivo ao excluir pasta: {e2}")
                        # Última tentativa: remover arquivos individualmente
                        try:
                            for item in os.listdir(pasta_pavimento):
                                item_path = os.path.join(pasta_pavimento, item)
                                try:
                                    if os.path.isdir(item_path):
                                        shutil.rmtree(item_path, ignore_errors=True)
                                    else:
                                        os.remove(item_path)
                                except:
                                    pass
                            print(f">>>   ⚠️ Arquivos removidos manualmente (pasta pode estar parcialmente limpa)")
                        except:
                            print(f">>>   ❌ Não foi possível limpar a pasta completamente")
                            messagebox.showerror("Erro Crítico", 
                                f"Não foi possível limpar a pasta do pavimento.\n\n"
                                f"Pasta: {pasta_pavimento}\n\n"
                                f"Feche o AutoCAD e qualquer programa que possa estar usando os arquivos, "
                                f"depois tente novamente.")
                            return
            else:
                print(f">>> ℹ️ Pasta não existe, será criada: {pasta_pavimento}")
            
            # RECRIAR A PASTA VAZIA (sem criar Combinados ainda - será criado depois pelo combinador)
            try:
                os.makedirs(pasta_pavimento, exist_ok=True)
                print(f">>> ✅ Pasta principal recriada vazia: {pasta_pavimento}")
                print(f">>> ℹ️ Pasta Combinados será criada automaticamente pelo combinador")
                
                # VERIFICAÇÃO FINAL CRÍTICA: Garantir que está realmente vazia ANTES de gerar qualquer script
                time.sleep(0.2)  # Aguardar sistema liberar arquivos
                itens_apos_criacao = os.listdir(pasta_pavimento)
                
                if itens_apos_criacao:
                    print(f"\n>>> ⚠️⚠️⚠️ PROBLEMA CRÍTICO: Ainda há {len(itens_apos_criacao)} itens na pasta após recriação!")
                    print(f">>>   Itens encontrados: {itens_apos_criacao}")
                    
                    # Listar detalhadamente o que foi encontrado
                    arquivos_scr_encontrados = [i for i in itens_apos_criacao if i.endswith('.scr')]
                    pastas_encontradas = [i for i in itens_apos_criacao if os.path.isdir(os.path.join(pasta_pavimento, i))]
                    
                    if arquivos_scr_encontrados:
                        print(f">>>   ⚠️ Encontrados {len(arquivos_scr_encontrados)} arquivo(s) .scr que DEVERIAM ter sido removidos:")
                        for scr in arquivos_scr_encontrados:
                            print(f">>>     - {scr}")
                    
                    if pastas_encontradas:
                        print(f">>>   ⚠️ Encontradas {len(pastas_encontradas)} subpasta(s):")
                        for subpasta in pastas_encontradas:
                            print(f">>>     - {subpasta}")
                    
                    # Remover manualmente TODOS os itens restantes
                    print(f">>> 🗑️ REMOVENDO MANUALMENTE todos os {len(itens_apos_criacao)} itens restantes...")
                    removidos_manual = 0
                    for item in itens_apos_criacao:
                        item_path = os.path.join(pasta_pavimento, item)
                        try:
                            if os.path.isdir(item_path):
                                shutil.rmtree(item_path, ignore_errors=True)
                            else:
                                os.remove(item_path)
                            print(f">>>   ✅ Removido: {item}")
                            removidos_manual += 1
                        except Exception as e:
                            print(f">>>   ❌ Erro ao remover {item}: {e}")
                    
                    if removidos_manual == len(itens_apos_criacao):
                        print(f">>> ✅ Todos os {removidos_manual} itens foram removidos manualmente")
                    else:
                        print(f">>> ⚠️ Apenas {removidos_manual} de {len(itens_apos_criacao)} itens foram removidos")
                    
                    # Verificar novamente após remoção manual
                    time.sleep(0.2)
                    itens_finais_verificacao = os.listdir(pasta_pavimento)
                    if itens_finais_verificacao:
                        print(f">>> ❌❌❌ CRÍTICO: Ainda há {len(itens_finais_verificacao)} itens após remoção manual!")
                        print(f">>>   Itens restantes: {itens_finais_verificacao}")
                        messagebox.showerror("Erro Crítico", 
                            f"Não foi possível limpar completamente a pasta do pavimento.\n\n"
                            f"Pasta: {pasta_pavimento}\n"
                            f"Itens restantes: {len(itens_finais_verificacao)}\n\n"
                            f"Feche o AutoCAD e qualquer programa usando esses arquivos,\n"
                            f"depois tente novamente.")
                        return
                    else:
                        print(f">>> ✅ VERIFICAÇÃO FINAL: Pasta está completamente vazia após remoção manual")
                else:
                    print(f">>> ✅✅✅ CONFIRMADO: Pasta está completamente vazia (0 itens)")
                    print(f">>> ✅ A pasta está PRONTA para receber novos scripts")
                
                # CONTAGEM FINAL: Contar arquivos SCR na pasta (deve ser 0)
                arquivos_scr_finais = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                if arquivos_scr_finais:
                    print(f"\n>>> ❌❌❌ ERRO CRÍTICO: Ainda há {len(arquivos_scr_finais)} arquivo(s) .scr na pasta!")
                    print(f">>>   Arquivos: {arquivos_scr_finais}")
                    for scr in arquivos_scr_finais:
                        scr_path = os.path.join(pasta_pavimento, scr)
                        try:
                            os.remove(scr_path)
                            print(f">>>   ✅ Removido: {scr}")
                        except Exception as e:
                            print(f">>>   ❌ Erro ao remover {scr}: {e}")
                else:
                    print(f">>> ✅ CONFIRMADO: 0 arquivos .scr na pasta - PRONTO PARA GERAR NOVOS SCRIPTS")
            except Exception as e:
                print(f">>> ❌ Erro ao recriar pasta: {e}")
                messagebox.showerror("Erro", f"Não foi possível criar a pasta do pavimento: {e}")
                return
            
            print(f">>> ============================================================")
            print(f">>> LIMPEZA CONCLUÍDA - PASTA PRONTA PARA NOVOS SCRIPTS")
            print(f">>> ============================================================\n")
            
            # Importar a mesma função usada pelo D.CAD 1 ITEM
            try:
                import sys
                sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'interfaces'))
                from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                
                print(f"\n>>> 🎯 USANDO MESMA LÓGICA DO D.CAD 1 ITEM")
                print(f">>> Processando {len(itens)} itens do pavimento {pavimento}")
                print(f">>> [DEBUG FINAL] Lista de itens que serão processados:")
                for idx, (num, d) in enumerate(itens, 1):
                    print(f">>>   [{idx}] Número: {num}, Nome: {d.get('nome', num)}, Pavimento: {d.get('pavimento', 'N/A')}")
                
                # Contador para verificar quantos itens foram realmente processados
                itens_processados_count = 0
                
                # Processar cada item usando a mesma lógica do D.CAD 1 ITEM
                # Usar set para garantir que cada item seja processado apenas uma vez
                itens_processados_numeros = set()
                
                for idx, (numero, dados) in enumerate(itens):
                    # VALIDAÇÃO: Verificar se já foi processado (proteção extra contra duplicatas)
                    if numero in itens_processados_numeros:
                        print(f">>> ⚠️ ITEM {numero} JÁ FOI PROCESSADO - PULANDO DUPLICATA")
                        continue
                    itens_processados_numeros.add(numero)
                    
                    # VALIDAÇÃO: Verificar novamente se o item pertence ao pavimento correto
                    pavimento_item_loop = dados.get("pavimento", "")
                    if pavimento_item_loop != pavimento:
                        print(f">>> ⚠️ PULANDO ITEM {numero}: pertence ao pavimento '{pavimento_item_loop}', não '{pavimento}'")
                        continue
                    
                    print(f"\n>>> PROCESSANDO ITEM {idx+1}/{len(itens)}: {numero}")
                    print(f">>> Pavimento do item: {pavimento_item_loop} (esperado: {pavimento})")
                    
                    nome = dados.get("nome", numero)
                    print(f">>> Nome do item: {nome}")
                    
                    # DEBUG: Verificar se é pilar especial
                    pilar_especial_data = dados.get('pilar_especial', {})
                    ativar_pilar_especial = dados.get('ativar_pilar_especial', False)
                    tipo_pilar_especial = dados.get('tipo_pilar_especial', '')
                    
                    print(f">>> [DEBUG] Pilar especial data completo: {pilar_especial_data}")
                    print(f">>> [DEBUG] Tipo de pilar_especial_data: {type(pilar_especial_data)}")
                    if isinstance(pilar_especial_data, dict):
                        print(f">>> [DEBUG] Chaves em pilar_especial_data: {list(pilar_especial_data.keys())}")
                        if 'comprimentos' in pilar_especial_data:
                            print(f">>> [DEBUG] Comprimentos: {pilar_especial_data.get('comprimentos', {})}")
                        if 'larguras' in pilar_especial_data:
                            print(f">>> [DEBUG] Larguras: {pilar_especial_data.get('larguras', {})}")
                    print(f">>> [DEBUG] Ativar pilar especial (direto): {ativar_pilar_especial}")
                    print(f">>> [DEBUG] Ativar pilar especial (aninhado): {pilar_especial_data.get('ativar_pilar_especial', False) if isinstance(pilar_especial_data, dict) else False}")
                    print(f">>> [DEBUG] Tipo pilar especial (direto): {tipo_pilar_especial}")
                    print(f">>> [DEBUG] Tipo pilar especial (aninhado): {pilar_especial_data.get('tipo_pilar_especial', '') if isinstance(pilar_especial_data, dict) else ''}")
                    
                    # Verificar se realmente é pilar especial
                    eh_pilar_especial = (
                        ativar_pilar_especial or 
                        pilar_especial_data.get('ativar_pilar_especial', False)
                    ) and (
                        tipo_pilar_especial in ['L', 'T', 'U'] or
                        pilar_especial_data.get('tipo_pilar_especial', '') in ['L', 'T', 'U']
                    )
                    
                    print(f">>> [DEBUG] É pilar especial: {eh_pilar_especial}")
                    
                    if eh_pilar_especial:
                        print(f">>> 🎯 PILAR ESPECIAL DETECTADO: {nome}")
                    else:
                        print(f">>> 📦 PILAR COMUM DETECTADO: {nome}")
                    
                    # Preparar dados do pilar especial com parafusos especiais dentro (mesma lógica de _dcad_item_cima)
                    pilar_especial_data = dados.get('pilar_especial', {}).copy() if dados.get('pilar_especial', {}) else {}
                    # Adicionar parafusos especiais dentro de pilar_especial
                    pilar_especial_data['parafusos_especiais'] = {
                        'parafusos_a': self._obter_parafusos_especiais_para_transferencia('a'),
                        'parafusos_e': self._obter_parafusos_especiais_para_transferencia('e')
                    }
                    
                    # Usar exatamente a mesma estrutura de dados do D.CAD 1 ITEM
                    dados_pilar = {
                        'interface_principal': self,  # Interface principal completa
                        'pavimento': pavimento,
                        'nome': nome,
                        'comprimento': dados.get('comprimento', ''),
                        'largura': dados.get('largura', ''),
                        'parafuso_p1_p2': int(float(dados.get('parafusos', {}).get('par_1_2', 0) or 0)),
                        'parafuso_p2_p3': int(float(dados.get('parafusos', {}).get('par_2_3', 0) or 0)),
                        'parafuso_p3_p4': int(float(dados.get('parafusos', {}).get('par_3_4', 0) or 0)),
                        'parafuso_p4_p5': int(float(dados.get('parafusos', {}).get('par_4_5', 0) or 0)),
                        'parafuso_p5_p6': int(float(dados.get('parafusos', {}).get('par_5_6', 0) or 0)),
                        'parafuso_p6_p7': int(float(dados.get('parafusos', {}).get('par_6_7', 0) or 0)),
                        'parafuso_p7_p8': int(float(dados.get('parafusos', {}).get('par_7_8', 0) or 0)),
                        'parafuso_p8_p9': int(float(dados.get('parafusos', {}).get('par_8_9', 0) or 0)),
                        # Grades grupo 1 (Grade A)
                        'grades_grupo1': dados.get('grades', {}),
                        'detalhes_grades': dados.get('detalhes_grades', {}),
                        # Grades grupo 2 (Grade B)
                        'grades_grupo2': dados.get('grades_grupo2', {}),
                        'detalhes_grades_grupo2': dados.get('detalhes_grades_grupo2', {}),
                        # Pilar especial completo (com parafusos especiais dentro)
                        'pilar_especial': pilar_especial_data,
                        # Dados do pilar especial (compatibilidade) - extrair corretamente da estrutura aninhada
                        'ativar_pilar_especial': pilar_especial_data.get('ativar_pilar_especial', False) if isinstance(pilar_especial_data, dict) else False,
                        'tipo_pilar_especial': pilar_especial_data.get('tipo_pilar_especial', '') if isinstance(pilar_especial_data, dict) else '',
                        # Extrair comprimentos e larguras da estrutura aninhada
                        'comp_1': pilar_especial_data.get('comprimentos', {}).get('comp_1', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('comprimentos', {}), dict) else '',
                        'comp_2': pilar_especial_data.get('comprimentos', {}).get('comp_2', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('comprimentos', {}), dict) else '',
                        'comp_3': pilar_especial_data.get('comprimentos', {}).get('comp_3', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('comprimentos', {}), dict) else '',
                        'larg_1': pilar_especial_data.get('larguras', {}).get('larg_1', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('larguras', {}), dict) else '',
                        'larg_2': pilar_especial_data.get('larguras', {}).get('larg_2', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('larguras', {}), dict) else '',
                        'larg_3': pilar_especial_data.get('larguras', {}).get('larg_3', '') if isinstance(pilar_especial_data, dict) and isinstance(pilar_especial_data.get('larguras', {}), dict) else '',
                        # Parafusos especiais para transferência (também no nível raiz para compatibilidade)
                        'parafusos_especiais': pilar_especial_data.get('parafusos_especiais', {})
                    }
                    
                    # Log detalhado dos dados que serão passados
                    print(f"\n>>> [DEBUG DADOS PILAR] Dados que serão passados para preencher_campos_diretamente_e_gerar_scripts:")
                    print(f">>>   - ativar_pilar_especial: {dados_pilar.get('ativar_pilar_especial')}")
                    print(f">>>   - tipo_pilar_especial: {dados_pilar.get('tipo_pilar_especial')}")
                    print(f">>>   - comp_1: '{dados_pilar.get('comp_1')}', comp_2: '{dados_pilar.get('comp_2')}', comp_3: '{dados_pilar.get('comp_3')}'")
                    print(f">>>   - larg_1: '{dados_pilar.get('larg_1')}', larg_2: '{dados_pilar.get('larg_2')}', larg_3: '{dados_pilar.get('larg_3')}'")
                    print(f">>>   - pilar_especial completo: {dados_pilar.get('pilar_especial')}")
                    
                    print(f">>> Executando preencher_campos_diretamente_e_gerar_scripts para {nome}...")
                    
                    # Verificar quantos arquivos SCR existem ANTES de processar este item
                    arquivos_antes = []
                    pastas_antes = []
                    if os.path.exists(pasta_pavimento):
                        for item in os.listdir(pasta_pavimento):
                            item_path = os.path.join(pasta_pavimento, item)
                            if os.path.isfile(item_path) and item.endswith('.scr'):
                                arquivos_antes.append(item)
                            elif os.path.isdir(item_path):
                                pastas_antes.append(item)
                                # Verificar arquivos dentro de subpastas também
                                try:
                                    for subitem in os.listdir(item_path):
                                        subitem_path = os.path.join(item_path, subitem)
                                        if os.path.isfile(subitem_path) and subitem.endswith('.scr'):
                                            arquivos_antes.append(f"{item}/{subitem}")
                                except:
                                    pass
                    
                    print(f"\n>>> ========================================")
                    print(f">>> ===== PROCESSANDO ITEM {idx+1}/{len(itens)} ===== ")
                    print(f">>> ========================================")
                    print(f">>> Item número: {numero}")
                    print(f">>> Item nome: {nome}")
                    print(f"\n>>> 📊 ESTADO ANTES DE PROCESSAR '{nome}':")
                    print(f">>>   Total de arquivos .scr na pasta: {len(arquivos_antes)}")
                    
                    if arquivos_antes:
                        print(f">>>   ⚠️⚠️⚠️ ATENÇÃO: HÁ {len(arquivos_antes)} ARQUIVO(S) .SCR NA PASTA ANTES DE GERAR!")
                        print(f">>>   Arquivos existentes:")
                        for arq in sorted(arquivos_antes)[:10]:
                            print(f">>>     - {arq}")
                        if len(arquivos_antes) > 10:
                            print(f">>>     ... e mais {len(arquivos_antes) - 10} arquivo(s)")
                    
                    if pastas_antes:
                        print(f">>>   Subpastas encontradas: {pastas_antes}")
                    
                    if len(arquivos_antes) > 0 and idx == 0:
                        print(f"\n>>> ❌❌❌ PROBLEMA CRÍTICO: HÁ {len(arquivos_antes)} ARQUIVO(S) NA PASTA ANTES DO PRIMEIRO ITEM!")
                        print(f">>>   Isso significa que a limpeza NÃO funcionou corretamente!")
                        print(f">>>   Arquivos que deveriam ter sido removidos:")
                        for arq in sorted(arquivos_antes):
                            print(f">>>     - {arq}")
                    
                    # REMOÇÃO ADICIONAL: Antes de processar este item, remover qualquer arquivo SCR relacionado
                    # Isso garante que não haverá arquivos antigos mesmo se a limpeza inicial falhou
                    nome_item_para_busca = nome
                    arquivos_relacionados_antes = []
                    if os.path.exists(pasta_pavimento):
                        arquivos_relacionados_antes = [
                            f for f in os.listdir(pasta_pavimento) 
                            if f.endswith('.scr') and (f.startswith(f"{nome}_CIMA") or f.startswith(f"{numero}_CIMA"))
                        ]
                        if arquivos_relacionados_antes:
                            print(f">>> 🗑️ Removendo {len(arquivos_relacionados_antes)} arquivo(s) SCR relacionado(s) a este item antes de processar:")
                            for arq_rel in arquivos_relacionados_antes:
                                caminho_arq = os.path.join(pasta_pavimento, arq_rel)
                                try:
                                    os.remove(caminho_arq)
                                    print(f">>>   ✅ Removido: {arq_rel}")
                                except Exception as e:
                                    print(f">>>   ⚠️ Erro ao remover {arq_rel}: {e}")
                    
                    # Chamar a mesma função usada pelo D.CAD 1 ITEM
                    print(f">>> Chamando preencher_campos_diretamente_e_gerar_scripts...")
                    preencher_campos_diretamente_e_gerar_scripts(dados_pilar)
                    print(f">>> Função retornou com sucesso")
                    
                    # Verificar quantos arquivos SCR existem DEPOIS de processar este item
                    time.sleep(0.1)  # Aguardar sistema salvar arquivos
                    arquivos_depois = []
                    pastas_depois = []
                    if os.path.exists(pasta_pavimento):
                        for item in os.listdir(pasta_pavimento):
                            item_path = os.path.join(pasta_pavimento, item)
                            if os.path.isfile(item_path) and item.endswith('.scr'):
                                arquivos_depois.append(item)
                            elif os.path.isdir(item_path):
                                pastas_depois.append(item)
                                # Verificar arquivos dentro de subpastas também
                                try:
                                    for subitem in os.listdir(item_path):
                                        subitem_path = os.path.join(item_path, subitem)
                                        if os.path.isfile(subitem_path) and subitem.endswith('.scr'):
                                            arquivos_depois.append(f"{item}/{subitem}")
                                except:
                                    pass
                    
                    novos_arquivos = [f for f in arquivos_depois if f not in arquivos_antes]
                    
                    print(f"\n>>> 📊 ESTADO DEPOIS DE PROCESSAR '{nome}':")
                    print(f">>>   Total de arquivos .scr na pasta: {len(arquivos_depois)}")
                    print(f">>>   Arquivos ANTES: {len(arquivos_antes)}")
                    print(f">>>   Arquivos DEPOIS: {len(arquivos_depois)}")
                    print(f">>>   NOVOS arquivos gerados: {len(novos_arquivos)}")
                    
                    if novos_arquivos:
                        print(f">>>   ✅ Arquivos gerados para '{nome}':")
                        for novo_arq in sorted(novos_arquivos):
                            print(f">>>     ✅ {novo_arq}")
                    
                    if len(novos_arquivos) == 0:
                        print(f">>>   ⚠️ ATENÇÃO: Nenhum arquivo SCR foi gerado para o item '{nome}'!")
                        print(f">>>   (Isso pode ser normal se o item não gerou scripts ou se houve erro)")
                    
                    if len(novos_arquivos) > 2 and not eh_pilar_especial:
                        print(f">>>   ⚠️ ATENÇÃO: {len(novos_arquivos)} arquivos gerados para um pilar COMUM!")
                        print(f">>>   (Pilares comuns devem gerar apenas 1 arquivo)")
                    
                    if eh_pilar_especial and len(novos_arquivos) > 4:
                        print(f">>>   ⚠️ ATENÇÃO: {len(novos_arquivos)} arquivos gerados para pilar ESPECIAL!")
                        print(f">>>   (Pilares especiais devem gerar partes individuais + 1 combinado)")
                        
                        # REGRA: Para pilares comuns, deve haver apenas 1 arquivo
                        # Para pilares especiais, pode haver um arquivo COMBINADO
                        pilar_especial = dados.get('ativar_pilar_especial', False)
                        
                        if not pilar_especial:
                            # PILAR COMUM: Manter apenas o arquivo principal e remover versões com sufixo
                            arquivo_principal = f"{nome}_CIMA.scr"
                            arquivos_para_manter = []
                            arquivos_para_remover_agora = []
                            
                            for novo_arq in novos_arquivos:
                                if novo_arq == arquivo_principal:
                                    arquivos_para_manter.append(novo_arq)
                                elif novo_arq.startswith(f"{nome}_CIMA") and novo_arq.endswith('.scr'):
                                    # É uma versão com sufixo (-1, -2, etc.) ou COMBINADO
                                    arquivos_para_remover_agora.append(novo_arq)
                            
                            # Se encontrou o arquivo principal, remover os extras
                            if arquivos_para_manter:
                                print(f">>> 🗑️ REMOVENDO {len(arquivos_para_remover_agora)} arquivo(s) extra(s) imediatamente...")
                                for arq_remover in arquivos_para_remover_agora:
                                    caminho_remover = os.path.join(pasta_pavimento, arq_remover)
                                    try:
                                        os.remove(caminho_remover)
                                        print(f">>>   ✅ Removido imediatamente: {arq_remover}")
                                    except Exception as e:
                                        print(f">>>   ❌ Erro ao remover {arq_remover}: {e}")
                            else:
                                print(f">>> ⚠️ Arquivo principal {arquivo_principal} não encontrado, mantendo o primeiro arquivo gerado")
                                # Manter apenas o primeiro e remover os demais
                                arquivo_manter = novos_arquivos[0]
                                for arq_remover in novos_arquivos[1:]:
                                    caminho_remover = os.path.join(pasta_pavimento, arq_remover)
                                    try:
                                        os.remove(caminho_remover)
                                        print(f">>>   ✅ Removido extra: {arq_remover} (mantido: {arquivo_manter})")
                                    except Exception as e:
                                        print(f">>>   ❌ Erro ao remover {arq_remover}: {e}")
                    
                    print(f">>> ===== FIM DO PROCESSAMENTO DO ITEM {nome} =====\n")
                    
                    itens_processados_count += 1
                    print(f">>> ✅ Item {nome} processado com sucesso ({itens_processados_count}/{len(itens)})")
                
                print(f"\n>>> 🎯 TODOS OS ITENS PROCESSADOS COM SUCESSO!")
                print(f">>> [DEBUG] Total de itens processados: {itens_processados_count} de {len(itens)} esperados")
                
                if itens_processados_count != len(itens):
                    print(f">>> ⚠️ ATENÇÃO: Foram processados {itens_processados_count} itens, mas {len(itens)} eram esperados!")
                
                # Verificar quantos arquivos SCR foram gerados na pasta antes de combinar
                pasta_pavimento_verificar = os.path.join(robo_dir, "output", "scripts", nome_pasta)
                if os.path.exists(pasta_pavimento_verificar):
                    arquivos_scr_gerados = [f for f in os.listdir(pasta_pavimento_verificar) if f.endswith('.scr')]
                    print(f"\n>>> ===== VERIFICAÇÃO FINAL ANTES DO COMBINADOR ===== ")
                    print(f">>> Total de itens processados: {itens_processados_count}")
                    print(f">>> Total de itens esperados: {len(itens)}")
                    print(f">>> Arquivos SCR encontrados na pasta: {len(arquivos_scr_gerados)}")
                    print(f">>> Lista completa de arquivos SCR:")
                    for i, arquivo in enumerate(sorted(arquivos_scr_gerados), 1):
                        # Verificar a qual item cada arquivo corresponde
                        corresponde_a_item = False
                        nome_item_correspondente = "DESCONHECIDO"
                        for num_item, dados_item in itens:
                            nome_item = dados_item.get("nome", num_item)
                            if (
                                arquivo.startswith(f"{nome_item}_CIMA") or
                                arquivo.startswith(f"{num_item}_CIMA") or
                                arquivo.startswith(f"{nome_item}_COMBINADO_CIMA") or
                                arquivo.startswith(f"{num_item}_COMBINADO_CIMA")
                            ):
                                corresponde_a_item = True
                                nome_item_correspondente = nome_item
                                break
                        
                        status = f"✅ Item: {nome_item_correspondente}" if corresponde_a_item else "❌ NÃO CORRESPONDE"
                        print(f">>>   [{i}] {arquivo} - {status}")
                    
                    if len(arquivos_scr_gerados) != len(itens):
                        print(f"\n>>> ⚠️ PROBLEMA: {len(arquivos_scr_gerados)} arquivos encontrados, mas apenas {len(itens)} itens foram processados")
                    else:
                        print(f"\n>>> ✅ CORRETO: {len(arquivos_scr_gerados)} arquivos correspondem aos {len(itens)} itens processados")
                    print(f">>> ===== FIM DA VERIFICAÇÃO =====\n")
                    
                    # CRÍTICO: Identificar quais arquivos SCR correspondem aos itens processados
                    print(f">>> [DEBUG] Verificando correspondência entre arquivos SCR e itens processados...")
                    
                    # Criar lista de nomes esperados baseados nos itens processados
                    nomes_itens_processados = {}
                    for numero_item, dados_item in itens:
                        nome_item = dados_item.get("nome", numero_item)
                        numero_item_str = str(numero_item)
                        nome_item_str = str(nome_item)
                        
                        # Criar múltiplas variações possíveis do nome
                        variacoes = [
                            nome_item_str,
                            numero_item_str,
                            nome_item_str.strip(),
                            numero_item_str.strip(),
                            nome_item_str.replace('P', ''),  # P8 -> 8
                            numero_item_str.replace('P', ''),
                        ]
                        
                        # Armazenar todas as variações como chaves que apontam para o mesmo item
                        for var in variacoes:
                            if var:
                                nomes_itens_processados[var.strip()] = (numero_item, nome_item)
                    
                    print(f">>> [DEBUG] Variações de nomes esperadas: {list(nomes_itens_processados.keys())}")
                    
                    # Verificar quais arquivos correspondem aos itens processados
                    # IMPORTANTE: Para pilares especiais, aceitar apenas arquivos COMBINADOS ou scripts normais
                    arquivos_validos = []
                    arquivos_para_remover = []
                    
                    for arquivo_scr in arquivos_scr_gerados:
                        # Regra prioritária: arquivos combinados são sempre válidos (pilares especiais)
                        if arquivo_scr.endswith('_COMBINADO_CIMA.scr'):
                            arquivos_validos.append(arquivo_scr)
                            print(f">>>   ✅ Arquivo COMBINADO MANTIDO (regra prioritária): {arquivo_scr}")
                            continue
                        nome_base_arquivo = arquivo_scr.replace('_CIMA.scr', '').replace('_COMBINADO_CIMA.scr', '')
                        # Remover números de versão do final (ex: -1, -2, -1-1)
                        # Pode ter múltiplos hífens: P8_CIMA-1-1.scr -> P8
                        partes = nome_base_arquivo.split('-')
                        nome_base_sem_versao = partes[0].strip()
                        
                        # Verificar se é um arquivo COMBINADO (para pilares especiais)
                        eh_combinado = arquivo_scr.endswith('_COMBINADO_CIMA.scr')
                        if eh_combinado:
                            # Para combinados, aceitar se o nome base corresponde a um item
                            corresponde = False
                            for nome_esperado in nomes_itens_processados.keys():
                                if (nome_base_sem_versao == nome_esperado or 
                                    nome_base_sem_versao.replace('P', '') == nome_esperado.replace('P', '') or
                                    nome_esperado.replace('P', '') == nome_base_sem_versao.replace('P', '')):
                                    corresponde = True
                                    break
                            
                            if corresponde:
                                arquivos_validos.append(arquivo_scr)
                                print(f">>>   ✅ Arquivo COMBINADO VÁLIDO: {arquivo_scr}")
                            else:
                                arquivos_para_remover.append(arquivo_scr)
                                print(f">>>   ❌ Arquivo COMBINADO INVÁLIDO: {arquivo_scr} (não corresponde a nenhum item processado)")
                        else:
                            # Para scripts normais (não combinados), verificar correspondência
                            # MAS: se houver um arquivo combinado para o mesmo item, REMOVER o script normal
                            corresponde = False
                            tem_combinado = False
                            
                            # Verificar se há arquivo combinado para este item
                            for nome_esperado in nomes_itens_processados.keys():
                                nome_combinado_esperado = f"{nome_esperado}_COMBINADO_CIMA.scr"
                                if nome_combinado_esperado in arquivos_scr_gerados:
                                    # Há um combinado, então este script normal é uma parte solta e deve ser removido
                                    if (nome_base_sem_versao == nome_esperado or 
                                        nome_base_sem_versao.replace('P', '') == nome_esperado.replace('P', '')):
                                        tem_combinado = True
                                        break
                                
                                # Verificar se corresponde ao item
                                if (nome_base_sem_versao == nome_esperado or 
                                    nome_base_sem_versao.startswith(nome_esperado) or 
                                    nome_esperado.startswith(nome_base_sem_versao) or
                                    nome_base_sem_versao.replace('P', '') == nome_esperado.replace('P', '')):
                                    corresponde = True
                                    break
                            
                            if tem_combinado:
                                # Se há combinado, remover este script normal (é parte de um pilar especial)
                                arquivos_para_remover.append(arquivo_scr)
                                print(f">>>   ❌ Arquivo PARTE SOLTA (há combinado): {arquivo_scr}")
                            elif corresponde:
                                arquivos_validos.append(arquivo_scr)
                                print(f">>>   ✅ Arquivo NORMAL VÁLIDO: {arquivo_scr}")
                            else:
                                arquivos_para_remover.append(arquivo_scr)
                                print(f">>>   ❌ Arquivo INVÁLIDO: {arquivo_scr} (não corresponde a nenhum item processado)")
                    
                    # Se há arquivos inválidos, removê-los
                    if arquivos_para_remover:
                        print(f">>> [CORREÇÃO] Removendo {len(arquivos_para_remover)} arquivos SCR que não correspondem aos itens processados...")
                        # os já está importado no topo do arquivo
                        for arquivo_remover in arquivos_para_remover:
                            caminho_arquivo = os.path.join(pasta_pavimento_verificar, arquivo_remover)
                            try:
                                os.remove(caminho_arquivo)
                                print(f">>>   ✅ Removido: {arquivo_remover}")
                            except Exception as e:
                                print(f">>>   ❌ Erro ao remover {arquivo_remover}: {str(e)}")
                        
                        print(f">>> ✅ Limpeza concluída: {len(arquivos_validos)} arquivos válidos mantidos")
                    else:
                        print(f">>> ✅ Todos os arquivos SCR correspondem aos itens processados")
                    
                    # Log final
                    arquivos_scr_apos_limpeza = [f for f in os.listdir(pasta_pavimento_verificar) if f.endswith('.scr')]
                    print(f"\n>>> ===== RESUMO FINAL DE ARQUIVOS SCR ===== ")
                    print(f">>> Total de arquivos SCR após limpeza: {len(arquivos_scr_apos_limpeza)}")
                    print(f">>> Arquivos válidos identificados: {len(arquivos_validos)}")
                    print(f">>> Esperados: {len(itens)} itens processados")
                    print(f">>> Lista completa de arquivos SCR na pasta:")
                    for i, arquivo in enumerate(sorted(arquivos_scr_apos_limpeza), 1):
                        status = "✅ VÁLIDO" if arquivo in arquivos_validos else "❌ INVÁLIDO (deveria ter sido removido)"
                        print(f">>>   [{i}] {arquivo} - {status}")
                    
                    # SEMPRE FAZER LIMPEZA FINAL: Mesmo se parecer correto, garantir que apenas arquivos válidos estejam presentes
                    arquivos_para_manter = set(arquivos_validos)
                    print(f">>> [LIMPEZA FINAL OBRIGATÓRIA] Arquivos que devem ser mantidos: {list(arquivos_para_manter)}")
                    
                    arquivos_removidos_final = 0
                    for arquivo_na_pasta in arquivos_scr_apos_limpeza:
                        if arquivo_na_pasta not in arquivos_para_manter:
                            caminho_remover = os.path.join(pasta_pavimento_verificar, arquivo_na_pasta)
                            try:
                                os.remove(caminho_remover)
                                print(f">>>   🗑️ Removido arquivo extra: {arquivo_na_pasta}")
                                arquivos_removidos_final += 1
                            except Exception as e:
                                print(f">>>   ❌ Erro ao remover {arquivo_na_pasta}: {e}")
                                # Tentar com delay
                                try:
                                    time.sleep(0.2)  # time já está importado no topo do arquivo
                                    os.remove(caminho_remover)
                                    arquivos_removidos_final += 1
                                    print(f">>>   ✅ Removido após retry: {arquivo_na_pasta}")
                                except Exception as e2:
                                    print(f">>>   ❌ Falha definitiva ao remover {arquivo_na_pasta}: {e2}")
                    
                    if arquivos_removidos_final > 0:
                        print(f">>> ✅ [LIMPEZA FINAL] Removidos {arquivos_removidos_final} arquivos extras")
                    else:
                        print(f">>> ✅ [LIMPEZA FINAL] Nenhum arquivo extra encontrado")
                    
                    # Verificar novamente após limpeza final
                    arquivos_finais = [f for f in os.listdir(pasta_pavimento_verificar) if f.endswith('.scr')]
                    print(f">>> [LIMPEZA FINAL] Arquivos SCR após limpeza final: {len(arquivos_finais)}")
                    
                    if len(arquivos_finais) != len(itens):
                        print(f"\n>>> ⚠️ PROBLEMA: {len(arquivos_finais)} arquivos encontrados, mas apenas {len(itens)} itens foram processados")
                        print(f">>> Isso indica que há arquivos extras sendo gerados ou não estão sendo removidos corretamente")
                        print(f">>> Lista de arquivos finais:")
                        for arq_final in sorted(arquivos_finais):
                            status = "✅ VÁLIDO" if arq_final in arquivos_para_manter else "❌ INVÁLIDO"
                            print(f">>>   - {arq_final} - {status}")
                    else:
                        print(f"\n>>> ✅ CORRETO: {len(arquivos_scr_apos_limpeza)} arquivos correspondem aos {len(itens)} itens processados")
                    print(f">>> ===== FIM DO RESUMO =====\n")
                
                print(f">>> Executando combinador...")
                
                # CORREÇÃO: Executar combinador via importação direta (compatível com frozen)
                # Em vez de usar subprocess, importar e chamar a função diretamente
                is_frozen = getattr(sys, 'frozen', False)
                if not is_frozen:
                    # Detecção alternativa para Nuitka
                    if hasattr(sys, 'executable') and sys.executable and sys.executable.endswith('.exe'):
                        exe_dir = os.path.dirname(sys.executable)
                        if '.dist' in exe_dir or os.path.basename(exe_dir) in ['run.dist', 'dist', 'dist_nuitka', 'dist_debug']:
                            is_frozen = True
                
                if is_frozen:
                    # Ambiente frozen: importar e executar diretamente
                    print(f">>> [COMBI] Ambiente frozen detectado - executando via importação direta")
                    try:
                        import importlib
                        
                        # Tentar múltiplos caminhos de importação
                        module = None
                        import_paths = [
                            'robots.Combinador_de_SCR _cima',
                            'src.robots.Combinador_de_SCR _cima',
                            'Combinador_de_SCR _cima'
                        ]
                        
                        for import_path in import_paths:
                            try:
                                module = importlib.import_module(import_path)
                                print(f">>> [OK] Módulo combinador importado: {import_path}")
                                break
                            except ImportError:
                                continue
                        
                        if not module:
                            raise ImportError("Não foi possível importar o módulo combinador")
                        
                        # Chamar a função processar_arquivos diretamente
                        if hasattr(module, 'processar_arquivos'):
                            print(f">>> [COMBI] Executando processar_arquivos na pasta: {pasta_pavimento}")
                            module.processar_arquivos(pasta_pavimento)
                            print(f">>> ✅ Combinador executado com sucesso (frozen)")
                        else:
                            raise AttributeError("Módulo combinador não possui função processar_arquivos")
                            
                    except Exception as e:
                        print(f">>> ❌ ERRO ao executar combinador via importação: {e}")
                        import traceback
                        traceback.print_exc()
                        # Fallback: tentar método antigo
                        print(f">>> [COMBI] Tentando método antigo como fallback...")
                        try:
                            self.conector_painel.combinar_codigos_cima()
                        except Exception:
                            pass
                else:
                    # Ambiente de desenvolvimento: usar subprocess como fallback
                    combinador_path = os.path.join(robo_dir, "src", "robots", "Combinador_de_SCR _cima.py")
                    
                    print(f">>> [COMBI] Verificando combinador...")
                    print(f">>> [COMBI] robo_dir: {robo_dir}")
                    print(f">>> [COMBI] combinador_path: {combinador_path}")
                    print(f">>> [COMBI] combinador existe: {os.path.exists(combinador_path)}")
                    
                    if os.path.exists(combinador_path):
                        import subprocess
                        import sys
                        print(f">>> [COMBI] Executando combinador via subprocess na pasta: {pasta_pavimento}")
                        print(f">>> [COMBI] Comando: {sys.executable} {combinador_path} {pasta_pavimento}")
                        try:
                            resultado = subprocess.run(
                                [sys.executable, combinador_path, pasta_pavimento], 
                                text=True,
                                capture_output=True,
                                timeout=300
                            )
                            if resultado.stdout:
                                print(f">>> [COMBI] stdout: {resultado.stdout[-500:]}")
                            if resultado.stderr:
                                print(f">>> [COMBI] stderr: {resultado.stderr[-500:]}")
                            
                            if resultado.returncode == 0:
                                print(f">>> ✅ Combinador executado com sucesso")
                            else:
                                print(f">>> ⚠️ Combinador retornou código {resultado.returncode}")
                        except Exception as e:
                            print(f">>> ❌ ERRO ao executar combinador: {e}")
                            import traceback
                            traceback.print_exc()
                            # Fallback: tentar método antigo
                            print(f">>> [COMBI] Tentando método antigo como fallback...")
                            self.conector_painel.combinar_codigos_cima()
                    else:
                        print(f">>> ⚠️ Caminho do combinador não encontrado: {combinador_path}")
                        # Fallback: tentar método antigo
                        print(f">>> [COMBI] Tentando método antigo como fallback...")
                        self.conector_painel.combinar_codigos_cima()
                
                # Executar o ordenador automaticamente (usando pasta_pavimento já normalizada)
                print(f">>> Executando ordenador...")
                print(f">>> [ORDENADOR] Usando pasta_pavimento normalizada: {pasta_pavimento}")
                combinados_dir_normalizado = os.path.join(pasta_pavimento, "Combinados")
                print(f">>> [ORDENADOR] pasta_combinados: {combinados_dir_normalizado}")

                # SINCRONIZAÇÃO: Garantir que arquivos *_COMBINADO_CIMA.scr estejam em Combinados
                try:
                    os.makedirs(combinados_dir_normalizado, exist_ok=True)
                    import re
                    existentes = [f for f in os.listdir(combinados_dir_normalizado) if f.endswith('.scr')]
                    ha_numericos = any(re.match(r"^\d+\.scr$", f) for f in existentes)
                    if ha_numericos:
                        print(f">>> [SYNC] Detectados arquivos numerados em Combinados; não copiar nomeados para evitar duplicidade")
                    else:
                        arquivos_pav = [f for f in os.listdir(pasta_pavimento) if f.endswith('_COMBINADO_CIMA.scr')]
                        copiados = 0
                        for f in arquivos_pav:
                            origem = os.path.join(pasta_pavimento, f)
                            destino = os.path.join(combinados_dir_normalizado, f)
                            if not os.path.exists(destino):
                                import shutil
                                shutil.copy2(origem, destino)
                                copiados += 1
                        if copiados:
                            print(f">>> [SYNC] Copiados {copiados} arquivos COMBINADO para Combinados")
                        else:
                            print(f">>> [SYNC] Nenhum arquivo COMBINADO novo para copiar")
                except Exception as e:
                    print(f">>> [SYNC] Erro ao sincronizar combinados: {e}")
                self._executar_ordenador_cima_automatico(combinados_dir_normalizado)
                print(f">>> ✅ Ordenador executado com sucesso")
                
            except Exception as e:
                print(f">>> ERRO ao processar itens: {str(e)}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Erro", f"Erro ao processar itens do pavimento: {str(e)}")
                return
            
            # Verificar se a pasta Combinados foi criada
            if not os.path.exists(combinados_dir):
                messagebox.showerror("Erro", f"Pasta Combinados não foi criada em:\n{combinados_dir}\n\nVerifique se o processamento foi executado corretamente.")
                return

            script_final = self._processar_scripts_combinados(combinados_dir)
            if not script_final:
                return
            
            # ========================================================
            # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
            # ========================================================
            self._limpar_linhas_em_branco_entre_semicolons(combinados_dir)
            
            # ========================================================
            # 🎯 ATUALIZAR SCRIPT_PAZ.scr COM O PRIMEIRO SCRIPT DA PASTA COMBINADOS
            # ========================================================
            script_1_path = os.path.join(combinados_dir, "1.scr")
            if os.path.exists(script_1_path):
                print(f">>> [PAZ] Atualizando script_PAZ.scr com o primeiro script da pasta Combinados (CIMA)")
                self._atualizar_script_paz(script_1_path)
                print(f">>> [PAZ] ✅ Script PAZ atualizado com 1.scr da pasta Combinados")
            else:
                print(f">>> [PAZ] ⚠️ Arquivo 1.scr não encontrado na pasta Combinados, pulando atualização do PAZ")

            print(f">>> Scripts combinados processados")

            # Perguntar se deseja executar no AutoCAD
            opcao = self._perguntar_execucao_autocad("desenho do pavimento CIMA")
            
            if opcao is None:  # Cancelado
                return
            
            if opcao == 1:
                # Opção 1: Desenhar no CAD diretamente
                self._executar_no_autocad(script_final, pavimento, "CIMA")
                print(f"✅ Desenho do pavimento CIMA executado com sucesso!")
            elif opcao == 2:
                # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                print(f"✅ Scripts gerados (não executados no AutoCAD)")
            elif opcao == 3:
                # Opção 3: Gerar DFX (em desenvolvimento)
                messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                return

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD pavimento CIMA: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar D.CAD pavimento CIMA: {str(e)}")
        finally:
            # Sempre liberar os locks ao final
            self.processing_lock = False
            self.dcad_lock = False

    def _limpar_linhas_em_branco_entre_semicolons(self, combinados_dir):
        """
        Remove linhas em branco entre dois ';' nas últimas 10 linhas dos scripts combinados.
        Compatível com ambiente frozen (PyInstaller).
        
        Args:
            combinados_dir: Caminho do diretório Combinados contendo os scripts .scr
        """
        print(f">>> [LIMPEZA FINAL] Verificando e removendo linhas em branco entre ';' nas últimas 10 linhas...")
        try:
            # Verificar se o diretório existe (compatível com frozen e desenvolvimento)
            if not os.path.exists(combinados_dir):
                print(f">>> [LIMPEZA FINAL] ⚠️ Diretório Combinados não encontrado: {combinados_dir}")
                print(f">>> [LIMPEZA FINAL] Pulando limpeza final...")
                return
            
            # Listar todos os arquivos .scr na pasta Combinados
            arquivos_scr = [f for f in os.listdir(combinados_dir) if f.endswith('.scr')]
            arquivos_processados = 0
            
            for arquivo_scr in arquivos_scr:
                caminho_arquivo = os.path.join(combinados_dir, arquivo_scr)
                
                try:
                    # Ler o arquivo
                    with open(caminho_arquivo, 'r', encoding='utf-16-le') as f:
                        linhas = f.readlines()
                    
                    # Verificar se tem pelo menos 10 linhas
                    if len(linhas) < 10:
                        # Se tiver menos de 10 linhas, verificar todas
                        linhas_verificar = linhas
                    else:
                        # Pegar apenas as últimas 10 linhas
                        linhas_verificar = linhas[-10:]
                        linhas_anteriores = linhas[:-10]
                    
                    # Verificar se há linhas em branco entre dois ";"
                    linhas_limpas = []
                    modificado = False
                    
                    for i, linha in enumerate(linhas_verificar):
                        linha_stripped = linha.rstrip('\n\r')
                        
                        # Se a linha anterior termina com ";" e esta linha está em branco e a próxima também termina com ";"
                        if i > 0 and i < len(linhas_verificar) - 1:
                            linha_anterior = linhas_verificar[i-1].rstrip('\n\r')
                            linha_proxima = linhas_verificar[i+1].rstrip('\n\r')
                            
                            # Verificar se: linha anterior termina com ";" E linha atual está vazia E linha próxima termina com ";"
                            # Usar strip() para remover espaços em branco antes de verificar
                            linha_anterior_limpa = linha_anterior.strip()
                            linha_atual_limpa = linha_stripped.strip()
                            linha_proxima_limpa = linha_proxima.strip()
                            
                            if (linha_anterior_limpa.endswith(';') and 
                                linha_atual_limpa == '' and 
                                linha_proxima_limpa.endswith(';')):
                                # Pular esta linha (não adicionar)
                                modificado = True
                                num_linha = len(linhas) - len(linhas_verificar) + i + 1
                                print(f">>>   [LIMPEZA] Removida linha em branco entre ';' no arquivo {arquivo_scr} (linha {num_linha})")
                                continue
                        
                        # Adicionar a linha normalmente
                        linhas_limpas.append(linha)
                    
                    # Se houve modificação, reescrever o arquivo
                    if modificado:
                        # Reconstruir todas as linhas
                        if len(linhas) >= 10:
                            linhas_finais = linhas_anteriores + linhas_limpas
                        else:
                            linhas_finais = linhas_limpas
                        
                        # Escrever de volta
                        with open(caminho_arquivo, 'w', encoding='utf-16-le') as f:
                            f.writelines(linhas_finais)
                        
                        arquivos_processados += 1
                        print(f">>>   [LIMPEZA] ✅ Arquivo {arquivo_scr} limpo com sucesso")
                
                except Exception as e:
                    print(f">>>   [LIMPEZA] ⚠️ Erro ao processar {arquivo_scr}: {e}")
                    continue
            
            # Verificação final após processar todos os arquivos
            if arquivos_processados > 0:
                print(f">>> [LIMPEZA FINAL] ✅ {arquivos_processados} arquivo(s) processado(s) e limpo(s)")
            else:
                print(f">>> [LIMPEZA FINAL] ℹ️ Nenhuma linha em branco encontrada entre ';' nos arquivos")
        
        except Exception as e:
            print(f">>> [LIMPEZA FINAL] ⚠️ Erro ao executar limpeza final: {e}")
            import traceback
            traceback.print_exc()
            # Não interromper o processo por causa deste erro
    
    def _executar_ordenador_cima_automatico(self, pasta_combinados):
        """Executa o ordenador CIMA automaticamente na pasta combinados"""
        try:
            import sys
            print(f">>> [ORDENADOR] Executando ordenador CIMA automaticamente...")
            print(f">>> [ORDENADOR] Pasta: {pasta_combinados}")
            
            if not os.path.exists(pasta_combinados):
                print(f">>> [ORDENADOR] ERRO: Pasta combinados não existe: {pasta_combinados}")
                return False
            
            # Adicionar diretório dos robôs ao path
            robots_dir = os.path.join(self._get_project_root(), 'src', 'robots')
            if robots_dir not in sys.path:
                sys.path.insert(0, robots_dir)
            
            # Importar ordenador CIMA
            from Ordenador_Cordenadas_cima import ProcessadorCoordenadasCima, atualizar_comando_pilar_cima
            from pathlib import Path
            import json
            
            # Carregar configurações
            cfg_path = Path('configuracao_ordenador_CIMA.json')
            if cfg_path.exists():
                with open(cfg_path, 'r') as f:
                    cfg = json.load(f)
                print(f">>> [ORDENADOR] Configurações carregadas: {cfg}")
            else:
                # Configurações padrão
                cfg = {
                    "numero_colunas": 4,
                    "distancia_x_colunas": 1585,
                    "distancia_y_linhas": -1148.6,
                    "distancia_y_extra": 0,
                    "linhas_para_extra": 0
                }
                print(f">>> [ORDENADOR] Usando configurações padrão: {cfg}")
            
            # Criar processador com as configurações
            processador = ProcessadorCoordenadasCima(
                numero_colunas=cfg.get("numero_colunas", 4),
                distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                distancia_y_extra=cfg.get("distancia_y_extra", 0),
                linhas_para_extra=cfg.get("linhas_para_extra", 0)
            )
            
            # Processar arquivos
            arquivos_scr = [f for f in os.listdir(pasta_combinados) if f.endswith('.scr')]
            # Se houver numerados, ignorar nomeados (usar apenas 1.scr, 2.scr ...)
            import re
            if any(re.match(r"^\d+\.scr$", f) for f in arquivos_scr):
                arquivos_scr = [f for f in arquivos_scr if re.match(r"^\d+\.scr$", f)]

            # Ordenação alfabética consistente por item com regras:
            # 1) Números puros primeiro (1,4,6,11...)
            # 2) Depois prefixos por letra com sufixo numérico em ordem natural (A1,A2,A3,A17, depois P3,P7...)
            # 3) Demais nomes alfabéticos
            import unicodedata
            def _ordenador_key(nome_arquivo: str):
                base = nome_arquivo
                # Remover sufixos de tipo
                if base.endswith('_COMBINADO_CIMA.scr'):
                    base = base[:-len('_COMBINADO_CIMA.scr')]
                elif base.endswith('_CIMA.scr'):
                    base = base[:-len('_CIMA.scr')]
                elif base.endswith('.scr'):
                    base = base[:-len('.scr')]
                # Remover versões como -1, -2
                base = base.split('-')[0]
                # Normalizar acentos
                base_norm = unicodedata.normalize('NFD', base)
                base_norm = ''.join(ch for ch in base_norm if unicodedata.category(ch) != 'Mn')
                # 1) Números puros
                if base_norm.isdigit():
                    return (0, int(base_norm), '')
                # 2) Letras seguidas de números (ex.: A17, P3)
                import re
                m = re.match(r'^([A-Za-z]+)(\d+)$', base_norm)
                if m:
                    prefix = m.group(1).upper()
                    num = int(m.group(2))
                    grupo = 1 if prefix == 'A' else (2 if prefix == 'P' else 3)
                    return (grupo, prefix, num)
                # 3) Alfabético puro
                return (4, base_norm.upper(), 0)

            arquivos_scr = sorted(arquivos_scr, key=_ordenador_key)
            print(f">>> [ORDENADOR] Arquivos encontrados: {len(arquivos_scr)}")
            for i, arq in enumerate(arquivos_scr, 1):
                print(f">>> [ORDENADOR] Ordem {i}: {arq}")

            # Fallback: se não houver arquivos em Combinados, copiar do diretório pai
            if len(arquivos_scr) == 0:
                try:
                    pasta_pai = os.path.dirname(pasta_combinados)
                    candidatos = [f for f in os.listdir(pasta_pai) if f.endswith('.scr')]
                    # Preferir combinados, senão todos
                    preferidos = [f for f in candidatos if f.endswith('_COMBINADO_CIMA.scr')]
                    copiar = preferidos if preferidos else candidatos
                    import shutil
                    copiados = 0
                    for f in copiar:
                        origem = os.path.join(pasta_pai, f)
                        destino = os.path.join(pasta_combinados, f)
                        if not os.path.exists(destino):
                            shutil.copy2(origem, destino)
                            copiados += 1
                    if copiados:
                        print(f">>> [ORDENADOR] Fallback: copiados {copiados} arquivos para Combinados")
                        arquivos_scr = sorted([f for f in os.listdir(pasta_combinados) if f.endswith('.scr')])
                        print(f">>> [ORDENADOR] Arquivos após fallback: {len(arquivos_scr)}")
                    else:
                        print(f">>> [ORDENADOR] Fallback: nenhum arquivo para copiar")
                except Exception as e:
                    print(f">>> [ORDENADOR] Fallback falhou: {e}")
            
            for i, arq in enumerate(arquivos_scr):
                caminho_completo = os.path.join(pasta_combinados, arq)
                print(f">>> [ORDENADOR] Processando arquivo {i+1}/{len(arquivos_scr)}: {arq}")
                processador.processar_arquivo(caminho_completo)
            
            # Atualizar comando pilar
            atualizar_comando_pilar_cima(pasta_combinados)
            print(f">>> [ORDENADOR] ✅ Ordenador executado com sucesso!")
            return True
            
        except Exception as e:
            print(f">>> [ORDENADOR] ERRO ao executar ordenador: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def _executar_ordenador_grades_automatico(self, pasta_combinados):
        """Executa o ordenador GRADES automaticamente na pasta combinados"""
        try:
            import sys
            print(f">>> [ORDENADOR] Executando ordenador GRADES automaticamente...")
            print(f">>> [ORDENADOR] Pasta: {pasta_combinados}")
            
            if not os.path.exists(pasta_combinados):
                print(f">>> [ORDENADOR] ERRO: Pasta combinados não existe: {pasta_combinados}")
                return False
            
            # Adicionar diretório dos robôs ao path
            robots_dir = os.path.join(self._get_project_root(), 'src', 'robots')
            if robots_dir not in sys.path:
                sys.path.insert(0, robots_dir)
            
            # Importar ordenador GRADES
            from Ordenador_Cordenadas_grades import ProcessadorCoordenadasGrade, atualizar_comando_pilar_grade
            from pathlib import Path
            import json
            
            # Carregar configurações
            cfg_path = Path('configuracao_ordenador_GRADES.json')
            if cfg_path.exists():
                with open(cfg_path, 'r') as f:
                    cfg = json.load(f)
                print(f">>> [ORDENADOR] Configurações carregadas: {cfg}")
            else:
                # Configurações padrão
                cfg = {
                    "numero_colunas": 4,
                    "distancia_x_colunas": 1585,
                    "distancia_y_linhas": -1148.6,
                    "distancia_y_extra": 0,
                    "linhas_para_extra": 0
                }
                print(f">>> [ORDENADOR] Usando configurações padrão: {cfg}")
            
            # Criar processador com as configurações
            processador = ProcessadorCoordenadasGrade(
                numero_colunas=cfg.get("numero_colunas", 4),
                distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                distancia_y_extra=cfg.get("distancia_y_extra", 0),
                linhas_para_extra=cfg.get("linhas_para_extra", 0)
            )
            
            # Processar arquivos
            arquivos_scr = [f for f in os.listdir(pasta_combinados) if f.endswith('.scr')]
            # Se houver numerados, ignorar nomeados (usar apenas 1.scr, 2.scr ...)
            import re
            if any(re.match(r"^\d+\.scr$", f) for f in arquivos_scr):
                arquivos_scr = [f for f in arquivos_scr if re.match(r"^\d+\.scr$", f)]

            # Ordenação alfabética consistente por item com regras:
            # 1) Números puros primeiro (1,4,6,11...)
            # 2) Depois prefixos por letra com sufixo numérico em ordem natural (A1,A2,A3,A17, depois P3,P7...)
            # 3) Demais nomes alfabéticos
            import unicodedata
            def _ordenador_key(nome_arquivo: str):
                base = nome_arquivo
                # Remover sufixos de tipo
                if base.endswith('_GRADES.scr'):
                    base = base[:-len('_GRADES.scr')]
                elif base.endswith('.scr'):
                    base = base[:-len('.scr')]
                # Remover versões como -1, -2
                base = base.split('-')[0]
                # Normalizar acentos
                base_norm = unicodedata.normalize('NFD', base)
                base_norm = ''.join(ch for ch in base_norm if unicodedata.category(ch) != 'Mn')
                # 1) Números puros
                if base_norm.isdigit():
                    return (0, int(base_norm), '')
                # 2) Letras seguidas de números (ex.: A17, P3)
                m = re.match(r'^([A-Za-z]+)(\d+)$', base_norm)
                if m:
                    prefix = m.group(1).upper()
                    num = int(m.group(2))
                    grupo = 1 if prefix == 'A' else (2 if prefix == 'P' else 3)
                    return (grupo, prefix, num)
                # 3) Alfabético puro
                return (4, base_norm.upper(), 0)

            arquivos_scr = sorted(arquivos_scr, key=_ordenador_key)
            print(f">>> [ORDENADOR] Arquivos encontrados: {len(arquivos_scr)}")
            for i, arq in enumerate(arquivos_scr, 1):
                print(f">>> [ORDENADOR] Ordem {i}: {arq}")

            # Processar cada arquivo
            for i, arq in enumerate(arquivos_scr, 1):
                print(f">>> [ORDENADOR] Processando arquivo {i}/{len(arquivos_scr)}: {arq}")
                processador.processar_arquivo(os.path.join(pasta_combinados, arq))
            
            # Atualizar comando pilar
            atualizar_comando_pilar_grade(pasta_combinados)
            print(f">>> ✅ Ordenador GRADES executado com sucesso!")
            return True
            
        except Exception as e:
            print(f">>> [ORDENADOR] ERRO ao executar ordenador GRADES: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _dcad_item_abcd(self):
        """Executa o desenho de um item no CAD para visão ABCD"""
        try:
            print(f"\n>>> EXECUTANDO D.CAD ITEM ABCD")
            print(f"\n>>> EXECUTANDO D.CAD ITEM ABCD")
            
            # Obter informações do item selecionado (com verificação segura)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            # Verificar se o sistema de créditos está disponível
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="ABCD"
                )
                
                # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                    return
                
                print(_get_obf_str("credit"))
                
                # Atualizar interface imediatamente
                self.atualizar_creditos_interface()

            # LIMPEZA ANTES DE INICIAR: Limpar pasta do pavimento para evitar resquícios
            print(f">>> LIMPANDO PASTA DO PAVIMENTO ANTES DE INICIAR...")
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            if os.path.exists(pasta_pavimento):
                print(f">>> Pasta encontrada, removendo conteúdo...")
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                    os.makedirs(pasta_pavimento, exist_ok=True)
                    print(f">>> ✅ Pasta '{pavimento}_ABCD' limpa e recriada com sucesso!")
                except Exception as e:
                    print(f">>> ❌ Erro ao limpar pasta: {e}")
                    messagebox.showwarning("Aviso", f"Erro ao limpar pasta do pavimento: {e}")
            else:
                print(f">>> Pasta não existe, criando nova pasta...")
                os.makedirs(pasta_pavimento, exist_ok=True)
                print(f">>> ✅ Nova pasta '{pavimento}_ABCD' criada com sucesso!")

            # Primeiro, executar o desenho normal para gerar o script
            if hasattr(self, 'conector_painel'):
                self.conector_painel.executar_abcd_excel()
            else:
                messagebox.showwarning("Aviso", "Conector não disponível")
                return

            # Obter o caminho do script gerado
            robo_dir = self._get_project_root()
            
            # Obter informações do item selecionado
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Construir o caminho do script
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            script_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Caminho do script: {script_path}")

            # VERIFICAR SE É PILAR ESPECIAL
            pilar_especial_ativo = dados.get("pilar_especial", {}).get("ativar_pilar_especial", False)
            print(f">>> Pilar especial ativo: {pilar_especial_ativo}")
            
            if pilar_especial_ativo:
                # PILAR ESPECIAL: Ler e concatenar 2 scripts
                print(f"\n[PILAR_ESPECIAL] ==========================================")
                print(f"[PILAR_ESPECIAL] DETECTADO PILAR ESPECIAL - CONCATENANDO 2 SCRIPTS")
                print(f"[PILAR_ESPECIAL] ==========================================\n")
                
                script1_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")
                script2_path = os.path.join(pasta_pavimento, f"{nome}_ABCD2.scr")
                
                print(f"[PILAR_ESPECIAL] Script 1: {script1_path}")
                print(f"[PILAR_ESPECIAL] Script 2: {script2_path}")
                
                # Verificar se ambos os scripts existem
                if not os.path.exists(script1_path):
                    print(f"[PILAR_ESPECIAL] ERRO: Script 1 não encontrado!")
                    messagebox.showerror("Erro", f"Script 1 não encontrado em:\n{script1_path}")
                    return
                    
                if not os.path.exists(script2_path):
                    print(f"[PILAR_ESPECIAL] ERRO: Script 2 não encontrado!")
                    messagebox.showerror("Erro", f"Script 2 não encontrado em:\n{script2_path}")
                    return
                
                # Ler e filtrar ambos os scripts
                print(f"[PILAR_ESPECIAL] Lendo script 1: {script1_path}")
                script1_content = self._ler_e_filtrar_script(script1_path)
                print(f"[PILAR_ESPECIAL] Script 1 lido: {len(script1_content) if script1_content else 0} caracteres")
                
                print(f"[PILAR_ESPECIAL] Lendo script 2: {script2_path}")
                script2_content = self._ler_e_filtrar_script(script2_path)
                print(f"[PILAR_ESPECIAL] Script 2 lido: {len(script2_content) if script2_content else 0} caracteres")
                
                if not script1_content or not script2_content:
                    print(f"[PILAR_ESPECIAL] ERRO: Falha ao ler/filtrar scripts")
                    print(f"[PILAR_ESPECIAL] Script 1 válido: {bool(script1_content)}")
                    print(f"[PILAR_ESPECIAL] Script 2 válido: {bool(script2_content)}")
                    return
                
                # Concatenar os scripts
                script_content = script1_content + "\n" + script2_content
                
                print(f"[PILAR_ESPECIAL] Scripts concatenados com sucesso!")
                print(f"[PILAR_ESPECIAL] Tamanho Script 1: {len(script1_content)} chars")
                print(f"[PILAR_ESPECIAL] Tamanho Script 2: {len(script2_content)} chars")
                print(f"[PILAR_ESPECIAL] Tamanho Total: {len(script_content)} chars")
                print(f"[PILAR_ESPECIAL] ==========================================\n")
                
                # ========================================================
                # 🎯 ATUALIZAR SCRIPT_PAZ.scr E ADICIONAR COMANDOS SCRIPT
                # ========================================================
                # PILAR ESPECIAL: 2 scripts (1 e 2)
                print(f">>> [PAZ] Pilar especial detectado - processando 2 scripts")
                
                # Atualizar PAZ com o primeiro script
                self._atualizar_script_paz(script1_path)
                
                # Adicionar comando SCRIPT no final do script 1 para executar o script 2
                self._adicionar_comando_script_ao_final(script1_path, script2_path)
                
                print(f">>> [PAZ] ✅ Scripts conectados: 1 -> 2")
                
            else:
                # COMPORTAMENTO NORMAL: Ler 1 script apenas
                if not os.path.exists(script_path):
                    print(f">>> ERRO: Script não encontrado!")
                    print(f">>> Verificando se a pasta existe: {os.path.exists(pasta_pavimento)}")
                    
                    if os.path.exists(pasta_pavimento):
                        print(f">>> Conteúdo da pasta {pasta_pavimento}:")
                        try:
                            for item in os.listdir(pasta_pavimento):
                                print(f">>>   - {item}")
                        except Exception as e:
                            print(f">>> Erro ao listar pasta: {e}")
                    
                    messagebox.showerror("Erro", f"Script não encontrado em:\n{script_path}\n\nVerifique se o item foi processado corretamente.")
                    return

                print(f">>> Script encontrado com sucesso!")

                # Ler e filtrar o script
                script_content = self._ler_e_filtrar_script(script_path)
                if not script_content:
                    print(f">>> ERRO: Falha ao ler/filtrar script")
                    return

                # ========================================================
                # 🎯 ATUALIZAR SCRIPT_PAZ.scr
                # ========================================================
                # PILAR COMUM: 1 script só
                print(f">>> [PAZ] Pilar comum detectado - processando 1 script")
                
                # Atualizar PAZ com o único script
                self._atualizar_script_paz(script_path)
                
                print(f">>> [PAZ] ✅ Script PAZ atualizado com script único")

            print(f">>> Script processado com sucesso")

            # Perguntar se deseja executar no AutoCAD
            opcao = self._perguntar_execucao_autocad("desenho visão ABCD")
            
            if opcao is None:  # Cancelado
                return
            
            if opcao == 1:
                # Opção 1: Desenhar no CAD diretamente
                self._executar_no_autocad(script_content, pavimento, "ABCD")
                print(f"✅ Desenho visão ABCD executado com sucesso!")
            elif opcao == 2:
                # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                print(f"✅ Scripts gerados (não executados no AutoCAD)")
            elif opcao == 3:
                # Opção 3: Gerar DFX (em desenvolvimento)
                messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                return

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD item ABCD: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar D.CAD item ABCD: {str(e)}")

    def _processar_item_abcd_para_pavimento(self, numero, dados):
        """
        Processa um item individual para o pavimento ABCD (mesma lógica do _dcad_item_abcd)
        Retorna o conteúdo do script processado ou None em caso de erro
        """
        try:
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Processando item: {nome} (Pavimento: {pavimento})")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido para item {nome}")
                return None
            
            # CARREGAR DADOS DO PILAR ESPECIAL NA INTERFACE PRINCIPAL ANTES DE CRIAR EXCEL
            # (mesma lógica do _dcad_item_abcd quando o item está selecionado)
            interface_principal = getattr(self.conector_painel, 'main_app', None)
            if interface_principal and hasattr(interface_principal, '_carregar_dados_pilar_especial'):
                try:
                    print(f">>> Carregando dados do pilar especial na interface principal para item {nome}...")
                    interface_principal._carregar_dados_pilar_especial(dados)
                    print(f">>> ✅ Dados do pilar especial carregados na interface principal")
                except Exception as e:
                    print(f">>> ⚠️ Aviso: Não foi possível carregar dados do pilar especial na interface principal: {e}")
            
            # Criar Excel temporário com apenas este item
            excel_temp = self.conector_painel.criar_excel_temporario([(numero, dados)])
            if not excel_temp:
                print(f">>> ERRO: Falha ao criar Excel temporário para item {nome}")
                return None
            
            try:
                # Coluna sempre inicia em 'E' para o primeiro item
                from openpyxl.utils import get_column_letter
                coluna_letra = get_column_letter(5)  # Coluna 5 = 'E'
                
                # Executar script ABCD usando função auxiliar compatível com frozen
                # Usar interface_principal já obtida acima (mesma referência usada em executar_abcd_excel)
                resultado = self.conector_painel._executar_excel_wrapper(
                    "abcd_excel", 
                    excel_temp, 
                    coluna_letra, 
                    pavimento, 
                    interface_principal=interface_principal,
                    gerar_pelo_pavimento=True
                )
                
                if not resultado or (hasattr(resultado, 'returncode') and resultado.returncode != 0):
                    print(f">>> ERRO: Falha ao gerar script para item {nome}")
                    return None
                
                # Aguardar um pouco para garantir que o script foi gerado
                time.sleep(0.5)  # Pequena espera para garantir que o arquivo foi criado
                
                # Obter o caminho do script gerado
                robo_dir = self._get_project_root()
                pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
                script_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")
                
                print(f">>> Procurando script: {script_path}")
                
                # VERIFICAR SE É PILAR ESPECIAL
                pilar_especial_ativo = dados.get("pilar_especial", {}).get("ativar_pilar_especial", False)
                print(f">>> Pilar especial ativo: {pilar_especial_ativo}")
                
                if pilar_especial_ativo:
                    # PILAR ESPECIAL: Verificar se os 2 scripts foram gerados (NÃO concatenar aqui)
                    # Os scripts devem ficar separados para o combinador processar individualmente
                    print(f"\n[PILAR_ESPECIAL] ==========================================")
                    print(f"[PILAR_ESPECIAL] DETECTADO PILAR ESPECIAL - VERIFICANDO 2 SCRIPTS")
                    print(f"[PILAR_ESPECIAL] Item: {nome}")
                    print(f"[PILAR_ESPECIAL] ==========================================\n")
                    
                    script1_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")
                    script2_path = os.path.join(pasta_pavimento, f"{nome}_ABCD2.scr")
                    
                    print(f"[PILAR_ESPECIAL] Script 1: {script1_path}")
                    print(f"[PILAR_ESPECIAL] Script 2: {script2_path}")
                    
                    # Verificar se ambos os scripts existem
                    if not os.path.exists(script1_path):
                        print(f"[PILAR_ESPECIAL] ERRO: Script 1 não encontrado!")
                        return None
                        
                    if not os.path.exists(script2_path):
                        print(f"[PILAR_ESPECIAL] ERRO: Script 2 não encontrado!")
                        return None
                    
                    # Verificar tamanho dos arquivos
                    tamanho1 = os.path.getsize(script1_path)
                    tamanho2 = os.path.getsize(script2_path)
                    
                    print(f"[PILAR_ESPECIAL] ✅ Scripts gerados com sucesso!")
                    print(f"[PILAR_ESPECIAL] Tamanho Script 1: {tamanho1} bytes")
                    print(f"[PILAR_ESPECIAL] Tamanho Script 2: {tamanho2} bytes")
                    print(f"[PILAR_ESPECIAL] Os scripts serão processados separadamente pelo combinador")
                    print(f"[PILAR_ESPECIAL] ==========================================\n")
                    
                    # Retornar uma tupla indicando que são 2 scripts separados
                    # O combinador vai pegar os arquivos diretamente da pasta
                    return (nome, True)  # True indica pilar especial com 2 scripts
                    
                else:
                    # COMPORTAMENTO NORMAL: Verificar se o script foi gerado
                    if not os.path.exists(script_path):
                        print(f">>> ERRO: Script não encontrado para item {nome}!")
                        print(f">>> Caminho esperado: {script_path}")
                        return None
                    
                    # Verificar tamanho do arquivo
                    tamanho = os.path.getsize(script_path)
                    print(f">>> ✅ Script encontrado com sucesso para item {nome}!")
                    print(f">>> Tamanho: {tamanho} bytes")
                    
                    # Retornar uma tupla indicando que é 1 script normal
                    return (nome, False)  # False indica pilar comum com 1 script
                    
            finally:
                # Limpar arquivo temporário
                if os.path.exists(excel_temp):
                    try:
                        os.remove(excel_temp)
                    except Exception as e:
                        print(f">>> Aviso: Não foi possível remover Excel temporário: {e}")
                        
        except Exception as e:
            print(f">>> ERRO ao processar item {numero}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _dcad_pavimento_abcd(self):
        """Executa o desenho do pavimento no CAD para visão ABCD - REFORMULADO"""
        # Proteção contra execução simultânea mais robusta
        if hasattr(self, 'dcad_lock') and self.dcad_lock:
            print(">>> Processamento D.CAD ABCD já em andamento. Aguarde...")
            return
        
        # Usar lock específico para D.CAD
        self.dcad_lock = True
        self.processing_lock = True
        
        try:
            print(f"\n>>> EXECUTANDO D.CAD PAVIMENTO ABCD (REFORMULADO)")
            print(f"\n>>> EXECUTANDO D.CAD PAVIMENTO ABCD (REFORMULADO)")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                messagebox.showwarning("Aviso", "Conector não disponível")
                return

            # Obter informações do pavimento selecionado (usar item selecionado como referência)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                messagebox.showerror("Erro", "Nenhum item selecionado. Selecione um item na lista para usar como referência do pavimento.")
                return
            
            pavimento = dados.get("pavimento", "")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Pavimento não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            # Obter todos os itens do pavimento primeiro (para contar créditos)
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            print(f">>> Total de itens encontrados para pavimento '{pavimento}': {len(itens)}")
            
            if not itens:
                messagebox.showwarning("Aviso", f"Nenhum item encontrado para o pavimento {pavimento}")
                return

            # LOG DETALHADO: Listar todos os itens encontrados
            print(f">>> [DEBUG] Itens encontrados:")
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                pavimento_item = dados_item.get("pavimento", "N/A")
                pilar_especial = dados_item.get("pilar_especial", {}).get("ativar_pilar_especial", False)
                print(f">>>   [{idx}] Número: {numero_item}, Nome: {nome_item}, Pavimento: {pavimento_item}, Especial: {pilar_especial}")
            
            # Verificar se o sistema de créditos está disponível
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                itens_dados = []
                for numero_item, dados_item in itens:
                    itens_dados.append(dados_item)
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_total = len(itens)
                quantidade_comuns = quantidade_total - quantidade_especiais
                
                # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    quantidade_total,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área total do pavimento
                area_total = credit_manager.calcular_area_pavimento(itens_dados)
                
                # Obter obra (do primeiro item)
                obra = itens_dados[0].get('obra', '') if itens_dados else ''
                
                print(f">>> Quantidade de itens no pavimento: {quantidade_total} ({quantidade_comuns} comuns, {quantidade_especiais} especiais)")
                print(_get_obf_str("credit"))
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento,
                    tipo="",  # Vazio para pavimento completo
                    area_m2=area_total,
                    parte_desenho=f"PAVIMENTO COMPLETO {quantidade_total} ITENS"
                )
                
                # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                    return
                
                print(_get_obf_str("credit"))
                
                # Atualizar interface imediatamente
                self.atualizar_creditos_interface()

            # LIMPEZA ANTES DE INICIAR: Limpar pasta do pavimento para evitar resquícios
            print(f">>> LIMPANDO PASTA DO PAVIMENTO ANTES DE INICIAR...")
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            if os.path.exists(pasta_pavimento):
                print(f">>> Pasta encontrada, removendo conteúdo...")
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                    os.makedirs(pasta_pavimento, exist_ok=True)
                    print(f">>> ✅ Pasta '{pavimento}_ABCD' limpa e recriada com sucesso!")
                except Exception as e:
                    print(f">>> ❌ Erro ao limpar pasta: {e}")
                    messagebox.showwarning("Aviso", f"Erro ao limpar pasta do pavimento: {e}")
            else:
                print(f">>> Pasta não existe, criando nova pasta...")
                os.makedirs(pasta_pavimento, exist_ok=True)
                print(f">>> ✅ Nova pasta '{pavimento}_ABCD' criada com sucesso!")

            # ========================================================
            # PROCESSAR CADA ITEM INDIVIDUALMENTE (MESMA LÓGICA DO D.CAD 1 ITEM)
            # ========================================================
            
            itens_processados = 0
            itens_com_erro = []
            
            print(f"\n>>> INICIANDO PROCESSAMENTO DE {len(itens)} ITENS...")
            
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                print(f"\n>>> [{idx}/{len(itens)}] Processando item: {nome_item}")
                
                # Processar item individualmente (mesma lógica do _dcad_item_abcd)
                # Retorna (nome, is_especial) ou None em caso de erro
                resultado = self._processar_item_abcd_para_pavimento(numero_item, dados_item)
                
                if resultado:
                    nome, is_especial = resultado
                    itens_processados += 1
                    if is_especial:
                        print(f">>> ✅ Item {nome_item} processado com sucesso! (2 scripts gerados)")
                    else:
                        print(f">>> ✅ Item {nome_item} processado com sucesso! (1 script gerado)")
                else:
                    itens_com_erro.append(nome_item)
                    print(f">>> ❌ ERRO ao processar item {nome_item}")
            
            print(f"\n>>> RESUMO DO PROCESSAMENTO:")
            print(f">>>   Total de itens: {len(itens)}")
            print(f">>>   Itens processados com sucesso: {itens_processados}")
            print(f">>>   Itens com erro: {len(itens_com_erro)}")
            
            if itens_com_erro:
                print(f">>>   Itens com erro: {', '.join(itens_com_erro)}")
            
            if itens_processados == 0:
                messagebox.showerror("Erro", "Nenhum item foi processado com sucesso. Verifique os logs para mais detalhes.")
                return

            if len(itens_com_erro) > 0:
                resposta = messagebox.askyesno(
                    "Aviso", 
                    f"{len(itens_com_erro)} item(ns) não foram processados com sucesso:\n{', '.join(itens_com_erro)}\n\nDeseja continuar com os itens processados?"
                )
                if not resposta:
                    return
            
            # ========================================================
            # EXECUTAR COMBINADOR E ORDENADOR
            # ========================================================
            
            print(f"\n>>> EXECUTANDO COMBINADOR...")
            print(f">>>   Scripts foram gerados na pasta do pavimento")
            print(f">>>   O combinador irá processar todos os arquivos .scr encontrados")
            
            # Executar combinador (cria pasta Combinados com scripts numerados)
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            
            # Executar combinador de forma compatível com frozen
            combinador_path = os.path.join(robo_dir, 'src', 'robots', "Combinador_de_SCR.py")
            is_frozen = getattr(self.conector_painel, '_is_frozen', lambda: False)()
            
            if is_frozen:
                # Em modo frozen: importar diretamente via importlib usando caminho do arquivo
                print(f">>> [COMBI] Ambiente frozen detectado - tentando importar combinador ABCD")
                try:
                    import importlib.util
                    import sys
                    
                    # Método 1: Tentar importar pelo caminho do arquivo
                    module = None
                    caminhos_possiveis = [
                        combinador_path,  # Caminho padrão
                        os.path.join(robo_dir, 'robots', "Combinador_de_SCR.py"),  # Alternativo
                        os.path.join(os.path.dirname(sys.executable) if hasattr(sys, 'executable') else robo_dir, 'src', 'robots', "Combinador_de_SCR.py"),  # Caminho do executável
                    ]
                    
                    for caminho_teste in caminhos_possiveis:
                        if os.path.exists(caminho_teste):
                            try:
                                print(f">>> [COMBI] Tentando importar pelo caminho do arquivo: {caminho_teste}")
                                spec = importlib.util.spec_from_file_location("Combinador_de_SCR_ABCD", caminho_teste)
                                if spec and spec.loader:
                                    module = importlib.util.module_from_spec(spec)
                                    sys.modules["Combinador_de_SCR_ABCD"] = module
                                    spec.loader.exec_module(module)
                                    print(f">>> [COMBI] ✅ Combinador importado via caminho do arquivo")
                                    break
                            except Exception as e:
                                print(f">>> [COMBI] ❌ Erro ao importar pelo caminho {caminho_teste}: {e}")
                                continue
                        else:
                            print(f">>> [COMBI] Caminho não existe: {caminho_teste}")
                    
                    if module is None:
                        print(f">>> [COMBI] ⚠️ Nenhum caminho físico encontrado, tentando importação por nome...")
                    
                    # Método 2: Tentar importar de múltiplos caminhos possíveis (fallback)
                    # Em modo frozen, os módulos são compilados e embutidos, então tentamos importar pelo nome
                    if module is None:
                        import importlib
                        # Tentar diferentes formas de importar o módulo
                        import_paths = [
                            'src.robots.Combinador_de_SCR',  # Caminho completo
                            'robots.Combinador_de_SCR',  # Caminho relativo
                            'Combinador_de_SCR',  # Nome direto
                            'automacaoexcel.Ordenamento.Combinador_de_SCR'  # Caminho usado pelo PyInstaller
                        ]
                        
                        print(f">>> [COMBI] Tentando importar de {len(import_paths)} caminhos possíveis...")
                        for import_path in import_paths:
                            try:
                                print(f">>> [COMBI] Tentando: {import_path}")
                                module = importlib.import_module(import_path)
                                print(f">>> [COMBI] ✅ Combinador importado via: {import_path}")
                                break
                            except ImportError as e:
                                print(f">>> [COMBI] ❌ Falha ao importar {import_path}: {e}")
                                continue
                            except Exception as e:
                                print(f">>> [COMBI] ❌ Erro inesperado ao importar {import_path}: {e}")
                                continue
                    
                    # Método 3: Tentar importar usando import direto (pode funcionar em frozen)
                    if module is None:
                        try:
                            # Tentar importar diretamente do pacote robots
                            from robots import Combinador_de_SCR as comb_module
                            module = comb_module
                            print(f">>> [COMBI] ✅ Combinador importado via 'from robots import'")
                        except (ImportError, SyntaxError):
                            try:
                                from src.robots import Combinador_de_SCR as comb_module
                                module = comb_module
                                print(f">>> [COMBI] ✅ Combinador importado via 'from src.robots import'")
                            except (ImportError, SyntaxError):
                                pass
                    
                    if module is None:
                        # Não levantar exceção - apenas avisar e continuar
                        print(f">>> [COMBI] ⚠️ Não foi possível importar módulo do combinador - continuando sem ele")
                        print(f">>> [COMBI] Os scripts individuais foram gerados, mas não serão combinados")
                        # Criar pasta Combinados manualmente e copiar scripts
                        pasta_combinados = os.path.join(pasta_pavimento, "Combinados")
                        os.makedirs(pasta_combinados, exist_ok=True)
                        # Copiar scripts individuais para Combinados
                        import shutil
                        scripts_gerados = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                        for i, script in enumerate(sorted(scripts_gerados), 1):
                            origem = os.path.join(pasta_pavimento, script)
                            destino = os.path.join(pasta_combinados, f"{i}.scr")
                            try:
                                shutil.copy2(origem, destino)
                                print(f">>> [COMBI] Script copiado: {script} -> {i}.scr")
                            except Exception as e:
                                print(f">>> [COMBI] Erro ao copiar {script}: {e}")
                        print(f">>> [COMBI] Scripts copiados manualmente para pasta Combinados")
                        module = None  # Não executar processar_arquivos
                    
                    if module is not None and hasattr(module, 'processar_arquivos'):
                        print(f">>> [COMBI] Executando processar_arquivos na pasta: {pasta_pavimento}")
                        # Garantir que a pasta existe
                        os.makedirs(pasta_pavimento, exist_ok=True)
                        try:
                            module.processar_arquivos(pasta_pavimento, mostrar_mensagem=False)
                            print(f">>> ✅ Combinador executado com sucesso (frozen)!")
                        except Exception as e:
                            print(f">>> [COMBI] ⚠️ Erro ao executar processar_arquivos: {e}")
                            # Criar pasta Combinados manualmente como fallback
                            pasta_combinados = os.path.join(pasta_pavimento, "Combinados")
                            os.makedirs(pasta_combinados, exist_ok=True)
                            import shutil
                            scripts_gerados = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                            for i, script in enumerate(sorted(scripts_gerados), 1):
                                origem = os.path.join(pasta_pavimento, script)
                                destino = os.path.join(pasta_combinados, f"{i}.scr")
                                try:
                                    shutil.copy2(origem, destino)
                                    print(f">>> [COMBI] Script copiado: {script} -> {i}.scr")
                                except Exception as e2:
                                    print(f">>> [COMBI] Erro ao copiar {script}: {e2}")
                            print(f">>> [COMBI] Scripts copiados manualmente para pasta Combinados (fallback)")
                    elif module is None:
                        # Já foi tratado acima - scripts foram copiados manualmente
                        pass
                    else:
                        print(f">>> [COMBI] ⚠️ Módulo importado mas não possui função processar_arquivos")
                        # Criar pasta Combinados manualmente como fallback
                        pasta_combinados = os.path.join(pasta_pavimento, "Combinados")
                        os.makedirs(pasta_combinados, exist_ok=True)
                        import shutil
                        scripts_gerados = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                        for i, script in enumerate(sorted(scripts_gerados), 1):
                            origem = os.path.join(pasta_pavimento, script)
                            destino = os.path.join(pasta_combinados, f"{i}.scr")
                            try:
                                shutil.copy2(origem, destino)
                                print(f">>> [COMBI] Script copiado: {script} -> {i}.scr")
                            except Exception as e2:
                                print(f">>> [COMBI] Erro ao copiar {script}: {e2}")
                        print(f">>> [COMBI] Scripts copiados manualmente para pasta Combinados (fallback)")
                        
                except Exception as e:
                    print(f">>> ERRO ao executar combinador: {e}")
                    import traceback
                    traceback.print_exc()
                    # Não retornar aqui - continuar mesmo se o combinador falhar
                    print(f">>> [COMBI] ⚠️ Continuando sem combinador (pode causar problemas)")
            else:
                # Em modo dev: usar subprocess
                if os.path.exists(combinador_path):
                    import subprocess
                    import sys
                    resultado = subprocess.run(
                        [sys.executable, combinador_path, pasta_pavimento], 
                        capture_output=True, 
                        text=True, 
                        encoding='utf-8', 
                        errors='ignore'
                    )
                    if resultado.returncode == 0:
                        print(f">>> ✅ Combinador executado com sucesso!")
                    else:
                        print(f">>> ERRO ao executar combinador: {resultado.stderr}")
                else:
                    print(f">>> ERRO: Arquivo combinador não encontrado: {combinador_path}")
            
            # Executar ordenador na pasta Combinados
            print(f"\n>>> EXECUTANDO ORDENADOR NA PASTA COMBINADOS...")
            pasta_combinados = os.path.join(pasta_pavimento, 'Combinados')
            
            if os.path.exists(pasta_combinados):
                try:
                    # Importar ordenador ABCD (compatível com frozen)
                    robots_path = os.path.join(robo_dir, 'src', 'robots')
                    if robots_path not in sys.path:
                        sys.path.insert(0, robots_path)
                    
                    # Tentar importação normal primeiro
                    try:
                        from Ordenador_Cordenadas_abcd import ProcessadorCoordenadasABCD, atualizar_comando_pilar, get_config_path_abcd
                    except ImportError:
                        # Fallback para modo frozen: importar pelo caminho do arquivo
                        if is_frozen:
                            import importlib.util
                            ordenador_path = os.path.join(robots_path, 'Ordenador_Cordenadas_abcd.py')
                            if os.path.exists(ordenador_path):
                                print(f">>> [ORDENADOR] Importando pelo caminho do arquivo (frozen)")
                                spec = importlib.util.spec_from_file_location("Ordenador_Cordenadas_abcd", ordenador_path)
                                if spec and spec.loader:
                                    ordenador_module = importlib.util.module_from_spec(spec)
                                    sys.modules["Ordenador_Cordenadas_abcd"] = ordenador_module
                                    spec.loader.exec_module(ordenador_module)
                                    ProcessadorCoordenadasABCD = ordenador_module.ProcessadorCoordenadasABCD
                                    atualizar_comando_pilar = ordenador_module.atualizar_comando_pilar
                                    get_config_path_abcd = ordenador_module.get_config_path_abcd
                                    print(f">>> [ORDENADOR] ✅ Ordenador importado com sucesso (frozen)")
                                else:
                                    raise ImportError("Não foi possível criar spec do módulo ordenador")
                            else:
                                raise ImportError(f"Arquivo ordenador não encontrado: {ordenador_path}")
                        else:
                            raise
                    
                    from pathlib import Path
                    import json
                    
                    # Carregar configurações
                    try:
                        cfg_path = get_config_path_abcd()
                        if cfg_path.exists():
                            with open(cfg_path, 'r', encoding='utf-8') as f:
                                cfg = json.load(f)
                        else:
                            cfg = {
                                "numero_colunas": 4,
                                "distancia_x_colunas": 1585,
                                "distancia_y_linhas": -1148.6,
                                "distancia_y_extra": 0,
                                "linhas_para_extra": 0
                            }
                    except Exception as e:
                        cfg = {
                            "numero_colunas": 4,
                            "distancia_x_colunas": 1585,
                            "distancia_y_linhas": -1148.6,
                            "distancia_y_extra": 0,
                            "linhas_para_extra": 0
                        }
                    
                    # Criar processador com as configurações
                    processador = ProcessadorCoordenadasABCD(
                        numero_colunas=cfg.get("numero_colunas", 4),
                        distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                        distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                        distancia_y_extra=cfg.get("distancia_y_extra", 0),
                        linhas_para_extra=cfg.get("linhas_para_extra", 0)
                    )
                    
                    # Processar arquivos na pasta Combinados
                    arquivos_scr = [f for f in os.listdir(pasta_combinados) if f.endswith('.scr')]
                    print(f">>> Arquivos encontrados na pasta Combinados: {len(arquivos_scr)}")
                    
                    for i, arq in enumerate(arquivos_scr, 1):
                        print(f">>> Ordenando arquivo {i}/{len(arquivos_scr)}: {arq}")
                        processador.processar_arquivo(os.path.join(pasta_combinados, arq))
                    
                    # Atualizar comando pilar
                    atualizar_comando_pilar(pasta_combinados)
                    print(f">>> ✅ Ordenador executado com sucesso!")
                    
                except Exception as e:
                    print(f">>> ERRO ao executar ordenador: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f">>> ERRO: Pasta Combinados não encontrada: {pasta_combinados}")
            
            # Executar no AutoCAD usando o script combinado da pasta Combinados
            print(f"\n>>> Executando scripts combinados no AutoCAD...")
            script_final = self._processar_scripts_combinados(pasta_combinados)
            if script_final:
                # ========================================================
                # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
                # ========================================================
                self._limpar_linhas_em_branco_entre_semicolons(pasta_combinados)
                
                # ========================================================
                # 🎯 ATUALIZAR SCRIPT_PAZ.scr COM O PRIMEIRO SCRIPT DA PASTA COMBINADOS
                # ========================================================
                script_1_path = os.path.join(pasta_combinados, "1.scr")
                if os.path.exists(script_1_path):
                    print(f">>> [PAZ] Atualizando script_PAZ.scr com o primeiro script da pasta Combinados (ABCD)")
                    self._atualizar_script_paz(script_1_path)
                    print(f">>> [PAZ] ✅ Script PAZ atualizado com 1.scr da pasta Combinados")
                else:
                    print(f">>> [PAZ] ⚠️ Arquivo 1.scr não encontrado na pasta Combinados, pulando atualização do PAZ")
                
                # Perguntar se deseja executar no AutoCAD
                opcao = self._perguntar_execucao_autocad("desenho do pavimento ABCD")
                
                if opcao is None:  # Cancelado
                    return
                
                if opcao == 1:
                    # Opção 1: Desenhar no CAD diretamente
                    self._executar_no_autocad(script_final, pavimento, "ABCD")
                    print(f"✅ Desenho visão ABCD do pavimento executado com sucesso!")
                elif opcao == 2:
                    # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                    print(f"✅ Scripts gerados (não executados no AutoCAD)")
                elif opcao == 3:
                    # Opção 3: Gerar DFX (em desenvolvimento)
                    messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                    return
            else:
                print(f">>> ERRO: Não foi possível processar scripts combinados")
                messagebox.showerror("Erro", "Não foi possível processar scripts combinados da pasta Combinados")
            print(f">>> Total de itens desenhados: {itens_processados}")

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD pavimento ABCD: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar D.CAD pavimento ABCD: {str(e)}")
        finally:
            # Sempre liberar os locks ao final
            self.processing_lock = False
            self.dcad_lock = False
            print(">>> [LOCK] Locks liberados após execução D.CAD pavimento ABCD")

    def _dcad_item_grades(self):
        """Executa o desenho de um item no CAD para GRADES"""
        print(f"\n>>> ========================================")
        print(f">>> [DEBUG GRADES] _dcad_item_grades INICIADO")
        print(f">>> ========================================")
        print(f">>> [DEBUG GRADES]   - processing_lock: {getattr(self, 'processing_lock', False)}")
        print(f">>> [DEBUG GRADES]   - dcad_lock: {getattr(self, 'dcad_lock', False)}")
        
        try:
            print(f"\n>>> EXECUTANDO D.CAD ITEM GRADES")
            print(f"\n>>> EXECUTANDO D.CAD ITEM GRADES")
            
            # Obter informações do item selecionado (com verificação segura)
            print(f">>> [DEBUG GRADES] Obtendo item selecionado...")
            numero, dados = self._get_item_selecionado_safe()
            print(f">>> [DEBUG GRADES]   - Número retornado: {numero}")
            print(f">>> [DEBUG GRADES]   - Dados retornados: {bool(dados)}")
            if dados:
                print(f">>> [DEBUG GRADES]   - Chaves dos dados: {list(dados.keys())[:10]}...")
            
            if not dados:
                print(f">>> [DEBUG GRADES] ❌ Nenhum item selecionado")
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> [DEBUG GRADES] ✅ Item selecionado encontrado")
            print(f">>> [DEBUG GRADES]   - Número: {numero}")
            print(f">>> [DEBUG GRADES]   - Nome: {nome}")
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            print(f">>> [DEBUG GRADES]   - Pavimento extraído: {pavimento}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            # Verificar se o sistema de créditos está disponível
            print(f">>> [DEBUG GRADES] Verificando sistema de créditos...")
            credit_manager = getattr(self, 'credit_manager', None)
            print(_get_obf_str("credit"))
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
                print(f">>> [DEBUG GRADES] ⚠️ Sistema de créditos não disponível")
            else:
                print(f">>> [DEBUG GRADES] ✅ Sistema de créditos disponível")
                
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                print(f">>> [DEBUG GRADES] Calculando créditos necessários...")
                print(f">>> [DEBUG GRADES]   - Item é especial: {eh_especial}")
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                print(_get_obf_str("credit"))
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="GRADES"
                )
                
                # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                print(_get_obf_str("credit"))
                print(_get_obf_str("credit"))
                print(f">>> [DEBUG GRADES]   - Descrição: {descricao}")
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                print(f">>> [DEBUG GRADES] Resultado do débito: sucesso={sucesso_debito}, mensagem={mensagem}")
                
                if not sucesso_debito:
                    print(f">>> [DEBUG GRADES] ❌ Débito falhou - retornando")
                    messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                    return
                
                print(f">>> [DEBUG GRADES] ✅ Débito bem-sucedido")
                print(_get_obf_str("credit"))
                
                # Atualizar interface imediatamente
                print(f">>> [DEBUG GRADES] Atualizando interface de créditos...")
                self.atualizar_creditos_interface()
                print(f">>> [DEBUG GRADES] ✅ Interface atualizada")

            # Primeiro, executar o desenho normal para gerar o script
            print(f">>> [DEBUG GRADES] Executando executar_grade_excel...")
            if hasattr(self, 'conector_painel'):
                print(f">>> [DEBUG GRADES] ✅ Conector disponível, chamando executar_grade_excel...")
                self.conector_painel.executar_grade_excel()
                print(f">>> [DEBUG GRADES] ✅ executar_grade_excel concluído")
            else:
                print(f">>> [DEBUG GRADES] ❌ Conector não disponível")
                messagebox.showwarning("Aviso", "Conector não disponível")
                return

            # Obter o caminho do script gerado
            print(f">>> [DEBUG GRADES] Obtendo caminho do script gerado...")
            robo_dir = self._get_project_root()
            print(f">>> [DEBUG GRADES]   - robo_dir: {robo_dir}")
            
            # Obter informações do item selecionado
            print(f">>> [DEBUG GRADES] Obtendo item selecionado novamente...")
            numero, dados = self._get_item_selecionado_safe()
            print(f">>> [DEBUG GRADES]   - Número retornado: {numero}")
            print(f">>> [DEBUG GRADES]   - Dados retornados: {bool(dados)}")
            if not dados:
                print(f">>> [DEBUG GRADES] ❌ Nenhum item selecionado")
                print(f">>> ERRO: Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> [DEBUG GRADES] ✅ Item selecionado obtido novamente")
            print(f">>> [DEBUG GRADES]   - Número: {numero}")
            print(f">>> [DEBUG GRADES]   - Nome: {nome}")
            print(f">>> [DEBUG GRADES]   - Pavimento: {pavimento}")
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Construir o caminho do script
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
            print(f">>> [DEBUG GRADES]   - pasta_pavimento: {pasta_pavimento}")
            print(f">>> [DEBUG GRADES]   - pasta_pavimento existe: {os.path.exists(pasta_pavimento)}")
            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            
            # Para GRADES, verificar se estamos no modo "pilar especial" (tem arquivos E, F, G, H)
            # Os arquivos gerados têm sufixos .A.scr e .B.scr (normais) ou .A, .B, .E, .F, .G, .H.scr (especiais)
            script_path_a = os.path.join(pasta_pavimento, f"{nome}.A.scr")
            script_path_b = os.path.join(pasta_pavimento, f"{nome}.B.scr")
            script_path_e = os.path.join(pasta_pavimento, f"{nome}.E.scr")
            script_path_f = os.path.join(pasta_pavimento, f"{nome}.F.scr")
            script_path_g = os.path.join(pasta_pavimento, f"{nome}.G.scr")
            script_path_h = os.path.join(pasta_pavimento, f"{nome}.H.scr")
            
            print(f">>> [DEBUG GRADES] Procurando arquivos GRADES...")
            print(f">>> Procurando arquivos GRADES:")
            print(f">>> Arquivo A: {script_path_a}")
            print(f">>> Arquivo B: {script_path_b}")
            print(f">>> [DEBUG GRADES]   - Arquivo A: {script_path_a}")
            print(f">>> [DEBUG GRADES]   - Arquivo B: {script_path_b}")
            print(f">>> [DEBUG GRADES]   - Arquivo E: {script_path_e}")
            print(f">>> [DEBUG GRADES]   - Arquivo F: {script_path_f}")
            print(f">>> [DEBUG GRADES]   - Arquivo G: {script_path_g}")
            print(f">>> [DEBUG GRADES]   - Arquivo H: {script_path_h}")
            
            # Verificar se arquivos existem
            print(f">>> [DEBUG GRADES] Verificando existência dos arquivos...")
            arquivo_a_existe = os.path.exists(script_path_a)
            arquivo_b_existe = os.path.exists(script_path_b)
            arquivo_e_existe = os.path.exists(script_path_e)
            arquivo_f_existe = os.path.exists(script_path_f)
            arquivo_g_existe = os.path.exists(script_path_g)
            arquivo_h_existe = os.path.exists(script_path_h)
            print(f">>> [DEBUG GRADES]   - A existe: {arquivo_a_existe}")
            print(f">>> [DEBUG GRADES]   - B existe: {arquivo_b_existe}")
            print(f">>> [DEBUG GRADES]   - E existe: {arquivo_e_existe}")
            print(f">>> [DEBUG GRADES]   - F existe: {arquivo_f_existe}")
            print(f">>> [DEBUG GRADES]   - G existe: {arquivo_g_existe}")
            print(f">>> [DEBUG GRADES]   - H existe: {arquivo_h_existe}")
            
            # Verificar se estamos no modo "pilar especial" (existem arquivos E ou F)
            pilar_especial_ativo = arquivo_e_existe or arquivo_f_existe
            print(f">>> [DEBUG GRADES]   - Pilar especial ativo: {pilar_especial_ativo}")
            
            if pilar_especial_ativo:
                print(f">>> MODO PILAR ESPECIAL detectado!")
                print(f">>> Arquivo E existe: {arquivo_e_existe}")
                print(f">>> Arquivo F existe: {arquivo_f_existe}")
                print(f">>> Arquivo G existe: {arquivo_g_existe}")
                print(f">>> Arquivo H existe: {arquivo_h_existe}")
            
            print(f">>> Arquivo A existe: {arquivo_a_existe}")
            print(f">>> Arquivo B existe: {arquivo_b_existe}")
            
            # Se nenhum dos dois existe E não está no modo especial, tentar alternativas
            if not arquivo_a_existe and not arquivo_b_existe and not pilar_especial_ativo:
                # Tentar com sufixo _GRADES (como no CIMA)
                script_path_alternativo = os.path.join(pasta_pavimento, f"{nome}_GRADES.scr")
                print(f">>> Tentando caminho alternativo _GRADES: {script_path_alternativo}")
                if os.path.exists(script_path_alternativo):
                    print(f">>> Script encontrado no caminho alternativo _GRADES!")
                    script_content = self._ler_e_filtrar_script(script_path_alternativo)
                    if not script_content:
                        print(f">>> ERRO: Falha ao ler/filtrar script")
                        return
                else:
                    # Tentar listar todos os arquivos .scr na pasta para debug
                    print(f">>> Verificando todos os arquivos .scr na pasta:")
                    try:
                        for arquivo in os.listdir(pasta_pavimento):
                            if arquivo.endswith('.scr'):
                                print(f">>>   - {arquivo}")
                    except Exception as e:
                        print(f">>> Erro ao listar arquivos: {e}")
                    
                    print(f">>> ERRO: Nenhum script GRADES encontrado!")
                    print(f">>> Verificando se a pasta existe: {os.path.exists(pasta_pavimento)}")
                    
                    if os.path.exists(pasta_pavimento):
                        print(f">>> Conteúdo da pasta {pasta_pavimento}:")
                        try:
                            for item in os.listdir(pasta_pavimento):
                                print(f">>>   - {item}")
                        except Exception as e:
                            print(f">>> Erro ao listar pasta: {e}")
                    
                    messagebox.showerror("Erro", f"Script não encontrado em:\n{script_path_a}\n{script_path_b}\n\nVerifique se o item foi processado corretamente.")
                    return
            else:
                # Combinar os scripts (A e B normais, ou A, B, E, F, G, H especiais)
                script_content = ""
                
                # Lista de arquivos para combinar (ordem: A, B, E, F, G, H)
                arquivos_para_combinar = []
                if pilar_especial_ativo:
                    # Modo pilar especial: combinar todos os 6 arquivos
                    arquivos_para_combinar = [
                        ('A', script_path_a, arquivo_a_existe),
                        ('B', script_path_b, arquivo_b_existe),
                        ('E', script_path_e, arquivo_e_existe),
                        ('F', script_path_f, arquivo_f_existe),
                        ('G', script_path_g, arquivo_g_existe),
                        ('H', script_path_h, arquivo_h_existe),
                    ]
                    print(f">>> [PILAR_ESPECIAL] Combinando arquivos A, B, E, F, G, H...")
                else:
                    # Modo normal: combinar apenas A e B
                    arquivos_para_combinar = [
                        ('A', script_path_a, arquivo_a_existe),
                        ('B', script_path_b, arquivo_b_existe),
                    ]
                    print(f">>> [MODO_NORMAL] Combinando arquivos A e B...")
                
                # Ler e combinar cada arquivo
                for letra, caminho, existe in arquivos_para_combinar:
                    if existe:
                        print(f">>> Lendo arquivo {letra}: {caminho}")
                        script_conteudo = self._ler_e_filtrar_script(caminho)
                        if script_conteudo:
                            script_content += script_conteudo
                            print(f">>> Arquivo {letra} adicionado ({len(script_conteudo)} caracteres)")
                        else:
                            print(f">>> ERRO: Falha ao ler arquivo {letra}")
                    else:
                        print(f">>> Arquivo {letra} não existe, pulando...")
                
                if not script_content:
                    print(f">>> ERRO: Nenhum conteúdo válido encontrado nos arquivos GRADES")
                    return
                
                # ========================================================
                # 🎯 ATUALIZAR SCRIPT_PAZ.scr E ADICIONAR COMANDOS SCRIPT
                # ========================================================
                # Lista de scripts válidos (existentes) na ordem correta
                scripts_validos = []
                for letra, caminho, existe in arquivos_para_combinar:
                    if existe:
                        scripts_validos.append((letra, caminho))
                
                if scripts_validos:
                    print(f">>> [PAZ] Processando {len(scripts_validos)} scripts GRADES")
                    
                    # Atualizar PAZ com o primeiro script (A)
                    primeiro_script = scripts_validos[0][1]
                    self._atualizar_script_paz(primeiro_script)
                    print(f">>> [PAZ] ✅ Script PAZ atualizado com primeiro script: {scripts_validos[0][0]}")
                    
                    # Adicionar comandos SCRIPT no final de cada script (exceto o último)
                    for i in range(len(scripts_validos) - 1):
                        script_atual = scripts_validos[i][1]
                        proximo_script = scripts_validos[i + 1][1]
                        letra_atual = scripts_validos[i][0]
                        letra_proximo = scripts_validos[i + 1][0]
                        
                        # Adicionar comando SCRIPT no final do script atual
                        self._adicionar_comando_script_ao_final(script_atual, proximo_script)
                        print(f">>> [PAZ] ✅ Script {letra_atual} conectado ao script {letra_proximo}")
                else:
                    print(f">>> [PAZ] ⚠️ Nenhum script válido encontrado, pulando atualização do PAZ")

            print(f">>> Script GRADES combinado processado com sucesso! ({len(script_content)} caracteres)")

            print(f">>> Script processado com sucesso")

            # Perguntar se deseja executar no AutoCAD
            opcao = self._perguntar_execucao_autocad("desenho visão Grades")
            
            if opcao is None:  # Cancelado
                return
            
            if opcao == 1:
                # Opção 1: Desenhar no CAD diretamente
                self._executar_no_autocad(script_content, pavimento, "GRADES")
                print(f"✅ Desenho visão Grades executado com sucesso!")
            elif opcao == 2:
                # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                print(f"✅ Scripts gerados (não executados no AutoCAD)")
            elif opcao == 3:
                # Opção 3: Gerar DFX (em desenvolvimento)
                messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                return

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD item GRADES: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar D.CAD item GRADES: {str(e)}")

    def _processar_item_grades_para_pavimento(self, numero, dados):
        """
        Processa um item individual para o pavimento GRADES (mesma lógica do _dcad_item_grades)
        Retorna uma tupla (nome, is_especial) indicando se é pilar especial ou None em caso de erro
        """
        print(f"\n>>> ========================================")
        print(f">>> [DEBUG GRADES] _processar_item_grades_para_pavimento INICIADO")
        print(f">>> [DEBUG GRADES]   - Número recebido: {numero}")
        print(f">>> [DEBUG GRADES]   - Dados recebidos: {list(dados.keys())[:10] if dados else 'None'}...")
        print(f">>> ========================================")
        
        try:
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> [DEBUG GRADES] Processando item GRADES: {nome} (Pavimento: {pavimento})")
            print(f">>> [DEBUG GRADES]   - Número: {numero}")
            print(f">>> [DEBUG GRADES]   - Nome extraído: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido para item {nome}")
                return None
            
            # CARREGAR DADOS DO PILAR ESPECIAL NA INTERFACE PRINCIPAL ANTES DE CRIAR EXCEL
            # (mesma lógica do _dcad_item_grades quando o item está selecionado)
            interface_principal = getattr(self.conector_painel, 'main_app', None)
            
            # DEBUG: Verificar se o item tem dados de pilar especial
            pilar_especial_item = dados.get('pilar_especial', {})
            ativar_pilar_especial = pilar_especial_item.get('ativar_pilar_especial', False) if pilar_especial_item else False
            print(f"\n>>> [DEBUG GRADES] Verificando pilar especial para item {nome}:")
            print(f">>> [DEBUG GRADES]   - Dados pilar_especial encontrados: {bool(pilar_especial_item)}")
            print(f">>> [DEBUG GRADES]   - ativar_pilar_especial: {ativar_pilar_especial}")
            
            if interface_principal and hasattr(interface_principal, '_carregar_dados_pilar_especial'):
                try:
                    print(f">>> Carregando dados do pilar especial na interface principal para item {nome}...")
                    interface_principal._carregar_dados_pilar_especial(dados)
                    print(f">>> ✅ Dados do pilar especial carregados na interface principal")
                    
                    # DEBUG: Verificar se o checkbox foi ativado na interface
                    if hasattr(interface_principal, 'ativar_pilar_especial'):
                        checkbox_ativo = interface_principal.ativar_pilar_especial.get()
                        print(f">>> [DEBUG GRADES]   - Checkbox na interface principal após carregar: {checkbox_ativo}")
                except Exception as e:
                    print(f">>> ⚠️ Aviso: Não foi possível carregar dados do pilar especial na interface principal: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f">>> [DEBUG GRADES] ⚠️ Interface principal não disponível ou sem método _carregar_dados_pilar_especial")
            
            # Criar Excel temporário com apenas este item
            print(f">>> [DEBUG GRADES] Criando Excel temporário para item {nome}...")
            print(f">>> [DEBUG GRADES]   - Número: {numero}")
            print(f">>> [DEBUG GRADES]   - Dados: {list(dados.keys())[:10]}...")  # Primeiras 10 chaves
            
            excel_temp = self.conector_painel.criar_excel_temporario([(numero, dados)])
            
            print(f">>> [DEBUG GRADES] Excel temporário criado: {excel_temp}")
            print(f">>> [DEBUG GRADES] Excel existe: {os.path.exists(excel_temp) if excel_temp else False}")
            
            if not excel_temp:
                print(f">>> ERRO: Falha ao criar Excel temporário para item {nome}")
                return None
            
            if not os.path.exists(excel_temp):
                print(f">>> ERRO: Excel temporário criado mas arquivo não existe: {excel_temp}")
                return None
            
            try:
                # Coluna sempre inicia em 'E' para o primeiro item
                from openpyxl.utils import get_column_letter
                coluna_letra = get_column_letter(5)  # Coluna 5 = 'E'
                
                # Executar script GRADES usando função auxiliar compatível com frozen
                print(f">>> [DEBUG GRADES] Executando Excel wrapper para item {nome}")
                print(f">>> [DEBUG GRADES]   - Excel temporário: {excel_temp}")
                print(f">>> [DEBUG GRADES]   - Coluna: {coluna_letra}")
                print(f">>> [DEBUG GRADES]   - Pavimento: {pavimento}")
                
                resultado = self.conector_painel._executar_excel_wrapper(
                    "grades_excel", 
                    excel_temp, 
                    coluna_letra, 
                    pavimento
                )
                
                print(f">>> [DEBUG GRADES] Resultado do Excel wrapper para item {nome}: {resultado}")
                if hasattr(resultado, 'returncode'):
                    print(f">>> [DEBUG GRADES]   - Return code: {resultado.returncode}")
                if hasattr(resultado, 'stdout'):
                    print(f">>> [DEBUG GRADES]   - Stdout: {resultado.stdout[:200] if resultado.stdout else 'None'}")
                if hasattr(resultado, 'stderr'):
                    print(f">>> [DEBUG GRADES]   - Stderr: {resultado.stderr[:200] if resultado.stderr else 'None'}")
                
                if not resultado or (hasattr(resultado, 'returncode') and resultado.returncode != 0):
                    print(f">>> ERRO: Falha ao gerar script para item {nome}")
                    if hasattr(resultado, 'stderr') and resultado.stderr:
                        print(f">>> ERRO detalhado: {resultado.stderr}")
                    return None
                
                # Aguardar um pouco para garantir que o script foi gerado
                print(f">>> [DEBUG GRADES] Aguardando 0.5s para garantir que o script foi gerado para item {nome}...")
                time.sleep(0.5)  # Pequena espera para garantir que o arquivo foi criado
                
                # Obter o caminho do script gerado
                robo_dir = self._get_project_root()
                pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
                print(f">>> [DEBUG GRADES] Pasta do pavimento: {pasta_pavimento}")
                print(f">>> [DEBUG GRADES] Pasta existe: {os.path.exists(pasta_pavimento)}")
                
                # Para GRADES, verificar se estamos no modo "pilar especial" (tem arquivos E, F, G, H)
                script_path_a = os.path.join(pasta_pavimento, f"{nome}.A.scr")
                script_path_b = os.path.join(pasta_pavimento, f"{nome}.B.scr")
                script_path_e = os.path.join(pasta_pavimento, f"{nome}.E.scr")
                script_path_f = os.path.join(pasta_pavimento, f"{nome}.F.scr")
                script_path_g = os.path.join(pasta_pavimento, f"{nome}.G.scr")
                script_path_h = os.path.join(pasta_pavimento, f"{nome}.H.scr")
                
                print(f">>> [DEBUG GRADES] Verificando scripts gerados para item {nome}:")
                print(f">>> [DEBUG GRADES]   - {nome}.A.scr: {script_path_a}")
                print(f">>> [DEBUG GRADES]   - {nome}.B.scr: {script_path_b}")
                print(f">>> [DEBUG GRADES]   - {nome}.E.scr: {script_path_e}")
                print(f">>> [DEBUG GRADES]   - {nome}.F.scr: {script_path_f}")
                print(f">>> [DEBUG GRADES]   - {nome}.G.scr: {script_path_g}")
                print(f">>> [DEBUG GRADES]   - {nome}.H.scr: {script_path_h}")
                
                # DEBUG: Listar TODOS os arquivos .scr na pasta para verificar o que foi gerado
                print(f"\n>>> [DEBUG GRADES] Listando TODOS os arquivos .scr na pasta do pavimento:")
                try:
                    todos_arquivos = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                    print(f">>> [DEBUG GRADES] Total de arquivos .scr encontrados: {len(todos_arquivos)}")
                    for arq in sorted(todos_arquivos):
                        caminho_completo = os.path.join(pasta_pavimento, arq)
                        tamanho = os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                        print(f">>> [DEBUG GRADES]   - {arq} ({tamanho} bytes)")
                    
                    # Verificar especificamente se os arquivos do P1 foram gerados
                    arquivos_p1 = [f for f in todos_arquivos if f.startswith(f"{nome}.")]
                    print(f">>> [DEBUG GRADES] Arquivos encontrados para item {nome}: {arquivos_p1}")
                    if not arquivos_p1:
                        print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Nenhum arquivo encontrado para item {nome}!")
                except Exception as e:
                    print(f">>> [DEBUG GRADES] ERRO ao listar arquivos: {e}")
                    import traceback
                    traceback.print_exc()
                
                # Verificar se arquivos existem
                arquivo_a_existe = os.path.exists(script_path_a)
                arquivo_b_existe = os.path.exists(script_path_b)
                arquivo_e_existe = os.path.exists(script_path_e)
                arquivo_f_existe = os.path.exists(script_path_f)
                arquivo_g_existe = os.path.exists(script_path_g)
                arquivo_h_existe = os.path.exists(script_path_h)
                
                print(f"\n>>> [DEBUG] Verificação de scripts para item {nome}:")
                print(f">>> [DEBUG]   A: {arquivo_a_existe} - {script_path_a}")
                print(f">>> [DEBUG]   B: {arquivo_b_existe} - {script_path_b}")
                print(f">>> [DEBUG]   E: {arquivo_e_existe} - {script_path_e}")
                print(f">>> [DEBUG]   F: {arquivo_f_existe} - {script_path_f}")
                print(f">>> [DEBUG]   G: {arquivo_g_existe} - {script_path_g}")
                print(f">>> [DEBUG]   H: {arquivo_h_existe} - {script_path_h}")
                
                # Verificar se estamos no modo "pilar especial" (existem arquivos E ou F)
                pilar_especial_ativo = arquivo_e_existe or arquivo_f_existe
                
                if pilar_especial_ativo:
                    # PILAR ESPECIAL: Verificar se os 6 scripts foram gerados (A, B, E, F, G, H)
                    # Os scripts devem ficar separados para o combinador processar individualmente
                    print(f"\n[PILAR_ESPECIAL_GRADES] ==========================================")
                    print(f"[PILAR_ESPECIAL_GRADES] DETECTADO PILAR ESPECIAL - VERIFICANDO SCRIPTS")
                    print(f"[PILAR_ESPECIAL_GRADES] Item: {nome}")
                    print(f"[PILAR_ESPECIAL_GRADES] ==========================================\n")
                    
                    # Verificar quais scripts existem (pode ter 4, 5 ou 6 scripts)
                    scripts_necessarios = [
                        ('A', script_path_a, arquivo_a_existe),
                        ('B', script_path_b, arquivo_b_existe),
                        ('E', script_path_e, arquivo_e_existe),
                        ('F', script_path_f, arquivo_f_existe),
                        ('G', script_path_g, arquivo_g_existe),
                        ('H', script_path_h, arquivo_h_existe),
                    ]
                    
                    scripts_existentes = [letra for letra, path, existe in scripts_necessarios if existe]
                    scripts_faltando = [letra for letra, path, existe in scripts_necessarios if not existe]
                    
                    print(f"[PILAR_ESPECIAL_GRADES] Scripts existentes: {', '.join(scripts_existentes)}")
                    if scripts_faltando:
                        # Para scripts G e H, não é um erro - eles só são gerados se tiverem dados válidos
                        if all(letra in ['G', 'H'] for letra in scripts_faltando):
                            print(f"[PILAR_ESPECIAL_GRADES] ℹ️ Scripts {', '.join(scripts_faltando)} não foram gerados (sem dados válidos - normal)")
                        else:
                            print(f"[PILAR_ESPECIAL_GRADES] ⚠️ Scripts faltando: {', '.join(scripts_faltando)}")
                    
                    # Verificar se pelo menos A e B existem (mínimo necessário)
                    if not arquivo_a_existe or not arquivo_b_existe:
                        print(f"[PILAR_ESPECIAL_GRADES] ERRO: Scripts A e/ou B não encontrados!")
                        return None
                    
                    print(f"[PILAR_ESPECIAL_GRADES] ✅ Scripts do pilar especial verificados!")
                    print(f"[PILAR_ESPECIAL_GRADES] Total de scripts: {len(scripts_existentes)}")
                    print(f"[PILAR_ESPECIAL_GRADES] Os scripts serão processados separadamente pelo combinador")
                    print(f"[PILAR_ESPECIAL_GRADES] ==========================================\n")
                    
                    # Retornar uma tupla indicando que são múltiplos scripts separados
                    return (nome, True)  # True indica pilar especial com múltiplos scripts
                    
                else:
                    # COMPORTAMENTO NORMAL: Verificar se os scripts A e B foram gerados
                    if not arquivo_a_existe and not arquivo_b_existe:
                        print(f">>> ERRO: Scripts A e B não encontrados para item {nome}!")
                        print(f">>> Caminho esperado A: {script_path_a}")
                        print(f">>> Caminho esperado B: {script_path_b}")
                        return None
                    
                    # Verificar tamanho dos arquivos
                    tamanho_a = os.path.getsize(script_path_a) if arquivo_a_existe else 0
                    tamanho_b = os.path.getsize(script_path_b) if arquivo_b_existe else 0
                    
                    print(f">>> ✅ Scripts encontrados com sucesso para item {nome}!")
                    print(f">>>   A: {tamanho_a} bytes - {script_path_a}")
                    print(f">>>   B: {tamanho_b} bytes - {script_path_b}")
                    
                    # Retornar uma tupla indicando que são 2 scripts normais
                    return (nome, False)  # False indica pilar comum com 2 scripts (A e B)
                    
            finally:
                # Limpar arquivo temporário
                if os.path.exists(excel_temp):
                    try:
                        os.remove(excel_temp)
                    except Exception as e:
                        print(f">>> Aviso: Não foi possível remover Excel temporário: {e}")
                        
        except Exception as e:
            print(f">>> ERRO ao processar item GRADES {numero}: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _dcad_pavimento_grades(self):
        """Executa o desenho do pavimento no CAD para GRADES - REFORMULADO"""
        print(f"\n>>> ========================================")
        print(f">>> [DEBUG GRADES] _dcad_pavimento_grades INICIADO")
        print(f">>> ========================================")
        print(f">>> [DEBUG GRADES]   - dcad_lock: {getattr(self, 'dcad_lock', False)}")
        print(f">>> [DEBUG GRADES]   - processing_lock: {getattr(self, 'processing_lock', False)}")
        
        # Proteção contra execução simultânea mais robusta
        if hasattr(self, 'dcad_lock') and self.dcad_lock:
            print(">>> Processamento D.CAD GRADES já em andamento. Aguarde...")
            print(f">>> [DEBUG GRADES] ⚠️ D.CAD lock ativo - retornando")
            return
        
        # Usar lock específico para D.CAD
        self.dcad_lock = True
        self.processing_lock = True
        print(f">>> [DEBUG GRADES] ✅ Locks ativados (dcad_lock=True, processing_lock=True)")
        
        try:
            print(f"\n>>> EXECUTANDO D.CAD PAVIMENTO GRADES (REFORMULADO)")
            print(f"\n>>> EXECUTANDO D.CAD PAVIMENTO GRADES (REFORMULADO)")
            
            # Verificar se o conector está disponível
            print(f">>> [DEBUG GRADES] Verificando conector_painel...")
            if not hasattr(self, 'conector_painel'):
                print(f">>> [DEBUG GRADES] ❌ Conector não disponível")
                messagebox.showwarning("Aviso", "Conector não disponível")
                return
            print(f">>> [DEBUG GRADES] ✅ Conector disponível")

            # Obter informações do pavimento selecionado
            print(f">>> [DEBUG GRADES] Obtendo item selecionado...")
            numero, dados = self._get_item_selecionado_safe()
            print(f">>> [DEBUG GRADES]   - Número retornado: {numero}")
            print(f">>> [DEBUG GRADES]   - Dados retornados: {bool(dados)}")
            if dados:
                print(f">>> [DEBUG GRADES]   - Chaves dos dados: {list(dados.keys())[:10]}...")
            
            if not dados:
                print(f">>> [DEBUG GRADES] ❌ Nenhum item selecionado")
                messagebox.showerror("Erro", "Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            print(f">>> [DEBUG GRADES] ✅ Item selecionado encontrado")
            print(f">>> [DEBUG GRADES]   - Número: {numero}")
            print(f">>> [DEBUG GRADES]   - Nome: {dados.get('nome', 'N/A')}")
            print(f">>> Pavimento: {pavimento}")
            print(f">>> [DEBUG GRADES]   - Pavimento extraído: {pavimento}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Pavimento não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            # Obter todos os itens do pavimento
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            if not itens:
                messagebox.showwarning("Aviso", f"Nenhum pilar encontrado para o pavimento {pavimento}")
                return

            # DEBUG: Listar todos os itens encontrados
            print(f"\n>>> [DEBUG GRADES] Itens encontrados para pavimento '{pavimento}': {len(itens)}")
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                pavimento_item = dados_item.get("pavimento", "N/A")
                print(f">>> [DEBUG GRADES]   [{idx}] Número: {numero_item}, Nome: {nome_item}, Pavimento: {pavimento_item}")
            
            print(f">>> [DEBUG GRADES] ✅ Passou pela listagem de itens")
            
            # Verificar se o sistema de créditos está disponível
            credit_manager = getattr(self, 'credit_manager', None)
            print(_get_obf_str("credit"))
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                print(f">>> [DEBUG GRADES] ✅ Sistema de créditos disponível, calculando créditos...")
                try:
                    # Contar itens especiais e comuns
                    quantidade_especiais = 0
                    itens_dados = []
                    for numero_item, dados_item in itens:
                        itens_dados.append(dados_item)
                        if self._verificar_se_item_especial(dados_item):
                            quantidade_especiais += 1
                    
                    quantidade_total = len(itens)
                    quantidade_comuns = quantidade_total - quantidade_especiais
                    
                    # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                    print(_get_obf_str("credit"))
                    print(f">>> [DEBUG GRADES]   - Total de itens: {quantidade_total}")
                    print(f">>> [DEBUG GRADES]   - Itens comuns: {quantidade_comuns}")
                    print(f">>> [DEBUG GRADES]   - Itens especiais: {quantidade_especiais}")
                    creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                        quantidade_total,
                        "item_simples",
                        quantidade_especiais=quantidade_especiais
                    )
                    print(_get_obf_str("credit"))
                    
                    # Calcular área total do pavimento
                    area_total = credit_manager.calcular_area_pavimento(itens_dados)
                    
                    # Obter obra (do primeiro item)
                    obra = itens_dados[0].get('obra', '') if itens_dados else ''
                    
                    # Gerar descrição detalhada
                    descricao = credit_manager.gerar_descricao_detalhada(
                        obra=obra,
                        pavimento=pavimento,
                        tipo="",  # Vazio para pavimento completo
                        area_m2=area_total,
                        parte_desenho=f"PAVIMENTO COMPLETO {quantidade_total} ITENS"
                    )
                    
                    # DEBITAR CRÉDITOS IMEDIATAMENTE ANTES DE EXECUTAR
                    print(_get_obf_str("credit"))
                    print(_get_obf_str("credit"))
                    print(f">>> [DEBUG GRADES]   - Descrição: {descricao}")
                    
                    sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                        creditos_necessarios, 
                        descricao
                    )
                    
                    print(f">>> [DEBUG GRADES] Resultado do débito: sucesso={sucesso_debito}, mensagem={mensagem}")
                    
                    if not sucesso_debito:
                        print(f">>> [DEBUG GRADES] ❌ Débito falhou - retornando")
                        messagebox.showerror("Erro de Créditos", f"Não foi possível debitar créditos:\n{mensagem}")
                        return
                    
                    print(f">>> [DEBUG GRADES] ✅ Débito bem-sucedido - continuando processamento")
                    print(_get_obf_str("credit"))
                    
                    # Atualizar interface imediatamente
                    print(f">>> [DEBUG GRADES] Atualizando interface de créditos...")
                    self.atualizar_creditos_interface()
                    print(f">>> [DEBUG GRADES] Interface atualizada - continuando para limpar pasta")
                except Exception as e:
                    print(f">>> [DEBUG GRADES] ❌ EXCEÇÃO durante débito de créditos: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao debitar créditos:\n{str(e)}")
                    return
            
            # ========================================================
            # LIMPAR PASTA DO PAVIMENTO ANTES DE GERAR SCRIPTS
            # ========================================================
            print(f">>> [DEBUG GRADES] Iniciando limpeza da pasta do pavimento...")
            
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
            
            print(f">>> Limpando pasta do pavimento: {pasta_pavimento}")
            print(f">>> [DEBUG GRADES]   - Pasta existe: {os.path.exists(pasta_pavimento)}")
            
            if os.path.exists(pasta_pavimento):
                # Listar arquivos ANTES de remover para debug
                print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Limpando pasta do pavimento!")
                print(f">>> [DEBUG GRADES]   - Arquivos na pasta ANTES da limpeza:")
                try:
                    arquivos_antes = os.listdir(pasta_pavimento)
                    for arq in arquivos_antes:
                        caminho_arq = os.path.join(pasta_pavimento, arq)
                        if os.path.isfile(caminho_arq):
                            tamanho = os.path.getsize(caminho_arq)
                            print(f">>> [DEBUG GRADES]     - {arq} ({tamanho} bytes)")
                        else:
                            print(f">>> [DEBUG GRADES]     - {arq} (pasta)")
                except Exception as e:
                    print(f">>> [DEBUG GRADES]     - Erro ao listar: {e}")
                
                try:
                    import shutil
                    print(f">>> [DEBUG GRADES]   - Removendo pasta completa...")
                    shutil.rmtree(pasta_pavimento)
                    print(f">>> [DEBUG GRADES] ✅ Pasta removida com sucesso")
                except Exception as e:
                    print(f">>> Aviso: Não foi possível limpar pasta: {e}")
                    print(f">>> [DEBUG GRADES] ❌ Erro ao remover pasta: {e}")
            
            print(f">>> [DEBUG GRADES]   - Criando pasta...")
            os.makedirs(pasta_pavimento, exist_ok=True)
            print(f">>> ✅ Pasta do pavimento limpa e criada")
            print(f">>> [DEBUG GRADES] ✅ Pasta criada: {pasta_pavimento}")

            # ========================================================
            # PROCESSAR CADA ITEM INDIVIDUALMENTE (MESMA LÓGICA DO D.CAD 1 ITEM)
            # ========================================================
            
            itens_processados = 0
            itens_com_erro = []
            
            print(f"\n>>> INICIANDO PROCESSAMENTO DE {len(itens)} ITENS GRADES...")
            print(f">>> [DEBUG GRADES] Lista de itens recebida:")
            for idx, (num, dados) in enumerate(itens, 1):
                nome = dados.get("nome", num)
                print(f">>> [DEBUG GRADES]   [{idx}] Número: {num}, Nome: {nome}")
            
            print(f">>> [DEBUG GRADES] ✅ Total de itens na lista: {len(itens)}")
            print(f">>> [DEBUG GRADES] ✅ Iniciando loop de processamento...")
            
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                print(f"\n>>> ========================================")
                print(f">>> [{idx}/{len(itens)}] INICIANDO PROCESSAMENTO DO ITEM: {nome_item}")
                print(f">>> [DEBUG GRADES] ✅ Loop iterando sobre item {idx}/{len(itens)}: {nome_item}")
                print(f">>> ========================================")
                print(f">>> [DEBUG GRADES] Dados do item {nome_item}:")
                print(f">>> [DEBUG GRADES]   - Número: {numero_item}")
                print(f">>> [DEBUG GRADES]   - Nome: {dados_item.get('nome', 'N/A')}")
                print(f">>> [DEBUG GRADES]   - Pavimento: {dados_item.get('pavimento', 'N/A')}")
                print(f">>> [DEBUG GRADES]   - Pilar especial: {bool(dados_item.get('pilar_especial', {}))}")
                
                # Processar item individualmente (mesma lógica do _dcad_item_grades)
                print(f">>> [DEBUG GRADES] Chamando _processar_item_grades_para_pavimento para {nome_item}...")
                try:
                    resultado = self._processar_item_grades_para_pavimento(numero_item, dados_item)
                    print(f">>> [DEBUG GRADES] Resultado retornado para {nome_item}: {resultado}")
                    
                    if resultado:
                        itens_processados += 1
                        nome, is_especial = resultado
                        print(f">>> ✅ Item {nome_item} processado com sucesso (especial: {is_especial})")
                    else:
                        itens_com_erro.append(nome_item)
                        print(f">>> ❌ ERRO ao processar item {nome_item} - resultado None")
                        print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Item {nome_item} retornou None - verificar logs acima")
                except Exception as e:
                    itens_com_erro.append(nome_item)
                    print(f">>> ❌ EXCEÇÃO ao processar item {nome_item}: {str(e)}")
                    print(f">>> [DEBUG GRADES] ⚠️ EXCEÇÃO CAPTURADA para {nome_item}")
                    import traceback
                    traceback.print_exc()
            
            print(f"\n>>> PROCESSAMENTO CONCLUÍDO:")
            print(f">>>   - Itens processados com sucesso: {itens_processados}/{len(itens)}")
            if itens_com_erro:
                print(f">>>   - Itens com erro: {', '.join(itens_com_erro)}")
            
            if itens_processados == 0:
                messagebox.showerror("Erro", "Nenhum item foi processado com sucesso!")
                return

            # DEBUG: Listar TODOS os arquivos .scr gerados na pasta do pavimento
            print(f"\n>>> [DEBUG] VERIFICAÇÃO FINAL - Listando TODOS os scripts gerados:")
            try:
                todos_arquivos_scr = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
                print(f">>> [DEBUG] Total de arquivos .scr encontrados: {len(todos_arquivos_scr)}")
                for arq in sorted(todos_arquivos_scr):
                    caminho_completo = os.path.join(pasta_pavimento, arq)
                    tamanho = os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                    print(f">>> [DEBUG]   - {arq} ({tamanho} bytes)")
            except Exception as e:
                print(f">>> [DEBUG] ERRO ao listar arquivos finais: {e}")

            # ========================================================
            # EXECUTAR COMBINADOR
            # ========================================================
            
            print(f"\n>>> EXECUTANDO COMBINADOR GRADES...")
            
            # Executar combinador de forma compatível com frozen
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR_GRADES.py")
            is_frozen = self.conector_painel._is_frozen()
            
            if is_frozen:
                # Em modo frozen: importar diretamente via importlib (sem arquivo físico)
                try:
                    import importlib
                    
                    # Tentar importar de múltiplos caminhos possíveis
                    combinador_module = None
                    import_paths = [
                        'robots.Combinador_de_SCR_GRADES',
                        'src.robots.Combinador_de_SCR_GRADES',
                        'Combinador_de_SCR_GRADES'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            combinador_module = importlib.import_module(import_path)
                            print(f">>> [COMBI] Combinador importado via: {import_path}")
                            break
                        except ImportError:
                            continue
                    
                    if combinador_module is None:
                        raise ImportError("Não foi possível importar módulo do combinador de nenhum caminho")
                    
                    if hasattr(combinador_module, 'processar_arquivos'):
                        combinador_module.processar_arquivos(pasta_pavimento)
                        print(f">>> ✅ Combinador GRADES executado com sucesso (frozen)!")
                    else:
                        raise AttributeError("Módulo não possui função processar_arquivos")
                        
                except Exception as e:
                    print(f">>> ERRO ao executar combinador GRADES: {e}")
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao executar combinador GRADES:\n{e}")
                    return
            else:
                # Em ambiente de desenvolvimento: usar subprocess
                try:
                    import subprocess
                    resultado_combinador = subprocess.run(
                        [sys.executable, combinador_path, pasta_pavimento],
                        capture_output=True,
                        text=True,
                        timeout=300
                    )
                    
                    if resultado_combinador.returncode == 0:
                        print(f">>> ✅ Combinador GRADES executado com sucesso!")
                    else:
                        print(f">>> ERRO no combinador GRADES: {resultado_combinador.stderr}")
                        messagebox.showerror("Erro", f"Erro ao executar combinador GRADES:\n{resultado_combinador.stderr}")
                        return
                except Exception as e:
                    print(f">>> ERRO ao executar combinador GRADES: {e}")
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao executar combinador GRADES:\n{e}")
                    return

            # ========================================================
            # EXECUTAR ORDENADOR NA PASTA COMBINADOS
            # ========================================================
            
            pasta_combinados = os.path.join(pasta_pavimento, "Combinados")
            
            if not os.path.exists(pasta_combinados):
                messagebox.showerror("Erro", f"Pasta Combinados não foi criada: {pasta_combinados}")
                return
            
            # DEBUG: Listar TODOS os scripts na pasta Combinados ANTES do ordenador
            print(f"\n>>> [DEBUG GRADES] Scripts na pasta Combinados ANTES do ordenador:")
            try:
                scripts_combinados = [f for f in os.listdir(pasta_combinados) if f.endswith('.scr')]
                print(f">>> [DEBUG GRADES] Total de scripts encontrados: {len(scripts_combinados)}")
                for arq in sorted(scripts_combinados):
                    caminho_completo = os.path.join(pasta_combinados, arq)
                    tamanho = os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                    print(f">>> [DEBUG GRADES]   - {arq} ({tamanho} bytes)")
            except Exception as e:
                print(f">>> [DEBUG GRADES] ERRO ao listar scripts: {e}")
            
            print(f"\n>>> EXECUTANDO ORDENADOR GRADES NA PASTA COMBINADOS...")
            print(f">>> Pasta Combinados: {pasta_combinados}")
            
            # Executar ordenador de forma compatível com frozen
            self._executar_ordenador_grades_automatico(pasta_combinados)
            
            # DEBUG: Listar TODOS os scripts na pasta Combinados DEPOIS do ordenador
            print(f"\n>>> [DEBUG GRADES] Scripts na pasta Combinados DEPOIS do ordenador:")
            try:
                scripts_combinados = [f for f in os.listdir(pasta_combinados) if f.endswith('.scr')]
                print(f">>> [DEBUG GRADES] Total de scripts encontrados: {len(scripts_combinados)}")
                for arq in sorted(scripts_combinados):
                    caminho_completo = os.path.join(pasta_combinados, arq)
                    tamanho = os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                    print(f">>> [DEBUG GRADES]   - {arq} ({tamanho} bytes)")
            except Exception as e:
                print(f">>> [DEBUG GRADES] ERRO ao listar scripts: {e}")

            # ========================================================
            # EXECUTAR SCRIPTS COMBINADOS NO AUTOCAD
            # ========================================================
            
            # Executar no AutoCAD usando o script combinado da pasta Combinados
            print(f"\n>>> Executando scripts combinados no AutoCAD...")
            script_final = self._processar_scripts_combinados(pasta_combinados)
            if script_final:
                # ========================================================
                # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
                # ========================================================
                self._limpar_linhas_em_branco_entre_semicolons(pasta_combinados)
                
                # ========================================================
                # 🎯 ATUALIZAR SCRIPT_PAZ.scr COM O PRIMEIRO SCRIPT DA PASTA COMBINADOS
                # ========================================================
                script_1_path = os.path.join(pasta_combinados, "1.scr")
                if os.path.exists(script_1_path):
                    print(f">>> [PAZ] Atualizando script_PAZ.scr com o primeiro script da pasta Combinados (GRADES)")
                    self._atualizar_script_paz(script_1_path)
                    print(f">>> [PAZ] ✅ Script PAZ atualizado com 1.scr da pasta Combinados")
                else:
                    print(f">>> [PAZ] ⚠️ Arquivo 1.scr não encontrado na pasta Combinados, pulando atualização do PAZ")
                
                # DEBUG: Verificar conteúdo do script final antes de enviar ao AutoCAD
                print(f"\n>>> [DEBUG GRADES] Verificando script final antes de enviar ao AutoCAD:")
                linhas_final = script_final.splitlines()
                linhas_nao_vazias_final = [l for l in linhas_final if l.strip()]
                print(f">>> [DEBUG GRADES]   - Total de linhas: {len(linhas_final)}")
                print(f">>> [DEBUG GRADES]   - Linhas não vazias: {len(linhas_nao_vazias_final)}")
                print(f">>> [DEBUG GRADES]   - Tamanho do script: {len(script_final)} caracteres")
                
                # Perguntar se deseja executar no AutoCAD
                opcao = self._perguntar_execucao_autocad("desenho do pavimento GRADES")
                
                if opcao is None:  # Cancelado
                    return
                
                if opcao == 1:
                    # Opção 1: Desenhar no CAD diretamente
                    # Garantir que o script final não tenha caracteres nulos antes de enviar ao AutoCAD
                    script_final = script_final.replace('\x00', '')
                    self._executar_no_autocad(script_final, pavimento, "GRADES")
                    print(f"✅ Desenho do pavimento GRADES executado com sucesso!")
                elif opcao == 2:
                    # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                    print(f"✅ Scripts gerados (não executados no AutoCAD)")
                elif opcao == 3:
                    # Opção 3: Gerar DFX (em desenvolvimento)
                    messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                    return
                
                # Contar quantos textos identificadores existem no script final
                textos_finais = []
                for i, linha in enumerate(linhas_final):
                    if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_final):
                        texto_id = linhas_final[i + 3].strip()
                        if texto_id and (texto_id.startswith('P') or texto_id.startswith('p')):
                            textos_finais.append(texto_id)
                print(f">>> [DEBUG GRADES]   - Textos identificadores no script final: {textos_finais}")
                print(f">>> [DEBUG GRADES]   - Total de textos identificadores: {len(textos_finais)}")
                
                # Verificar se P4.E e P4.F estão presentes
                if 'P4.E' in textos_finais and 'P4.F' in textos_finais:
                    print(f">>> [DEBUG GRADES] ✅ P4.E e P4.F estão presentes no script final!")
                else:
                    print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: P4.E e/ou P4.F NÃO estão presentes no script final!")
                    if 'P4.E' not in textos_finais:
                        print(f">>> [DEBUG GRADES]   - P4.E está faltando!")
                    if 'P4.F' not in textos_finais:
                        print(f">>> [DEBUG GRADES]   - P4.F está faltando!")
                
                # Procurar onde P4.B termina e onde P4.E deveria começar
                print(f">>> [DEBUG GRADES]   - Procurando posição de P4.B e P4.E no script final...")
                posicao_p4b = None
                posicao_p4e = None
                for i, linha in enumerate(linhas_final):
                    if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_final):
                        texto_id = linhas_final[i + 3].strip()
                        if texto_id == 'P4.B':
                            posicao_p4b = i
                            print(f">>> [DEBUG GRADES]     - P4.B encontrado na linha {i+1}")
                        elif texto_id == 'P4.E':
                            posicao_p4e = i
                            print(f">>> [DEBUG GRADES]     - P4.E encontrado na linha {i+1}")
                
                if posicao_p4b is not None and posicao_p4e is None:
                    print(f">>> [DEBUG GRADES] ⚠️ P4.B encontrado na linha {posicao_p4b+1}, mas P4.E não foi encontrado!")
                    print(f">>> [DEBUG GRADES]   - Verificando linhas após P4.B...")
                    # Mostrar últimas 20 linhas do script final
                    print(f">>> [DEBUG GRADES]   - Últimas 20 linhas do script final:")
                    for i, linha in enumerate(linhas_final[-20:], start=len(linhas_final)-19):
                        print(f">>> [DEBUG GRADES]     [{i}] {linha[:80] if len(linha) > 80 else linha}")
                
                # DEBUG: Salvar script final em arquivo temporário para depuração
                try:
                    from utils.robust_path_resolver import robust_path_resolver
                    pasta_debug = os.path.join(robust_path_resolver.get_project_root(), "output", "debug")
                    os.makedirs(pasta_debug, exist_ok=True)
                    arquivo_debug = os.path.join(pasta_debug, f"script_final_GRADES_{pavimento}.scr")
                    # Garantir que o script final não tenha caracteres nulos antes de salvar
                    script_final_limpo = script_final.replace('\x00', '')
                    # Salvar com UTF-16 LE (com BOM) para compatibilidade com AutoCAD
                    with open(arquivo_debug, 'wb') as f:
                        # Adicionar BOM UTF-16 LE
                        f.write(b'\xFF\xFE')
                        # Converter conteúdo para UTF-16 LE
                        f.write(script_final_limpo.encode('utf-16-le'))
                    print(f">>> [DEBUG GRADES] ✅ Script final salvo em: {arquivo_debug}")
                    print(f">>> [DEBUG GRADES]   - Tamanho do arquivo: {len(script_final_limpo)} caracteres")
                    
                    # Verificar se P4.E e P4.F estão no arquivo salvo
                    with open(arquivo_debug, 'r', encoding='utf-16-le') as f:
                        conteudo_arquivo = f.read()
                        if 'P4.E' in conteudo_arquivo:
                            print(f">>> [DEBUG GRADES] ✅ P4.E encontrado no arquivo salvo")
                        else:
                            print(f">>> [DEBUG GRADES] ⚠️ P4.E NÃO encontrado no arquivo salvo!")
                        if 'P4.F' in conteudo_arquivo:
                            print(f">>> [DEBUG GRADES] ✅ P4.F encontrado no arquivo salvo")
                        else:
                            print(f">>> [DEBUG GRADES] ⚠️ P4.F NÃO encontrado no arquivo salvo!")
                    
                    # Mostrar uma amostra do conteúdo onde P4.E deveria estar
                    if posicao_p4e is not None:
                        print(f">>> [DEBUG GRADES] Mostrando amostra do conteúdo ao redor de P4.E (linha {posicao_p4e+1}):")
                        inicio = max(0, posicao_p4e - 5)
                        fim = min(len(linhas_final), posicao_p4e + 20)
                        for i in range(inicio, fim):
                            linha = linhas_final[i]
                            marcador = ">>>" if i == posicao_p4e else "   "
                            print(f">>> [DEBUG GRADES] {marcador} [{i+1}] {linha[:100] if len(linha) > 100 else linha}")
                    
                    # Mostrar uma amostra do conteúdo onde P4.F deveria estar
                    posicao_p4f = None
                    for i, linha in enumerate(linhas_final):
                        if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_final):
                            texto_id = linhas_final[i + 3].strip()
                            if texto_id == 'P4.F':
                                posicao_p4f = i
                                break
                    
                    if posicao_p4f is not None:
                        print(f">>> [DEBUG GRADES] Mostrando amostra do conteúdo ao redor de P4.F (linha {posicao_p4f+1}):")
                        inicio = max(0, posicao_p4f - 5)
                        fim = min(len(linhas_final), posicao_p4f + 20)
                        for i in range(inicio, fim):
                            linha = linhas_final[i]
                            marcador = ">>>" if i == posicao_p4f else "   "
                            print(f">>> [DEBUG GRADES] {marcador} [{i+1}] {linha[:100] if len(linha) > 100 else linha}")
                    else:
                        print(f">>> [DEBUG GRADES] ⚠️ P4.F não encontrado no script final!")
                        
                except Exception as e:
                    print(f">>> [DEBUG GRADES] ⚠️ Erro ao salvar script final para depuração: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f">>> ERRO: Não foi possível processar scripts combinados")
                messagebox.showerror("Erro", "Não foi possível processar scripts combinados da pasta Combinados")

        except Exception as e:
            print(f">>> ERRO ao executar D.CAD pavimento GRADES: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar D.CAD pavimento GRADES: {str(e)}")
        finally:
            # Liberar locks
            self.processing_lock = False
            self.dcad_lock = False
            print(">>> [LOCK] Locks liberados após execução D.CAD pavimento GRADES")

    def _segunda_revisao_style(self, script_content):
        """
        Segunda revisão específica para comandos STYLE - garante exatamente 9 linhas
        """
        lines = script_content.splitlines()
        edited_lines = []
        i = 0
        
        while i < len(lines):
            line = lines[i].strip()
            
            # Se encontrar um comando STYLE
            if line.upper() == '-STYLE':
                # Adicionar o comando -STYLE
                edited_lines.append('-STYLE')
                i += 1
                
                # Coletar os próximos valores do comando STYLE
                font_name = "Romans"  # valor padrão
                font_size = "12"      # valor padrão
                
                # Coletar o nome da fonte (primeira linha após -STYLE)
                if i < len(lines):
                    font_line = lines[i].strip()
                    if font_line and not font_line.startswith(('-', '_')):
                        font_name = font_line
                        i += 1
                
                # Pular linhas em branco até encontrar o tamanho da fonte
                while i < len(lines):
                    current_line = lines[i].strip()
                    # Se encontrar um comando principal, para
                    if current_line.startswith(('-', '_')) and len(current_line) > 1:
                        break
                    # Se encontrar um valor numérico, é o tamanho da fonte
                    if current_line and current_line.replace('.', '').replace(',', '').isdigit():
                        font_size = current_line
                        i += 1
                        break
                    i += 1
                
                # Pular todas as linhas restantes até encontrar o próximo comando
                while i < len(lines):
                    current_line = lines[i].strip()
                    # Se encontrar um comando principal (começa com - ou _), para
                    if current_line.startswith(('-', '_')) and len(current_line) > 1:
                        break
                    i += 1
                
                # Adicionar estrutura do STYLE com valores coletados dinamicamente
                # Regra especial: Arial tem 4 linhas em branco, demais fontes têm 5
                blank_lines_count = 4 if font_name.lower() == "arial" else 5
                
                style_lines = [
                    font_name,       # linha 2: tipo de texto coletado
                    "",             # linha 3: linha em branco
                    font_size,      # linha 4: valor numérico coletado
                ]
                
                # Adicionar linhas em branco conforme a regra
                for _ in range(blank_lines_count):
                    style_lines.append("")
                
                edited_lines.extend(style_lines)
                continue
            else:
                # Adicionar linha normal
                edited_lines.append(lines[i])
                i += 1
        
        return '\n'.join(edited_lines)

    def _has_linetype_in_last_lines(self, edited_lines, max_lines=4):
        """
        Verifica se há um comando -LINETYPE nas últimas 'max_lines' linhas
        para evitar remover linhas em branco que são necessárias após -LINETYPE
        """
        if not edited_lines:
            return False
        
        # Verifica as últimas 'max_lines' linhas não vazias
        lines_checked = 0
        for i in range(len(edited_lines) - 1, -1, -1):
            line = edited_lines[i].strip()
            if line:  # Só conta linhas não vazias
                if line == '-LINETYPE':
                    return True
                lines_checked += 1
                if lines_checked >= max_lines:
                    break
        return False

    def _ler_e_filtrar_script(self, script_path_or_content, is_content=False):
        """Lê e filtra o conteúdo do script. Pode receber um caminho de arquivo ou o conteúdo direto."""
        try:
            # Se recebeu conteúdo direto, usa ele. Senão, lê do arquivo
            if is_content:
                script_content = script_path_or_content
            else:
                # Tentar diferentes encodings (priorizar UTF-16 LE que é o padrão para .scr)
                # Em ambiente frozen, garantir que o arquivo existe antes de tentar ler
                script_content = None
                if not os.path.exists(script_path_or_content):
                    raise FileNotFoundError(f"Arquivo não encontrado: {script_path_or_content}")
                
                for encoding in ['utf-16-le', 'utf-16', 'utf-8', 'latin-1', 'cp1252']:
                    try:
                        # Tentar ler com encoding especificado
                        with open(script_path_or_content, "r", encoding=encoding) as f:
                            script_content = f.read()
                        print(f"[PILAR_ESPECIAL] Arquivo lido com encoding: {encoding}")
                        break
                    except UnicodeDecodeError as e:
                        print(f"[PILAR_ESPECIAL] Falha ao ler com {encoding}: {e}")
                        continue
                    except Exception as e:
                        print(f"[PILAR_ESPECIAL] Erro ao ler arquivo com {encoding}: {e}")
                        continue
                if script_content is None:
                    raise Exception("Não foi possível ler o arquivo com os encodings suportados.")

            # Filtrar o script
            lines = script_content.splitlines()
            edited_lines = []
            skip_next_blank = False
            i = 0

            while i < len(lines):
                line = lines[i]
                # Remover caracteres invisíveis e espaços em branco
                # Usar strip() para remover espaços e caracteres especiais
                original_line = line.strip()
                # Remover caracteres de controle (incluindo \x00 que aparecem em UTF-16 mal lido)
                # Remover caracteres nulos e outros caracteres de controle
                original_line = ''.join(char for char in original_line if ord(char) >= 32 or char in '\t\n\r')
                original_line = original_line.replace('\x00', '')  # Remover caracteres nulos
                original_line = original_line.strip()
                lower_line = original_line.lower()

                # Remover linhas que começam com 'c:' (case-insensitive, cobre c:\ e c:)
                # IMPORTANTE: Remover também linhas que contenham caminhos completos (C:\Users\...)
                # Verificar se a linha contém um caminho de arquivo Windows
                # Verificar também se a linha contém apenas o caminho (sem outros caracteres)
                is_c_path = (lower_line.startswith('c:') or 
                            'c:\\' in lower_line or 
                            'c:/' in lower_line or
                            original_line.startswith('C:') or
                            'output\\scripts' in lower_line or
                            'output/scripts' in lower_line or
                            ('combinados' in lower_line and '.scr' in lower_line) or
                            (lower_line.count('\\') > 2 and '.scr' in lower_line))  # Caminho completo com .scr
                
                if is_c_path:
                    skip_next_blank = True
                    i += 1
                    continue

                # Remover linhas que começam com 'SCRIPT' (case-insensitive)
                # IMPORTANTE: Remover também linhas que sejam apenas 'SCRIPT' (comando do AutoCAD)
                # Verificar se a linha é exatamente 'SCRIPT' ou começa com 'SCRIPT'
                # Remover também linhas que contenham apenas 'SCRIPT' (sem outros caracteres)
                is_script = (lower_line == 'script' or 
                            lower_line.strip() == 'script' or
                            lower_line.startswith('script') or
                            original_line == 'SCRIPT' or
                            original_line.strip() == 'SCRIPT' or
                            original_line.startswith('SCRIPT'))
                
                if is_script:
                    skip_next_blank = True
                    i += 1
                    continue

                # Remover linhas que comecem com ';'
                if original_line.startswith(';'):
                    i += 1
                    continue

                # Remover apenas linhas que sejam apenas ';' (comentários vazios)
                if line.strip() == ';':
                    i += 1
                    continue

                # Remover linhas que começam com 'C:' (case-sensitive também)
                if original_line.startswith('C:'):
                    skip_next_blank = True
                    i += 1
                    continue

                # Remover linhas que começam com '_SCRIPT'
                if original_line.startswith('_SCRIPT'):
                    skip_next_blank = True
                    i += 1
                    continue

                # Se a linha anterior era 'c:...' e esta está em branco, pula
                if skip_next_blank and original_line == '':
                    skip_next_blank = False
                    i += 1
                    continue

                # Verificar se a linha é exatamente LAYER, TEXT ou LINETYPE
                if original_line.upper().endswith('LAYER'):
                    # Remove quaisquer linhas em branco anteriores ao -LAYER
                    # EXCEÇÃO: Não remove se há -LINETYPE nas últimas 4 linhas
                    if not self._has_linetype_in_last_lines(edited_lines, 4):
                        while edited_lines and not edited_lines[-1].strip():
                            edited_lines.pop()
                    
                    # Adiciona o comando -LAYER
                    edited_lines.append('-LAYER')
                    
                    # Processa a próxima linha (que deve conter S e possivelmente o nome do layer)
                    i += 1
                    if i < len(lines):
                        next_line = lines[i].strip()
                        if next_line.startswith('S'):
                            # Se a linha contém "S" e o nome do layer junto
                            if len(next_line) > 1:
                                layer_name = next_line[1:].strip()  # Remove o S e espaços
                                edited_lines.append('S')  # Adiciona S sozinho
                                edited_lines.append(layer_name)  # Adiciona nome do layer em linha separada
                            else:
                                # Se é só o S, adiciona ele e pega o nome do layer da próxima linha
                                edited_lines.append('S')
                                i += 1
                                if i < len(lines):
                                    layer_name = lines[i].strip()
                                    if layer_name:  # Se não for vazia
                                        edited_lines.append(layer_name)
                    
                    # Garante EXATAMENTE uma linha em branco após o nome do layer
                    edited_lines.append('')
                    
                    # Avança o índice até encontrar uma linha não vazia
                    i += 1
                    while i < len(lines) and not lines[i].strip():
                        i += 1
                    i -= 1  # Volta uma posição para o loop principal pegar o próximo comando
                    
                    continue
                elif original_line.upper() == '-STYLE':
                    # NOVO TRATAMENTO PARA STYLE - Coletar valores dinamicamente
                    edited_lines.append('-STYLE')
                    i += 1
                    
                    # Coletar os próximos valores do comando STYLE
                    font_name = "Romans"  # valor padrão
                    font_size = "12"      # valor padrão
                    
                    # Coletar o nome da fonte (primeira linha após -STYLE)
                    if i < len(lines):
                        font_line = lines[i].strip()
                        if font_line and not font_line.startswith(('-', '_')):
                            font_name = font_line
                            i += 1
                    
                    # Pular linhas em branco até encontrar o tamanho da fonte
                    while i < len(lines):
                        current_line = lines[i].strip()
                        # Se encontrar um comando principal, para
                        if current_line.startswith(('-', '_')) and len(current_line) > 1:
                            break
                        # Se encontrar um valor numérico, é o tamanho da fonte
                        if current_line and current_line.replace('.', '').replace(',', '').isdigit():
                            font_size = current_line
                            i += 1
                            break
                        i += 1
                    
                    # Pular todas as linhas restantes até encontrar o próximo comando
                    while i < len(lines):
                        current_line = lines[i].strip()
                        # Se encontrar um comando principal (começa com - ou _), para
                        if current_line.startswith(('-', '_')) and len(current_line) > 1:
                            break
                        i += 1
                    
                    # Adicionar estrutura do STYLE com valores coletados dinamicamente
                    # Regra especial: Arial tem 4 linhas em branco, demais fontes têm 5
                    blank_lines_count = 4 if font_name.lower() == "arial" else 5
                    
                    style_lines = [
                        font_name,       # linha 2: tipo de texto coletado
                        "",             # linha 3: linha em branco
                        font_size,      # linha 4: valor numérico coletado
                    ]
                    
                    # Adicionar linhas em branco conforme a regra
                    for _ in range(blank_lines_count):
                        style_lines.append("")
                    
                    edited_lines.extend(style_lines)
                    
                    # Voltar uma posição para o loop principal processar o próximo comando
                    i -= 1
                    continue
                elif 'TEXT' in original_line.upper():
                    # Tratamento específico para TEXT: garantir que sempre comece com -
                    line = '-TEXT'
                elif original_line.upper().endswith('LINETYPE'):
                    # NÃO remove linhas em branco anteriores ao -LINETYPE
                    # As linhas em branco são importantes para o funcionamento correto
                    
                    # Adiciona o comando -LINETYPE
                    edited_lines.append('-LINETYPE')
                    
                    # Processa a próxima linha (que deve conter S e possivelmente o nome do linetype)
                    i += 1
                    if i < len(lines):
                        next_line = lines[i].strip()
                        if next_line.startswith('S'):
                            # Se a linha contém "S" e o nome do linetype junto
                            if len(next_line) > 1:
                                linetype_name = next_line[1:].strip()  # Remove o S e espaços
                                edited_lines.append('S')  # Adiciona S sozinho
                                edited_lines.append(linetype_name)  # Adiciona nome do linetype em linha separada
                            else:
                                # Se é só o S, adiciona ele e pega o nome do linetype da próxima linha
                                edited_lines.append('S')
                                i += 1
                                if i < len(lines):
                                    linetype_name = lines[i].strip()
                                    if linetype_name:  # Se não for vazia
                                        edited_lines.append(linetype_name)
                    
                    # GARANTIR EXATAMENTE uma linha em branco após o nome do linetype
                    # Verifica se já existe uma linha em branco no script original
                    i += 1
                    has_blank_in_original = False
                    while i < len(lines) and not lines[i].strip():
                        has_blank_in_original = True
                        i += 1
                    
                    # Adiciona apenas uma linha em branco se não existir no original
                    if not has_blank_in_original:
                        edited_lines.append('')
                    
                    i -= 1  # Volta uma posição para o loop principal pegar o próximo comando
                    
                    continue

                # Tratamento especial para -INSERT
                if original_line == '-INSERT':
                    edited_lines.append(line)
                    # Adiciona as próximas 4 linhas (nome, coordenadas, escala, rotação)
                    lines_added = 0
                    for j in range(4):
                        if i + 1 + j < len(lines):
                            next_line = lines[i + 1 + j].strip()
                            if next_line:  # Só adiciona se não for vazia
                                edited_lines.append(next_line)
                                lines_added += 1
                    # Avança o índice para depois do último item adicionado
                    i += lines_added + 1
                    continue

                edited_lines.append(line)
                i += 1
                skip_next_blank = False

            # Varredura final para remover linhas em branco após -INSERT
            i = 0
            while i < len(edited_lines):
                if edited_lines[i].strip() == '-INSERT':
                    # Pula as 4 linhas obrigatórias após o -INSERT (nome, coordenadas, escala, rotação)
                    next_lines = 4
                    i += next_lines + 1
                    # Verifica as próximas 6 linhas
                    j = i
                    while j < min(i + 6, len(edited_lines)):
                        # Se for uma linha em branco, remove
                        # EXCEÇÃO: Não remove se há -LINETYPE nas proximidades
                        if edited_lines[j].strip() == '' and not self._has_linetype_in_last_lines(edited_lines[:j+1], 4):
                            edited_lines.pop(j)
                        # Se for um "0", mantém e avança
                        elif edited_lines[j].strip() == '0':
                            j += 1
                        # Se for qualquer outra coisa, avança
                        else:
                            j += 1
                else:
                    i += 1

            # Varredura final para remover linha em branco após -TEXT (mesma lógica do -INSERT)
            i = 0
            while i < len(edited_lines):
                if edited_lines[i].strip() == '-TEXT':
                    # Pula as 4 linhas obrigatórias após o -TEXT (nome, coordenadas, altura, rotação, por exemplo)
                    next_lines = 4
                    i += next_lines + 1
                    # Verifica as próximas 6 linhas
                    j = i
                    while j < min(i + 6, len(edited_lines)):
                        # Se for uma linha em branco, remove
                        # EXCEÇÃO: Não remove se há -LINETYPE nas proximidades
                        if edited_lines[j].strip() == '' and not self._has_linetype_in_last_lines(edited_lines[:j+1], 4):
                            edited_lines.pop(j)
                        # Se for um "0", mantém e avança
                        elif edited_lines[j].strip() == '0':
                            j += 1
                        # Se for qualquer outra coisa, avança
                        else:
                            j += 1
                else:
                    i += 1

            # Varredura final para preservar linhas em branco após -LINETYPE
            i = 0
            while i < len(edited_lines):
                if edited_lines[i].strip() == '-LINETYPE':
                    # Pula o comando -LINETYPE
                    i += 1
                    # Pula a linha S se existir
                    if i < len(edited_lines) and edited_lines[i].strip() == 'S':
                        i += 1
                    # Pula o nome do linetype se existir
                    if i < len(edited_lines) and edited_lines[i].strip() and not edited_lines[i].strip().startswith('-'):
                        i += 1
                    # Preserva todas as linhas em branco que vêm depois - NÃO REMOVE NENHUMA
                    # Apenas avança para o próximo comando
                    while i < len(edited_lines) and edited_lines[i].strip() == '':
                        i += 1
                    # Se não for linha em branco, volta uma posição para o loop principal processar
                    if i < len(edited_lines) and edited_lines[i].strip() != '':
                        i -= 1
                else:
                    i += 1

            # Varredura para remover espaços da 4ª linha após -TEXT
            i = 0
            while i < len(edited_lines):
                if edited_lines[i].strip() == '-TEXT':
                    # Verifica se há pelo menos 4 linhas após o -TEXT
                    if i + 4 < len(edited_lines):
                        # Pega a 4ª linha (índice i + 3)
                        fourth_line = edited_lines[i + 3]
                        # Remove todos os espaços da linha
                        edited_lines[i + 3] = fourth_line.replace(' ', '')
                i += 1
            
            # Varredura final para normalizar comandos LAYER (tanto -LAYER quanto LAYER)
            # Garante que após cada comando LAYER completo haja EXATAMENTE 1 linha em branco
            i = 0
            while i < len(edited_lines):
                line_stripped = edited_lines[i].strip().upper()
                # Verifica se é comando LAYER (com ou sem hífen)
                if line_stripped == '-LAYER' or line_stripped == 'LAYER':
                    # Guarda o índice onde encontrou o LAYER e o formato original
                    layer_start = i
                    layer_command = edited_lines[i].strip()  # Mantém formato original (-LAYER ou LAYER)
                    
                    # Avança para processar as próximas linhas
                    i += 1
                    layer_lines = [layer_command]  # Começa com o comando no formato original
                    
                    # Processa o S e o nome do layer
                    layer_name = None
                    if i < len(edited_lines):
                        next_line = edited_lines[i].strip()
                        if next_line.startswith('S'):
                            if len(next_line) > 1:
                                # Se S e layer estão na mesma linha: "S NOME_LAYER"
                                # Manter na mesma linha
                                layer_lines.append(next_line)  # Mantém "S NOME_LAYER" intacto
                                layer_name = next_line[1:].strip()
                                i += 1
                            else:
                                # Se é só o S, pega o nome do layer da próxima linha
                                layer_lines.append('S')
                                i += 1
                                if i < len(edited_lines):
                                    layer_name = edited_lines[i].strip()
                                    if layer_name:
                                        # Combinar S com o nome do layer na mesma linha
                                        layer_lines[-1] = f"S {layer_name}"
                                        i += 1
                    
                    # Agora normaliza as linhas em branco após o comando LAYER completo
                    # Conta quantas linhas em branco existem
                    blank_count = 0
                    original_i = i
                    while i < len(edited_lines) and not edited_lines[i].strip():
                        blank_count += 1
                        i += 1
                    
                    # Garante EXATAMENTE 1 linha em branco
                    layer_lines.append('')
                    
                    # Substitui as linhas antigas pelas novas
                    # Remove todas as linhas do comando LAYER até o próximo comando não vazio
                    edited_lines[layer_start:original_i + blank_count] = layer_lines
                    # Ajusta o índice para continuar após o comando normalizado
                    i = layer_start + len(layer_lines)
                else:
                    i += 1
            
            # Varredura final para normalizar comandos _LINE
            # Garante que após cada comando _LINE completo (comando + coordenadas) haja EXATAMENTE 1 linha em branco
            i = 0
            while i < len(edited_lines):
                line_stripped = edited_lines[i].strip()
                # Verifica se é comando _LINE
                if line_stripped == '_LINE':
                    # Guarda o índice onde encontrou o _LINE
                    line_start = i
                    line_lines = ['_LINE']  # Começa com o comando
                    
                    # Avança para processar as coordenadas (coleta todas as linhas não vazias até encontrar linha em branco ou outro comando)
                    i += 1
                    while i < len(edited_lines):
                        next_line = edited_lines[i].strip()
                        # Se a linha está vazia, para de coletar coordenadas
                        if not next_line:
                            break
                        # Se encontrar outro comando (começa com _ ou -), para
                        if next_line.startswith('_') or next_line.startswith('-'):
                            break
                        # Adiciona a linha (coordenada ou outro conteúdo)
                        line_lines.append(next_line)
                        i += 1
                    
                    # Agora normaliza as linhas em branco após o comando _LINE completo
                    # Conta quantas linhas em branco existem
                    blank_count = 0
                    original_i = i
                    while i < len(edited_lines) and not edited_lines[i].strip():
                        blank_count += 1
                        i += 1
                    
                    # Garante EXATAMENTE 1 linha em branco
                    line_lines.append('')
                    
                    # Substitui as linhas antigas pelas novas
                    # Remove todas as linhas do comando _LINE até o próximo comando não vazio
                    edited_lines[line_start:original_i + blank_count] = line_lines
                    # Ajusta o índice para continuar após o comando normalizado
                    i = line_start + len(line_lines)
                else:
                    i += 1
            
            # Varredura final para normalizar comandos _ZOOM
            # Garante que após cada comando _ZOOM completo haja EXATAMENTE 0 linhas em branco
            i = 0
            while i < len(edited_lines):
                line_stripped = edited_lines[i].strip()
                # Verifica se é comando _ZOOM
                if line_stripped == '_ZOOM':
                    # Guarda o índice onde encontrou o _ZOOM
                    zoom_start = i
                    zoom_lines = ['_ZOOM']  # Começa com o comando
                    
                    # Avança para processar as linhas do comando (C, coordenadas, escala, ;)
                    i += 1
                    while i < len(edited_lines):
                        next_line = edited_lines[i].strip()
                        # Se a linha está vazia, para de coletar
                        if not next_line:
                            break
                        # Se encontrar outro comando (começa com _ ou -), para
                        if (next_line.startswith('_') or next_line.startswith('-')) and next_line != '_ZOOM':
                            break
                        # Adiciona a linha (C, coordenadas, escala, ;, etc)
                        zoom_lines.append(next_line)
                        i += 1
                        # Se encontrou o ;, o comando terminou
                        if next_line == ';':
                            break
                    
                    # Agora normaliza as linhas em branco após o comando _ZOOM completo
                    # Conta quantas linhas em branco existem
                    blank_count = 0
                    original_i = i
                    while i < len(edited_lines) and not edited_lines[i].strip():
                        blank_count += 1
                        i += 1
                    
                    # Garante EXATAMENTE 0 linhas em branco (não adiciona linha em branco)
                    # Não adiciona linha em branco ao zoom_lines
                    
                    # Substitui as linhas antigas pelas novas
                    # Remove todas as linhas do comando _ZOOM até o próximo comando não vazio, incluindo linhas em branco
                    edited_lines[zoom_start:original_i + blank_count] = zoom_lines
                    # Ajusta o índice para continuar após o comando normalizado
                    i = zoom_start + len(zoom_lines)
                else:
                    i += 1
            
            # Retornar o script filtrado
            script_filtrado = '\n'.join(edited_lines)
            
            # Aplicar segunda revisão específica para comandos STYLE
            script_final = self._segunda_revisao_style(script_filtrado)
            
            return script_final

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler/filtrar script: {str(e)}")
            return None

    def _processar_scripts_combinados(self, combinados_dir):
        """Processa todos os scripts da pasta Combinados"""
        try:
            print(f"\n>>> PROCESSANDO SCRIPTS COMBINADOS")
            print(f">>> Diretório: {combinados_dir}")
            
            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não encontrada!")
                print(f">>> Caminho completo: {os.path.abspath(combinados_dir)}")
                
                # Tentar listar o diretório pai para debug
                dir_pai = os.path.dirname(combinados_dir)
                if os.path.exists(dir_pai):
                    print(f">>> Conteúdo do diretório pai ({dir_pai}):")
                    try:
                        for item in os.listdir(dir_pai):
                            print(f">>>   - {item}")
                    except Exception as e:
                        print(f">>> Erro ao listar diretório pai: {e}")
                else:
                    print(f">>> Diretório pai também não existe: {dir_pai}")
                
                messagebox.showerror("Erro", f"Pasta Combinados não encontrada em:\n{combinados_dir}\n\nVerifique se o pavimento foi processado corretamente.")
                return None

            print(f">>> Pasta Combinados encontrada com sucesso!")

            # Buscar e ordenar arquivos .scr (alfabético, ignorando acentos e sufixos)
            try:
                nomes = [f for f in os.listdir(combinados_dir) if f.endswith('.scr')]
            except Exception as e:
                print(f">>> Erro ao listar arquivos: {e}")
                messagebox.showerror("Erro", f"Erro ao listar arquivos na pasta:\n{str(e)}")
                return None

            import unicodedata
            import re
            # Detectar presença de nomeados e numerados
            has_numeric = any(re.match(r"^\d+\.scr$", f) for f in nomes)
            has_named = any(not re.match(r"^\d+\.scr$", f) for f in nomes)
            if has_numeric:
                nomes = [f for f in nomes if re.match(r"^\d+\.scr$", f)]
            def _ordenador_key(nome_arquivo: str):
                base = nome_arquivo
                if base.endswith('_COMBINADO_CIMA.scr'):
                    base = base[:-len('_COMBINADO_CIMA.scr')]
                elif base.endswith('_CIMA.scr'):
                    base = base[:-len('_CIMA.scr')]
                elif base.endswith('.scr'):
                    base = base[:-len('.scr')]
                base = base.split('-')[0]
                base_norm = unicodedata.normalize('NFD', base)
                base_norm = ''.join(ch for ch in base_norm if unicodedata.category(ch) != 'Mn')
                # 1) Números puros
                if base_norm.isdigit():
                    return (0, int(base_norm), 0 if nome_arquivo.endswith('_COMBINADO_CIMA.scr') else 1)
                # 2) Letras seguidas de números (A1, P3)
                import re
                m = re.match(r'^([A-Za-z]+)(\d+)$', base_norm)
                if m:
                    prefix = m.group(1).upper()
                    num = int(m.group(2))
                    grupo = 1 if prefix == 'A' else (2 if prefix == 'P' else 3)
                    return (grupo, prefix, num)
                # 3) Alfabético puro
                return (4, base_norm.upper(), 0)

            arquivos_scr = sorted(nomes, key=_ordenador_key)
            print(f">>> Total de arquivos .scr encontrados: {len(arquivos_scr)}")
            
            # DEBUG: Listar TODOS os arquivos encontrados
            print(f"\n>>> [DEBUG GRADES] Arquivos .scr encontrados na pasta Combinados:")
            for i, nome in enumerate(arquivos_scr, 1):
                caminho_completo = os.path.join(combinados_dir, nome)
                tamanho = os.path.getsize(caminho_completo) if os.path.exists(caminho_completo) else 0
                print(f">>> [DEBUG GRADES]   [{i}] {nome} ({tamanho} bytes)")

            if not arquivos_scr:
                print(f">>> ERRO: Nenhum arquivo .scr encontrado!")
                messagebox.showerror("Erro", f"Nenhum arquivo .scr encontrado em:\n{combinados_dir}\n\nVerifique se os scripts foram gerados corretamente.")
                return None

            # Listar arquivos encontrados para debug
            print(f">>> Arquivos encontrados:")
            for i, nome in enumerate(arquivos_scr, 1):
                print(f">>>   {i}: {nome}")

            # Processar apenas arquivos presentes em Combinados (já ordenados e com offset aplicado pelo ordenador)
            import re
            scripts_combinados = []
            
            # CORREÇÃO: Para GRADES, os arquivos na pasta Combinados são numerados (1.scr, 2.scr, etc.)
            # e já foram processados e ordenados pelo combinador. Não devemos usar detecção de duplicatas
            # baseada em base_core porque esses números podem ser os mesmos de execuções anteriores.
            # Devemos processar TODOS os arquivos da pasta Combinados.
            
            vistos = set()  # Apenas para evitar duplicatas dentro desta execução
            scripts_ignorados = []
            for nome_arquivo in arquivos_scr:
                base = nome_arquivo
                if base.endswith('_COMBINADO_CIMA.scr'):
                    base_core = base[:-len('_COMBINADO_CIMA.scr')]
                elif base.endswith('_CIMA.scr'):
                    base_core = base[:-len('_CIMA.scr')]
                else:
                    # Para GRADES, os arquivos são numerados (1.scr, 2.scr, etc)
                    # Não devemos filtrar por base_core para GRADES
                    base_core = base.replace('.scr', '')
                
                # CORREÇÃO: Para GRADES, não verificar self._itens_desenhados_sessao
                # porque os arquivos na pasta Combinados já foram processados pelo combinador
                # e devem ser processados novamente nesta execução.
                # Apenas verificar duplicatas dentro desta execução (vistos).
                if base_core in vistos:
                    print(f">>> [DEBUG GRADES] ⚠️ Ignorando duplicata dentro desta execução: {nome_arquivo} (base_core: {base_core})")
                    scripts_ignorados.append(nome_arquivo)
                    continue
                
                vistos.add(base_core)
                
                # DEBUG: Confirmar que está sendo processado
                print(f">>> [DEBUG GRADES] ✅ Processando script: {nome_arquivo} (base_core: {base_core})")
                
                # Processar o script
                script_path = os.path.join(combinados_dir, nome_arquivo)
                print(f"\n>>> Processando arquivo: {nome_arquivo}")
                script_content = self._ler_e_filtrar_script(script_path)
                if script_content:
                    # DEBUG: Verificar se o script contém conteúdo relevante
                    linhas_script = script_content.splitlines()
                    linhas_nao_vazias = [l for l in linhas_script if l.strip()]
                    print(f">>> Arquivo {nome_arquivo} processado com sucesso")
                    print(f">>> [DEBUG GRADES]   - Total de linhas: {len(linhas_script)}")
                    print(f">>> [DEBUG GRADES]   - Linhas não vazias: {len(linhas_nao_vazias)}")
                    # Verificar se contém texto identificador (P1, P2, P4, etc.)
                    # Formato: -TEXT (linha i), coordenadas (linha i+1), ângulo (linha i+2), texto (linha i+3)
                    texto_identificador = None
                    for i, linha in enumerate(linhas_script[:100]):  # Verificar primeiras 100 linhas
                        if '-TEXT' in linha.upper() or '_TEXT' in linha.upper():
                            # Procurar linha com nome do pilar na linha i+3 (após coordenadas e ângulo)
                            if i + 3 < len(linhas_script):
                                texto_identificador = linhas_script[i + 3].strip()
                                if texto_identificador and (texto_identificador.startswith('P') or texto_identificador.startswith('p')):
                                    break
                    if texto_identificador:
                        print(f">>> [DEBUG GRADES]   - Texto identificador encontrado: {texto_identificador}")
                    else:
                        # Tentar procurar em todo o script
                        for i, linha in enumerate(linhas_script):
                            if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_script):
                                texto_candidato = linhas_script[i + 3].strip()
                                if texto_candidato and (texto_candidato.startswith('P') or texto_candidato.startswith('p')):
                                    texto_identificador = texto_candidato
                                    print(f">>> [DEBUG GRADES]   - Texto identificador encontrado (busca completa): {texto_identificador}")
                                    break
                    scripts_combinados.append(script_content)
                else:
                    print(f">>> ⚠️ Arquivo {nome_arquivo} não pôde ser processado")
            
            # DEBUG: Resumo de scripts ignorados
            if scripts_ignorados:
                print(f"\n>>> [DEBUG GRADES] ⚠️ Total de scripts ignorados: {len(scripts_ignorados)}")
                for script_ignorado in scripts_ignorados:
                    print(f">>> [DEBUG GRADES]   - {script_ignorado}")

            print(f"\n>>> [DEBUG GRADES] Total de scripts processados: {len(scripts_combinados)}")
            
            # DEBUG: Verificar conteúdo antes da combinação
            print(f">>> [DEBUG GRADES] Verificando conteúdo dos scripts antes da combinação:")
            for idx, script in enumerate(scripts_combinados, 1):
                linhas = script.splitlines()
                linhas_nao_vazias = [l for l in linhas if l.strip()]
                # Procurar texto identificador
                # Formato: -TEXT (linha i), coordenadas (linha i+1), ângulo (linha i+2), texto (linha i+3)
                texto_id = None
                for i, linha in enumerate(linhas[:100]):
                    if '-TEXT' in linha.upper() or '_TEXT' in linha.upper():
                        if i + 3 < len(linhas):
                            texto_candidato = linhas[i + 3].strip()
                            if texto_candidato and (texto_candidato.startswith('P') or texto_candidato.startswith('p')):
                                texto_id = texto_candidato
                                break
                if not texto_id:
                    # Tentar procurar em todo o script
                    for i, linha in enumerate(linhas):
                        if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas):
                            texto_candidato = linhas[i + 3].strip()
                            if texto_candidato and (texto_candidato.startswith('P') or texto_candidato.startswith('p')):
                                texto_id = texto_candidato
                                break
                print(f">>> [DEBUG GRADES]   Script {idx}: {len(linhas)} linhas, {len(linhas_nao_vazias)} não vazias, texto: {texto_id}")

            # Combina os scripts
            script_final = '\n\n'.join(scripts_combinados) if scripts_combinados else None
            
            # DEBUG: Verificar conteúdo após combinação (antes da filtragem final)
            if script_final:
                linhas_antes = script_final.splitlines()
                linhas_nao_vazias_antes = [l for l in linhas_antes if l.strip()]
                print(f">>> [DEBUG GRADES] Script combinado ANTES da filtragem final:")
                print(f">>> [DEBUG GRADES]   - Total de linhas: {len(linhas_antes)}")
                print(f">>> [DEBUG GRADES]   - Linhas não vazias: {len(linhas_nao_vazias_antes)}")
                # Contar quantos textos identificadores existem
                # Formato: -TEXT (linha i), coordenadas (linha i+1), ângulo (linha i+2), texto (linha i+3)
                textos_encontrados = []
                for i, linha in enumerate(linhas_antes):
                    if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_antes):
                        texto_id = linhas_antes[i + 3].strip()
                        if texto_id and (texto_id.startswith('P') or texto_id.startswith('p')):
                            textos_encontrados.append(texto_id)
                print(f">>> [DEBUG GRADES]   - Textos identificadores encontrados: {textos_encontrados}")
                print(f">>> [DEBUG GRADES]   - Total de textos identificadores: {len(textos_encontrados)}")
            
            # Aplicar filtragem final apenas uma vez no resultado combinado
            if script_final:
                print(f">>> Aplicando filtragem final no script combinado...")
                script_final_antes = script_final
                tamanho_antes = len(script_final_antes)
                print(f">>> [DEBUG GRADES] Tamanho ANTES da filtragem: {tamanho_antes} caracteres")
                
                # DEBUG: Verificar quantas ocorrências de SCRIPT e C: existem antes da filtragem
                linhas_antes = script_final_antes.splitlines()
                script_count_antes = sum(1 for line in linhas_antes if 'SCRIPT' in line.upper() or 'script' in line.lower())
                c_count_antes = sum(1 for line in linhas_antes if 'C:' in line or 'c:' in line.lower())
                print(f">>> [DEBUG GRADES] Ocorrências ANTES da filtragem: SCRIPT={script_count_antes}, C:={c_count_antes}")
                
                script_final = self._ler_e_filtrar_script(script_final, is_content=True)
                
                # DEBUG: Verificar quantas ocorrências de SCRIPT e C: existem depois da filtragem
                if script_final:
                    linhas_depois = script_final.splitlines()
                    script_count_depois = sum(1 for line in linhas_depois if 'SCRIPT' in line.upper() or 'script' in line.lower())
                    c_count_depois = sum(1 for line in linhas_depois if 'C:' in line or 'c:' in line.lower())
                    print(f">>> [DEBUG GRADES] Ocorrências DEPOIS da filtragem: SCRIPT={script_count_depois}, C:={c_count_depois}")
                    if script_count_depois > 0 or c_count_depois > 0:
                        print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Ainda há resquícios de SCRIPT ou C: após a filtragem!")
                        print(f">>> [DEBUG GRADES]   - SCRIPT removidos: {script_count_antes - script_count_depois}")
                        print(f">>> [DEBUG GRADES]   - C: removidos: {c_count_antes - c_count_depois}")
                
                # DEBUG: Verificar conteúdo após filtragem final
                if script_final:
                    tamanho_depois = len(script_final)
                    print(f">>> [DEBUG GRADES] Tamanho DEPOIS da filtragem: {tamanho_depois} caracteres")
                    if tamanho_depois < tamanho_antes:
                        print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Tamanho reduziu em {tamanho_antes - tamanho_depois} caracteres!")
                    
                    # Verificar onde P4.B termina e onde P4.E deveria começar
                    linhas_antes_filtragem = script_final_antes.splitlines()
                    linhas_depois_filtragem = script_final.splitlines()
                    
                    # Procurar P4.B e P4.E nas linhas ANTES da filtragem
                    posicao_p4b_antes = None
                    posicao_p4e_antes = None
                    for i, linha in enumerate(linhas_antes_filtragem):
                        if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_antes_filtragem):
                            texto_id = linhas_antes_filtragem[i + 3].strip()
                            if texto_id == 'P4.B':
                                posicao_p4b_antes = i
                            elif texto_id == 'P4.E':
                                posicao_p4e_antes = i
                    
                    # Procurar P4.B e P4.E nas linhas DEPOIS da filtragem
                    posicao_p4b_depois = None
                    posicao_p4e_depois = None
                    for i, linha in enumerate(linhas_depois_filtragem):
                        if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_depois_filtragem):
                            texto_id = linhas_depois_filtragem[i + 3].strip()
                            if texto_id == 'P4.B':
                                posicao_p4b_depois = i
                            elif texto_id == 'P4.E':
                                posicao_p4e_depois = i
                    
                    print(f">>> [DEBUG GRADES] Posições ANTES da filtragem:")
                    print(f">>> [DEBUG GRADES]   - P4.B: linha {posicao_p4b_antes+1 if posicao_p4b_antes is not None else 'NÃO ENCONTRADO'}")
                    print(f">>> [DEBUG GRADES]   - P4.E: linha {posicao_p4e_antes+1 if posicao_p4e_antes is not None else 'NÃO ENCONTRADO'}")
                    print(f">>> [DEBUG GRADES] Posições DEPOIS da filtragem:")
                    print(f">>> [DEBUG GRADES]   - P4.B: linha {posicao_p4b_depois+1 if posicao_p4b_depois is not None else 'NÃO ENCONTRADO'}")
                    print(f">>> [DEBUG GRADES]   - P4.E: linha {posicao_p4e_depois+1 if posicao_p4e_depois is not None else 'NÃO ENCONTRADO'}")
                    
                    if posicao_p4e_antes is not None and posicao_p4e_depois is None:
                        print(f">>> [DEBUG GRADES] ⚠️ PROBLEMA: P4.E estava na linha {posicao_p4e_antes+1} ANTES da filtragem, mas NÃO está DEPOIS!")
                        print(f">>> [DEBUG GRADES]   - Verificando linhas entre P4.B e P4.E ANTES da filtragem...")
                        if posicao_p4b_antes is not None and posicao_p4e_antes is not None:
                            linhas_entre = linhas_antes_filtragem[posicao_p4b_antes:posicao_p4e_antes+10]
                            print(f">>> [DEBUG GRADES]   - Linhas entre P4.B e P4.E (primeiras 20):")
                            for i, linha in enumerate(linhas_entre[:20], start=posicao_p4b_antes+1):
                                print(f">>> [DEBUG GRADES]     [{i}] {linha[:80] if len(linha) > 80 else linha}")
                    
                    linhas_depois = script_final.splitlines()
                    linhas_nao_vazias_depois = [l for l in linhas_depois if l.strip()]
                    print(f">>> [DEBUG GRADES] Script combinado DEPOIS da filtragem final:")
                    print(f">>> [DEBUG GRADES]   - Total de linhas: {len(linhas_depois)}")
                    print(f">>> [DEBUG GRADES]   - Linhas não vazias: {len(linhas_nao_vazias_depois)}")
                    # Contar quantos textos identificadores existem
                    # Formato: -TEXT (linha i), coordenadas (linha i+1), ângulo (linha i+2), texto (linha i+3)
                    textos_encontrados_depois = []
                    for i, linha in enumerate(linhas_depois):
                        if ('-TEXT' in linha.upper() or '_TEXT' in linha.upper()) and i + 3 < len(linhas_depois):
                            texto_id = linhas_depois[i + 3].strip()
                            if texto_id and (texto_id.startswith('P') or texto_id.startswith('p')):
                                textos_encontrados_depois.append(texto_id)
                    print(f">>> [DEBUG GRADES]   - Textos identificadores encontrados: {textos_encontrados_depois}")
                    print(f">>> [DEBUG GRADES]   - Total de textos identificadores: {len(textos_encontrados_depois)}")
                    if len(textos_encontrados) != len(textos_encontrados_depois):
                        print(f">>> [DEBUG GRADES] ⚠️ ATENÇÃO: Textos identificadores foram removidos pela filtragem!")
                        print(f">>> [DEBUG GRADES]   - Antes: {len(textos_encontrados)} textos")
                        print(f">>> [DEBUG GRADES]   - Depois: {len(textos_encontrados_depois)} textos")
                
                print(f">>> Script final gerado com sucesso!")
            return script_final


        except Exception as e:
            print(f">>> ERRO GERAL ao processar scripts combinados: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao processar scripts combinados:\n{str(e)}")
            return None

    def _gerar_script_item_cima(self, numero, dados):
        """Gera o script para um item CIMA sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT ITEM CIMA")
            print(f">>> Item: {numero}")
            print(f">>> Dados: {dados}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho direto para gerar o script
            self._dcad_item_cima()

            # Obter o caminho do script gerado (usar nome normalizado do pavimento)
            robo_dir = self._get_project_root()
            import unicodedata
            def _normalizar_nome_pasta_local(texto):
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(ch for ch in texto_normalizado if unicodedata.category(ch) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                return texto_limpo.strip('_')

            pav_norm = _normalizar_nome_pasta_local(pavimento)
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pav_norm}_CIMA")
            
            # Primeiro, tentar encontrar o script combinado (para pilares especiais)
            script_combinado_path = os.path.join(pasta_pavimento, f"{nome}_COMBINADO_CIMA.scr")
            script_normal_path = os.path.join(pasta_pavimento, f"{nome}_CIMA.scr")
            
            # Verificar qual script usar
            if os.path.exists(script_combinado_path):
                script_path = script_combinado_path
                print(f">>> Usando script combinado (pilar especial)")
            else:
                script_path = script_normal_path
                print(f">>> Usando script normal")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Caminho do script: {script_path}")

            if not os.path.exists(script_path):
                print(f">>> ERRO: Script não encontrado!")
                return None

            print(f">>> Script encontrado com sucesso!")

            # Ler e filtrar o script
            script_content = self._ler_e_filtrar_script(script_path)
            if not script_content:
                print(f">>> ERRO: Falha ao ler/filtrar script")
                return None

            print(f">>> Script CIMA gerado com sucesso!")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao gerar script CIMA: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _gerar_script_item_abcd(self, numero, dados):
        """Gera o script para um item ABCD sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT ITEM ABCD")
            print(f">>> Item: {numero}")
            print(f">>> Dados: {dados}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho normal para gerar o script
            self.conector_painel.executar_abcd_excel()

            # Obter o caminho do script gerado
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            script_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Caminho do script: {script_path}")

            if not os.path.exists(script_path):
                print(f">>> ERRO: Script não encontrado!")
                return None

            print(f">>> Script encontrado com sucesso!")

            # Ler e filtrar o script
            script_content = self._ler_e_filtrar_script(script_path)
            if not script_content:
                print(f">>> ERRO: Falha ao ler/filtrar script")
                return None

            print(f">>> Script ABCD gerado com sucesso!")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao gerar script ABCD: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _gerar_script_item_grades(self, numero, dados):
        """Gera o script para um item GRADES sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT ITEM GRADES")
            print(f">>> Item: {numero}")
            print(f">>> Dados: {dados}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho normal para gerar o script
            self.conector_painel.executar_grade_excel()

            # Obter o caminho do script gerado
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
            script_path = os.path.join(pasta_pavimento, f"{nome}.scr")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Caminho do script: {script_path}")
            
            # Para GRADES, sempre tentar combinar os arquivos .A e .B
            # Os arquivos gerados têm sufixos .A.scr e .B.scr
            script_path_a = os.path.join(pasta_pavimento, f"{nome}.A.scr")
            script_path_b = os.path.join(pasta_pavimento, f"{nome}.B.scr")
            
            print(f">>> Procurando arquivos GRADES:")
            print(f">>> Arquivo A: {script_path_a}")
            print(f">>> Arquivo B: {script_path_b}")
            
            # Verificar se ambos os arquivos existem
            arquivo_a_existe = os.path.exists(script_path_a)
            arquivo_b_existe = os.path.exists(script_path_b)
            
            print(f">>> Arquivo A existe: {arquivo_a_existe}")
            print(f">>> Arquivo B existe: {arquivo_b_existe}")
            
            # Se nenhum dos dois existe, tentar alternativas
            if not arquivo_a_existe and not arquivo_b_existe:
                # Tentar com sufixo _GRADES (como no CIMA)
                script_path_alternativo = os.path.join(pasta_pavimento, f"{nome}_GRADES.scr")
                print(f">>> Tentando caminho alternativo _GRADES: {script_path_alternativo}")
                if os.path.exists(script_path_alternativo):
                    print(f">>> Script encontrado no caminho alternativo _GRADES!")
                    script_content = self._ler_e_filtrar_script(script_path_alternativo)
                    if script_content:
                        print(f">>> Script GRADES gerado com sucesso!")
                        return script_content
                else:
                    # Tentar listar todos os arquivos .scr na pasta para debug
                    print(f">>> Verificando todos os arquivos .scr na pasta:")
                    try:
                        for arquivo in os.listdir(pasta_pavimento):
                            if arquivo.endswith('.scr'):
                                print(f">>>   - {arquivo}")
                    except Exception as e:
                        print(f">>> Erro ao listar arquivos: {e}")
                
                print(f">>> ERRO: Nenhum script GRADES encontrado!")
                return None
            
            # Combinar os scripts A e B
            script_content_combined = ""
            
            # Ler arquivo A se existir
            if arquivo_a_existe:
                print(f">>> Lendo arquivo A: {script_path_a}")
                script_content_a = self._ler_e_filtrar_script(script_path_a)
                if script_content_a:
                    script_content_combined += script_content_a
                    print(f">>> Arquivo A adicionado ({len(script_content_a)} caracteres)")
                else:
                    print(f">>> ERRO: Falha ao ler arquivo A")
            
            # Ler arquivo B se existir e aplicar deslocamento de 500cm para a direita
            if arquivo_b_existe:
                print(f">>> Lendo arquivo B: {script_path_b}")
                script_content_b = self._ler_e_filtrar_script(script_path_b)
                if script_content_b:
                    # Aplicar deslocamento de 500cm para a direita no script B
                    print(f">>> Aplicando deslocamento de 500cm para a direita no script B...")
                    script_content_b_deslocado = self._aplicar_deslocamento_coordenadas(script_content_b, 500.0, 0.0)
                    script_content_combined += script_content_b_deslocado
                    print(f">>> Arquivo B adicionado com deslocamento ({len(script_content_b_deslocado)} caracteres)")
                else:
                    print(f">>> ERRO: Falha ao ler arquivo B")
            
            if not script_content_combined:
                print(f">>> ERRO: Nenhum conteúdo válido encontrado nos arquivos GRADES")
                return None

            print(f">>> Script GRADES combinado gerado com sucesso! ({len(script_content_combined)} caracteres)")
            return script_content_combined

        except Exception as e:
            print(f">>> ERRO ao gerar script GRADES: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _aplicar_deslocamento_coordenadas(self, script_content, deslocamento_x, deslocamento_y):
        """
        Aplica deslocamento nas coordenadas do script, similar aos ordenadores.
        
        Args:
            script_content (str): Conteúdo do script
            deslocamento_x (float): Deslocamento no eixo X
            deslocamento_y (float): Deslocamento no eixo Y
            
        Returns:
            str: Script com coordenadas deslocadas
        """
        import re
        
        print(f">>> Aplicando deslocamento: X={deslocamento_x}, Y={deslocamento_y}")
        
        # Dividir o script em linhas
        linhas = script_content.split('\n')
        novas_linhas = []
        
        for i, linha in enumerate(linhas):
            # Ignora linhas com PD: ou NÍVEL DE CHEGADA
            if any(padrao in linha for padrao in ['PD:', 'NÍVEL DE CHEGADA']):
                nova_linha = linha
            else:
                # Verifica se a quarta linha acima contém "_TEXT"
                if i >= 4 and "_TEXT" in linhas[i-4].upper():
                    nova_linha = linha  # Não processa se _TEXT está 4 linhas acima
                else:
                    # Aplicar deslocamento nas coordenadas usando regex
                    nova_linha = re.sub(
                        r'(-?\d+\.?\d*),(-?\d+\.?\d*)',
                        lambda m: f"{float(m.group(1)) + deslocamento_x:.4f},{float(m.group(2)) + deslocamento_y:.4f}",
                        linha
                    )
            novas_linhas.append(nova_linha)
        
        # Reconstruir o script
        script_deslocado = '\n'.join(novas_linhas)
        print(f">>> Deslocamento aplicado com sucesso")
        
        return script_deslocado

    def _obter_script_item_cima(self, numero, dados):
        """
        Obtém o script CIMA para um item sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_item_cima, incluindo sistema de créditos.
        Usa a mesma lógica de _dcad_item_cima: preencher_campos_diretamente_e_gerar_scripts (sem Excel temporário).
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT ITEM CIMA")
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="CIMA"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Limpar pasta do pavimento
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_CIMA")
            
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                except Exception as e:
                    print(f">>> ⚠️ Erro ao excluir pasta: {str(e)}")
            
            # Criar pasta se não existir
            os.makedirs(pasta_pavimento, exist_ok=True)

            # Gerar script usando a mesma lógica de _dcad_item_cima (sem Excel temporário)
            try:
                # Importar a função com múltiplos fallbacks para ambiente frozen
                import sys
                
                try:
                    from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                except ImportError:
                    try:
                        from interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                    except ImportError:
                        try:
                            from src.interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                        except ImportError:
                            # Último fallback: adicionar ao path e tentar novamente
                            current_file = os.path.abspath(__file__)
                            utils_dir = os.path.dirname(current_file)
                            interfaces_dir = os.path.join(utils_dir, '..', 'interfaces')
                            interfaces_dir = os.path.abspath(interfaces_dir)
                            if interfaces_dir not in sys.path:
                                sys.path.insert(0, interfaces_dir)
                            from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                
                # Preparar dados do pilar (mesma lógica de _dcad_item_cima)
                dados_pilar = {
                    'interface_principal': self,  # Interface principal completa
                    'pavimento': pavimento,
                    'nome': nome,
                    'comprimento': dados.get('comprimento', ''),
                    'largura': dados.get('largura', ''),
                    'parafuso_p1_p2': int(float(dados.get('parafusos', {}).get('par_1_2', 0) or 0)),
                    'parafuso_p2_p3': int(float(dados.get('parafusos', {}).get('par_2_3', 0) or 0)),
                    'parafuso_p3_p4': int(float(dados.get('parafusos', {}).get('par_3_4', 0) or 0)),
                    'parafuso_p4_p5': int(float(dados.get('parafusos', {}).get('par_4_5', 0) or 0)),
                    'parafuso_p5_p6': int(float(dados.get('parafusos', {}).get('par_5_6', 0) or 0)),
                    'parafuso_p6_p7': int(float(dados.get('parafusos', {}).get('par_6_7', 0) or 0)),
                    'parafuso_p7_p8': int(float(dados.get('parafusos', {}).get('par_7_8', 0) or 0)),
                    'parafuso_p8_p9': int(float(dados.get('parafusos', {}).get('par_8_9', 0) or 0)),
                    'grades_grupo1': dados.get('grades', {}),
                    'detalhes_grades': dados.get('detalhes_grades', {}),
                    'grades_grupo2': dados.get('grades_grupo2', {}),
                    'detalhes_grades_grupo2': dados.get('detalhes_grades_grupo2', {}),
                    'pilar_especial': dados.get('pilar_especial', {}),
                    'ativar_pilar_especial': dados.get('pilar_especial', {}).get('ativar_pilar_especial', False),
                    'tipo_pilar_especial': dados.get('pilar_especial', {}).get('tipo_pilar_especial', ''),
                    'comp_1': dados.get('pilar_especial', {}).get('comprimentos', {}).get('comp_1', ''),
                    'comp_2': dados.get('pilar_especial', {}).get('comprimentos', {}).get('comp_2', ''),
                    'comp_3': dados.get('pilar_especial', {}).get('comprimentos', {}).get('comp_3', ''),
                    'larg_1': dados.get('pilar_especial', {}).get('larguras', {}).get('larg_1', ''),
                    'larg_2': dados.get('pilar_especial', {}).get('larguras', {}).get('larg_2', ''),
                    'larg_3': dados.get('pilar_especial', {}).get('larguras', {}).get('larg_3', ''),
                    'parafusos_especiais': {
                        'parafusos_a': self._obter_parafusos_especiais_para_transferencia('a'),
                        'parafusos_e': self._obter_parafusos_especiais_para_transferencia('e')
                    }
                }
                
                # Chamar a função diretamente (sem Excel temporário)
                preencher_campos_diretamente_e_gerar_scripts(dados_pilar)
                
                print(f">>> ✅ Função CIMA executada com sucesso")
                
            except Exception as e:
                print(f">>> ERRO na execução direta do CIMA: {str(e)}")
                import traceback
                traceback.print_exc()
                return None

            # Obter caminho do script gerado (mesma lógica de _dcad_item_cima)
            import unicodedata
            def _normalizar_nome_pasta_local(texto):
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(ch for ch in texto_normalizado if unicodedata.category(ch) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                return texto_limpo.strip('_')

            pav_norm = _normalizar_nome_pasta_local(pavimento)
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pav_norm}_CIMA")
            
            # Verificar qual script usar (combinado ou normal)
            script_combinado_path = os.path.join(pasta_pavimento, f"{nome}_COMBINADO_CIMA.scr")
            script_normal_path = os.path.join(pasta_pavimento, f"{nome}_CIMA.scr")
            
            if os.path.exists(script_combinado_path):
                script_path = script_combinado_path
                print(f">>> Usando script combinado (pilar especial)")
            else:
                script_path = script_normal_path
                print(f">>> Usando script normal")
            
            if not os.path.exists(script_path):
                print(f">>> ERRO: Script não encontrado: {script_path}")
                return None

            # Ler e filtrar o script
            script_content = self._ler_e_filtrar_script(script_path)
            if not script_content:
                print(f">>> ERRO: Falha ao ler/filtrar script")
                return None

            print(f">>> ✅ Script CIMA obtido com sucesso ({len(script_content)} caracteres)")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script CIMA: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _obter_script_item_abcd(self, numero, dados):
        """
        Obtém o script ABCD para um item sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_item_abcd, incluindo sistema de créditos.
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT ITEM ABCD")
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="ABCD"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Limpar pasta do pavimento
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                    os.makedirs(pasta_pavimento, exist_ok=True)
                except Exception as e:
                    print(f">>> ❌ Erro ao limpar pasta: {e}")

            # Gerar script usando a mesma lógica de _dcad_item_abcd
            if hasattr(self, 'conector_painel'):
                self.conector_painel.executar_abcd_excel()
            else:
                print(f">>> ERRO: Conector não disponível")
                return None

            # Obter caminho do script gerado
            script_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")
            
            # Verificar se é pilar especial
            pilar_especial_ativo = dados.get("pilar_especial", {}).get("ativar_pilar_especial", False)
            
            if pilar_especial_ativo:
                # PILAR ESPECIAL: Ler e concatenar 2 scripts
                script1_path = os.path.join(pasta_pavimento, f"{nome}_ABCD.scr")
                script2_path = os.path.join(pasta_pavimento, f"{nome}_ABCD2.scr")
                
                if not os.path.exists(script1_path) or not os.path.exists(script2_path):
                    print(f">>> ERRO: Scripts de pilar especial não encontrados")
                    return None
                
                script1_content = self._ler_e_filtrar_script(script1_path)
                script2_content = self._ler_e_filtrar_script(script2_path)
                
                if not script1_content or not script2_content:
                    print(f">>> ERRO: Falha ao ler/filtrar scripts de pilar especial")
                    return None
                
                script_content = script1_content + "\n" + script2_content
                print(f">>> ✅ Script ABCD (pilar especial) obtido com sucesso ({len(script_content)} caracteres)")
                return script_content
            else:
                # COMPORTAMENTO NORMAL: Ler 1 script apenas
                if not os.path.exists(script_path):
                    print(f">>> ERRO: Script não encontrado: {script_path}")
                    return None

                script_content = self._ler_e_filtrar_script(script_path)
                if not script_content:
                    print(f">>> ERRO: Falha ao ler/filtrar script")
                    return None

                print(f">>> ✅ Script ABCD obtido com sucesso ({len(script_content)} caracteres)")
                return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script ABCD: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _obter_script_item_grades(self, numero, dados):
        """
        Obtém o script GRADES para um item sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_item_grades, incluindo sistema de créditos.
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT ITEM GRADES")
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos (1 comum = 1 crédito, 1 especial = 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área do item
                area_item = credit_manager.calcular_area_pilar(dados)
                
                # Obter informações do item
                obra = dados.get('obra', '')
                pavimento_item = dados.get('pavimento', pavimento)
                nivel_saida = dados.get('nivel_saida', '')
                nivel_chegada = dados.get('nivel_chegada', '')
                numero_item = dados.get('numero', numero)
                nome_item = dados.get('nome', nome)
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento_item,
                    nivel_saida=nivel_saida,
                    nivel_chegada=nivel_chegada,
                    numero_item=numero_item,
                    nome=nome_item,
                    tipo=tipo_item,
                    area_m2=area_item,
                    parte_desenho="GRADES"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Limpar pasta do pavimento
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
            
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                    os.makedirs(pasta_pavimento, exist_ok=True)
                except Exception as e:
                    print(f">>> ❌ Erro ao limpar pasta: {e}")

            # Gerar script usando a mesma lógica de _dcad_item_grades
            if hasattr(self, 'conector_painel'):
                self.conector_painel.executar_grade_excel()
            else:
                print(f">>> ERRO: Conector não disponível")
                return None

            # Obter caminhos dos scripts gerados
            script_path_a = os.path.join(pasta_pavimento, f"{nome}.A.scr")
            script_path_b = os.path.join(pasta_pavimento, f"{nome}.B.scr")
            script_path_e = os.path.join(pasta_pavimento, f"{nome}.E.scr")
            script_path_f = os.path.join(pasta_pavimento, f"{nome}.F.scr")
            script_path_g = os.path.join(pasta_pavimento, f"{nome}.G.scr")
            script_path_h = os.path.join(pasta_pavimento, f"{nome}.H.scr")
            
            # Verificar quais arquivos existem
            arquivo_a_existe = os.path.exists(script_path_a)
            arquivo_b_existe = os.path.exists(script_path_b)
            arquivo_e_existe = os.path.exists(script_path_e)
            arquivo_f_existe = os.path.exists(script_path_f)
            arquivo_g_existe = os.path.exists(script_path_g)
            arquivo_h_existe = os.path.exists(script_path_h)
            
            pilar_especial_ativo = arquivo_e_existe or arquivo_f_existe
            
            # Combinar os scripts
            script_content = ""
            
            if pilar_especial_ativo:
                # Modo pilar especial: combinar todos os 6 arquivos
                arquivos_para_combinar = [
                    ('A', script_path_a, arquivo_a_existe),
                    ('B', script_path_b, arquivo_b_existe),
                    ('E', script_path_e, arquivo_e_existe),
                    ('F', script_path_f, arquivo_f_existe),
                    ('G', script_path_g, arquivo_g_existe),
                    ('H', script_path_h, arquivo_h_existe),
                ]
            else:
                # Modo normal: combinar apenas A e B
                arquivos_para_combinar = [
                    ('A', script_path_a, arquivo_a_existe),
                    ('B', script_path_b, arquivo_b_existe),
                ]
            
            # Ler e combinar cada arquivo
            for letra, caminho, existe in arquivos_para_combinar:
                if existe:
                    script_conteudo = self._ler_e_filtrar_script(caminho)
                    if script_conteudo:
                        script_content += script_conteudo
                    else:
                        print(f">>> ERRO: Falha ao ler arquivo {letra}")
            
            if not script_content:
                # Tentar caminho alternativo
                script_path_alternativo = os.path.join(pasta_pavimento, f"{nome}_GRADES.scr")
                if os.path.exists(script_path_alternativo):
                    script_content = self._ler_e_filtrar_script(script_path_alternativo)
                    if not script_content:
                        print(f">>> ERRO: Falha ao ler/filtrar script alternativo")
                        return None
                else:
                    print(f">>> ERRO: Nenhum script GRADES encontrado")
                    return None

            print(f">>> ✅ Script GRADES obtido com sucesso ({len(script_content)} caracteres)")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script GRADES: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _desenhar_1_item_completo(self):
        """
        Desenha 1 item completo - executa os 3 scripts (CIMA, ABCD, GRADES) sequencialmente e unifica.
        Usa as funções D.CAD 1 item que já existem, mas retorna o script sem enviar ao CAD.
        Cada função debita seus próprios créditos (1 crédito por função = 3 créditos total).
        No final, junta tudo e envia uma única vez ao CAD.
        """
        try:
            print(f"\n>>> DESENHAR 1 ITEM COMPLETO")
            
            # Obter informações do item selecionado (com verificação segura)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                messagebox.showerror("Erro", "Nenhum item selecionado")
                return
            
            pavimento = dados.get("pavimento", "")
            nome = dados.get("nome", numero)
            
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Nome do item: {nome}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão 'Desenhar 1 Item Completo' não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 PRÉ-APROVAÇÃO DE CRÉDITOS (MULTIPLICAR POR 3)
            # ========================================================
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is not None:
                # Verificar se é especial
                eh_especial = self._verificar_se_item_especial(dados)
                tipo_item = "especial" if eh_especial else "comum"
                quantidade_especiais = 1 if eh_especial else 0
                
                # Calcular créditos para 1 tipo (CIMA, ABCD ou GRADES)
                creditos_1_tipo = credit_manager.calcular_creditos_necessarios(
                    1, 
                    "item_simples", 
                    quantidade_especiais=quantidade_especiais
                )
                
                # Multiplicar por 3 (CIMA, ABCD, GRADES)
                creditos_necessarios_pre = creditos_1_tipo * 3
                
                # Verificar saldo disponível
                sucesso_saldo, saldo_atual = credit_manager.consultar_saldo()
                if not sucesso_saldo:
                    messagebox.showerror("Erro", "Não foi possível verificar o saldo de créditos.")
                    return
                
                if float(saldo_atual) < creditos_necessarios_pre:
                    messagebox.showerror(
                        "Créditos Insuficientes",
                        f"Créditos insuficientes para executar esta operação.\n\n"
                        _get_obf_str("credit")
                        _get_obf_str("saldo")
                        _get_obf_str("credit")
                    )
                    return
                
                print(_get_obf_str("credit"))
                print(_get_obf_str("saldo"))
                print(f">>> [PRÉ-APROVAÇÃO] ✅ Saldo suficiente, prosseguindo...")

            # Proteção contra execução simultânea
            if hasattr(self, 'item_completo_lock') and self.item_completo_lock:
                print(">>> Processamento de item completo já em andamento. Aguarde...")
                return
            
            self.item_completo_lock = True
            
            try:
                # Obter scripts sequencialmente usando as funções que já existem
                # Cada função debita seus próprios créditos automaticamente
                print(f"\n>>> Iniciando obtenção sequencial dos 3 scripts...")
                print(f">>> Cada função debitará seus próprios créditos (1 crédito por função)")
                
                # Obter scripts (isso gera os arquivos .scr)
                print(f"\n>>> [1/3] Gerando script CIMA...")
                self._obter_script_item_cima(numero, dados)
                
                print(f"\n>>> [2/3] Gerando script ABCD...")
                self._obter_script_item_abcd(numero, dados)
                
                print(f"\n>>> [3/3] Gerando script GRADES...")
                self._obter_script_item_grades(numero, dados)
                
                # Obter caminhos dos scripts gerados
                robo_dir = self._get_project_root()
                import unicodedata
                def _normalizar_nome_pasta_local(texto):
                    if not texto:
                        return ""
                    texto = str(texto).strip()
                    if not texto:
                        return ""
                    try:
                        texto_normalizado = unicodedata.normalize('NFD', texto)
                        texto_sem_acentos = ''.join(ch for ch in texto_normalizado if unicodedata.category(ch) != 'Mn')
                    except Exception:
                        texto_sem_acentos = texto
                    texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                    texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                    while '__' in texto_limpo:
                        texto_limpo = texto_limpo.replace('__', '_')
                    return texto_limpo.strip('_')
                
                pav_norm = _normalizar_nome_pasta_local(pavimento)
                
                # Caminhos dos scripts
                pasta_cima = os.path.join(robo_dir, "output", "scripts", f"{pav_norm}_CIMA")
                pasta_abcd = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
                pasta_grades = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
                
                # Verificar scripts CIMA (pode ser especial ou comum)
                script_cima_path = None
                script_cima_1_path = None
                script_cima_2_path = None
                script_cima_combinado = os.path.join(pasta_cima, f"{nome}_COMBINADO_CIMA.scr")
                script_cima_normal = os.path.join(pasta_cima, f"{nome}_CIMA.scr")
                script_cima_1 = os.path.join(pasta_cima, f"{nome}_CIMA_1.scr")
                script_cima_2 = os.path.join(pasta_cima, f"{nome}_CIMA_2.scr")
                
                # Verificar se há scripts CIMA separados (1 e 2)
                if os.path.exists(script_cima_1) and os.path.exists(script_cima_2):
                    script_cima_1_path = script_cima_1
                    script_cima_2_path = script_cima_2
                    script_cima_path = script_cima_1_path  # Usar o primeiro como referência principal
                elif os.path.exists(script_cima_combinado):
                    script_cima_path = script_cima_combinado
                elif os.path.exists(script_cima_normal):
                    script_cima_path = script_cima_normal
                
                # Verificar scripts ABCD (pode ser especial ou comum)
                # Padrão de nomenclatura: {nome}_ABCD.scr (script 1) e {nome}_ABCD2.scr (script 2)
                script_abcd_path = None
                script_abcd_2_path = None
                script_abcd_1 = os.path.join(pasta_abcd, f"{nome}_ABCD.scr")
                script_abcd_2 = os.path.join(pasta_abcd, f"{nome}_ABCD2.scr")
                
                # Verificar se é item especial (tem ambos os scripts)
                if os.path.exists(script_abcd_1) and os.path.exists(script_abcd_2):
                    # Item especial: tem ambos os scripts
                    script_abcd_path = script_abcd_1
                    script_abcd_2_path = script_abcd_2
                    print(f">>> Item especial ABCD detectado: {script_abcd_1} e {script_abcd_2}")
                elif os.path.exists(script_abcd_1):
                    # Item comum: apenas script 1
                    script_abcd_path = script_abcd_1
                    print(f">>> Item comum ABCD: {script_abcd_1}")
                
                # Verificar scripts GRADES (pode ter A, B, E, F, G, H) - ordenar corretamente
                ordem_grades = {'A': 1, 'B': 2, 'E': 3, 'F': 4, 'G': 5, 'H': 6}
                scripts_grades_paths = []
                for letra in ['A', 'B', 'E', 'F', 'G', 'H']:
                    script_grade_path = os.path.join(pasta_grades, f"{nome}.{letra}.scr")
                    if os.path.exists(script_grade_path):
                        scripts_grades_paths.append((ordem_grades[letra], script_grade_path))
                
                # Ordenar scripts GRADES pela ordem correta
                scripts_grades_paths.sort(key=lambda x: x[0])
                scripts_grades_paths = [path for _, path in scripts_grades_paths]
                
                # Verificar se pelo menos um script foi gerado
                scripts_encontrados = []
                # Adicionar scripts CIMA (pode ser separados ou combinado/normal)
                if script_cima_1_path and script_cima_2_path:
                    scripts_encontrados.append(('CIMA1', script_cima_1_path))
                    scripts_encontrados.append(('CIMA2', script_cima_2_path))
                elif script_cima_path:
                    scripts_encontrados.append(('CIMA', script_cima_path))
                # Adicionar scripts ABCD
                if script_abcd_path:
                    scripts_encontrados.append(('ABCD', script_abcd_path))
                    if script_abcd_2_path:
                        scripts_encontrados.append(('ABCD2', script_abcd_2_path))
                # Adicionar scripts GRADES
                if scripts_grades_paths:
                    for path in scripts_grades_paths:
                        scripts_encontrados.append(('GRADES', path))
                
                if not scripts_encontrados:
                    messagebox.showerror("Erro", "Nenhum script foi gerado com sucesso.\n\nVerifique se o item foi processado corretamente.")
                    return
                
                print(f"\n>>> ✅ {len(scripts_encontrados)} script(s) encontrado(s)")
                
                # Combinar todos os scripts em um único conteúdo (SEM FILTRO para PAZ)
                print(f"\n>>> Combinando todos os scripts em um único conteúdo (SEM FILTRO para PAZ)...")
                script_content_completo = ""
                script_content_completo_filtrado = ""  # Versão filtrada para execução direta no CAD
                
                # Função auxiliar para ler script sem filtro
                def _ler_script_sem_filtro(script_path):
                    """Lê o script sem aplicar filtro"""
                    try:
                        for encoding in ['utf-16-le', 'utf-16', 'utf-8', 'latin-1', 'cp1252']:
                            try:
                                with open(script_path, "r", encoding=encoding) as f:
                                    return f.read()
                            except (UnicodeDecodeError, Exception):
                                continue
                        return None
                    except Exception as e:
                        print(f">>> ⚠️ Erro ao ler script sem filtro: {e}")
                        return None
                
                for tipo, script_path in scripts_encontrados:
                    print(f">>> Lendo script {tipo}: {script_path}")
                    # Ler SEM filtro para PAZ
                    script_content_sem_filtro = _ler_script_sem_filtro(script_path)
                    # Ler COM filtro para execução direta no CAD
                    script_content_filtrado = self._ler_e_filtrar_script(script_path)
                    
                    if script_content_sem_filtro:
                        # Adicionar ao conteúdo completo SEM filtro (para PAZ)
                        if script_content_completo:  # Se já tem conteúdo, adicionar linha em branco antes
                            script_content_completo += "\n"
                        script_content_completo += script_content_sem_filtro
                        print(f">>> ✅ Script {tipo} adicionado SEM filtro ({len(script_content_sem_filtro)} caracteres)")
                    else:
                        print(f">>> ⚠️ Falha ao ler script {tipo} (sem filtro)")
                    
                    if script_content_filtrado:
                        # Adicionar ao conteúdo completo COM filtro (para execução direta no CAD)
                        if script_content_completo_filtrado:  # Se já tem conteúdo, adicionar linha em branco antes
                            script_content_completo_filtrado += "\n"
                        script_content_completo_filtrado += script_content_filtrado
                        print(f">>> ✅ Script {tipo} adicionado COM filtro ({len(script_content_filtrado)} caracteres)")
                    else:
                        print(f">>> ⚠️ Falha ao ler/filtrar script {tipo}")
                
                if not script_content_completo:
                    messagebox.showerror("Erro", "Nenhum script foi lido com sucesso.\n\nVerifique se os scripts foram gerados corretamente.")
                    return
                
                # Atualizar script_PAZ.scr com apenas o comando SCRIPT apontando para o script CIMA
                print(f"\n>>> Atualizando script_PAZ.scr com comando SCRIPT para script CIMA...")
                # Em frozen mode, usar diretório do executável (não _MEIPASS)
                is_frozen = getattr(sys, 'frozen', False)
                if is_frozen:
                    exe_dir = os.path.dirname(sys.executable)
                    output_dir = os.path.join(exe_dir, "output")
                else:
                    robo_dir = self._get_project_root()
                    output_dir = os.path.join(robo_dir, "output")
                os.makedirs(output_dir, exist_ok=True)
                script_paz_path = os.path.join(output_dir, "script_PAZ.scr")
                
                # Verificar se o script CIMA foi gerado
                # Para script_PAZ.scr, usar sempre o primeiro script CIMA (não o segundo)
                script_cima_para_paz = None
                if script_cima_1_path:
                    # Se houver scripts separados, usar o primeiro
                    script_cima_para_paz = script_cima_1_path
                elif script_cima_path:
                    # Usar o script CIMA encontrado (normal ou combinado)
                    script_cima_para_paz = script_cima_path
                
                if not script_cima_para_paz:
                    messagebox.showerror("Erro", "Script CIMA não foi gerado.\n\nVerifique se o item foi processado corretamente.")
                    return
                
                # Obter o caminho absoluto completo do script CIMA
                script_cima_abs_path = os.path.abspath(script_cima_para_paz)
                
                # Criar o conteúdo do script_PAZ.scr com apenas o comando SCRIPT
                script_paz_content = f";\nSCRIPT\n{script_cima_abs_path}\n"
                
                # Salvar apenas o comando SCRIPT no script_PAZ.scr
                try:
                    with open(script_paz_path, 'wb') as f:
                        # Adicionar BOM UTF-16 LE
                        f.write(b'\xFF\xFE')
                        # Converter conteúdo para UTF-16 LE
                        f.write(script_paz_content.encode('utf-16-le'))
                    
                    print(f">>> ✅ script_PAZ.scr atualizado com comando SCRIPT!")
                    print(f">>>    Arquivo: {script_paz_path}")
                    print(f">>>    Script CIMA: {script_cima_abs_path}")
                except Exception as e:
                    print(f">>> ❌ ERRO ao salvar script_PAZ.scr: {str(e)}")
                    import traceback
                    traceback.print_exc()
                
                # Encadear os scripts: CIMA -> ABCD -> GRADES
                # Para itens especiais: comando SCRIPT apenas no segundo script de cada tipo
                print(f"\n>>> Encadeando scripts individuais...")
                
                # Verificar se é item especial (tem COMBINADO_CIMA, CIMA_1/CIMA_2, ou ABCD_1/ABCD_2)
                eh_item_especial = (
                    os.path.exists(script_cima_combinado) or 
                    (script_cima_1_path and script_cima_2_path) or
                    (script_abcd_path and script_abcd_2_path)
                )
                
                # 1. CIMA -> ABCD
                # Para item especial: adicionar SCRIPT apenas no segundo script CIMA (COMBINADO ou CIMA_2)
                # Para item comum: adicionar SCRIPT no CIMA normal
                if script_cima_path and script_abcd_path:
                    script_cima_para_encadear = None
                    if eh_item_especial:
                        # Item especial: usar o segundo script CIMA
                        if script_cima_2_path:
                            # Tem scripts CIMA separados (1 e 2)
                            script_cima_para_encadear = script_cima_2_path
                            print(f">>> Item especial detectado: usando CIMA_2 para encadeamento")
                        elif os.path.exists(script_cima_combinado):
                            # Tem COMBINADO_CIMA
                            script_cima_para_encadear = script_cima_combinado
                            print(f">>> Item especial detectado: usando COMBINADO_CIMA para encadeamento")
                        else:
                            # Item especial mas sem segundo script identificado, usar o normal
                            script_cima_para_encadear = script_cima_path
                            print(f">>> Item especial detectado mas sem segundo script: usando CIMA normal")
                    else:
                        # Item comum: usar CIMA normal
                        script_cima_para_encadear = script_cima_path
                        print(f">>> Item comum: usando CIMA normal")
                    
                    if script_cima_para_encadear:
                        # CIMA -> ABCD_1 (sempre aponta para o primeiro ABCD)
                        self._adicionar_comando_script_ao_final(script_cima_para_encadear, script_abcd_path)
                        print(f">>> ✅ CIMA -> ABCD_1")
                
                # 2. ABCD_1 -> ABCD_2 (apenas para itens especiais)
                # Para item especial: encadear ABCD_1 -> ABCD_2
                if eh_item_especial and script_abcd_path and script_abcd_2_path:
                    self._adicionar_comando_script_ao_final(script_abcd_path, script_abcd_2_path)
                    print(f">>> ✅ ABCD_1 -> ABCD_2")
                
                # 3. ABCD -> GRADES (primeiro script de GRADES)
                # Para item especial: adicionar SCRIPT apenas no ABCD_2
                # Para item comum: adicionar SCRIPT no ABCD normal
                if script_abcd_path and scripts_grades_paths:
                    primeiro_grade = scripts_grades_paths[0]  # Primeiro script de GRADES
                    script_abcd_para_encadear = None
                    
                    if eh_item_especial and script_abcd_2_path:
                        # Item especial: usar ABCD_2 para apontar para GRADES
                        script_abcd_para_encadear = script_abcd_2_path
                        print(f">>> Item especial: usando ABCD_2 para encadeamento com GRADES")
                    else:
                        # Item comum: usar ABCD normal
                        script_abcd_para_encadear = script_abcd_path
                        print(f">>> Item comum: usando ABCD normal")
                    
                    if script_abcd_para_encadear:
                        self._adicionar_comando_script_ao_final(script_abcd_para_encadear, primeiro_grade)
                        print(f">>> ✅ ABCD -> GRADES (primeiro)")
                
                # 4. Encadear scripts GRADES entre si (A -> B -> E -> F -> G -> H)
                if len(scripts_grades_paths) > 1:
                    for i in range(len(scripts_grades_paths) - 1):
                        script_grade_atual = scripts_grades_paths[i]
                        script_grade_proximo = scripts_grades_paths[i + 1]
                        self._adicionar_comando_script_ao_final(script_grade_atual, script_grade_proximo)
                        print(f">>> ✅ GRADES[{i}] -> GRADES[{i+1}]")
                
                # Perguntar se deseja executar no AutoCAD
                opcao = self._perguntar_execucao_autocad("desenho item completo")
                
                if opcao is None:  # Cancelado
                    return
                
                if opcao == 1:
                    # Opção 1: Desenhar no CAD diretamente
                    # Executar o script completo combinado COM FILTRO no AutoCAD (copiar/colar)
                    if script_content_completo_filtrado:
                        self._executar_no_autocad(script_content_completo_filtrado, pavimento, "COMPLETO")
                        print(f">>> ✅ Script completo FILTRADO executado no AutoCAD com sucesso!")
                    elif script_content_completo:
                        # Fallback: usar versão sem filtro se a filtrada não estiver disponível
                        print(f">>> ⚠️ Usando versão sem filtro como fallback")
                        self._executar_no_autocad(script_content_completo, pavimento, "COMPLETO")
                        print(f">>> ✅ Script completo executado no AutoCAD com sucesso!")
                    else:
                        messagebox.showwarning("Aviso", "Script completo não encontrado. Use o comando PAZ no AutoCAD para executar os scripts.")
                elif opcao == 2:
                    # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                    print(f">>> ✅ Scripts gerados (não executados no AutoCAD)")
                elif opcao == 3:
                    # Opção 3: Gerar DFX (em desenvolvimento)
                    messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                    return
                
                
            except Exception as e:
                print(f">>> ❌ ERRO durante obtenção/unificação dos scripts: {str(e)}")
                import traceback
                traceback.print_exc()
                raise e
                
            finally:
                # Sempre liberar o lock
                self.item_completo_lock = False
                
        except Exception as e:
            print(f">>> ❌ ERRO ao executar desenhar 1 item completo: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar desenhar 1 item completo:\n{str(e)}")
            # Liberar lock em caso de erro
            if hasattr(self, 'item_completo_lock'):
                self.item_completo_lock = False

    def _gerar_script_pavimento_cima(self, pavimento):
        """Gera o script para um pavimento CIMA sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT PAVIMENTO CIMA")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho normal para gerar os scripts
            self.conector_painel.executar_todos_cima_excel()

            # Obter o caminho dos scripts combinados
            robo_dir = self._get_project_root()
            nome_pasta = f"{pavimento.replace(' ', '_')}_CIMA"
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Diretório combinados: {combinados_dir}")

            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não encontrada!")
                return None

            # Processar scripts combinados
            script_content = self._processar_scripts_combinados(combinados_dir)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados")
                return None

            print(f">>> Script CIMA do pavimento gerado com sucesso!")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao gerar script CIMA do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _gerar_script_pavimento_abcd(self, pavimento):
        """Gera o script para um pavimento ABCD sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT PAVIMENTO ABCD")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho normal para gerar os scripts
            self.conector_painel.executar_todos_abcd_excel()

            # Obter o caminho dos scripts combinados
            robo_dir = self._get_project_root()
            nome_pasta = f"{pavimento.replace(' ', '_')}_ABCD"
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Diretório combinados: {combinados_dir}")

            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não encontrada!")
                return None

            # Processar scripts combinados
            script_content = self._processar_scripts_combinados(combinados_dir)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados")
                return None

            print(f">>> Script ABCD do pavimento gerado com sucesso!")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao gerar script ABCD do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _gerar_script_pavimento_grades(self, pavimento):
        """Gera o script para um pavimento GRADES sem executar no CAD"""
        try:
            print(f"\n>>> GERANDO SCRIPT PAVIMENTO GRADES")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                print(f">>> ERRO: Pavimento 'Todos' não é válido")
                return None

            # Executar o desenho normal para gerar os scripts
            self.conector_painel.executar_todos_grade_excel()

            # Obter o caminho dos scripts combinados
            robo_dir = self._get_project_root()
            nome_pasta = f"{pavimento.replace(' ', '_')}_GRADES"
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")

            print(f">>> Pasta do pavimento: {pasta_pavimento}")
            print(f">>> Diretório combinados: {combinados_dir}")

            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não encontrada!")
                return None

            # Processar scripts combinados
            script_content = self._processar_scripts_combinados(combinados_dir)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados")
                return None

            print(f">>> Script GRADES do pavimento gerado com sucesso!")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao gerar script GRADES do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _obter_script_pavimento_cima(self, pavimento):
        """
        Obtém o script CIMA para um pavimento sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_pavimento_cima, incluindo sistema de créditos.
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT PAVIMENTO CIMA")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Obter todos os itens do pavimento
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            if not itens:
                print(f">>> ERRO: Nenhum item encontrado para o pavimento {pavimento}")
                return None

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                itens_dados = []
                for numero_item, dados_item in itens:
                    itens_dados.append(dados_item)
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_total = len(itens)
                quantidade_comuns = quantidade_total - quantidade_especiais
                
                # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    quantidade_total,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área total do pavimento
                area_total = credit_manager.calcular_area_pavimento(itens_dados)
                
                # Obter obra (do primeiro item)
                obra = itens_dados[0].get('obra', '') if itens_dados else ''
                
                print(f">>> Quantidade de itens no pavimento: {quantidade_total} ({quantidade_comuns} comuns, {quantidade_especiais} especiais)")
                print(_get_obf_str("credit"))
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento,
                    tipo="",  # Vazio para pavimento completo
                    area_m2=area_total,
                    parte_desenho=f"PAVIMENTO COMPLETO {quantidade_total} ITENS"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Usar a mesma lógica de _dcad_pavimento_cima para gerar scripts
            # Mas em vez de executar no CAD, retornar o script
            robo_dir = self._get_project_root()
            import unicodedata
            
            def normalizar_nome_pasta(texto):
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                return texto_limpo.strip('_')
            
            pavimento_normalizado = normalizar_nome_pasta(pavimento)
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento_normalizado}_CIMA")
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")
            
            # Limpar pasta antes de gerar scripts
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                except Exception as e:
                    print(f">>> ⚠️ Erro ao excluir pasta: {str(e)}")
            
            os.makedirs(pasta_pavimento, exist_ok=True)
            
            # Processar cada item individualmente (mesma lógica de _dcad_pavimento_cima)
            for numero_item, dados_item in itens:
                nome_item = dados_item.get("nome", numero_item)
                print(f">>> Processando item CIMA: {nome_item}")
                
                # Usar a mesma lógica de _dcad_item_cima para gerar script
                # Mas sem executar no CAD
                try:
                    from CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                except ImportError:
                    try:
                        from interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                    except ImportError:
                        from src.interfaces.CIMA_FUNCIONAL_EXCEL import preencher_campos_diretamente_e_gerar_scripts
                
                dados_pilar = {
                    'interface_principal': self,
                    'pavimento': pavimento,
                    'nome': nome_item,
                    'comprimento': dados_item.get('comprimento', ''),
                    'largura': dados_item.get('largura', ''),
                    'parafuso_p1_p2': int(float(dados_item.get('parafusos', {}).get('par_1_2', 0) or 0)),
                    'parafuso_p2_p3': int(float(dados_item.get('parafusos', {}).get('par_2_3', 0) or 0)),
                    'parafuso_p3_p4': int(float(dados_item.get('parafusos', {}).get('par_3_4', 0) or 0)),
                    'parafuso_p4_p5': int(float(dados_item.get('parafusos', {}).get('par_4_5', 0) or 0)),
                    'parafuso_p5_p6': int(float(dados_item.get('parafusos', {}).get('par_5_6', 0) or 0)),
                    'parafuso_p6_p7': int(float(dados_item.get('parafusos', {}).get('par_6_7', 0) or 0)),
                    'parafuso_p7_p8': int(float(dados_item.get('parafusos', {}).get('par_7_8', 0) or 0)),
                    'parafuso_p8_p9': int(float(dados_item.get('parafusos', {}).get('par_8_9', 0) or 0)),
                    'grades_grupo1': dados_item.get('grades', {}),
                    'detalhes_grades': dados_item.get('detalhes_grades', {}),
                    'grades_grupo2': dados_item.get('grades_grupo2', {}),
                    'detalhes_grades_grupo2': dados_item.get('detalhes_grades_grupo2', {}),
                    'pilar_especial': dados_item.get('pilar_especial', {}),
                    'ativar_pilar_especial': dados_item.get('pilar_especial', {}).get('ativar_pilar_especial', False),
                    'tipo_pilar_especial': dados_item.get('pilar_especial', {}).get('tipo_pilar_especial', ''),
                    'comp_1': dados_item.get('pilar_especial', {}).get('comprimentos', {}).get('comp_1', ''),
                    'comp_2': dados_item.get('pilar_especial', {}).get('comprimentos', {}).get('comp_2', ''),
                    'comp_3': dados_item.get('pilar_especial', {}).get('comprimentos', {}).get('comp_3', ''),
                    'larg_1': dados_item.get('pilar_especial', {}).get('larguras', {}).get('larg_1', ''),
                    'larg_2': dados_item.get('pilar_especial', {}).get('larguras', {}).get('larg_2', ''),
                    'larg_3': dados_item.get('pilar_especial', {}).get('larguras', {}).get('larg_3', ''),
                    'parafusos_especiais': {
                        'parafusos_a': self._obter_parafusos_especiais_para_transferencia('a'),
                        'parafusos_e': self._obter_parafusos_especiais_para_transferencia('e')
                    }
                }
                
                preencher_campos_diretamente_e_gerar_scripts(dados_pilar)
            
            # Executar combinador (mesma lógica de _dcad_pavimento_cima)
            print(f">>> Executando combinador CIMA...")
            is_frozen = getattr(self.conector_painel, '_is_frozen', lambda: False)()
            
            if is_frozen:
                try:
                    import importlib
                    module = None
                    import_paths = [
                        'robots.Combinador_de_SCR _cima',
                        'src.robots.Combinador_de_SCR _cima',
                        'Combinador_de_SCR _cima'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if module and hasattr(module, 'processar_arquivos'):
                        module.processar_arquivos(pasta_pavimento)
                        print(f">>> ✅ Combinador CIMA executado com sucesso (frozen)")
                    else:
                        print(f">>> ⚠️ Combinador CIMA não encontrado, tentando método antigo...")
                        self.conector_painel.combinar_codigos_cima()
                except Exception as e:
                    print(f">>> ⚠️ Erro ao executar combinador CIMA (frozen): {e}")
                    try:
                        self.conector_painel.combinar_codigos_cima()
                    except:
                        pass
            else:
                try:
                    import subprocess
                    import sys
                    combinador_path = os.path.join(robo_dir, "src", "robots", "Combinador_de_SCR _cima.py")
                    if os.path.exists(combinador_path):
                        resultado = subprocess.run(
                            [sys.executable, combinador_path, pasta_pavimento], 
                            text=True,
                            capture_output=True,
                            timeout=300
                        )
                        if resultado.returncode == 0:
                            print(f">>> ✅ Combinador CIMA executado com sucesso")
                        else:
                            print(f">>> ⚠️ Combinador CIMA retornou código {resultado.returncode}")
                            self.conector_painel.combinar_codigos_cima()
                    else:
                        self.conector_painel.combinar_codigos_cima()
                except Exception as e:
                    print(f">>> ⚠️ Erro ao executar combinador CIMA: {e}")
                    try:
                        self.conector_painel.combinar_codigos_cima()
                    except:
                        pass
            
            # Executar ordenador (mesma lógica de _dcad_pavimento_cima)
            print(f">>> Executando ordenador CIMA...")
            os.makedirs(combinados_dir, exist_ok=True)
            self._executar_ordenador_cima_automatico(combinados_dir)
            print(f">>> ✅ Ordenador CIMA executado com sucesso")
            
            # Verificar se a pasta Combinados foi criada
            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não foi criada: {combinados_dir}")
                return None
            
            # ========================================================
            # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
            # ========================================================
            self._limpar_linhas_em_branco_entre_semicolons(combinados_dir)
            
            # Processar scripts combinados (mesma lógica de _dcad_pavimento_cima)
            script_content = self._processar_scripts_combinados(combinados_dir)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados CIMA")
                return None

            print(f">>> ✅ Script CIMA do pavimento obtido com sucesso ({len(script_content)} caracteres)")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script CIMA do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _obter_script_pavimento_abcd(self, pavimento):
        """
        Obtém o script ABCD para um pavimento sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_pavimento_abcd, incluindo sistema de créditos.
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT PAVIMENTO ABCD")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Obter todos os itens do pavimento
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            if not itens:
                print(f">>> ERRO: Nenhum item encontrado para o pavimento {pavimento}")
                return None

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                itens_dados = []
                for numero_item, dados_item in itens:
                    itens_dados.append(dados_item)
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_total = len(itens)
                quantidade_comuns = quantidade_total - quantidade_especiais
                
                # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    quantidade_total,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área total do pavimento
                area_total = credit_manager.calcular_area_pavimento(itens_dados)
                
                # Obter obra (do primeiro item)
                obra = itens_dados[0].get('obra', '') if itens_dados else ''
                
                print(f">>> Quantidade de itens no pavimento: {quantidade_total} ({quantidade_comuns} comuns, {quantidade_especiais} especiais)")
                print(_get_obf_str("credit"))
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento,
                    tipo="",  # Vazio para pavimento completo
                    area_m2=area_total,
                    parte_desenho=f"PAVIMENTO COMPLETO {quantidade_total} ITENS"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Limpar pasta do pavimento
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_ABCD")
            
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                except Exception as e:
                    print(f">>> ⚠️ Erro ao excluir pasta: {str(e)}")
            
            os.makedirs(pasta_pavimento, exist_ok=True)

            # Processar cada item individualmente (mesma lógica de _dcad_pavimento_abcd)
            itens_processados = 0
            itens_com_erro = []
            
            print(f"\n>>> INICIANDO PROCESSAMENTO DE {len(itens)} ITENS ABCD...")
            for idx, (numero_item, dados_item) in enumerate(itens, 1):
                nome_item = dados_item.get("nome", numero_item)
                print(f">>> [{idx}/{len(itens)}] Processando item ABCD: {nome_item}")
                
                # Usar a mesma lógica de _dcad_item_abcd para gerar script
                # Mas sem executar no CAD
                try:
                    resultado = self._processar_item_abcd_para_pavimento(numero_item, dados_item)
                    if resultado:
                        itens_processados += 1
                        print(f">>> ✅ Item {nome_item} processado com sucesso")
                    else:
                        itens_com_erro.append(nome_item)
                        print(f">>> ❌ ERRO ao processar item {nome_item}")
                except Exception as e:
                    itens_com_erro.append(nome_item)
                    print(f">>> ❌ EXCEÇÃO ao processar item {nome_item}: {str(e)}")
                    import traceback
                    traceback.print_exc()
            
            print(f"\n>>> RESUMO DO PROCESSAMENTO ABCD:")
            print(f">>>   Total de itens: {len(itens)}")
            print(f">>>   Itens processados com sucesso: {itens_processados}")
            print(f">>>   Itens com erro: {len(itens_com_erro)}")
            
            if itens_processados == 0:
                print(f">>> ERRO: Nenhum item foi processado com sucesso!")
                return None
            
            # Executar combinador (mesma lógica de _dcad_pavimento_abcd)
            print(f">>> Executando combinador ABCD...")
            is_frozen = getattr(self.conector_painel, '_is_frozen', lambda: False)()
            
            if is_frozen:
                try:
                    import importlib
                    module = None
                    import_paths = [
                        'robots.Combinador_de_SCR',
                        'src.robots.Combinador_de_SCR',
                        'Combinador_de_SCR',
                        'automacaoexcel.Ordenamento.Combinador_de_SCR'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if module and hasattr(module, 'processar_arquivos'):
                        module.processar_arquivos(pasta_pavimento, mostrar_mensagem=False)
                        print(f">>> ✅ Combinador ABCD executado com sucesso (frozen)")
                    else:
                        print(f">>> ⚠️ Combinador ABCD não encontrado")
                        return None
                except Exception as e:
                    print(f">>> ERRO ao executar combinador ABCD (frozen): {e}")
                    return None
            else:
                try:
                    import subprocess
                    import sys
                    combinador_path = os.path.join(robo_dir, 'src', 'robots', "Combinador_de_SCR.py")
                    if os.path.exists(combinador_path):
                        resultado = subprocess.run(
                            [sys.executable, combinador_path, pasta_pavimento],
                            capture_output=True,
                            text=True,
                            timeout=300
                        )
                        if resultado.returncode == 0:
                            print(f">>> ✅ Combinador ABCD executado com sucesso")
                        else:
                            print(f">>> ERRO no combinador ABCD: {resultado.stderr}")
                            return None
                    else:
                        print(f">>> ERRO: Combinador ABCD não encontrado: {combinador_path}")
                        return None
                except Exception as e:
                    print(f">>> ERRO ao executar combinador ABCD: {e}")
                    return None
            
            # Executar ordenador (mesma lógica de _dcad_pavimento_abcd)
            print(f">>> Executando ordenador ABCD...")
            combinados_dir = os.path.join(pasta_pavimento, "Combinados")
            if not os.path.exists(combinados_dir):
                print(f">>> ERRO: Pasta Combinados não foi criada: {combinados_dir}")
                return None
            
            # ========================================================
            # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
            # ========================================================
            self._limpar_linhas_em_branco_entre_semicolons(combinados_dir)
            
            # Executar ordenador diretamente (mesma lógica de _dcad_pavimento_abcd)
            try:
                import sys
                robots_path = os.path.join(robo_dir, 'src', 'robots')
                if robots_path not in sys.path:
                    sys.path.insert(0, robots_path)
                
                from Ordenador_Cordenadas_abcd import ProcessadorCoordenadasABCD, get_config_path_abcd
                from pathlib import Path
                import json
                
                # Carregar configurações
                try:
                    cfg_path = get_config_path_abcd()
                    if cfg_path.exists():
                        with open(cfg_path, 'r', encoding='utf-8') as f:
                            cfg = json.load(f)
                    else:
                        cfg = {
                            "numero_colunas": 4,
                            "distancia_x_colunas": 1585,
                            "distancia_y_linhas": -1148.6,
                            "distancia_y_extra": 0,
                            "linhas_para_extra": 0
                        }
                except Exception as e:
                    cfg = {
                        "numero_colunas": 4,
                        "distancia_x_colunas": 1585,
                        "distancia_y_linhas": -1148.6,
                        "distancia_y_extra": 0,
                        "linhas_para_extra": 0
                    }
                
                # Criar processador com as configurações
                processador = ProcessadorCoordenadasABCD(
                    numero_colunas=cfg.get("numero_colunas", 4),
                    distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                    distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                    distancia_y_extra=cfg.get("distancia_y_extra", 0),
                    linhas_para_extra=cfg.get("linhas_para_extra", 0)
                )
                
                # Processar arquivos na pasta Combinados
                arquivos_scr = [f for f in os.listdir(combinados_dir) if f.endswith('.scr')]
                print(f">>> Arquivos encontrados na pasta Combinados: {len(arquivos_scr)}")
                
                for arquivo in arquivos_scr:
                    caminho_arquivo = os.path.join(combinados_dir, arquivo)
                    try:
                        processador.processar_arquivo(caminho_arquivo)
                    except Exception as e:
                        print(f">>> ⚠️ Erro ao processar arquivo {arquivo}: {e}")
                
                print(f">>> ✅ Ordenador ABCD executado com sucesso")
            except Exception as e:
                print(f">>> ⚠️ Erro ao executar ordenador ABCD: {e}")
                import traceback
                traceback.print_exc()
                # Continuar mesmo se o ordenador falhar
            
            # Processar scripts combinados (mesma lógica de _dcad_pavimento_abcd)
            script_content = self._processar_scripts_combinados(combinados_dir)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados ABCD")
                return None

            print(f">>> ✅ Script ABCD do pavimento obtido com sucesso ({len(script_content)} caracteres)")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script ABCD do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _obter_script_pavimento_grades(self, pavimento):
        """
        Obtém o script GRADES para um pavimento sem enviar ao CAD.
        Retorna o conteúdo do script ou None em caso de erro.
        Mantém toda a funcionalidade de _dcad_pavimento_grades, incluindo sistema de créditos.
        """
        try:
            print(f"\n>>> OBTENDO SCRIPT PAVIMENTO GRADES")
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                print(f">>> ERRO: Conector não disponível")
                return None

            # Obter todos os itens do pavimento
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            if not itens:
                print(f">>> ERRO: Nenhum item encontrado para o pavimento {pavimento}")
                return None

            # ========================================================
            # 💳 INTEGRAÇÃO COM SISTEMA DE CRÉDITOS
            # ========================================================
            
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is None:
                print("⚠️ Sistema de créditos não disponível - continuando sem controle de créditos")
            else:
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                itens_dados = []
                for numero_item, dados_item in itens:
                    itens_dados.append(dados_item)
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_total = len(itens)
                quantidade_comuns = quantidade_total - quantidade_especiais
                
                # Calcular créditos (comuns: 1 crédito, especiais: 2 créditos)
                creditos_necessarios = credit_manager.calcular_creditos_necessarios(
                    quantidade_total,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Calcular área total do pavimento
                area_total = credit_manager.calcular_area_pavimento(itens_dados)
                
                # Obter obra (do primeiro item)
                obra = itens_dados[0].get('obra', '') if itens_dados else ''
                
                print(f">>> Quantidade de itens no pavimento: {quantidade_total} ({quantidade_comuns} comuns, {quantidade_especiais} especiais)")
                print(_get_obf_str("credit"))
                
                # Gerar descrição detalhada
                descricao = credit_manager.gerar_descricao_detalhada(
                    obra=obra,
                    pavimento=pavimento,
                    tipo="",  # Vazio para pavimento completo
                    area_m2=area_total,
                    parte_desenho=f"PAVIMENTO COMPLETO {quantidade_total} ITENS"
                )
                
                sucesso_debito, mensagem = credit_manager.debitar_creditos_imediato(
                    creditos_necessarios, 
                    descricao
                )
                
                if not sucesso_debito:
                    print(f">>> ERRO: Não foi possível debitar créditos: {mensagem}")
                    return None
                
                print(_get_obf_str("credit"))
                self.atualizar_creditos_interface()

            # Limpar pasta do pavimento
            robo_dir = self._get_project_root()
            pasta_pavimento = os.path.join(robo_dir, "output", "scripts", f"{pavimento.replace(' ', '_')}_GRADES")
            
            if os.path.exists(pasta_pavimento):
                try:
                    import shutil
                    shutil.rmtree(pasta_pavimento)
                except Exception as e:
                    print(f">>> ⚠️ Erro ao excluir pasta: {str(e)}")
            
            os.makedirs(pasta_pavimento, exist_ok=True)

            # Processar cada item individualmente (mesma lógica de _dcad_pavimento_grades)
            for numero_item, dados_item in itens:
                nome_item = dados_item.get("nome", numero_item)
                print(f">>> Processando item GRADES: {nome_item}")
                
                # Usar a mesma lógica de _dcad_item_grades para gerar script
                # Mas sem executar no CAD
                resultado = self._processar_item_grades_para_pavimento(numero_item, dados_item)
                if not resultado:
                    print(f">>> ⚠️ Erro ao processar item GRADES: {nome_item}")
            
            # Executar combinador (mesma lógica de _dcad_pavimento_grades)
            is_frozen = self.conector_painel._is_frozen()
            
            if is_frozen:
                try:
                    import importlib
                    combinador_module = None
                    import_paths = [
                        'robots.Combinador_de_SCR_GRADES',
                        'src.robots.Combinador_de_SCR_GRADES',
                        'Combinador_de_SCR_GRADES'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            combinador_module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if combinador_module and hasattr(combinador_module, 'processar_arquivos'):
                        combinador_module.processar_arquivos(pasta_pavimento)
                except Exception as e:
                    print(f">>> ERRO ao executar combinador GRADES: {e}")
                    return None
            else:
                try:
                    import subprocess
                    import sys
                    combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR_GRADES.py")
                    resultado_combinador = subprocess.run(
                        [sys.executable, combinador_path, pasta_pavimento],
                        capture_output=True,
                        text=True,
                        timeout=300
                    )
                    
                    if resultado_combinador.returncode != 0:
                        print(f">>> ERRO no combinador GRADES: {resultado_combinador.stderr}")
                        return None
                except Exception as e:
                    print(f">>> ERRO ao executar combinador GRADES: {e}")
                    return None
            
            # Executar ordenador
            pasta_combinados = os.path.join(pasta_pavimento, "Combinados")
            self._executar_ordenador_grades_automatico(pasta_combinados)
            
            # Processar scripts combinados
            script_content = self._processar_scripts_combinados(pasta_combinados)
            if not script_content:
                print(f">>> ERRO: Falha ao processar scripts combinados GRADES")
                return None

            print(f">>> ✅ Script GRADES do pavimento obtido com sucesso ({len(script_content)} caracteres)")
            return script_content

        except Exception as e:
            print(f">>> ERRO ao obter script GRADES do pavimento: {str(e)}")
            import traceback
            traceback.print_exc()
            return None

    def _desenhar_1_pavimento_completo(self):
        """Desenha 1 pavimento completo - executa os 3 scripts (ABCD, CIMA, GRADES) sequencialmente e unifica"""
        try:
            print(f"\n>>> DESENHAR 1 PAVIMENTO COMPLETO")
            
            # Verificar se o conector está disponível
            if not hasattr(self, 'conector_painel'):
                messagebox.showwarning("Aviso", "Conector não disponível")
                return

            # Obter informações do pavimento selecionado (usar item selecionado como referência)
            numero, dados = self._get_item_selecionado_safe()
            if not dados:
                messagebox.showerror("Erro", "Nenhum item selecionado. Selecione um item na lista para usar como referência do pavimento.")
                return
            
            pavimento = dados.get("pavimento", "")
            
            print(f">>> Pavimento: {pavimento}")
            
            # Verificar se o pavimento não é "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão 'Desenhar 1 Pavimento Completo' não funciona quando 'Todos' está selecionado.")
                return

            # ========================================================
            # 💳 PRÉ-APROVAÇÃO DE CRÉDITOS (MULTIPLICAR POR 3)
            # ========================================================
            credit_manager = getattr(self, 'credit_manager', None)
            if credit_manager is not None:
                # Obter todos os itens do pavimento para calcular créditos
                itens = self.conector_painel.get_itens_por_pavimento(pavimento)
                if not itens:
                    messagebox.showerror("Erro", f"Nenhum item encontrado para o pavimento {pavimento}")
                    return
                
                # Contar itens especiais e comuns
                quantidade_especiais = 0
                for numero_item, dados_item in itens:
                    if self._verificar_se_item_especial(dados_item):
                        quantidade_especiais += 1
                
                quantidade_total = len(itens)
                quantidade_comuns = quantidade_total - quantidade_especiais
                
                # Calcular créditos para 1 tipo (CIMA, ABCD ou GRADES)
                creditos_1_tipo = credit_manager.calcular_creditos_necessarios(
                    quantidade_total,
                    "item_simples",
                    quantidade_especiais=quantidade_especiais
                )
                
                # Multiplicar por 3 (CIMA, ABCD, GRADES)
                creditos_necessarios_pre = creditos_1_tipo * 3
                
                # Verificar saldo disponível
                sucesso_saldo, saldo_atual = credit_manager.consultar_saldo()
                if not sucesso_saldo:
                    messagebox.showerror("Erro", "Não foi possível verificar o saldo de créditos.")
                    return
                
                if float(saldo_atual) < creditos_necessarios_pre:
                    messagebox.showerror(
                        "Créditos Insuficientes",
                        f"Créditos insuficientes para executar esta operação.\n\n"
                        _get_obf_str("credit")
                        _get_obf_str("saldo")
                        _get_obf_str("credit")
                        f"Pavimento: {quantidade_total} itens ({quantidade_comuns} comuns, {quantidade_especiais} especiais)"
                    )
                    return
                
                print(_get_obf_str("credit"))
                print(_get_obf_str("saldo"))
                print(f">>> [PRÉ-APROVAÇÃO] ✅ Saldo suficiente, prosseguindo...")

            # NOTA: Sistema de créditos é gerenciado individualmente por cada função
            # (_obter_script_pavimento_cima, _obter_script_pavimento_abcd, _obter_script_pavimento_grades)
            # Cada função debita seus próprios créditos antes de gerar o script
            print(f">>> Sistema de créditos será gerenciado individualmente por cada função D.CAD")

            # Proteção contra execução simultânea
            if hasattr(self, 'pavimento_completo_lock') and self.pavimento_completo_lock:
                print(">>> Processamento de pavimento completo já em andamento. Aguarde...")
                return
            
            self.pavimento_completo_lock = True
            
            # LIMPEZA DAS PASTAS ANTIGAS DO PAVIMENTO
            print(f">>> Iniciando limpeza das pastas antigas do pavimento '{pavimento}'...")
            
            # Obter o diretório raiz
            robo_dir = self._get_project_root()
            pastas_scripts_dir = os.path.join(robo_dir, "output", "scripts")
            
            # Definir os nomes das pastas que devem ser removidas
            pastas_para_remover = [
                f"{pavimento.replace(' ', '_')}_ABCD",
                f"{pavimento.replace(' ', '_')}_CIMA", 
                f"{pavimento.replace(' ', '_')}_GRADES"
            ]
            
            print(f">>> Pastas que serão verificadas e removidas:")
            for pasta in pastas_para_remover:
                print(f">>>   - {pasta}")
            
            # Verificar e remover cada pasta
            pastas_removidas = 0
            for nome_pasta in pastas_para_remover:
                caminho_pasta = os.path.join(pastas_scripts_dir, nome_pasta)
                
                if os.path.exists(caminho_pasta):
                    try:
                        print(f">>> Removendo pasta: {nome_pasta}")
                        import shutil
                        shutil.rmtree(caminho_pasta)
                        print(f">>> ✅ Pasta '{nome_pasta}' removida com sucesso")
                        pastas_removidas += 1
                    except Exception as e:
                        print(f">>> ⚠️ ERRO ao remover pasta '{nome_pasta}': {str(e)}")
                        # Continuar mesmo se uma pasta falhar
                else:
                    print(f">>> ℹ️ Pasta '{nome_pasta}' não existe, pulando...")
            
            print(f">>> Limpeza concluída: {pastas_removidas} pastas removidas")
            print(f">>> As pastas serão recriadas durante a geração dos scripts...")
            
            try:
                # Dicionário para armazenar os resultados
                resultados = {
                    'cima': None,
                    'abcd': None,
                    'grades': None
                }
                
                print(f">>> Iniciando geração sequencial dos 3 scripts de pavimento...")
                print(f">>> Usando funções D.CAD pavimento (sem enviar ao CAD)...")
                
                # Obter scripts sequencialmente usando as funções D.CAD (sem enviar ao CAD)
                # Cada função já debita seus próprios créditos
                print(f">>> Obtendo script CIMA do pavimento...")
                resultados['cima'] = self._obter_script_pavimento_cima(pavimento)
                print(f">>> Script CIMA: {'✅ Sucesso' if resultados['cima'] else '❌ Falha'}")
                
                print(f">>> Obtendo script ABCD do pavimento...")
                resultados['abcd'] = self._obter_script_pavimento_abcd(pavimento)
                print(f">>> Script ABCD: {'✅ Sucesso' if resultados['abcd'] else '❌ Falha'}")
                
                print(f">>> Obtendo script GRADES do pavimento...")
                resultados['grades'] = self._obter_script_pavimento_grades(pavimento)
                print(f">>> Script GRADES: {'✅ Sucesso' if resultados['grades'] else '❌ Falha'}")
                
                print(f">>> Todos os scripts de pavimento obtidos!")
                print(f">>> Resultados: CIMA={resultados['cima'] is not None}, ABCD={resultados['abcd'] is not None}, GRADES={resultados['grades'] is not None}")
                
                # Verificar se pelo menos um script foi gerado
                scripts_gerados = [s for s in resultados.values() if s is not None]
                if not scripts_gerados:
                    messagebox.showerror("Erro", "Nenhum script de pavimento foi gerado com sucesso.\n\nVerifique se o pavimento foi processado corretamente.")
                    return
                
                # Obter caminhos dos scripts combinados (1.scr, 2.scr, etc.) de cada tipo
                import unicodedata
                def _normalizar_nome_pasta_local(texto):
                    if not texto:
                        return ""
                    texto = str(texto).strip()
                    if not texto:
                        return ""
                    try:
                        texto_normalizado = unicodedata.normalize('NFD', texto)
                        texto_sem_acentos = ''.join(ch for ch in texto_normalizado if unicodedata.category(ch) != 'Mn')
                    except Exception:
                        texto_sem_acentos = texto
                    texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                    texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                    while '__' in texto_limpo:
                        texto_limpo = texto_limpo.replace('__', '_')
                    return texto_limpo.strip('_')
                
                pav_norm = _normalizar_nome_pasta_local(pavimento)
                pav_norm_espacos = pavimento.replace(' ', '_')
                
                # Pastas dos scripts combinados
                pasta_cima_combinados = os.path.join(robo_dir, "output", "scripts", f"{pav_norm}_CIMA", "Combinados")
                pasta_abcd_combinados = os.path.join(robo_dir, "output", "scripts", f"{pav_norm_espacos}_ABCD", "Combinados")
                pasta_grades_combinados = os.path.join(robo_dir, "output", "scripts", f"{pav_norm_espacos}_GRADES", "Combinados")
                
                # ========================================================
                # 🎯 TRATAMENTO FINAL: REMOVER LINHAS EM BRANCO ENTRE ";" NAS ÚLTIMAS 10 LINHAS
                # ========================================================
                print(f">>> [LIMPEZA FINAL] Aplicando limpeza final em todas as pastas Combinados...")
                
                # Limpar scripts CIMA
                if resultados['cima'] and os.path.exists(pasta_cima_combinados):
                    print(f">>> [LIMPEZA FINAL] Limpando scripts CIMA...")
                    self._limpar_linhas_em_branco_entre_semicolons(pasta_cima_combinados)
                
                # Limpar scripts ABCD
                if resultados['abcd'] and os.path.exists(pasta_abcd_combinados):
                    print(f">>> [LIMPEZA FINAL] Limpando scripts ABCD...")
                    self._limpar_linhas_em_branco_entre_semicolons(pasta_abcd_combinados)
                
                # Limpar scripts GRADES
                if resultados['grades'] and os.path.exists(pasta_grades_combinados):
                    print(f">>> [LIMPEZA FINAL] Limpando scripts GRADES...")
                    self._limpar_linhas_em_branco_entre_semicolons(pasta_grades_combinados)
                
                print(f">>> [LIMPEZA FINAL] ✅ Limpeza final concluída para todas as pastas!")
                
                # Função para encontrar o último script numerado em uma pasta
                def encontrar_ultimo_script(pasta):
                    if not os.path.exists(pasta):
                        return None
                    scripts = []
                    for f in os.listdir(pasta):
                        if f.endswith('.scr'):
                            try:
                                num = int(f.replace('.scr', ''))
                                scripts.append((num, os.path.join(pasta, f)))
                            except ValueError:
                                continue
                    if scripts:
                        scripts.sort(key=lambda x: x[0])
                        return scripts[-1][1]  # Retorna o caminho do último script
                    return None
                
                # Função para encontrar o primeiro script numerado em uma pasta
                def encontrar_primeiro_script(pasta):
                    if not os.path.exists(pasta):
                        return None
                    scripts = []
                    for f in os.listdir(pasta):
                        if f.endswith('.scr'):
                            try:
                                num = int(f.replace('.scr', ''))
                                scripts.append((num, os.path.join(pasta, f)))
                            except ValueError:
                                continue
                    if scripts:
                        scripts.sort(key=lambda x: x[0])
                        return scripts[0][1]  # Retorna o caminho do primeiro script
                    return None
                
                # Encontrar scripts
                print(f"\n>>> Procurando scripts combinados...")
                print(f">>> Pasta CIMA: {pasta_cima_combinados} (existe: {os.path.exists(pasta_cima_combinados)})")
                print(f">>> Pasta ABCD: {pasta_abcd_combinados} (existe: {os.path.exists(pasta_abcd_combinados)})")
                print(f">>> Pasta GRADES: {pasta_grades_combinados} (existe: {os.path.exists(pasta_grades_combinados)})")
                
                ultimo_cima = encontrar_ultimo_script(pasta_cima_combinados) if resultados['cima'] else None
                primeiro_abcd = encontrar_primeiro_script(pasta_abcd_combinados) if resultados['abcd'] else None
                ultimo_abcd = encontrar_ultimo_script(pasta_abcd_combinados) if resultados['abcd'] else None
                primeiro_grades = encontrar_primeiro_script(pasta_grades_combinados) if resultados['grades'] else None
                primeiro_cima = encontrar_primeiro_script(pasta_cima_combinados) if resultados['cima'] else None
                
                print(f"\n>>> Scripts encontrados:")
                print(f">>>   - Último CIMA: {ultimo_cima}")
                print(f">>>   - Primeiro ABCD: {primeiro_abcd}")
                print(f">>>   - Último ABCD: {ultimo_abcd}")
                print(f">>>   - Primeiro GRADES: {primeiro_grades}")
                print(f">>>   - Primeiro CIMA: {primeiro_cima}")
                
                print(f"\n>>> Encadeando scripts entre tipos...")
                
                # Encadear: último CIMA -> primeiro ABCD
                if ultimo_cima and primeiro_abcd:
                    self._adicionar_comando_script_ao_final(ultimo_cima, primeiro_abcd)
                    print(f">>> ✅ Último CIMA -> Primeiro ABCD")
                else:
                    if not ultimo_cima:
                        print(f">>> ⚠️ Último CIMA não encontrado")
                    if not primeiro_abcd:
                        print(f">>> ⚠️ Primeiro ABCD não encontrado")
                
                # Encadear: último ABCD -> primeiro GRADES
                if ultimo_abcd and primeiro_grades:
                    self._adicionar_comando_script_ao_final(ultimo_abcd, primeiro_grades)
                    print(f">>> ✅ Último ABCD -> Primeiro GRADES")
                else:
                    if not ultimo_abcd:
                        print(f">>> ⚠️ Último ABCD não encontrado")
                    if not primeiro_grades:
                        print(f">>> ⚠️ Primeiro GRADES não encontrado")
                
                # Atualizar PAZ com o primeiro script de CIMA (1.scr)
                if primeiro_cima:
                    self._atualizar_script_paz(primeiro_cima)
                    print(f">>> ✅ PAZ atualizado com primeiro script: CIMA (1.scr)")
                elif primeiro_abcd:
                    self._atualizar_script_paz(primeiro_abcd)
                    print(f">>> ✅ PAZ atualizado com primeiro script: ABCD (1.scr)")
                elif primeiro_grades:
                    self._atualizar_script_paz(primeiro_grades)
                    print(f">>> ✅ PAZ atualizado com primeiro script: GRADES (1.scr)")
                
                # Perguntar se deseja executar no AutoCAD
                opcao = self._perguntar_execucao_autocad("desenho pavimento completo")
                
                if opcao is None:  # Cancelado
                    return
                
                if opcao == 1:
                    # Opção 1: Desenhar no CAD diretamente
                    # Executar o primeiro script no AutoCAD (os outros serão chamados via SCRIPT)
                    if primeiro_cima:
                        script_content = self._ler_e_filtrar_script(primeiro_cima)
                        if script_content:
                            self._executar_no_autocad(script_content, pavimento, "COMPLETO")
                            print(f">>> ✅ Script executado no AutoCAD com sucesso!")
                    else:
                        messagebox.showwarning("Aviso", "Script CIMA não encontrado. Use o comando PAZ no AutoCAD para executar os scripts encadeados.")
                elif opcao == 2:
                    # Opção 2: Gerar comando PAZ (não mostrar messagebox de sucesso)
                    print(f">>> ✅ Scripts gerados (não executados no AutoCAD)")
                elif opcao == 3:
                    # Opção 3: Gerar DFX (em desenvolvimento)
                    messagebox.showinfo("Em Desenvolvimento", "A opção de gerar DFX está em desenvolvimento.")
                    return
                
            except Exception as e:
                # Erro durante execução - créditos já foram debitados pelas funções individuais
                print(f">>> Erro durante execução: {str(e)}")
                raise e  # Re-raise para ser tratado no bloco externo
                
            finally:
                # Sempre liberar o lock
                self.pavimento_completo_lock = False
                
        except Exception as e:
            print(f">>> ❌ ERRO ao executar desenhar 1 pavimento completo: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar desenhar 1 pavimento completo:\n{str(e)}")
            # Liberar lock em caso de erro
            if hasattr(self, 'pavimento_completo_lock'):
                self.pavimento_completo_lock = False

    def _criar_comando_lisp(self):
        """Cria os arquivos LISP e SCR para o comando PAZ"""
        try:
            print(f"\n>>> CRIAR COMANDO LISP - INICIADO")
            
            # Detectar se está em frozen (PC do cliente)
            is_frozen = False
            if hasattr(self, 'conector_painel'):
                is_frozen = self.conector_painel._is_frozen()
            else:
                # Fallback: detectar manualmente
                is_frozen = getattr(sys, 'frozen', False)
                if not is_frozen:
                    if hasattr(sys, 'executable') and sys.executable and sys.executable.endswith('.exe'):
                        exe_dir = os.path.dirname(sys.executable)
                        if '.dist' in exe_dir or os.path.basename(exe_dir) in ['run.dist', 'dist', 'dist_nuitka', 'dist_debug']:
                            is_frozen = True
            
            print(f">>> [LISP] Ambiente frozen detectado: {is_frozen}")
            
            # Determinar o caminho da pasta output
            if is_frozen:
                # Em frozen, a pasta output DEVE estar no mesmo diretório do executável (não em _MEIPASS)
                # _MEIPASS é temporário e será deletado quando o programa fechar
                exe_dir = os.path.dirname(sys.executable)
                output_dir = os.path.join(exe_dir, "output")
                print(f">>> [LISP] Frozen mode: usando diretório do executável")
                print(f">>> [LISP] Executable dir: {exe_dir}")
            else:
                # Em desenvolvimento, usar a estrutura normal
                project_root = self._get_project_root()
                output_dir = os.path.join(project_root, "output")
                print(f">>> [LISP] Development mode: usando project root")
                print(f">>> [LISP] Project root: {project_root}")
            
            # Garantir que a pasta output existe
            os.makedirs(output_dir, exist_ok=True)
            print(f">>> [LISP] Output dir: {output_dir}")
            
            # Caminho do arquivo SCR
            script_scr_path = os.path.join(output_dir, "script_PAZ.scr")
            # Converter para formato absoluto e usar barras normais (forward slashes)
            # IMPORTANTE: No LISP, usar barras normais (/) é mais seguro que barras invertidas
            # O Windows aceita ambos os formatos, mas barras normais evitam problemas de escape
            script_scr_path_absoluto = os.path.abspath(script_scr_path)
            # Converter barras invertidas para barras normais (forward slashes)
            script_scr_path_lisp = script_scr_path_absoluto.replace('\\', '/')
            
            # Caminho do arquivo LISP
            comando_lisp_path = os.path.join(output_dir, "comando_PAZ.lsp")
            
            # Criar o conteúdo do arquivo LISP
            # IMPORTANTE: Usar barras normais (forward slashes) no caminho para evitar problemas de escape
            lisp_content = f""";; Comando para executar script SCR PAZ
(defun c:PAZ ()
  (command "_SCRIPT" "{script_scr_path_lisp}")
  (princ)
)
"""
            
            # Criar o arquivo LISP
            with open(comando_lisp_path, 'w', encoding='utf-8') as f:
                f.write(lisp_content)
            print(f">>> [LISP] ✅ Arquivo LISP criado: {comando_lisp_path}")
            
            # Criar o arquivo SCR vazio (ou com conteúdo básico)
            with open(script_scr_path, 'w', encoding='utf-16-le') as f:
                # Adicionar BOM UTF-16 LE
                f.write('\ufeff')
                # Arquivo vazio por enquanto (será editado depois)
                f.write("")
            print(f">>> [LISP] ✅ Arquivo SCR criado: {script_scr_path}")
            
            messagebox.showinfo("Sucesso", 
                f"Arquivos criados com sucesso!\n\n"
                f"LISP: {comando_lisp_path}\n"
                f"SCR: {script_scr_path}\n\n"
                f"O comando PAZ está pronto para uso no AutoCAD.")
            
        except Exception as e:
            print(f">>> ❌ ERRO ao criar comando LISP: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao criar comando LISP:\n{str(e)}")

    def _perguntar_execucao_autocad(self, tipo_operacao="desenho"):
        """
        Pergunta ao usuário como deseja executar o desenho.
        
        Args:
            tipo_operacao: Tipo de operação (ex: "desenho", "pavimento")
        
        Returns:
            1 se Opção 1 (Desenhar no CAD diretamente)
            2 se Opção 2 (Gerar comando PAZ)
            3 se Opção 3 (Gerar DFX - em desenvolvimento)
            None se Cancelar
        """
        try:
            import tkinter as tk
            from tkinter import messagebox as mb
            
            # Obter janela principal (root) - múltiplas tentativas
            root = None
            try:
                # Método 1: Tentar obter do conector_painel -> main_app -> root
                if hasattr(self, 'conector_painel'):
                    conector = self.conector_painel
                    if hasattr(conector, 'main_app'):
                        main_app = conector.main_app
                        if hasattr(main_app, 'root'):
                            root = main_app.root
                        elif hasattr(main_app, 'master'):
                            root = main_app.master
                
                # Método 2: Tentar obter diretamente de self
                if root is None:
                    if hasattr(self, 'root'):
                        root = self.root
                    elif hasattr(self, 'master'):
                        root = self.master
                        # Subir até encontrar root
                        while root and hasattr(root, 'master') and root.master:
                            root = root.master
                
                # Método 3: Tentar obter root global
                if root is None:
                    root = tk._default_root
                    
            except Exception as e:
                print(f">>> [OPÇÃO] Erro ao obter root: {e}")
                root = None
            
            # Se não encontrou root, criar um temporário
            if root is None or not hasattr(root, 'winfo_exists') or not root.winfo_exists():
                print(f">>> [OPÇÃO] Criando root temporário")
                root = tk.Tk()
                root.withdraw()  # Esconder janela temporária
                root.attributes('-topmost', True)  # Garantir que apareça
            
            # Verificar se a janela principal está ativa/focada
            janela_principal_ativa = False
            if root and hasattr(root, 'winfo_exists') and root.winfo_exists():
                try:
                    # Verificar se a janela está visível e não está minimizada
                    root.update_idletasks()
                    if root.winfo_viewable():
                        # Tentar trazer a janela principal para frente primeiro
                        try:
                            root.lift()
                            root.focus_force()
                            root.update()
                            # Aguardar um pouco para a janela principal ganhar foco
                            root.after(100, lambda: None)
                            root.update()
                            janela_principal_ativa = True
                            print(f">>> [OPÇÃO] Janela principal trazida para frente")
                        except Exception as e:
                            print(f">>> [OPÇÃO] Erro ao trazer janela principal para frente: {e}")
                            janela_principal_ativa = False
                except Exception as e:
                    print(f">>> [OPÇÃO] Erro ao verificar estado da janela principal: {e}")
                    janela_principal_ativa = False
            
            # Criar janela customizada
            dialog = tk.Toplevel(root)
            dialog.title("Opções de Execução")
            dialog.geometry("550x520")
            dialog.resizable(False, False)
            dialog.transient(root)  # Fazer janela modal
            
            # Só usar grab_set se a janela principal estiver ativa
            # Caso contrário, a janela pode bloquear interação quando o usuário está em outra janela
            if janela_principal_ativa:
                try:
                    dialog.grab_set()  # Bloquear interação com outras janelas apenas se a principal estiver ativa
                    print(f">>> [OPÇÃO] Dialog modal ativado (janela principal ativa)")
                except Exception as e:
                    print(f">>> [OPÇÃO] Erro ao ativar grab_set: {e}")
            else:
                print(f">>> [OPÇÃO] Dialog não-modal (janela principal não está ativa)")
                # Não usar grab_set para evitar bloqueio quando usuário está em outra janela
            
            dialog.attributes('-topmost', True)  # Garantir que apareça no topo
            dialog.lift()  # Trazer para frente
            dialog.focus_force()  # Forçar foco
            
            # Centralizar janela
            dialog.update_idletasks()
            width = dialog.winfo_width()
            height = dialog.winfo_height()
            x = (dialog.winfo_screenwidth() // 2) - (width // 2)
            y = (dialog.winfo_screenheight() // 2) - (height // 2)
            dialog.geometry(f'{width}x{height}+{x}+{y}')
            
            # Variável para armazenar resultado
            resultado = [None]
            
            # Frame principal
            main_frame = tk.Frame(dialog, padx=20, pady=15)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo = tk.Label(
                main_frame, 
                text=f"Como deseja executar o {tipo_operacao}?",
                font=("Arial", 11, "bold"),
                wraplength=500
            )
            titulo.pack(pady=(0, 15))
            
            # Frame para instruções
            instrucoes_frame = tk.Frame(main_frame, relief=tk.SUNKEN, bd=1, bg="#f0f0f0")
            instrucoes_frame.pack(fill=tk.X, pady=(0, 15))
            
            instrucoes_texto = (
                "⚠️ IMPORTANTE - Antes de executar:\n\n"
                "• NÃO mexa no AutoCAD enquanto o desenho está sendo executado\n"
                "• Desative o OSNAP no AutoCAD (F3) antes de iniciar\n"
                "• Para usar o comando PAZ: gere e carregue o comando LISP no AutoCAD\n"
                "• Garanta que todos os itens do pavimento estejam no mesmo modo de desenho (INI ou NOVA)\n"
                "• Aguarde a conclusão completa do desenho antes de interagir com o AutoCAD"
            )
            
            label_instrucoes = tk.Label(
                instrucoes_frame,
                text=instrucoes_texto,
                font=("Arial", 9),
                justify=tk.LEFT,
                bg="#f0f0f0",
                wraplength=480,
                padx=10,
                pady=10
            )
            label_instrucoes.pack(anchor="w")
            
            # Frame para os botões de opções
            opcoes_frame = tk.Frame(main_frame)
            opcoes_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
            
            # Opção 1: Desenhar no CAD Diretamente
            def opcao1():
                resultado[0] = 1
                dialog.destroy()
            
            btn_opcao1 = tk.Button(
                opcoes_frame,
                text="Opção 1: Desenhar no CAD Diretamente\n(Modo lento mas seguro de não interromper-se e qualidade)",
                font=("Arial", 9),
                width=55,
                height=2,
                command=opcao1,
                anchor="w",
                justify="left",
                relief=tk.RAISED,
                bd=1,
                wraplength=480
            )
            btn_opcao1.pack(pady=4, fill=tk.X)
            
            # Opção 2: Gerar comando PAZ
            def opcao2():
                resultado[0] = 2
                dialog.destroy()
            
            btn_opcao2 = tk.Button(
                opcoes_frame,
                text="Opção 2: Gerar comando PAZ\n(rapidez e qualidade, mas chance de interromper-se se algum dado mal preenchido)",
                font=("Arial", 9),
                width=55,
                height=2,
                command=opcao2,
                anchor="w",
                justify="left",
                relief=tk.RAISED,
                bd=1,
                wraplength=480
            )
            btn_opcao2.pack(pady=4, fill=tk.X)
            
            # Opção 3: Gerar DFX (desativada)
            def opcao3():
                resultado[0] = 3
                dialog.destroy()
            
            btn_opcao3 = tk.Button(
                opcoes_frame,
                text="Opção 3: Gerar DFX (EM DESENVOLVIMENTO)",
                font=("Arial", 9),
                width=55,
                height=2,
                command=opcao3,
                anchor="w",
                justify="left",
                relief=tk.RAISED,
                bd=1,
                state=tk.DISABLED,
                wraplength=480
            )
            btn_opcao3.pack(pady=4, fill=tk.X)
            
            # Botão Cancelar
            def cancelar():
                resultado[0] = None
                dialog.destroy()
            
            btn_cancelar = tk.Button(
                main_frame,
                text="Cancelar",
                font=("Arial", 9),
                width=15,
                height=1,
                command=cancelar,
                relief=tk.RAISED,
                bd=1
            )
            btn_cancelar.pack(pady=(5, 0))
            
            # Sempre permitir fechar o dialog (mesmo quando janela principal não está ativa)
            def on_close_dialog_always(event=None):
                """Permite fechar o dialog sempre, mesmo quando a janela principal não está ativa"""
                try:
                    if dialog.winfo_exists():
                        if resultado[0] is None:  # Se ainda não foi escolhida uma opção
                            resultado[0] = None  # Cancelar
                        dialog.destroy()
                except:
                    pass
            
            # Adicionar protocolo de fechamento sempre ativo
            dialog.protocol("WM_DELETE_WINDOW", on_close_dialog_always)
            dialog.bind('<Escape>', on_close_dialog_always)
            
            # Garantir que a janela seja visível
            dialog.update_idletasks()
            dialog.deiconify()  # Garantir que não esteja escondida
            dialog.lift()  # Trazer para frente novamente
            dialog.focus_set()  # Focar na janela
            
            # Processar eventos para garantir que a janela apareça
            root.update_idletasks()
            
            # Aguardar fechamento da janela
            dialog.wait_window()
            
            # Retornar resultado
            if resultado[0] == 1:
                print(f">>> [OPÇÃO] Usuário escolheu: Opção 1 - Desenhar no CAD Diretamente")
                return 1
            elif resultado[0] == 2:
                print(f">>> [OPÇÃO] Usuário escolheu: Opção 2 - Gerar comando PAZ")
                return 2
            elif resultado[0] == 3:
                print(f">>> [OPÇÃO] Usuário escolheu: Opção 3 - Gerar DFX")
                return 3
            else:
                print(f">>> [OPÇÃO] Operação cancelada pelo usuário")
                return None
                
        except Exception as e:
            print(f">>> ❌ ERRO ao perguntar execução AutoCAD: {str(e)}")
            import traceback
            traceback.print_exc()
            # Em caso de erro, assumir que deve executar (comportamento padrão)
            return 1
    
    def _atualizar_script_paz(self, caminho_primeiro_script):
        """
        Atualiza o script_PAZ.scr com o caminho do primeiro script a ser executado.
        Funciona em ambiente frozen e desenvolvimento.
        
        Args:
            caminho_primeiro_script: Caminho absoluto do primeiro script .scr a ser executado
        """
        try:
            print(f"\n>>> ATUALIZANDO SCRIPT_PAZ.scr")
            print(f">>> Caminho do primeiro script: {caminho_primeiro_script}")
            
            # Determinar o diretório de output
            # Em frozen mode, DEVE ser o diretório do executável (não _MEIPASS)
            is_frozen = getattr(sys, 'frozen', False)
            if is_frozen:
                # Em frozen, output DEVE estar no mesmo diretório do executável
                exe_dir = os.path.dirname(sys.executable)
                output_dir = os.path.join(exe_dir, "output")
                print(f">>> [PAZ] Frozen mode: usando diretório do executável")
                print(f">>> [PAZ] Executable dir: {exe_dir}")
            else:
                # Em desenvolvimento, usar a estrutura normal
                project_root = self._get_project_root()
                output_dir = os.path.join(project_root, "output")
                print(f">>> [PAZ] Development mode: usando project root")
                print(f">>> [PAZ] Project root: {project_root}")
            
            # Garantir que a pasta output existe
            os.makedirs(output_dir, exist_ok=True)
            
            # Caminho do arquivo script_PAZ.scr
            script_paz_path = os.path.join(output_dir, "script_PAZ.scr")
            
            # Converter caminho do primeiro script para formato Windows (barras invertidas)
            caminho_absoluto = os.path.abspath(caminho_primeiro_script)
            # Garantir que o caminho usa barras invertidas (Windows)
            caminho_windows = caminho_absoluto.replace('/', '\\')
            
            # Criar conteúdo do script_PAZ.scr (apenas o comando SCRIPT com o caminho)
            # O script_PAZ.scr deve executar o primeiro script
            # IMPORTANTE: Usar raw string ou escapar corretamente para preservar barras invertidas
            conteudo_script = "SCRIPT" + "\n" + caminho_windows + "\n"
            
            # Salvar o arquivo com UTF-16 LE e BOM
            # Usar modo binário para garantir que as barras invertidas sejam preservadas
            with open(script_paz_path, 'wb') as f:
                # Adicionar BOM UTF-16 LE
                f.write(b'\xFF\xFE')
                # Converter conteúdo para UTF-16 LE
                f.write(conteudo_script.encode('utf-16-le'))
            
            print(f">>> ✅ script_PAZ.scr atualizado com sucesso!")
            print(f">>>    Arquivo: {script_paz_path}")
            print(f">>>    Conteúdo: SCRIPT -> {caminho_windows}")
            
        except Exception as e:
            print(f">>> ❌ ERRO ao atualizar script_PAZ.scr: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _adicionar_comando_script_ao_final(self, caminho_script, caminho_proximo_script):
        """
        Adiciona comando SCRIPT no final de um script para conectar ao próximo script.
        Funciona em ambiente frozen e desenvolvimento.
        
        Args:
            caminho_script: Caminho do script atual (onde adicionar o comando)
            caminho_proximo_script: Caminho do próximo script (para o comando SCRIPT)
        """
        try:
            print(f"\n>>> ADICIONANDO COMANDO SCRIPT AO FINAL")
            print(f">>> Script atual: {caminho_script}")
            print(f">>> Próximo script: {caminho_proximo_script}")
            
            # Verificar se o arquivo existe
            if not os.path.exists(caminho_script):
                print(f">>> ❌ ERRO: Arquivo não encontrado: {caminho_script}")
                raise FileNotFoundError(f"Arquivo não encontrado: {caminho_script}")
            
            # Ler o conteúdo atual do script
            try:
                with open(caminho_script, 'r', encoding='utf-16-le') as f:
                    conteudo = f.read()
                print(f">>> ✅ Arquivo lido com encoding: utf-16-le")
            except UnicodeDecodeError:
                # Fallback para outros encodings
                print(f">>> ⚠️ Tentando fallback para utf-8...")
                try:
                    with open(caminho_script, 'r', encoding='utf-8') as f:
                        conteudo = f.read()
                    print(f">>> ✅ Arquivo lido com encoding: utf-8")
                except UnicodeDecodeError:
                    # Último fallback para latin-1
                    print(f">>> ⚠️ Tentando fallback para latin-1...")
                    with open(caminho_script, 'r', encoding='latin-1') as f:
                        conteudo = f.read()
                    print(f">>> ✅ Arquivo lido com encoding: latin-1")
            
            # Remover BOM se existir
            if conteudo.startswith('\ufeff'):
                conteudo = conteudo[1:]
                print(f">>> ✅ BOM removido")
            
            print(f">>> Tamanho do conteúdo: {len(conteudo)} caracteres")
            print(f">>> Número de linhas: {len(conteudo.split(chr(10)))} linhas")
            
            # Verificar se já tem comando SCRIPT no final (evitar duplicação)
            linhas = conteudo.rstrip().split('\n')
            # Verificar se as últimas 2 linhas são "SCRIPT" seguido de um caminho
            if len(linhas) >= 2:
                ultima_linha = linhas[-1].strip()
                penultima_linha = linhas[-2].strip().upper()
                # Verificar se penúltima linha é "SCRIPT" ou ";SCRIPT" e última linha é um caminho
                if (penultima_linha == 'SCRIPT' or penultima_linha == ';SCRIPT' or 
                    penultima_linha.endswith('SCRIPT')) and (ultima_linha and 
                    (os.path.sep in ultima_linha or ':' in ultima_linha)):
                    print(f">>> ⚠️ Script já tem comando SCRIPT no final, removendo antes de adicionar novo")
                    # Remover as últimas 2 linhas (SCRIPT e caminho)
                    linhas = linhas[:-2]
                    # Remover linhas em branco e comentários no final
                    while linhas and (linhas[-1].strip() == '' or linhas[-1].strip().startswith(';')):
                        linhas.pop()
                    conteudo = '\n'.join(linhas)
            
            # Converter caminho do próximo script para formato Windows (barras invertidas)
            caminho_absoluto = os.path.abspath(caminho_proximo_script)
            # Garantir que o caminho usa barras invertidas (Windows)
            caminho_windows = caminho_absoluto.replace('/', '\\')
            
            # Verificar se o último comando é _LAYER ou -LAYER seguido de S e nome do layer
            # Padrão: _LAYER\nS\n(LAYER)\n ou -LAYER\nS\n(LAYER)\n
            linhas_finais = conteudo.rstrip().split('\n')
            ultimo_comando_layer = False
            
            # Log das últimas linhas para debug
            print(f">>> Últimas 5 linhas do arquivo:")
            for i, linha in enumerate(linhas_finais[-5:]):
                print(f">>>   [{len(linhas_finais) - 5 + i}] '{linha}'")
            
            # Verificar se as últimas 3 linhas são _LAYER/-LAYER, S, e nome do layer
            if len(linhas_finais) >= 3:
                # Remover linhas em branco do final para verificar o padrão
                linhas_limpas = [l for l in linhas_finais if l.strip()]
                if len(linhas_limpas) >= 3:
                    ultima_linha = linhas_limpas[-1].strip()
                    penultima_linha = linhas_limpas[-2].strip().upper()
                    antepenultima_linha = linhas_limpas[-3].strip().upper()
                    
                    # Verificar se é o padrão _LAYER/-LAYER, S, (LAYER)
                    # Aceita tanto _LAYER quanto -LAYER
                    if (antepenultima_linha == '_LAYER' or antepenultima_linha == '-LAYER') and penultima_linha == 'S':
                        ultimo_comando_layer = True
                        print(f">>> ✅ Detectado último comando LAYER: {antepenultima_linha} -> {penultima_linha} -> {ultima_linha}")
            
            # Verificar também se o último comando não vazio é _LAYER ou -LAYER (caso não tenha S explícito)
            if not ultimo_comando_layer and len(linhas_finais) >= 1:
                linhas_limpas = [l for l in linhas_finais if l.strip()]
                if len(linhas_limpas) >= 1:
                    ultima_linha_limpa = linhas_limpas[-1].strip().upper()
                    # Verificar se a última linha não vazia é _LAYER ou -LAYER
                    if ultima_linha_limpa == '_LAYER' or ultima_linha_limpa == '-LAYER':
                        ultimo_comando_layer = True
                        print(f">>> ✅ Detectado último comando LAYER (sem S): {ultima_linha_limpa}")
            
            # Adicionar comando SCRIPT no final
            conteudo_novo = conteudo.rstrip()
            
            # Se o último comando é _LAYER/-LAYER, adicionar linha em branco antes do ;
            if ultimo_comando_layer:
                # Garantir que há pelo menos uma linha em branco antes do ;
                if not conteudo_novo.endswith('\n'):
                    conteudo_novo += '\n'
                # Adicionar linha em branco extra antes do ;
                conteudo_novo += '\n;\nSCRIPT\n'
            else:
                # IMPORTANTE: Usar concatenação de strings para preservar barras invertidas
                if not conteudo_novo.endswith('\n'):
                    conteudo_novo += '\n'
                conteudo_novo += ';\nSCRIPT\n'
            
            conteudo_novo += caminho_windows + '\n'
            
            # Log das últimas linhas do conteúdo que será salvo
            linhas_finais_novo = conteudo_novo.rstrip().split('\n')
            print(f">>> Últimas 5 linhas do conteúdo que será salvo:")
            for i, linha in enumerate(linhas_finais_novo[-5:]):
                print(f">>>   [{len(linhas_finais_novo) - 5 + i}] '{linha}'")
            
            # Salvar o arquivo com UTF-16 LE e BOM
            # Usar modo binário para garantir que as barras invertidas sejam preservadas
            with open(caminho_script, 'wb') as f:
                # Adicionar BOM UTF-16 LE
                f.write(b'\xFF\xFE')
                # Converter conteúdo para UTF-16 LE
                f.write(conteudo_novo.encode('utf-16-le'))
            
            print(f">>> ✅ Comando SCRIPT adicionado com sucesso!")
            print(f">>>    Script atualizado: {caminho_script}")
            print(f">>>    Próximo script: {caminho_windows}")
            print(f">>>    Último comando era LAYER: {ultimo_comando_layer}")
            
        except Exception as e:
            print(f">>> ❌ ERRO ao adicionar comando SCRIPT: {str(e)}")
            import traceback
            traceback.print_exc()

    def _limpar_todas_pastas_scripts(self):
        """
        Remove todas as pastas de scripts para limpeza completa do sistema
        """
        try:
            print(f"\n>>> LIMPANDO TODAS AS PASTAS DE SCRIPTS")
            
            # Obter o diretório raiz
            robo_dir = self._get_project_root()
            pastas_scripts_dir = os.path.join(robo_dir, "output", "scripts")
            
            print(f">>> Diretório de pastas: {pastas_scripts_dir}")
            
            if not os.path.exists(pastas_scripts_dir):
                print(f">>> ℹ️ Diretório Pastas_SCRIPTS não existe")
                return
            
            # Listar todas as pastas
            pastas_encontradas = []
            try:
                for item in os.listdir(pastas_scripts_dir):
                    item_path = os.path.join(pastas_scripts_dir, item)
                    if os.path.isdir(item_path):
                        pastas_encontradas.append(item)
            except Exception as e:
                print(f">>> ❌ ERRO ao listar pastas: {str(e)}")
                return
            
            if not pastas_encontradas:
                print(f">>> ℹ️ Nenhuma pasta de scripts encontrada")
                return
            
            print(f">>> Pastas encontradas ({len(pastas_encontradas)}):")
            for pasta in pastas_encontradas:
                print(f">>>   - {pasta}")
            
            # Confirmar com o usuário
            resposta = messagebox.askyesno(
                "Limpeza Completa", 
                f"Encontradas {len(pastas_encontradas)} pastas de scripts.\n\n"
                f"Deseja remover TODAS as pastas?\n\n"
                f"Isso irá limpar completamente o sistema de scripts."
            )
            
            if not resposta:
                print(f">>> Limpeza cancelada pelo usuário")
                return
            
            # Remover cada pasta
            pastas_removidas = 0
            for pasta in pastas_encontradas:
                pasta_completa = os.path.join(pastas_scripts_dir, pasta)
                try:
                    import shutil
                    shutil.rmtree(pasta_completa)
                    print(f">>> ✅ Pasta removida: {pasta}")
                    pastas_removidas += 1
                except Exception as e:
                    print(f">>> ❌ Erro ao remover pasta {pasta}: {str(e)}")
            
            print(f">>> Resumo: {pastas_removidas}/{len(pastas_encontradas)} pastas removidas")
            
            if pastas_removidas > 0:
                messagebox.showinfo(
                    "Limpeza Concluída", 
                    f"Limpeza concluída com sucesso!\n\n"
                    f"Pastas removidas: {pastas_removidas}/{len(pastas_encontradas)}"
                )
            
        except Exception as e:
            print(f">>> ❌ ERRO GERAL ao limpar todas as pastas: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao limpar pastas:\n{str(e)}")

    def _limpar_pasta_pavimento(self, pavimento, tipo_visao, confirmar=True):
        """
        Remove a pasta do pavimento processado para evitar acúmulo de scripts
        
        Args:
            pavimento: Nome do pavimento processado
            tipo_visao: Tipo de visão (CIMA, ABCD, GRADES)
            confirmar: Se True, pergunta ao usuário antes de remover
        """
        try:
            print(f"\n>>> LIMPANDO PASTA DO PAVIMENTO PROCESSADO")
            print(f">>> Pavimento: {pavimento}")
            print(f">>> Tipo de visão: {tipo_visao}")
            
            # Obter o diretório raiz
            robo_dir = self._get_project_root()
            
            # Construir nome da pasta baseado no pavimento e tipo de visão
            nome_pasta = f"{pavimento.replace(' ', '_')}_{tipo_visao}"
            pasta_completa = os.path.join(robo_dir, "output", "scripts", nome_pasta)
            
            print(f">>> Pasta a ser removida: {pasta_completa}")
            
            if os.path.exists(pasta_completa):
                # Verificar se a pasta não está vazia antes de remover
                try:
                    conteudo = os.listdir(pasta_completa)
                    print(f">>> Conteúdo da pasta antes da remoção:")
                    for item in conteudo:
                        print(f">>>   - {item}")
                    
                    # Confirmar com o usuário se solicitado
                    if confirmar:
                        resposta = messagebox.askyesno(
                            "Limpar Pasta", 
                            f"Deseja remover a pasta do pavimento processado?\n\n"
                            f"Pavimento: {pavimento}\n"
                            f"Tipo: {tipo_visao}\n"
                            f"Arquivos: {len(conteudo)} itens\n\n"
                            f"Isso irá liberar espaço e evitar acúmulo de scripts."
                        )
                        
                        if not resposta:
                            print(f">>> Limpeza cancelada pelo usuário")
                            return
                    
                    # Remover a pasta e todo seu conteúdo
                    import shutil
                    shutil.rmtree(pasta_completa)
                    print(f">>> ✅ Pasta removida com sucesso: {nome_pasta}")
                    
                    # Verificar se foi realmente removida
                    if not os.path.exists(pasta_completa):
                        print(f">>> ✅ Confirmação: Pasta não existe mais")
                    else:
                        print(f">>> ⚠️ ATENÇÃO: Pasta ainda existe após tentativa de remoção")
                        
                except Exception as e:
                    print(f">>> ❌ ERRO ao remover pasta: {str(e)}")
                    # Tentar remover arquivos individualmente
                    try:
                        for root, dirs, files in os.walk(pasta_completa, topdown=False):
                            for file in files:
                                file_path = os.path.join(root, file)
                                try:
                                    os.remove(file_path)
                                    print(f">>>   Removido arquivo: {file}")
                                except Exception as file_error:
                                    print(f">>>   Erro ao remover arquivo {file}: {str(file_error)}")
                            
                            for dir_name in dirs:
                                dir_path = os.path.join(root, dir_name)
                                try:
                                    os.rmdir(dir_path)
                                    print(f">>>   Removido diretório: {dir_name}")
                                except Exception as dir_error:
                                    print(f">>>   Erro ao remover diretório {dir_name}: {str(dir_error)}")
                        
                        # Tentar remover a pasta principal novamente
                        os.rmdir(pasta_completa)
                        print(f">>> ✅ Pasta removida com limpeza manual")
                    except Exception as manual_error:
                        print(f">>> ❌ ERRO na limpeza manual: {str(manual_error)}")
            else:
                print(f">>> ℹ️ Pasta não existe: {nome_pasta}")
                
        except Exception as e:
            print(f">>> ❌ ERRO GERAL ao limpar pasta: {str(e)}")
            import traceback
            traceback.print_exc()

    def _executar_no_autocad(self, script_content, pavimento=None, tipo_visao=None):
        """Executa o script filtrado no AutoCAD e limpa a pasta após execução"""
        try:
            # Copiar para área de transferência
            try:
                import pyperclip
                pyperclip.copy(script_content)
            except ImportError:
                # Se pyperclip não estiver disponível, usar método alternativo
                print("[AVISO] pyperclip não disponível - tentando método alternativo")
                try:
                    import subprocess
                    # Usar PowerShell para copiar para área de transferência
                    subprocess.run(
                        ['powershell', '-command', f'Set-Clipboard -Value @"""{script_content}"""'],
                        check=False,
                        capture_output=True
                    )
                except Exception as e:
                    print(f"[AVISO] Não foi possível copiar para área de transferência: {e}")
                    # Continuar mesmo sem copiar para área de transferência

            # Importar função robusta de inicialização do AutoCAD
            try:
                # Tentar importar a função do aa_corrigido copy 10.py
                import sys
                import os
                import importlib.util
                
                current_dir = os.path.dirname(os.path.abspath(__file__))
                module_path = os.path.join(current_dir, "aa_corrigido copy 10.py")
                
                if os.path.exists(module_path):
                    spec = importlib.util.spec_from_file_location("aa_corrigido_copy_10", module_path)
                    aa_module = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(aa_module)
                    
                    # Inicializar AutoCAD de forma robusta com tratamento de erro
                    try:
                        print("[DEBUG] Tentando inicializar AutoCAD...")
                        acad, doc = aa_module.inicializar_autocad()
                        print(f"[DEBUG] AutoCAD inicializado: acad={acad is not None}, doc={doc is not None}")
                        if not acad or not doc:
                            print("[DEBUG] AutoCAD não inicializado corretamente")
                            messagebox.showerror("Erro", "Não foi possível conectar ao AutoCAD")
                            return
                    except Exception as e:
                        print(f"[DEBUG] Erro ao inicializar AutoCAD: {e}")
                        import traceback
                        traceback.print_exc()
                        messagebox.showerror("Erro", f"Não foi possível conectar ao AutoCAD: {str(e)}")
                        return
                else:
                    raise ImportError("Arquivo não encontrado")
                
            except (ImportError, Exception):
                # Fallback para método tradicional
                pythoncom.CoInitialize()
                try:
                    if self.ac is None:
                        self.ac = win32com.client.Dispatch("AutoCAD.Application")
                    acad = self.ac
                    doc = acad.ActiveDocument
                    if not doc:
                        messagebox.showerror("Erro", "Nenhum documento ativo encontrado no AutoCAD")
                        return
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao conectar com AutoCAD: {str(e)}")
                    return
                finally:
                    pythoncom.CoUninitialize()

            # Trazer AutoCAD para frente
            try:
                # Buscar janela do AutoCAD por diferentes nomes possíveis
                acad_windows = [
                    win32gui.FindWindow(None, "AutoCAD"),
                    win32gui.FindWindow("Afx:00400000:8:00010003:00000006:11CE1031", None),  # Classe específica do AutoCAD
                ]
                
                # Buscar qualquer janela que contenha "AutoCAD" no título
                def enum_windows_callback(hwnd, windows):
                    if win32gui.IsWindowVisible(hwnd):
                        window_text = win32gui.GetWindowText(hwnd)
                        if "AutoCAD" in window_text:
                            windows.append(hwnd)
                
                autocad_windows = []
                win32gui.EnumWindows(enum_windows_callback, autocad_windows)
                
                # Tentar ativar qualquer janela do AutoCAD encontrada
                for hwnd in autocad_windows:
                    if hwnd:
                        try:
                            win32gui.ShowWindow(hwnd, 3)  # SW_MAXIMIZE
                            win32gui.SetForegroundWindow(hwnd)
                            win32gui.BringWindowToTop(hwnd)
                            time.sleep(0.2)
                            break
                        except Exception as e:
                            print(f"[DEBUG] Erro ao ativar janela {hwnd}: {str(e)}")
                            continue

            except Exception as e:
                print(f"[DEBUG] Erro ao trazer janela do AutoCAD para frente: {str(e)}")

            # Executar o script
            try:
                if doc:
                    # Garantir que o AutoCAD está visível
                    acad.Visible = True
                    
                    # Enviar comando
                    doc.SendCommand(script_content + "\n")
                    time.sleep(0.5)

                    # Limpar área de transferência
                    # pyperclip.copy('')  # DESATIVADO TEMPORARIAMENTE PARA DEPURAÇÃO
                    # print("[DEBUG] Área de transferência limpa")

                    # Messagebox de sucesso removida para tornar o processo mais dinâmico
                    # A instrução já foi mostrada antes da execução
                    print(f">>> ✅ Script executado no AutoCAD com sucesso!")
                    
                    # DEPURAÇÃO: manter scripts na pasta do pavimento — limpeza desativada
                    # if pavimento and tipo_visao:
                    #     print(f">>> Iniciando limpeza da pasta após execução bem-sucedida...")
                    #     time.sleep(1.0)
                    #     self._limpar_pasta_pavimento(pavimento, tipo_visao, confirmar=False)
                    # else:
                    #     print(f">>> Informações de pavimento não fornecidas, pulando limpeza")
                        
                else:
                    messagebox.showerror("Erro", "Nenhum documento ativo encontrado no AutoCAD")

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao executar comando no AutoCAD: {str(e)}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao executar no AutoCAD: {str(e)}")


class ControleDesenhoManager:
    """
    Classe para gerenciar funcionalidades avançadas do painel de controle de desenho
    """
    
    def __init__(self, parent_app):
        self.parent_app = parent_app
        self.configuracoes = {}
        self.carregar_configuracoes()
    
    def carregar_configuracoes(self):
        """Carrega configurações salvas do painel de controle"""
        try:
            import json
            # Usar path resolver para encontrar o arquivo de configuração
            project_root = self._get_project_root()
            config_file = os.path.join(project_root, "config", "controle_desenho_config.json")
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    self.configuracoes = json.load(f)
            else:
                # Configurações padrão
                self.configuracoes = {
                    "visao_cima": {
                        "escala": 1.0,
                        "orientacao": "norte",
                        "mostrar_cotas": True
                    },
                    "visao_abcd": {
                        "escala": 1.0,
                        "mostrar_detalhes": True,
                        "incluir_aberturas": True
                    },
                    "grades": {
                        "tipo_representacao": "simplificada",
                        "mostrar_parafusos": True,
                        "incluir_medidas": True
                    }
                }
                self.salvar_configuracoes()
        except Exception as e:
            print(f"Erro ao carregar configurações: {e}")
            self.configuracoes = {}
    
    def salvar_configuracoes(self):
        """Salva configurações do painel de controle"""
        try:
            import json
            # Usar path resolver para encontrar o arquivo de configuração
            project_root = self._get_project_root()
            config_file = os.path.join(project_root, "config", "controle_desenho_config.json")
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(self.configuracoes, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar configurações: {e}")
    
    def obter_configuracao(self, categoria, chave, padrao=None):
        """Obtém uma configuração específica"""
        return self.configuracoes.get(categoria, {}).get(chave, padrao)
    
    def definir_configuracao(self, categoria, chave, valor):
        """Define uma configuração específica"""
        if categoria not in self.configuracoes:
            self.configuracoes[categoria] = {}
        self.configuracoes[categoria][chave] = valor
        self.salvar_configuracoes()     
# ========================================================
    # 💳 SISTEMA DE CONFIRMAÇÃO DE CRÉDITOS
    # ========================================================
    
    def _confirmar_consumo_creditos(self, area_necessaria, saldo_atual, detalhes_operacao, tipo_operacao="item"):
        """
        Exibe diálogo de confirmação para consumo de créditos com interface detalhada.
        
        Args:
            area_necessaria: Área em m² que será consumida
            saldo_atual: Saldo atual de créditos em m²
            detalhes_operacao: Dict com detalhes da operação (nome, pavimento, pilares, etc.)
            tipo_operacao: "item" ou "pavimento"
            
        Returns:
            bool: True se o usuário confirmou, False se cancelou
        """
        import tkinter as tk
        from tkinter import ttk
        
        try:
            print(f">>> Exibindo diálogo de confirmação de créditos...")
            print(f">>> Área necessária: {area_necessaria:.4f} m²")
            print(_get_obf_str("saldo"))
            print(f">>> Tipo de operação: {tipo_operacao}")
            
            # Verificar se há saldo suficiente
            saldo_suficiente = saldo_atual >= area_necessaria
            saldo_restante = saldo_atual - area_necessaria
            
            # Criar janela de confirmação
            janela = tk.Toplevel(self)
            janela.title("💳 Confirmação de Créditos")
            janela.geometry("500x600")
            janela.resizable(False, False)
            janela.transient(self)
            janela.grab_set()
            
            # Centralizar na tela
            janela.update_idletasks()
            width = janela.winfo_width()
            height = janela.winfo_height()
            x = (janela.winfo_screenwidth() // 2) - (width // 2)
            y = (janela.winfo_screenheight() // 2) - (height // 2)
            janela.geometry(f'{width}x{height}+{x}+{y}')
            
            # Variável para armazenar resultado
            resultado = {"confirmado": False}
            
            # Frame principal
            main_frame = ttk.Frame(janela, padding=20)
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Título
            titulo_frame = ttk.Frame(main_frame)
            titulo_frame.pack(fill=tk.X, pady=(0, 15))
            
            ttk.Label(titulo_frame, text="💳 Confirmação de Uso de Créditos", 
                     font=("Arial", 14, "bold")).pack()
            
            if tipo_operacao == "item":
                subtitulo = f"Desenho de Item Completo"
            else:
                subtitulo = f"Desenho de Pavimento Completo"
            
            ttk.Label(titulo_frame, text=subtitulo, font=("Arial", 10)).pack()
            
            # Frame de informações da operação
            info_frame = ttk.LabelFrame(main_frame, text="Detalhes da Operação", padding=10)
            info_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Informações básicas
            if tipo_operacao == "item":
                ttk.Label(info_frame, text=f"Pilar: {detalhes_operacao.get('nome', 'N/A')}", 
                         font=("Arial", 10, "bold")).pack(anchor=tk.W)
                ttk.Label(info_frame, text=f"Pavimento: {detalhes_operacao.get('pavimento', 'N/A')}").pack(anchor=tk.W)
                ttk.Label(info_frame, text=f"Dimensões: {detalhes_operacao.get('comprimento', 'N/A')}cm × {detalhes_operacao.get('largura', 'N/A')}cm").pack(anchor=tk.W)
            else:
                ttk.Label(info_frame, text=f"Pavimento: {detalhes_operacao.get('pavimento', 'N/A')}", 
                         font=("Arial", 10, "bold")).pack(anchor=tk.W)
                ttk.Label(info_frame, text=f"Total de pilares: {detalhes_operacao.get('total_pilares', 0)}").pack(anchor=tk.W)
                
                # Lista de pilares (se for pavimento)
                if 'pilares' in detalhes_operacao and detalhes_operacao['pilares']:
                    ttk.Label(info_frame, text="Pilares inclusos:", font=("Arial", 9, "bold")).pack(anchor=tk.W, pady=(5, 0))
                    
                    # Frame com scroll para lista de pilares
                    pilares_frame = ttk.Frame(info_frame)
                    pilares_frame.pack(fill=tk.X, pady=(2, 0))
                    
                    # Scrollbar
                    scrollbar = ttk.Scrollbar(pilares_frame)
                    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                    
                    # Listbox para pilares
                    pilares_listbox = tk.Listbox(pilares_frame, height=6, font=("Arial", 8),
                                               yscrollcommand=scrollbar.set)
                    pilares_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    scrollbar.config(command=pilares_listbox.yview)
                    
                    # Adicionar pilares à lista
                    for pilar in detalhes_operacao['pilares']:
                        pilares_listbox.insert(tk.END, f"• {pilar['nome']}: {pilar['area']:.4f} m²")
            
            # Frame de créditos
            creditos_frame = ttk.LabelFrame(main_frame, text="Informações de Créditos", padding=10)
            creditos_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Área necessária
            area_label = ttk.Label(creditos_frame, text=f"Área necessária: {area_necessaria:.4f} m²", 
                                  font=("Arial", 11, "bold"))
            area_label.pack(anchor=tk.W)
            
            # Saldo atual
            saldo_label = ttk.Label(creditos_frame, text=_get_obf_str("saldo"), 
                                   font=("Arial", 10))
            saldo_label.pack(anchor=tk.W)
            
            # Saldo após operação
            if saldo_suficiente:
                saldo_restante_label = ttk.Label(creditos_frame, 
                                               text=_get_obf_str("saldo"), 
                                               font=("Arial", 10), foreground="green")
            else:
                saldo_restante_label = ttk.Label(creditos_frame, 
                                               text=_get_obf_str("saldo"), 
                                               font=("Arial", 10, "bold"), foreground="red")
            saldo_restante_label.pack(anchor=tk.W)
            
            # Frame de status
            status_frame = ttk.Frame(main_frame)
            status_frame.pack(fill=tk.X, pady=(0, 15))
            
            if saldo_suficiente:
                status_icon = "✅"
                status_text = "Saldo suficiente para executar a operação"
                status_color = "green"
            else:
                status_icon = "❌"
                status_text = "Saldo insuficiente! Entre em contato para recarregar seus créditos."
                status_color = "red"
            
            status_label = ttk.Label(status_frame, text=f"{status_icon} {status_text}", 
                                   font=("Arial", 10, "bold"), foreground=status_color)
            status_label.pack()
            
            # Frame de botões
            btn_frame = ttk.Frame(main_frame)
            btn_frame.pack(fill=tk.X, pady=(10, 0))
            
            def confirmar():
                resultado["confirmado"] = True
                janela.destroy()
            
            def cancelar():
                resultado["confirmado"] = False
                janela.destroy()
            
            # Botões
            if saldo_suficiente:
                btn_confirmar = ttk.Button(btn_frame, text="✅ Confirmar e Prosseguir", 
                                         command=confirmar, style="Accent.TButton")
                btn_confirmar.pack(side=tk.LEFT, padx=(0, 10))
            
            btn_cancelar = ttk.Button(btn_frame, text="❌ Cancelar", command=cancelar)
            btn_cancelar.pack(side=tk.RIGHT)
            
            # Se não há saldo suficiente, desabilitar confirmação
            if not saldo_suficiente:
                ttk.Label(btn_frame, text="⚠️ Operação não pode ser executada devido ao saldo insuficiente", 
                         font=("Arial", 9), foreground="red").pack(side=tk.LEFT)
            
            # Focar no botão apropriado
            if saldo_suficiente:
                btn_confirmar.focus_set()
            else:
                btn_cancelar.focus_set()
            
            # Vincular teclas
            def on_enter(event):
                if saldo_suficiente:
                    confirmar()
            
            def on_escape(event):
                cancelar()
            
            janela.bind('<Return>', on_enter)
            janela.bind('<Escape>', on_escape)
            
            # Executar loop da janela
            janela.wait_window()
            
            print(f">>> Resultado da confirmação: {'Confirmado' if resultado['confirmado'] else 'Cancelado'}")
            return resultado["confirmado"]
            
        except Exception as e:
            print(f">>> ERRO ao exibir diálogo de confirmação: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Fallback para messagebox simples
            if saldo_suficiente:
                return messagebox.askyesno(
                    "Confirmar Operação",
                    f"Confirmar uso de {area_necessaria:.4f} m² de créditos?\n\n"
                    _get_obf_str("saldo")
                    _get_obf_str("saldo")
                )
            else:
                messagebox.showerror(
                    "Saldo Insuficiente",
                    f"Saldo insuficiente para executar esta operação.\n\n"
                    f"Área necessária: {area_necessaria:.4f} m²\n"
                    _get_obf_str("saldo")
                    _get_obf_str("saldo")
                )
                return False

# Funções auxiliares para testes
def criar_excel_temporario(dados_teste, nome_arquivo):
    """
    Cria um arquivo Excel temporário com os dados de teste
    """
    import tempfile
    from openpyxl import Workbook
    import os
    
    try:
        print(f"[DEBUG] Criando Excel temporário com dados: {dados_teste}")
        
        # Criar arquivo temporário
        if nome_arquivo:
            temp_path = f"temp_{nome_arquivo}.xlsx"
        else:
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='pilares_temp_')
            os.close(temp_fd)
        
        # Criar Excel
        wb = Workbook()
        ws = wb.active
        
        # Mapeamento básico para os dados
        excel_mapping = {
            "nome": 4,
            "comprimento": 6,
            "largura": 7,
            "pavimento": 3,
            
            # Parafusos
            "par_1_2": 173,
            "par_2_3": 174,
            "par_3_4": 175,
            "par_4_5": 176,
            "par_5_6": 177,
            "par_6_7": 178,
            "par_7_8": 179,
            "par_8_9": 180,
            
            # Grades Grupo 1 (Grade A)
            "grade_1": 181,
            "distancia_1": 182,
            "grade_2": 183,
            "distancia_2": 184,
            "grade_3": 185,
            
            # Grades Grupo 2 (Grade B)
            "grade_1_grupo2": 219,
            "distancia_1_grupo2": 220,
            "grade_2_grupo2": 221,
            "distancia_2_grupo2": 222,
            "grade_3_grupo2": 223,
        }
        
        # Preencher dados básicos
        ws.cell(row=1, column=5, value=1)  # Número do pilar na coluna E
        ws.cell(row=excel_mapping["nome"] + 1, column=5, value=dados_teste.get('nome', ''))
        ws.cell(row=excel_mapping["comprimento"] + 1, column=5, value=float(dados_teste.get('comprimento', 0)))
        ws.cell(row=excel_mapping["largura"] + 1, column=5, value=float(dados_teste.get('largura', 0)))
        ws.cell(row=excel_mapping["pavimento"] + 1, column=5, value=dados_teste.get('pavimento', ''))
        
        # Preencher parafusos
        if 'parafusos' in dados_teste:
            for par_key, valor in dados_teste['parafusos'].items():
                if par_key in excel_mapping:
                    ws.cell(row=excel_mapping[par_key] + 1, column=5, value=float(valor) if valor else 0)
        
        # Preencher grades grupo 1 (Grade A)
        if 'grades' in dados_teste:
            for grade_key, valor in dados_teste['grades'].items():
                if grade_key in excel_mapping and valor:
                    ws.cell(row=excel_mapping[grade_key] + 1, column=5, value=float(valor))
        
        # Preencher grades grupo 2 (Grade B) - ESTA É A PARTE IMPORTANTE!
        if 'grades_grupo2' in dados_teste:
            print(f"[DEBUG] Preenchendo grades grupo 2: {dados_teste['grades_grupo2']}")
            for grade_key, valor in dados_teste['grades_grupo2'].items():
                if grade_key in excel_mapping and valor:
                    linha = excel_mapping[grade_key] + 1
                    print(f"[DEBUG] Preenchendo {grade_key} na linha {linha} com valor {valor}")
                    ws.cell(row=linha, column=5, value=float(valor))
        
        # Salvar arquivo
        wb.save(temp_path)
        print(f"[DEBUG] Excel temporário criado: {temp_path}")
        
        return temp_path
        
    except Exception as e:
        print(f"[ERROR] Erro ao criar Excel temporário: {e}")
        import traceback
        traceback.print_exc()
        return None

    def _contar_itens_pavimento(self, pavimento):
        """Conta a quantidade de itens em um pavimento específico"""
        try:
            if not hasattr(self, 'conector_painel'):
                return 0
            
            # Obter todos os itens do pavimento
            itens = self.conector_painel.get_itens_por_pavimento(pavimento)
            return len(itens) if itens else 0
            
        except Exception as e:
            print(f"Erro ao contar itens do pavimento: {str(e)}")
            return 0

def executar_dcad_cima(excel_path, coluna, pavimento):
    """
    Executa o robô DCAD CIMA com o arquivo Excel
    """
    try:
        print(f"[DEBUG] Executando DCAD CIMA com Excel: {excel_path}")
        print(f"[DEBUG] Coluna: {coluna}, Pavimento: {pavimento}")
        
        # Importar o robô CIMA
        import sys
        import os
        
        # Adicionar caminhos necessários
        current_dir = os.path.dirname(os.path.abspath(__file__))
        robots_dir = os.path.join(current_dir, '..', 'robots')
        sys.path.append(robots_dir)
        
        from robo_cima_wrapper import executar_robo_cima
        
        # Executar o robô
        print(f"[DEBUG] Chamando executar_robo_cima...")
        resultado = executar_robo_cima()
        
        if resultado:
            print(f"[DEBUG] ✅ DCAD CIMA executado com sucesso!")
            return True
        else:
            print(f"[DEBUG] ❌ Falha ao executar DCAD CIMA")
            return False
        
    except Exception as e:
        print(f"[ERROR] Erro ao executar DCAD CIMA: {e}")
        import traceback
        traceback.print_exc()
        return False