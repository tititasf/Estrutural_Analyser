
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

"""
Módulo Conector Interface-Painel de Controle

Este módulo conecta a interface principal com as funcionalidades do Painel de Controle,
permitindo que os 12 botões utilizem os dados da lista de pilares em vez de depender
diretamente do Excel.

Os 12 botões implementados são:
- ABCD: Executar Individual, Executar Todos + Combinar, Ajustes Scripts, Configurar Ordenador
- CIMA: Executar Individual, Executar Todos + Combinar, Ajustes Scripts, Configurar Ordenador
- GRADES: Executar Individual, Executar Todos + Combinar, Ajustes Scripts, Configurar Ordenador
"""

import os
import sys
import subprocess
import tempfile
import shutil
import time
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook, load_workbook
from glob import glob
from openpyxl.utils import get_column_letter
import json
import tkinter as tk
from tkinter import ttk

# IMPORTAR HELPER FROZEN GLOBAL - garante que paths estão configurados
try:
    from _frozen_helper import ensure_paths
    ensure_paths()
except ImportError:
    try:
        from src._frozen_helper import ensure_paths
        ensure_paths()
    except ImportError:
        # Fallback manual
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
            if script_dir not in sys.path:
                sys.path.insert(0, script_dir)
            src_dir = os.path.join(script_dir, 'src')
            if os.path.exists(src_dir) and src_dir not in sys.path:
                sys.path.insert(0, src_dir)
        else:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            parent_dir = os.path.dirname(current_dir)
            if parent_dir not in sys.path:
                sys.path.insert(0, parent_dir)

# Importar robust_path_resolver com múltiplos fallbacks
try:
    from ..utils.robust_path_resolver import robust_path_resolver, get_directories, get_all_robots
except ImportError:
    try:
        from utils.robust_path_resolver import robust_path_resolver, get_directories, get_all_robots
    except ImportError:
        try:
            from src.utils.robust_path_resolver import robust_path_resolver, get_directories, get_all_robots
        except ImportError:
            try:
                from robust_path_resolver import robust_path_resolver, get_directories, get_all_robots
            except ImportError:
                # Último fallback: importlib direto
                try:
                    import importlib.util
                    current_file = os.path.abspath(__file__)
                    utils_dir = os.path.join(os.path.dirname(os.path.dirname(current_file)), 'utils')
                    robust_path = os.path.join(utils_dir, 'robust_path_resolver.py')
                    if os.path.exists(robust_path):
                        spec = importlib.util.spec_from_file_location("robust_path_resolver", robust_path)
                        if spec and spec.loader:
                            robust_module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(robust_module)
                            robust_path_resolver = robust_module.robust_path_resolver
                            get_directories = robust_module.get_directories
                            get_all_robots = robust_module.get_all_robots
                except Exception:
                    raise ImportError("Não foi possível importar robust_path_resolver")


class ConectorPainelControle:
    """
    Classe que conecta a interface principal com as funcionalidades do Painel de Controle
    """
    
    def __init__(self, main_app):
        """
        Inicializa o conector com referência à aplicação principal
        
        Args:
            main_app: Instância da aplicação principal (PilarAnalyzer)
        """
        self.main_app = main_app
        
        # Usar path resolver robusto para encontrar diretórios
        dirs = get_directories()
        robots = get_all_robots()
        
        self.painel_dir = str(dirs.get("ordenamento", ""))
        self.scripts_dir = str(dirs.get("automacao", ""))
        self.robos_dir = str(dirs.get("robots", ""))
        self.automacao_dir = str(dirs.get("automacao", ""))
        self.utils_dir = str(dirs.get("utils", ""))
        
        # Cache dos caminhos dos robôs
        self.robot_paths = robots
        

    def _get_project_root(self):
        """Retorna o diretório raiz do projeto"""
        try:
            from ..utils import robust_path_resolver
            return robust_path_resolver.get_project_root()
        except Exception:
            # Fallback manual
            return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        # Validar instalação
        valid, missing = robust_path_resolver.validate_installation()
        if not valid:
            pass
    
    def _is_frozen(self):
        """Detecta se estamos em ambiente frozen (compilado)"""
        is_frozen = getattr(sys, 'frozen', False)
        if not is_frozen:
            # Detecção alternativa para Nuitka: verificar se sys.executable é .exe
            if hasattr(sys, 'executable') and sys.executable and sys.executable.endswith('.exe'):
                exe_dir = os.path.dirname(sys.executable)
                if '.dist' in exe_dir or os.path.basename(exe_dir) in ['run.dist', 'dist', 'dist_nuitka', 'dist_debug']:
                    is_frozen = True
        return is_frozen
    
    def _executar_excel_wrapper(self, wrapper_name, excel_path, colunas, pavimento=None, interface_principal=None, gerar_pelo_pavimento=False):
        """
        Executa um wrapper Excel (CIMA, ABCD ou GRADES) de forma compatível com frozen e dev
        
        Args:
            wrapper_name: Nome do wrapper ('cima_excel', 'abcd_excel', 'grades_excel')
            excel_path: Caminho do arquivo Excel temporário
            colunas: String de colunas (ex: 'E' ou 'E,F,G')
            pavimento: Nome do pavimento (opcional)
            interface_principal: Instância da interface principal (opcional, apenas para ABCD)
            
        Returns:
            subprocess.CompletedProcess ou None em caso de erro
        """
        print(f"\n>>> [DEBUG WRAPPER] _executar_excel_wrapper INICIADO")
        print(f">>> [DEBUG WRAPPER]   - wrapper_name: {wrapper_name}")
        print(f">>> [DEBUG WRAPPER]   - excel_path: {excel_path}")
        print(f">>> [DEBUG WRAPPER]   - colunas: {colunas}")
        print(f">>> [DEBUG WRAPPER]   - pavimento: {pavimento}")
        print(f">>> [DEBUG WRAPPER]   - interface_principal: {interface_principal}")
        print(f">>> [DEBUG WRAPPER]   - gerar_pelo_pavimento: {gerar_pelo_pavimento}")
        
        wrapper_path = self.robot_paths.get(wrapper_name)
        print(f">>> [DEBUG WRAPPER]   - wrapper_path: {wrapper_path}")
        
        if not wrapper_path:
            messagebox.showerror("Erro", f"Wrapper {wrapper_name} não encontrado")
            return None
        
        is_frozen = self._is_frozen()
        
        # Se em ambiente frozen e o path é um nome de módulo (não um caminho físico)
        if is_frozen and not os.path.exists(str(wrapper_path)):
            try:
                # Importar o módulo
                import importlib
                
                # Tentar importar de múltiplos caminhos
                module = None
                import_paths = [
                    wrapper_path,  # Ex: 'src.interfaces.CIMA_FUNCIONAL_EXCEL'
                    wrapper_path.replace('src.', ''),  # Ex: 'interfaces.CIMA_FUNCIONAL_EXCEL'
                    wrapper_path.split('.')[-1]  # Ex: 'CIMA_FUNCIONAL_EXCEL'
                ]
                
                for import_path in import_paths:
                    try:
                        module = importlib.import_module(import_path)
                        break
                    except ImportError:
                        continue
                
                if not module:
                    raise ImportError(f"Não foi possível importar o módulo {wrapper_name}")
                
                # Chamar a função preencher_campos_e_gerar_scripts diretamente
                if hasattr(module, 'preencher_campos_e_gerar_scripts'):
                    func = module.preencher_campos_e_gerar_scripts
                    
                    # Se colunas contém vírgula, é múltiplas colunas
                    if ',' in str(colunas):
                        colunas_str = str(colunas)
                        # ABCD precisa de interface_principal, outros não
                        if wrapper_name == 'abcd_excel' and interface_principal:
                            func(excel_path, coluna_especifica=colunas_str, interface_principal=interface_principal, gerar_pelo_pavimento=gerar_pelo_pavimento)
                        else:
                            func(excel_path, coluna_especifica=colunas_str)
                    else:
                        coluna_str = str(colunas)
                        # ABCD precisa de interface_principal, outros não
                        if wrapper_name == 'abcd_excel' and interface_principal:
                            func(excel_path, coluna_especifica=coluna_str, interface_principal=interface_principal, gerar_pelo_pavimento=gerar_pelo_pavimento)
                        else:
                            func(excel_path, coluna_especifica=coluna_str)
                    
                    # Retornar um objeto compatível com subprocess.CompletedProcess
                    class FakeResult:
                        def __init__(self):
                            self.returncode = 0
                            self.stdout = ""
                            self.stderr = ""
                    
                    return FakeResult()
                else:
                    raise AttributeError(f"Módulo {import_path} não possui função preencher_campos_e_gerar_scripts")
                    
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao executar {wrapper_name}:\n{e}")
                return None
        
        # Em ambiente de desenvolvimento
        else:
            # Sempre usar importação direta em dev (mais confiável)
            try:
                import importlib
                import importlib.util
                
                module = None
                
                # Se wrapper_path é um caminho físico, carregar do arquivo
                if os.path.exists(str(wrapper_path)) and str(wrapper_path).endswith('.py'):
                    spec = importlib.util.spec_from_file_location(wrapper_name, wrapper_path)
                    if spec and spec.loader:
                        module = importlib.util.module_from_spec(spec)
                        spec.loader.exec_module(module)
                else:
                    # Tentar importar de múltiplos caminhos
                    import_paths = [
                        f'interfaces.{wrapper_path.split(".")[-1]}' if '.' in str(wrapper_path) else f'interfaces.{wrapper_path}',
                        f'src.interfaces.{wrapper_path.split(".")[-1]}' if '.' in str(wrapper_path) else f'src.interfaces.{wrapper_path}',
                        wrapper_path.split('.')[-1] if '.' in str(wrapper_path) else str(wrapper_path)
                    ]
                    
                    # Mapear nomes de wrappers para nomes de módulos
                    wrapper_to_module = {
                        'abcd_excel': 'Abcd_Excel',
                        'cima_excel': 'CIMA_FUNCIONAL_EXCEL',
                        'grades_excel': 'GRADE_EXCEL'
                    }
                    
                    if wrapper_name in wrapper_to_module:
                        module_name = wrapper_to_module[wrapper_name]
                        import_paths = [
                            f'interfaces.{module_name}',
                            f'src.interfaces.{module_name}',
                            module_name
                        ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                
                if not module:
                    raise ImportError(f"Não foi possível importar o módulo {wrapper_name}")
                
                # Chamar a função diretamente
                if hasattr(module, 'preencher_campos_e_gerar_scripts'):
                    func = module.preencher_campos_e_gerar_scripts
                    
                    if ',' in str(colunas):
                        colunas_str = str(colunas)
                        if wrapper_name == 'abcd_excel' and interface_principal:
                            func(excel_path, coluna_especifica=colunas_str, interface_principal=interface_principal, gerar_pelo_pavimento=gerar_pelo_pavimento)
                        else:
                            func(excel_path, coluna_especifica=colunas_str)
                    else:
                        coluna_str = str(colunas)
                        if wrapper_name == 'abcd_excel' and interface_principal:
                            func(excel_path, coluna_especifica=coluna_str, interface_principal=interface_principal, gerar_pelo_pavimento=gerar_pelo_pavimento)
                        else:
                            func(excel_path, coluna_especifica=coluna_str)
                    
                    class FakeResult:
                        def __init__(self):
                            self.returncode = 0
                            self.stdout = ""
                            self.stderr = ""
                    
                    return FakeResult()
                else:
                    raise AttributeError(f"Módulo não possui função preencher_campos_e_gerar_scripts")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao executar {wrapper_name}:\n{e}")
                return None
    
    def get_item_selecionado(self):
        """
        Obtém o item atualmente selecionado na lista
        ATUALIZADO: Busca na estrutura obras[obra][pavimento][numero]
        
        Returns:
            tuple: (numero, dados_pilar) ou (None, None) se nenhum item selecionado
        """
        # Primeiro verificar se o filtro do combobox está em "Todos"
        filtro_pavimento = self.main_app.pavimento_filtro_var.get()
        if filtro_pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro antes de usar os botões D.CAD.\n\nO sistema não funciona quando 'Todos' está selecionado no filtro.")
            return None, None
        
        selection = self.main_app.lista_pilares.selection()
        if not selection:
            messagebox.showwarning("Aviso", "Selecione um pilar na lista primeiro.")
            return None, None
            
        item = self.main_app.lista_pilares.item(selection[0])
        valores = item['values']
        numero = str(valores[0]).strip()
        pavimento_lista = str(valores[2]).strip()  # Pavimento da lista
        
        # Obter obra atual do filtro
        obra_atual = self.main_app.obra_filtro_var.get()
        if not obra_atual:
            obra_atual = "Obra Padrão"
        
        # ACESSO DIRETO: Buscar na estrutura obras[obra][pavimento][numero]
        dados = None
        
        try:
            # Verificar se a estrutura obras existe
            if not hasattr(self.main_app, 'obras'):
                error_msg = f"Estrutura 'obras' não encontrada na aplicação principal."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            if not self.main_app.obras:
                error_msg = f"Estrutura 'obras' está vazia."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            # Verificar se a obra existe
            if obra_atual not in self.main_app.obras:
                error_msg = f"Obra '{obra_atual}' não encontrada na estrutura."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                print(f"[get_item_selecionado] Obras disponíveis: {list(self.main_app.obras.keys())}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            obra_dados = self.main_app.obras[obra_atual]
            
            # Verificar se é um dicionário válido
            if not isinstance(obra_dados, dict):
                error_msg = f"Dados da obra '{obra_atual}' não são um dicionário válido."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            # Verificar se o pavimento existe na obra
            if pavimento_lista not in obra_dados:
                error_msg = f"Pavimento '{pavimento_lista}' não encontrado na obra '{obra_atual}'."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                print(f"[get_item_selecionado] Pavimentos disponíveis na obra '{obra_atual}': {[k for k in obra_dados.keys() if k != 'pavimentos_data']}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            pav_dados = obra_dados[pavimento_lista]
            
            # Verificar se é um dicionário válido
            if not isinstance(pav_dados, dict):
                error_msg = f"Dados do pavimento '{pavimento_lista}' não são um dicionário válido."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            # Verificar se o número existe no pavimento
            if numero not in pav_dados:
                error_msg = f"Pilar '{numero}' não encontrado no pavimento '{pavimento_lista}' da obra '{obra_atual}'."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                print(f"[get_item_selecionado] Itens disponíveis no pavimento '{pavimento_lista}': {list(pav_dados.keys())}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            item_dados = pav_dados[numero]
            
            # Verificar se é um dicionário válido e contém 'dados'
            if not isinstance(item_dados, dict):
                error_msg = f"Dados do pilar '{numero}' não são um dicionário válido."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            if "dados" not in item_dados:
                error_msg = f"Chave 'dados' não encontrada no pilar '{numero}'."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                print(f"[get_item_selecionado] Chaves disponíveis no item: {list(item_dados.keys())}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            dados = item_dados["dados"]
            
            # Verificar se dados é um dicionário válido
            if not isinstance(dados, dict):
                error_msg = f"Conteúdo da chave 'dados' do pilar '{numero}' não é um dicionário válido."
                print(f"[get_item_selecionado] ERRO: {error_msg}")
                messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
                return None, None
            
            print(f"[get_item_selecionado] ✅ Dados encontrados: obra='{obra_atual}', pavimento='{pavimento_lista}', numero='{numero}'")
            
        except Exception as e:
            import traceback
            error_msg = f"Erro inesperado ao buscar dados na estrutura obras: {str(e)}"
            print(f"[get_item_selecionado] ERRO: {error_msg}")
            print(f"[get_item_selecionado] Traceback:\n{traceback.format_exc()}")
            messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
            return None, None
        
        if not dados:
            error_msg = f"Dados do pilar '{numero}' não foram encontrados após busca completa."
            print(f"[get_item_selecionado] ERRO: {error_msg}")
            messagebox.showerror("Erro", f"Dados do pilar {numero} não encontrados.\n\n{error_msg}\n\nObra: {obra_atual}\nPavimento: {pavimento_lista}")
            return None, None
        
        pavimento = dados.get("pavimento", "")
        
        # Verificar se o pavimento é válido (não é "Todos")
        if not pavimento or pavimento == "Todos":
            messagebox.showwarning("Aviso", "O item selecionado não possui um pavimento válido. Selecione um item de um pavimento específico.")
            return None, None
            
        return numero, dados
    
    def get_pavimento_selecionado(self):
        """
        Obtém o pavimento do item selecionado ou do filtro do combobox
        ATUALIZADO: Busca na estrutura obras[obra][pavimento][numero]
        
        Returns:
            str: Nome do pavimento ou None
        """
        
        # Primeiro verificar se o filtro do combobox está em "Todos"
        filtro_pavimento = self.main_app.pavimento_filtro_var.get()
        
        if filtro_pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro antes de usar os botões D.CAD.\n\nO sistema não funciona quando 'Todos' está selecionado no filtro.")
            return None
        
        # Verificar se há um item selecionado na lista
        selection = self.main_app.lista_pilares.selection()
        
        if selection:
            item = self.main_app.lista_pilares.item(selection[0])
            valores = item['values']
            numero = str(valores[0]).strip()
            pavimento_lista = str(valores[2]).strip()  # Pavimento da lista
            
            # Obter obra atual do filtro
            obra_atual = self.main_app.obra_filtro_var.get()
            if not obra_atual:
                obra_atual = "Obra Padrão"
            
            # ACESSO DIRETO: Buscar na estrutura obras[obra][pavimento][numero]
            try:
                # Verificar se a estrutura obras existe
                if not hasattr(self.main_app, 'obras') or not self.main_app.obras:
                    print(f"[get_pavimento_selecionado] AVISO: Estrutura 'obras' não encontrada ou vazia.")
                    # Continuar para usar o filtro como fallback
                elif obra_atual in self.main_app.obras:
                    obra_dados = self.main_app.obras[obra_atual]
                    if isinstance(obra_dados, dict) and pavimento_lista in obra_dados:
                        pav_dados = obra_dados[pavimento_lista]
                        if isinstance(pav_dados, dict) and numero in pav_dados:
                            item_dados = pav_dados[numero]
                            if isinstance(item_dados, dict) and "dados" in item_dados:
                                dados = item_dados["dados"]
                                if isinstance(dados, dict):
                                    pavimento = dados.get("pavimento", "")
                                    if pavimento and pavimento != "Todos":
                                        print(f"[get_pavimento_selecionado] ✅ Pavimento encontrado: '{pavimento}'")
                                        return pavimento
            except Exception as e:
                print(f"[get_pavimento_selecionado] ERRO ao buscar na estrutura obras: {e}")
                import traceback
                traceback.print_exc()
        
        # Se não há item selecionado ou o pavimento é "Todos", verificar o filtro do combobox
        if filtro_pavimento and filtro_pavimento != "Todos":
            return filtro_pavimento
        
        # Se chegou até aqui, não há pavimento válido selecionado
        messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
        return None
    
    def get_itens_por_pavimento(self, pavimento=None):
        """
        Obtém todos os itens de um pavimento específico
        CORREÇÃO: Busca APENAS da estrutura de OBRAS (mesma fonte da interface)
        
        Args:
            pavimento: Nome do pavimento (se None, usa o pavimento selecionado)
            
        Returns:
            list: Lista de tuplas (numero, dados) do pavimento
        """
        if not pavimento:
            pavimento = self.get_pavimento_selecionado()
            if not pavimento:
                return []
        
        # Obter obra atual DO COMBOBOX (não da seleção, mas do filtro!)
        obra_atual = self.main_app.obra_filtro_var.get()
        obra_atual_str = str(obra_atual).strip()
        
        
        itens_pavimento = []
        pavimento_normalizado = str(pavimento).strip()
        
        # CORREÇÃO CRÍTICA: A interface mostra itens da estrutura OBRAS, não de pilares_salvos
        # Portanto, devemos buscar APENAS da estrutura de obras (mesma fonte da interface)
        # Isso garante que processamos exatamente os mesmos itens que aparecem na lista
        
        try:
            if obra_atual_str and hasattr(self.main_app, 'obras') and self.main_app.obras:
                if obra_atual_str in self.main_app.obras:
                    obra_dados = self.main_app.obras[obra_atual_str]
                    
                    if isinstance(obra_dados, dict):
                        # Procurar o pavimento na obra (pode ter nome exato ou similar)
                        pav_dados = None
                        pavimento_encontrado = None
                        
                        # Tentar primeiro com nome exato
                        if pavimento_normalizado in obra_dados:
                            pav_dados = obra_dados[pavimento_normalizado]
                            pavimento_encontrado = pavimento_normalizado
                        else:
                            # Tentar com normalização (sem acentos, etc)
                            for pav_nome, pav_data in obra_dados.items():
                                if pav_nome == "pavimentos_data":
                                    continue
                                pav_nome_limpo = str(pav_nome).strip()
                                if pav_nome_limpo == pavimento_normalizado:
                                    pav_dados = pav_data
                                    pavimento_encontrado = pav_nome_limpo
                                    break
                        
                        if pav_dados is not None and isinstance(pav_dados, dict):
                            for numero, item_data in pav_dados.items():
                                try:
                                    if isinstance(item_data, dict) and "dados" in item_data:
                                        dados = item_data.get("dados", {})
                                        
                                        if not isinstance(dados, dict):
                                            continue
                                        
                                        # Extrair número limpo
                                        numero_limpo = str(numero).split("_")[0] if "_" in str(numero) else str(numero)
                                        
                                        # VALIDAÇÃO: Verificar se o item realmente pertence ao pavimento correto
                                        dados_pavimento = str(dados.get("pavimento", "")).strip()
                                        if dados_pavimento == pavimento_normalizado:
                                            itens_pavimento.append((numero_limpo, dados))
                                except Exception as e:
                                    continue
        except Exception as e:
            import traceback
        
        # REMOVER DUPLICATAS: Usar set para identificar itens únicos por número e nome
        itens_unicos = []
        itens_vistos = set()  # (numero, nome) para identificar duplicatas
        
        for numero_item, dados_item in itens_pavimento:
            nome_item = dados_item.get("nome", numero_item)
            chave_item = (str(numero_item).strip(), str(nome_item).strip())
            
            if chave_item not in itens_vistos:
                itens_unicos.append((numero_item, dados_item))
                itens_vistos.add(chave_item)
        
        if len(itens_unicos) != len(itens_pavimento):
            pass
        
        # Ordenar por número usando natsort para ordenação natural
        from natsort import natsorted
        itens_unicos = natsorted(itens_unicos, key=lambda x: x[0])
        return itens_unicos
    
    def criar_excel_temporario(self, itens_dados, nome_arquivo=None):
        """
        Cria um arquivo Excel temporário com os dados dos itens
        
        Args:
            itens_dados: Lista de tuplas (numero, dados) 
            nome_arquivo: Nome do arquivo (se None, cria temporário)
            
        Returns:
            str: Caminho do arquivo Excel criado
        """
        print(f"\n>>> ========================================")
        print(f">>> [DEBUG FUNCOES_AUX_2] criar_excel_temporario INICIADO")
        print(f">>> ========================================")
        print(f">>> [DEBUG FUNCOES_AUX_2]   - Total de itens: {len(itens_dados)}")
        print(f">>> [DEBUG FUNCOES_AUX_2]   - nome_arquivo: {nome_arquivo}")
        
        if nome_arquivo is None:
            # Criar arquivo temporário
            print(f">>> [DEBUG FUNCOES_AUX_2] Criando arquivo temporário...")
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='pilares_temp_')
            os.close(temp_fd)
            nome_arquivo = temp_path
            print(f">>> [DEBUG FUNCOES_AUX_2]   - Arquivo temporário criado: {nome_arquivo}")
        else:
            # Garantir que o nome do arquivo tenha extensão .xlsx
            if not nome_arquivo.endswith('.xlsx'):
                nome_arquivo = nome_arquivo + '.xlsx'
            print(f">>> [DEBUG FUNCOES_AUX_2]   - Usando arquivo fornecido: {nome_arquivo}")
        
        try:
            # Criar Excel básico sempre - mais confiável
            print(f">>> [DEBUG FUNCOES_AUX_2] Criando Workbook...")
            wb = Workbook()
            print(f">>> [DEBUG FUNCOES_AUX_2] ✅ Workbook criado")
            
            ws = wb.active
            print(f">>> [DEBUG FUNCOES_AUX_2] ✅ Worksheet ativo obtido")
            
            # Usar mapeamento Excel da aplicação principal
            print(f">>> [DEBUG FUNCOES_AUX_2] Carregando excel_mapping...")
            excel_mapping = getattr(self.main_app, 'EXCEL_MAPPING', {})
            print(f">>> [DEBUG FUNCOES_AUX_2]   - excel_mapping da aplicação principal: {len(excel_mapping) if excel_mapping else 0} campos")
            
            # FORÇAR uso do excel_mapping.py - sempre importar dele para garantir campos atualizados
            excel_mapping = None
            print(f">>> [DEBUG FUNCOES_AUX_2] Tentando importar excel_mapping do módulo...")
            try:
                # Tentar import relativo primeiro
                from ..utils.excel_mapping import EXCEL_MAPPING
                excel_mapping = EXCEL_MAPPING
                print(f">>> [DEBUG FUNCOES_AUX_2] ✅ excel_mapping importado (relativo): {len(excel_mapping) if excel_mapping else 0} campos")
            except ImportError as e:
                print(f">>> [DEBUG FUNCOES_AUX_2] ⚠️ Falha ao importar relativo: {e}")
                try:
                    # Fallback para import direto
                    utils_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'utils')
                    if utils_path not in sys.path:
                        sys.path.insert(0, utils_path)
                    from excel_mapping import EXCEL_MAPPING
                    excel_mapping = EXCEL_MAPPING
                    print(f">>> [DEBUG FUNCOES_AUX_2] ✅ excel_mapping importado (direto): {len(excel_mapping) if excel_mapping else 0} campos")
                except ImportError as e2:
                    print(f">>> [DEBUG FUNCOES_AUX_2] ⚠️ Falha ao importar direto: {e2}")
                    excel_mapping = None
            
            if not excel_mapping:
                # Fallback para mapeamento completo - incluindo alturas dos sarrafos
                excel_mapping = {
                    # Dados Gerais
                    "nome": 4,
                    "comprimento": 6,
                    "largura": 7,
                    "pavimento": 3,
                    "nivel_saida": 8,
                    "nivel_chegada": 9,
                    "altura": 12,
                    
                    # Parafusos
                    "par_1_2": 173,
                    "par_2_3": 174,
                    "par_3_4": 175,
                    "par_4_5": 176,
                    "par_5_6": 177,
                    "par_6_7": 178,
                    "par_7_8": 179,
                    "par_8_9": 180,
                    
                    # Grades - Grupo 1 (corrigido conforme excel_mapping.py)
                    "grade_1": 180,
                    "distancia_1": 182,
                    "grade_2": 183,
                    "distancia_2": 184,
                    "grade_3": 185,
                    
                    # Alturas dos Sarrafos (para GRADES)
                    "sarr_esquerda": 187,
                    "sarr1_altura": 188,
                    "sarr2_altura": 189,
                    "sarr3_altura": 190,
                    "sarr_direita": 191,
                    
                    # Detalhes das Grades
                    "detalhe_grade1_1": 192,
                    "detalhe_grade1_2": 193,
                    "detalhe_grade1_3": 194,
                    "detalhe_grade1_4": 195,
                    "detalhe_grade1_5": 196,
                    "detalhe_grade2_1": 197,
                    "detalhe_grade2_2": 198,
                    "detalhe_grade2_3": 199,
                    "detalhe_grade2_4": 200,
                    "detalhe_grade2_5": 201,
                    "detalhe_grade3_1": 202,
                    "detalhe_grade3_2": 203,
                    "detalhe_grade3_3": 204,
                    "detalhe_grade3_4": 205,
                    "detalhe_grade3_5": 206,
                    
                    # === ALTURAS DOS DETALHES - CONJUNTO 1 (GRADE A) ===
                    # Grade A - 1 (linhas 207-212)
                    "altura_detalhe_grade_a_1_0": 207,
                    "altura_detalhe_grade_a_1_1": 208,
                    "altura_detalhe_grade_a_1_2": 209,
                    "altura_detalhe_grade_a_1_3": 210,
                    "altura_detalhe_grade_a_1_4": 211,
                    "altura_detalhe_grade_a_1_5": 212,
                    
                    # Grade A - 2 (linhas 213-218)
                    "altura_detalhe_grade_a_2_0": 213,
                    "altura_detalhe_grade_a_2_1": 214,
                    "altura_detalhe_grade_a_2_2": 215,
                    "altura_detalhe_grade_a_2_3": 216,
                    "altura_detalhe_grade_a_2_4": 217,
                    "altura_detalhe_grade_a_2_5": 218,
                    
                    # Grade A - 3 (linhas 250-255)
                    "altura_detalhe_grade_a_3_0": 250,
                    "altura_detalhe_grade_a_3_1": 251,
                    "altura_detalhe_grade_a_3_2": 252,
                    "altura_detalhe_grade_a_3_3": 253,
                    "altura_detalhe_grade_a_3_4": 254,
                    "altura_detalhe_grade_a_3_5": 255,
                    
                    # === ALTURAS DOS DETALHES - CONJUNTO 2 (GRADE B) ===
                    # Grade B - 1 (linhas 256-261)
                    "altura_detalhe_grade_b_1_0": 256,
                    "altura_detalhe_grade_b_1_1": 257,
                    "altura_detalhe_grade_b_1_2": 258,
                    "altura_detalhe_grade_b_1_3": 259,
                    "altura_detalhe_grade_b_1_4": 260,
                    "altura_detalhe_grade_b_1_5": 261,
                    
                    # Grade B - 2 (linhas 262-267)
                    "altura_detalhe_grade_b_2_0": 262,
                    "altura_detalhe_grade_b_2_1": 263,
                    "altura_detalhe_grade_b_2_2": 264,
                    "altura_detalhe_grade_b_2_3": 265,
                    "altura_detalhe_grade_b_2_4": 266,
                    "altura_detalhe_grade_b_2_5": 267,
                    
                    # Grade B - 3 (linhas 268-273)
                    "altura_detalhe_grade_b_3_0": 268,
                    "altura_detalhe_grade_b_3_1": 269,
                    "altura_detalhe_grade_b_3_2": 270,
                    "altura_detalhe_grade_b_3_3": 271,
                    "altura_detalhe_grade_b_3_4": 272,
                    "altura_detalhe_grade_b_3_5": 273,
                    
                    # Grade Grupo 2 - Adicionado para suporte ao segundo grupo de grades
                    "grade_1_grupo2": 219,
                    "distancia_1_grupo2": 220,
                    "grade_2_grupo2": 221,
                    "distancia_2_grupo2": 222,
                    "grade_3_grupo2": 223,
                    
                    # Detalhes das Grades do Grupo 2
                    "detalhe_grade1_1_grupo2": 224,
                    "detalhe_grade1_2_grupo2": 225,
                    "detalhe_grade1_3_grupo2": 226,
                    "detalhe_grade1_4_grupo2": 227,
                    "detalhe_grade1_5_grupo2": 228,
                    "detalhe_grade2_1_grupo2": 229,
                    "detalhe_grade2_2_grupo2": 230,
                    "detalhe_grade2_3_grupo2": 231,
                    "detalhe_grade2_4_grupo2": 232,
                    "detalhe_grade2_5_grupo2": 233,
                    "detalhe_grade3_1_grupo2": 234,
                    "detalhe_grade3_2_grupo2": 235,
                    "detalhe_grade3_3_grupo2": 236,
                    "detalhe_grade3_4_grupo2": 237,
                    "detalhe_grade3_5_grupo2": 238,
                    
                    # Aberturas de Laje A (1=laje, 0=normal)
                    "abertura_laje_esq1_a": 224,
                    "abertura_laje_esq2_a": 225,
                    "abertura_laje_dir1_a": 226,
                    "abertura_laje_dir2_a": 227,
                    
                    # Aberturas de Laje B (1=laje, 0=normal)
                    "abertura_laje_esq1_b": 228,
                    "abertura_laje_esq2_b": 229,
                    "abertura_laje_dir1_b": 230,
                    "abertura_laje_dir2_b": 231,
                    
                    # Aberturas de Laje C (1=laje, 0=normal)
                    "abertura_laje_esq1_c": 232,
                    "abertura_laje_esq2_c": 233,
                    "abertura_laje_dir1_c": 234,
                    "abertura_laje_dir2_c": 235,
                    
                    # Aberturas de Laje D (1=laje, 0=normal)
                    "abertura_laje_esq1_d": 236,
                    "abertura_laje_esq2_d": 237,
                    "abertura_laje_dir1_d": 238,
                    "abertura_laje_dir2_d": 239,
                }
            
            # Exportar dados sequencialmente a partir da coluna E (índice 4)
            print(f"\n>>> [DEBUG FUNCOES_AUX_2] criar_excel_temporario - Iniciando loop de processamento de {len(itens_dados)} itens...")
            col_base = 4
            for idx, (numero, dados) in enumerate(itens_dados):
                print(f"\n>>> [DEBUG FUNCOES_AUX_2] ========================================")
                print(f">>> [DEBUG FUNCOES_AUX_2] [{idx+1}/{len(itens_dados)}] Processando item: {numero}")
                print(f">>> [DEBUG FUNCOES_AUX_2] ========================================")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Número: {numero}")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Dados disponíveis: {list(dados.keys())[:10]}...")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Nome: {dados.get('nome', 'N/A')}")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Pavimento: {dados.get('pavimento', 'N/A')}")
                
                coluna_idx = col_base + idx
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Coluna Excel: {get_column_letter(coluna_idx + 1)} (índice {coluna_idx})")
                
                # Número do pilar na linha 1
                ws.cell(row=1, column=coluna_idx + 1, value=numero)
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Número escrito na célula {get_column_letter(coluna_idx + 1)}1: {numero}")
                
                # DEBUG: Verificar estrutura dos dados
                
                # DEBUG: Verificar se altura_detalhe_grade_a_1_1 está no excel_mapping
                
                # DEBUG: Verificar campo altura 0
                
                # Preencher dados usando o mapeamento
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Iniciando preenchimento de campos...")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - excel_mapping carregado: {len(excel_mapping) if excel_mapping else 0} campos")
                campos_preenchidos = 0
                for campo, linha in excel_mapping.items():
                    valor = self._extrair_valor_campo(dados, campo)
                    
                    # DEBUG específico para altura_detalhe_grade_a_1_0
                    
                    # DEBUG específico para altura_detalhe_grade_a_1_1
                    
                    # Logar campos importantes
                    if campo in ['nome', 'pavimento'] or 'grade' in campo.lower() or 'nome' in campo.lower() or 'detalhe' in campo.lower():
                        if valor is not None:
                            print(f">>> [DEBUG FUNCOES_AUX_2]     - Campo '{campo}' (linha {linha}): valor encontrado = {valor}")
                    
                    # Escrever valor se não for None (incluindo 0)
                    if valor is not None:
                        try:
                            # Tentar converter para número
                            valor_numerico = float(str(valor).replace(",", "."))
                            if abs(round(valor_numerico) - valor_numerico) < 0.001:
                                valor_escrito = int(round(valor_numerico))
                                ws.cell(row=linha, column=coluna_idx + 1, value=valor_escrito)
                                # Log específico para detalhes_grupo2
                                if 'detalhe_grade' in campo and '_grupo2' in campo:
                                    print(f">>> [DEBUG ESCRITA EXCEL] Campo '{campo}' escrito na linha {linha}, coluna {get_column_letter(coluna_idx + 1)} com valor {valor_escrito}")
                                campos_preenchidos += 1
                            else:
                                valor_escrito = round(valor_numerico, 2)
                                ws.cell(row=linha, column=coluna_idx + 1, value=valor_escrito)
                                # Log específico para detalhes_grupo2
                                if 'detalhe_grade' in campo and '_grupo2' in campo:
                                    print(f">>> [DEBUG ESCRITA EXCEL] Campo '{campo}' escrito na linha {linha}, coluna {get_column_letter(coluna_idx + 1)} com valor {valor_escrito}")
                                campos_preenchidos += 1
                        except (ValueError, TypeError):
                            # Se não conseguir converter para número, escrever como string
                            valor_str = str(valor).strip()
                            if valor_str:  # Só escrever se não for string vazia
                                ws.cell(row=linha, column=coluna_idx + 1, value=valor_str)
                                campos_preenchidos += 1
                
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Total de campos preenchidos para {numero}: {campos_preenchidos}")
                print(f">>> [DEBUG FUNCOES_AUX_2]   - ✅ Item {numero} exportado para coluna {get_column_letter(coluna_idx + 1)}")

                # Escrita explícita dos campos de abertura de laje por item com base no 'tipo' salvo
                # para garantir consistência no modo pavimento
                try:
                    laje_map = {
                        'a': {'esq1': 224, 'esq2': 225, 'dir1': 226, 'dir2': 227},
                        'b': {'esq1': 228, 'esq2': 229, 'dir1': 230, 'dir2': 231},
                        'c': {'esq1': 232, 'esq2': 233, 'dir1': 234, 'dir2': 235},
                        'd': {'esq1': 236, 'esq2': 237, 'dir1': 238, 'dir2': 239},
                    }
                    if 'paineis' in dados:
                        for letra in ['A','B','C','D']:
                            painel = dados['paineis'].get(letra, {})
                            if not painel:
                                continue
                            for abertura_id in ['esq1','esq2','dir1','dir2']:
                                # CORREÇÃO: Derivar valor do tipo salvo da interface - NÃO usar valores fixos
                                valor_tipo = None
                                try:
                                    lado_comp = 'esquerda' if abertura_id.startswith('esq') else 'direita'
                                    num = abertura_id[-1]
                                    if (
                                        'aberturas' in painel and
                                        lado_comp in painel['aberturas'] and
                                        num in painel['aberturas'][lado_comp]
                                    ):
                                        # Ler valor da interface - não usar '0' fixo como padrão
                                        tipo_valor = painel['aberturas'][lado_comp][num].get('tipo')
                                        if tipo_valor is not None:
                                            valor_tipo = str(tipo_valor).strip()
                                except Exception:
                                    pass
                                
                                # Escrever na planilha na coluna do item
                                # Se não houver valor na interface, usar 0 (não ativo)
                                linha_dest = laje_map[letra.lower()][abertura_id]
                                valor_final = int(1 if valor_tipo == '1' else 0)
                                ws.cell(row=linha_dest, column=coluna_idx + 1, value=valor_final)
                except Exception:
                    pass
            
            # Adicionar dados do pilar especial em linhas especiais (linha 1000+)
            try:
                
                # Verificar se pilar especial está ativo
                # PRIORIDADE 1: Ler dos dados do item (quando processando item por item)
                # PRIORIDADE 2: Ler da interface principal (comportamento padrão do executar_abcd_excel)
                pilar_especial_ativo = False
                dados_pilar_especial = {}
                
                # PRIORIDADE 1: Tentar ler dos dados do item primeiro
                if itens_dados and len(itens_dados) > 0:
                    primeiro_item_dados = itens_dados[0][1] if len(itens_dados) > 0 else {}
                    if 'pilar_especial' in primeiro_item_dados:
                        pilar_especial_item = primeiro_item_dados.get('pilar_especial', {})
                        pilar_especial_ativo = pilar_especial_item.get('ativar_pilar_especial', False)
                        if pilar_especial_ativo:
                            dados_pilar_especial = {
                                'ativar': pilar_especial_ativo,
                                'tipo': pilar_especial_item.get('tipo_pilar_especial', ''),
                                'comp_1': pilar_especial_item.get('comp_1', ''),
                                'comp_2': pilar_especial_item.get('comp_2', ''),
                                'comp_3': pilar_especial_item.get('comp_3', ''),
                                'larg_1': pilar_especial_item.get('larg_1', ''),
                                'larg_2': pilar_especial_item.get('larg_2', ''),
                                'larg_3': pilar_especial_item.get('larg_3', ''),
                            }
                            print(f"[PILAR_ESPECIAL] Dados do pilar especial lidos do item: {dados_pilar_especial}")
                            
                            # Se os dados estão vazios, tentar ler da interface principal como fallback
                            # (mesma lógica do executar_abcd_excel que usa self.main_app)
                            if hasattr(self.main_app, 'ativar_pilar_especial'):
                                # Verificar se algum dado está vazio e tentar preencher da interface principal
                                if not dados_pilar_especial.get('tipo', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'tipo_pilar_especial'):
                                            dados_pilar_especial['tipo'] = self.main_app.tipo_pilar_especial.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('comp_1', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'comp_1'):
                                            dados_pilar_especial['comp_1'] = self.main_app.comp_1.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('comp_2', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'comp_2'):
                                            dados_pilar_especial['comp_2'] = self.main_app.comp_2.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('comp_3', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'comp_3'):
                                            dados_pilar_especial['comp_3'] = self.main_app.comp_3.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('larg_1', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'larg_1'):
                                            dados_pilar_especial['larg_1'] = self.main_app.larg_1.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('larg_2', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'larg_2'):
                                            dados_pilar_especial['larg_2'] = self.main_app.larg_2.get()
                                    except:
                                        pass
                                
                                if not dados_pilar_especial.get('larg_3', '').strip():
                                    try:
                                        if hasattr(self.main_app, 'larg_3'):
                                            dados_pilar_especial['larg_3'] = self.main_app.larg_3.get()
                                    except:
                                        pass
                                
                                print(f"[PILAR_ESPECIAL] Dados do pilar especial após fallback da interface principal: {dados_pilar_especial}")
                
                # PRIORIDADE 2: Se não encontrou nos dados do item, ler da interface principal
                # (mesma lógica do executar_abcd_excel que usa self.main_app)
                if not pilar_especial_ativo and hasattr(self.main_app, 'ativar_pilar_especial'):
                    pilar_especial_ativo = self.main_app.ativar_pilar_especial.get()
                    
                    if pilar_especial_ativo:
                        dados_pilar_especial = {
                            'ativar': pilar_especial_ativo,
                            'tipo': '',
                            'comp_1': '',
                            'comp_2': '',
                            'comp_3': '',
                            'larg_1': '',
                            'larg_2': '',
                            'larg_3': '',
                        }
                        
                        # Coletar valores SOMENTE se os atributos existirem na interface
                        if hasattr(self.main_app, 'tipo_pilar_especial'):
                            try:
                                dados_pilar_especial['tipo'] = self.main_app.tipo_pilar_especial.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'comp_1'):
                            try:
                                dados_pilar_especial['comp_1'] = self.main_app.comp_1.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'comp_2'):
                            try:
                                dados_pilar_especial['comp_2'] = self.main_app.comp_2.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'comp_3'):
                            try:
                                dados_pilar_especial['comp_3'] = self.main_app.comp_3.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'larg_1'):
                            try:
                                dados_pilar_especial['larg_1'] = self.main_app.larg_1.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'larg_2'):
                            try:
                                dados_pilar_especial['larg_2'] = self.main_app.larg_2.get()
                            except:
                                pass
                        
                        if hasattr(self.main_app, 'larg_3'):
                            try:
                                dados_pilar_especial['larg_3'] = self.main_app.larg_3.get()
                            except:
                                pass
                        print(f"[PILAR_ESPECIAL] Dados do pilar especial lidos da interface principal: {dados_pilar_especial}")
                
                # Salvar dados do pilar especial em linhas especiais (999+ para compensar linha adicional)
                linha_base = 999  # Ajustado para compensar linha adicional
                valor_checkbox = 1 if pilar_especial_ativo else 0
                ws.cell(row=linha_base + 1, column=1, value=valor_checkbox)  # Ativo/Inativo (linha 1000, coluna 1)
                ws.cell(row=linha_base + 2, column=1, value=dados_pilar_especial.get('tipo', ''))
                ws.cell(row=linha_base + 3, column=1, value=dados_pilar_especial.get('comp_1', ''))
                ws.cell(row=linha_base + 4, column=1, value=dados_pilar_especial.get('comp_2', ''))
                ws.cell(row=linha_base + 5, column=1, value=dados_pilar_especial.get('comp_3', ''))
                ws.cell(row=linha_base + 6, column=1, value=dados_pilar_especial.get('larg_1', ''))
                ws.cell(row=linha_base + 7, column=1, value=dados_pilar_especial.get('larg_2', ''))
                ws.cell(row=linha_base + 8, column=1, value=dados_pilar_especial.get('larg_3', ''))
                print(f"[PILAR_ESPECIAL] ✅ Checkbox salvo no Excel (linha {linha_base + 1}, coluna 1): {valor_checkbox}")
                print(f"[PILAR_ESPECIAL] ✅ Pilar especial ativo: {pilar_especial_ativo}")
                print(f"[PILAR_ESPECIAL] ✅ GRADE_EXCEL.py deve ler este valor para gerar scripts E, F, G, H")
                
                # Se pilar especial ativo, salvar dados das 6 grades especiais (A, B, E, F, G, H)
                if pilar_especial_ativo and hasattr(self.main_app, 'ativar_pilar_especial'):
                    try:
                        
                        # Importar EXCEL_MAPPING de forma robusta
                        try:
                            from ..utils.excel_mapping import EXCEL_MAPPING
                        except ImportError:
                            # Fallback para import direto
                            utils_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'utils')
                            if utils_path not in sys.path:
                                sys.path.insert(0, utils_path)
                            from excel_mapping import EXCEL_MAPPING
                        
                        # Contadores para estatísticas
                        total_campos_encontrados = 0
                        total_campos_salvos = 0
                        total_campos_nao_encontrados = 0
                        campos_nao_encontrados = []
                        
                        # Obter dados do item primeiro (se existirem)
                        dados_item_grades = {}
                        dados_item_detalhes = {}
                        dados_item_alturas = {}
                        primeiro_item_dados = {}
                        if itens_dados and len(itens_dados) > 0:
                            primeiro_item_dados = itens_dados[0][1] if len(itens_dados) > 0 else {}
                            # Verificar se há dados de grades salvos no item
                            if 'grades' in primeiro_item_dados:
                                dados_item_grades = primeiro_item_dados.get('grades', {})
                            if 'detalhes_grades' in primeiro_item_dados:
                                dados_item_detalhes = primeiro_item_dados.get('detalhes_grades', {})
                            if 'altura_detalhes_grades' in primeiro_item_dados:
                                dados_item_alturas = primeiro_item_dados.get('altura_detalhes_grades', {})
                        
                        # Para cada letra das grades especiais
                        for letra in ['a', 'b', 'e', 'f', 'g', 'h']:
                            letra_campos_encontrados = 0
                            letra_campos_salvos = 0
                            
                            # Grades principais (grade_{letra}_1, grade_{letra}_2, grade_{letra}_3)
                            for num in range(1, 4):
                                campo_grade = f"grade_{letra}_{num}"
                                total_campos_encontrados += 1
                                valor = None
                                
                                # PRIORIDADE 1: Ler dos dados salvos do item
                                if dados_item_grades and campo_grade in dados_item_grades:
                                    valor = dados_item_grades.get(campo_grade)
                                
                                # PRIORIDADE 2: Ler da interface principal (fallback)
                                if valor is None and hasattr(self.main_app, campo_grade):
                                    try:
                                        valor = getattr(self.main_app, campo_grade).get()
                                    except:
                                        pass
                                
                                # Salvar no Excel se encontrou valor
                                if valor is not None:
                                    linha_grade = EXCEL_MAPPING.get(campo_grade)
                                    if linha_grade:
                                        ws.cell(row=linha_grade, column=5, value=valor)  # Coluna E
                                        total_campos_salvos += 1
                                        letra_campos_salvos += 1
                                    else:
                                        campos_nao_encontrados.append(campo_grade)
                                else:
                                    total_campos_nao_encontrados += 1
                                    campos_nao_encontrados.append(campo_grade)
                            
                            # Distâncias (dist_{letra}_1, dist_{letra}_2)
                            for num in range(1, 3):
                                campo_dist = f"dist_{letra}_{num}"
                                total_campos_encontrados += 1
                                valor = None
                                
                                # PRIORIDADE 1: Ler dos dados salvos do item
                                if dados_item_grades and campo_dist in dados_item_grades:
                                    valor = dados_item_grades.get(campo_dist)
                                
                                # PRIORIDADE 2: Ler da interface principal (fallback)
                                if valor is None and hasattr(self.main_app, campo_dist):
                                    try:
                                        valor = getattr(self.main_app, campo_dist).get()
                                    except:
                                        pass
                                
                                # Salvar no Excel se encontrou valor
                                if valor is not None:
                                    linha_dist = EXCEL_MAPPING.get(campo_dist)
                                    if linha_dist:
                                        ws.cell(row=linha_dist, column=5, value=valor)  # Coluna E
                                        total_campos_salvos += 1
                                        letra_campos_salvos += 1
                                    else:
                                        campos_nao_encontrados.append(campo_dist)
                                else:
                                    total_campos_nao_encontrados += 1
                                    campos_nao_encontrados.append(campo_dist)
                            
                            # Detalhes das grades (detalhe_{letra}_{grade_num}_{detalhe_num})
                            for grade_num in range(1, 4):
                                for detalhe_num in range(1, 6):
                                    campo_detalhe = f"detalhe_{letra}_{grade_num}_{detalhe_num}"
                                    total_campos_encontrados += 1
                                    valor = None
                                    
                                    # PRIORIDADE 1: Ler dos dados salvos do item
                                    if dados_item_detalhes and campo_detalhe in dados_item_detalhes:
                                        valor = dados_item_detalhes.get(campo_detalhe)
                                    
                                    # PRIORIDADE 2: Ler da interface principal (fallback)
                                    if valor is None and hasattr(self.main_app, campo_detalhe):
                                        try:
                                            valor = getattr(self.main_app, campo_detalhe).get()
                                        except:
                                            pass
                                    
                                    # Salvar no Excel se encontrou valor
                                    if valor is not None:
                                        linha_detalhe = EXCEL_MAPPING.get(campo_detalhe)
                                        if linha_detalhe:
                                            ws.cell(row=linha_detalhe, column=5, value=valor)  # Coluna E
                                            total_campos_salvos += 1
                                            letra_campos_salvos += 1
                                        else:
                                            campos_nao_encontrados.append(campo_detalhe)
                                    else:
                                        total_campos_nao_encontrados += 1
                                        campos_nao_encontrados.append(campo_detalhe)
                                
                            # Alturas dos detalhes (altura_detalhe_{letra}_{grade_num}_{detalhe_num})
                            for grade_num in range(1, 4):
                                for altura_num in range(0, 6):  # 0 a 5
                                    campo_altura = f"altura_detalhe_{letra}_{grade_num}_{altura_num}"
                                    total_campos_encontrados += 1
                                    valor = None
                                    
                                    # PRIORIDADE 1: Ler dos dados salvos do item
                                    if dados_item_alturas and campo_altura in dados_item_alturas:
                                        valor = dados_item_alturas.get(campo_altura)
                                    
                                    # PRIORIDADE 2: Ler da interface principal (fallback)
                                    if valor is None and hasattr(self.main_app, campo_altura):
                                        try:
                                            valor = getattr(self.main_app, campo_altura).get()
                                        except:
                                            pass
                                    
                                    # Salvar no Excel se encontrou valor
                                    if valor is not None:
                                        linha_altura = EXCEL_MAPPING.get(campo_altura)
                                        if linha_altura:
                                            ws.cell(row=linha_altura, column=5, value=valor)  # Coluna E
                                            if valor and str(valor).strip():
                                                total_campos_salvos += 1
                                                letra_campos_salvos += 1
                                        else:
                                            if valor and str(valor).strip():
                                                campos_nao_encontrados.append(campo_altura)
                                    else:
                                        total_campos_nao_encontrados += 1
                                        campos_nao_encontrados.append(campo_altura)
                        
                        # CORREÇÃO: Salvar alturas e lajes dos painéis E, F, G, H no Excel temporário
                        # PRIORIDADE 1: Ler dos dados salvos do item (quando processando item por item)
                        # PRIORIDADE 2: Ler da interface principal (comportamento padrão do executar_abcd_excel)
                        for letra in ['E', 'F', 'G', 'H']:
                            # Obter dados do item primeiro (se existirem)
                            dados_painel_efgh = {}
                            if itens_dados and len(itens_dados) > 0:
                                primeiro_item_dados = itens_dados[0][1] if len(itens_dados) > 0 else {}
                                # Verificar se há dados dos painéis E, F, G, H salvos no item
                                if 'paineis' in primeiro_item_dados:
                                    painel_letra = letra.upper()
                                    if painel_letra in primeiro_item_dados['paineis']:
                                        dados_painel_efgh = primeiro_item_dados['paineis'][painel_letra]
                            
                            # Salvar lajes (laje_E, laje_F, etc.)
                            campo_laje = f"laje_{letra}"
                            valor_laje = None
                            
                            # PRIORIDADE 1: Ler dos dados salvos do item
                            if dados_painel_efgh and 'laje' in dados_painel_efgh:
                                valor_laje = dados_painel_efgh.get('laje')
                            
                            # PRIORIDADE 2: Ler da interface principal (fallback)
                            if valor_laje is None and hasattr(self.main_app, campo_laje):
                                try:
                                    valor_laje = getattr(self.main_app, campo_laje).get()
                                except:
                                    pass
                            
                            # Salvar no Excel se encontrou valor
                            if valor_laje is not None:
                                try:
                                    linha_laje = EXCEL_MAPPING.get(campo_laje)
                                    if linha_laje:
                                        ws.cell(row=linha_laje, column=5, value=valor_laje)  # Coluna E
                                except:
                                    pass
                            
                            # Salvar posição da laje (posicao_laje_E, posicao_laje_F, etc.)
                            campo_pos_laje = f"posicao_laje_{letra}"
                            valor_pos_laje = None
                            
                            # PRIORIDADE 1: Ler dos dados salvos do item
                            if dados_painel_efgh and 'posicao_laje' in dados_painel_efgh:
                                valor_pos_laje = dados_painel_efgh.get('posicao_laje')
                            
                            # PRIORIDADE 2: Ler da interface principal (fallback)
                            if valor_pos_laje is None and hasattr(self.main_app, campo_pos_laje):
                                try:
                                    valor_pos_laje = getattr(self.main_app, campo_pos_laje).get()
                                except:
                                    pass
                            
                            # Salvar no Excel se encontrou valor
                            if valor_pos_laje is not None:
                                try:
                                    linha_pos_laje = EXCEL_MAPPING.get(campo_pos_laje)
                                    if linha_pos_laje:
                                        ws.cell(row=linha_pos_laje, column=5, value=valor_pos_laje)  # Coluna E
                                except:
                                    pass
                            
                            # Salvar alturas (h1_E, h2_E, h3_E, h4_E, h5_E, etc.)
                            for altura_num in range(1, 6):  # h1 até h5
                                campo_altura = f"h{altura_num}_{letra}"
                                valor_altura = None
                                
                                # PRIORIDADE 1: Ler dos dados salvos do item
                                if dados_painel_efgh:
                                    # Verificar se há alturas salvas no item
                                    altura_key = f"h{altura_num}"
                                    if altura_key in dados_painel_efgh:
                                        valor_altura = dados_painel_efgh.get(altura_key)
                                
                                # PRIORIDADE 2: Ler da interface principal (fallback)
                                if valor_altura is None and hasattr(self.main_app, campo_altura):
                                    try:
                                        valor_altura = getattr(self.main_app, campo_altura).get()
                                    except:
                                        pass
                                
                                # Salvar no Excel se encontrou valor
                                if valor_altura is not None:
                                    try:
                                        linha_altura = EXCEL_MAPPING.get(campo_altura)
                                        if linha_altura:
                                            ws.cell(row=linha_altura, column=5, value=valor_altura)  # Coluna E
                                    except:
                                        pass
                        
                        if campos_nao_encontrados:
                            for campo in campos_nao_encontrados[:20]:
                                pass
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                
                # Calcular e salvar globais para pilar especial tipo L
                if pilar_especial_ativo and dados_pilar_especial.get('tipo') == 'L':
                    globais = self.calcular_globais_pilar_especial_L(
                        dados_pilar_especial.get('comp_1', ''),
                        dados_pilar_especial.get('comp_2', ''),
                        dados_pilar_especial.get('larg_1', ''),
                        dados_pilar_especial.get('larg_2', '')
                    )
                    
                    if globais:
                        # Salvar globais em linhas adicionais (1100+)
                        linha_globais = 1100
                        ws.cell(row=linha_globais + 1, column=1, value=globais.get('tipo_pilar', ''))
                        ws.cell(row=linha_globais + 2, column=1, value=globais.get('comp_1', ''))
                        ws.cell(row=linha_globais + 3, column=1, value=globais.get('comp_2', ''))
                        ws.cell(row=linha_globais + 4, column=1, value=globais.get('larg_1', ''))
                        ws.cell(row=linha_globais + 5, column=1, value=globais.get('larg_2', ''))
                        
                        # CORREÇÃO: PILAR 1 - Posições - usar valores calculados da interface (sem 0 fixo)
                        # Valores são calculados pela função calcular_globais_pilar_especial_L
                        ws.cell(row=linha_globais + 10, column=1, value=globais.get('pilar1_paia_posicao') or 0)
                        ws.cell(row=linha_globais + 11, column=1, value=globais.get('pilar1_paib_posicao') or 0)
                        ws.cell(row=linha_globais + 12, column=1, value=globais.get('pilar1_gradea_posicao') or 0)
                        ws.cell(row=linha_globais + 13, column=1, value=globais.get('pilar1_gradeb_posicao') or 0)
                        ws.cell(row=linha_globais + 14, column=1, value=globais.get('pilar1_parafuso_posicao') or 0)
                        ws.cell(row=linha_globais + 15, column=1, value=globais.get('pilar1_metala_posicao') or 0)
                        ws.cell(row=linha_globais + 16, column=1, value=globais.get('pilar1_metalb_posicao') or 0)
                        
                        # CORREÇÃO: PILAR 1 - Tamanhos - usar valores calculados da interface (sem 0 fixo)
                        ws.cell(row=linha_globais + 20, column=1, value=globais.get('pilar1_paia_tamanho') or 0)
                        ws.cell(row=linha_globais + 21, column=1, value=globais.get('pilar1_paib_tamanho') or 0)
                        ws.cell(row=linha_globais + 22, column=1, value=globais.get('pilar1_gradea_tamanho') or 0)
                        ws.cell(row=linha_globais + 23, column=1, value=globais.get('pilar1_gradeb_tamanho') or 0)
                        ws.cell(row=linha_globais + 24, column=1, value=globais.get('pilar1_parafuso_tamanho') or 0)
                        ws.cell(row=linha_globais + 25, column=1, value=globais.get('pilar1_metala_tamanho') or 0)
                        ws.cell(row=linha_globais + 26, column=1, value=globais.get('pilar1_metalb_tamanho') or 0)
                        
                        # CORREÇÃO: PILAR 2 - Posições - usar valores calculados da interface (sem 0 fixo)
                        ws.cell(row=linha_globais + 30, column=1, value=globais.get('pilar2_paia_posicao') or 0)
                        ws.cell(row=linha_globais + 31, column=1, value=globais.get('pilar2_paib_posicao') or 0)
                        ws.cell(row=linha_globais + 32, column=1, value=globais.get('pilar2_gradea_posicao') or 0)
                        ws.cell(row=linha_globais + 33, column=1, value=globais.get('pilar2_gradeb_posicao') or 0)
                        ws.cell(row=linha_globais + 34, column=1, value=globais.get('pilar2_parafuso_posicao') or 0)
                        ws.cell(row=linha_globais + 35, column=1, value=globais.get('pilar2_metala_posicao') or 0)
                        ws.cell(row=linha_globais + 36, column=1, value=globais.get('pilar2_metalb_posicao') or 0)
                        
                        # CORREÇÃO: PILAR 2 - Tamanhos - usar valores calculados da interface (sem 0 fixo)
                        ws.cell(row=linha_globais + 40, column=1, value=globais.get('pilar2_paia_tamanho') or 0)
                        ws.cell(row=linha_globais + 41, column=1, value=globais.get('pilar2_paib_tamanho') or 0)
                        ws.cell(row=linha_globais + 42, column=1, value=globais.get('pilar2_gradea_tamanho') or 0)
                        ws.cell(row=linha_globais + 43, column=1, value=globais.get('pilar2_gradeb_tamanho') or 0)
                        ws.cell(row=linha_globais + 44, column=1, value=globais.get('pilar2_parafuso_tamanho') or 0)
                        ws.cell(row=linha_globais + 45, column=1, value=globais.get('pilar2_metala_tamanho') or 0)
                        ws.cell(row=linha_globais + 46, column=1, value=globais.get('pilar2_metalb_tamanho') or 0)
                        
                        # CORREÇÃO: PERFIS METÁLICOS - Posições - usar valores calculados da interface (sem 0 fixo)
                        ws.cell(row=linha_globais + 50, column=1, value=globais.get('perfil_metalico_a_posicao') or 0)
                        ws.cell(row=linha_globais + 51, column=1, value=globais.get('perfil_metalico_b_posicao') or 0)
                        
                        # CORREÇÃO: PERFIS METÁLICOS - Tamanhos - usar valores calculados da interface (sem 0 fixo)
                        ws.cell(row=linha_globais + 55, column=1, value=globais.get('perfil_metalico_a_tamanho') or 0)
                        ws.cell(row=linha_globais + 56, column=1, value=globais.get('perfil_metalico_b_tamanho') or 0)
                        
                        
                        # Enviar globais para o robô
                        self.enviar_globais_para_robo(globais)
                
                
            except Exception as e:
                pass
            
            print(f">>> [DEBUG FUNCOES_AUX_2] Salvando arquivo Excel...")
            print(f">>> [DEBUG FUNCOES_AUX_2]   - Caminho: {nome_arquivo}")
            wb.save(nome_arquivo)
            print(f">>> [DEBUG FUNCOES_AUX_2] ✅ Arquivo Excel salvo com sucesso")
            print(f">>> [DEBUG FUNCOES_AUX_2]   - Arquivo existe: {os.path.exists(nome_arquivo)}")
            print(f">>> [DEBUG FUNCOES_AUX_2]   - Tamanho: {os.path.getsize(nome_arquivo) if os.path.exists(nome_arquivo) else 0} bytes")
            
            # Verificar se os valores detalhe_grade*_*_grupo2 foram salvos corretamente
            try:
                from openpyxl import load_workbook
                wb_verificacao = load_workbook(nome_arquivo, data_only=False)
                ws_verificacao = wb_verificacao.active
                coluna_letra = get_column_letter(coluna_idx + 1)
                print(f">>> [DEBUG VERIFICACAO] Verificando valores detalhe_grade*_*_grupo2 na coluna {coluna_letra}...")
                for grade_num in range(1, 4):
                    for detalhe_num in range(1, 6):
                        campo_key = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                        if campo_key in excel_mapping:
                            linha = excel_mapping[campo_key]
                            celula = f'{coluna_letra}{linha}'
                            valor_lido = ws_verificacao[celula].value
                            print(f">>> [DEBUG VERIFICACAO] {campo_key} ({celula}): {valor_lido} (tipo: {type(valor_lido)})")
                wb_verificacao.close()
            except Exception as e:
                print(f">>> [DEBUG VERIFICACAO] Erro ao verificar valores salvos: {e}")
            
            return nome_arquivo
            
        except Exception as e:
            print(f">>> [DEBUG FUNCOES_AUX_2] ❌ EXCEÇÃO ao criar Excel temporário: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao criar Excel temporário: {str(e)}")
            return None
    
    def _extrair_valor_campo(self, dados, campo):
        """
        Extrai valor de um campo dos dados do pilar
        
        Args:
            dados: Dicionário com dados do pilar
            campo: Nome do campo a extrair
            
        Returns:
            Valor do campo ou None
        """
        # Logar apenas campos importantes para não poluir muito
        logar_campo = campo in ['nome', 'pavimento'] or 'grade' in campo.lower() or 'nome' in campo.lower()
        
        if logar_campo:
            print(f">>> [DEBUG FUNCOES_AUX_2] _extrair_valor_campo: campo='{campo}'")
        
        # Verificar campos básicos
        if campo in dados:
            valor = dados[campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em dados básicos: {valor}")
            return valor
        
        # Verificar parafusos
        if "parafusos" in dados and campo in dados["parafusos"]:
            valor = dados["parafusos"][campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em parafusos: {valor}")
            return valor
        
        # Verificar grades
        if "grades" in dados and campo in dados["grades"]:
            valor = dados["grades"][campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em grades: {valor}")
            return valor
        
        # Verificar detalhes das grades
        if "detalhes_grades" in dados and campo in dados["detalhes_grades"]:
            valor = dados["detalhes_grades"][campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em detalhes_grades: {valor}")
            return valor
        
        # Verificar grades do Grupo 2
        if "grades_grupo2" in dados and campo in dados["grades_grupo2"]:
            valor = dados["grades_grupo2"][campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em grades_grupo2: {valor}")
            return valor
        
        # Verificar detalhes das grades do Grupo 2
        if "detalhes_grades_grupo2" in dados and campo in dados["detalhes_grades_grupo2"]:
            valor = dados["detalhes_grades_grupo2"][campo]
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em detalhes_grades_grupo2: {valor}")
            return valor
        
        # Verificar alturas dos detalhes das grades A e B
        # CORREÇÃO: Converter campos altura_detalhe_h_X_Y para altura_detalhe_grade_b_X_Y
        campo_corrigido = campo
        if campo.startswith("altura_detalhe_h_"):
            # Converter altura_detalhe_h_X_Y para altura_detalhe_grade_b_X_Y
            campo_corrigido = campo.replace("altura_detalhe_h_", "altura_detalhe_grade_b_", 1)
        
        # CORREÇÃO: Verificar primeiro altura_detalhes_grades_b se o campo for grade_b
        # Isso garante que os dados do conjunto 2 sejam usados corretamente
        if campo_corrigido.startswith("altura_detalhe_grade_b_"):
            if "altura_detalhes_grades_b" in dados:
                if campo_corrigido in dados["altura_detalhes_grades_b"]:
                    valor = dados["altura_detalhes_grades_b"][campo_corrigido]
                    if logar_campo:
                        print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em altura_detalhes_grades_b: {valor}")
                    return valor
        
        # Verificar altura_detalhes_grades (conjunto 1 - Grade A)
        if "altura_detalhes_grades" in dados:
            # DEBUG específico para campo 0
            if "altura_detalhe_grade_a_1_0" in campo_corrigido or "altura_detalhe_grade_b_1_0" in campo_corrigido:
                pass
            
            if campo_corrigido in dados["altura_detalhes_grades"]:
                valor = dados["altura_detalhes_grades"][campo_corrigido]
                if "altura_detalhe_grade_a_1_0" in campo_corrigido or "altura_detalhe_grade_b_1_0" in campo_corrigido:
                    pass
                if logar_campo:
                    print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em altura_detalhes_grades: {valor}")
                return valor
            else:
                if "_0" in campo_corrigido and "altura_detalhe" in campo_corrigido:
                    pass
        
        # Fallback: verificar altura_detalhes_grades_b novamente (caso o campo não comece com grade_b_)
        if "altura_detalhes_grades_b" in dados:
            if campo_corrigido in dados["altura_detalhes_grades_b"]:
                valor = dados["altura_detalhes_grades_b"][campo_corrigido]
                if logar_campo:
                    print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em altura_detalhes_grades_b (fallback): {valor}")
                return valor
        
        # VALORES para alturas dos sarrafos (para GRADES) - usar altura do pilar da interface
        # CORREÇÃO: NÃO usar valores fixos ou fictícios - sempre usar valores da interface
        if campo in ["sarr_esquerda", "sarr1_altura", "sarr2_altura", "sarr3_altura", "sarr_direita"]:
            altura_pilar = dados.get("altura")  # Usar altura do pilar da interface (sem fallback fixo)
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - Campo de sarrafo, altura_pilar: {altura_pilar}")
            if altura_pilar is None or altura_pilar == '':
                if logar_campo:
                    print(f">>> [DEBUG FUNCOES_AUX_2]   - ⚠️ Altura do pilar não encontrada, retornando None")
                return None  # Retornar None se não houver altura na interface
            if logar_campo:
                print(f">>> [DEBUG FUNCOES_AUX_2]   - ✅ Valor encontrado (altura do pilar): {altura_pilar}")
            return altura_pilar
        
        # Verificar painéis
        if "paineis" in dados:
            for letra in ['A', 'B', 'C', 'D']:
                if campo.endswith(f"_{letra}") and letra in dados["paineis"]:
                    painel = dados["paineis"][letra]
                    base = campo[:-2]
                    if base in painel:
                        valor = painel[base]
                        if logar_campo:
                            print(f">>> [DEBUG FUNCOES_AUX_2]   - Valor encontrado em painéis[{letra}]: {valor}")
                        return valor
                    elif "aberturas" in painel:
                        # Processar aberturas
                        for lado in ["esquerda", "direita"]:
                            if lado not in painel["aberturas"]:
                                continue
                            prefixo = "esq" if lado == "esquerda" else "dir"
                            for num in ["1", "2"]:
                                if num not in painel["aberturas"][lado]:
                                    continue
                                
                                # Mapeamento de nomes de campos
                                mapeamentos_campo = [
                                    (f"distancia_{prefixo}_{num}_{letra}", "distancia"),
                                    (f"largura_{prefixo}_{num}_{letra}", "largura"),
                                    (f"profundidade_{prefixo}_{num}_{letra}", "profundidade"),
                                    (f"posicao_{prefixo}_{num}_{letra}", "posicao"),
                                    (f"dist_{prefixo}_{num}_{letra}", "distancia"),
                                    (f"larg_{prefixo}_{num}_{letra}", "largura"),
                                    (f"prof_{prefixo}_{num}_{letra}", "profundidade"),
                                    (f"pos_{prefixo}_{num}_{letra}", "posicao")
                                ]
                                
                                for nome_campo, nome_interno in mapeamentos_campo:
                                    if campo == nome_campo:
                                        if nome_interno in painel["aberturas"][lado][num]:
                                            return painel["aberturas"][lado][num][nome_interno]
        
        # Verificar campos de abertura de laje
        if campo.startswith("abertura_laje_"):
            # Extrair painel e posição do campo (ex: abertura_laje_esq1_a)
            partes = campo.split("_")
            if len(partes) >= 4:
                posicao = partes[2]  # esq1, esq2, dir1, dir2
                painel = partes[3]   # a, b, c, d
                
                # Verificar se existe dados do painel
                if "paineis" in dados and painel.upper() in dados["paineis"]:
                    painel_data = dados["paineis"][painel.upper()]
                    # 1) Preferir estrutura nova 'aberturas_laje'
                    # CORREÇÃO: NÃO usar valores fixos - sempre usar valores da interface
                    if "aberturas_laje" in painel_data:
                        valor = painel_data["aberturas_laje"].get(posicao)
                        if valor is None or valor == '':
                            return None  # Retornar None se não houver valor na interface
                        return valor
                    # 2) Fallback: derivar do 'aberturas' usando o campo 'tipo'
                    if "aberturas" in painel_data:
                        try:
                            lado_comp = "esquerda" if posicao.startswith("esq") else "direita"
                            num = posicao[-1]
                            if (
                                lado_comp in painel_data["aberturas"] and
                                num in painel_data["aberturas"][lado_comp] and
                                "tipo" in painel_data["aberturas"][lado_comp][num]
                            ):
                                return str(painel_data["aberturas"][lado_comp][num]["tipo"]).strip() or "0"
                        except Exception:
                            pass
                    return "0"
        
        # Se não encontrou em nenhum lugar, retornar None
        if logar_campo:
            print(f">>> [DEBUG FUNCOES_AUX_2]   - ⚠️ Campo não encontrado em nenhum lugar, retornando None")
        return None
    
    def get_colunas_por_pavimento(self, pavimento):
        """
        Obtém lista de colunas (números de pilares) por pavimento
        
        Args:
            pavimento: Nome do pavimento
            
        Returns:
            list: Lista de números de pilares do pavimento
        """
        itens = self.get_itens_por_pavimento(pavimento)
        return [numero for numero, _ in itens]
    
    def get_nomes_por_pavimento(self, pavimento):
        """
        Obtém mapeamento de nomes por pavimento
        
        Args:
            pavimento: Nome do pavimento
            
        Returns:
            dict: Mapeamento {nome: numero}
        """
        itens = self.get_itens_por_pavimento(pavimento)
        return {dados.get("nome", numero): numero for numero, dados in itens}
    
    def sobrescrever_arquivo_comando(self, caminho_scr_gerado, caminho_destino):
        """
        Sobrescreve arquivo de comando
        
        Args:
            caminho_scr_gerado: Caminho do arquivo SCR gerado
            caminho_destino: Caminho de destino
        """
        try:
            if os.path.exists(caminho_scr_gerado):
                shutil.copy2(caminho_scr_gerado, caminho_destino)
        except Exception as e:
            pass
    
    # ===== FUNÇÕES ABCD =====
    
    def executar_abcd_excel(self):
        """1. ABCD - Executar Individual (desenhar item)"""
        numero, dados = self.get_item_selecionado()
        if not dados:
            return
        
        pavimento = dados.get("pavimento", "")
        nome = dados.get("nome", numero)
        
        # Verificar se o pavimento não é "Todos"
        if pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
            return
        
        # Criar Excel temporário com apenas este item
        excel_temp = self.criar_excel_temporario([(numero, dados)])
        if not excel_temp:
            return
        
        # Coluna sempre inicia em 'E' para o primeiro item
        coluna_letra = get_column_letter(5)  # Coluna 5 = 'E'
        
        try:
            # Executar script ABCD usando função auxiliar compatível com frozen
            resultado = self._executar_excel_wrapper("abcd_excel", excel_temp, coluna_letra, pavimento, interface_principal=self.main_app)
            
            if resultado and resultado.returncode == 0:
                # Messagebox removida para tornar o processo mais dinâmico
                print(f">>> ✅ ABCD Individual executado com sucesso para {nome}!")
                
                # Procurar e sobrescrever comando
                pasta_pavimento = os.path.join(self.scripts_dir, f"{pavimento.replace(' ', '_')}_ABCD")
                arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome}*.scr"))
                if arquivos_scr:
                    arquivo_recente = max(arquivos_scr, key=os.path.getmtime)
                    destino = os.path.join(robust_path_resolver.get_project_root(), "output", "comando_pilar_combinado_tpatpa.scr")
                    self.sobrescrever_arquivo_comando(arquivo_recente, destino)
            elif resultado:
                messagebox.showerror("Erro", f"Erro ao executar ABCD Individual:\n{resultado.stderr if hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
        
        finally:
            # Limpar arquivo temporário
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
    
    def limpar_pasta_pavimento(self, pavimento):
        """
        Limpa a pasta do pavimento antes de gerar novos scripts
        
        Args:
            pavimento: Nome do pavimento
            
        Returns:
            bool: True se a limpeza foi bem-sucedida, False caso contrário
        """
        try:
            import shutil
            import os
            
            # Definir caminho da pasta do pavimento
            pasta_pavimento = os.path.join(self.scripts_dir, f"{pavimento}_ABCD")
            
            
            if os.path.exists(pasta_pavimento):
                
                # Listar arquivos antes da remoção para debug
                arquivos_antes = []
                for root, dirs, files in os.walk(pasta_pavimento):
                    for file in files:
                        arquivos_antes.append(os.path.join(root, file))
                
                for arquivo in arquivos_antes:
                    pass
                
                # Remover toda a pasta e recriar
                try:
                    shutil.rmtree(pasta_pavimento)
                except Exception:
                    pass
                os.makedirs(pasta_pavimento, exist_ok=True)
                
                return True
            else:
                os.makedirs(pasta_pavimento, exist_ok=True)
                return True
                
        except Exception as e:
            return False

    def executar_todos_abcd_excel(self):
        """Executa script ABCD para todos os itens do pavimento"""
        return self.executar_pipeline_abcd()
    
    def executar_pipeline_abcd(self):
        """2. ABCD - Executar Todos + Combinar (desenhar pavimento)"""
        
        progress_window = None
        excel_temp = None
        
        try:
            pavimento = self.get_pavimento_selecionado()
            
            if not pavimento:
                return
            
            # Verificação adicional para garantir que o pavimento não seja "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
                return
            
            # Limpar pasta do pavimento antes de gerar novos scripts
            if not self.limpar_pasta_pavimento(pavimento):
                return
            
            itens = self.get_itens_por_pavimento(pavimento)
            
            if not itens:
                messagebox.showwarning("Aviso", f"Nenhum pilar encontrado para o pavimento {pavimento}")
                return
            
            # Criar Excel temporário com todos os itens do pavimento
            excel_temp = self.criar_excel_temporario(itens)
            if not excel_temp:
                return
            
            # Criar janela de progresso
            progress_window = ProgressWindow(self.main_app, "Processando Pavimento ABCD")
            progress_window.update_progress(0, "Iniciando geração de scripts...")
            # Executar pipeline completo ABCD (em processo)
            from glob import glob
            total_steps = len(itens) + 2  # Total de itens + combinador + ordenador
            current_step = 0

            # 1. Executar geração para todas as colunas diretamente
            total = len(itens)
            colunas_letras = [get_column_letter(5 + i) for i in range(total)]
            colunas_str = ','.join(colunas_letras)

            progress_window.update_progress(10, f"Gerando scripts para {total} itens...")
            
            # Executar script ABCD usando função auxiliar compatível com frozen
            resultado = self._executar_excel_wrapper("abcd_excel", excel_temp, colunas_str, pavimento, interface_principal=self.main_app)
            
            if not resultado or resultado.returncode != 0:
                messagebox.showerror("Erro", f"Erro ao executar ABCD Pipeline:\n{resultado.stderr if resultado and hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
                return
            
            current_step = len(itens)
            progress_window.update_progress(60, "Scripts gerados com sucesso!")
            
            # 2. Executar combinador
            progress_window.update_progress(70, "Executando combinador de scripts...")
            pasta_scripts = self.scripts_dir
            nome_pasta = str(pavimento).replace(' ', '_').replace('.0', '') + '_ABCD'
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            # Verificar se a pasta do pavimento existe
            if not os.path.exists(pasta_pavimento):
                messagebox.showerror("Erro", f"Pasta do pavimento não foi criada: {pasta_pavimento}")
                return
            
            # Verificar se há arquivos .scr na pasta
            arquivos_scr = [f for f in os.listdir(pasta_pavimento) if f.endswith('.scr')]
            for arquivo in arquivos_scr:
                pass
            
            # Executar combinador de forma compatível com frozen
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR.py")
            is_frozen = self._is_frozen()
            
            if is_frozen:
                # Em modo frozen: importar diretamente para evitar reinicializar o executável
                try:
                    import importlib
                    import importlib.util
                    
                    # Tentar múltiplas formas de importação devido ao espaço no nome do arquivo
                    combinador_filename = "Combinador_de_SCR.py"
                    combinador_full_path = os.path.join(self._get_project_root(), 'src', 'robots', combinador_filename)
                    
                    # Método 1: Importação via spec (funciona com espaços no nome)
                    if os.path.exists(combinador_full_path):
                        spec = importlib.util.spec_from_file_location("Combinador_de_SCR", combinador_full_path)
                        if spec and spec.loader:
                            module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(module)
                        else:
                            raise ImportError("Não foi possível criar spec do módulo")
                    else:
                        # Método 2: Tentar importação normal (pode não funcionar com espaços)
                        import_paths = [
                            'src.robots.Combinador_de_SCR',
                            'robots.Combinador_de_SCR',
                            'Combinador_de_SCR'
                        ]
                        
                        module = None
                        for import_path in import_paths:
                            try:
                                module = importlib.import_module(import_path)
                                break
                            except ImportError:
                                continue
                        
                        if not module:
                            raise ImportError("Não foi possível importar o módulo combinador")
                    
                    if module and hasattr(module, 'processar_arquivos'):
                        module.processar_arquivos(pasta_pavimento, mostrar_mensagem=False)
                        current_step += 1
                        progress_window.update_progress(85, "Scripts combinados com sucesso!")
                    else:
                        pass
                except Exception as e:
                    pass
                    import traceback
                    traceback.print_exc()
            else:
                # Em modo dev: usar subprocess normalmente
                if os.path.exists(combinador_path):
                    resultado = subprocess.run([sys.executable, combinador_path, pasta_pavimento], 
                                            capture_output=True, text=True, encoding='utf-8', errors='ignore')
                    if resultado.stdout:
                        pass
                    if resultado.stderr:
                        pass
                    current_step += 1
                    progress_window.update_progress(85, "Scripts combinados com sucesso!")
            
            # 3. Executar ordenador automaticamente
            progress_window.update_progress(90, "Executando ordenador...")
            pasta_combinados = os.path.join(pasta_pavimento, 'Combinados')
            if os.path.exists(pasta_combinados):
                # Importar ordenador ABCD usando path absoluto
                robots_path = os.path.join(self._get_project_root(), 'src', 'robots')
                if robots_path not in sys.path:
                    sys.path.insert(0, robots_path)
                from Ordenador_Cordenadas_abcd import ProcessadorCoordenadasABCD, atualizar_comando_pilar, get_config_path_abcd
                from pathlib import Path

                # Carregar configurações usando path resolver do ordenador
                try:
                    cfg_path = get_config_path_abcd()
                    if cfg_path.exists():
                        with open(cfg_path, 'r', encoding='utf-8') as f:
                            cfg = json.load(f)
                    else:
                        cfg = {
                            "numero_colunas": 4,
                            "distancia_x_colunas": 1585,
                            "distancia_y_linhas": -1148.6,
                            "distancia_y_extra": 0,
                            "linhas_para_extra": 0
                        }
                except Exception as e:
                    cfg = {
                        "numero_colunas": 4,
                        "distancia_x_colunas": 1585,
                        "distancia_y_linhas": -1148.6,
                        "distancia_y_extra": 0,
                        "linhas_para_extra": 0
                    }

                # Criar processador com as configurações
                processador = ProcessadorCoordenadasABCD(
                    numero_colunas=cfg.get("numero_colunas", 4),
                    distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                    distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                    distancia_y_extra=cfg.get("distancia_y_extra", 0),
                    linhas_para_extra=cfg.get("linhas_para_extra", 0)
                )

                # Processar arquivos
                arquivos_scr = sorted([f for f in os.listdir(pasta_combinados) if f.endswith('.scr')])
                for i, arq in enumerate(arquivos_scr):
                    progress = 90 + (i / len(arquivos_scr) * 8)  # Progresso de 90% a 98%
                    progress_window.update_progress(progress, f"Ordenando arquivo {i+1} de {len(arquivos_scr)}...")
                    processador.processar_arquivo(os.path.join(pasta_combinados, arq))

                # Atualizar comando pilar
                progress_window.update_progress(99, "Atualizando comando pilar...")
                atualizar_comando_pilar(pasta_combinados)
                current_step += 1
            else:
                messagebox.showerror("Erro", f"Pasta Combinados não encontrada em: {pasta_combinados}")
                return
            
            progress_window.update_progress(100, "Processo concluído com sucesso!")
            messagebox.showinfo("Sucesso", f"Pipeline ABCD completo executado para o pavimento {pavimento}!")
        
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro", f"Erro ao executar pipeline ABCD: {str(e)}")
        finally:
            # Limpar arquivo temporário e fechar janela de progresso
            if excel_temp and os.path.exists(excel_temp):
                try:
                    os.remove(excel_temp)
                except Exception as e:
                    pass
            if progress_window:
                try:
                    progress_window.close()
                except Exception as e:
                    pass
    
    def combinar_codigos_abcd(self):
        """Executa o combinador de scripts ABCD"""
        try:
            # Obter informações do pavimento selecionado
            pavimento = self.get_pavimento_selecionado()
            if not pavimento:
                return
            
            # Verificação adicional para garantir que o pavimento não seja "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
                return
            
            # Executar combinador
            pasta_scripts = self.scripts_dir
            nome_pasta = str(pavimento).replace(' ', '_').replace('.0', '') + '_ABCD'
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            # Executar combinador de forma compatível com frozen
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR.py")
            is_frozen = self._is_frozen()
            
            if is_frozen:
                # Em modo frozen: importar diretamente para evitar reinicializar o executável
                try:
                    import importlib
                    import importlib.util
                    
                    combinador_filename = "Combinador_de_SCR.py"
                    combinador_full_path = os.path.join(self._get_project_root(), 'src', 'robots', combinador_filename)
                    
                    if os.path.exists(combinador_full_path):
                        spec = importlib.util.spec_from_file_location("Combinador_de_SCR", combinador_full_path)
                        if spec and spec.loader:
                            module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(module)
                            
                            if hasattr(module, 'processar_arquivos'):
                                module.processar_arquivos(pasta_pavimento, mostrar_mensagem=False)
                                messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
                            else:
                                messagebox.showerror("Erro", "Função processar_arquivos não encontrada")
                        else:
                            messagebox.showerror("Erro", "Não foi possível importar o combinador")
                    else:
                        messagebox.showerror("Erro", f"Arquivo não encontrado: {combinador_path}")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao executar combinador:\n{e}")
            else:
                # Em modo dev: usar subprocess normalmente
                if os.path.exists(combinador_path):
                    resultado = subprocess.run([sys.executable, combinador_path, pasta_pavimento], cwd=self.utils_dir)
                    if resultado.returncode == 0:
                        messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
                    else:
                        messagebox.showerror("Erro", "Erro ao combinar scripts")
                else:
                    messagebox.showerror("Erro", f"Arquivo não encontrado: {combinador_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao combinar scripts ABCD:\n{e}")
    
    def ordenar_coordenadas_abcd(self):
        """Abre somente a janela de configurações do Ordenador ABCD"""
        try:
            sys.path.append(self.painel_dir)
            from interface_ordenador_abcd import JanelaConfiguracaoABCD, ConfiguracaoOrdenadorABCD

            root = tk.Tk()
            root.attributes('-alpha', 0)
            root.attributes('-toolwindow', True)
            root.geometry('1x1+0+0')

            config_mgr = ConfiguracaoOrdenadorABCD()
            JanelaConfiguracaoABCD(root, config_mgr)
            root.mainloop()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador ABCD:\n{e}")
    
    def abrir_robo_abcd(self):
        """3. ABCD - Ajustes Scripts (configurações)"""
        self._verificar_configuracoes_atuais("ABCD")
        
        is_frozen = self._is_frozen()
        
        try:
            if is_frozen:
                # Ambiente frozen: importar diretamente e abrir interface
                try:
                    import importlib
                    
                    # Tentar múltiplos caminhos de importação
                    module = None
                    import_paths = [
                        'robots.Robo_Pilar_ABCD',
                        'src.robots.Robo_Pilar_ABCD',
                        'Robo_Pilar_ABCD'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if not module:
                        raise ImportError("Não foi possível importar o módulo Robo_Pilar_ABCD")
                    
                    # Tentar abrir interface de configuração
                    # ABCD usa GeradorPilares (com "s"), não GeradorPilar
                    if hasattr(module, 'GeradorPilares') or hasattr(module, 'GeradorPilar') or hasattr(module, 'main'):
                        # Se o módulo tem classe ou função main, tentar executar
                        import tkinter as tk
                        # ABCD usa GeradorPilares (com "s")
                        if hasattr(module, 'GeradorPilares'):
                            # Criar janela root corretamente - não pode ser None
                            # Ocultar a janela principal (como no --config do robô original)
                            root = tk.Tk()
                            root.withdraw()  # Ocultar janela principal
                            # Passar root como parent (primeiro parâmetro) - obrigatório
                            gerador = module.GeradorPilares(root, None)
                            # Chamar método para mostrar configurações
                            if hasattr(gerador, 'criar_janela_configuracoes'):
                                gerador.criar_janela_configuracoes()
                            elif hasattr(gerador, 'mostrar_configuracao'):
                                gerador.mostrar_configuracao()
                            # Não chamar root.mainloop() aqui - isso bloqueia a execução
                            # A janela de configurações já tem seu próprio loop quando criada
                            messagebox.showinfo("Sucesso", "Configurações do Robô ABCD abertas!")
                        elif hasattr(module, 'GeradorPilar'):
                            # Fallback para GeradorPilar (sem "s") caso exista
                            root = tk.Tk()
                            root.withdraw()
                            gerador = module.GeradorPilar(root, None)
                            if hasattr(gerador, 'criar_janela_configuracoes'):
                                gerador.criar_janela_configuracoes()
                            elif hasattr(gerador, 'mostrar_configuracao'):
                                gerador.mostrar_configuracao()
                            messagebox.showinfo("Sucesso", "Configurações do Robô ABCD abertas!")
                        else:
                            # Tentar executar main com --config
                            import sys
                            old_argv = sys.argv[:]
                            sys.argv = ['Robo_Pilar_ABCD.py', '--config']
                            try:
                                if hasattr(module, 'main'):
                                    module.main()
                                messagebox.showinfo("Sucesso", "Configurações do Robô ABCD abertas!")
                            finally:
                                sys.argv = old_argv
                    else:
                        messagebox.showwarning("Aviso", "Interface de configuração do robô ABCD não disponível em ambiente frozen.")
                            
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao abrir configurações do robô ABCD:\n{e}")
            else:
                # Ambiente dev: usar subprocess
                import sys  # Garantir que sys está importado no escopo local
                robo_path = os.path.join(self.robos_dir, "Robo_Pilar_ABCD.py")
                if os.path.exists(robo_path):
                    subprocess.Popen([sys.executable, robo_path, "--config"])
                    messagebox.showinfo("Sucesso", "Configurações do Robô ABCD abertas!")
                else:
                    messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do robô ABCD:\n{e}")
    
    def abrir_configuracao_abcd(self):
        """4. ABCD - Configurar Ordenador (ordenamento)"""
        try:
            # Tentar importar e executar a interface do ordenador ABCD
            interfaces_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
            if interfaces_dir not in sys.path:
                sys.path.insert(0, interfaces_dir)
            
            from interface_ordenador_abcd import InterfaceOrdenador
            app = InterfaceOrdenador()
            app.iniciar()
            messagebox.showinfo("Sucesso", "Configurações do Ordenador ABCD abertas!")
        except Exception as e:
            # Tentar abrir como subprocess se o import falhar
            try:
                ordenador_path = os.path.join(interfaces_dir, "interface_ordenador_abcd.py")
                if os.path.exists(ordenador_path):
                    subprocess.Popen([sys.executable, ordenador_path])
                    messagebox.showinfo("Sucesso", "Configurações do Ordenador ABCD abertas!")
                else:
                    messagebox.showerror("Erro", f"Interface do ordenador não encontrada: {ordenador_path}")
            except Exception as e2:
                messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador ABCD:\n{e2}")
    
    # ===== FUNÇÕES CIMA =====
    
    def executar_cima_excel(self):
        """5. CIMA - Executar Individual (desenhar item)"""
        numero, dados = self.get_item_selecionado()
        if not dados:
            return
        
        pavimento = dados.get("pavimento", "")
        nome = dados.get("nome", numero)
        
        # Verificar se o pavimento não é "Todos"
        if pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
            return
        
        # Criar Excel temporário com apenas este item
        excel_temp = self.criar_excel_temporario([(numero, dados)])
        if not excel_temp:
            return
        
        # Coluna sempre inicia em 'E' para o primeiro item
        coluna_letra = get_column_letter(5)  # Coluna 5 = 'E'
        
        try:
            # Executar script CIMA usando função auxiliar compatível com frozen
            resultado = self._executar_excel_wrapper("cima_excel", excel_temp, coluna_letra, pavimento)
            
            if resultado and resultado.returncode == 0:
                messagebox.showinfo("Sucesso", f"CIMA Individual executado com sucesso para {nome}!")
                
                # Procurar e sobrescrever comando
                pasta_pavimento = os.path.join(self.scripts_dir, f"{pavimento.replace(' ', '_')}_CIMA")
                arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome}*.scr"))
                if arquivos_scr:
                    arquivo_recente = max(arquivos_scr, key=os.path.getmtime)
                    destino = os.path.join(robust_path_resolver.get_project_root(), "output", "comando_pilar_combinado.scr")
                    self.sobrescrever_arquivo_comando(arquivo_recente, destino)
            elif resultado:
                messagebox.showerror("Erro", f"Erro ao executar CIMA Individual:\n{resultado.stderr if hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
        
        finally:
            # Limpar arquivo temporário
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
    
    def executar_todos_cima_excel(self):
        """Executa script CIMA para todos os itens do pavimento"""
        resultado = self.executar_pipeline_cima()
        return resultado
    
    def executar_pipeline_cima(self):
        """6. CIMA - Executar Todos + Combinar (desenhar pavimento)"""
        
        pavimento = self.get_pavimento_selecionado()
        if not pavimento:
            return
        
        # Verificação adicional para garantir que o pavimento não seja "Todos"
        if pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
            return
        
        itens = self.get_itens_por_pavimento(pavimento)
        for idx, (numero, dados) in enumerate(itens):
            pass
        
        if not itens:
            messagebox.showwarning("Aviso", f"Nenhum pilar encontrado para o pavimento {pavimento}")
            return
        
        # Criar Excel temporário com todos os itens do pavimento
        excel_temp = self.criar_excel_temporario(itens)
        if not excel_temp:
            return
        
        # Criar janela de progresso
        progress_window = ProgressWindow(self.main_app, "Processando Pavimento CIMA")
        progress_window.update_progress(0, "Iniciando geração de scripts...")
        
        try:
            # Executar pipeline completo CIMA
            from glob import glob
            total_steps = len(itens) + 2  # Total de itens + combinador + ordenador
            current_step = 0
            
            # 1. Executar script para todos usando função auxiliar compatível com frozen
            total = len(itens)
            colunas_letras = [get_column_letter(5 + i) for i in range(total)]
            colunas_str = ','.join(colunas_letras)
            
            
            progress_window.update_progress(10, f"Gerando scripts para {total} itens...")
            
            resultado = self._executar_excel_wrapper("cima_excel", excel_temp, colunas_str, pavimento)
            
            if not resultado or resultado.returncode != 0:
                messagebox.showerror("Erro", f"Erro ao executar CIMA Pipeline:\n{resultado.stderr if resultado and hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
                return
            
            current_step = len(itens)
            progress_window.update_progress(60, "Scripts gerados com sucesso!")
            
            # 2. Executar combinador
            progress_window.update_progress(70, "Executando combinador de scripts...")
            pasta_scripts = self.scripts_dir
            
            # CORREÇÃO: Usar a mesma normalização que _dcad_pavimento_cima usa
            import unicodedata
            def normalizar_nome_pasta(texto):
                """Normaliza o nome removendo acentos (mesma função usada em _dcad_pavimento_cima)"""
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                texto_limpo = texto_limpo.strip('_')
                return texto_limpo
            
            pavimento_normalizado = normalizar_nome_pasta(pavimento)
            nome_pasta = f"{pavimento_normalizado}_CIMA"
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR _cima.py")
            if os.path.exists(combinador_path):
                subprocess.run([sys.executable, combinador_path, pasta_pavimento])
                current_step += 1
                progress_window.update_progress(85, "Scripts combinados com sucesso!")
            
            # 3. Executar ordenador automaticamente
            progress_window.update_progress(90, "Executando ordenador...")
            pasta_combinados = os.path.join(pasta_pavimento, 'Combinados')
            if os.path.exists(pasta_combinados):
                # Adicionar diretório dos robôs ao path
                robots_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "robots")
                if robots_dir not in sys.path:
                    sys.path.insert(0, robots_dir)
                # Importar ordenador CIMA usando path absoluto
                robots_path = os.path.join(self._get_project_root(), 'src', 'robots')
                if robots_path not in sys.path:
                    sys.path.insert(0, robots_path)
                from Ordenador_Cordenadas_cima import ProcessadorCoordenadasCima, atualizar_comando_pilar_cima
                from pathlib import Path

                # Carregar configurações
                cfg_path = Path('configuracao_ordenador_CIMA.json')
                if cfg_path.exists():
                    with open(cfg_path, 'r') as f:
                        cfg = json.load(f)
                else:
                    cfg = {
                        "numero_colunas": 4,
                        "distancia_x_colunas": 1585,
                        "distancia_y_linhas": -1148.6,
                        "distancia_y_extra": 0,
                        "linhas_para_extra": 0
                    }

                # Criar processador com as configurações
                processador = ProcessadorCoordenadasCima(
                    numero_colunas=cfg.get("numero_colunas", 4),
                    distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                    distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                    distancia_y_extra=cfg.get("distancia_y_extra", 0),
                    linhas_para_extra=cfg.get("linhas_para_extra", 0)
                )

                # Processar arquivos
                arquivos_scr = sorted([f for f in os.listdir(pasta_combinados) if f.endswith('.scr')])
                for i, arq in enumerate(arquivos_scr):
                    progress = 90 + (i / len(arquivos_scr) * 8)  # Progresso de 90% a 98%
                    progress_window.update_progress(progress, f"Ordenando arquivo {i+1} de {len(arquivos_scr)}...")
                    processador.processar_arquivo(os.path.join(pasta_combinados, arq))

                # Atualizar comando pilar
                progress_window.update_progress(99, "Atualizando comando pilar...")
                atualizar_comando_pilar_cima(pasta_combinados)
                current_step += 1
            else:
                messagebox.showerror("Erro", f"Pasta Combinados não encontrada em: {pasta_combinados}")
                return
            
            progress_window.update_progress(100, "Processo concluído com sucesso!")
            messagebox.showinfo("Sucesso", f"Pipeline CIMA completo executado para o pavimento {pavimento}!")
        
        finally:
            # Limpar arquivo temporário e fechar janela de progresso
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
            progress_window.close()
    
    def combinar_codigos_cima(self):
        """Executa o combinador de scripts CIMA"""
        try:
            # Obter informações do pavimento selecionado
            pavimento = self.get_pavimento_selecionado()
            if not pavimento:
                return
            
            # Verificação adicional para garantir que o pavimento não seja "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
                return
            
            # CORREÇÃO: Usar a mesma normalização que _dcad_pavimento_cima usa
            import unicodedata
            def normalizar_nome_pasta(texto):
                """Normaliza o nome removendo acentos (mesma função usada em _dcad_pavimento_cima)"""
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                texto_limpo = texto_limpo.strip('_')
                return texto_limpo
            
            # Executar combinador
            pasta_scripts = self.scripts_dir
            pavimento_normalizado = normalizar_nome_pasta(pavimento)
            nome_pasta = f"{pavimento_normalizado}_CIMA"
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            
            # CORREÇÃO: Executar combinador via importação direta (compatível com frozen)
            is_frozen = self._is_frozen()
            
            if is_frozen:
                # Ambiente frozen: importar e executar diretamente
                try:
                    import importlib
                    
                    # Tentar múltiplos caminhos de importação
                    module = None
                    import_paths = [
                        'robots.Combinador_de_SCR _cima',
                        'src.robots.Combinador_de_SCR _cima',
                        'Combinador_de_SCR _cima'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if not module:
                        raise ImportError("Não foi possível importar o módulo combinador")
                    
                    # Chamar a função processar_arquivos diretamente
                    if hasattr(module, 'processar_arquivos'):
                        module.processar_arquivos(pasta_pavimento)
                        messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
                    else:
                        raise AttributeError("Módulo combinador não possui função processar_arquivos")
                        
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao combinar scripts:\n{e}")
            else:
                # Ambiente de desenvolvimento: usar subprocess como fallback
                combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR _cima.py")
                if os.path.exists(combinador_path):
                    resultado = subprocess.run([sys.executable, combinador_path, pasta_pavimento], cwd=self.utils_dir)
                    if resultado.returncode == 0:
                        messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
                    else:
                        messagebox.showerror("Erro", "Erro ao combinar scripts")
                else:
                    messagebox.showerror("Erro", f"Arquivo não encontrado: {combinador_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao combinar scripts CIMA:\n{e}")
    
    def ordenar_coordenadas_cima(self):
        """Executa ordenador de coordenadas CIMA"""
        try:
            # Importar e abrir diretamente a janela de configurações
            sys.path.append(self.painel_dir)
            from interface_ordenador_cima import JanelaConfiguracaoCima, ConfiguracaoOrdenadorCima
            
            # Criar uma janela Tk temporária
            root = tk.Tk()
            root.attributes('-alpha', 0)  # Torna a janela completamente transparente
            root.attributes('-toolwindow', True)  # Remove a barra de título
            root.attributes('-topmost', True)  # Mantém a janela sempre no topo
            root.geometry('1x1+0+0')  # Define um tamanho mínimo e posiciona no canto
            
            # Criar o gerenciador de configuração
            config_manager = ConfiguracaoOrdenadorCima()
            
            # Abrir a janela de configurações
            janela_config = JanelaConfiguracaoCima(root, config_manager)
            
            # Iniciar o loop principal
            root.mainloop()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do ordenador CIMA:\n{e}")
    
    def abrir_robo_cima(self):
        """7. CIMA - Ajustes Scripts (configurações)"""
        self._verificar_configuracoes_atuais("CIMA")
        
        is_frozen = self._is_frozen()
        
        try:
            if is_frozen:
                # Ambiente frozen: importar diretamente e abrir interface
                try:
                    import importlib
                    
                    # Tentar múltiplos caminhos de importação
                    module = None
                    import_paths = [
                        'robots.Robo_Pilar_Visao_Cima',
                        'src.robots.Robo_Pilar_Visao_Cima',
                        'Robo_Pilar_Visao_Cima'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if not module:
                        raise ImportError("Não foi possível importar o módulo Robo_Pilar_Visao_Cima")
                    
                    # Tentar abrir interface de configuração
                    if hasattr(module, 'GeradorPilar'):
                        # Se o módulo tem a classe GeradorPilar, criar instância e mostrar janela de configuração
                        import tkinter as tk
                        # Criar janela root corretamente - não pode ser None
                        # Ocultar a janela principal (como no --config do robô original)
                        root = tk.Tk()
                        root.withdraw()  # Ocultar janela principal
                        # Passar root como parent (primeiro parâmetro) - obrigatório
                        gerador = module.GeradorPilar(root, None)
                        # Chamar método para mostrar configurações
                        if hasattr(gerador, 'criar_janela_configuracoes'):
                            gerador.criar_janela_configuracoes()
                        elif hasattr(gerador, 'mostrar_configuracao'):
                            gerador.mostrar_configuracao()
                        elif hasattr(gerador, 'config_window'):
                            gerador.config_window()
                        # Não chamar root.mainloop() aqui - isso bloqueia a execução
                        # A janela de configurações já tem seu próprio loop quando criada
                        messagebox.showinfo("Sucesso", "Configurações do Robô CIMA abertas!")
                    else:
                        # Fallback: tentar executar como script
                        if hasattr(module, '__main__'):
                            import sys
                            old_argv = sys.argv[:]
                            sys.argv = ['Robo_Pilar_Visao_Cima.py', '--config']
                            try:
                                if hasattr(module, 'main'):
                                    module.main()
                                messagebox.showinfo("Sucesso", "Configurações do Robô CIMA abertas!")
                            finally:
                                sys.argv = old_argv
                        else:
                            messagebox.showwarning("Aviso", "Interface de configuração do robô CIMA não disponível em ambiente frozen.")
                            
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao abrir configurações do robô CIMA:\n{e}")
            else:
                # Ambiente dev: usar subprocess
                import sys  # Garantir que sys está importado no escopo local
                robo_path = os.path.join(self.robos_dir, "Robo_Pilar_Visao_Cima.py")
                if os.path.exists(robo_path):
                    subprocess.Popen([sys.executable, robo_path, "--config"])
                    messagebox.showinfo("Sucesso", "Configurações do Robô CIMA abertas!")
                else:
                    messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do robô CIMA:\n{e}")
    
    def _verificar_configuracoes_atuais(self, robo):
        """Verifica as configurações atuais de um robô"""
        try:
            from pathlib import Path
            import json
            
            # Mapear nomes dos robôs para os arquivos corretos
            config_files = {
                "CIMA": "config_cima.json",
                "ABCD": "config_abcd.json", 
                "GRADE": "config_grades.json"
            }
            
            # Tentar múltiplos caminhos para encontrar config em frozen e dev
            config_name = config_files.get(robo, f"config_{robo.lower()}.json")
            config_file = None
            
            # Tentar paths em ordem de prioridade
            possible_paths = [
                os.path.join(robust_path_resolver.get_project_root(), "config", config_name),
                os.path.join(self._get_project_root(), "config", config_name),
                os.path.join("config", config_name),
                config_name
            ]
            
            for config_path in possible_paths:
                if os.path.exists(config_path):
                    config_file = config_path
                    break
            
            if config_file and os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                pass
        except Exception as e:
            pass
            import traceback
            traceback.print_exc()
    
    def abrir_configuracao_cima(self):
        """8. CIMA - Configurar Ordenador (ordenamento)"""
        try:
            # Tentar importar e executar a interface do ordenador CIMA
            interfaces_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
            if interfaces_dir not in sys.path:
                sys.path.insert(0, interfaces_dir)
            
            from interface_ordenador_cima import InterfaceOrdenadorCima
            app = InterfaceOrdenadorCima()
            app.iniciar()
            messagebox.showinfo("Sucesso", "Configurações do Ordenador CIMA abertas!")
        except Exception as e:
            # Tentar abrir como subprocess se o import falhar
            try:
                ordenador_path = os.path.join(interfaces_dir, "interface_ordenador_cima.py")
                if os.path.exists(ordenador_path):
                    subprocess.Popen([sys.executable, ordenador_path])
                    messagebox.showinfo("Sucesso", "Configurações do Ordenador CIMA abertas!")
                else:
                    messagebox.showerror("Erro", f"Interface do ordenador não encontrada: {ordenador_path}")
            except Exception as e2:
                messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador CIMA:\n{e2}")
    
    # ===== FUNÇÕES GRADES =====
    
    def executar_grade_excel(self):
        """9. GRADES - Executar Individual (desenhar item)"""
        print(f"\n>>> ========================================")
        print(f">>> [DEBUG GRADES] executar_grade_excel INICIADO")
        print(f">>> ========================================")
        
        print(f">>> [DEBUG GRADES] Obtendo item selecionado...")
        numero, dados = self.get_item_selecionado()
        print(f">>> [DEBUG GRADES]   - Número retornado: {numero}")
        print(f">>> [DEBUG GRADES]   - Dados retornados: {bool(dados)}")
        if dados:
            print(f">>> [DEBUG GRADES]   - Chaves dos dados: {list(dados.keys())[:10]}...")
        
        if not dados:
            print(f">>> [DEBUG GRADES] ❌ Nenhum item selecionado - retornando")
            return
        
        pavimento = dados.get("pavimento", "")
        nome = dados.get("nome", numero)
        print(f">>> [DEBUG GRADES] ✅ Item selecionado encontrado")
        print(f">>> [DEBUG GRADES]   - Número: {numero}")
        print(f">>> [DEBUG GRADES]   - Nome: {nome}")
        print(f">>> [DEBUG GRADES]   - Pavimento: {pavimento}")
        
        # Verificar se o pavimento não é "Todos"
        if pavimento == "Todos":
            print(f">>> [DEBUG GRADES] ❌ Pavimento é 'Todos' - retornando")
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.\n\nO botão D.CAD Item não funciona quando 'Todos' está selecionado.")
            return
        
        # Criar Excel temporário com apenas este item
        print(f">>> [DEBUG GRADES] Criando Excel temporário...")
        print(f">>> [DEBUG GRADES]   - Número: {numero}")
        print(f">>> [DEBUG GRADES]   - Dados: {list(dados.keys())[:10]}...")
        excel_temp = self.criar_excel_temporario([(numero, dados)])
        print(f">>> [DEBUG GRADES]   - Excel temporário criado: {excel_temp}")
        print(f">>> [DEBUG GRADES]   - Excel existe: {os.path.exists(excel_temp) if excel_temp else False}")
        
        if not excel_temp:
            print(f">>> [DEBUG GRADES] ❌ Falha ao criar Excel temporário - retornando")
            return
        
        # Coluna sempre inicia em 'E' para o primeiro item
        coluna_letra = get_column_letter(5)  # Coluna 5 = 'E'
        print(f">>> [DEBUG GRADES]   - Coluna: {coluna_letra}")
        
        try:
            # Executar script GRADES usando função auxiliar compatível com frozen
            print(f">>> [DEBUG GRADES] Chamando _executar_excel_wrapper...")
            print(f">>> [DEBUG GRADES]   - wrapper_name: grades_excel")
            print(f">>> [DEBUG GRADES]   - excel_temp: {excel_temp}")
            print(f">>> [DEBUG GRADES]   - coluna_letra: {coluna_letra}")
            print(f">>> [DEBUG GRADES]   - pavimento: {pavimento}")
            resultado = self._executar_excel_wrapper("grades_excel", excel_temp, coluna_letra, pavimento)
            print(f">>> [DEBUG GRADES]   - Resultado retornado: {resultado}")
            if resultado:
                print(f">>> [DEBUG GRADES]   - Return code: {getattr(resultado, 'returncode', 'N/A')}")
            
            if resultado and resultado.returncode == 0:
                # Messagebox removida para tornar o processo mais dinâmico
                print(f">>> ✅ GRADES Individual executado com sucesso para {nome}!")
                
                # Procurar e sobrescrever comando
                pasta_pavimento = os.path.join(self.scripts_dir, f"{pavimento.replace(' ', '_')}_GRADES")
                arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome}*.scr"))
                if arquivos_scr:
                    arquivo_recente = max(arquivos_scr, key=os.path.getmtime)
                    destino = os.path.join(robust_path_resolver.get_project_root(), "output", "comando_pilar_combinado.scr")
                    self.sobrescrever_arquivo_comando(arquivo_recente, destino)
            elif resultado:
                messagebox.showerror("Erro", f"Erro ao executar GRADES Individual:\n{resultado.stderr if hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
        
        finally:
            # Limpar arquivo temporário
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
    
    def executar_todos_grade_excel(self):
        """Executa script GRADE para todos os itens do pavimento"""
        return self.executar_pipeline_grades()
    
    def executar_pipeline_grades(self):
        """10. GRADES - Executar Todos + Combinar (desenhar pavimento)"""
        pavimento = self.get_pavimento_selecionado()
        if not pavimento:
            return
        
        # Verificação adicional para garantir que o pavimento não seja "Todos"
        if pavimento == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
            return
        
        itens = self.get_itens_por_pavimento(pavimento)
        if not itens:
            messagebox.showwarning("Aviso", f"Nenhum pilar encontrado para o pavimento {pavimento}")
            return
        
        # Criar Excel temporário com todos os itens do pavimento
        excel_temp = self.criar_excel_temporario(itens)
        if not excel_temp:
            return
        
        # Criar janela de progresso
        progress_window = ProgressWindow(self.main_app, "Processando Pavimento GRADES")
        progress_window.update_progress(0, "Iniciando geração de scripts...")
        
        try:
            # Executar pipeline completo GRADES
            from glob import glob
            total_steps = len(itens) + 2  # Total de itens + combinador + ordenador
            current_step = 0
            
            # 1. Executar script para todos usando função auxiliar compatível com frozen
            total = len(itens)
            colunas_letras = [get_column_letter(5 + i) for i in range(total)]
            colunas_str = ','.join(colunas_letras)
            
            progress_window.update_progress(10, f"Gerando scripts para {total} itens...")
            
            resultado = self._executar_excel_wrapper("grades_excel", excel_temp, colunas_str, pavimento)
            
            if not resultado or resultado.returncode != 0:
                messagebox.showerror("Erro", f"Erro ao executar GRADES Pipeline:\n{resultado.stderr if resultado and hasattr(resultado, 'stderr') else 'Erro desconhecido'}")
                return
            
            current_step = len(itens)
            progress_window.update_progress(60, "Scripts gerados com sucesso!")
            
            # 2. Executar combinador
            progress_window.update_progress(70, "Executando combinador de scripts...")
            pasta_scripts = self.scripts_dir
            nome_pasta = str(pavimento).replace(' ', '_').replace('.0', '') + '_GRADES'
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR_GRADES.py")
            if os.path.exists(combinador_path):
                subprocess.run([sys.executable, combinador_path, pasta_pavimento])
                current_step += 1
                progress_window.update_progress(85, "Scripts combinados com sucesso!")
            
            # 3. Executar ordenador automaticamente
            progress_window.update_progress(90, "Executando ordenador...")
            pasta_combinados = os.path.join(pasta_pavimento, 'Combinados')
            if os.path.exists(pasta_combinados):
                sys.path.append(self.painel_dir)
                # Importar ordenador GRADES usando path absoluto
                robots_path = os.path.join(self._get_project_root(), 'src', 'robots')
                if robots_path not in sys.path:
                    sys.path.insert(0, robots_path)
                from Ordenador_Cordenadas_grades import ProcessadorCoordenadasGrade, atualizar_comando_pilar_grade
                from pathlib import Path

                # Carregar configurações
                cfg_path = Path('configuracao_ordenador_GRADE.json')
                if cfg_path.exists():
                    with open(cfg_path, 'r') as f:
                        cfg = json.load(f)
                else:
                    cfg = {
                        "numero_colunas": 4,
                        "distancia_x_colunas": 1585,
                        "distancia_y_linhas": -1148.6,
                        "distancia_y_extra": 0,
                        "linhas_para_extra": 0
                    }

                # Criar processador com as configurações
                processador = ProcessadorCoordenadasGrade(
                    numero_colunas=cfg.get("numero_colunas", 4),
                    distancia_x_colunas=cfg.get("distancia_x_colunas", 1585),
                    distancia_y_linhas=cfg.get("distancia_y_linhas", -1148.6),
                    distancia_y_extra=cfg.get("distancia_y_extra", 0),
                    linhas_para_extra=cfg.get("linhas_para_extra", 0)
                )

                # Processar arquivos
                arquivos_scr = sorted([f for f in os.listdir(pasta_combinados) if f.endswith('.scr')])
                for i, arq in enumerate(arquivos_scr):
                    progress = 90 + (i / len(arquivos_scr) * 8)  # Progresso de 90% a 98%
                    progress_window.update_progress(progress, f"Ordenando arquivo {i+1} de {len(arquivos_scr)}...")
                    processador.processar_arquivo(os.path.join(pasta_combinados, arq))

                # Atualizar comando pilar
                progress_window.update_progress(99, "Atualizando comando pilar...")
                atualizar_comando_pilar_grade(pasta_combinados)
                current_step += 1
            else:
                messagebox.showerror("Erro", f"Pasta Combinados não encontrada em: {pasta_combinados}")
                return
            
            progress_window.update_progress(100, "Processo concluído com sucesso!")
            messagebox.showinfo("Sucesso", f"Pipeline GRADES completo executado para o pavimento {pavimento}!")
        
        finally:
            # Limpar arquivo temporário e fechar janela de progresso
            if os.path.exists(excel_temp):
                os.remove(excel_temp)
            progress_window.close()
    
    def executar_combinador_grades(self):
        """Executa o combinador de scripts GRADES"""
        try:
            # Obter informações do pavimento selecionado
            pavimento = self.get_pavimento_selecionado()
            if not pavimento:
                return
            
            # Verificação adicional para garantir que o pavimento não seja "Todos"
            if pavimento == "Todos":
                messagebox.showwarning("Aviso", "Selecione um pavimento específico no filtro ou um item de um pavimento específico na lista.")
                return
            
            # Executar combinador
            pasta_scripts = self.scripts_dir
            nome_pasta = str(pavimento).replace(' ', '_').replace('.0', '') + '_GRADES'
            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
            
            combinador_path = os.path.join(self._get_project_root(), 'src', 'robots', "Combinador_de_SCR_GRADES.py")
            if os.path.exists(combinador_path):
                resultado = subprocess.run([sys.executable, combinador_path, pasta_pavimento], cwd=self.utils_dir)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
                else:
                    messagebox.showerror("Erro", "Erro ao combinar scripts")
            else:
                messagebox.showerror("Erro", f"Arquivo não encontrado: {combinador_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao combinar scripts GRADES:\n{e}")
    
    def executar_ordenador_grades(self):
        """Abre somente a janela de configurações do Ordenador GRADES"""
        try:
            sys.path.append(self.painel_dir)
            from interface_ordenador_grades import JanelaConfiguracaoGrade, ConfiguracaoOrdenadorGrade

            root = tk.Tk()
            root.withdraw()  # Ocultar janela principal - não usar alpha=0 pois pode causar problemas
            root.attributes('-topmost', True)

            config_manager = ConfiguracaoOrdenadorGrade()
            JanelaConfiguracaoGrade(root, config_manager)
            
            # Usar threading para executar root.mainloop() sem bloquear
            import threading
            def run_mainloop():
                root.mainloop()
            thread = threading.Thread(target=run_mainloop, daemon=True)
            thread.start()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador GRADES:\n{e}")
    
    def abrir_robo_grades(self):
        """11. GRADES - Ajustes Scripts (configurações)"""
        self._verificar_configuracoes_atuais("GRADE")
        
        is_frozen = self._is_frozen()
        
        try:
            if is_frozen:
                # Ambiente frozen: importar diretamente e abrir interface
                try:
                    import importlib
                    
                    # Tentar múltiplos caminhos de importação
                    module = None
                    import_paths = [
                        'robots.ROBO_GRADES',
                        'src.robots.ROBO_GRADES',
                        'ROBO_GRADES'
                    ]
                    
                    for import_path in import_paths:
                        try:
                            module = importlib.import_module(import_path)
                            break
                        except ImportError:
                            continue
                    
                    if not module:
                        raise ImportError("Não foi possível importar o módulo ROBO_GRADES")
                    
                    # Tentar abrir interface de configuração
                    # GRADES não tem classe GeradorPilar - apenas função criar_janela_configuracoes()
                    if hasattr(module, 'criar_janela_configuracoes'):
                        # GRADES usa função criar_janela_configuracoes() que precisa de root global
                        import tkinter as tk
                        # Criar janela root corretamente - não pode ser None
                        # Ocultar a janela principal (como no --config do robô original)
                        root = tk.Tk()
                        root.withdraw()  # Ocultar janela principal
                        # IMPORTANTE: Definir root no módulo ANTES de chamar criar_janela_configuracoes()
                        # A função criar_janela_configuracoes() usa root como variável global
                        module.root = root
                        # Chamar função criar_janela_configuracoes diretamente
                        module.criar_janela_configuracoes()
                        # Usar threading para executar root.mainloop() sem bloquear
                        import threading
                        def run_mainloop():
                            root.mainloop()
                        thread = threading.Thread(target=run_mainloop, daemon=True)
                        thread.start()
                        messagebox.showinfo("Sucesso", "Configurações do Robô GRADES abertas!")
                    elif hasattr(module, 'GeradorPilar'):
                        # Fallback para GeradorPilar caso exista
                        import tkinter as tk
                        root = tk.Tk()
                        root.withdraw()
                        gerador = module.GeradorPilar(root, None)
                        if hasattr(gerador, 'criar_janela_configuracoes'):
                            gerador.criar_janela_configuracoes()
                        elif hasattr(gerador, 'mostrar_configuracao'):
                            gerador.mostrar_configuracao()
                        messagebox.showinfo("Sucesso", "Configurações do Robô GRADES abertas!")
                    else:
                        messagebox.showwarning("Aviso", "Interface de configuração do robô GRADES não disponível em ambiente frozen.")
                            
                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    messagebox.showerror("Erro", f"Erro ao abrir configurações do robô GRADES:\n{e}")
            else:
                # Ambiente dev: usar subprocess
                robo_path = os.path.join(self.robos_dir, "ROBO_GRADES.py")
                if os.path.exists(robo_path):
                    subprocess.Popen([sys.executable, robo_path, "--config"])
                    messagebox.showinfo("Sucesso", "Configurações do Robô GRADES abertas!")
                else:
                    messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir configurações do robô GRADES:\n{e}")
    
    def abrir_configuracao_grades(self):
        """12. GRADES - Configurar Ordenador (ordenamento)"""
        try:
            # Usar o mesmo método que executar_ordenador_grades usa
            # Tentar importar JanelaConfiguracaoGrade e ConfiguracaoOrdenadorGrade
            interfaces_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)))
            if interfaces_dir not in sys.path:
                sys.path.insert(0, interfaces_dir)
            
            # CORREÇÃO: A classe é InterfaceOrdenadorGrade (sem "s"), não InterfaceOrdenadorGrades
            try:
                from interface_ordenador_grades import JanelaConfiguracaoGrade, ConfiguracaoOrdenadorGrade
                import tkinter as tk
                
                # Criar janela root oculta
                root = tk.Tk()
                root.withdraw()  # Ocultar janela principal
                
                # Criar gerenciador de configuração
                config_manager = ConfiguracaoOrdenadorGrade()
                
                # Abrir janela de configurações
                janela_config = JanelaConfiguracaoGrade(root, config_manager)
                
                # Chamar root.mainloop() em thread separada para não bloquear
                import threading
                def run_mainloop():
                    root.mainloop()
                thread = threading.Thread(target=run_mainloop, daemon=True)
                thread.start()
                
                messagebox.showinfo("Sucesso", "Configurações do Ordenador GRADES abertas!")
            except ImportError as ie:
                # Se não conseguir importar, tentar usar executar_ordenador_grades
                # Usar executar_ordenador_grades que já tem threading
                try:
                    self.executar_ordenador_grades()
                    messagebox.showinfo("Sucesso", "Configurações do Ordenador GRADES abertas!")
                except Exception as e2:
                    messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador GRADES:\n{e2}")
        except Exception as e:
            import traceback
            traceback.print_exc()
            # Tentar usar executar_ordenador_grades como fallback
            try:
                self.executar_ordenador_grades()
            except Exception as e2:
                messagebox.showerror("Erro", f"Erro ao abrir configurações do Ordenador GRADES:\n{e2}")

    def calcular_globais_pilar_especial_L(self, comp_1, comp_2, larg_1, larg_2):
        """
        Calcula as globais para pilar especial tipo L conforme especificações.
        
        Args:
            comp_1: Comprimento do pilar 1
            comp_2: Comprimento do pilar 2
            larg_1: Largura do pilar 1
            larg_2: Largura do pilar 2
            
        Returns:
            dict: Dicionário com todas as globais calculadas
        """
        try:
            
            # Converter para float para cálculos
            comp_1 = float(comp_1) if comp_1 else 0
            comp_2 = float(comp_2) if comp_2 else 0
            larg_1 = float(larg_1) if larg_1 else 0
            larg_2 = float(larg_2) if larg_2 else 0
            
            globais = {
                # PILAR 1 - Alterações de Posição
                'pilar1_paia_posicao': 0,  # PAI.A = Comprimento do 1 + 22
                'pilar1_paib_posicao': 0,  # PAI.B = Comprimento 1 - largura do 2 - 18.5 + 11 + 4.4
                'pilar1_gradea_posicao': 0,  # Grade A = Comprimento do 1 + 22
                'pilar1_gradeb_posicao': 0,  # Grade B = Comprimento do 1 - largura do 2 - 18.5 + 11
                'pilar1_parafuso_posicao': 0,  # Parafuso 1 = Comprimento do 1 - largura do 2 - 30 + 22
                'pilar1_metala_posicao': 0,  # Metal A = Padrão
                'pilar1_metalb_posicao': 0,  # Metal B: ponto de início normal (sem offset em X)
                
                # PILAR 1 - Alterações de Tamanho
                'pilar1_paia_tamanho': comp_1 + 22,
                'pilar1_paib_tamanho': comp_1 - larg_2 - 18.5 + 11 + 4.4,
                'pilar1_gradea_tamanho': comp_1 + 22,
                'pilar1_gradeb_tamanho': comp_1 - larg_2 - 18.5 + 11,
                'pilar1_parafuso_tamanho': comp_1 - larg_2 - 30 + 22,
                'pilar1_metala_tamanho': 0,  # Padrão (Metal A normal)
                'pilar1_metalb_tamanho': -58.5,  # Metal B: -58.5 no tamanho total
                
                # PILAR 2 - Alterações de Posição
                'pilar2_paia_posicao': 0,  # PAI.A = Comprimento do 2 + largura do 1 + 11
                'pilar2_paib_posicao': 0,  # PAI.B = Comprimento 2 + 11
                'pilar2_gradea_posicao': 0,  # Grade A - posição removida
                'pilar2_gradeb_posicao': 0,  # Grade B - posição removida
                'pilar2_parafuso_posicao': 41,  # Parafuso 1 começa 41 para a direita
                'pilar2_metala_posicao': 58.5,  # METAL A: inicia mais para a direita pelo mesmo valor (+58.5)
                'pilar2_metalb_posicao': 0,  # Metal B normal
                
                # PILAR 2 - Alterações de Tamanho
                'pilar2_paia_tamanho': comp_2 + larg_1 + 11,
                'pilar2_paib_tamanho': comp_2 + 11,
                'pilar2_gradea_tamanho': comp_2 + larg_1 + 11,
                'pilar2_gradeb_tamanho': comp_2 + 11 - 18.5,
                'pilar2_parafuso_tamanho': comp_1 - larg_2 - 30 - 11,
                'pilar2_metala_tamanho': -58.5,  # METAL A: -58.5 no tamanho total
                'pilar2_metalb_tamanho': 0,  # Metal B normal
                
                # PERFIS METÁLICOS - Alterações de Posição
                'perfil_metalico_a_posicao': 47.5,  # Começa 47.5 para a direita
                'perfil_metalico_b_posicao': 47.5,  # Começa 47.5 para a direita
                
                # PERFIS METÁLICOS - Alterações de Tamanho
                'perfil_metalico_a_tamanho': -47.5,  # -47.5 em seu tamanho
                'perfil_metalico_b_tamanho': -47.5,  # -47.5 em seu tamanho
                
                # Informações gerais
                'tipo_pilar': 'L',
                'comp_1': comp_1,
                'comp_2': comp_2,
                'larg_1': larg_1,
                'larg_2': larg_2
            }
            
            
            return globais
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {}

    def enviar_globais_para_robo(self, globais):
        """
        Envia as globais calculadas para o robô visão cima.
        
        Args:
            globais: Dicionário com as globais calculadas
        """
        try:
            
            # Tentar importar e usar a função do robô
            try:
                from src.robots.Robo_Pilar_Visao_Cima import definir_globais_pilar_especial
                sucesso = definir_globais_pilar_especial(globais)
                if sucesso:
                    pass
                else:
                    pass
            except ImportError:
                pass
            except Exception as e:
                pass
                
        except Exception as e:
            pass


def inicializar_conector_painel_controle(main_app):
    """
    Inicializa o conector do painel de controle sem criar interface adicional
    
    Args:
        main_app: Instância da aplicação principal
        
    Returns:
        ConectorPainelControle: Instância do conector
    """
    return ConectorPainelControle(main_app)


# Exemplo de uso - para integrar na aplicação principal
if __name__ == "__main__":
    pass


class ProgressWindow:
    def __init__(self, parent, title="Processando..."):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.geometry("400x150")
        self.window.transient(parent)
        self.window.grab_set()
        
        # Centralizar a janela
        self.window.update_idletasks()
        width = self.window.winfo_width()
        height = self.window.winfo_height()
        x = (self.window.winfo_screenwidth() // 2) - (width // 2)
        y = (self.window.winfo_screenheight() // 2) - (height // 2)
        self.window.geometry(f"{width}x{height}+{x}+{y}")
        
        # Frame principal
        frame = ttk.Frame(self.window, padding="20")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Label de status
        self.status_label = ttk.Label(frame, text="Iniciando...", wraplength=350)
        self.status_label.grid(row=0, column=0, pady=(0, 10))
        
        # Barra de progresso
        self.progress = ttk.Progressbar(frame, length=350, mode='determinate')
        self.progress.grid(row=1, column=0, pady=(0, 10))
        
        # Label de porcentagem
        self.percent_label = ttk.Label(frame, text="0%")
        self.percent_label.grid(row=2, column=0)
        
        self.window.protocol("WM_DELETE_WINDOW", lambda: None)  # Desabilita o botão fechar
        
        # Configurar o tema
        style = ttk.Style()
        style.configure("TProgressbar", thickness=20)
        
        # Manter a janela sempre no topo
        self.window.attributes('-topmost', True)
        
        # Configurar o grid
        self.window.columnconfigure(0, weight=1)
        self.window.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
    
    def update_progress(self, value, status_text=None):
        """Atualiza o progresso e opcionalmente o texto de status"""
        self.progress['value'] = value
        self.percent_label['text'] = f"{int(value)}%"
        if status_text:
            self.status_label['text'] = status_text
        self.window.update()
    
    def close(self):
        """Fecha a janela de progresso"""
        self.window.destroy()
