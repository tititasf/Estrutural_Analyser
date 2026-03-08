
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

import openpyxl
import os
import sys
import time
import traceback
import argparse

# IMPORTAR HELPER FROZEN GLOBAL - garante que paths estão configurados
try:
    from _frozen_helper import ensure_paths
    ensure_paths()
except ImportError:
    try:
        from src._frozen_helper import ensure_paths
        ensure_paths()
    except ImportError:
        try:
            from _ensure_frozen import ensure
            ensure()
        except ImportError:
            # Fallback manual
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
                if script_dir not in sys.path:
                    sys.path.insert(0, script_dir)
                src_dir = os.path.join(script_dir, 'src')
                if os.path.exists(src_dir) and src_dir not in sys.path:
                    sys.path.insert(0, src_dir)
            else:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                parent_dir = os.path.dirname(current_dir)
                if parent_dir not in sys.path:
                    sys.path.insert(0, parent_dir)

# Configurar caminhos usando sistema robusto com múltiplos fallbacks
def setup_robots_path():
    """Configura o path para os robôs de forma robusta"""
    try:
        from config_paths import ROBOTS_DIR
        return ROBOTS_DIR
    except ImportError:
        try:
            from src.config_paths import ROBOTS_DIR
            return ROBOTS_DIR
        except ImportError:
            # Fallback para nova estrutura
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
                robots_dir = os.path.join(script_dir, 'src', 'robots')
                if os.path.exists(robots_dir):
                    return robots_dir
                return os.path.join(script_dir, 'robots')
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                return os.path.abspath(os.path.join(script_dir, "..", "robots"))

robos_dir = setup_robots_path()

# Adicionar diretório dos robôs ao path de forma confiável
if robos_dir and robos_dir not in sys.path:
    sys.path.insert(0, robos_dir)

print(f"Diretório de robôs adicionado ao path: {robos_dir}")
print(f"Path atual: {sys.path}")

# Verificar se o arquivo existe fisicamente
robo_path = os.path.join(robos_dir, "Robo_Pilar_Visao_Cima.py")
if not os.path.exists(robo_path):
    error_msg = f"ERRO: O arquivo Robo_Pilar_Visao_Cima.py não foi encontrado em {robos_dir}"
    print(error_msg)
    print(f"Arquivos disponíveis no diretório:")
    try:
        for arquivo in os.listdir(robos_dir):
            print(f"  - {arquivo}")
    except Exception as e:
        print(f"  Erro ao listar diretório: {e}")
    
    # Em vez de sys.exit(1), levantar exceção para ser tratada pelo código chamador
    # Isso evita fechar o programa inesperadamente
    raise FileNotFoundError(f"Robo_Pilar_Visao_Cima.py não encontrado em {robos_dir}")
else:
    print(f"Arquivo Robo_Pilar_Visao_Cima.py encontrado em: {robo_path}")

# Importar bibliotecas necessárias
try:
    # Importar módulos necessários que o Robo_Pilar_Visao_Cima.py usa
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import logging
    import matplotlib
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import math
    import json
    from datetime import datetime
    
    # Tentar importar win32com.client, mas continuar mesmo se falhar
    try:
        import win32com.client
        import pyautogui
        print("Módulos win32com e pyautogui importados com sucesso.")
    except ImportError:
        print("AVISO: Módulos win32com.client ou pyautogui não estão instalados.")
        print("Algumas funcionalidades relacionadas ao AutoCAD podem não funcionar.")
    
    # Configuração de logging
    log_file = os.path.join(os.path.dirname(robos_dir), "pilares.log")
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file), 
            logging.StreamHandler()
        ]
    )
    
    # Realizar a importação do módulo Robo_Pilar_Visao_Cima
    print("Tentando importar AplicacaoUnificada do módulo Robo_Pilar_Visao_Cima...")
    
    # Usar importlib para importação mais controlada
    import importlib.util
    spec = importlib.util.spec_from_file_location("Robo_Pilar_Visao_Cima", robo_path)
    robo_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(robo_module)
    
    # Agora extrair a classe AplicacaoUnificada do módulo carregado
    AplicacaoUnificada = robo_module.AplicacaoUnificada
    print("Importação realizada com sucesso!")
    
except ImportError as e:
    print(f"Erro ao importar: {e}")
    traceback.print_exc()
    sys.exit(1)
except Exception as e:
    print(f"Erro desconhecido: {e}")
    traceback.print_exc()
    sys.exit(1)

def modificar_x_inicial_script(script_content, novo_x):
    """
    Modifica todas as coordenadas X no script para usar o novo valor X inicial.
    Aplica um offset de novo_x a todas as coordenadas X encontradas.
    
    Args:
        script_content: Conteúdo do script original
        novo_x: Novo valor X inicial (ex: -196)
        
    Returns:
        str: Script modificado com as novas coordenadas X
    """
    try:
        import re
        
        # Verificar se script_content é válido
        if script_content is None:
            print("  Erro: script_content é None")
            return ""
        
        if not isinstance(script_content, str):
            print(f"  Erro: script_content não é string, é {type(script_content)}")
            return str(script_content) if script_content is not None else ""
        
        # Padrão mais específico para encontrar coordenadas X,Y no script
        # Procura por números seguidos de vírgula e outro número (coordenadas)
        # Evita modificar outros números que não sejam coordenadas
        padrao_coordenadas = r'(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)'
        
        coordenadas_modificadas = 0
        
        def substituir_coordenada(match):
            nonlocal coordenadas_modificadas
            
            x_original = float(match.group(1))
            y_original = float(match.group(2))
            
            # Aplicar offset X a todas as coordenadas
            novo_x_coord = x_original + novo_x
            
            coordenadas_modificadas += 1
            
            # Manter formato original (inteiro ou decimal)
            if '.' in match.group(1):
                x_formatado = f"{novo_x_coord:.1f}" if novo_x_coord != int(novo_x_coord) else f"{int(novo_x_coord)}.0"
            else:
                x_formatado = str(int(novo_x_coord)) if novo_x_coord == int(novo_x_coord) else f"{novo_x_coord:.1f}"
                
            if '.' in match.group(2):
                y_formatado = match.group(2)  # Manter Y original
            else:
                y_formatado = match.group(2)  # Manter Y original
            
            return f"{x_formatado},{y_formatado}"
        
        # Aplicar a substituição
        script_modificado = re.sub(padrao_coordenadas, substituir_coordenada, script_content)
        
        print(f"  Script modificado: {coordenadas_modificadas} coordenadas X deslocadas por {novo_x}")
        return script_modificado
        
    except Exception as e:
        print(f"  Erro ao modificar X inicial do script: {e}")
        import traceback
        traceback.print_exc()
        return script_content if script_content is not None else ""  # Retornar script original em caso de erro

def modificar_y_inicial_script(script_content, offset_y):
    """
    Modifica todas as coordenadas Y no script aplicando um offset.
    
    Args:
        script_content: Conteúdo do script original
        offset_y: Valor a ser somado às coordenadas Y (ex: -150 para descer 150)
        
    Returns:
        str: Script modificado com as novas coordenadas Y
    """
    try:
        import re
        
        # Verificar se script_content é válido
        if script_content is None:
            print("  Erro: script_content é None")
            return ""
        
        if not isinstance(script_content, str):
            print(f"  Erro: script_content não é string, é {type(script_content)}")
            return str(script_content) if script_content is not None else ""
        
        # Padrão para encontrar coordenadas X,Y no script
        padrao_coordenadas = r'(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)'
        
        coordenadas_modificadas = 0
        
        def substituir_coordenada(match):
            nonlocal coordenadas_modificadas
            
            x_original = float(match.group(1))
            y_original = float(match.group(2))
            
            # Aplicar offset Y a todas as coordenadas
            novo_y_coord = y_original + offset_y
            
            coordenadas_modificadas += 1
            
            # Manter formato original X
            x_formatado = match.group(1)  # Manter X original
                
            # Formatar Y
            if '.' in match.group(2):
                y_formatado = f"{novo_y_coord:.1f}" if novo_y_coord != int(novo_y_coord) else f"{int(novo_y_coord)}.0"
            else:
                y_formatado = str(int(novo_y_coord)) if novo_y_coord == int(novo_y_coord) else f"{novo_y_coord:.1f}"
            
            return f"{x_formatado},{y_formatado}"
        
        # Aplicar a substituição
        script_modificado = re.sub(padrao_coordenadas, substituir_coordenada, script_content)
        
        print(f"  Script modificado: {coordenadas_modificadas} coordenadas Y deslocadas por {offset_y}")
        return script_modificado
        
    except Exception as e:
        print(f"  Erro ao modificar Y inicial do script: {e}")
        import traceback
        traceback.print_exc()
        return script_content if script_content is not None else ""  # Retornar script original em caso de erro


def ajustar_zoom_inicial_y(script_content, novo_y):
    """
    Ajusta o Y do primeiro comando _ZOOM C x,y 10 encontrado no script.
    
    Mantém o X e o fator de zoom, alterando apenas o Y base do centro.
    """
    try:
        import re
        if not isinstance(script_content, str) or not script_content:
            return script_content or ""
        # Captura o primeiro bloco _ZOOM seguido de C x,y (com ou sem fator depois)
        padrao = re.compile(r"(_ZOOM\s*\r?\nC\s*(-?\d+(?:\.\d+)?),)(-?\d+(?:\.\d+)?)(\b)")
        def repl(m):
            prefixo = m.group(1)
            sufixo = m.group(4)
            # Formatar como inteiro ou decimal .0 para manter padrão do arquivo
            y = float(novo_y)
            y_fmt = f"{y:.1f}" if y != int(y) else (f"{int(y)}.0" if "." in m.group(3) else f"{int(y)}")
            return f"{prefixo}{y_fmt}{sufixo}"
        return padrao.sub(repl, script_content, count=1)
    except Exception:
        return script_content or ""



def calcular_globais_pilar_especial_L(comp_1, comp_2, larg_1, larg_2, detalhes_grades_especiais=None):
    """
    Calcula as globais para pilar especial tipo L conforme especificações.
    
    Args:
        comp_1: Comprimento do pilar 1
        comp_2: Comprimento do pilar 2
        larg_1: Largura do pilar 1
        larg_2: Largura do pilar 2
        detalhes_grades_especiais: Dicionário com os detalhes das grades A-H (opcional)
        
    Returns:
        dict: Dicionário com todas as globais calculadas
    """
    try:
        print(f"[GLOBAIS] Calculando globais para pilar especial tipo L...")
        print(f"   Dimensões: comp_1={comp_1}, comp_2={comp_2}, larg_1={larg_1}, larg_2={larg_2}")
        
        # Converter para float para cálculos
        comp_1 = float(comp_1) if comp_1 else 0
        comp_2 = float(comp_2) if comp_2 else 0
        larg_1 = float(larg_1) if larg_1 else 0
        larg_2 = float(larg_2) if larg_2 else 0
        
        globais = {
            # PILAR 1 - Alterações de Posição
            'pilar1_paia_posicao': 0,  # PAI.A = Comprimento do 1 + 22
            'pilar1_paib_posicao': 0,  # PAI.B = Comprimento 1 - largura do 2 - 18.5 + 11 + 4.4
            'pilar1_gradea_posicao': 0,  # Grade A = Comprimento do 1 + 22
            'pilar1_gradeb_posicao': 0,  # Grade B = Comprimento do 1 - largura do 2 - 18.5 + 11
            'pilar1_parafuso_posicao': 0,  # Parafuso 1 = Comprimento do 1 - largura do 2 - 30 + 22
            'pilar1_metala_posicao': 0,  # Metal A = Padrão
            'pilar1_metalb_posicao': 0,  # Metal B: ponto de início normal (sem offset em X)
            
            # PILAR 1 - Alterações de Tamanho
            'pilar1_paia_tamanho': comp_1 + 22,
            'pilar1_paib_tamanho': comp_1 - larg_2 - 18.5 + 11 + 4.4,
            'pilar1_gradea_tamanho': comp_1 + 22,
            'pilar1_gradeb_tamanho': comp_1 - larg_2 - 18.5 + 11,
            'pilar1_parafuso_tamanho': comp_1 - larg_2 - 30 + 22,
            'pilar1_metala_tamanho': 0,  # Padrão (Metal A normal)
            'pilar1_metalb_tamanho': -58.5,  # Metal B: -58.5 no tamanho total
            
            # PILAR 2 - Alterações de Posição
            'pilar2_paia_posicao': 0,  # PAI.A = Comprimento do 2 + largura do 1 + 11
            'pilar2_paib_posicao': 0,  # PAI.B = Comprimento 2 + 11
            'pilar2_gradea_posicao': 0,  # Grade A - posição removida
            'pilar2_gradeb_posicao': 0,  # Grade B - posição removida
            'pilar2_parafuso_posicao': 41,  # Parafuso 1 começa 41 para a direita
            'pilar2_metala_posicao': 58.5,  # METAL A: inicia mais para a direita pelo mesmo valor (+58.5)
            'pilar2_metalb_posicao': 0,  # Metal B normal
            
            # PILAR 2 - Alterações de Tamanho
            'pilar2_paia_tamanho': comp_2 + larg_1 + 11,
            'pilar2_paib_tamanho': comp_2 + 11,
            'pilar2_gradea_tamanho': comp_2 + larg_1 + 11,
            'pilar2_gradeb_tamanho': comp_2 + 11 - 18.5,
            'pilar2_parafuso_tamanho': comp_1 - larg_2 - 30 - 11,
            'pilar2_metala_tamanho': -58.5,  # METAL A: -58.5 no tamanho total
            'pilar2_metalb_tamanho': 0,  # Metal B normal
            
            # PERFIS METÁLICOS - Alterações de Posição
            'perfil_metalico_a_posicao': 47.5,  # Começa 47.5 para a direita
            'perfil_metalico_b_posicao': 47.5,  # Começa 47.5 para a direita
            
            # PERFIS METÁLICOS - Alterações de Tamanho
            'perfil_metalico_a_tamanho': -47.5,  # -47.5 em seu tamanho
            'perfil_metalico_b_tamanho': -47.5,  # -47.5 em seu tamanho
            
            # Informações gerais
            'tipo_pilar': 'L',
            'comp_1': comp_1,
            'comp_2': comp_2,
            'larg_1': larg_1,
            'larg_2': larg_2
        }
        
        # Adicionar detalhes das grades especiais se fornecidos
        if detalhes_grades_especiais:
            globais['detalhes_grades_especiais'] = detalhes_grades_especiais
            print(f"[GLOBAIS] Detalhes de grades especiais adicionados: {len(detalhes_grades_especiais)} valores")
        
        print(f"[GLOBAIS] Globais calculadas para pilar tipo L:")
        print(f"   Pilar 1 - PAI.A: pos={globais['pilar1_paia_posicao']}, tam={globais['pilar1_paia_tamanho']}")
        print(f"   Pilar 1 - PAI.B: pos={globais['pilar1_paib_posicao']}, tam={globais['pilar1_paib_tamanho']}")
        print(f"   Pilar 1 - Grade A: pos={globais['pilar1_gradea_posicao']}, tam={globais['pilar1_gradea_tamanho']}")
        print(f"   Pilar 1 - Grade B: pos={globais['pilar1_gradeb_posicao']}, tam={globais['pilar1_gradeb_tamanho']}")
        print(f"   Pilar 2 - PAI.A: pos={globais['pilar2_paia_posicao']}, tam={globais['pilar2_paia_tamanho']}")
        print(f"   Pilar 2 - PAI.B: pos={globais['pilar2_paib_posicao']}, tam={globais['pilar2_paib_tamanho']}")
        print(f"   Pilar 2 - Grade A: pos={globais['pilar2_gradea_posicao']}, tam={globais['pilar2_gradea_tamanho']}")
        print(f"   Pilar 2 - Grade B: pos={globais['pilar2_gradeb_posicao']}, tam={globais['pilar2_gradeb_tamanho']}")
        
        return globais
        
    except Exception as e:
        print(f"[GLOBAIS] Erro ao calcular globais do pilar especial tipo L: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}


def modificar_x_inicial_script(script_content, offset_x):
    """
    Modifica o X inicial do script aplicando um offset.
    """
    try:
        import re
        # Aplicar offset a coordenadas específicas
        def ajustar_coordenada(match):
            x_original = float(match.group(1))
            y_original = float(match.group(2))
            x_novo = x_original + offset_x
            return f"C {x_novo},{y_original}"
        
        # Aplicar modificação nas coordenadas
        script_modificado = re.sub(r'C\s+(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', ajustar_coordenada, script_content)
        return script_modificado
    except Exception as e:
        print(f"[OFFSET_X] Erro ao modificar X inicial: {e}")
        return script_content


def modificar_y_inicial_script(script_content, offset_y):
    """
    Modifica o Y inicial do script aplicando um offset.
    """
    try:
        import re
        # Aplicar offset a coordenadas específicas
        def ajustar_coordenada(match):
            x_original = float(match.group(1))
            y_original = float(match.group(2))
            y_novo = y_original + offset_y
            return f"C {x_original},{y_novo}"
        
        # Aplicar modificação nas coordenadas
        script_modificado = re.sub(r'C\s+(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)', ajustar_coordenada, script_content)
        return script_modificado
    except Exception as e:
        print(f"[OFFSET_Y] Erro ao modificar Y inicial: {e}")
        return script_content


def aplicar_offsets_metais_ao_script(script_content, gerador, globais_pilar_especial=None):
    """
    Aplica os offsets dos metais ao script para pilares especiais.
    
    Esta função verifica se há offsets de metais disponíveis e os aplica
    ao script usando as funções de modificação de coordenadas existentes.
    
    Args:
        script_content: Conteúdo do script original
        gerador: Instância do gerador de pilares
        globais_pilar_especial: Dicionário com globais do pilar especial
        
    Returns:
        str: Script modificado com os offsets dos metais aplicados
    """
    try:
        print("[VERIFICAR] Verificando offsets dos metais para aplicar ao script...")
        
        script_modificado = script_content
        
        # Usar globais do pilar especial se disponíveis
        if globais_pilar_especial and len(globais_pilar_especial) > 0:
            print(f"[GLOBAIS] Aplicando globais do pilar especial: {len(globais_pilar_especial)} valores")
            
            # PILAR 1 - Metal B: manter ponto de início normal (não deslocar o script inteiro)
            # Offsets de metais não devem deslocar todo o script; ajustes são feitos a nível robô nas coordenadas
            
            # PILAR 2 - Metal A: não deslocar o script inteiro (pos aplicaremos a nível de coordenada no robô)
            
            # PERFIS METÁLICOS: manter deslocamentos que movem script inteiro (conforme orientação)
            if 'perfil_metalico_a_posicao' in globais_pilar_especial and globais_pilar_especial['perfil_metalico_a_posicao'] != 0:
                offset_x = globais_pilar_especial['perfil_metalico_a_posicao']
                print(f"[OFFSET] Aplicando offset Perfil Metálico A (global): {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            if 'perfil_metalico_b_posicao' in globais_pilar_especial and globais_pilar_especial['perfil_metalico_b_posicao'] != 0:
                offset_x = globais_pilar_especial['perfil_metalico_b_posicao']
                print(f"[OFFSET] Aplicando offset Perfil Metálico B (global): {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            
            print("[OK] Globais do pilar especial aplicadas ao script")
            return script_modificado
            
        # Fallback para método antigo
        elif hasattr(gerador, 'offsets_metais_recebidos') and gerador.offsets_metais_recebidos:
            offsets = gerador.offsets_metais_recebidos
            print(f"[OFFSETS] Offsets dos metais encontrados (método antigo): {offsets}")
            
            # Aplicar offset de posição do Metal B Pilar 1 (-11)
            if 'metal_b_1_posicao' in offsets and offsets['metal_b_1_posicao'] != 0:
                offset_x = offsets['metal_b_1_posicao']
                print(f"[OFFSET] Aplicando offset Metal B Pilar 1: {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            
            # Aplicar offset do Metal A Pilar 2 (valor negativo)
            if 'metal_a_2_offset' in offsets and offsets['metal_a_2_offset'] != 0:
                offset_x = offsets['metal_a_2_offset']
                print(f"[OFFSET] Aplicando offset Metal A Pilar 2: {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            
            # Aplicar offset de posição dos Perfis Metálicos (+47.5)
            if 'perfil_metalico_a_offset' in offsets and offsets['perfil_metalico_a_offset'] != 0:
                offset_x = offsets['perfil_metalico_a_offset']
                print(f"[OFFSET] Aplicando offset Perfil Metálico A: {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            
            if 'perfil_metalico_b_offset' in offsets and offsets['perfil_metalico_b_offset'] != 0:
                offset_x = offsets['perfil_metalico_b_offset']
                print(f"[OFFSET] Aplicando offset Perfil Metálico B: {offset_x}")
                script_modificado = modificar_x_inicial_script(script_modificado, offset_x)
            
            print("[OK] Offsets dos metais aplicados ao script com sucesso (método antigo)")
            return script_modificado
            
        else:
            print("[INFO] Nenhum offset de metal encontrado para aplicar")
            return script_content
            
    except Exception as e:
        print(f"[ERRO] Erro ao aplicar offsets dos metais: {str(e)}")
        return script_content


def preencher_campos_diretamente_e_gerar_scripts(dados_pilar):
    """
    Preenche os campos da interface VISAO CIMA diretamente com os dados fornecidos
    e gera os scripts .scr correspondentes.

    Args:
        dados_pilar: Dicionário com os dados do pilar, incluindo interface_principal
    """
    
    # ADICIONAR LOG DE DEPURAÇÃO
    import time
    import os
    import sys
    timestamp_func_inicio = time.time()
    print(f"[CIMA_SCRIPT] ===== INÍCIO DA FUNÇÃO preencher_campos_diretamente_e_gerar_scripts =====")
    print(f"[CIMA_SCRIPT] Timestamp: {timestamp_func_inicio}")
    print(f"[CIMA_SCRIPT] PID: {os.getpid()}")
    print(f"[CIMA_SCRIPT] Dados recebidos: {dados_pilar}")
    print(f"[DEBUG] FUNÇÃO CHAMADA - dados_pilar: {dados_pilar}")
    print(f"[DEBUG] FUNÇÃO CHAMADA - tipo dados_pilar: {type(dados_pilar)}")
    print(f"[DEBUG] FUNÇÃO CHAMADA - chaves disponíveis: {list(dados_pilar.keys()) if isinstance(dados_pilar, dict) else 'NÃO É DICT'}")

    try:
        # CORREÇÃO: Criar a aplicação UMA ÚNICA VEZ
        print("[DEBUG-FUNC] Iniciando aplicação VISAO CIMA...")
        app = AplicacaoUnificada()
        gerador = app.gerador
        print("[DEBUG-FUNC] Aplicação VISAO CIMA criada com sucesso")
        
        print(f"[DEBUG-FUNC] === PROCESSANDO DADOS DIRETOS ===")
        
        # 1. Preencher dados básicos
        print(f"[DEBUG] Chegou na seção 1 - preencher dados básicos")
        campos_basicos = ['pavimento', 'nome', 'comprimento', 'largura']
        for campo in campos_basicos:
            valor = dados_pilar.get(campo, "")
            print(f"  Preenchendo {campo}: {valor}")
            
            if campo == 'pavimento':
                gerador.pavimento_entry.delete(0, 'end')
                gerador.pavimento_entry.insert(0, valor if valor is not None else "")
            elif campo == 'nome':
                gerador.nome_pilar_entry.delete(0, 'end')
                gerador.nome_pilar_entry.insert(0, valor if valor is not None else "")
            elif campo == 'comprimento':
                gerador.comprimento_pilar_entry.delete(0, 'end')
                gerador.comprimento_pilar_entry.insert(0, valor if valor is not None else "")
            elif campo == 'largura':
                gerador.largura_pilar_entry.delete(0, 'end')
                gerador.largura_pilar_entry.insert(0, valor if valor is not None else "")

        # 2. Calcular valores iniciais
        print("Calculando valores iniciais...")
        print(f"[DEBUG] Chegou na seção 2 - calcular valores iniciais")
        gerador.calcular_valores()

        # 3. Preencher parafusos
        print("Preenchendo parafusos...")
        print(f"[DEBUG] Chegou na seção 3 - preencher parafusos")
        
        # Verificar se há parafusos especiais para usar
        parafusos_especiais = dados_pilar.get('parafusos_especiais', {})
        pilar_especial_ativo = dados_pilar.get('ativar_pilar_especial', False)
        
        # Limpar todos os campos de parafusos primeiro (sem disparar traces)
        print("[DEBUG] Limpando campos de parafusos - desabilitando traces temporariamente")
        
        # Desabilitar traces temporariamente para evitar erros durante limpeza
        traces_salvos = []
        for entry in gerador.parafuso_entries:
            # Salvar e remover traces
            if hasattr(entry, 'valor_var') and entry.valor_var:
                traces = entry.valor_var.trace_info()
                traces_salvos.append((entry, traces))
                for trace in traces:
                    entry.valor_var.trace_remove(trace[0], trace[1])
            
            # Limpar campo
            entry.delete(0, 'end')
            entry.insert(0, "0")
        
        # Restaurar traces
        for entry, traces in traces_salvos:
            for trace in traces:
                entry.valor_var.trace_add(trace[0], trace[1])
        
        print("[DEBUG] Limpeza concluída - traces restaurados")
        
        # Decidir quais parafusos usar baseado no pilar especial
        # CORREÇÃO: Buscar parafusos_especiais tanto no nível raiz quanto dentro de pilar_especial
        parafusos_especiais_raiz = dados_pilar.get('parafusos_especiais', {})
        parafusos_especiais_pilar = dados_pilar.get('pilar_especial', {}).get('parafusos_especiais', {})
        # Usar o que estiver disponível (prioridade para pilar_especial)
        parafusos_especiais = parafusos_especiais_pilar if parafusos_especiais_pilar else parafusos_especiais_raiz
        
        print(f"[PARAFUSOS_ESPECIAIS] 🔍 Verificando parafusos especiais...")
        print(f"[PARAFUSOS_ESPECIAIS] Pilar especial ativo: {pilar_especial_ativo}")
        print(f"[PARAFUSOS_ESPECIAIS] Parafusos especiais (raiz): {parafusos_especiais_raiz}")
        print(f"[PARAFUSOS_ESPECIAIS] Parafusos especiais (pilar_especial): {parafusos_especiais_pilar}")
        print(f"[PARAFUSOS_ESPECIAIS] Parafusos especiais (usado): {parafusos_especiais}")
        
        if pilar_especial_ativo and parafusos_especiais:
            print(f"[PARAFUSOS_ESPECIAIS] 🔍 Pilar especial ativo - usando parafusos especiais")
            
            # Para pilares especiais, usar parafusos A (todos os campos)
            # O robô decidirá internamente se usar A ou E baseado no pilar_atual
            parafusos_a = parafusos_especiais.get('parafusos_a', [])
            
            if parafusos_a:
                print(f"[PARAFUSOS_ESPECIAIS] Preenchendo campos com parafusos especiais A: {parafusos_a}")
                # Preencher com parafusos especiais A (todos os campos)
                for i, valor in enumerate(parafusos_a):
                    if i < len(gerador.parafuso_entries):
                        gerador.parafuso_entries[i].delete(0, 'end')
                        gerador.parafuso_entries[i].insert(0, str(int(valor)) if valor > 0 else "0")
                        print(f"  Preenchendo parafuso {i+1} com valor especial A: {valor}")
                        
                # Transferir também os parafusos especiais para o gerador
                # CORREÇÃO: Buscar parafusos_especiais tanto no nível raiz quanto dentro de pilar_especial
                dados_parafusos_especiais_pilar = dados_pilar.get('pilar_especial', {}).get('parafusos_especiais', {})
                dados_parafusos_especiais_raiz = dados_pilar.get('parafusos_especiais', {})
                # Usar o que estiver disponível (prioridade para pilar_especial)
                dados_parafusos_especiais = dados_parafusos_especiais_pilar if dados_parafusos_especiais_pilar else dados_parafusos_especiais_raiz
                
                try:
                    from Robo_Pilar_Visao_Cima import definir_dados_parafusos_especiais
                    print(f"[PARAFUSOS_ESPECIAIS] Estrutura dos dados antes da transferência: {type(dados_parafusos_especiais)}")
                    if isinstance(dados_parafusos_especiais, dict):
                        print(f"[PARAFUSOS_ESPECIAIS] Chaves disponíveis: {list(dados_parafusos_especiais.keys())}")
                        if 'parafusos_e' in dados_parafusos_especiais:
                            print(f"[PARAFUSOS_ESPECIAIS] Parafusos E antes da transferência: {dados_parafusos_especiais.get('parafusos_e')}")
                        if 'parafusos_a' in dados_parafusos_especiais:
                            print(f"[PARAFUSOS_ESPECIAIS] Parafusos A antes da transferência: {dados_parafusos_especiais.get('parafusos_a')}")
                    definir_dados_parafusos_especiais(dados_parafusos_especiais)
                    print(f"[PARAFUSOS_ESPECIAIS] Dados de parafusos especiais transferidos para o gerador via função global")
                except ImportError:
                    # Fallback: definir diretamente no gerador
                    print(f"[PARAFUSOS_ESPECIAIS] Estrutura dos dados (fallback): {type(dados_parafusos_especiais)}")
                    if isinstance(dados_parafusos_especiais, dict):
                        print(f"[PARAFUSOS_ESPECIAIS] Chaves disponíveis (fallback): {list(dados_parafusos_especiais.keys())}")
                    gerador._dados_parafusos_especiais = dados_parafusos_especiais
                    print(f"[PARAFUSOS_ESPECIAIS] Dados de parafusos especiais definidos diretamente no gerador")
            else:
                print(f"[PARAFUSOS_ESPECIAIS] Nenhum parafuso especial A encontrado - usando parafusos normais")
        else:
            # Usar parafusos normais se não houver especiais ou pilar especial não estiver ativo
            print(f"[PARAFUSOS] Usando parafusos normais")
            parafusos_ordem = [
                'parafuso_p1_p2',
                'parafuso_p2_p3',
                'parafuso_p3_p4',
                'parafuso_p4_p5',
                'parafuso_p5_p6',
                'parafuso_p6_p7',
                'parafuso_p7_p8',
                'parafuso_p8_p9'
            ]

            # Preencher parafusos na ordem
            for i, parafuso in enumerate(parafusos_ordem):
                if i < len(gerador.parafuso_entries):
                    valor = dados_pilar.get(parafuso, 0)
                    if valor is not None:
                        gerador.parafuso_entries[i].delete(0, 'end')
                        gerador.parafuso_entries[i].insert(0, valor)
                        print(f"  Preenchendo {parafuso}: {valor}")

        # 4. NÃO recalcular após parafusos - manter valores originais
        print("Mantendo valores originais após parafusos...")
        print(f"[DEBUG] Chegou na seção 4 - mantendo valores originais")

        # 4.5. Preencher parafusos do Pilar 2 se necessário (parafusos E)
        print("Verificando parafusos do Pilar 2...")
        print(f"[DEBUG] Chegou na seção 4.5 - parafusos pilar 2")
        
        # Verificar se há parafusos especiais E para Pilar 2
        if parafusos_especiais and parafusos_especiais.get('parafusos_e'):
            print(f"[PARAFUSOS_PILAR2] 🔍 Usando parafusos especiais E para Pilar 2")
            parafusos_e = parafusos_especiais.get('parafusos_e', [])
            
            # Para Pilar 2, usar parafusos E (desconsiderando primeiro campo)
            # Os parafusos E já foram coletados a partir do segundo campo (par_e_2 a par_e_9)
            print(f"[PARAFUSOS_PILAR2] Parafusos E coletados para Pilar 2: {parafusos_e}")
            
            # NOTA: O robô usará automaticamente estes parafusos para Pilar 2
            # baseado na lógica interna de detecção do tipo de pilar
        else:
            print(f"[PARAFUSOS_PILAR2] Nenhum parafuso especial E encontrado para Pilar 2")

        # 5. CORREÇÃO: Preencher grades do Grupo 1 (Grade A) usando dados recebidos diretamente
        print("Preenchendo grades do Grupo 1 (Grade A)...")
        print(f"[DEBUG] Chegou na seção 5 - grade grupo 1")
        
        # Desabilitar eventos temporariamente para evitar recálculos automáticos
        print("  Desabilitando eventos automáticos...")
        gerador.grade1_entry.unbind("<KeyRelease>")
        gerador.grade1_entry.unbind("<FocusOut>")
        gerador.grade2_entry.unbind("<KeyRelease>")
        gerador.grade2_entry.unbind("<FocusOut>")
        gerador.grade3_entry.unbind("<KeyRelease>")
        gerador.grade3_entry.unbind("<FocusOut>")
        gerador.distancia1_entry.unbind("<KeyRelease>")
        gerador.distancia1_entry.unbind("<FocusOut>")
        gerador.distancia2_entry.unbind("<KeyRelease>")
        gerador.distancia2_entry.unbind("<FocusOut>")
            
        # CORREÇÃO: Usar dados recebidos diretamente em dados_pilar
        print("  Usando dados recebidos diretamente...")
        grades_grupo1 = dados_pilar.get('grades_grupo1', {})
        print(f"  Dados grades_grupo1 recebidos: {grades_grupo1}")
        
        if grades_grupo1:
            # Grade Grupo 1 - valores dos dados recebidos
            grade1_valor = grades_grupo1.get('grade_1', '')
            grade2_valor = grades_grupo1.get('grade_2', '')
            grade3_valor = grades_grupo1.get('grade_3', '')
            distancia1_valor = grades_grupo1.get('distancia_1', '')
            distancia2_valor = grades_grupo1.get('distancia_2', '')
            
            print(f"  Valores extraídos dos dados (Grupo 1):")
            print(f"    Grade 1: {grade1_valor}")
            print(f"    Distância 1: {distancia1_valor}")
            print(f"    Grade 2: {grade2_valor}")
            print(f"    Distância 2: {distancia2_valor}")
            print(f"    Grade 3: {grade3_valor}")
            
            # Preencher campos do robô com valores dos dados
            gerador.grade1_entry.delete(0, 'end')
            gerador.grade1_entry.insert(0, str(grade1_valor) if grade1_valor else "")
            print(f"  Preenchendo grade1_entry: {grade1_valor}")
            
            gerador.distancia1_entry.delete(0, 'end')
            gerador.distancia1_entry.insert(0, str(distancia1_valor) if distancia1_valor else "")
            print(f"  Preenchendo distancia1_entry: {distancia1_valor}")
            
            gerador.grade2_entry.delete(0, 'end')
            gerador.grade2_entry.insert(0, str(grade2_valor) if grade2_valor else "")
            print(f"  Preenchendo grade2_entry: {grade2_valor}")
            
            gerador.distancia2_entry.delete(0, 'end')
            gerador.distancia2_entry.insert(0, str(distancia2_valor) if distancia2_valor else "")
            print(f"  Preenchendo distancia2_entry: {distancia2_valor}")
            
            gerador.grade3_entry.delete(0, 'end')
            gerador.grade3_entry.insert(0, str(grade3_valor) if grade3_valor else "")
            print(f"  Preenchendo grade3_entry: {grade3_valor}")
        else:
            print("  Dados grades_grupo1 não encontrados!")

        # 6. Preencher detalhes das grades do Grupo 1
        print("Preenchendo detalhes das grades do Grupo 1...")
        print(f"[DEBUG] Chegou na seção 6 - detalhes grades grupo 1")

        # Verificar se há pilares especiais ativos
        pilar_especial = dados_pilar.get('pilar_especial', {})
        if pilar_especial and pilar_especial.get('ativar_pilar_especial'):
            print(f"[PILARES_ESPECIAIS] 🔍 Pilares especiais detectados, preservando detalhes especiais")
            print(f"[PILARES_ESPECIAIS] DEBUG: dados_pilar detalhes_grades_especiais = {dados_pilar.get('detalhes_grades_especiais', {})}")

            # Usar detalhes especiais diretamente de dados_pilar
            detalhes_especiais = dados_pilar.get('detalhes_grades_especiais', {})

            # Mapear detalhes especiais para o formato esperado pelo Grupo 1
            detalhes_grupo1 = {}
            for chave, valor in detalhes_especiais.items():
                if chave.startswith('detalhe_a_'):
                    # Mapear detalhe_a_1_1 -> detalhe_grade1_1, detalhe_a_1_2 -> detalhe_grade1_2, etc.
                    partes = chave.split('_')
                    if len(partes) >= 4 and partes[1] == 'a':
                        grade_num = partes[2]  # 1, 2, 3
                        detalhe_num = partes[3]  # 1, 2, 3
                        nova_chave = f'detalhe_grade{grade_num}_{detalhe_num}'
                        detalhes_grupo1[nova_chave] = valor
        else:
            # Usar detalhes normais se não houver pilares especiais
            detalhes_grupo1 = dados_pilar.get('detalhes_grades', {})
        
        # Mapear os campos de detalhes para as entradas correspondentes
        detalhes_grade1 = [
            ('detalhe_grade1_1', 0),
            ('detalhe_grade1_2', 1),
            ('detalhe_grade1_3', 2),
            ('detalhe_grade1_4', 3),
            ('detalhe_grade1_5', 4)
        ]
        
        detalhes_grade2 = [
            ('detalhe_grade2_1', 0),
            ('detalhe_grade2_2', 1),
            ('detalhe_grade2_3', 2),
            ('detalhe_grade2_4', 3),
            ('detalhe_grade2_5', 4)
        ]
        
        detalhes_grade3 = [
            ('detalhe_grade3_1', 0),
            ('detalhe_grade3_2', 1),
            ('detalhe_grade3_3', 2),
            ('detalhe_grade3_4', 3),
            ('detalhe_grade3_5', 4)
        ]

        # Preencher detalhes da Grade 1
        for campo, idx in detalhes_grade1:
            if idx < len(gerador.detalhe_grade1_entries):
                valor = detalhes_grupo1.get(campo, "")
                gerador.detalhe_grade1_entries[idx].delete(0, 'end')
                gerador.detalhe_grade1_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo}: {valor}")

        # Preencher detalhes da Grade 2
        for campo, idx in detalhes_grade2:
            if idx < len(gerador.detalhe_grade2_entries):
                valor = detalhes_grupo1.get(campo, "")
                gerador.detalhe_grade2_entries[idx].delete(0, 'end')
                gerador.detalhe_grade2_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo}: {valor}")

        # Preencher detalhes da Grade 3
        for campo, idx in detalhes_grade3:
            if idx < len(gerador.detalhe_grade3_entries):
                valor = detalhes_grupo1.get(campo, "")
                gerador.detalhe_grade3_entries[idx].delete(0, 'end')
                gerador.detalhe_grade3_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo}: {valor}")

        # 7. CORREÇÃO: Preencher Grade Grupo 2 (Grade B) usando dados recebidos
        print("Preenchendo Grade Grupo 2...")
        print(f"[DEBUG] Chegou na seção 7 - grade grupo 2")
        
        # Desabilitar eventos temporariamente para evitar recálculos automáticos
        print("  Desabilitando eventos automáticos do Grupo 2...")
        try:
            gerador.grade1_grupo2_entry.unbind("<KeyRelease>")
            gerador.grade1_grupo2_entry.unbind("<FocusOut>")
            gerador.grade2_grupo2_entry.unbind("<KeyRelease>")
            gerador.grade2_grupo2_entry.unbind("<FocusOut>")
            gerador.grade3_grupo2_entry.unbind("<KeyRelease>")
            gerador.grade3_grupo2_entry.unbind("<FocusOut>")
            gerador.distancia1_grupo2_entry.unbind("<KeyRelease>")
            gerador.distancia1_grupo2_entry.unbind("<FocusOut>")
            gerador.distancia2_grupo2_entry.unbind("<KeyRelease>")
            gerador.distancia2_grupo2_entry.unbind("<FocusOut>")
        except:
            pass
        
        # CORREÇÃO: Usar dados das grades_grupo2 recebidos em dados_pilar
        grades_grupo2 = dados_pilar.get('grades_grupo2', {})
        print(f"  Dados grades_grupo2 recebidos: {grades_grupo2}")
        
        if grades_grupo2:
            # Grade Grupo 2 - valores dos dados recebidos
            grade1_grupo2_valor = grades_grupo2.get('grade_1_grupo2', '')
            grade2_grupo2_valor = grades_grupo2.get('grade_2_grupo2', '')
            grade3_grupo2_valor = grades_grupo2.get('grade_3_grupo2', '')
            distancia1_grupo2_valor = grades_grupo2.get('distancia_1_grupo2', '')
            distancia2_grupo2_valor = grades_grupo2.get('distancia_2_grupo2', '')
            
            print(f"  Valores extraídos dos dados (Grupo 2 - ANTES DA INVERSÃO):")
            print(f"    Grade 1 Grupo 2: {grade1_grupo2_valor}")
            print(f"    Distância 1 Grupo 2: {distancia1_grupo2_valor}")
            print(f"    Grade 2 Grupo 2: {grade2_grupo2_valor}")
            print(f"    Distância 2 Grupo 2: {distancia2_grupo2_valor}")
            print(f"    Grade 3 Grupo 2: {grade3_grupo2_valor}")
            
            # ===== INVERSÃO PARA LADO B (PILARES COMUNS) =====
            # Inverter ordem das grades: Grade 1 <-> Grade 3, Grade 2 permanece
            # Inverter distâncias: Dist 1 <-> Dist 2
            print(f"\n  [INVERSÃO LADO B] Aplicando inversão para pilares comuns...")
            grade1_invertido = grade3_grupo2_valor  # Grade 1 recebe valor da Grade 3
            grade2_invertido = grade2_grupo2_valor  # Grade 2 permanece
            grade3_invertido = grade1_grupo2_valor  # Grade 3 recebe valor da Grade 1
            distancia1_invertido = distancia2_grupo2_valor  # Dist 1 recebe valor da Dist 2
            distancia2_invertido = distancia1_grupo2_valor  # Dist 2 recebe valor da Dist 1
            
            print(f"  [INVERSÃO LADO B] Valores após inversão:")
            print(f"    Grade 1 (era Grade 3): {grade1_invertido}")
            print(f"    Distância 1 (era Dist 2): {distancia1_invertido}")
            print(f"    Grade 2 (permanece): {grade2_invertido}")
            print(f"    Distância 2 (era Dist 1): {distancia2_invertido}")
            print(f"    Grade 3 (era Grade 1): {grade3_invertido}")
            
            # Preencher campos do robô com valores INVERTIDOS
            gerador.grade1_grupo2_entry.delete(0, 'end')
            gerador.grade1_grupo2_entry.insert(0, str(grade1_invertido) if grade1_invertido else "")
            print(f"  Preenchendo grade1_grupo2 (INVERTIDO): {grade1_invertido}")
            
            gerador.distancia1_grupo2_entry.delete(0, 'end')
            gerador.distancia1_grupo2_entry.insert(0, str(distancia1_invertido) if distancia1_invertido else "")
            print(f"  Preenchendo distancia1_grupo2 (INVERTIDO): {distancia1_invertido}")
            
            gerador.grade2_grupo2_entry.delete(0, 'end')
            gerador.grade2_grupo2_entry.insert(0, str(grade2_invertido) if grade2_invertido else "")
            print(f"  Preenchendo grade2_grupo2 (permanece): {grade2_invertido}")
            
            gerador.distancia2_grupo2_entry.delete(0, 'end')
            gerador.distancia2_grupo2_entry.insert(0, str(distancia2_invertido) if distancia2_invertido else "")
            print(f"  Preenchendo distancia2_grupo2 (INVERTIDO): {distancia2_invertido}")
            
            gerador.grade3_grupo2_entry.delete(0, 'end')
            gerador.grade3_grupo2_entry.insert(0, str(grade3_invertido) if grade3_invertido else "")
            print(f"  Preenchendo grade3_grupo2 (INVERTIDO): {grade3_invertido}")
                
        else:
            print("  Dados grades_grupo2 não encontrados!")

        # 8. Preencher detalhes das grades do Grupo 2
        print("Preenchendo detalhes das grades do Grupo 2...")
        print(f"[DEBUG] Chegou na seção 8 - detalhes grades grupo 2")

        # Verificar se há pilares especiais ativos para Grupo 2
        if pilar_especial and pilar_especial.get('ativar_pilar_especial'):
            print(f"[PILARES_ESPECIAIS] 🔍 Pilares especiais detectados, preservando detalhes especiais do Grupo 2")

            # Usar detalhes especiais diretamente de dados_pilar
            detalhes_especiais = dados_pilar.get('detalhes_grades_especiais', {})

            # Mapear detalhes especiais para o formato esperado pelo Grupo 2
            detalhes_grupo2 = {}
            for chave, valor in detalhes_especiais.items():
                if chave.startswith('detalhe_b_'):
                    # Mapear detalhe_b_1_1 -> detalhe_grade1_1_grupo2, detalhe_b_1_2 -> detalhe_grade1_2_grupo2, etc.
                    partes = chave.split('_')
                    if len(partes) >= 4 and partes[1] == 'b':
                        grade_num = partes[2]  # 1, 2, 3
                        detalhe_num = partes[3]  # 1, 2, 3
                        nova_chave = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                        detalhes_grupo2[nova_chave] = valor
        else:
            # Usar detalhes normais se não houver pilares especiais
            detalhes_grupo2 = dados_pilar.get('detalhes_grades_grupo2', {})
        
        detalhes_grade1_grupo2 = [
            ('detalhe_grade1_1_grupo2', 0),
            ('detalhe_grade1_2_grupo2', 1),
            ('detalhe_grade1_3_grupo2', 2),
            ('detalhe_grade1_4_grupo2', 3),
            ('detalhe_grade1_5_grupo2', 4)
        ]
        
        detalhes_grade2_grupo2 = [
            ('detalhe_grade2_1_grupo2', 0),
            ('detalhe_grade2_2_grupo2', 1),
            ('detalhe_grade2_3_grupo2', 2),
            ('detalhe_grade2_4_grupo2', 3),
            ('detalhe_grade2_5_grupo2', 4)
        ]
        
        detalhes_grade3_grupo2 = [
            ('detalhe_grade3_1_grupo2', 0),
            ('detalhe_grade3_2_grupo2', 1),
            ('detalhe_grade3_3_grupo2', 2),
            ('detalhe_grade3_4_grupo2', 3),
            ('detalhe_grade3_5_grupo2', 4)
        ]

        # ===== INVERSÃO DOS DETALHES PARA LADO B (PILARES COMUNS) =====
        # Grade 1 usa detalhes da Grade 3 (invertidos)
        # Grade 2 permanece (mas detalhes invertidos)
        # Grade 3 usa detalhes da Grade 1 (invertidos)
        # Dentro de cada grade: Detalhe 1 <-> Detalhe 4, Detalhe 2 <-> Detalhe 3
        # Detalhe 0 e Detalhe 5 permanecem
        
        print(f"\n  [INVERSÃO DETALHES LADO B] Aplicando inversão dos detalhes...")
        
        # Mapeamento de inversão dos detalhes dentro de cada grade
        # Detalhe 0 (idx 0) -> Detalhe 0 (idx 0) - permanece
        # Detalhe 1 (idx 1) -> Detalhe 4 (idx 3) - INVERTIDO
        # Detalhe 2 (idx 2) -> Detalhe 3 (idx 2) - INVERTIDO
        # Detalhe 3 (idx 3) -> Detalhe 2 (idx 1) - INVERTIDO
        # Detalhe 4 (idx 4) -> Detalhe 1 (idx 0) - INVERTIDO
        # Detalhe 5 (idx 5) -> Detalhe 5 (idx 4) - permanece (mas não existe idx 5, então não há)
        
        # Função auxiliar para inverter detalhes dentro de uma grade
        def inverter_detalhes_grade(detalhes_originais):
            """
            Inverte detalhes: D1<->D4, D2<->D3, D5 permanece
            Índices: idx 0=D1, idx 1=D2, idx 2=D3, idx 3=D4, idx 4=D5
            """
            if len(detalhes_originais) < 5:
                return detalhes_originais
            detalhes_invertidos = [None] * 5
            detalhes_invertidos[0] = detalhes_originais[3]  # D1 (idx 0) recebe D4 (idx 3)
            detalhes_invertidos[1] = detalhes_originais[2]  # D2 (idx 1) recebe D3 (idx 2)
            detalhes_invertidos[2] = detalhes_originais[1]  # D3 (idx 2) recebe D2 (idx 1)
            detalhes_invertidos[3] = detalhes_originais[0]  # D4 (idx 3) recebe D1 (idx 0)
            detalhes_invertidos[4] = detalhes_originais[4]  # D5 (idx 4) permanece
            return detalhes_invertidos
        
        # Coletar detalhes originais de cada grade
        detalhes_g1_orig = []
        detalhes_g2_orig = []
        detalhes_g3_orig = []
        
        for campo, idx in detalhes_grade1_grupo2:
            valor = detalhes_grupo2.get(campo, "")
            detalhes_g1_orig.append(valor)
        
        for campo, idx in detalhes_grade2_grupo2:
            valor = detalhes_grupo2.get(campo, "")
            detalhes_g2_orig.append(valor)
        
        for campo, idx in detalhes_grade3_grupo2:
            valor = detalhes_grupo2.get(campo, "")
            detalhes_g3_orig.append(valor)
        
        print(f"  [INVERSÃO DETALHES] Detalhes originais:")
        print(f"    Grade 1: {detalhes_g1_orig}")
        print(f"    Grade 2: {detalhes_g2_orig}")
        print(f"    Grade 3: {detalhes_g3_orig}")
        
        # Inverter detalhes dentro de cada grade
        detalhes_g1_inv = inverter_detalhes_grade(detalhes_g1_orig)
        detalhes_g2_inv = inverter_detalhes_grade(detalhes_g2_orig)
        detalhes_g3_inv = inverter_detalhes_grade(detalhes_g3_orig)
        
        # Grade 1 usa detalhes da Grade 3 (invertidos)
        # Grade 2 permanece (mas detalhes invertidos)
        # Grade 3 usa detalhes da Grade 1 (invertidos)
        detalhes_g1_final = detalhes_g3_inv  # Grade 1 recebe Grade 3 invertida
        detalhes_g2_final = detalhes_g2_inv  # Grade 2 permanece (mas invertida)
        detalhes_g3_final = detalhes_g1_inv  # Grade 3 recebe Grade 1 invertida
        
        print(f"  [INVERSÃO DETALHES] Detalhes após inversão:")
        print(f"    Grade 1 (era Grade 3): {detalhes_g1_final}")
        print(f"    Grade 2 (permanece): {detalhes_g2_final}")
        print(f"    Grade 3 (era Grade 1): {detalhes_g3_final}")
        
        # Preencher detalhes da Grade 1 Grupo 2 (usando Grade 3 invertida)
        for campo, idx in detalhes_grade1_grupo2:
            if idx < len(gerador.detalhe_grade1_grupo2_entries) and idx < len(detalhes_g1_final):
                valor = detalhes_g1_final[idx]
                gerador.detalhe_grade1_grupo2_entries[idx].delete(0, 'end')
                gerador.detalhe_grade1_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo} (INVERTIDO - era Grade 3): {valor}")

        # Preencher detalhes da Grade 2 Grupo 2 (permanece, mas invertida)
        for campo, idx in detalhes_grade2_grupo2:
            if idx < len(gerador.detalhe_grade2_grupo2_entries) and idx < len(detalhes_g2_final):
                valor = detalhes_g2_final[idx]
                gerador.detalhe_grade2_grupo2_entries[idx].delete(0, 'end')
                gerador.detalhe_grade2_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo} (INVERTIDO): {valor}")

        # Preencher detalhes da Grade 3 Grupo 2 (usando Grade 1 invertida)
        for campo, idx in detalhes_grade3_grupo2:
            if idx < len(gerador.detalhe_grade3_grupo2_entries) and idx < len(detalhes_g3_final):
                valor = detalhes_g3_final[idx]
                gerador.detalhe_grade3_grupo2_entries[idx].delete(0, 'end')
                gerador.detalhe_grade3_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                print(f"  Preenchendo {campo} (INVERTIDO - era Grade 1): {valor}")

        # 9. NÃO recalcular - manter os valores preenchidos
        print("Mantendo valores preenchidos sem recalcular...")
        print(f"[DEBUG] Chegou na seção 9 - antes dos pilares especiais")
        
        # DEBUG: Print dados recebidos na função
        print(f"\n>>> DADOS RECEBIDOS NA FUNCAO:")
        print(f">>> dados_pilar completo: {dados_pilar}")
        print(f">>> grades_grupo1 recebido: {dados_pilar.get('grades_grupo1', 'NÃO ENCONTRADO')}")
        print(f">>> grades_grupo2 recebido: {dados_pilar.get('grades_grupo2', 'NÃO ENCONTRADO')}")
        
        # DEBUG: Print grades being received
        print(f"\n>>> GRADES RECEBIDAS:")
        print(f">>> Grade Grupo 1 (Grade A) - Valores inseridos:")
        print(f"    grade1_entry: {gerador.grade1_entry.get()}")
        print(f"    distancia1_entry: {gerador.distancia1_entry.get()}")
        print(f"    grade2_entry: {gerador.grade2_entry.get()}")
        print(f"    distancia2_entry: {gerador.distancia2_entry.get()}")
        print(f"    grade3_entry: {gerador.grade3_entry.get()}")
        
        print(f">>> Grade Grupo 2 (Grade B) - Valores inseridos:")
        print(f"    grade1_grupo2_entry: {gerador.grade1_grupo2_entry.get()}")
        print(f"    distancia1_grupo2_entry: {gerador.distancia1_grupo2_entry.get()}")
        print(f"    grade2_grupo2_entry: {gerador.grade2_grupo2_entry.get()}")
        print(f"    distancia2_grupo2_entry: {gerador.distancia2_grupo2_entry.get()}")
        print(f"    grade3_grupo2_entry: {gerador.grade3_grupo2_entry.get()}")
        
        print(f">>> Detalhes Grade Grupo 1 - Primeiros valores:")
        if hasattr(gerador, 'detalhe_grade1_entries') and len(gerador.detalhe_grade1_entries) > 0:
            print(f"    detalhe_grade1_1: {gerador.detalhe_grade1_entries[0].get()}")
        if hasattr(gerador, 'detalhe_grade2_entries') and len(gerador.detalhe_grade2_entries) > 0:
            print(f"    detalhe_grade2_1: {gerador.detalhe_grade2_entries[0].get()}")
        if hasattr(gerador, 'detalhe_grade3_entries') and len(gerador.detalhe_grade3_entries) > 0:
            print(f"    detalhe_grade3_1: {gerador.detalhe_grade3_entries[0].get()}")
            
        print(f">>> Detalhes Grade Grupo 2 - Primeiros valores:")
        if hasattr(gerador, 'detalhe_grade1_grupo2_entries') and len(gerador.detalhe_grade1_grupo2_entries) > 0:
            print(f"    detalhe_grade1_1_grupo2: {gerador.detalhe_grade1_grupo2_entries[0].get()}")
        if hasattr(gerador, 'detalhe_grade2_grupo2_entries') and len(gerador.detalhe_grade2_grupo2_entries) > 0:
            print(f"    detalhe_grade2_1_grupo2: {gerador.detalhe_grade2_grupo2_entries[0].get()}")
        if hasattr(gerador, 'detalhe_grade3_grupo2_entries') and len(gerador.detalhe_grade3_grupo2_entries) > 0:
            print(f"    detalhe_grade3_1_grupo2: {gerador.detalhe_grade3_grupo2_entries[0].get()}")
        
        print(f">>> FIM GRADES RECEBIDAS\n")

        # 10. Verificar e ativar "pilar rotacionado" se "pilar especial" estiver ativo
        print("Verificando status do pilar especial...")
        try:
            if hasattr(gerador, 'pilar_rotacionado_var'):
                status_atual = gerador.pilar_rotacionado_var.get()
                print(f"  Status atual do checkbox: {status_atual}")
                
                # Verificar se existe uma instância da interface principal com pilar especial ativo
                pilar_especial_ativo = False
                
                # Buscar globalmente por instâncias com pilar especial
                import gc
                print("  Buscando globalmente por instancias com pilar especial...")
                objetos_encontrados = 0
                for obj in gc.get_objects():
                    try:
                        if hasattr(obj, 'ativar_pilar_especial'):
                            objetos_encontrados += 1
                            attr = getattr(obj, 'ativar_pilar_especial', None)
                            if attr and hasattr(attr, 'get'):
                                valor = attr.get()
                                print(f"    Objeto {objetos_encontrados}: pilar_especial = {valor}")
                                if valor:
                                    pilar_especial_ativo = True
                                    break
                    except Exception:
                        continue
                print(f"  Total de objetos com ativar_pilar_especial encontrados: {objetos_encontrados}")
                
                # Se pilar especial estiver ativo, ativar o pilar rotacionado
                if pilar_especial_ativo:
                    print("PILAR ESPECIAL ATIVO - Ativando 'Pilar Rotacionado' automaticamente")
                    gerador.pilar_rotacionado_var.set(True)
                    novo_status = gerador.pilar_rotacionado_var.get()
                    print(f"  Checkbox ativado diretamente. Novo status: {novo_status}")
                else:
                    print("Pilar especial nao esta ativo - mantendo configuracao atual")
            else:
                print("  Gerador nao tem pilar_rotacionado_var")
                
        except Exception as e:
            print("Erro ao verificar pilar especial: " + str(e))

        # 11. VERIFICAR E PROCESSAR PILARES ESPECIAIS DIRETAMENTE
        print("\n[DEBUG] Verificando se é pilar especial...")
        print(f"[DEBUG] Chegou na seção de pilares especiais")
        
        # Log detalhado dos dados recebidos
        print(f"[PILARES_ESPECIAIS] === DIAGNÓSTICO DOS DADOS RECEBIDOS ===")
        print(f"[PILARES_ESPECIAIS] dados_pilar keys: {list(dados_pilar.keys())}")
        print(f"[PILARES_ESPECIAIS] dados_pilar['ativar_pilar_especial']: {dados_pilar.get('ativar_pilar_especial')}")
        print(f"[PILARES_ESPECIAIS] dados_pilar['tipo_pilar_especial']: {dados_pilar.get('tipo_pilar_especial')}")
        print(f"[PILARES_ESPECIAIS] dados_pilar['comp_1']: '{dados_pilar.get('comp_1')}', comp_2: '{dados_pilar.get('comp_2')}', comp_3: '{dados_pilar.get('comp_3')}'")
        print(f"[PILARES_ESPECIAIS] dados_pilar['larg_1']: '{dados_pilar.get('larg_1')}', larg_2: '{dados_pilar.get('larg_2')}', larg_3: '{dados_pilar.get('larg_3')}'")
        
        # Verificar se é pilar especial usando dados_pilar
        # Suporte para estruturas aninhadas (dados_pilar['pilar_especial']) e diretas
        pilar_especial_data = dados_pilar.get('pilar_especial', {})
        print(f"[PILARES_ESPECIAIS] pilar_especial_data: {pilar_especial_data}")
        print(f"[PILARES_ESPECIAIS] Tipo de pilar_especial_data: {type(pilar_especial_data)}")
        if isinstance(pilar_especial_data, dict):
            print(f"[PILARES_ESPECIAIS] Chaves em pilar_especial_data: {list(pilar_especial_data.keys())}")
        
        pilar_especial_ativo = (
            dados_pilar.get('ativar_pilar_especial', False) or
            (pilar_especial_data.get('ativar_pilar_especial', False) if isinstance(pilar_especial_data, dict) else False)
        )
        print(f"[PILARES_ESPECIAIS] === FIM DO DIAGNÓSTICO ===")

        dados_pilar_especial = {}
        globais_pilar_especial = {}
        dimensoes_especiais = []

        print(f"[PILARES_ESPECIAIS] Pilar especial ativo: {pilar_especial_ativo}")
        print(f"[PILARES_ESPECIAIS] Tipo do valor: {type(pilar_especial_ativo)}")
        print(f"[PILARES_ESPECIAIS] Dados pilar_especial: {pilar_especial_data}")
        print(f"[PILARES_ESPECIAIS] Verificação booleana: {bool(pilar_especial_ativo)}")

        if pilar_especial_ativo:
            # Coletar dados do pilar especial dos dados_pilar
            dados_pilar_especial = {
                'ativo': pilar_especial_ativo,
                'tipo': (
                    dados_pilar.get('tipo_pilar_especial', '') or
                    pilar_especial_data.get('tipo_pilar_especial', '')
                ),
                'comp_1': (
                    dados_pilar.get('comp_1', '') or
                    pilar_especial_data.get('comp_1', '')
                ),
                'comp_2': (
                    dados_pilar.get('comp_2', '') or
                    pilar_especial_data.get('comp_2', '')
                ),
                'comp_3': (
                    dados_pilar.get('comp_3', '') or
                    pilar_especial_data.get('comp_3', '')
                ),
                'larg_1': (
                    dados_pilar.get('larg_1', '') or
                    pilar_especial_data.get('larg_1', '')
                ),
                'larg_2': (
                    dados_pilar.get('larg_2', '') or
                    pilar_especial_data.get('larg_2', '')
                ),
                'larg_3': (
                    dados_pilar.get('larg_3', '') or
                    pilar_especial_data.get('larg_3', '')
                ),
            }
            print(f"[PILARES_ESPECIAIS] Dados coletados: {dados_pilar_especial}")
            
            # Calcular globais para pilar especial tipo L
            if dados_pilar_especial.get('tipo') == 'L':
                    print("[PILARES_ESPECIAIS] Calculando globais para pilar tipo L...")
                    
                    # Obter detalhes das grades especiais
                    detalhes_grades_especiais = dados_pilar.get('detalhes_grades_especiais', {})
                    print(f"[PILARES_ESPECIAIS] Detalhes de grades especiais encontrados: {len(detalhes_grades_especiais)} valores")
                    
                    globais_pilar_especial = calcular_globais_pilar_especial_L(
                        dados_pilar_especial.get('comp_1', ''),
                        dados_pilar_especial.get('comp_2', ''),
                        dados_pilar_especial.get('larg_1', ''),
                        dados_pilar_especial.get('larg_2', ''),
                        detalhes_grades_especiais
                    )
                    
                    if globais_pilar_especial:
                        print(f"[PILARES_ESPECIAIS] Globais calculadas: {len(globais_pilar_especial)} valores")
                        
                        # Definir globais no gerador
                        gerador._globais_pilar_especial = globais_pilar_especial
                        print(f"[PILARES_ESPECIAIS] Globais definidas no gerador: {len(gerador._globais_pilar_especial)} valores")
            
            # CORREÇÃO CRÍTICA: Coletar dimensões válidas para múltiplos scripts
            # Esta parte deve estar FORA do bloco if tipo == 'L', para funcionar com TODOS os tipos de pilar especial
            dimensoes_dados = [
                (dados_pilar_especial.get('comp_1'), dados_pilar_especial.get('larg_1'), 1),
                (dados_pilar_especial.get('comp_2'), dados_pilar_especial.get('larg_2'), 2),
                (dados_pilar_especial.get('comp_3'), dados_pilar_especial.get('larg_3'), 3)
            ]
            
            print(f"[PILARES_ESPECIAIS] Verificando dimensões: {dimensoes_dados}")
            
            for comp_val, larg_val, indice in dimensoes_dados:
                comp_str = str(comp_val).strip() if comp_val else ''
                larg_str = str(larg_val).strip() if larg_val else ''
                
                print(f"[PILARES_ESPECIAIS] Dimensão {indice}: comp_val='{comp_val}' => comp_str='{comp_str}', larg_val='{larg_val}' => larg_str='{larg_str}'")
                
                if comp_str and larg_str and comp_str != '0' and larg_str != '0':
                    dimensoes_especiais.append({
                        'comprimento': comp_str,
                        'largura': larg_str,
                        'indice': indice
                    })
                    print(f"[PILARES_ESPECIAIS] ✅ Dimensão {indice} VÁLIDA adicionada: comp={comp_str}, larg={larg_str}")
                else:
                    print(f"[PILARES_ESPECIAIS] ⚠️ Dimensão {indice} INVÁLIDA ignorada (comp_str='{comp_str}', larg_str='{larg_str}')")
            
            print(f"[PILARES_ESPECIAIS] Total de dimensões válidas coletadas: {len(dimensoes_especiais)}")
        
        # 12. GERAR SCRIPTS (ÚNICO OU MÚLTIPLOS)
        print(f"[DEBUG] Verificando condições: pilar_especial_ativo={pilar_especial_ativo}, len(dimensoes_especiais)={len(dimensoes_especiais)}")
        print(f"[DEBUG] dimensoes_especiais: {dimensoes_especiais}")
        
        if pilar_especial_ativo and dimensoes_especiais:
            print(f"[PILARES_ESPECIAIS] Gerando {len(dimensoes_especiais)} scripts para pilar especial...")
            
            # Salvar dimensões originais
            comp_original = gerador.comprimento_pilar_entry.get()
            larg_original = gerador.largura_pilar_entry.get()
            
            # Definir diretórios uma vez para todos os scripts
            nome_pilar = dados_pilar.get('nome', 'PILAR_SEM_NOME')
            pavimento = dados_pilar.get('pavimento', 'SEM_PAVIMENTO')
            
            # CORREÇÃO: Usar normalização para remover acentos (mesma função usada em _dcad_pavimento_cima)
            import unicodedata
            def normalizar_nome_pasta_especial(texto):
                """Normaliza o nome removendo acentos"""
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                texto_limpo = texto_limpo.strip('_')
                return texto_limpo
            
            pavimento_normalizado = normalizar_nome_pasta_especial(pavimento)
            # CORREÇÃO: Usar robust_path_resolver para path correto em frozen e dev
            try:
                from utils.robust_path_resolver import robust_path_resolver
                diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
            except ImportError:
                try:
                    from src.utils.robust_path_resolver import robust_path_resolver
                    diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
                except ImportError:
                    # Fallback manual (não recomendado, mas necessário em alguns casos)
                    diretorio_base = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "output", "scripts")
            nome_pasta_pavimento = f"{pavimento_normalizado}_CIMA"
            diretorio_pavimento = os.path.join(diretorio_base, nome_pasta_pavimento)
            os.makedirs(diretorio_pavimento, exist_ok=True)
            
            print(f"[PILARES_ESPECIAIS] Pavimento original: '{pavimento}'")
            print(f"[PILARES_ESPECIAIS] Pavimento normalizado: '{pavimento_normalizado}'")
            print(f"[PILARES_ESPECIAIS] Diretório: {diretorio_pavimento}")
            
            for idx, dimensao in enumerate(dimensoes_especiais):
                print(f"[PILARES_ESPECIAIS] Gerando script {idx + 1}/{len(dimensoes_especiais)}...")
                
                # Definir configurações para cada script
                if idx == 0:
                    # Primeiro script: ROTATE ativo
                    gerador.pilar_rotacionado_var.set(True)
                    gerador.usar_scale_var.set(False)
                    gerador.pilar_especial_ativo_var.set(False)
                    print(f"[PILARES_ESPECIAIS] Configuração PILAR 1: ROTATE=True, SCALE=False, ESPECIAL=False")
                elif idx == len(dimensoes_especiais) - 1:
                    # Último script: SCALE ativo com valores especiais
                    gerador.pilar_rotacionado_var.set(False)
                    gerador.usar_scale_var.set(True)
                    gerador.pilar_especial_ativo_var.set(True)
                    print(f"[PILARES_ESPECIAIS] Configuração PILAR 2: ROTATE=False, SCALE=True, ESPECIAL=True")
                else:
                    # Scripts intermediários: sem ROTATE nem SCALE
                    gerador.pilar_rotacionado_var.set(False)
                    gerador.usar_scale_var.set(False)
                    gerador.pilar_especial_ativo_var.set(False)
                    print(f"[PILARES_ESPECIAIS] Configuração PILAR {idx + 1}: ROTATE=False, SCALE=False, ESPECIAL=False")
                
                # Alterar dimensões temporariamente
                gerador.comprimento_pilar_entry.delete(0, 'end')
                gerador.comprimento_pilar_entry.insert(0, dimensao['comprimento'])
                gerador.largura_pilar_entry.delete(0, 'end')
                gerador.largura_pilar_entry.insert(0, dimensao['largura'])
                
                print(f"[PILARES_ESPECIAIS] Usando dimensões: comp={dimensao['comprimento']}, larg={dimensao['largura']}")
                
                # Verificar se globais estão definidas antes da geração
                if globais_pilar_especial and len(globais_pilar_especial) > 0:
                    gerador._globais_pilar_especial = globais_pilar_especial
                    print(f"[PILARES_ESPECIAIS] Globais redefinidas no gerador: {len(gerador._globais_pilar_especial)} valores")
                
                # Adicionar dados dos parafusos especiais para o gerador
                if dados_pilar.get('pilar_especial', {}).get('parafusos_especiais'):
                    parafusos_especiais_data = dados_pilar['pilar_especial']['parafusos_especiais']
                    gerador._dados_parafusos_especiais = parafusos_especiais_data
                    print(f"[PILARES_ESPECIAIS] Dados de parafusos especiais definidos no gerador: {len(parafusos_especiais_data)} grupos")
                
                # Definir pilar_atual baseado no índice (para pilares especiais)
                gerador.pilar_atual = idx + 1
                print(f"[PILARES_ESPECIAIS] Definindo pilar_atual = {gerador.pilar_atual}")
                
                # ===== APLICAR PARAFUSOS ESPECIAIS PARA CADA SCRIPT =====
                if (dados_pilar.get('ativar_pilar_especial', False) and 
                    dados_pilar.get('tipo_pilar_especial', '') == 'L'):
                    
                    # CORREÇÃO: Buscar parafusos_especiais tanto no nível raiz quanto dentro de pilar_especial
                    parafusos_especiais_pilar = dados_pilar.get('pilar_especial', {}).get('parafusos_especiais', {})
                    parafusos_especiais_raiz = dados_pilar.get('parafusos_especiais', {})
                    # Usar o que estiver disponível (prioridade para pilar_especial)
                    parafusos_especiais_data = parafusos_especiais_pilar if parafusos_especiais_pilar else parafusos_especiais_raiz
                    
                    print(f"[PILARES_ESPECIAIS] Parafusos especiais (pilar_especial): {parafusos_especiais_pilar}")
                    print(f"[PILARES_ESPECIAIS] Parafusos especiais (raiz): {parafusos_especiais_raiz}")
                    print(f"[PILARES_ESPECIAIS] Parafusos especiais (usado): {parafusos_especiais_data}")
                    
                    if gerador.pilar_atual == 1:
                        # PILAR 1: usar parafusos A
                        # Verificar formato: lista (novo) ou dict (antigo)
                        if 'parafusos_a' in parafusos_especiais_data:
                            # Formato novo: {'parafusos_a': [lista]}
                            parafusos_a_lista = parafusos_especiais_data.get('parafusos_a', [])
                            print(f"[PILARES_ESPECIAIS] Aplicando parafusos A para Pilar 1 (formato lista): {parafusos_a_lista}")
                            
                            # Preencher campos de parafusos com valores especiais A
                            for i, valor in enumerate(parafusos_a_lista):
                                if i < len(gerador.parafuso_entries) and valor and float(valor) > 0:
                                    gerador.parafuso_entries[i].delete(0, 'end')
                                    gerador.parafuso_entries[i].insert(0, str(int(float(valor))))
                                    print(f"[PILARES_ESPECIAIS] Parafuso {i+1}: {valor}")
                        else:
                            # Formato antigo: {'par_a': {dict}}
                            parafusos_a = parafusos_especiais_data.get('par_a', {})
                            print(f"[PILARES_ESPECIAIS] Aplicando parafusos A para Pilar 1 (formato dict): {parafusos_a}")
                            
                            # Preencher campos de parafusos com valores especiais A
                            for i in range(1, 10):
                                campo_name = f"par_a_{i}"
                                if campo_name in parafusos_a and i-1 < len(gerador.parafuso_entries):
                                    valor = parafusos_a[campo_name]
                                    if valor and str(valor).strip():
                                        gerador.parafuso_entries[i-1].delete(0, 'end')
                                        gerador.parafuso_entries[i-1].insert(0, str(valor))
                                        print(f"[PILARES_ESPECIAIS] Parafuso {i}: {valor}")
                                    
                    elif gerador.pilar_atual == 2:
                        # PILAR 2: usar parafusos E (ignorar primeiro campo)
                        # Verificar formato: lista (novo) ou dict (antigo)
                        if 'parafusos_e' in parafusos_especiais_data:
                            # Formato novo: {'parafusos_e': [lista]} - lista já vem sem o primeiro campo (par_e_2 a par_e_9)
                            parafusos_e_lista = parafusos_especiais_data.get('parafusos_e', [])
                            print(f"[PILARES_ESPECIAIS] Aplicando parafusos E para Pilar 2 (formato lista): {parafusos_e_lista}")
                            
                            # Preencher campos de parafusos com valores especiais E (a lista já vem sem o primeiro campo)
                            for i, valor in enumerate(parafusos_e_lista):
                                if i < len(gerador.parafuso_entries) and valor and float(valor) > 0:
                                    gerador.parafuso_entries[i].delete(0, 'end')
                                    gerador.parafuso_entries[i].insert(0, str(int(float(valor))))
                                    print(f"[PILARES_ESPECIAIS] Parafuso E {i+1} (corresponde a par_e_{i+2}): {valor}")
                        else:
                            # Formato antigo: {'par_e': {dict}}
                            parafusos_e = parafusos_especiais_data.get('par_e', {})
                            print(f"[PILARES_ESPECIAIS] Aplicando parafusos E para Pilar 2 (formato dict): {parafusos_e}")
                            
                            # Preencher campos de parafusos com valores especiais E (a partir do segundo campo)
                            for i in range(2, 10):  # par_e_2 até par_e_9
                                campo_name = f"par_e_{i}"
                                if campo_name in parafusos_e and i-2 < len(gerador.parafuso_entries):
                                    valor = parafusos_e[campo_name]
                                    if valor and str(valor).strip():
                                        gerador.parafuso_entries[i-2].delete(0, 'end')
                                        gerador.parafuso_entries[i-2].insert(0, str(valor))
                                        print(f"[PILARES_ESPECIAIS] Parafuso {i-1}: {valor}")
                
                # ===== NOVA FUNCIONALIDADE: PREENCHIMENTO AUTOMÁTICO DE GRADES PARA PILARES ESPECIAIS TIPO "L" =====
                # Aplicar grades específicas baseadas no pilar_atual (1 = Grade A/B, 2 = Grade E/F)
                if (dados_pilar.get('ativar_pilar_especial', False) and 
                    dados_pilar.get('tipo_pilar_especial', '') == 'L'):
                    
                    print(f"[PILARES_ESPECIAIS] 🎯 PILAR ESPECIAL TIPO 'L' DETECTADO - PREENCHENDO GRADES COM VALORES EXATOS DA ABA")
                    
                    try:
                        # Obter dados EXATOS dos campos da aba Pilares Especiais
                        dados_especiais = dados_pilar.get('pilar_especial', {})
                        grades_especiais = dados_especiais.get('grades', {})
                        detalhes_especiais = dados_especiais.get('detalhes_grades_especiais', {})
                        
                        print(f"[PILARES_ESPECIAIS] Grades especiais encontradas: {len(grades_especiais)} valores")
                        print(f"[PILARES_ESPECIAIS] Detalhes especiais encontrados: {len(detalhes_especiais)} valores")
                        
                        # DETECTAR QUAL PILAR ESTÁ SENDO PROCESSADO
                        pilar_atual = gerador.pilar_atual
                        nome_pilar = dados_pilar.get('nome', '').upper()
                        
                        print(f"[PILARES_ESPECIAIS] 🔍 DEBUG DETALHADO - Pilar Especial:")
                        print(f"[PILARES_ESPECIAIS]   Nome original: '{dados_pilar.get('nome', '')}'")
                        print(f"[PILARES_ESPECIAIS]   Nome em maiúsculo: '{nome_pilar}'")
                        print(f"[PILARES_ESPECIAIS]   Pilar atual: {pilar_atual}")
                        
                        if grades_especiais:
                            if pilar_atual == 1:
                                # PILAR 1: Conjunto 1 = Grade A, Conjunto 2 = Grade B
                                print(f"[PILARES_ESPECIAIS] 🎯 PILAR 1 DETECTADO - Usando Grade A e Grade B")
                                print(f"[PILARES_ESPECIAIS]   Pilar atual: {pilar_atual}")
                                print(f"[PILARES_ESPECIAIS]   Condição executada: pilar_atual == 1 = {pilar_atual == 1}")
                                
                                # Grade A (Conjunto 1) - valores exatos dos campos grade_a_1, grade_a_2, grade_a_3, dist_a_1, dist_a_2
                                grade_a = {
                                    'grade_1': grades_especiais.get('grade_a_1', ''),
                                    'grade_2': grades_especiais.get('grade_a_2', ''),
                                    'grade_3': grades_especiais.get('grade_a_3', ''),
                                    'distancia_1': grades_especiais.get('dist_a_1', ''),
                                    'distancia_2': grades_especiais.get('dist_a_2', '')
                                }
                                
                                # Grade B (Conjunto 2) - valores exatos dos campos grade_b_1, grade_b_2, grade_b_3, dist_b_1, dist_b_2
                                grade_b = {
                                    'grade_1_grupo2': grades_especiais.get('grade_b_1', ''),
                                    'grade_2_grupo2': grades_especiais.get('grade_b_2', ''),
                                    'grade_3_grupo2': grades_especiais.get('grade_b_3', ''),
                                    'distancia_1_grupo2': grades_especiais.get('dist_b_1', ''),
                                    'distancia_2_grupo2': grades_especiais.get('dist_b_2', '')
                                }
                                
                                # Detalhes Grade A (Conjunto 1) - valores exatos dos campos detalhe_a_X_Y
                                detalhes_grade_a = {}
                                for grade_num in range(1, 4):
                                    for detalhe_num in range(1, 6):  # Corrigido: 5 detalhes (1-5)
                                        chave_original = f'detalhe_a_{grade_num}_{detalhe_num}'
                                        chave_nova = f'detalhe_grade{grade_num}_{detalhe_num}'
                                        detalhes_grade_a[chave_nova] = detalhes_especiais.get(chave_original, '')
                                
                                # Detalhes Grade B (Conjunto 2) - valores exatos dos campos detalhe_b_X_Y
                                detalhes_grade_b = {}
                                for grade_num in range(1, 4):
                                    for detalhe_num in range(1, 6):  # Corrigido: 5 detalhes (1-5)
                                        chave_original = f'detalhe_b_{grade_num}_{detalhe_num}'
                                        chave_nova = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                                        detalhes_grade_b[chave_nova] = detalhes_especiais.get(chave_original, '')
                                
                                print(f"[PILARES_ESPECIAIS] Grade A (Pilar 1 - Conjunto 1) - VALORES EXATOS DA ABA:")
                                print(f"[PILARES_ESPECIAIS]   grade_1={grade_a['grade_1']} (campo grade_a_1)")
                                print(f"[PILARES_ESPECIAIS]   grade_2={grade_a['grade_2']} (campo grade_a_2)")
                                print(f"[PILARES_ESPECIAIS]   grade_3={grade_a['grade_3']} (campo grade_a_3)")
                                print(f"[PILARES_ESPECIAIS]   dist_1={grade_a['distancia_1']} (campo dist_a_1)")
                                print(f"[PILARES_ESPECIAIS]   dist_2={grade_a['distancia_2']} (campo dist_a_2)")
                                
                                print(f"[PILARES_ESPECIAIS] Grade B (Pilar 1 - Conjunto 2) - VALORES EXATOS DA ABA:")
                                print(f"[PILARES_ESPECIAIS]   grade_1_grupo2={grade_b['grade_1_grupo2']} (campo grade_b_1)")
                                print(f"[PILARES_ESPECIAIS]   grade_2_grupo2={grade_b['grade_2_grupo2']} (campo grade_b_2)")
                                print(f"[PILARES_ESPECIAIS]   grade_3_grupo2={grade_b['grade_3_grupo2']} (campo grade_b_3)")
                                print(f"[PILARES_ESPECIAIS]   dist_1_grupo2={grade_b['distancia_1_grupo2']} (campo dist_b_1)")
                                print(f"[PILARES_ESPECIAIS]   dist_2_grupo2={grade_b['distancia_2_grupo2']} (campo dist_b_2)")
                                
                                # Aplicar as grades diretamente nos campos de entrada (SIMPLES)
                                # Grade A (Conjunto 1) - aplicar nos campos normais
                                gerador.grade1_entry.delete(0, 'end')
                                gerador.grade1_entry.insert(0, str(grade_a['grade_1']) if grade_a['grade_1'] else "")
                                gerador.grade2_entry.delete(0, 'end')
                                gerador.grade2_entry.insert(0, str(grade_a['grade_2']) if grade_a['grade_2'] else "")
                                gerador.grade3_entry.delete(0, 'end')
                                gerador.grade3_entry.insert(0, str(grade_a['grade_3']) if grade_a['grade_3'] else "")
                                gerador.distancia1_entry.delete(0, 'end')
                                gerador.distancia1_entry.insert(0, str(grade_a['distancia_1']) if grade_a['distancia_1'] else "")
                                gerador.distancia2_entry.delete(0, 'end')
                                gerador.distancia2_entry.insert(0, str(grade_a['distancia_2']) if grade_a['distancia_2'] else "")
                                
                                # Grade B (Conjunto 2) - aplicar nos campos do grupo 2
                                gerador.grade1_grupo2_entry.delete(0, 'end')
                                gerador.grade1_grupo2_entry.insert(0, str(grade_b['grade_1_grupo2']) if grade_b['grade_1_grupo2'] else "")
                                gerador.grade2_grupo2_entry.delete(0, 'end')
                                gerador.grade2_grupo2_entry.insert(0, str(grade_b['grade_2_grupo2']) if grade_b['grade_2_grupo2'] else "")
                                gerador.grade3_grupo2_entry.delete(0, 'end')
                                gerador.grade3_grupo2_entry.insert(0, str(grade_b['grade_3_grupo2']) if grade_b['grade_3_grupo2'] else "")
                                gerador.distancia1_grupo2_entry.delete(0, 'end')
                                gerador.distancia1_grupo2_entry.insert(0, str(grade_b['distancia_1_grupo2']) if grade_b['distancia_1_grupo2'] else "")
                                gerador.distancia2_grupo2_entry.delete(0, 'end')
                                gerador.distancia2_grupo2_entry.insert(0, str(grade_b['distancia_2_grupo2']) if grade_b['distancia_2_grupo2'] else "")
                                
                                # Aplicar detalhes das grades aos campos da interface
                                # Detalhes Grade A (Conjunto 1) - aplicar nos campos normais
                                if hasattr(gerador, 'detalhe_grade1_entries') and gerador.detalhe_grade1_entries:
                                    # Aplicar os 5 valores da grade 1 aos campos detalhe_grade1_entries
                                    valores_grade1 = [
                                        detalhes_grade_a.get('detalhe_grade1_1', ''),
                                        detalhes_grade_a.get('detalhe_grade1_2', ''),
                                        detalhes_grade_a.get('detalhe_grade1_3', ''),
                                        detalhes_grade_a.get('detalhe_grade1_4', ''),
                                        detalhes_grade_a.get('detalhe_grade1_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade1):
                                        if i < len(gerador.detalhe_grade1_entries):
                                            gerador.detalhe_grade1_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade1_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade2_entries') and gerador.detalhe_grade2_entries:
                                    # Aplicar os 5 valores da grade 2 aos campos detalhe_grade2_entries
                                    valores_grade2 = [
                                        detalhes_grade_a.get('detalhe_grade2_1', ''),
                                        detalhes_grade_a.get('detalhe_grade2_2', ''),
                                        detalhes_grade_a.get('detalhe_grade2_3', ''),
                                        detalhes_grade_a.get('detalhe_grade2_4', ''),
                                        detalhes_grade_a.get('detalhe_grade2_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade2):
                                        if i < len(gerador.detalhe_grade2_entries):
                                            gerador.detalhe_grade2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade3_entries') and gerador.detalhe_grade3_entries:
                                    # Aplicar os 5 valores da grade 3 aos campos detalhe_grade3_entries
                                    valores_grade3 = [
                                        detalhes_grade_a.get('detalhe_grade3_1', ''),
                                        detalhes_grade_a.get('detalhe_grade3_2', ''),
                                        detalhes_grade_a.get('detalhe_grade3_3', ''),
                                        detalhes_grade_a.get('detalhe_grade3_4', ''),
                                        detalhes_grade_a.get('detalhe_grade3_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade3):
                                        if i < len(gerador.detalhe_grade3_entries):
                                            gerador.detalhe_grade3_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade3_entries[i].insert(0, valor if valor else "")

                                # Detalhes Grade B (Conjunto 2) - aplicar nos campos do grupo 2
                                if hasattr(gerador, 'detalhe_grade1_grupo2_entries') and gerador.detalhe_grade1_grupo2_entries:
                                    # Aplicar os 5 valores da grade 1 do grupo 2 aos campos detalhe_grade1_grupo2_entries
                                    valores_grade1_g2 = [
                                        detalhes_grade_b.get('detalhe_grade1_1_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade1_2_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade1_3_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade1_4_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade1_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade1_g2):
                                        if i < len(gerador.detalhe_grade1_grupo2_entries):
                                            gerador.detalhe_grade1_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade1_grupo2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade2_grupo2_entries') and gerador.detalhe_grade2_grupo2_entries:
                                    # Aplicar os 5 valores da grade 2 do grupo 2 aos campos detalhe_grade2_grupo2_entries
                                    valores_grade2_g2 = [
                                        detalhes_grade_b.get('detalhe_grade2_1_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade2_2_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade2_3_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade2_4_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade2_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade2_g2):
                                        if i < len(gerador.detalhe_grade2_grupo2_entries):
                                            gerador.detalhe_grade2_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade2_grupo2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade3_grupo2_entries') and gerador.detalhe_grade3_grupo2_entries:
                                    # Aplicar os 5 valores da grade 3 do grupo 2 aos campos detalhe_grade3_grupo2_entries
                                    valores_grade3_g2 = [
                                        detalhes_grade_b.get('detalhe_grade3_1_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade3_2_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade3_3_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade3_4_grupo2', ''),
                                        detalhes_grade_b.get('detalhe_grade3_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade3_g2):
                                        if i < len(gerador.detalhe_grade3_grupo2_entries):
                                            gerador.detalhe_grade3_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade3_grupo2_entries[i].insert(0, valor if valor else "")

                                # Aplicar detalhes das grades
                                gerador.grades_grupo1 = grade_a
                                gerador.grades_grupo2 = grade_b
                                gerador.detalhes_grades = detalhes_grade_a
                                gerador.detalhes_grades_grupo2 = detalhes_grade_b

                                print(f"[PILARES_ESPECIAIS] ✅ Pilar 1 - Grades aplicadas ao gerador:")
                                print(f"[PILARES_ESPECIAIS]   Conjunto 1 (Grade A): {gerador.grades_grupo1}")
                                print(f"[PILARES_ESPECIAIS]   Conjunto 2 (Grade B): {gerador.grades_grupo2}")
                                print(f"[PILARES_ESPECIAIS]   Detalhes Conjunto 1: {len(gerador.detalhes_grades)} valores")
                                print(f"[PILARES_ESPECIAIS]   Detalhes Conjunto 2: {len(gerador.detalhes_grades_grupo2)} valores")
                                
                            elif pilar_atual == 2:
                                # PILAR 2: Conjunto 1 = Grade E, Conjunto 2 = Grade F
                                print(f"[PILARES_ESPECIAIS] 🎯 PILAR 2 DETECTADO - Usando Grade E e Grade F")
                                print(f"[PILARES_ESPECIAIS]   Pilar atual: {pilar_atual}")
                                print(f"[PILARES_ESPECIAIS]   Condição executada: pilar_atual == 2 = {pilar_atual == 2}")
                                
                                # Grade E (Conjunto 1) - valores exatos dos campos grade_e_1, grade_e_2, grade_e_3, dist_e_1, dist_e_2
                                grade_e = {
                                    'grade_1': grades_especiais.get('grade_e_1', ''),
                                    'grade_2': grades_especiais.get('grade_e_2', ''),
                                    'grade_3': grades_especiais.get('grade_e_3', ''),
                                    'distancia_1': grades_especiais.get('dist_e_1', ''),
                                    'distancia_2': grades_especiais.get('dist_e_2', '')
                                }
                                
                                # Grade F (Conjunto 2) - valores exatos dos campos grade_f_1, grade_f_2, grade_f_3, dist_f_1, dist_f_2
                                grade_f = {
                                    'grade_1_grupo2': grades_especiais.get('grade_f_1', ''),
                                    'grade_2_grupo2': grades_especiais.get('grade_f_2', ''),
                                    'grade_3_grupo2': grades_especiais.get('grade_f_3', ''),
                                    'distancia_1_grupo2': grades_especiais.get('dist_f_1', ''),
                                    'distancia_2_grupo2': grades_especiais.get('dist_f_2', '')
                                }
                                
                                # Detalhes Grade E (Conjunto 1) - valores exatos dos campos detalhe_e_X_Y
                                detalhes_grade_e = {}
                                for grade_num in range(1, 4):
                                    for detalhe_num in range(1, 6):  # Corrigido: 5 detalhes (1-5)
                                        chave_original = f'detalhe_e_{grade_num}_{detalhe_num}'
                                        chave_nova = f'detalhe_grade{grade_num}_{detalhe_num}'
                                        detalhes_grade_e[chave_nova] = detalhes_especiais.get(chave_original, '')
                                
                                # Detalhes Grade F (Conjunto 2) - valores exatos dos campos detalhe_f_X_Y
                                detalhes_grade_f = {}
                                for grade_num in range(1, 4):
                                    for detalhe_num in range(1, 6):  # Corrigido: 5 detalhes (1-5)
                                        chave_original = f'detalhe_f_{grade_num}_{detalhe_num}'
                                        chave_nova = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                                        detalhes_grade_f[chave_nova] = detalhes_especiais.get(chave_original, '')
                                
                                print(f"[PILARES_ESPECIAIS] Grade E (Pilar 2 - Conjunto 1) - VALORES EXATOS DA ABA:")
                                print(f"[PILARES_ESPECIAIS]   grade_1={grade_e['grade_1']} (campo grade_e_1)")
                                print(f"[PILARES_ESPECIAIS]   grade_2={grade_e['grade_2']} (campo grade_e_2)")
                                print(f"[PILARES_ESPECIAIS]   grade_3={grade_e['grade_3']} (campo grade_e_3)")
                                print(f"[PILARES_ESPECIAIS]   dist_1={grade_e['distancia_1']} (campo dist_e_1)")
                                print(f"[PILARES_ESPECIAIS]   dist_2={grade_e['distancia_2']} (campo dist_e_2)")
                                
                                print(f"[PILARES_ESPECIAIS] Grade F (Pilar 2 - Conjunto 2) - VALORES EXATOS DA ABA:")
                                print(f"[PILARES_ESPECIAIS]   grade_1_grupo2={grade_f['grade_1_grupo2']} (campo grade_f_1)")
                                print(f"[PILARES_ESPECIAIS]   grade_2_grupo2={grade_f['grade_2_grupo2']} (campo grade_f_2)")
                                print(f"[PILARES_ESPECIAIS]   grade_3_grupo2={grade_f['grade_3_grupo2']} (campo grade_f_3)")
                                print(f"[PILARES_ESPECIAIS]   dist_1_grupo2={grade_f['distancia_1_grupo2']} (campo dist_f_1)")
                                print(f"[PILARES_ESPECIAIS]   dist_2_grupo2={grade_f['distancia_2_grupo2']} (campo dist_f_2)")
                                
                                # Aplicar as grades diretamente nos campos de entrada (SIMPLES)
                                # Grade E (Conjunto 1) - aplicar nos campos normais
                                gerador.grade1_entry.delete(0, 'end')
                                gerador.grade1_entry.insert(0, str(grade_e['grade_1']) if grade_e['grade_1'] else "")
                                gerador.grade2_entry.delete(0, 'end')
                                gerador.grade2_entry.insert(0, str(grade_e['grade_2']) if grade_e['grade_2'] else "")
                                gerador.grade3_entry.delete(0, 'end')
                                gerador.grade3_entry.insert(0, str(grade_e['grade_3']) if grade_e['grade_3'] else "")
                                gerador.distancia1_entry.delete(0, 'end')
                                gerador.distancia1_entry.insert(0, str(grade_e['distancia_1']) if grade_e['distancia_1'] else "")
                                gerador.distancia2_entry.delete(0, 'end')
                                gerador.distancia2_entry.insert(0, str(grade_e['distancia_2']) if grade_e['distancia_2'] else "")
                                
                                # Grade F (Conjunto 2) - aplicar nos campos do grupo 2
                                gerador.grade1_grupo2_entry.delete(0, 'end')
                                gerador.grade1_grupo2_entry.insert(0, str(grade_f['grade_1_grupo2']) if grade_f['grade_1_grupo2'] else "")
                                gerador.grade2_grupo2_entry.delete(0, 'end')
                                gerador.grade2_grupo2_entry.insert(0, str(grade_f['grade_2_grupo2']) if grade_f['grade_2_grupo2'] else "")
                                gerador.grade3_grupo2_entry.delete(0, 'end')
                                gerador.grade3_grupo2_entry.insert(0, str(grade_f['grade_3_grupo2']) if grade_f['grade_3_grupo2'] else "")
                                gerador.distancia1_grupo2_entry.delete(0, 'end')
                                gerador.distancia1_grupo2_entry.insert(0, str(grade_f['distancia_1_grupo2']) if grade_f['distancia_1_grupo2'] else "")
                                gerador.distancia2_grupo2_entry.delete(0, 'end')
                                gerador.distancia2_grupo2_entry.insert(0, str(grade_f['distancia_2_grupo2']) if grade_f['distancia_2_grupo2'] else "")

                                # Aplicar detalhes das grades aos campos da interface
                                # Detalhes Grade E (Conjunto 1) - aplicar nos campos normais
                                if hasattr(gerador, 'detalhe_grade1_entries') and gerador.detalhe_grade1_entries:
                                    # Aplicar os 5 valores da grade 1 aos campos detalhe_grade1_entries
                                    valores_grade1 = [
                                        detalhes_grade_e.get('detalhe_grade1_1', ''),
                                        detalhes_grade_e.get('detalhe_grade1_2', ''),
                                        detalhes_grade_e.get('detalhe_grade1_3', ''),
                                        detalhes_grade_e.get('detalhe_grade1_4', ''),
                                        detalhes_grade_e.get('detalhe_grade1_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade1):
                                        if i < len(gerador.detalhe_grade1_entries):
                                            gerador.detalhe_grade1_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade1_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade2_entries') and gerador.detalhe_grade2_entries:
                                    # Aplicar os 5 valores da grade 2 aos campos detalhe_grade2_entries
                                    valores_grade2 = [
                                        detalhes_grade_e.get('detalhe_grade2_1', ''),
                                        detalhes_grade_e.get('detalhe_grade2_2', ''),
                                        detalhes_grade_e.get('detalhe_grade2_3', ''),
                                        detalhes_grade_e.get('detalhe_grade2_4', ''),
                                        detalhes_grade_e.get('detalhe_grade2_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade2):
                                        if i < len(gerador.detalhe_grade2_entries):
                                            gerador.detalhe_grade2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade3_entries') and gerador.detalhe_grade3_entries:
                                    # Aplicar os 5 valores da grade 3 aos campos detalhe_grade3_entries
                                    valores_grade3 = [
                                        detalhes_grade_e.get('detalhe_grade3_1', ''),
                                        detalhes_grade_e.get('detalhe_grade3_2', ''),
                                        detalhes_grade_e.get('detalhe_grade3_3', ''),
                                        detalhes_grade_e.get('detalhe_grade3_4', ''),
                                        detalhes_grade_e.get('detalhe_grade3_5', '')
                                    ]
                                    for i, valor in enumerate(valores_grade3):
                                        if i < len(gerador.detalhe_grade3_entries):
                                            gerador.detalhe_grade3_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade3_entries[i].insert(0, valor if valor else "")

                                # Detalhes Grade F (Conjunto 2) - aplicar nos campos do grupo 2
                                if hasattr(gerador, 'detalhe_grade1_grupo2_entries') and gerador.detalhe_grade1_grupo2_entries:
                                    # Aplicar os 5 valores da grade 1 do grupo 2 aos campos detalhe_grade1_grupo2_entries
                                    valores_grade1_g2 = [
                                        detalhes_grade_f.get('detalhe_grade1_1_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade1_2_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade1_3_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade1_4_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade1_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade1_g2):
                                        if i < len(gerador.detalhe_grade1_grupo2_entries):
                                            gerador.detalhe_grade1_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade1_grupo2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade2_grupo2_entries') and gerador.detalhe_grade2_grupo2_entries:
                                    # Aplicar os 5 valores da grade 2 do grupo 2 aos campos detalhe_grade2_grupo2_entries
                                    valores_grade2_g2 = [
                                        detalhes_grade_f.get('detalhe_grade2_1_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade2_2_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade2_3_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade2_4_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade2_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade2_g2):
                                        if i < len(gerador.detalhe_grade2_grupo2_entries):
                                            gerador.detalhe_grade2_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade2_grupo2_entries[i].insert(0, valor if valor else "")

                                if hasattr(gerador, 'detalhe_grade3_grupo2_entries') and gerador.detalhe_grade3_grupo2_entries:
                                    # Aplicar os 5 valores da grade 3 do grupo 2 aos campos detalhe_grade3_grupo2_entries
                                    valores_grade3_g2 = [
                                        detalhes_grade_f.get('detalhe_grade3_1_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade3_2_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade3_3_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade3_4_grupo2', ''),
                                        detalhes_grade_f.get('detalhe_grade3_5_grupo2', '')
                                    ]
                                    for i, valor in enumerate(valores_grade3_g2):
                                        if i < len(gerador.detalhe_grade3_grupo2_entries):
                                            gerador.detalhe_grade3_grupo2_entries[i].delete(0, 'end')
                                            gerador.detalhe_grade3_grupo2_entries[i].insert(0, valor if valor else "")

                                # Aplicar detalhes das grades
                                gerador.grades_grupo1 = grade_e
                                gerador.grades_grupo2 = grade_f
                                gerador.detalhes_grades = detalhes_grade_e
                                gerador.detalhes_grades_grupo2 = detalhes_grade_f
                                
                                print(f"[PILARES_ESPECIAIS] ✅ Pilar 2 - Grades aplicadas ao gerador:")
                                print(f"[PILARES_ESPECIAIS]   Conjunto 1 (Grade E): {gerador.grades_grupo1}")
                                print(f"[PILARES_ESPECIAIS]   Conjunto 2 (Grade F): {gerador.grades_grupo2}")
                                print(f"[PILARES_ESPECIAIS]   Detalhes Conjunto 1: {len(gerador.detalhes_grades)} valores")
                                print(f"[PILARES_ESPECIAIS]   Detalhes Conjunto 2: {len(gerador.detalhes_grades_grupo2)} valores")
                                
                            else:
                                print(f"[PILARES_ESPECIAIS] ⚠️ Pilar não identificado como Pilar 1 ou Pilar 2: pilar_atual={pilar_atual}")
                                print(f"[PILARES_ESPECIAIS] Mantendo dados originais das grades")
                                
                        else:
                            print(f"[PILARES_ESPECIAIS] ⚠️ Grades especiais não encontradas - mantendo dados originais")
                            
                    except Exception as e:
                        print(f"[PILARES_ESPECIAIS] ⚠️ Erro ao preencher grades automaticamente: {str(e)}")
                        print(f"[PILARES_ESPECIAIS] Mantendo dados originais das grades")
                        import traceback
                        traceback.print_exc()
                
                # Gerar script usando o método personalizado
                # Usar apenas o nome normal do pilar, sem sufixos
                gerador.nome_pilar_entry.delete(0, 'end')
                gerador.nome_pilar_entry.insert(0, nome_pilar)
                
                print(f"[PILARES_ESPECIAIS] Gerando script para {nome_pilar}...")
                
                # IMPORTANTE: NÃO recalcular valores - usar valores da aba Pilares Especiais
                print(f"[PILARES_ESPECIAIS] Mantendo valores da aba Pilares Especiais sem recálculo")
                
                # Gerar script
                script_cima = gerador.gerar_script()
                
                # Salvar script manualmente
                
                nome_arquivo_base = os.path.join(diretorio_pavimento, nome_pilar)
                if idx == 0:
                    nome_arquivo_cima = f"{nome_arquivo_base}_CIMA.scr"
                else:
                    nome_arquivo_cima = f"{nome_arquivo_base}_CIMA-{idx + 1}.scr"
                
                # Contador para evitar sobrescrita
                contador = 1
                while os.path.exists(nome_arquivo_cima):
                    nome_arquivo_cima = f"{nome_arquivo_base}_CIMA-{idx + 1}-{contador}.scr"
                    contador += 1
                
                with open(nome_arquivo_cima, "w", encoding="utf-16") as f:
                    f.write(script_cima)
                
                print(f"[PILARES_ESPECIAIS] Script salvo: {nome_arquivo_cima}")
                
                # Restaurar nome original (já está correto, sem sufixo)
                # gerador.nome_pilar_entry já contém o nome correto
        
            # Restaurar dimensões originais e configurações
            gerador.comprimento_pilar_entry.delete(0, 'end')
            gerador.comprimento_pilar_entry.insert(0, comp_original)
            gerador.largura_pilar_entry.delete(0, 'end')
            gerador.largura_pilar_entry.insert(0, larg_original)
            
            # Restaurar configurações padrão
            gerador.pilar_rotacionado_var.set(False)
            gerador.usar_scale_var.set(True)
            gerador.pilar_especial_ativo_var.set(False)
            print("[PILARES_ESPECIAIS] Configurações restauradas para o padrão")
            
            # COMBINAR OS SCRIPTS EM UM ÚNICO ARQUIVO
            print(f"[PILARES_ESPECIAIS] Combinando {len(dimensoes_especiais)} scripts em um único arquivo...")
            
            script_combinado = ""
            scripts_gerados = []
            
            # Coletar todos os scripts gerados
            for idx, dimensao in enumerate(dimensoes_especiais):
                # Usar nome único para cada script do pilar especial
                if idx == 0:
                    nome_script = f"{nome_pilar}_CIMA.scr"
                else:
                    nome_script = f"{nome_pilar}_CIMA-{idx + 1}.scr"
                caminho_script = os.path.join(diretorio_pavimento, nome_script)
                
                if os.path.exists(caminho_script):
                    try:
                        with open(caminho_script, 'r', encoding='utf-16') as f:
                            conteudo = f.read()
                            script_combinado += conteudo + "\n"
                            scripts_gerados.append(nome_script)
                            print(f"[PILARES_ESPECIAIS] Script {nome_script} adicionado ao combinado")
                    except Exception as e:
                        print(f"[PILARES_ESPECIAIS] Erro ao ler {nome_script}: {e}")
            
            # Salvar script combinado
            if script_combinado:
                nome_combinado = f"{nome_pilar}_COMBINADO_CIMA.scr"
                caminho_combinado = os.path.join(diretorio_pavimento, nome_combinado)
                
                try:
                    with open(caminho_combinado, 'w', encoding='utf-16') as f:
                        f.write(script_combinado)
                    print(f"[PILARES_ESPECIAIS] Script combinado salvo: {caminho_combinado}")
                    print(f"[PILARES_ESPECIAIS] Scripts incluídos: {', '.join(scripts_gerados)}")
                    
                    # REMOVER SCRIPTS INDIVIDUAIS APÓS COMBINAR
                    # Eles não são mais necessários, apenas o combinado deve permanecer
                    scripts_removidos = 0
                    for nome_script in scripts_gerados:
                        caminho_script = os.path.join(diretorio_pavimento, nome_script)
                        if os.path.exists(caminho_script):
                            try:
                                os.remove(caminho_script)
                                print(f"[PILARES_ESPECIAIS] ✅ Script individual removido: {nome_script}")
                                scripts_removidos += 1
                            except Exception as e:
                                print(f"[PILARES_ESPECIAIS] ⚠️ Erro ao remover {nome_script}: {e}")
                    
                    if scripts_removidos > 0:
                        print(f"[PILARES_ESPECIAIS] ✅ Removidos {scripts_removidos} scripts individuais (mantido apenas o combinado)")
                except Exception as e:
                    print(f"[PILARES_ESPECIAIS] Erro ao salvar script combinado: {e}")
            
            print(f"[PILARES_ESPECIAIS] Scripts múltiplos gerados e combinados para pilar especial '{nome_pilar}'")
        
        else:
            # PILAR COMUM: Gerar script único
            print(f"[PILAR_COMUM] Gerando script único para pilar comum...")
            
            # Definir diretórios
            nome_pilar = dados_pilar.get('nome', 'PILAR_SEM_NOME')
            pavimento = dados_pilar.get('pavimento', 'SEM_PAVIMENTO')
            
            # CORREÇÃO CRÍTICA: Usar a mesma normalização que _dcad_pavimento_cima usa (remove acentos)
            import unicodedata
            def normalizar_nome_pasta(texto):
                """Normaliza o nome removendo acentos (mesma função usada em _dcad_pavimento_cima)"""
                if not texto:
                    return ""
                texto = str(texto).strip()
                if not texto:
                    return ""
                try:
                    texto_normalizado = unicodedata.normalize('NFD', texto)
                    texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                except Exception:
                    texto_sem_acentos = texto
                texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                while '__' in texto_limpo:
                    texto_limpo = texto_limpo.replace('__', '_')
                texto_limpo = texto_limpo.strip('_')
                return texto_limpo
            
            pavimento_normalizado = normalizar_nome_pasta(pavimento)
            
            # CORREÇÃO: Usar robust_path_resolver para path correto em frozen e dev
            try:
                from utils.robust_path_resolver import robust_path_resolver
                diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
            except ImportError:
                try:
                    from src.utils.robust_path_resolver import robust_path_resolver
                    diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
                except ImportError:
                    # Fallback manual (não recomendado, mas necessário em alguns casos)
                    diretorio_base = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "output", "scripts")
            nome_pasta_pavimento = f"{pavimento_normalizado}_CIMA"
            diretorio_pavimento = os.path.join(diretorio_base, nome_pasta_pavimento)
            
            print(f"[PILAR_COMUM] Pavimento original: '{pavimento}'")
            print(f"[PILAR_COMUM] Pavimento normalizado: '{pavimento_normalizado}'")
            print(f"[PILAR_COMUM] Nome da pasta: '{nome_pasta_pavimento}'")
            print(f"[PILAR_COMUM] Diretório: {diretorio_pavimento}")
            
            os.makedirs(diretorio_pavimento, exist_ok=True)
            
            # Configurar para pilar comum (sem pilar especial)
            gerador.pilar_rotacionado_var.set(False)
            gerador.usar_scale_var.set(True)
            gerador.pilar_especial_ativo_var.set(False)
            
            # Limpar globais do pilar especial
            if hasattr(gerador, '_globais_pilar_especial'):
                gerador._globais_pilar_especial = {}
            
            print(f"[PILAR_COMUM] Configuração: ROTATE=False, SCALE=True, ESPECIAL=False")
            
            # Gerar script
            script_cima = gerador.gerar_script()
            
            # Salvar script
            nome_arquivo_cima = os.path.join(diretorio_pavimento, f"{nome_pilar}_CIMA.scr")
            base_nome = f"{nome_pilar}_CIMA"
            
            # REMOÇÃO AGRESSIVA: Remover TODOS os arquivos SCR relacionados a este item ANTES de gerar
            # Isso garante que não haverá arquivos antigos acumulando
            print(f"[PILAR_COMUM] 🗑️ Removendo TODOS os arquivos SCR antigos para '{nome_pilar}'...")
            
            arquivos_removidos = []
            if os.path.exists(diretorio_pavimento):
                # Listar TODOS os arquivos SCR primeiro (para evitar problemas de iteração)
                todos_arquivos = []
                try:
                    todos_arquivos = os.listdir(diretorio_pavimento)
                except Exception as e:
                    print(f"[PILAR_COMUM] ⚠️ Erro ao listar diretório: {e}")
                
                # Filtrar apenas arquivos SCR que correspondem a este pilar
                for arquivo_na_pasta in todos_arquivos:
                    if not arquivo_na_pasta.endswith('.scr'):
                        continue
                    
                    # Verificar se o arquivo corresponde a este pilar (pode ter sufixos -1, -2, etc.)
                    # Exemplos: P8_CIMA.scr, P8_CIMA-1.scr, P8_CIMA-2.scr, etc.
                    nome_sem_ext = arquivo_na_pasta.replace('.scr', '')
                    
                    # Verificar se começa com o nome base (ex: "P8_CIMA")
                    corresponde = (
                        arquivo_na_pasta == f"{base_nome}.scr" or  # Exato
                        nome_sem_ext.startswith(f"{base_nome}-") or  # Com sufixo -1, -2
                        arquivo_na_pasta.startswith(f"{nome_pilar}_")  # Qualquer variação do nome
                    )
                    
                    if corresponde:
                        caminho_arquivo = os.path.join(diretorio_pavimento, arquivo_na_pasta)
                        try:
                            os.remove(caminho_arquivo)
                            arquivos_removidos.append(arquivo_na_pasta)
                            print(f"[PILAR_COMUM] ✅ Removido: {arquivo_na_pasta}")
                        except Exception as e:
                            print(f"[PILAR_COMUM] ❌ Erro ao remover {arquivo_na_pasta}: {e}")
                            # Tentar novamente com um pequeno delay
                            try:
                                time.sleep(0.1)  # time já está importado no topo do arquivo
                                os.remove(caminho_arquivo)
                                arquivos_removidos.append(arquivo_na_pasta)
                                print(f"[PILAR_COMUM] ✅ Removido após retry: {arquivo_na_pasta}")
                            except Exception as e2:
                                print(f"[PILAR_COMUM] ❌ Falha definitiva ao remover {arquivo_na_pasta}: {e2}")
            
            if arquivos_removidos:
                print(f"[PILAR_COMUM] ✅ Removidos {len(arquivos_removidos)} arquivo(s) antigo(s) para '{nome_pilar}'")
            else:
                print(f"[PILAR_COMUM] ℹ️ Nenhum arquivo antigo encontrado para '{nome_pilar}' (normal se for primeira execução)")
            
            # Agora salvar o novo arquivo (sempre com o nome base, sem sufixos)
            with open(nome_arquivo_cima, "w", encoding="utf-16") as f:
                f.write(script_cima)
            
            print(f"[PILAR_COMUM] Script salvo: {nome_arquivo_cima}")
            print(f"[PILAR_COMUM] Script único gerado para pilar comum '{nome_pilar}'")

        # CORREÇÃO: Destruir a aplicação apenas no final
        print("[DEBUG-FUNC] Finalizando aplicação VISAO CIMA...")
        app.root.destroy()
        
        timestamp_func_fim = time.time()
        duracao_func = timestamp_func_fim - timestamp_func_inicio
        
        print(f"[CIMA_SCRIPT] === PROCESSAMENTO CONCLUÍDO ===")
        print(f"[CIMA_SCRIPT] Duração da função: {duracao_func:.2f} segundos")
        print(f"[CIMA_SCRIPT] Fim da função preencher_campos_diretamente_e_gerar_scripts: {timestamp_func_fim}")
        print(f"[CIMA_SCRIPT] ===== FIM DA FUNÇÃO preencher_campos_diretamente_e_gerar_scripts =====")
        
        print("Fim do processamento dos dados.")

    except Exception as e:
        print(f"[ERROR-FUNC] Erro ao processar os dados: {e}")
        import traceback
        traceback.print_exc()

def preencher_campos_e_gerar_scripts(caminho_arquivo_excel, coluna_especifica=None):
    """
    Lê os dados da primeira aba de uma planilha Excel, preenche os campos da interface
    VISAO CIMA e gera os scripts .scr correspondentes.

    Args:
        caminho_arquivo_excel: O caminho para o arquivo Excel.
        coluna_especifica: A coluna específica a ser processada.
    """

    # ADICIONAR LOG DE DEPURAÇÃO
    import time
    import os  # CORREÇÃO: Importar os no início da função
    import sys
    timestamp_func_inicio = time.time()
    print(f"[CIMA_SCRIPT] ===== INÍCIO DA FUNÇÃO preencher_campos_e_gerar_scripts =====")
    print(f"[CIMA_SCRIPT] Timestamp: {timestamp_func_inicio}")
    print(f"[CIMA_SCRIPT] PID: {os.getpid()}")
    print(f"[CIMA_SCRIPT] Argumentos: caminho_arquivo_excel={caminho_arquivo_excel}, coluna_especifica={coluna_especifica}")

    try:
        # Carrega a planilha
        workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        sheet = workbook.worksheets[0]  # Obtém a primeira aba
        print(f"[DEBUG-FUNC] Planilha carregada: {caminho_arquivo_excel}")

        # Mapeamento de linhas para campos da interface VISAO CIMA
        linhas_cima = {
            'pavimento': 4,  # Linha 2: Pavimento/CIMA
            'nome': 5,  # Linha 4: Nome
            'comprimento': 7,  # Linha 6: COMPRIMENTO
            'largura': 8,  # Linha 7: Largura
            'parafuso_p1_p2': 174,  # Atualizado para P1-P2
            'parafuso_p2_p3': 175,  
            'parafuso_p3_p4': 176,  
            'parafuso_p4_p5': 177,  
            'parafuso_p5_p6': 178,  
            'parafuso_p6_p7': 179,  
            'parafuso_p7_p8': 180,  
            'parafuso_p8_p9': 181,  # Adicionado para P8-P9
            'grade1': 182,  
            'distancia1': 183,  
            'grade2': 184,  
            'distancia2': 185,  
            'grade3': 186,  
            # Detalhes das Grades
            'detalhe_grade1_1': 193,
            'detalhe_grade1_2': 194,
            'detalhe_grade1_3': 195,
            'detalhe_grade1_4': 196,
            'detalhe_grade1_5': 197,
            'detalhe_grade2_1': 198,
            'detalhe_grade2_2': 199,
            'detalhe_grade2_3': 200,
            'detalhe_grade2_4': 201,
            'detalhe_grade2_5': 202,
            'detalhe_grade3_1': 203,
            'detalhe_grade3_2': 204,
            'detalhe_grade3_3': 205,
            'detalhe_grade3_4': 206,
            'detalhe_grade3_5': 207,
        }

        coluna_base = 'E'  # Coluna inicial para os dados dos pilares
        coluna_base_idx = ord(coluna_base) - ord('A')
        colunas_vazias = 0
        parar_busca = False

        # Função para converter índice numérico para letra de coluna (suporta AA, AB, etc.)
        def coluna_para_letra(col_idx):
            letra = ''
            while col_idx >= 0:
                letra = chr(ord('A') + (col_idx % 26)) + letra
                col_idx = col_idx // 26 - 1
            return letra
            
        # Função para converter letra de coluna para índice numérico
        def letra_para_coluna(letra):
            col_idx = 0
            for i, char in enumerate(reversed(letra)):
                col_idx += (ord(char) - ord('A') + 1) * (26 ** i)
            return col_idx - 1

        # Se uma coluna específica foi fornecida, processar apenas essa coluna
        if coluna_especifica:
            # Verificar se é uma string com múltiplas colunas (ex: "E,F,G")
            if ',' in coluna_especifica:
                # Processar múltiplas colunas
                colunas_lista = [col.strip() for col in coluna_especifica.split(',')]
                colunas_a_processar = [letra_para_coluna(col) for col in colunas_lista]
                print(f"[DEBUG-FUNC] Processando múltiplas colunas: {colunas_lista}")
            else:
                # Processar uma única coluna
                colunas_a_processar = [letra_para_coluna(coluna_especifica)]
                print(f"[DEBUG-FUNC] Processando apenas a coluna {coluna_especifica}")
        else:
            # Caso contrário, iterar por todas as colunas a partir da coluna base
            colunas_a_processar = range(coluna_base_idx, sheet.max_column)
            print(f"[DEBUG-FUNC] Processando todas as colunas a partir de {coluna_base}")

        print(f"[DEBUG-FUNC] Total de colunas a processar: {len(colunas_a_processar)}")
        print(f"[DEBUG-FUNC] Colunas a processar: {list(colunas_a_processar)}")

        # Itera pelas colunas a processar
        colunas_processadas = set()  # Conjunto para rastrear colunas já processadas
        
        # CORREÇÃO: Criar a aplicação UMA ÚNICA VEZ fora do loop
        print("[DEBUG-FUNC] Iniciando aplicação VISAO CIMA...")
        app = AplicacaoUnificada()
        gerador = app.gerador
        print("[DEBUG-FUNC] Aplicação VISAO CIMA criada com sucesso")
        

        
        print(f"[DEBUG-FUNC] === INICIANDO PROCESSAMENTO DE {len(colunas_a_processar)} COLUNAS ===")
        
        for col_idx in colunas_a_processar:
            if parar_busca:
                break

            coluna = coluna_para_letra(col_idx)
            
            # Verificar se a coluna já foi processada
            if coluna in colunas_processadas:
                print(f"[DEBUG-FUNC] Pulando coluna {coluna}: já foi processada anteriormente.")
                continue
            
            colunas_processadas.add(coluna)  # Marcar como processada
            print(f"[DEBUG-FUNC] Processando coluna {coluna} (índice {col_idx})")
            
            # Verifica se a coluna tem um nome de pilar
            nome_pilar = sheet[f'{coluna}{linhas_cima["nome"]}'].value
            if nome_pilar:
                colunas_vazias = 0  # Reinicia a contagem de colunas vazias
                
                print(f"[DEBUG-FUNC] === PROCESSANDO COLUNA {coluna} - PILAR: {nome_pilar} ===")
                
                # CORREÇÃO: Usar a aplicação já criada, não criar uma nova
                # app = AplicacaoUnificada()  # REMOVIDO - já criado fora do loop
                # gerador = app.gerador  # REMOVIDO - já obtido fora do loop

                # 1. Primeiro preencher dados básicos
                campos_basicos = ['pavimento', 'nome', 'comprimento', 'largura']
                for campo in campos_basicos:
                    valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                    print(f"  Preenchendo {campo}: {valor}")
                    
                    if campo == 'pavimento':
                        gerador.pavimento_entry.delete(0, 'end')
                        gerador.pavimento_entry.insert(0, valor if valor is not None else "")
                    elif campo == 'nome':
                        gerador.nome_pilar_entry.delete(0, 'end')
                        gerador.nome_pilar_entry.insert(0, valor if valor is not None else "")
                    elif campo == 'comprimento':
                        gerador.comprimento_pilar_entry.delete(0, 'end')
                        gerador.comprimento_pilar_entry.insert(0, valor if valor is not None else "")
                    elif campo == 'largura':
                        gerador.largura_pilar_entry.delete(0, 'end')
                        gerador.largura_pilar_entry.insert(0, valor if valor is not None else "")

                # 2. Calcular valores iniciais
                print("Calculando valores iniciais...")
                gerador.calcular_valores()

                # 3. Limpar e preencher parafusos na ordem correta
                print("Preenchendo parafusos...")
                parafusos_ordem = [
                    'parafuso_p1_p2',  # Atualizado para P1-P2
                    'parafuso_p2_p3',
                    'parafuso_p3_p4',
                    'parafuso_p4_p5',
                    'parafuso_p5_p6',
                    'parafuso_p6_p7',
                    'parafuso_p7_p8',
                    'parafuso_p8_p9'   # Adicionado P8-P9
                ]

                # Limpar todos os campos de parafusos primeiro (sem disparar traces)
                print("[DEBUG] Limpando campos de parafusos pilares comuns - desabilitando traces temporariamente")
                
                # Desabilitar traces temporariamente para evitar erros durante limpeza
                traces_salvos = []
                for entry in gerador.parafuso_entries:
                    # Salvar e remover traces
                    if hasattr(entry, 'valor_var') and entry.valor_var:
                        traces = entry.valor_var.trace_info()
                        traces_salvos.append((entry, traces))
                        for trace in traces:
                            entry.valor_var.trace_remove(trace[0], trace[1])
                    
                    # Limpar campo
                    entry.delete(0, 'end')
                    entry.insert(0, "0")  # Inserir zero como valor padrão
                
                # Restaurar traces
                for entry, traces in traces_salvos:
                    for trace in traces:
                        entry.valor_var.trace_add(trace[0], trace[1])
                
                print("[DEBUG] Limpeza pilares comuns concluída - traces restaurados")

                # Preencher parafusos na ordem
                for i, parafuso in enumerate(parafusos_ordem):
                    if i < len(gerador.parafuso_entries):  # Verificar se o índice está dentro dos limites
                        valor = sheet[f'{coluna}{linhas_cima[parafuso]}'].value if parafuso in linhas_cima else None
                        if valor is not None:
                            gerador.parafuso_entries[i].delete(0, 'end')
                            gerador.parafuso_entries[i].insert(0, valor)
                            print(f"  Preenchendo {parafuso}: {valor}")

                # 4. Recalcular após parafusos
                print("Recalculando após parafusos...")
                gerador.calcular_valores()

                # 5. Preencher grades na ordem correta
                print("Preenchendo grades e distâncias...")
                grades_ordem = [
                    ('grade1', gerador.grade1_entry),
                    ('distancia1', gerador.distancia1_entry),
                    ('grade2', gerador.grade2_entry),
                    ('distancia2', gerador.distancia2_entry),
                    ('grade3', gerador.grade3_entry)
                ]

                for campo, entry in grades_ordem:
                    valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                    entry.delete(0, 'end')
                    entry.insert(0, valor if valor is not None else "")
                    print(f"  Preenchendo {campo}: {valor}")

                # 6. NÃO recalcular - manter valores preenchidos
                print("Mantendo valores preenchidos sem recalcular...")

                # 7. Verificar e ativar "pilar rotacionado" se "pilar especial" estiver ativo
                print("Verificando status do pilar especial...")
                try:
                    # Verificar se o gerador atual tem o checkbox pilar_rotacionado_var
                    print(f"  Gerador atual: {gerador}")
                    print(f"  Tipo do gerador: {type(gerador)}")
                    print(f"  Tem pilar_rotacionado_var: {hasattr(gerador, 'pilar_rotacionado_var')}")
                    
                    if hasattr(gerador, 'pilar_rotacionado_var'):
                        status_atual = gerador.pilar_rotacionado_var.get()
                        print(f"  Status atual do checkbox: {status_atual}")
                        
                        # Verificar se existe uma instância da interface principal com pilar especial ativo
                        pilar_especial_ativo = False
                        
                        # Buscar globalmente por instâncias com pilar especial
                        import gc
                        print("  Buscando globalmente por instancias com pilar especial...")
                        objetos_encontrados = 0
                        for obj in gc.get_objects():
                            try:
                                if hasattr(obj, 'ativar_pilar_especial'):
                                    objetos_encontrados += 1
                                    attr = getattr(obj, 'ativar_pilar_especial', None)
                                    if attr and hasattr(attr, 'get'):
                                        valor = attr.get()
                                        print(f"    Objeto {objetos_encontrados}: pilar_especial = {valor}")
                                        if valor:
                                            pilar_especial_ativo = True
                                            break
                            except Exception:
                                continue
                        print(f"  Total de objetos com ativar_pilar_especial encontrados: {objetos_encontrados}")
                        
                        # Se pilar especial estiver ativo, ativar o pilar rotacionado
                        if pilar_especial_ativo:
                            print("PILAR ESPECIAL ATIVO - Ativando 'Pilar Rotacionado' automaticamente")
                            gerador.pilar_rotacionado_var.set(True)
                            novo_status = gerador.pilar_rotacionado_var.get()
                            print(f"  Checkbox ativado diretamente. Novo status: {novo_status}")
                        else:
                            print("Pilar especial nao esta ativo - mantendo configuracao atual")
                    else:
                        print("  Gerador nao tem pilar_rotacionado_var")
                        
                except Exception as e:
                    print("Erro ao verificar pilar especial: " + str(e))

                # 8. Preencher parafusos novamente com valores da planilha
                print("Preenchendo parafusos pela última vez...")
                for i, parafuso in enumerate(parafusos_ordem):
                    if i < len(gerador.parafuso_entries):  # Verificar se o índice está dentro dos limites
                        valor = sheet[f'{coluna}{linhas_cima[parafuso]}'].value if parafuso in linhas_cima else None
                        if valor is not None:
                            gerador.parafuso_entries[i].delete(0, 'end')
                            gerador.parafuso_entries[i].insert(0, valor)
                            print(f"  Preenchendo final {parafuso}: {valor}")

                # 8.1 Preencher grades novamente
                print("Preenchendo grades pela última vez...")
                for campo, entry in grades_ordem:
                    valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                    entry.delete(0, 'end')
                    entry.insert(0, valor if valor is not None else "")
                    print(f"  Preenchendo final {campo}: {valor}")

                # 8.2 Preencher detalhes das grades
                print("Preenchendo detalhes das grades...")
                
                # Mapear os campos de detalhes para as entradas correspondentes
                detalhes_grade1 = [
                    ('detalhe_grade1_1', 0),
                    ('detalhe_grade1_2', 1),
                    ('detalhe_grade1_3', 2),
                    ('detalhe_grade1_4', 3),
                    ('detalhe_grade1_5', 4)
                ]
                
                detalhes_grade2 = [
                    ('detalhe_grade2_1', 0),
                    ('detalhe_grade2_2', 1),
                    ('detalhe_grade2_3', 2),
                    ('detalhe_grade2_4', 3),
                    ('detalhe_grade2_5', 4)
                ]
                
                detalhes_grade3 = [
                    ('detalhe_grade3_1', 0),
                    ('detalhe_grade3_2', 1),
                    ('detalhe_grade3_3', 2),
                    ('detalhe_grade3_4', 3),
                    ('detalhe_grade3_5', 4)
                ]

                # Preencher detalhes da Grade 1
                for campo, idx in detalhes_grade1:
                    if idx < len(gerador.detalhe_grade1_entries):
                        valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                        gerador.detalhe_grade1_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade1_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")

                # Preencher detalhes da Grade 2
                for campo, idx in detalhes_grade2:
                    if idx < len(gerador.detalhe_grade2_entries):
                        valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                        gerador.detalhe_grade2_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade2_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")

                # Preencher detalhes da Grade 3
                for campo, idx in detalhes_grade3:
                    if idx < len(gerador.detalhe_grade3_entries):
                        valor = sheet[f'{coluna}{linhas_cima[campo]}'].value
                        gerador.detalhe_grade3_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade3_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")

                # 8.3 Preencher Grade Grupo 2 - NOVO
                print("Preenchendo Grade Grupo 2...")
                
                # Mapeamento das linhas do Grade Grupo 2
                # CORREÇÃO: Compensar o deslocamento +1 do Excel temporário
                linhas_grupo2 = {
                    'grade1_grupo2': 220,  # 219 + 1 (compensação do Excel temporário)
                    'distancia1_grupo2': 221,  # 220 + 1 (compensação do Excel temporário)
                    'grade2_grupo2': 222,  # 221 + 1 (compensação do Excel temporário)
                    'distancia2_grupo2': 223,  # 222 + 1 (compensação do Excel temporário)
                    'grade3_grupo2': 224,  # 223 + 1 (compensação do Excel temporário)
                    # Detalhes das grades do Grupo 2
                    'detalhe_grade1_1_grupo2': 225,  # 224 + 1 (compensação do Excel temporário)
                    'detalhe_grade1_2_grupo2': 226,  # 225 + 1 (compensação do Excel temporário)
                    'detalhe_grade1_3_grupo2': 227,  # 226 + 1 (compensação do Excel temporário)
                    'detalhe_grade1_4_grupo2': 228,  # 227 + 1 (compensação do Excel temporário)
                    'detalhe_grade1_5_grupo2': 229,  # 228 + 1 (compensação do Excel temporário)
                    'detalhe_grade2_1_grupo2': 230,  # 229 + 1 (compensação do Excel temporário)
                    'detalhe_grade2_2_grupo2': 231,  # 230 + 1 (compensação do Excel temporário)
                    'detalhe_grade2_3_grupo2': 232,  # 231 + 1 (compensação do Excel temporário)
                    'detalhe_grade2_4_grupo2': 233,  # 232 + 1 (compensação do Excel temporário)
                    'detalhe_grade2_5_grupo2': 234,  # 233 + 1 (compensação do Excel temporário)
                    'detalhe_grade3_1_grupo2': 235,  # 234 + 1 (compensação do Excel temporário)
                    'detalhe_grade3_2_grupo2': 236,  # 235 + 1 (compensação do Excel temporário)
                    'detalhe_grade3_3_grupo2': 237,  # 236 + 1 (compensação do Excel temporário)
                    'detalhe_grade3_4_grupo2': 238,  # 237 + 1 (compensação do Excel temporário)
                    'detalhe_grade3_5_grupo2': 239   # 238 + 1 (compensação do Excel temporário)
                }

                # Preencher grades do Grupo 2
                grades_grupo2_ordem = [
                    ('grade1_grupo2', gerador.grade1_grupo2_entry),
                    ('distancia1_grupo2', gerador.distancia1_grupo2_entry),
                    ('grade2_grupo2', gerador.grade2_grupo2_entry),
                    ('distancia2_grupo2', gerador.distancia2_grupo2_entry),
                    ('grade3_grupo2', gerador.grade3_grupo2_entry)
                ]

                for campo, entry in grades_grupo2_ordem:
                    linha = linhas_grupo2[campo]
                    celula = f'{coluna}{linha}'
                    valor = sheet[celula].value
                    entry.delete(0, 'end')
                    entry.insert(0, valor if valor is not None else "")
                    print(f"  Preenchendo {campo} (linha {linha}, célula {celula}): {valor}")
                    
                # CORREÇÃO: Verificar se os valores estão sendo preenchidos corretamente
                print(f"  VERIFICAÇÃO - Valores preenchidos:")
                print(f"    grade1_grupo2_entry: {gerador.grade1_grupo2_entry.get()}")
                print(f"    distancia1_grupo2_entry: {gerador.distancia1_grupo2_entry.get()}")
                print(f"    grade2_grupo2_entry: {gerador.grade2_grupo2_entry.get()}")
                print(f"    distancia2_grupo2_entry: {gerador.distancia2_grupo2_entry.get()}")
                print(f"    grade3_grupo2_entry: {gerador.grade3_grupo2_entry.get()}")

                # Preencher detalhes das grades do Grupo 2
                detalhes_grade1_grupo2 = [
                    ('detalhe_grade1_1_grupo2', 0),
                    ('detalhe_grade1_2_grupo2', 1),
                    ('detalhe_grade1_3_grupo2', 2),
                    ('detalhe_grade1_4_grupo2', 3),
                    ('detalhe_grade1_5_grupo2', 4)
                ]
                
                detalhes_grade2_grupo2 = [
                    ('detalhe_grade2_1_grupo2', 0),
                    ('detalhe_grade2_2_grupo2', 1),
                    ('detalhe_grade2_3_grupo2', 2),
                    ('detalhe_grade2_4_grupo2', 3),
                    ('detalhe_grade2_5_grupo2', 4)
                ]
                
                detalhes_grade3_grupo2 = [
                    ('detalhe_grade3_1_grupo2', 0),
                    ('detalhe_grade3_2_grupo2', 1),
                    ('detalhe_grade3_3_grupo2', 2),
                    ('detalhe_grade3_4_grupo2', 3),
                    ('detalhe_grade3_5_grupo2', 4)
                ]

                # Preencher detalhes da Grade 1 Grupo 2
                for campo, idx in detalhes_grade1_grupo2:
                    if idx < len(gerador.detalhe_grade1_grupo2_entries):
                        valor = sheet[f'{coluna}{linhas_grupo2[campo]}'].value
                        gerador.detalhe_grade1_grupo2_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade1_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")

                # Preencher detalhes da Grade 2 Grupo 2
                for campo, idx in detalhes_grade2_grupo2:
                    if idx < len(gerador.detalhe_grade2_grupo2_entries):
                        valor = sheet[f'{coluna}{linhas_grupo2[campo]}'].value
                        gerador.detalhe_grade2_grupo2_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade2_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")

                # Preencher detalhes da Grade 3 Grupo 2
                for campo, idx in detalhes_grade3_grupo2:
                    if idx < len(gerador.detalhe_grade3_grupo2_entries):
                        valor = sheet[f'{coluna}{linhas_grupo2[campo]}'].value
                        gerador.detalhe_grade3_grupo2_entries[idx].delete(0, 'end')
                        gerador.detalhe_grade3_grupo2_entries[idx].insert(0, valor if valor is not None else "")
                        print(f"  Preenchendo {campo}: {valor}")



                # 10. Verificar se pilar especial está ativo para gerar múltiplos scripts
                print("Verificando se deve gerar multiplos scripts...")
                
                # Ler dados do pilar especial do Excel temporário
                pilar_especial_ativo = False
                dimensoes_especiais = []
                
                try:
                    print("Lendo dados do pilar especial do Excel temporario...")
                    linha_base = 999  # Ajustado para compensar linha adicional
                    
                    # Ler se pilar especial está ativo
                    ativo_cell = sheet.cell(row=linha_base + 1, column=1).value
                    pilar_especial_ativo = bool(ativo_cell) if ativo_cell is not None else False
                    
                    print(f"Pilar especial ativo (do Excel): {pilar_especial_ativo}")
                    
                    if pilar_especial_ativo:
                        print("Pilar especial ativo - coletando dimensoes especiais do Excel...")
                        
                        # Ler dados do pilar especial
                        tipo = sheet.cell(row=linha_base + 2, column=1).value or ''
                        comp_1 = sheet.cell(row=linha_base + 3, column=1).value or ''
                        comp_2 = sheet.cell(row=linha_base + 4, column=1).value or ''
                        comp_3 = sheet.cell(row=linha_base + 5, column=1).value or ''
                        larg_1 = sheet.cell(row=linha_base + 6, column=1).value or ''
                        larg_2 = sheet.cell(row=linha_base + 7, column=1).value or ''
                        larg_3 = sheet.cell(row=linha_base + 8, column=1).value or ''
                        
                        print(f"Dados lidos: tipo={tipo}, comp_1={comp_1}, larg_1={larg_1}, comp_2={comp_2}, larg_2={larg_2}, comp_3={comp_3}, larg_3={larg_3}")
                        
                        # Coletar dimensões válidas
                        dimensoes_dados = [
                            (comp_1, larg_1, 1),
                            (comp_2, larg_2, 2),
                            (comp_3, larg_3, 3)
                        ]
                        
                        for comp_val, larg_val, indice in dimensoes_dados:
                            comp_str = str(comp_val).strip()
                            larg_str = str(larg_val).strip()
                            
                            if comp_str and larg_str and comp_str != '0' and larg_str != '0':
                                dimensoes_especiais.append({
                                    'comprimento': comp_str,
                                    'largura': larg_str,
                                    'indice': indice
                                })
                                print(f"  Dimensao {indice}: comp={comp_str}, larg={larg_str}")
                        
                        print(f"Total de dimensoes validas encontradas: {len(dimensoes_especiais)}")
                        
                        # Ler globais do pilar especial (linhas 1100+)
                        globais_pilar_especial = {}
                        try:
                            linha_globais = 1100
                            
                            # Informações básicas
                            tipo_global = sheet.cell(row=linha_globais + 1, column=1).value or ''
                            comp_1_global = sheet.cell(row=linha_globais + 2, column=1).value or ''
                            comp_2_global = sheet.cell(row=linha_globais + 3, column=1).value or ''
                            larg_1_global = sheet.cell(row=linha_globais + 4, column=1).value or ''
                            larg_2_global = sheet.cell(row=linha_globais + 5, column=1).value or ''
                            
                            print(f"[DEBUG-GLOBAIS] Lendo globais do Excel temporário...")
                            print(f"[DEBUG-GLOBAIS] Linha base: {linha_globais}")
                            print(f"[DEBUG-GLOBAIS] Tipo global: '{tipo_global}'")
                            print(f"[DEBUG-GLOBAIS] Comp_1: '{comp_1_global}', Comp_2: '{comp_2_global}'")
                            print(f"[DEBUG-GLOBAIS] Larg_1: '{larg_1_global}', Larg_2: '{larg_2_global}'")
                            
                            # Verificar se o pilar especial está ativo (linha 1000)
                            pilar_especial_ativo = sheet.cell(row=1000, column=1).value
                            print(f"[DEBUG-GLOBAIS] Pilar especial ativo (linha 1000): {pilar_especial_ativo}")
                            
                            # Verificar dados básicos do pilar especial (linhas 1001-1008)
                            tipo_basico = sheet.cell(row=1001, column=1).value
                            comp_1_basico = sheet.cell(row=1002, column=1).value
                            comp_2_basico = sheet.cell(row=1003, column=1).value
                            larg_1_basico = sheet.cell(row=1005, column=1).value
                            larg_2_basico = sheet.cell(row=1006, column=1).value
                            print(f"[DEBUG-GLOBAIS] Dados básicos - Tipo: '{tipo_basico}', Comp_1: '{comp_1_basico}', Comp_2: '{comp_2_basico}'")
                            print(f"[DEBUG-GLOBAIS] Dados básicos - Larg_1: '{larg_1_basico}', Larg_2: '{larg_2_basico}'")
                            
                            if tipo_global == 'L':
                                print("[GLOBAIS] Lendo globais para pilar tipo L...")
                                
                                # Incluir também chaves básicas necessárias para ativação no robô
                                globais_pilar_especial['tipo_pilar'] = (tipo_global or tipo_basico or tipo or '').strip()
                                # Converter para float onde possível
                                try:
                                    globais_pilar_especial['comp_1'] = float(str(comp_1_global or comp_1_basico or comp_1 or 0).replace(',', '.'))
                                except Exception:
                                    globais_pilar_especial['comp_1'] = 0.0
                                try:
                                    globais_pilar_especial['comp_2'] = float(str(comp_2_global or comp_2_basico or comp_2 or 0).replace(',', '.'))
                                except Exception:
                                    globais_pilar_especial['comp_2'] = 0.0
                                try:
                                    globais_pilar_especial['larg_1'] = float(str(larg_1_global or larg_1_basico or larg_1 or 0).replace(',', '.'))
                                except Exception:
                                    globais_pilar_especial['larg_1'] = 0.0
                                try:
                                    globais_pilar_especial['larg_2'] = float(str(larg_2_global or larg_2_basico or larg_2 or 0).replace(',', '.'))
                                except Exception:
                                    globais_pilar_especial['larg_2'] = 0.0
                                
                                # PILAR 1 - Posições
                                globais_pilar_especial['pilar1_paia_posicao'] = sheet.cell(row=linha_globais + 10, column=1).value or 0
                                globais_pilar_especial['pilar1_paib_posicao'] = sheet.cell(row=linha_globais + 11, column=1).value or 0
                                globais_pilar_especial['pilar1_gradea_posicao'] = sheet.cell(row=linha_globais + 12, column=1).value or 0
                                globais_pilar_especial['pilar1_gradeb_posicao'] = sheet.cell(row=linha_globais + 13, column=1).value or 0
                                globais_pilar_especial['pilar1_parafuso_posicao'] = sheet.cell(row=linha_globais + 14, column=1).value or 0
                                globais_pilar_especial['pilar1_metala_posicao'] = sheet.cell(row=linha_globais + 15, column=1).value or 0
                                globais_pilar_especial['pilar1_metalb_posicao'] = sheet.cell(row=linha_globais + 16, column=1).value or 0
                                
                                # PILAR 1 - Tamanhos
                                globais_pilar_especial['pilar1_paia_tamanho'] = sheet.cell(row=linha_globais + 20, column=1).value or 0
                                globais_pilar_especial['pilar1_paib_tamanho'] = sheet.cell(row=linha_globais + 21, column=1).value or 0
                                globais_pilar_especial['pilar1_gradea_tamanho'] = sheet.cell(row=linha_globais + 22, column=1).value or 0
                                globais_pilar_especial['pilar1_gradeb_tamanho'] = sheet.cell(row=linha_globais + 23, column=1).value or 0
                                globais_pilar_especial['pilar1_parafuso_tamanho'] = sheet.cell(row=linha_globais + 24, column=1).value or 0
                                globais_pilar_especial['pilar1_metala_tamanho'] = sheet.cell(row=linha_globais + 25, column=1).value or 0
                                globais_pilar_especial['pilar1_metalb_tamanho'] = sheet.cell(row=linha_globais + 26, column=1).value or 0
                                
                                # PILAR 2 - Posições
                                globais_pilar_especial['pilar2_paia_posicao'] = sheet.cell(row=linha_globais + 30, column=1).value or 0
                                globais_pilar_especial['pilar2_paib_posicao'] = sheet.cell(row=linha_globais + 31, column=1).value or 0
                                globais_pilar_especial['pilar2_gradea_posicao'] = sheet.cell(row=linha_globais + 32, column=1).value or 0
                                globais_pilar_especial['pilar2_gradeb_posicao'] = sheet.cell(row=linha_globais + 33, column=1).value or 0
                                globais_pilar_especial['pilar2_parafuso_posicao'] = sheet.cell(row=linha_globais + 34, column=1).value or 0
                                globais_pilar_especial['pilar2_metala_posicao'] = sheet.cell(row=linha_globais + 35, column=1).value or 0
                                globais_pilar_especial['pilar2_metalb_posicao'] = sheet.cell(row=linha_globais + 36, column=1).value or 0
                                
                                # PILAR 2 - Tamanhos
                                globais_pilar_especial['pilar2_paia_tamanho'] = sheet.cell(row=linha_globais + 40, column=1).value or 0
                                globais_pilar_especial['pilar2_paib_tamanho'] = sheet.cell(row=linha_globais + 41, column=1).value or 0
                                globais_pilar_especial['pilar2_gradea_tamanho'] = sheet.cell(row=linha_globais + 42, column=1).value or 0
                                globais_pilar_especial['pilar2_gradeb_tamanho'] = sheet.cell(row=linha_globais + 43, column=1).value or 0
                                globais_pilar_especial['pilar2_parafuso_tamanho'] = sheet.cell(row=linha_globais + 44, column=1).value or 0
                                globais_pilar_especial['pilar2_metala_tamanho'] = sheet.cell(row=linha_globais + 45, column=1).value or 0
                                globais_pilar_especial['pilar2_metalb_tamanho'] = sheet.cell(row=linha_globais + 46, column=1).value or 0
                                
                                # PERFIS METÁLICOS - Posições
                                globais_pilar_especial['perfil_metalico_a_posicao'] = sheet.cell(row=linha_globais + 50, column=1).value or 0
                                globais_pilar_especial['perfil_metalico_b_posicao'] = sheet.cell(row=linha_globais + 51, column=1).value or 0
                                
                                # PERFIS METÁLICOS - Tamanhos
                                globais_pilar_especial['perfil_metalico_a_tamanho'] = sheet.cell(row=linha_globais + 55, column=1).value or 0
                                globais_pilar_especial['perfil_metalico_b_tamanho'] = sheet.cell(row=linha_globais + 56, column=1).value or 0
                                
                                print(f"[OK] Globais lidas do Excel: {len(globais_pilar_especial)} valores")
                                print(f"   Exemplo - Pilar 1 PAI.A: pos={globais_pilar_especial.get('pilar1_paia_posicao')}, tam={globais_pilar_especial.get('pilar1_paia_tamanho')}")
                                
                                # DEFINIR GLOBAIS NO ROBÔ
                                try:
                                    # CORREÇÃO: Importação relativa correta para quando executado como script
                                    # (sys e os já importados no início da função)
                                    
                                    # Adicionar diretório dos robôs ao path se não estiver
                                    robots_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'robots')
                                    if robots_dir not in sys.path:
                                        sys.path.insert(0, robots_dir)
                                    
                                    # Tentar várias formas de importação
                                    try:
                                        definir_globais_pilar_especial = robo_module.definir_globais_pilar_especial
                                        print("[OK] Importação via robo_module bem-sucedida")
                                    except AttributeError:
                                        try:
                                            definir_globais_pilar_especial = robo_module.definir_globais_pilar_especial
                                            print("[OK] Importação via robo_module bem-sucedida")
                                        except ImportError:
                                            print("[ERRO] Todas as tentativas de importação falharam")
                                            raise
                                    
                                    sucesso = definir_globais_pilar_especial(globais_pilar_especial)
                                    if sucesso:
                                        print("[OK] Globais definidas no robô com sucesso!")
                                    else:
                                        print("[AVISO] Falha ao definir globais no robô")
                                        
                                except ImportError as ie:
                                    print(f"[ERRO] Não foi possível importar definir_globais_pilar_especial: {ie}")
                                except Exception as ge:
                                    print(f"[ERRO] Erro ao definir globais no robô: {ge}")
                                
                        except Exception as e:
                            print(f"[ERRO] Erro ao ler globais do pilar especial: {e}")
                        
                except Exception as e:
                    print(f"Erro ao ler dados do pilar especial do Excel: {e}")
                    import traceback
                    traceback.print_exc()

                # Criar diretórios base
                # CORREÇÃO: Usar robust_path_resolver para path correto em frozen e dev
                try:
                    from utils.robust_path_resolver import robust_path_resolver
                    diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
                except ImportError:
                    try:
                        from src.utils.robust_path_resolver import robust_path_resolver
                        diretorio_base = os.path.join(robust_path_resolver.get_project_root(), "output", "scripts")
                    except ImportError:
                        # Fallback manual (não recomendado, mas necessário em alguns casos)
                        diretorio_base = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "output", "scripts")
                pavimento = gerador.pavimento_entry.get().strip()
                if not pavimento:
                    pavimento = "SEM_PAVIMENTO"
                
                # CORREÇÃO: Usar normalização para remover acentos (mesma função usada em _dcad_pavimento_cima)
                import unicodedata
                def normalizar_nome_pasta_excel(texto):
                    """Normaliza o nome removendo acentos"""
                    if not texto:
                        return ""
                    texto = str(texto).strip()
                    if not texto:
                        return ""
                    try:
                        texto_normalizado = unicodedata.normalize('NFD', texto)
                        texto_sem_acentos = ''.join(char for char in texto_normalizado if unicodedata.category(char) != 'Mn')
                    except Exception:
                        texto_sem_acentos = texto
                    texto_sem_acentos = texto_sem_acentos.replace(' ', '_').replace('-', '_')
                    texto_limpo = ''.join(c if (c.isalnum() or c == '_') else '' for c in texto_sem_acentos)
                    while '__' in texto_limpo:
                        texto_limpo = texto_limpo.replace('__', '_')
                    texto_limpo = texto_limpo.strip('_')
                    return texto_limpo
                
                pavimento_normalizado = normalizar_nome_pasta_excel(pavimento)
                nome_pasta_pavimento = f"{pavimento_normalizado}_CIMA"
                diretorio_pavimento = os.path.join(diretorio_base, nome_pasta_pavimento)
                os.makedirs(diretorio_pavimento, exist_ok=True)
                print(f"Salvando na pasta do pavimento: {diretorio_pavimento} (normalizado de '{pavimento}')")

                scripts_gerados = []
                
                if pilar_especial_ativo and dimensoes_especiais:
                    print(f"Gerando {len(dimensoes_especiais)} scripts para pilar especial...")
                    
                    # Salvar dimensões originais
                    comp_original = gerador.comprimento_pilar_entry.get()
                    larg_original = gerador.largura_pilar_entry.get()
                    
                    for idx, dimensao in enumerate(dimensoes_especiais):
                        print(f"Gerando script {idx + 1}/{len(dimensoes_especiais)}...")
                        
                        # Definir configurações para cada script
                        if idx == 0:
                            # Primeiro script: ROTATE ativo
                            gerador.pilar_rotacionado_var.set(True)
                            gerador.usar_scale_var.set(False)
                            gerador.pilar_especial_ativo_var.set(False)  # Primeiro script não usa valores especiais
                            print(f"  Configuracao PILAR 1: ROTATE=True, SCALE=False, ESPECIAL=False")
                        elif idx == len(dimensoes_especiais) - 1:
                            # Último script: SCALE ativo com valores especiais
                            gerador.pilar_rotacionado_var.set(False)
                            gerador.usar_scale_var.set(True)
                            gerador.pilar_especial_ativo_var.set(True)  # Ativar valores especiais para SCALE
                            print(f"  Configuracao PILAR 2: ROTATE=False, SCALE=True, ESPECIAL=True")
                        else:
                            # Scripts intermediários: sem ROTATE nem SCALE
                            gerador.pilar_rotacionado_var.set(False)
                            gerador.usar_scale_var.set(False)
                            gerador.pilar_especial_ativo_var.set(False)
                            print(f"  Configuracao PILAR {idx + 1}: ROTATE=False, SCALE=False, ESPECIAL=False")
                        
                        # Alterar dimensões temporariamente
                        gerador.comprimento_pilar_entry.delete(0, 'end')
                        gerador.comprimento_pilar_entry.insert(0, dimensao['comprimento'])
                        gerador.largura_pilar_entry.delete(0, 'end')
                        gerador.largura_pilar_entry.insert(0, dimensao['largura'])
                        
                        print(f"  Usando dimensoes: comp={dimensao['comprimento']}, larg={dimensao['largura']}")
                        
                        # Verificar configuração atual
                        status_rotate = gerador.pilar_rotacionado_var.get()
                        status_scale = gerador.usar_scale_var.get()
                        print(f"  Configuracao: rotate={status_rotate}, scale={status_scale}")
                        
                        # DEFINIR GLOBAIS ANTES DE CADA GERAÇÃO DE SCRIPT
                        # SOLUÇÃO DEFINITIVA: Definir globais diretamente no objeto gerador
                        if globais_pilar_especial and len(globais_pilar_especial) > 0:
                            try:
                                print(f"  [SOLUCAO-DEFINITIVA] Processo atual PID: {os.getpid()}")
                                print(f"  [SOLUCAO-DEFINITIVA] Globais a definir: {len(globais_pilar_especial)} itens")
                                
                                # SOLUÇÃO DEFINITIVA: Definir globais diretamente no objeto gerador
                                print(f"  [SOLUCAO-DEFINITIVA] Gerador ID: {id(gerador)}")
                                print(f"  [SOLUCAO-DEFINITIVA] Gerador tipo: {type(gerador)}")
                                
                                # Definir globais diretamente no objeto gerador
                                gerador._globais_pilar_especial = globais_pilar_especial
                                print(f"  [SOLUCAO-DEFINITIVA] Globais definidas diretamente no gerador: {len(gerador._globais_pilar_especial)} valores")
                                
                                # Verificar se funcionou
                                if hasattr(gerador, '_globais_pilar_especial') and gerador._globais_pilar_especial:
                                    print(f"  [OK] Globais definidas no gerador para script {idx + 1}: {len(gerador._globais_pilar_especial)} valores")
                                else:
                                    print(f"  [ERRO] Falha na definição no gerador")
                                
                                # TAMBÉM definir na importação direta como backup
                                robo_module._globais_pilar_especial = globais_pilar_especial
                                print(f"  [SOLUCAO-DEFINITIVA] Backup definido também")
                                
                            except Exception as e:
                                print(f"  [ERRO] Erro ao definir globais para script {idx + 1}: {e}")
                                import traceback
                                traceback.print_exc()
                        
                        # Gerar script
                        script_cima = gerador.gerar_script()
                        
                        # APLICAR OFFSETS DOS METAIS PARA PILARES ESPECIAIS
                        script_cima = aplicar_offsets_metais_ao_script(script_cima, gerador, globais_pilar_especial)
                        
                        # Para o segundo e terceiro script, modificar X inicial para -196
                        if idx > 0:  # Não é o primeiro script
                            print(f"  Modificando X inicial para -196 (script {idx + 1})")
                            script_cima = modificar_x_inicial_script(script_cima, -196)
                        
                        # Requisito: no SEGUNDO script (pilar 2) descer em Y usando fórmula: -comp_1 + larg_2
                        if idx == 1 and len(dimensoes_especiais) >= 2:
                            try:
                                comprimento_primeiro = float(dimensoes_especiais[0]['comprimento'])
                                largura_segundo = float(dimensoes_especiais[1]['largura'])
                                offset_y = -comprimento_primeiro + largura_segundo
                                print(f"  Modificando Y inicial para {offset_y} (segundo script: -comp_1 + larg_2 = -{comprimento_primeiro} + {largura_segundo})")
                                script_cima = modificar_y_inicial_script(script_cima, offset_y)
                                # Garantir ajuste explícito do _ZOOM inicial (base 70 + offset)
                                y_base_esperado = 70 + offset_y
                                script_cima = ajustar_zoom_inicial_y(script_cima, y_base_esperado)
                                # Depuração: extrair Y amostras após modificação
                                try:
                                    import re
                                    coords = re.findall(r"\nC\s*(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)", script_cima)
                                    if coords:
                                        ys = sorted({float(y) for _, y in coords})
                                        print(f"  [DEBUG] Y únicos (head) PARTE2 após offset: {ys[:10]}")
                                        # Verificar se o y_base_esperado está presente
                                        hit = any(abs(float(y)-y_base_esperado) < 1e-6 for y in ys)
                                        print(f"  [DEBUG] Encontrou y_base_esperado={y_base_esperado}? {hit}")
                                    # Mostrar algumas linhas-chave de INSERTs
                                    for chave in ("PAR_BAIXO2", "PAR_CIMA2"):
                                        m = re.findall(rf"-INSERT\s+{chave}\s+(-?\\d+(?:\\.\\d+)?),(-?\\d+(?:\\.\\d+)?)", script_cima)
                                        if m:
                                            print(f"  [DEBUG] {chave} coords (amostra): {m[:2]}")
                                except Exception as __e:
                                    print(f"  [DEBUG] Falha ao coletar amostras de coordenadas: {__e}")
                            except Exception as _e:
                                print(f"  [AVISO] Falha ao ajustar Y do segundo script: {_e}")
                        
                        # Para o terceiro script, descer em Y usando fórmula: -comp_1 + larg_3
                        if idx == 2 and len(dimensoes_especiais) >= 3:  # É o terceiro script
                            comprimento_primeiro = float(dimensoes_especiais[0]['comprimento'])
                            largura_terceiro = float(dimensoes_especiais[2]['largura'])
                            offset_y = -comprimento_primeiro + largura_terceiro
                            print(f"  Modificando Y inicial para {offset_y} (terceiro script: -comp_1 + larg_3 = -{comprimento_primeiro} + {largura_terceiro})")
                            script_cima = modificar_y_inicial_script(script_cima, offset_y)
                        
                        # Verificar se o script é válido antes de salvar
                        if script_cima is None:
                            print(f"  ERRO: Script gerado é None para script {idx + 1}")
                            script_cima = ""
                        elif not isinstance(script_cima, str):
                            print(f"  ERRO: Script gerado não é string, é {type(script_cima)}")
                            script_cima = str(script_cima) if script_cima is not None else ""
                        
                        # Salvar script
                        nome_arquivo_base = os.path.join(diretorio_pavimento, str(nome_pilar))
                        if idx == 0:
                            nome_arquivo_cima = f"{nome_arquivo_base}_CIMA.scr"
                        else:
                            nome_arquivo_cima = f"{nome_arquivo_base}_CIMA-{idx + 1}.scr"
                        
                        contador = 1
                        while os.path.exists(nome_arquivo_cima):
                            nome_arquivo_cima = f"{nome_arquivo_base}_CIMA-{idx + 1}-{contador}.scr"
                            contador += 1
                        
                        with open(nome_arquivo_cima, "w", encoding="utf-16") as f:
                            f.write(script_cima)
                        
                        scripts_gerados.append(nome_arquivo_cima)
                        print(f"  Script salvo: {nome_arquivo_cima}")
                    
                    # Restaurar dimensões originais e configurações
                    gerador.comprimento_pilar_entry.delete(0, 'end')
                    gerador.comprimento_pilar_entry.insert(0, comp_original)
                    gerador.largura_pilar_entry.delete(0, 'end')
                    gerador.largura_pilar_entry.insert(0, larg_original)
                    
                    # Restaurar configurações padrão
                    gerador.pilar_rotacionado_var.set(False)
                    gerador.usar_scale_var.set(True)
                    gerador.pilar_especial_ativo_var.set(False)
                    print("Configuracoes restauradas para o padrao")
                    
                    print(f"Scripts multiplos gerados para pilar especial '{nome_pilar}':")
                    for script in scripts_gerados:
                        print(f"  - {script}")
                    
                    # Criar script combinado para execução única no CAD
                    print("Criando script combinado para execucao unica no CAD...")
                    nome_arquivo_combinado = os.path.join(diretorio_pavimento, f"{nome_pilar}_COMBINADO_CIMA.scr")
                    
                    with open(nome_arquivo_combinado, "w", encoding="utf-16") as f_combinado:
                        for script_path in scripts_gerados:
                            with open(script_path, "r", encoding="utf-16") as f_individual:
                                conteudo = f_individual.read()
                                f_combinado.write(conteudo)
                                f_combinado.write("\n")  # Separador entre scripts
                    
                    print(f"Script combinado criado: {nome_arquivo_combinado}")
                        
                else:
                    # Gerar script normal (único)
                    print("Gerando script normal (unico)...")
                    
                    # DEFINIR GLOBAIS TAMBÉM PARA CASOS NORMAIS
                    if 'globais_pilar_especial' in locals() and globais_pilar_especial and len(globais_pilar_especial) > 0:
                        try:
                            print(f"  [CASO-NORMAL] Definindo globais para script normal...")
                            print(f"  [CASO-NORMAL] Globais a definir: {len(globais_pilar_especial)} itens")
                            
                            # Definir globais diretamente no objeto gerador
                            gerador._globais_pilar_especial = globais_pilar_especial
                            print(f"  [CASO-NORMAL] Globais definidas diretamente no gerador: {len(gerador._globais_pilar_especial)} valores")
                            
                            # Verificar se funcionou
                            if hasattr(gerador, '_globais_pilar_especial') and gerador._globais_pilar_especial:
                                print(f"  [OK] Globais definidas no gerador para script normal: {len(gerador._globais_pilar_especial)} valores")
                            else:
                                print(f"  [ERRO] Falha na definição no gerador para script normal")
                            
                            # TAMBÉM definir na importação direta como backup
                            robo_module._globais_pilar_especial = globais_pilar_especial
                            print(f"  [CASO-NORMAL] Backup definido também")
                            
                        except Exception as e:
                            print(f"  [ERRO] Erro ao definir globais para script normal: {e}")
                            import traceback
                            traceback.print_exc()
                    
                    script_cima = gerador.gerar_script()
                    
                    # Verificar se o script é válido antes de salvar
                    if script_cima is None:
                        print(f"  ERRO: Script gerado é None para pilar {nome_pilar}")
                        script_cima = ""
                    elif not isinstance(script_cima, str):
                        print(f"  ERRO: Script gerado não é string, é {type(script_cima)}")
                        script_cima = str(script_cima) if script_cima is not None else ""
                    
                    nome_arquivo_base = os.path.join(diretorio_pavimento, str(nome_pilar))
                    nome_arquivo_cima = f"{nome_arquivo_base}_CIMA.scr"
                    
                    # Sobrescrever se existir (sem contador)
                    
                    with open(nome_arquivo_cima, "w", encoding="utf-16") as f:
                        f.write(script_cima)
                    
                    scripts_gerados.append(nome_arquivo_cima)
                    print(f"Script VISAO CIMA gerado para o pilar '{nome_pilar}' no pavimento '{pavimento}':")
                    print(f"  - {nome_arquivo_cima}")
                
                # CORREÇÃO: NÃO destruir a aplicação aqui - apenas no final
                # app.root.destroy()  # REMOVIDO - será destruído no final

            else:
                print(f"Pulando coluna {coluna}: nome do pilar em branco.")
                colunas_vazias += 1
                if colunas_vazias >= 10:
                    print(f"Parando a busca após 10 colunas vazias consecutivas a partir da coluna {coluna}.")
                    parar_busca = True

        # CORREÇÃO: Destruir a aplicação apenas no final, após processar todas as colunas
        print("[DEBUG-FUNC] Finalizando aplicação VISAO CIMA...")
        app.root.destroy()
        
        timestamp_func_fim = time.time()
        duracao_func = timestamp_func_fim - timestamp_func_inicio
        
        print(f"[CIMA_SCRIPT] === PROCESSAMENTO CONCLUÍDO ===")
        print(f"[CIMA_SCRIPT] Total de colunas processadas: {len(colunas_processadas)}")
        print(f"[CIMA_SCRIPT] Colunas processadas: {sorted(colunas_processadas)}")
        print(f"[CIMA_SCRIPT] Duração da função: {duracao_func:.2f} segundos")
        print(f"[CIMA_SCRIPT] Fim da função preencher_campos_e_gerar_scripts: {timestamp_func_fim}")
        print(f"[CIMA_SCRIPT] ===== FIM DA FUNÇÃO preencher_campos_e_gerar_scripts =====")
        
        print("Fim do processamento da planilha.")

    except Exception as e:
        print(f"[ERROR-FUNC] Erro ao processar a planilha: {e}")
        import traceback
        traceback.print_exc()

# Exemplo de uso:
if __name__ == "__main__":
    print(f"[CIMA_SCRIPT] ===== INÍCIO DA EXECUÇÃO =====")
    print(f"[CIMA_SCRIPT] Timestamp: {time.time()}")
    print(f"[CIMA_SCRIPT] PID: {os.getpid()}")
    print(f"[CIMA_SCRIPT] Argumentos: {sys.argv}")
    print(f"[CIMA_SCRIPT] Diretório atual: {os.getcwd()}")
    print(f"[CIMA_SCRIPT] __file__: {__file__}")
    print(f"[CIMA_SCRIPT] Robos dir: {robos_dir}")
    print(f"[CIMA_SCRIPT] Robos dir existe: {os.path.exists(robos_dir)}")
    
    parser = argparse.ArgumentParser(description="Automação CIMA - Geração de Scripts SCR")
    parser.add_argument("excel", nargs="?", help="Caminho do arquivo Excel")
    parser.add_argument("colunas", nargs="?", help="Colunas a serem processadas (ex: E,F,G)")
    parser.add_argument("pavimento", nargs="?", help="Pavimento a ser processado (opcional)")
    args = parser.parse_args()

    if args.excel:
        caminho_arquivo_excel = args.excel
        colunas = args.colunas.split(",") if args.colunas else None
        pavimento = args.pavimento
        print(f"[CIMA_SCRIPT] Processando arquivo: {caminho_arquivo_excel}")
        print(f"[CIMA_SCRIPT] Colunas: {colunas}")
        print(f"[CIMA_SCRIPT] Pavimento: {pavimento}")
        
        # ADICIONAR LOG DE DEPURAÇÃO PARA IDENTIFICAR DUPLICAÇÃO
        import time
        timestamp_inicio = time.time()
        print(f"[DEBUG] Início da execução: {timestamp_inicio}")
        print(f"[DEBUG] PID do processo: {os.getpid()}")
        print(f"[DEBUG] Argumentos recebidos: excel={caminho_arquivo_excel}, colunas={colunas}, pavimento={pavimento}")
        
        # VERIFICAÇÃO DE DUPLICAÇÃO: Criar um arquivo de lock para evitar execuções simultâneas
        lock_file = f"cima_lock_{os.getpid()}_{timestamp_inicio}.tmp"
        lock_path = os.path.join(os.path.dirname(caminho_arquivo_excel), lock_file)
        
        try:
            # Tentar criar arquivo de lock
            with open(lock_path, 'w') as f:
                f.write(f"PID: {os.getpid()}\nTimestamp: {timestamp_inicio}\n")
            print(f"[DEBUG] Arquivo de lock criado: {lock_path}")
            
            # CORREÇÃO: Processar todas as colunas de uma vez, não individualmente
            if colunas:
                # Se múltiplas colunas foram fornecidas, processar todas de uma vez
                # passando a string completa de colunas para evitar duplicação
                colunas_str = ','.join(colunas)
                print(f"[DEBUG] Processando múltiplas colunas: {colunas_str}")
                preencher_campos_e_gerar_scripts(caminho_arquivo_excel, coluna_especifica=colunas_str)
            else:
                # Se nenhuma coluna específica foi fornecida, processar todas as colunas
                print(f"[DEBUG] Processando todas as colunas")
                preencher_campos_e_gerar_scripts(caminho_arquivo_excel)
            
            timestamp_fim = time.time()
            duracao = timestamp_fim - timestamp_inicio
            print(f"[CIMA_SCRIPT] Fim da execução: {timestamp_fim}")
            print(f"[CIMA_SCRIPT] Duração total: {duracao:.2f} segundos")
            print(f"[CIMA_SCRIPT] ===== FIM DA EXECUÇÃO =====")
            print("Processamento concluído com sucesso!")
            
        except Exception as e:
            print(f"[ERROR] Erro durante execução: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # Remover arquivo de lock
            try:
                if os.path.exists(lock_path):
                    os.remove(lock_path)
                    print(f"[DEBUG] Arquivo de lock removido: {lock_path}")
            except Exception as e:
                print(f"[WARN] Erro ao remover arquivo de lock: {e}")
    else:
        # Modo antigo: interface gráfica
        pass  # main() removido para evitar erro de variável indefinida