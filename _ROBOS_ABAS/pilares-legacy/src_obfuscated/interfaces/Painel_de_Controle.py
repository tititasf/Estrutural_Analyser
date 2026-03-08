
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

import openpyxl
import os
import time
import pyautogui
import threading
import queue
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import keyboard
import sys
import subprocess
from openpyxl.utils import get_column_letter
import shutil
from glob import glob

# IMPORTAR HELPER FROZEN GLOBAL - garante que paths estão configurados
try:
    from _frozen_helper import ensure_paths
    ensure_paths()
except ImportError:
    try:
        from src._frozen_helper import ensure_paths
        ensure_paths()
    except ImportError:
        # Fallback manual
        if getattr(sys, 'frozen', False):
            script_dir = os.path.dirname(sys.executable)
            if script_dir not in sys.path:
                sys.path.insert(0, script_dir)
            src_dir = os.path.join(script_dir, 'src')
            if os.path.exists(src_dir) and src_dir not in sys.path:
                sys.path.insert(0, src_dir)
        else:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            parent_dir = os.path.dirname(current_dir)
            if parent_dir not in sys.path:
                sys.path.insert(0, parent_dir)

# Imports relativos para a estrutura modularizada - com múltiplos fallbacks
try:
    from ..utils.funcoes_auxiliares import *
except ImportError:
    try:
        from utils.funcoes_auxiliares import *
    except ImportError:
        try:
            from src.utils.funcoes_auxiliares import *
        except ImportError:
            pass

try:
    from ..core.pilar_analyzer import *
except ImportError:
    try:
        from core.pilar_analyzer import *
    except ImportError:
        try:
            from src.core.pilar_analyzer import *
        except ImportError:
            pass


# Constantes para as coordenadas dos botões
COORD_BOTAO_CALCULAR = (68, 534)
COORD_BOTAO_GERAR_SCRIPT = (64, 593)
COORD_BOTAO_LIMPAR = (56, 560)
COORD_INICIO_CAMPOS = (116, 80)

def clicar_botao(coordenadas):
    """Move o mouse até as coordenadas e simula um clique instantâneo."""
    try:
        pyautogui.PAUSE = 0.05  # Reduz ainda mais o delay entre comandos
        pyautogui.click(coordenadas[0], coordenadas[1])  # Click direto nas coordenadas
        print(f"  Botão nas coordenadas {coordenadas} clicado.")
    except Exception as e:
        print(f"  Erro ao clicar no botão nas coordenadas {coordenadas}: {e}")

def formatar_valor(valor):
    """
    Formata o valor da célula, substituindo vírgulas por pontos.
    """
    if valor is None:
        return None
    
    # Se for texto, retorna diretamente
    if isinstance(valor, str):
        # Remove espaços extras e verifica se não está vazio
        valor_limpo = valor.strip()
        # Substitui vírgula por ponto
        valor_limpo = valor_limpo.replace(',', '.')
        return valor_limpo if valor_limpo else None
    
    try:
        # Tenta converter para float (para números e resultados de fórmulas)
        num = float(str(valor).replace(',', '.'))
        
        # Se for um número inteiro, retorna sem decimais
        if num.is_integer():
            return str(int(num))
        
        # Se for decimal, formata com no máximo 2 casas e usa ponto
        return f"{num:.2f}"  # Removido o replace para manter o ponto
        
    except (ValueError, TypeError, AttributeError):
        # Se não for possível converter para número, tenta retornar como texto
        try:
            texto = str(valor).strip()
            # Substitui vírgula por ponto mesmo em texto
            texto = texto.replace(',', '.')
            return texto if texto else None
        except:
            return None

def digitar_valor(valor):
    """Digita um valor usando pyautogui com velocidade máxima, usando ponto ao invés de vírgula."""
    if valor is not None:
        valor_formatado = formatar_valor(valor)
        if valor_formatado is not None:
            valor_para_digitar = str(valor_formatado).replace(',', '.')
            print(f"Digitando valor: {valor_para_digitar}")
            pyautogui.typewrite(valor_para_digitar, interval=0.001)
            time.sleep(0.02)

def preencher_campos_e_gerar_scripts(caminho_arquivo_excel):
    """Versão ultra otimizada da função principal"""
    try:
        # Configurações extremas do pyautogui para velocidade máxima
        pyautogui.PAUSE = 0.05  # Delay mínimo entre comandos
        pyautogui.MINIMUM_DURATION = 0
        pyautogui.MINIMUM_SLEEP = 0
        
        # Otimização para pressionar tab múltiplas vezes
        def multi_tab(count):
            if count > 0:
                print(f"Pressionando TAB {count} vezes")
                pyautogui.write(['tab'] * count, interval=0.001)
                for _ in range(count):
                    print("TAB PRESSIONADO")

        # Função otimizada para backspace múltiplo
        def multi_backspace(count):
            if count > 0:
                print(f"Pressionando BACKSPACE {count} vezes")
                pyautogui.write(['backspace'] * count, interval=0.001)
                for _ in range(count):
                    print("BACKSPACE PRESSIONADO")

        # Carrega a planilha com data_only=True para obter valores calculados ao invés de fórmulas
        workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        sheet = workbook.worksheets[0]

        # Mapeamento de linhas para campos da interface VISAO ABCD
        linhas_abcd = {
            'nome': 5,
            'pavimento': 4,
            'nivel_saida': 8,
            'nivel_chegada': 9,
            'comprimento': 6,
            'largura': 7,
            'altura': 12    ,
            'laje_a': 13,
            'laje_b': 55,
            'laje_c': 97,
            'laje_d': 131,
            'posicao_laje_a': 14,
            'posicao_laje_b': 56,
            'posicao_laje_c': 98,
            'posicao_laje_d': 132,
            'parafusos_p1_p2': 165,
            'parafusos_p2_p3': 166,
            'parafusos_p3_p4': 167,
            'parafusos_p4_p5': 168,
            'parafusos_p5_p6': 169,
            'parafusos_p6_p7': 170,
            'parafusos_p7_p8': 171,
            'largura_1_painel_a': 15,
            'largura_2_painel_a': 16,
            'largura_3_painel_a': 17,
            'altura_5_painel_a': 22,
            'altura_4_painel_a': 21,
            'altura_3_painel_a': 20,
            'altura_2_painel_a': 19,
            'altura_1_painel_a': 18,
            'abertura_esq1_dist_a': 25,
            'abertura_esq1_prof_a': 26,
            'abertura_esq1_larg_a': 27,
            'abertura_esq1_pos_a': 28,
            'abertura_esq2_dist_a': 29,
            'abertura_esq2_prof_a': 30,
            'abertura_esq2_larg_a': 31,
            'abertura_esq2_pos_a': 32,
            'abertura_dir1_dist_a': 33,
            'abertura_dir1_prof_a': 34,
            'abertura_dir1_larg_a': 35,
            'abertura_dir1_pos_a': 36,
            'abertura_dir2_dist_a': 37,
            'abertura_dir2_prof_a': 38,
            'abertura_dir2_larg_a': 39,
            'abertura_dir2_pos_a': 40,
            'hatch_l1_h2_a': 42,
            'hatch_l1_h3_a': 43,
            'hatch_l1_h4_a': 44,
            'hatch_l1_h5_a': 45,
            'hatch_l2_h2_a': 46,
            'hatch_l2_h3_a': 47,
            'hatch_l2_h4_a': 48,
            'hatch_l2_h5_a': 49,
            'hatch_l3_h2_a': 50,
            'hatch_l3_h3_a': 51,
            'hatch_l3_h4_a': 52,
            'hatch_l3_h5_a': 53,
            'largura_1_painel_b': 57,
            'largura_2_painel_b': 58,
            'largura_3_painel_b': 59,
            'altura_5_painel_b': 64,
            'altura_4_painel_b': 63,
            'altura_3_painel_b': 62,
            'altura_2_painel_b': 61,
            'altura_1_painel_b': 60,
            'abertura_esq1_dist_b': 67,
            'abertura_esq1_prof_b': 68,
            'abertura_esq1_larg_b': 69,
            'abertura_esq1_pos_b': 70,
            'abertura_esq2_dist_b': 71,
            'abertura_esq2_prof_b': 72,
            'abertura_esq2_larg_b': 73,
            'abertura_esq2_pos_b': 74,
            'abertura_dir1_dist_b': 75,
            'abertura_dir1_prof_b': 76,
            'abertura_dir1_larg_b': 77,
            'abertura_dir1_pos_b': 78,
            'abertura_dir2_dist_b': 79,
            'abertura_dir2_prof_b': 80,
            'abertura_dir2_larg_b': 81,
            'abertura_dir2_pos_b': 82,
            'hatch_l1_h2_b': 84,
            'hatch_l1_h3_b': 85,
            'hatch_l1_h4_b': 86,
            'hatch_l1_h5_b': 87,
            'hatch_l2_h2_b': 88,
            'hatch_l2_h3_b': 89,
            'hatch_l2_h4_b': 90,
            'hatch_l2_h5_b': 91,
            'hatch_l3_h2_b': 92,
            'hatch_l3_h3_b': 93,
            'hatch_l3_h4_b': 94,
            'hatch_l3_h5_b': 95,
            'largura_1_painel_c': 98,
            'largura_2_painel_c': 99,
            'altura_4_painel_c': 104,
            'altura_3_painel_c': 103,
            'altura_2_painel_c': 102,
            'altura_1_painel_c': 101,            
            'abertura_esq1_dist_c': 107,
            'abertura_esq1_prof_c': 108,
            'abertura_esq1_larg_c': 109,
            'abertura_esq1_pos_c': 110,
            'abertura_esq2_dist_c': 111,
            'abertura_esq2_prof_c': 112,
            'abertura_esq2_larg_c': 113,
            'abertura_esq2_pos_c': 114,
            'abertura_dir1_dist_c': 115,
            'abertura_dir1_prof_c': 116,
            'abertura_dir1_larg_c': 117,
            'abertura_dir1_pos_c': 118,
            'abertura_dir2_dist_c': 119,
            'abertura_dir2_prof_c': 120,
            'abertura_dir2_larg_c': 121,
            'abertura_dir2_pos_c': 122,
            'hatch_l1_h2_c': 124,
            'hatch_l1_h3_c': 125,
            'hatch_l1_h4_c': 126,
            'hatch_l2_h2_c': 127,
            'hatch_l2_h3_c': 128,
            'hatch_l2_h4_c': 129,
            'largura_1_painel_d': 133,
            'largura_2_painel_d': 134,
            'altura_4_painel_d': 138,
            'altura_3_painel_d': 137,
            'altura_2_painel_d': 136,
            'altura_1_painel_d': 135,            
            'abertura_esq1_dist_d': 141,
            'abertura_esq1_prof_d': 142,
            'abertura_esq1_larg_d': 143,
            'abertura_esq1_pos_d': 144,
            'abertura_esq2_dist_d': 145,
            'abertura_esq2_prof_d': 146,
            'abertura_esq2_larg_d': 147,
            'abertura_esq2_pos_d': 148,
            'abertura_dir1_dist_d': 149,
            'abertura_dir1_prof_d': 150,
            'abertura_dir1_larg_d': 151,
            'abertura_dir1_pos_d': 152,
            'abertura_dir2_dist_d': 153,
            'abertura_dir2_prof_d': 154,
            'abertura_dir2_larg_d': 155,
            'abertura_dir2_pos_d': 156,
            'hatch_l1_h2_d': 158,
            'hatch_l1_h3_d': 159,
            'hatch_l1_h4_d': 160,
            'hatch_l2_h2_d': 161,
            'hatch_l2_h3_d': 162,
            'hatch_l2_h4_d': 163,
        }

        coluna_base = 'E'
        coluna_base_idx = ord(coluna_base) - ord('A')
        colunas_vazias = 0
        parar_busca = False
        parar_automacao = False

        def monitorar_esc():
            """Monitora a tecla Esc e define a variável parar_automacao."""
            nonlocal parar_automacao
            while not parar_automacao:
                if keyboard.is_pressed('esc'):
                    parar_automacao = True
                    print("Tecla Esc pressionada. Parando automação...")
                time.sleep(0.1)  # Pequeno intervalo para evitar uso excessivo da CPU

        thread_monitor_esc = threading.Thread(target=monitorar_esc)
        thread_monitor_esc.daemon = True
        thread_monitor_esc.start()

        def automacao_thread():
            nonlocal parar_busca, colunas_vazias
            for coluna_idx in range(coluna_base_idx, sheet.max_column):
                if parar_busca or parar_automacao:
                    break

                coluna = chr(coluna_idx + ord('A'))
                
                # Verifica se há dados na próxima coluna usando os valores formatados
                proxima_coluna_tem_dados = any(
                    formatar_valor(sheet.cell(row=linhas_abcd[campo], column=coluna_idx + 2).value) is not None
                    for campo in ['nome', 'pavimento', 'nivel_saida', 'nivel_chegada', 'altura']
                )

                # Verifica se há dados na coluna atual usando os valores formatados
                if any(
                    formatar_valor(sheet.cell(row=linhas_abcd[campo], column=coluna_idx + 1).value) is not None
                    for campo in ['nome', 'pavimento', 'nivel_saida', 'nivel_chegada', 'altura']
                ):
                    colunas_vazias = 0

                    # --- Preenchimento dos campos: Parte 1 ---
                    clicar_botao(COORD_INICIO_CAMPOS)
                    sequencia_campos_parte_1 = [
                        ('nome', 4), ('tab',),
                        ('pavimento', 4), ('tab',),
                        ('nivel_saida', 8), ('tab',),
                        ('nivel_chegada', 9), ('tab',),
                        ('comprimento', 6), ('tab',),
                        ('largura', 7), ('tab',),
                        ('altura', 12), ('tab',),
                        ('laje_a', 13), ('tab',),
                        ('posicao_laje_a', 14, 'ref_1A'), ('tab',),
                        ('posicao_laje_a', 14, 'ref_2A'), ('tab',),
                        ('posicao_laje_a', 14, 'ref_3A'), ('tab',),
                        ('posicao_laje_a', 14, 'ref_4A'), ('tab',),
                        ('posicao_laje_a', 14, 'ref_5A'), ('tab',),
                        ('posicao_laje_a', 14, 'ref_6A'), ('tab',),
                        ('laje_b', 55), ('tab',),
                        ('posicao_laje_b', 56, 'ref_1B'), ('tab',),
                        ('posicao_laje_b', 56, 'ref_2B'), ('tab',),
                        ('posicao_laje_b', 56, 'ref_3B'), ('tab',),
                        ('posicao_laje_b', 56, 'ref_4B'), ('tab',),
                        ('posicao_laje_b', 56, 'ref_5B'), ('tab',),
                        ('posicao_laje_b', 56, 'ref_6B'), ('tab',),
                        ('laje_c', 97), ('tab',),
                        ('posicao_laje_c', 98, 'ref_1C'), ('tab',),
                        ('posicao_laje_c', 98, 'ref_2C'), ('tab',),
                        ('posicao_laje_c', 98, 'ref_3C'), ('tab',),
                        ('posicao_laje_c', 98, 'ref_4C'), ('tab',),
                        ('posicao_laje_c', 98, 'ref_5C'), ('tab',),
                        ('laje_d', 131), ('tab',),
                        ('posicao_laje_d', 132, 'ref_1D'), ('tab',),
                        ('posicao_laje_d', 132, 'ref_2D'), ('tab',),
                        ('posicao_laje_d', 132, 'ref_3D'), ('tab',),
                        ('posicao_laje_d', 132, 'ref_4D'), ('tab',),
                        ('posicao_laje_d', 132, 'ref_5D'), ('tab',),
                        ('parafusos_p1_p2', 165), ('tab',), ('tab',),
                        ('parafusos_p2_p3', 166), ('tab',), ('tab',),
                        ('parafusos_p3_p4', 167), ('tab',), ('tab',),
                        ('parafusos_p4_p5', 168), ('tab',), ('tab',),
                        ('parafusos_p5_p6', 169), ('tab',), ('tab',),
                        ('parafusos_p6_p7', 170), ('tab',), ('tab',),
                        ('parafusos_p7_p8', 171), ('tab',), ('tab',)
                    ]

                    for item in sequencia_campos_parte_1:
                        if parar_automacao:
                            break
                        if item[0] == 'tab':
                            pyautogui.press('tab')
                            print("TAB PRESSIONADO")
                        elif len(item) == 3:
                            campo, linha, ref = item
                            valor = formatar_valor(sheet[f'{coluna}{linha}'].value)
                            print(f"Campo: {campo} | Linha: {linha} | Valor: {valor}")
                            if ref == 'ref_1A' and valor == 0:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_2A' and valor == 1:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_3A' and valor == 2:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_4A' and valor == 3:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_5A' and valor == 4:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_6A' and valor == 5:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_1B' and valor == 0:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_2B' and valor == 1:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_3B' and valor == 2:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_4B' and valor == 3:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_5B' and valor == 4:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_6B' and valor == 5:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_1C' and valor == 0:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_2C' and valor == 1:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_3C' and valor == 2:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_4C' and valor == 3:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_5C' and valor == 4:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_1D' and valor == 0:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_2D' and valor == 1:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_3D' and valor == 2:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_4D' and valor == 3:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                            elif ref == 'ref_5D' and valor == 4:
                                pyautogui.press('space')
                                print("ESPAÇO PRESSIONADO")
                        else:
                            campo, linha = item
                            valor = formatar_valor(sheet[f'{coluna}{linha}'].value)
                            print(f"Campo: {campo} | Linha: {linha} | Valor: {valor}")
                            digitar_valor(valor)

                    if parar_automacao:
                        break

                    # Clicar no botão "Calcular"
                    clicar_botao(COORD_BOTAO_CALCULAR)
                    time.sleep(0.5)

                    if parar_automacao:
                        break

                    # --- Preenchimento dos campos: Parte 2 (Refatorado e Dividido) ---
                    clicar_botao((700, 110))
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    pyautogui.press('backspace')
                    
                    
                    # Parte 2A - Versão Completa com Aberturas Integradas
                    sequencia_campos_parte_2A = [
                        # Painel A
                        ('largura_1_painel_a', 15), ('tab',), ('tab',),
                        ('largura_2_painel_a', 16), ('tab',), ('tab',),
                        ('largura_3_painel_a', 17), ('tab',), ('tab',),
                        ('altura_5_painel_a', 22), ('tab',), ('tab',), ('hatch_l1_h5_a', 45), ('tab',), ('hatch_l2_h5_a', 49), ('tab',), ('hatch_l3_h5_a', 53), ('tab',),
                        ('altura_4_painel_a', 21), ('tab',), ('tab',), ('hatch_l1_h4_a', 44), ('tab',), ('hatch_l2_h4_a', 48), ('tab',), ('hatch_l3_h4_a', 52), ('tab',),
                        ('altura_3_painel_a', 20), ('tab',), ('tab',), ('hatch_l1_h3_a', 43), ('tab',), ('hatch_l2_h3_a', 47), ('tab',), ('hatch_l3_h3_a', 51), ('tab',),
                        ('altura_2_painel_a', 19), ('tab',), ('tab',), ('hatch_l1_h2_a', 42), ('tab',), ('hatch_l2_h2_a', 46), ('tab',), ('hatch_l3_h2_a', 50), ('tab',),
                        ('altura_1_painel_a', 18), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',),

                        # Aberturas Painel A (Modificado para 0 em negativo)
                        ('abertura_esq1_pos_a', 28, 2, 'space', '0'), ('tab',),('abertura_esq1_pos_a', 28, 1, 'space', '0'), ('tab',),
                        ('abertura_esq1_dist_a', 25), ('tab',), ('abertura_esq1_prof_a', 26), ('tab',), ('abertura_esq1_larg_a', 27), ('tab',),
                        ('abertura_esq2_pos_a', 32, 2, 'space', '0'), ('tab',),('abertura_esq2_pos_a', 32, 1, 'space', '0'), ('tab',),
                        ('abertura_esq2_dist_a', 29), ('tab',), ('abertura_esq2_prof_a', 30), ('tab',), ('abertura_esq2_larg_a', 31), ('tab',),
                        ('abertura_dir1_pos_a', 36, 2, 'space', '0'), ('tab',),('abertura_dir1_pos_a', 36, 1, 'space', '0'), ('tab',),
                        ('abertura_dir1_dist_a', 33), ('tab',), ('abertura_dir1_prof_a', 34), ('tab',), ('abertura_dir1_larg_a', 35), ('tab',),
                        ('abertura_dir2_pos_a', 40, 2, 'space', '0'), ('tab',),('abertura_dir2_pos_a', 40, 1, 'space', '0'), ('tab',),
                        ('abertura_dir2_dist_a', 37), ('tab',), ('abertura_dir2_prof_a', 38), ('tab',), ('abertura_dir2_larg_a', 39), ('tab',), ('tab',), ('tab',),

                        # Painel B
                        ('largura_1_painel_b', 57), ('tab',), ('tab',),
                        ('largura_2_painel_b', 58), ('tab',), ('tab',),
                        ('largura_3_painel_b', 59), ('tab',), ('tab',),
                        ('altura_5_painel_b', 64), ('tab',), ('tab',), ('hatch_l1_h5_b', 87), ('tab',), ('hatch_l2_h5_b', 91), ('tab',), ('hatch_l3_h5_b', 95), ('tab',),
                        ('altura_4_painel_b', 63), ('tab',), ('tab',), ('hatch_l1_h4_b', 86), ('tab',), ('hatch_l2_h4_b', 90), ('tab',), ('hatch_l3_h4_b', 94), ('tab',),
                        ('altura_3_painel_b', 62), ('tab',), ('tab',), ('hatch_l1_h3_b', 85), ('tab',), ('hatch_l2_h3_b', 89), ('tab',), ('hatch_l3_h3_b', 93), ('tab',),
                        ('altura_2_painel_b', 61), ('tab',), ('tab',), ('hatch_l1_h2_b', 84), ('tab',), ('hatch_l2_h2_b', 88), ('tab',), ('hatch_l3_h2_b', 92), ('tab',),
                        ('altura_1_painel_b', 60), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',),

                        # Aberturas Painel B (Modificado para 0 em negativo)
                        ('abertura_esq1_pos_b', 70, 1, 'space', '0'), ('tab',),('abertura_esq1_pos_b', 70, 2, 'space', '0'), ('tab',),
                        ('abertura_esq1_dist_b', 67), ('tab',), ('abertura_esq1_prof_b', 68), ('tab',), ('abertura_esq1_larg_b', 69), ('tab',),
                        ('abertura_esq2_pos_b', 74, 1, 'space', '0'), ('tab',),('abertura_esq2_pos_b', 74, 2, 'space', '0'), ('tab',),
                        ('abertura_esq2_dist_b', 71), ('tab',), ('abertura_esq2_prof_b', 72), ('tab',), ('abertura_esq2_larg_b', 73), ('tab',),
                        ('abertura_dir1_pos_b', 78, 1, 'space', '0'), ('tab',),('abertura_dir1_pos_b', 78, 2, 'space', '0'), ('tab',),
                        ('abertura_dir1_dist_b', 75), ('tab',), ('abertura_dir1_prof_b', 76), ('tab',), ('abertura_dir1_larg_b', 77), ('tab',),
                        ('abertura_dir2_pos_b', 82, 1, 'space', '0'), ('tab',),('abertura_dir2_pos_b', 82, 2, 'space', '0'), ('tab',),
                        ('abertura_dir2_dist_b', 79), ('tab',), ('abertura_dir2_prof_b', 80), ('tab',), ('abertura_dir2_larg_b', 81), ('tab',), ('tab',), ('tab',),

                        # Painel C
                        ('largura_1_painel_c', 99), ('tab',), ('tab',),
                        ('largura_2_painel_c', 100), ('tab',), ('tab',),
                        ('altura_4_painel_c', 104), ('tab',), ('tab',), ('hatch_l1_h4_c', 126), ('tab',), ('hatch_l2_h4_c', 129), ('tab',),
                        ('altura_3_painel_c', 103), ('tab',), ('tab',), ('hatch_l1_h3_c', 125), ('tab',), ('hatch_l2_h3_c', 128), ('tab',),
                        ('altura_2_painel_c', 102), ('tab',), ('tab',), ('hatch_l1_h2_c', 124), ('tab',), ('hatch_l2_h2_c', 127), ('tab',),
                        ('altura_1_painel_c', 101), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',),

                        # Painel D
                        ('largura_1_painel_d', 133), ('tab',), ('tab',),
                        ('largura_2_painel_d', 134), ('tab',), ('tab',),
                        ('altura_4_painel_d', 138), ('tab',), ('tab',), ('hatch_l1_h4_d', 160), ('tab',), ('hatch_l2_h4_d', 163), ('tab',),
                        ('altura_3_painel_d', 137), ('tab',), ('tab',), ('hatch_l1_h3_d', 159), ('tab',), ('hatch_l2_h3_d', 162), ('tab',),
                        ('altura_2_painel_d', 136), ('tab',), ('tab',), ('hatch_l1_h2_d', 158), ('tab',), ('hatch_l2_h2_d', 161), ('tab',),
                        ('altura_1_painel_d', 135), ('tab',), ('tab',), ('tab',), ('tab',), ('tab',),
                    ]

                    # Otimização dos delays após ações principais
                    DELAY_APOS_CALCULAR = 0.2
                    DELAY_APOS_GERAR = 0.2
                    DELAY_APOS_LIMPAR = 0.2
                    DELAY_ENTRE_COLUNAS = 0.2

                    for item in sequencia_campos_parte_2A:
                        if parar_automacao:
                            break
                        if item[0] == 'tab':
                            continue
                        else:
                            tabs_count = 0
                            idx = sequencia_campos_parte_2A.index(item) + 1
                            while idx < len(sequencia_campos_parte_2A) and sequencia_campos_parte_2A[idx][0] == 'tab':
                                tabs_count += 1
                                idx += 1
                            
                            # Lógica unificada para campos normais e condicionais
                            if len(item) == 5:  # Campos condicionais (abertura)
                                campo, linha, valor_esperado, acao_true, acao_false = item
                                valor = formatar_valor(sheet[f'{coluna}{linha}'].value)
                                
                                if valor is not None and str(valor) == str(valor_esperado):
                                    pyautogui.press(acao_true)
                                else:
                                    pyautogui.press(acao_false)
                            else:  # Campos normais
                                campo, linha = item
                                valor = formatar_valor(sheet[f'{coluna}{linha}'].value)
                                if 'hatch' in campo:
                                    print(f"[DEBUG] (ANTES) Tentando digitar no campo de hatch: {campo} | Valor: {valor}")
                                    digitar_valor(valor)
                                    time.sleep(0.08)  # Delay maior para garantir preenchimento
                                    print(f"[DEBUG] (DEPOIS) Digitado no campo de hatch: {campo} | Valor: {valor}")
                                else:
                                    digitar_valor(valor)
                            
                            if tabs_count > 0:
                                multi_tab(tabs_count)

                    if parar_automacao:
                        break

                    # --- Cliques adicionais antes de "Gerar Script" ---
                    # Mapeamento de linhas para coordenadas de clique
                    coordenadas_clique = {
                        14: {0: (46, 290), 1: (86, 290), 2: (122, 290), 3: (166, 290), 4: (202, 290), 5: (242, 290)},
                        56: {0: (46, 338), 1: (86, 338), 2: (122, 338), 3: (166, 338), 4: (202, 338), 5: (242, 388)},
                        98: {0: (46, 389), 1: (86, 389), 2: (122, 389), 3: (166, 389), 4: (202, 389)},
                        131: {0: (46, 429), 1: (86, 429), 2: (122, 429), 3: (166, 429), 4: (202, 429)},
                    }

                    # Iterar pelas linhas e clicar com base nos valores
                    for linha, coordenadas in coordenadas_clique.items():
                        if parar_automacao:
                            break
                        valor = formatar_valor(sheet[f'{coluna}{linha}'].value)
                        # Forçar conversão para inteiro se o valor não for None
                        if valor is not None:
                            try:
                                valor = int(valor)
                            except ValueError:
                                valor = None  # Mantém None se a conversão falhar
                        if valor is not None and valor in coordenadas:
                            clicar_botao(coordenadas[valor])

                    if parar_automacao:
                        break

                    # Clicar no botão "Gerar Script"
                    clicar_botao(COORD_BOTAO_GERAR_SCRIPT)
                    time.sleep(DELAY_APOS_GERAR)

                    if parar_automacao:
                        break

                    # Clicar no botão "Limpar Campos"
                    clicar_botao(COORD_BOTAO_LIMPAR)
                    time.sleep(DELAY_APOS_LIMPAR)

                    if proxima_coluna_tem_dados:
                        time.sleep(DELAY_ENTRE_COLUNAS)

                else:
                    print(f"Coluna {coluna}: sem dados relevantes para VISAO ABCD.")
                    colunas_vazias += 1
                    if colunas_vazias >= 5:
                        print(f"Parando a busca após 5 colunas vazias consecutivas a partir da coluna {coluna}.")
                        parar_busca = True

            if not parar_automacao:
                print("Fim do processamento da planilha.")
                # Fechar a aplicação, se necessário

        # Iniciar a thread de automação
        thread_automacao = threading.Thread(target=automacao_thread)
        thread_automacao.start()

    except Exception as e:
        print(f"Erro ao processar a planilha: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar a planilha:\n{e}")

def iniciar_automacao_abcd():
    """Versão otimizada da função de início"""
    global caminho_arquivo_excel
    if caminho_arquivo_excel:
        print(f"Iniciando automação ABCD com o arquivo: {caminho_arquivo_excel}")
        
        # Configurar fail-safe do pyautogui para false para maior velocidade
        pyautogui.FAILSAFE = False
        
        # Iniciar a thread de automação
        preencher_campos_e_gerar_scripts(caminho_arquivo_excel)
        
        # Restaurar fail-safe após a execução
        pyautogui.FAILSAFE = True
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def ler_pavimentos_disponiveis(caminho_arquivo_excel):
    """Lê todos os pavimentos disponíveis no arquivo Excel."""
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        sheet = workbook.worksheets[0]
        
        pavimentos = set()  # Usar set para evitar duplicatas
        coluna_base = 'E'
        coluna_base_idx = ord(coluna_base) - ord('A')
        
        # Linha do pavimento é 4 conforme mapeamento em linhas_abcd
        linha_pavimento = 4
        
        for coluna_idx in range(coluna_base_idx, sheet.max_column):
            coluna = get_column_letter(coluna_idx + 1)
            valor = formatar_valor(sheet[f'{coluna}{linha_pavimento}'].value)
            if valor is not None and valor.strip():
                pavimentos.add(valor)
        
        return sorted(list(pavimentos))  # Converter para lista ordenada
    except Exception as e:
        print(f"Erro ao ler pavimentos: {e}")
        return []

def atualizar_lista_pavimentos():
    """Atualiza a lista de pavimentos no combobox."""
    global caminho_arquivo_excel, combo_pavimentos
    if caminho_arquivo_excel:
        pavimentos = ler_pavimentos_disponiveis(caminho_arquivo_excel)
        if pavimentos:
            combo_pavimentos['values'] = pavimentos
            combo_pavimentos.set(pavimentos[0])  # Seleciona o primeiro pavimento
            messagebox.showinfo("Pavimentos", f"Foram encontrados {len(pavimentos)} pavimentos:\n{', '.join(pavimentos)}")
            atualizar_combos_nomes()
        else:
            combo_pavimentos['values'] = []
            combo_pavimentos.set('')
            atualizar_combos_nomes()
            messagebox.showwarning("Aviso", "Nenhum pavimento encontrado no arquivo Excel.")

def abrir_arquivo_excel():
    """Abre uma janela de diálogo para selecionar um arquivo Excel."""
    global caminho_arquivo_excel
    caminho_arquivo_excel = filedialog.askopenfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo_excel:
        print(f"Arquivo selecionado: {caminho_arquivo_excel}")
        atualizar_lista_pavimentos()  # Atualiza a lista de pavimentos após selecionar o arquivo
    else:
        # Caminho padrão se nenhum argumento for fornecido
        script_dir = os.path.dirname(os.path.abspath(__file__))
        caminho_arquivo_excel = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(script_dir))), 
                                           "EXCELLS-PRODUCAO", "TEMPLATE3.0-PILARES-TORRE1- FIM.xlsx")

def atualizar_e_executar_automacao_cima():
    """Atualiza o caminho do arquivo Excel no script CIMA e o executa."""
    global caminho_arquivo_excel
    if not caminho_arquivo_excel:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
        return

    # Caminho para o novo script CIMA FUNCIONAL EXCEL.py
    caminho_script_cima = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CIMA FUNCIONAL EXCEL.py")
    
    try:
        # Verificar se o arquivo existe
        if not os.path.exists(caminho_script_cima):
            messagebox.showerror("Erro", f"O arquivo {caminho_script_cima} não foi encontrado.")
            return
            
        print(f"Executando automação CIMA com o arquivo: {caminho_arquivo_excel}")
        
        # Em vez de modificar o arquivo, vamos passar o caminho do Excel como argumento
        # Configurar o comando para executar o script com o caminho do Excel como argumento
        comando = f'python "{caminho_script_cima}" "{caminho_arquivo_excel}"'
        print(f"Executando comando: {comando}")
        
        # Executar o comando
        subprocess.run([sys.executable, caminho_script_cima, caminho_arquivo_excel], check=False)

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao executar o script CIMA:\n{e}")
        print(f"Erro detalhado: {str(e)}")

def executar_combinador_scripts_abcd():
    """Executa o script Combinador de SCR .py (para ABCD)."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    caminho_script_combinador = os.path.join(script_dir, "Combinador_de_SCR.py")
    try:
        print(f"Executando Combinador de SCR ABCD...")
        subprocess.run([sys.executable, caminho_script_combinador], check=False)
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao executar o script Combinador de SCR ABCD:\n{e}")

def executar_combinador_scripts_cima():
    """Executa o script combinador_de_scr_cima.py (para CIMA)."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    caminho_script_combinador_cima = os.path.join(script_dir, "Combinador_de_SCR _cima.py")
    try:
        print(f"Executando Combinador de SCR CIMA...")
        subprocess.run([sys.executable, caminho_script_combinador_cima], check=False)
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao executar o script Combinador de SCR CIMA:\n{e}")

def executar_ordenador_coordenadas_abcd(pasta_automatica=None):
    """Executa o ordenador de coordenadas ABCD"""
    try:
        if pasta_automatica:
            pasta_selecionada = pasta_automatica
        else:
            pasta_selecionada = filedialog.askdirectory(title="Selecione a pasta com os arquivos .scr")
        
        if pasta_selecionada:
            try:
                from .interface_ordenador_abcd import InterfaceOrdenador
                app = InterfaceOrdenador()
                app.pasta_selecionada = pasta_selecionada
                app.iniciar()
            except ImportError:
                # Fallback para subprocess se não conseguir importar
                import subprocess
                import sys
                ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_abcd.py")
                if os.path.exists(ordenador_path):
                    subprocess.Popen([sys.executable, ordenador_path])
                else:
                    raise ImportError("Não foi possível encontrar interface_ordenador_abcd")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao executar ordenador ABCD: {str(e)}")

def executar_ordenador_coordenadas_cima(pasta_automatica=None):
    """Executa o ordenador de coordenadas CIMA"""
    try:
        if pasta_automatica:
            pasta_selecionada = pasta_automatica
        else:
            pasta_selecionada = filedialog.askdirectory(title="Selecione a pasta com os arquivos .scr")
        
        if pasta_selecionada:
            try:
                from .interface_ordenador_cima import InterfaceOrdenadorCima
                app = InterfaceOrdenadorCima()
                app.pasta_selecionada = pasta_selecionada
                app.iniciar()
            except ImportError:
                # Fallback para subprocess se não conseguir importar
                import subprocess
                import sys
                ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_cima.py")
                if os.path.exists(ordenador_path):
                    subprocess.Popen([sys.executable, ordenador_path])
                else:
                    raise ImportError("Não foi possível encontrar interface_ordenador_cima")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao executar ordenador CIMA: {str(e)}")

def executar_ordenador_coordenadas_grades(pasta_automatica=None):
    """Executa o ordenador de coordenadas GRADES"""
    try:
        if pasta_automatica:
            pasta_selecionada = pasta_automatica
        else:
            pasta_selecionada = filedialog.askdirectory(title="Selecione a pasta com os arquivos .scr")
        
        if pasta_selecionada:
            try:
                from .interface_ordenador_grades import InterfaceOrdenadorGrades
                app = InterfaceOrdenadorGrades()
                app.pasta_selecionada = pasta_selecionada
                app.iniciar()
            except ImportError:
                # Fallback para subprocess se não conseguir importar
                import subprocess
                import sys
                ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_grades.py")
                if os.path.exists(ordenador_path):
                    subprocess.Popen([sys.executable, ordenador_path])
                else:
                    raise ImportError("Não foi possível encontrar interface_ordenador_grades")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao executar ordenador GRADES: {str(e)}")

def abrir_configuracao_abcd():
    """Abre a interface completa do ordenador ABCD com suas configurações"""
    try:
        try:
            from .interface_ordenador_abcd import InterfaceOrdenador
            app = InterfaceOrdenador()
            app.iniciar()
        except ImportError:
            # Fallback para subprocess se não conseguir importar
            import subprocess
            import sys
            ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_abcd.py")
            if os.path.exists(ordenador_path):
                subprocess.Popen([sys.executable, ordenador_path])
            else:
                raise ImportError("Não foi possível encontrar interface_ordenador_abcd")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do ABCD: {str(e)}")

def abrir_configuracao_cima():
    """Abre a interface completa do ordenador CIMA com suas configurações"""
    try:
        try:
            from .interface_ordenador_cima import InterfaceOrdenadorCima
            app = InterfaceOrdenadorCima()
            app.iniciar()
        except ImportError:
            # Fallback para subprocess se não conseguir importar
            import subprocess
            import sys
            ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_cima.py")
            if os.path.exists(ordenador_path):
                subprocess.Popen([sys.executable, ordenador_path])
            else:
                raise ImportError("Não foi possível encontrar interface_ordenador_cima")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do CIMA: {str(e)}")

def abrir_configuracao_grades():
    """Abre a interface completa do ordenador GRADES com suas configurações"""
    try:
        try:
            from .interface_ordenador_grades import InterfaceOrdenadorGrades
            app = InterfaceOrdenadorGrades()
            app.iniciar()
        except ImportError:
            # Fallback para subprocess se não conseguir importar
            import subprocess
            import sys
            ordenador_path = os.path.join(os.path.dirname(__file__), "interface_ordenador_grades.py")
            if os.path.exists(ordenador_path):
                subprocess.Popen([sys.executable, ordenador_path])
            else:
                raise ImportError("Não foi possível encontrar interface_ordenador_grades")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do GRADES: {str(e)}")

def abrir_configuracoes_ordenadores():
    """Abre a janela de configurações dos ordenadores"""
    try:
        # Abre as configurações de cada ordenador em sequência
        abrir_configuracao_abcd()
        abrir_configuracao_cima()
        abrir_configuracao_grades()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações: {str(e)}")

def get_colunas_por_pavimento(caminho_arquivo_excel, pavimento_selecionado):
    """Retorna as colunas que contêm o pavimento selecionado."""
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        sheet = workbook.worksheets[0]
        
        colunas = []
        coluna_base = 'E'
        coluna_base_idx = ord(coluna_base) - ord('A')
        
        # Linha do pavimento é 4 conforme mapeamento em linhas_abcd
        linha_pavimento = 4
        
        for coluna_idx in range(coluna_base_idx, sheet.max_column):
            coluna = get_column_letter(coluna_idx + 1)
            valor = formatar_valor(sheet[f'{coluna}{linha_pavimento}'].value)
            if valor == pavimento_selecionado:
                colunas.append(coluna)
        
        return colunas
    except Exception as e:
        print(f"Erro ao buscar colunas por pavimento: {e}")
        return []

def sobrescrever_arquivo_comando(caminho_scr_gerado, caminho_destino):
    try:
        with open(caminho_scr_gerado, 'r', encoding='utf-16') as f:
            conteudo = f.read()
        with open(caminho_destino, 'w', encoding='utf-16') as f:
            f.write(conteudo)
        print(f"Arquivo de comando sobrescrito com sucesso: {caminho_destino}")
    except Exception as e:
        print(f"Erro ao sobrescrever arquivo de comando: {e}")

def executar_abcd_excel():
    """Executa o script Abcd_Excel.py"""
    # Configurar logging de erros
    from datetime import datetime
    is_frozen = getattr(sys, 'frozen', False)
    error_log_file = os.path.join(os.path.dirname(sys.executable) if is_frozen else os.path.dirname(os.path.abspath(__file__)), "interface_errors.log")
    
    def log_error(msg):
        try:
            with open(error_log_file, 'a', encoding='utf-8') as f:
                f.write(f"[{datetime.now()}] {msg}\n")
        except Exception:
            pass
    
    try:
        log_error("executar_abcd_excel chamado")
        global caminho_arquivo_excel
        nome_selecionado = combo_nome_abcd.get()
        pavimento_selecionado = combo_pavimentos.get()
        if not pavimento_selecionado:
            messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
            return
        if not nome_selecionado:
            messagebox.showwarning("Aviso", "Selecione um nome primeiro.")
            return
        nomes_dict = get_nomes_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
        coluna_especificada = nomes_dict.get(nome_selecionado, None)
        if not coluna_especificada:
            messagebox.showwarning("Aviso", "Nome não encontrado para o pavimento selecionado.")
            return
        if caminho_arquivo_excel:
            print(f"Executando Abcd_Excel.py com o arquivo: {caminho_arquivo_excel} (Coluna: {coluna_especificada}, Pavimento: {pavimento_selecionado})")
            log_error(f"Frozen: {is_frozen}, sys.executable: {sys.executable}, Arquivo Excel: {caminho_arquivo_excel}")
            
            if is_frozen:
                # Em executável, tentar importar e executar diretamente
                try:
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    # Tentar encontrar o módulo no _MEIPASS ou no diretório do executável
                    if hasattr(sys, '_MEIPASS'):
                        script_dir = sys._MEIPASS
                        # Tentar src/interfaces
                        potential_path = os.path.join(script_dir, 'src', 'interfaces', 'Abcd_Excel.py')
                        if not os.path.exists(potential_path):
                            potential_path = os.path.join(script_dir, 'interfaces', 'Abcd_Excel.py')
                    else:
                        potential_path = os.path.join(script_dir, "Abcd_Excel.py")
                    
                    # Tentar importar diretamente
                    try:
                        from ..interfaces.Abcd_Excel import main as abcd_main
                        # Executar a função main com os argumentos
                        abcd_main(caminho_arquivo_excel, coluna_especificada, pavimento_selecionado)
                        messagebox.showinfo("Sucesso", f"Abcd_Excel executado com sucesso para a coluna {coluna_especificada} do pavimento {pavimento_selecionado}!")
                    except ImportError:
                        try:
                            from interfaces.Abcd_Excel import main as abcd_main
                            abcd_main(caminho_arquivo_excel, coluna_especificada, pavimento_selecionado)
                            messagebox.showinfo("Sucesso", f"Abcd_Excel executado com sucesso!")
                        except ImportError:
                            messagebox.showerror("Erro", f"Não foi possível importar Abcd_Excel no executável.\n\nTente usar a versão de desenvolvimento.")
                except Exception as e:
                    import traceback
                    error_msg = f"Erro ao executar Abcd_Excel no executável:\n{str(e)}\n\n{traceback.format_exc()}"
                    print(error_msg)
                    log_error(f"ERRO executar_abcd_excel: {error_msg}")
                    messagebox.showerror("Erro", f"Erro ao executar Abcd_Excel:\n{str(e)}")
            else:
                # Em desenvolvimento, usar subprocess normalmente
                script_dir = os.path.dirname(os.path.abspath(__file__))
                abcd_excel_path = os.path.join(script_dir, "Abcd_Excel.py")
                if os.path.exists(abcd_excel_path):
                    try:
                        resultado = subprocess.run(
                            [sys.executable, abcd_excel_path, caminho_arquivo_excel, coluna_especificada, pavimento_selecionado], 
                            capture_output=True, 
                            text=True
                        )
                        print("Saída do Abcd_Excel.py:")
                        print(resultado.stdout)
                        if resultado.returncode == 0:
                            messagebox.showinfo("Sucesso", f"Abcd_Excel.py executado com sucesso para a coluna {coluna_especificada} do pavimento {pavimento_selecionado}!")
                            # --- NOVO: Procurar o último .scr gerado e sobrescrever o comando ---
                            from glob import glob
                            import time
                            # Usar path resolver para obter caminho dinâmico
                            from ..utils.robust_path_resolver import robust_path_resolver
                            pasta_scripts = os.path.join(robust_path_resolver.get_project_root(), 'output')
                            nome_pasta = str(pavimento_selecionado).replace(' ', '_') + '_ABCD'
                            pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
                            arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome_selecionado}*.scr"))
                            if arquivos_scr:
                                arquivo_scr_mais_recente = max(arquivos_scr, key=os.path.getmtime)
                                caminho_destino = os.path.join(robust_path_resolver.get_project_root(), 'output', 'comando_pilar_combinado_tpatpa.scr')
                                sobrescrever_arquivo_comando(arquivo_scr_mais_recente, caminho_destino)
                        else:
                            print("Erro:", resultado.stderr)
                            messagebox.showerror("Erro", f"Erro ao executar Abcd_Excel.py para a coluna {coluna_especificada}:\n{resultado.stderr}")
                    except Exception as e:
                        print(f"Erro ao executar Abcd_Excel.py: {e}")
                        log_error(f"ERRO subprocess Abcd_Excel: {str(e)}")
                        messagebox.showerror("Erro", f"Erro ao executar Abcd_Excel.py:\n{e}")
                else:
                    print(f"Arquivo não encontrado: {abcd_excel_path}")
                    log_error(f"Arquivo nao encontrado: {abcd_excel_path}")
                    messagebox.showerror("Erro", f"O arquivo Abcd_Excel.py não foi encontrado em:\n{abcd_excel_path}")
        else:
            messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
    except Exception as e:
        import traceback
        error_msg = f"ERRO GERAL em executar_abcd_excel: {str(e)}\n{traceback.format_exc()}"
        log_error(error_msg)
        messagebox.showerror("Erro Fatal", f"Erro ao executar Abcd_Excel:\n{str(e)}")

def executar_todos_abcd_excel():
    """Executa o script Abcd_Excel.py para todas as colunas do pavimento selecionado."""
    global caminho_arquivo_excel
    
    pavimento_selecionado = combo_pavimentos.get()
    
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    
    if caminho_arquivo_excel:
        # Buscar todas as colunas do pavimento
        colunas = get_colunas_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
        
        if not colunas:
            messagebox.showwarning("Aviso", f"Nenhuma coluna encontrada para o pavimento {pavimento_selecionado}")
            return
        
        print(f"Executando Abcd_Excel.py para todas as colunas do pavimento {pavimento_selecionado}")
        
        # Definir o caminho para o script Abcd_Excel.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        abcd_excel_path = os.path.join(script_dir, "Abcd_Excel.py")
        
        # Verificar se o arquivo existe
        if os.path.exists(abcd_excel_path):
            try:
                # Passar todas as colunas como argumento único, separadas por vírgula
                colunas_str = ','.join(colunas)
                resultado = subprocess.run(
                    [sys.executable, abcd_excel_path, caminho_arquivo_excel, colunas_str, pavimento_selecionado], 
                    capture_output=True, 
                    text=True
                )
                print("Saída do Abcd_Excel.py:")
                print(resultado.stdout)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", f"Abcd_Excel.py executado com sucesso para todas as colunas do pavimento {pavimento_selecionado}!")
                else:
                    print("Erro:", resultado.stderr)
                    messagebox.showerror("Erro", f"Erro ao executar Abcd_Excel.py para as colunas:\n{resultado.stderr}")
            except Exception as e:
                print(f"Erro ao executar Abcd_Excel.py: {e}")
                messagebox.showerror("Erro", f"Erro ao executar Abcd_Excel.py:\n{e}")
        else:
            print(f"Arquivo não encontrado: {abcd_excel_path}")
            messagebox.showerror("Erro", f"O arquivo Abcd_Excel.py não foi encontrado em:\n{abcd_excel_path}")
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def executar_cima_excel():
    global caminho_arquivo_excel
    nome_selecionado = combo_nome_cima.get()
    pavimento_selecionado = combo_pavimentos.get()
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    if not nome_selecionado:
        messagebox.showwarning("Aviso", "Selecione um nome primeiro.")
        return
    nomes_dict = get_nomes_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
    coluna_especificada = nomes_dict.get(nome_selecionado, None)
    if not coluna_especificada:
        messagebox.showwarning("Aviso", "Nome não encontrado para o pavimento selecionado.")
        return
    if caminho_arquivo_excel:
        print(f"Executando CIMA_FUNCIONAL_EXCEL.py com o arquivo: {caminho_arquivo_excel} (Coluna: {coluna_especificada}, Pavimento: {pavimento_selecionado})")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        cima_excel_path = os.path.join(script_dir, "CIMA_FUNCIONAL_EXCEL.py")
        if os.path.exists(cima_excel_path):
            try:
                resultado = subprocess.run(
                    [sys.executable, cima_excel_path, caminho_arquivo_excel, coluna_especificada, pavimento_selecionado], 
                    capture_output=True, 
                    text=True
                )
                print("Saída do CIMA_FUNCIONAL_EXCEL.py:")
                print(resultado.stdout)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", f"CIMA_FUNCIONAL_EXCEL.py executado com sucesso para a coluna {coluna_especificada} do pavimento {pavimento_selecionado}!")
                    # --- NOVO: Procurar o último .scr gerado e sobrescrever o comando ---
                    from glob import glob
                    # Usar path resolver para obter caminho dinâmico
                    from ..utils.robust_path_resolver import robust_path_resolver
                    pasta_scripts = os.path.join(robust_path_resolver.get_project_root(), 'output')
                    nome_pasta = str(pavimento_selecionado).replace(' ', '_') + '_CIMA'
                    pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
                    arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome_selecionado}*.scr"))
                    if arquivos_scr:
                        arquivo_scr_mais_recente = max(arquivos_scr, key=os.path.getmtime)
                        caminho_destino = os.path.join(robust_path_resolver.get_project_root(), 'output', 'comando_pilar_combinado.scr')
                        sobrescrever_arquivo_comando(arquivo_scr_mais_recente, caminho_destino)
                else:
                    print("Erro:", resultado.stderr)
                    messagebox.showerror("Erro", f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py para a coluna {coluna_especificada}:\n{resultado.stderr}")
            except Exception as e:
                print(f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py: {e}")
                messagebox.showerror("Erro", f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py:\n{e}")
        else:
            print(f"Arquivo não encontrado: {cima_excel_path}")
            messagebox.showerror("Erro", f"O arquivo CIMA_FUNCIONAL_EXCEL.py não foi encontrado em:\n{cima_excel_path}")
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def executar_todos_cima_excel():
    """Executa o script CIMA_FUNCIONAL_EXCEL.py para todas as colunas do pavimento selecionado."""
    global caminho_arquivo_excel
    
    pavimento_selecionado = combo_pavimentos.get()
    
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    
    if caminho_arquivo_excel:
        # Buscar todas as colunas do pavimento
        colunas = get_colunas_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
        
        if not colunas:
            messagebox.showwarning("Aviso", f"Nenhuma coluna encontrada para o pavimento {pavimento_selecionado}")
            return
        
        print(f"Executando CIMA_FUNCIONAL_EXCEL.py para todas as colunas do pavimento {pavimento_selecionado}")
        
        # Definir o caminho para o script CIMA_FUNCIONAL_EXCEL.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        cima_excel_path = os.path.join(script_dir, "CIMA_FUNCIONAL_EXCEL.py")
        
        # Verificar se o arquivo existe
        if os.path.exists(cima_excel_path):
            try:
                colunas_str = ','.join(colunas)
                resultado = subprocess.run(
                    [sys.executable, cima_excel_path, caminho_arquivo_excel, colunas_str, pavimento_selecionado], 
                    capture_output=True, 
                    text=True
                )
                print("Saída do CIMA_FUNCIONAL_EXCEL.py:")
                print(resultado.stdout)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", f"CIMA_FUNCIONAL_EXCEL.py executado com sucesso para todas as colunas do pavimento {pavimento_selecionado}!")
                else:
                    print("Erro:", resultado.stderr)
                    messagebox.showerror("Erro", f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py para as colunas:\n{resultado.stderr}")
            except Exception as e:
                print(f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py: {e}")
                messagebox.showerror("Erro", f"Erro ao executar CIMA_FUNCIONAL_EXCEL.py:\n{e}")
        else:
            print(f"Arquivo não encontrado: {cima_excel_path}")
            messagebox.showerror("Erro", f"O arquivo CIMA_FUNCIONAL_EXCEL.py não foi encontrado em:\n{cima_excel_path}")
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def executar_grade_excel():
    global caminho_arquivo_excel
    nome_selecionado = combo_nome_grades.get()
    pavimento_selecionado = combo_pavimentos.get()
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    if not nome_selecionado:
        messagebox.showwarning("Aviso", "Selecione um nome primeiro.")
        return
    nomes_dict = get_nomes_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
    coluna_especificada = nomes_dict.get(nome_selecionado, None)
    if not coluna_especificada:
        messagebox.showwarning("Aviso", "Nome não encontrado para o pavimento selecionado.")
        return
    if caminho_arquivo_excel:
        print(f"Executando GRADE_EXCEL.py com o arquivo: {caminho_arquivo_excel} (Coluna: {coluna_especificada}, Pavimento: {pavimento_selecionado})")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        grade_excel_path = os.path.join(script_dir, "GRADE_EXCEL.py")
        if os.path.exists(grade_excel_path):
            try:
                resultado = subprocess.run(
                    [sys.executable, grade_excel_path, caminho_arquivo_excel, coluna_especificada, pavimento_selecionado], 
                    capture_output=True, 
                    text=True
                )
                print("Saída do GRADE_EXCEL.py:")
                print(resultado.stdout)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", f"GRADE_EXCEL.py executado com sucesso para a coluna {coluna_especificada} do pavimento {pavimento_selecionado}!")
                    # --- NOVO: Procurar o último .scr gerado e sobrescrever o comando ---
                    from glob import glob
                    # Usar path resolver para obter caminho dinâmico
                    from ..utils.robust_path_resolver import robust_path_resolver
                    pasta_scripts = os.path.join(robust_path_resolver.get_project_root(), 'output')
                    nome_pasta = str(pavimento_selecionado).replace(' ', '_') + '_GRADES'
                    pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)
                    arquivos_scr = glob(os.path.join(pasta_pavimento, f"*{nome_selecionado}*.scr"))
                    if arquivos_scr:
                        arquivo_scr_mais_recente = max(arquivos_scr, key=os.path.getmtime)
                        caminho_destino = os.path.join(robust_path_resolver.get_project_root(), 'output', 'comando_pilar_combinado.scr')
                        sobrescrever_arquivo_comando(arquivo_scr_mais_recente, caminho_destino)
                else:
                    print("Erro:", resultado.stderr)
                    messagebox.showerror("Erro", f"Erro ao executar GRADE_EXCEL.py para a coluna {coluna_especificada}:\n{resultado.stderr}")
            except Exception as e:
                print(f"Erro ao executar GRADE_EXCEL.py: {e}")
                messagebox.showerror("Erro", f"Erro ao executar GRADE_EXCEL.py:\n{e}")
        else:
            print(f"Arquivo não encontrado: {grade_excel_path}")
            messagebox.showerror("Erro", f"O arquivo GRADE_EXCEL.py não foi encontrado em:\n{grade_excel_path}")
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def executar_todos_grade_excel():
    """Executa o script GRADE_EXCEL.py para todas as colunas do pavimento selecionado."""
    global caminho_arquivo_excel
    
    pavimento_selecionado = combo_pavimentos.get()
    
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    
    if caminho_arquivo_excel:
        # Buscar todas as colunas do pavimento
        colunas = get_colunas_por_pavimento(caminho_arquivo_excel, pavimento_selecionado)
        
        if not colunas:
            messagebox.showwarning("Aviso", f"Nenhuma coluna encontrada para o pavimento {pavimento_selecionado}")
            return
        
        print(f"Executando GRADE_EXCEL.py para todas as colunas do pavimento {pavimento_selecionado}")
        
        # Definir o caminho para o script GRADE_EXCEL.py
        script_dir = os.path.dirname(os.path.abspath(__file__))
        grade_excel_path = os.path.join(script_dir, "GRADE_EXCEL.py")
        
        # Verificar se o arquivo existe
        if os.path.exists(grade_excel_path):
            try:
                colunas_str = ','.join(colunas)
                resultado = subprocess.run(
                    [sys.executable, grade_excel_path, caminho_arquivo_excel, colunas_str, pavimento_selecionado], 
                    capture_output=True, 
                    text=True
                )
                print("Saída do GRADE_EXCEL.py:")
                print(resultado.stdout)
                if resultado.returncode == 0:
                    messagebox.showinfo("Sucesso", f"GRADE_EXCEL.py executado com sucesso para todas as colunas do pavimento {pavimento_selecionado}!")
                else:
                    print("Erro:", resultado.stderr)
                    messagebox.showerror("Erro", f"Erro ao executar GRADE_EXCEL.py para as colunas:\n{resultado.stderr}")
            except Exception as e:
                print(f"Erro ao executar GRADE_EXCEL.py: {e}")
                messagebox.showerror("Erro", f"Erro ao executar GRADE_EXCEL.py:\n{e}")
        else:
            print(f"Arquivo não encontrado: {grade_excel_path}")
            messagebox.showerror("Erro", f"O arquivo GRADE_EXCEL.py não foi encontrado em:\n{grade_excel_path}")
    else:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")

def executar_combinador_grades():
    """Executa o script Combinador_de_SCR_GRADES.py."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    combinador_grades_path = os.path.join(script_dir, "Combinador_de_SCR_GRADES.py")
    try:
        print(f"Executando Combinador_de_SCR_GRADES.py...")
        subprocess.run([sys.executable, combinador_grades_path], check=False)
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao executar o script Combinador_de_SCR_GRADES.py:\n{e}")

def executar_ordenador_grades():
    """Executa o script Ordenador_Cordenadas_grades.py."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ordenador_grades_path = os.path.join(script_dir, "Ordenador_Cordenadas_grades.py")
    try:
        print(f"Executando Ordenador_Cordenadas_grades.py...")
        subprocess.run([sys.executable, ordenador_grades_path], check=False)
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao executar o script Ordenador_Cordenadas_grades.py:\n{e}")

# Funções para os outros botões (a serem implementadas)
def iniciar_automacao_cima():
    """Inicia a automação CIMA."""
    atualizar_e_executar_automacao_cima()

def combinar_codigos_abcd():
    """Combina códigos ABCD."""
    executar_combinador_scripts_abcd()

def combinar_codigos_cima():
    """Combina códigos CIMA."""
    executar_combinador_scripts_cima()

def ordenar_coordenadas_abcd():
    """Ordena coordenadas ABCD."""
    executar_ordenador_coordenadas_abcd()

def ordenar_coordenadas_cima():
    """Ordena coordenadas CIMA."""
    executar_ordenador_coordenadas_cima()

def pipeline_automatizado(tipo):
    """
    Executa o pipeline completo: geração de scripts, combinador e ordenador.
    tipo: 'ABCD', 'CIMA' ou 'GRADES'
    """
    global caminho_arquivo_excel
    pavimento_selecionado = combo_pavimentos.get()
    if not pavimento_selecionado:
        messagebox.showwarning("Aviso", "Selecione um pavimento primeiro.")
        return
    if not caminho_arquivo_excel:
        messagebox.showwarning("Aviso", "Selecione um arquivo Excel primeiro.")
        return
    # Definir sufixos e scripts
    sufixo = {'ABCD': '_ABCD', 'CIMA': '_CIMA', 'GRADES': '_GRADES'}[tipo]
    combinador = {
        'ABCD': 'Combinador_de_SCR.py',
        'CIMA': 'Combinador_de_SCR _cima.py',
        'GRADES': 'Combinador_de_SCR_GRADES.py'
    }[tipo]
    # Caminho da pasta do pavimento
    # Usar path resolver para obter caminho dinâmico
    from ..utils.robust_path_resolver import robust_path_resolver
    pasta_scripts = os.path.join(robust_path_resolver.get_project_root(), 'output')
    nome_pasta = str(pavimento_selecionado).replace(' ', '_')
    if not nome_pasta.upper().endswith(sufixo):
        nome_pasta += sufixo
    pasta_pavimento = os.path.join(pasta_scripts, nome_pasta)

    # NOVO: Verificar e excluir a pasta se ela já existir
    if os.path.exists(pasta_pavimento):
        print(f"Excluindo pasta existente: {pasta_pavimento}")
        try:
            shutil.rmtree(pasta_pavimento)
            print(f"Pasta {pasta_pavimento} excluída com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível excluir a pasta {pasta_pavimento}:\n{e}")
            print(f"Erro ao excluir pasta: {e}")
            return # Parar a execução se não conseguir excluir a pasta

    # 1. Gerar scripts
    if tipo == 'ABCD':
        executar_todos_abcd_excel()
    elif tipo == 'CIMA':
        executar_todos_cima_excel()
    elif tipo == 'GRADES':
        executar_todos_grade_excel()
    
    # 2. Executar combinador (passando o path da pasta)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    combinador_path = os.path.join(script_dir, combinador)
    print(f"Executando combinador: {combinador_path} na pasta {pasta_pavimento}")
    subprocess.run([sys.executable, combinador_path, pasta_pavimento], cwd=script_dir)
    
    # 3. Executar ordenador na subpasta Combinados (chamando função diretamente)
    pasta_combinados = os.path.join(pasta_pavimento, 'Combinados')
    print(f"Executando ordenador {tipo} na pasta {pasta_combinados}")
    
    # Chamar a função do ordenador correspondente passando a pasta automaticamente
    if tipo == 'ABCD':
        executar_ordenador_coordenadas_abcd(pasta_combinados)
    elif tipo == 'CIMA':
        executar_ordenador_coordenadas_cima(pasta_combinados)
    elif tipo == 'GRADES':
        executar_ordenador_grades(pasta_combinados)
    
    messagebox.showinfo("Pipeline Concluído", f"Pipeline {tipo} finalizado para o pavimento {pavimento_selecionado}!")

def executar_pipeline_abcd():
    pipeline_automatizado('ABCD')
def executar_pipeline_cima():
    pipeline_automatizado('CIMA')
def executar_pipeline_grades():
    pipeline_automatizado('GRADES')

def abrir_robo_abcd():
    """Abre a janela de configurações do robô ABCD"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        robo_path = os.path.join(script_dir, "..", "robots", "Robo_Pilar_ABCD.py")
        if os.path.exists(robo_path):
            subprocess.Popen([sys.executable, robo_path, "--config"])
        else:
            messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do robô ABCD: {str(e)}")

def abrir_robo_cima():
    """Abre a janela de configurações do robô CIMA"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        robo_path = os.path.join(script_dir, "..", "robots", "Robo_Pilar_Visao_Cima.py")
        if os.path.exists(robo_path):
            subprocess.Popen([sys.executable, robo_path, "--config"])
        else:
            messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do robô CIMA: {str(e)}")

def abrir_robo_grades():
    """Abre a janela de configurações do robô GRADES"""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        robo_path = os.path.join(script_dir, "..", "robots", "ROBO_GRADES.py")
        if os.path.exists(robo_path):
            subprocess.Popen([sys.executable, robo_path, "--config"])
        else:
            messagebox.showerror("Erro", f"Arquivo não encontrado: {robo_path}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir configurações do robô GRADES: {str(e)}")

# Configuração da interface Tkinter
root = tk.Tk()
root.title("Automação de Pilares")
root.geometry("800x600")  # Aumentado para melhor organização

# Frame principal para seleção de arquivo e pavimento
frame_principal = ttk.LabelFrame(root, text="Seleção de Arquivo e Pavimento", padding="10")
frame_principal.pack(fill=tk.X, padx=10, pady=5)

# Subframe para seleção de arquivo
frame_arquivo = ttk.Frame(frame_principal)
frame_arquivo.pack(fill=tk.X, pady=5)

botao_abrir_excel = ttk.Button(frame_arquivo, text="Selecionar Excel", command=abrir_arquivo_excel)
botao_abrir_excel.pack(side=tk.LEFT, padx=5)

label_pavimentos = ttk.Label(frame_arquivo, text="Pavimento:")
label_pavimentos.pack(side=tk.LEFT, padx=5)

combo_pavimentos = ttk.Combobox(frame_arquivo, width=30, state="readonly")
combo_pavimentos.pack(side=tk.LEFT, padx=5)

botao_atualizar_pavimentos = ttk.Button(frame_arquivo, text="Atualizar Pavimentos", command=atualizar_lista_pavimentos)
botao_atualizar_pavimentos.pack(side=tk.LEFT, padx=5)

# Frame para ABCD
frame_abcd = ttk.LabelFrame(root, text="ABCD - Visão Lateral", padding="10")
frame_abcd.pack(fill=tk.X, padx=10, pady=5)

frame_abcd_controles = ttk.Frame(frame_abcd)
frame_abcd_controles.pack(fill=tk.X, pady=5)

ttk.Label(frame_abcd_controles, text="Nome:").pack(side=tk.LEFT, padx=5)
combo_nome_abcd = ttk.Combobox(frame_abcd_controles, width=30, state="readonly")
combo_nome_abcd.pack(side=tk.LEFT, padx=5)

frame_abcd_botoes = ttk.Frame(frame_abcd)
frame_abcd_botoes.pack(fill=tk.X, pady=5)

ttk.Button(frame_abcd_botoes, text="Executar Individual", command=executar_abcd_excel).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_abcd_botoes, text="Executar Todos + Combinar", command=executar_pipeline_abcd).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_abcd_botoes, text="Ajustes Scripts", command=abrir_robo_abcd).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_abcd_botoes, text="Configurar Ordenador", command=abrir_configuracao_abcd).pack(side=tk.LEFT, padx=5)

# Frame para CIMA
frame_cima = ttk.LabelFrame(root, text="CIMA - Visão Superior", padding="10")
frame_cima.pack(fill=tk.X, padx=10, pady=5)

frame_cima_controles = ttk.Frame(frame_cima)
frame_cima_controles.pack(fill=tk.X, pady=5)

ttk.Label(frame_cima_controles, text="Nome:").pack(side=tk.LEFT, padx=5)
combo_nome_cima = ttk.Combobox(frame_cima_controles, width=30, state="readonly")
combo_nome_cima.pack(side=tk.LEFT, padx=5)

frame_cima_botoes = ttk.Frame(frame_cima)
frame_cima_botoes.pack(fill=tk.X, pady=5)

ttk.Button(frame_cima_botoes, text="Executar Individual", command=executar_cima_excel).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_cima_botoes, text="Executar Todos + Combinar", command=executar_pipeline_cima).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_cima_botoes, text="Ajustes Scripts", command=abrir_robo_cima).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_cima_botoes, text="Configurar Ordenador", command=abrir_configuracao_cima).pack(side=tk.LEFT, padx=5)

# Frame para GRADES
frame_grades = ttk.LabelFrame(root, text="GRADES - Detalhamento", padding="10")
frame_grades.pack(fill=tk.X, padx=10, pady=5)

frame_grades_controles = ttk.Frame(frame_grades)
frame_grades_controles.pack(fill=tk.X, pady=5)

ttk.Label(frame_grades_controles, text="Nome:").pack(side=tk.LEFT, padx=5)
combo_nome_grades = ttk.Combobox(frame_grades_controles, width=30, state="readonly")
combo_nome_grades.pack(side=tk.LEFT, padx=5)

frame_grades_botoes = ttk.Frame(frame_grades)
frame_grades_botoes.pack(fill=tk.X, pady=5)

ttk.Button(frame_grades_botoes, text="Executar Individual", command=executar_grade_excel).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_grades_botoes, text="Executar Todos + Combinar", command=executar_pipeline_grades).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_grades_botoes, text="Ajustes Scripts", command=abrir_robo_grades).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_grades_botoes, text="Configurar Ordenador", command=abrir_configuracao_grades).pack(side=tk.LEFT, padx=5)

# Atualizar combos de nomes ao selecionar pavimento
combo_pavimentos.bind("<<ComboboxSelected>>", lambda e: atualizar_combos_nomes())

root.mainloop()

# Botão para executar o script CIMA_FUNCIONAL_EXCEL.py para todas as colunas
botao_executar_todos_cima_excel = tk.Button(frame_cima, text="Executar CIMA (Todos)", command=executar_pipeline_cima, bg="#03A9F4", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5)
botao_executar_todos_cima_excel.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

# Botão para abrir o robô CIMA
botao_ajustes_cima = tk.Button(frame_cima, text="Ajustes Scripts Cima", command=abrir_robo_cima, bg="#FF9800", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5)
botao_ajustes_cima.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

# Frame para GRADES Excel
frame_grades = tk.Frame(frame_executar_scripts, bg="#f0f0f0")
frame_grades.pack(fill=tk.X, padx=10, pady=5)

# Campo para inserir coluna GRADES
label_coluna_grades = tk.Label(frame_grades, text="Nome GRADES:", bg="#f0f0f0")
label_coluna_grades.pack(side=tk.LEFT, padx=(0, 5))

combo_nome_grades = ttk.Combobox(frame_grades, width=30, state="readonly")
combo_nome_grades.pack(side=tk.LEFT, padx=(0, 10))

# Botão para executar o script GRADE_EXCEL.py para uma coluna específica
botao_executar_grades_excel = tk.Button(frame_grades, text="Executar GRADES Excel", command=executar_grade_excel, bg="#9C27B0", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5)
botao_executar_grades_excel.pack(side=tk.LEFT, fill=tk.X, expand=True)

# Botão para executar o script GRADE_EXCEL.py para todas as colunas
botao_executar_todos_grades_excel = tk.Button(frame_grades, text="Executar GRADES (Todos)", command=executar_pipeline_grades, bg="#BA68C8", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5)
botao_executar_todos_grades_excel.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

# Botão para abrir o robô GRADES
botao_ajustes_grades = tk.Button(frame_grades, text="Ajustes Scripts Grades", command=abrir_robo_grades, bg="#FF9800", fg="white", font=("Arial", 10, "bold"), padx=10, pady=5)
botao_ajustes_grades.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))

# Função para obter nomes da linha 5 por pavimento

def get_nomes_por_pavimento(caminho_arquivo_excel, pavimento_selecionado):
    """Retorna um dicionário {nome: coluna} para o pavimento selecionado, onde nome é o valor da linha 5."""
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        sheet = workbook.worksheets[0]
        nomes_colunas = {}
        coluna_base = 'E'
        coluna_base_idx = ord(coluna_base) - ord('A')
        linha_pavimento = 4
        linha_nome = 5
        for coluna_idx in range(coluna_base_idx, sheet.max_column):
            coluna = get_column_letter(coluna_idx + 1)
            valor_pav = formatar_valor(sheet[f'{coluna}{linha_pavimento}'].value)
            valor_nome = formatar_valor(sheet[f'{coluna}{linha_nome}'].value)
            if valor_pav == pavimento_selecionado and valor_nome:
                nomes_colunas[valor_nome] = coluna
        return nomes_colunas
    except Exception as e:
        print(f"Erro ao buscar nomes por pavimento: {e}")
        return {}

# Função para atualizar os combos de nomes ao selecionar pavimento ou Excel

def atualizar_combos_nomes():
    global caminho_arquivo_excel, combo_pavimentos, combo_nome_abcd, combo_nome_cima, combo_nome_grades
    # Só atualiza se os combos já existirem
    if not all([combo_nome_abcd, combo_nome_cima, combo_nome_grades]):
        return
    pavimento = combo_pavimentos.get()
    if caminho_arquivo_excel and pavimento:
        nomes_dict = get_nomes_por_pavimento(caminho_arquivo_excel, pavimento)
        nomes = list(nomes_dict.keys())
        for combo in [combo_nome_abcd, combo_nome_cima, combo_nome_grades]:
            combo['values'] = nomes
            if nomes:
                combo.set(nomes[0])
            else:
                combo.set('')
    else:
        for combo in [combo_nome_abcd, combo_nome_cima, combo_nome_grades]:
            combo['values'] = []
            combo.set('')

# Atualizar combos de nomes ao selecionar pavimento
combo_pavimentos.bind("<<ComboboxSelected>>", lambda e: atualizar_combos_nomes())

# Atualizar combos de nomes ao selecionar Excel
# (já é chamado em atualizar_lista_pavimentos, mas garantir aqui também)
def atualizar_lista_pavimentos():
    global caminho_arquivo_excel, combo_pavimentos
    if caminho_arquivo_excel:
        pavimentos = ler_pavimentos_disponiveis(caminho_arquivo_excel)
        if pavimentos:
            combo_pavimentos['values'] = pavimentos
            combo_pavimentos.set(pavimentos[0])
            messagebox.showinfo("Pavimentos", f"Foram encontrados {len(pavimentos)} pavimentos:\n{', '.join(pavimentos)}")
            atualizar_combos_nomes()
        else:
            combo_pavimentos['values'] = []
            combo_pavimentos.set('')
            atualizar_combos_nomes()
            messagebox.showwarning("Aviso", "Nenhum pavimento encontrado no arquivo Excel.")

# Adicionar botões de configuração dos ordenadores
frame_ordenadores = ttk.LabelFrame(root, text="Ordenadores", padding="5")
frame_ordenadores.pack(fill=tk.X, padx=5, pady=5)

# Frame para botões principais
frame_botoes_principais = ttk.Frame(frame_ordenadores)
frame_botoes_principais.pack(fill=tk.X, padx=5, pady=2)

# Botões existentes ajustados
ttk.Button(frame_botoes_principais, text="Ordenador ABCD", command=executar_ordenador_coordenadas_abcd).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_botoes_principais, text="Ordenador CIMA", command=executar_ordenador_coordenadas_cima).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_botoes_principais, text="Ordenador GRADES", command=executar_ordenador_coordenadas_grades).pack(side=tk.LEFT, padx=5)

# Frame para botões de configuração
frame_botoes_config = ttk.Frame(frame_ordenadores)
frame_botoes_config.pack(fill=tk.X, padx=5, pady=2)

# Botões de configuração individuais
ttk.Button(frame_botoes_config, text="Configurar ABCD", command=abrir_configuracao_abcd).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_botoes_config, text="Configurar CIMA", command=abrir_configuracao_cima).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_botoes_config, text="Configurar GRADES", command=abrir_configuracao_grades).pack(side=tk.LEFT, padx=5)

root.mainloop() 