
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

import openpyxl
import os
import sys
import time
import traceback
import math
import shutil
# import xlwings as xw  # Comentado temporariamente - usando openpyxl

# IMPORTAR HELPER FROZEN GLOBAL - garante que paths estão configurados
try:
    from _frozen_helper import ensure_paths
    ensure_paths()
except ImportError:
    try:
        from src._frozen_helper import ensure_paths
        ensure_paths()
    except ImportError:
        try:
            from _ensure_frozen import ensure
            ensure()
        except ImportError:
            # Fallback manual
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
                if script_dir not in sys.path:
                    sys.path.insert(0, script_dir)
                src_dir = os.path.join(script_dir, 'src')
                if os.path.exists(src_dir) and src_dir not in sys.path:
                    sys.path.insert(0, src_dir)
            else:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                parent_dir = os.path.dirname(current_dir)
                if parent_dir not in sys.path:
                    sys.path.insert(0, parent_dir)

# Configurar caminhos usando sistema robusto com múltiplos fallbacks
def setup_robots_path():
    """Configura o path para os robôs de forma robusta"""
    try:
        from config_paths import ROBOTS_DIR
        return ROBOTS_DIR
    except ImportError:
        try:
            from src.config_paths import ROBOTS_DIR
            return ROBOTS_DIR
        except ImportError:
            # Fallback para nova estrutura
            if getattr(sys, 'frozen', False):
                script_dir = os.path.dirname(sys.executable)
                robots_dir = os.path.join(script_dir, 'src', 'robots')
                if os.path.exists(robots_dir):
                    return robots_dir
                return os.path.join(script_dir, 'robots')
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                return os.path.abspath(os.path.join(script_dir, "..", "robots"))

robos_dir = setup_robots_path()

# Adicionar diretório dos robôs ao path de forma confiável
if robos_dir and robos_dir not in sys.path:
    sys.path.insert(0, robos_dir)

print(f"Diretório de robôs adicionado ao path: {robos_dir}")
print(f"Path atual: {sys.path}")

# Verificar se o arquivo existe fisicamente
robo_path = os.path.join(robos_dir, "ROBO_GRADES.py")
if not os.path.exists(robo_path):
    print(f"ERRO: O arquivo ROBO_GRADES.py não foi encontrado em {robos_dir}")
    print(f"Arquivos disponíveis no diretório:")
    for arquivo in os.listdir(robos_dir):
        print(f"  - {arquivo}")
    sys.exit(1)
else:
    print(f"Arquivo ROBO_GRADES.py encontrado em: {robo_path}")

# Importar bibliotecas necessárias
try:
    # Importar módulos necessários que o ROBO_GRADES.py usa
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import logging
    import matplotlib
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import json
    from datetime import datetime
    
    # Tentar importar win32com.client, mas continuar mesmo se falhar
    try:
        import win32com.client
        import pyautogui
        print("Módulos win32com e pyautogui importados com sucesso.")
    except ImportError:
        print("AVISO: Módulos win32com.client ou pyautogui não estão instalados.")
        print("Algumas funcionalidades relacionadas ao AutoCAD podem não funcionar.")
    
    # Configuração de logging
    log_file = os.path.join(os.path.dirname(robos_dir), "pilares.log")
    logging.basicConfig(
        level=logging.INFO, 
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file), 
            logging.StreamHandler()
        ]
    )
    
    # Realizar a importação do módulo ROBO_GRADES
    print("Tentando importar módulos do ROBO_GRADES.py...")
    
    # Usar importlib para importação mais controlada
    import importlib.util
    spec = importlib.util.spec_from_file_location("ROBO_GRADES", robo_path)
    robo_module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(robo_module)
    
    print("Importação realizada com sucesso!")
    
except ImportError as e:
    print(f"Erro ao importar: {e}")
    traceback.print_exc()
    sys.exit(1)
except Exception as e:
    print(f"Erro desconhecido: {e}")
    traceback.print_exc()
    sys.exit(1)

class GradeAutomation:
    def __init__(self):
        self.linhas_grades = {
            # Campos Gerais
            'pavimento': 3,               # linha do campo "Pavimento"
            'nome_arquivo': 4,            # linha do campo "Nome do Arquivo"
            'altura_base': 12,            # linha do campo "Altura" (sarrafos verticais)

            # Larguras das Grades - Grupo 1
            'grade1_largura': 180,        # linha da largura da Grade 1 (Grupo 1)
            'grade2_largura': 183,        # linha da largura da Grade 2 (Grupo 1)
            'grade3_largura': 185,        # linha da largura da Grade 3 (Grupo 1)

            # Distâncias entre Grades - Grupo 1
            'distancia_1': 182,           # linha da distância entre Grade 1 e Grade 2 (Grupo 1)
            'distancia_2': 184,           # linha da distância entre Grade 2 e Grade 3 (Grupo 1)
            
            # Mantidos para compatibilidade
            'distancia_grade1_2': 182,    # Alias para distancia_1
            'distancia_grade2_3': 184,    # Alias para distancia_2
            
            # Larguras das Grades - Grupo 2
            'grade1_largura_grupo2': 219,    # Grade 1 do Grupo 2
            'grade2_largura_grupo2': 221,    # Grade 2 do Grupo 2
            'grade3_largura_grupo2': 223,    # Grade 3 do Grupo 2
            
            # Distâncias entre Grades - Grupo 2 (Grade B)
            'distancia_1_grupo2': 220,    # Distância 1 do Grupo 2
            'distancia_2_grupo2': 222,    # Distância 2 do Grupo 2

            # Detalhes das Grades - Distâncias dos Sarrafos
            'distancia_1_grade_1': 193,   # linha da distância do sarrafo 1 da grade 1
            'distancia_2_grade_1': 194,   # linha da distância do sarrafo 2 da grade 1
            'distancia_3_grade_1': 195,   # linha da distância do sarrafo 3 da grade 1
            'distancia_1_grade_2': 198,   # linha da distância do sarrafo 1 da grade 2
            'distancia_2_grade_2': 199,   # linha da distância do sarrafo 2 da grade 2
            'distancia_3_grade_2': 200,   # linha da distância do sarrafo 3 da grade 2
            'distancia_1_grade_3': 203,   # linha da distância do sarrafo 1 da grade 3
            'distancia_2_grade_3': 204,   # linha da distância do sarrafo 2 da grade 3
            'distancia_3_grade_3': 205,   # linha da distância do sarrafo 3 da grade 3

            # Grade 1
            'grade1': {
                'largura': 180,           # linha do campo Largura (Grade 1)
                'sarr_esquerda': 13,     # linha do campo Sarrafo Esquerdo (altura)
                'sarr1_altura': 13,      # linha do campo Altura Sarrafo 1
                'sarr2_altura': 13,      # linha do campo Altura Sarrafo 2
                'sarr3_altura': 13,      # linha do campo Altura Sarrafo 3
                'sarr_direita': 13,      # linha do campo Sarrafo Direito (altura)
            },

            # Grade 2
            'grade2': { 
                'largura': 183,           # linha do campo Largura (Grade 2) - CORRIGIDO
                'sarr_esquerda': 13,     # linha do campo Sarrafo Esquerdo (altura)
                'sarr1_altura': 13,      # linha do campo Altura Sarrafo 1
                'sarr2_altura': 13,      # linha do campo Altura Sarrafo 2
                'sarr3_altura': 13,      # linha do campo Altura Sarrafo 3
                'sarr_direita': 13,      # linha do campo Sarrafo Direito (altura)
            },

            # Grade 3
            'grade3': {
                'largura': 185,           # linha do campo Largura (Grade 3) - CORRIGIDO
                'sarr_esquerda': 13,     # linha do campo Sarrafo Esquerdo (altura)
                'sarr1_altura': 13,      # linha do campo Altura Sarrafo 1
                'sarr2_altura': 13,      # linha do campo Altura Sarrafo 2
                'sarr3_altura': 13,      # linha do campo Altura Sarrafo 3
                'sarr_direita': 13,      # linha do campo Sarrafo Direito (altura)
            },
            
            # Detalhes das Grades (larguras dos retângulos verticais)
            'detalhe_grade1_1': 192,
            'detalhe_grade1_2': 193,
            'detalhe_grade1_3': 194,
            'detalhe_grade1_4': 195,
            'detalhe_grade1_5': 196,
            'detalhe_grade2_1': 197,
            'detalhe_grade2_2': 198,
            'detalhe_grade2_3': 199,
            'detalhe_grade2_4': 200,
            'detalhe_grade2_5': 201,
            'detalhe_grade3_1': 202,
            'detalhe_grade3_2': 203,
            'detalhe_grade3_3': 204,
            'detalhe_grade3_4': 205,
            'detalhe_grade3_5': 206,
            
            # Detalhes das Grades - Grupo 2 (Grade B)
            'detalhe_grade1_1_grupo2': 518,
            'detalhe_grade1_2_grupo2': 519,
            'detalhe_grade1_3_grupo2': 520,
            'detalhe_grade1_4_grupo2': 521,
            'detalhe_grade1_5_grupo2': 522,
            'detalhe_grade2_1_grupo2': 523,
            'detalhe_grade2_2_grupo2': 524,
            'detalhe_grade2_3_grupo2': 525,
            'detalhe_grade2_4_grupo2': 526,
            'detalhe_grade2_5_grupo2': 527,
            'detalhe_grade3_1_grupo2': 528,
            'detalhe_grade3_2_grupo2': 529,
            'detalhe_grade3_3_grupo2': 530,
            'detalhe_grade3_4_grupo2': 531,
            'detalhe_grade3_5_grupo2': 532,
            
            # Altura dos Detalhes das Grades - Grade A
            'altura_detalhe_grade_a_1_0': 207,  # Campo 0 - Altura do sarrafo da extremidade esquerda
            'altura_detalhe_grade_a_1_1': 208,
            'altura_detalhe_grade_a_1_2': 209,
            'altura_detalhe_grade_a_1_3': 210,
            'altura_detalhe_grade_a_1_4': 211,
            'altura_detalhe_grade_a_1_5': 212,
            'altura_detalhe_grade_a_2_0': 213,  # Campo 0 - Altura do sarrafo da extremidade esquerda
            'altura_detalhe_grade_a_2_1': 214,
            'altura_detalhe_grade_a_2_2': 215,
            'altura_detalhe_grade_a_2_3': 216,
            'altura_detalhe_grade_a_2_4': 217,
            'altura_detalhe_grade_a_2_5': 218,
            'altura_detalhe_grade_a_3_0': 250,  # Campo 0 - Altura do sarrafo da extremidade esquerda
            'altura_detalhe_grade_a_3_1': 251,
            'altura_detalhe_grade_a_3_2': 252,
            'altura_detalhe_grade_a_3_3': 253,
            'altura_detalhe_grade_a_3_4': 254,
            'altura_detalhe_grade_a_3_5': 255,
            
            # === ALTURAS DOS DETALHES - CONJUNTO 2 (GRADE B) ===
            # Grade B - 1 (linhas 256-261)
            'altura_detalhe_grade_b_1_0': 256,
            'altura_detalhe_grade_b_1_1': 257,
            'altura_detalhe_grade_b_1_2': 258,
            'altura_detalhe_grade_b_1_3': 259,
            'altura_detalhe_grade_b_1_4': 260,
            'altura_detalhe_grade_b_1_5': 261,
            
            # Grade B - 2 (linhas 262-267)
            'altura_detalhe_grade_b_2_0': 262,
            'altura_detalhe_grade_b_2_1': 263,
            'altura_detalhe_grade_b_2_2': 264,
            'altura_detalhe_grade_b_2_3': 265,
            'altura_detalhe_grade_b_2_4': 266,
            'altura_detalhe_grade_b_2_5': 267,
            
            # Grade B - 3 (linhas 268-273)
            'altura_detalhe_grade_b_3_0': 268,
            'altura_detalhe_grade_b_3_1': 269,
            'altura_detalhe_grade_b_3_2': 270,
            'altura_detalhe_grade_b_3_3': 271,
            'altura_detalhe_grade_b_3_4': 272,
            'altura_detalhe_grade_b_3_5': 273
        }
        self.dados_grades = {}  # Inicializa o dicionário de dados das grades
        # Removido self.app pois não usamos mais xlwings

    def __del__(self):
        """Destrutor - não precisa fechar Excel com openpyxl"""
        pass

    def coluna_para_letra(self, col_idx):
        letra = ''
        while col_idx >= 0:
            letra = chr(ord('A') + (col_idx % 26)) + letra
            col_idx = col_idx // 26 - 1
        return letra
            
    def letra_para_coluna(self, letra):
        col_idx = 0
        for i, char in enumerate(reversed(letra)):
            col_idx += (ord(char) - ord('A') + 1) * (26 ** i)
        return col_idx - 1

    def formatar_valor(self, valor):
        """Formata um valor numérico para string com 2 casas decimais."""
        if valor is None:
            return "0.00"
        try:
            return f"{float(valor):.2f}"
        except (ValueError, TypeError):
            return "0.00"

    def gerar_script_grade(self, dados_grade):
        """Gera o script SQL para inserir os dados da grade."""
        script = []
        script.append("-- Script gerado automaticamente para inserção de dados da grade")
        script.append("-- Data de geração: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        script.append("")
        
        for grade, dados in dados_grade.items():
            script.append(f"-- Dados da Grade {grade}")
            script.append(f"INSERT INTO grade (numero_grade, altura, largura) VALUES")
            script.append(f"({grade}, {dados['altura']}, {dados['largura']});")
            script.append("")
        
        return "\n".join(script)

    def gerar_script_automatico(self, dados_grade):
        """Gera o script SQL para inserir os dados da grade."""
        script = []
        script.append("-- Script gerado automaticamente para inserção de dados da grade")
        script.append("-- Data de geração: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        script.append("")
        
        for grade, dados in dados_grade.items():
            script.append(f"-- Dados da Grade {grade}")
            script.append(f"INSERT INTO grade (numero_grade, altura, largura) VALUES")
            script.append(f"({grade}, {dados['altura']}, {dados['largura']});")
            script.append("")
        
        return "\n".join(script)

    def calcular_valor_formula(self, sheet, formula, coluna):
        """Calcula o valor de uma fórmula do Excel usando openpyxl."""
        try:
            # Se não for uma fórmula, retorna o valor direto
            if not formula or not isinstance(formula, str):
                print(f"Valor não é uma fórmula: {formula}")
                # Se for string com vírgula, converter para float
                if isinstance(formula, str) and ',' in formula:
                    try:
                        return float(formula.replace(',', '.'))
                    except:
                        return formula
                return formula
            
            # Se não começar com =, retorna o valor direto
            if not formula.startswith('='):
                print(f"Valor não começa com =: {formula}")
                # Se for string com vírgula, converter para float
                if isinstance(formula, str) and ',' in formula:
                    try:
                        return float(formula.replace(',', '.'))
                    except:
                        return formula
                return formula
            
            # Para openpyxl, não podemos calcular fórmulas automaticamente
            # Vamos tentar extrair valores numéricos simples
            try:
                # Remover o = e tentar avaliar como expressão Python simples
                formula_clean = formula[1:]  # Remove o =
                
                # Substituir referências de células por valores
                # Exemplo: A1+B1 -> valor_celula_A1 + valor_celula_B1
                import re
                cell_refs = re.findall(r'[A-Z]+\d+', formula_clean)
                
                for ref in cell_refs:
                    # Converter referência de célula para coordenadas
                    col = ref.rstrip('0123456789')
                    row = int(ref[len(col):])
                    
                    # Obter valor da célula
                    cell_value = sheet[f'{col}{row}'].value
                    if cell_value is not None:
                        formula_clean = formula_clean.replace(ref, str(cell_value))
                
                # Tentar avaliar a expressão
                try:
                    valor_calculado = eval(formula_clean)
                    print(f"Fórmula {formula} calculada: {valor_calculado}")
                    return valor_calculado
                except:
                    # Se não conseguir calcular, retorna a fórmula original
                    print(f"Não foi possível calcular fórmula {formula}, retornando original")
                    return formula
            except:
                # Se não conseguir processar a fórmula, retorna a fórmula original
                print(f"Erro ao processar fórmula {formula}, retornando original")
                return formula
        except:
            # Se não conseguir processar a fórmula, retorna a fórmula original
            print(f"Erro geral ao processar fórmula {formula}, retornando original")
            return formula

    def calcular_posicoes_sarrafos(self, largura):
        """
        Calcula as posições dos sarrafos centrais baseado na largura da grade.
        Nova regra: tamanho da grade / 30.3 arredondado para baixo
        Exceção: grades acima de 120cm permitem até 4 sarrafos centrais
        
        Args:
            largura (float): Largura total da grade
            
        Returns:
            list: Lista com as posições dos sarrafos [pos1, pos2, pos3] ou [pos1, pos2, pos3, pos4]
        """
        try:
            largura = float(largura)
            
            # Calcular quantidade de sarrafos centrais: largura / 30.3 arredondado para baixo
            quantidade_sarrafos = math.floor(largura / 30.3)
            
            # Aplicar limite: 3 para grades <= 120, 4 para grades > 120
            if largura > 120:
                quantidade_sarrafos = min(4, quantidade_sarrafos)
            else:
                quantidade_sarrafos = min(3, quantidade_sarrafos)
            
            print(f"Grade de {largura}cm: {largura}/30.3 = {largura/30.3:.2f} -> {quantidade_sarrafos} sarrafos centrais")
            
            if quantidade_sarrafos <= 0:
                return [0, 0, 0, 0]  # Nenhum sarrafo central
                
            elif quantidade_sarrafos == 1:
                # 1 sarrafo central - posição central
                centro = largura / 2
                return [centro, 0, 0, 0]
                
            elif quantidade_sarrafos == 2:
                # 2 sarrafos centrais - dividir em 3 partes
                pos1 = largura / 3
                pos2 = 2 * largura / 3
                return [pos1, pos2, 0, 0]
                
            elif quantidade_sarrafos == 3:
                # 3 sarrafos centrais - usar 3 posições
                pos1 = largura / 4
                pos2 = largura / 2  # Central
                pos3 = 3 * largura / 4
                return [pos1, pos2, pos3, 0]
                
            else:  # quantidade_sarrafos == 4 (apenas para grades > 120)
                # 4 sarrafos centrais - dividir em 5 partes
                pos1 = largura / 5
                pos2 = 2 * largura / 5
                pos3 = 3 * largura / 5
                pos4 = 4 * largura / 5
                return [pos1, pos2, pos3, pos4]
                
        except (ValueError, TypeError):
            return [0, 0, 0, 0]

    def calcular_quantidade_sarrafos_centrais(self, largura):
        """
        Calcula apenas a quantidade de sarrafos centrais baseada na largura da grade.
        Regra: tamanho da grade / 30.3 arredondado para baixo
        Exceção: grades acima de 120cm permitem até 4 sarrafos centrais
        
        Args:
            largura (float): Largura total da grade
            
        Returns:
            int: Quantidade de sarrafos centrais (0, 1, 2, 3 ou 4 se largura > 120)
        """
        try:
            largura = float(largura)
            quantidade = math.floor(largura / 30.3)
            # Exceção: grades acima de 120cm permitem até 4 sarrafos
            if largura > 120:
                return max(0, min(4, quantidade))  # Limitar entre 0 e 4 para grades > 120
            else:
                return max(0, min(3, quantidade))  # Limitar entre 0 e 3 para grades <= 120
        except (ValueError, TypeError):
            return 0

    def processar_planilha(self, caminho_arquivo_excel, coluna_especifica=None):
        print(f"\n>>> [DEBUG GRADE_EXCEL] processar_planilha INICIADO")
        print(f">>> [DEBUG GRADE_EXCEL]   - caminho_arquivo_excel: {caminho_arquivo_excel}")
        print(f">>> [DEBUG GRADE_EXCEL]   - coluna_especifica: {coluna_especifica}")
        
        try:
            # Abrir o Excel com openpyxl
            print(f">>> [DEBUG GRADE_EXCEL] Abrindo Excel com openpyxl...")
            wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
            sheet = wb.active
            print(f">>> [DEBUG GRADE_EXCEL] ✅ Excel aberto com sucesso")
            
            print("Planilha carregada com sucesso")
            print(f"Número de linhas na planilha: {sheet.max_row}")
            print(f"Número de colunas na planilha: {sheet.max_column}")
            
            coluna_base = 'E'
            coluna_base_idx = ord(coluna_base) - ord('A')
            colunas_vazias = 0
            parar_busca = False

            if coluna_especifica:
                colunas_a_processar = [self.letra_para_coluna(coluna_especifica)]
                print(f"Processando apenas a coluna {coluna_especifica}")
                print(f">>> [DEBUG GRADE_EXCEL] Processando apenas a coluna {coluna_especifica}")
            else:
                colunas_a_processar = range(coluna_base_idx, sheet.max_column)
                print(f"Processando todas as colunas a partir de {coluna_base}")
                print(f">>> [DEBUG GRADE_EXCEL] Processando todas as colunas a partir de {coluna_base} (total: {len(list(colunas_a_processar))})")

            for col_idx in colunas_a_processar:
                if parar_busca:
                    print(f">>> [DEBUG GRADE_EXCEL] ⚠️ Parando busca (parar_busca=True)")
                    break

                coluna = self.coluna_para_letra(col_idx)
                print(f">>> [DEBUG GRADE_EXCEL] Processando coluna {coluna} (índice {col_idx})")
                nome_arquivo = sheet[f'{coluna}{self.linhas_grades["nome_arquivo"]}'].value
                print(f">>> [DEBUG GRADE_EXCEL]   - nome_arquivo lido: {nome_arquivo}")
                
                if nome_arquivo:
                    colunas_vazias = 0
                    print(f"\nProcessando coluna {coluna} para o arquivo '{nome_arquivo}'")
                    print(f">>> [DEBUG GRADE_EXCEL] ✅ Processando coluna {coluna} para o arquivo '{nome_arquivo}'")
                    
                    # Ler altura base
                    try:
                        # Ler a fórmula do campo 12
                        formula_altura = sheet[f'{coluna}12'].value
                        print(f"\nLendo fórmula da altura (campo 12):")
                        print(f"Fórmula {coluna}12: {formula_altura}")
                        
                        # Calcular o valor da altura
                        altura_base = self.calcular_valor_formula(sheet, formula_altura, coluna)
                        print(f"Altura base calculada: {altura_base}")
                            
                    except Exception as e:
                        print(f"Erro ao ler altura base: {e}")
                        altura_base = 0.0
                    
                    # Usar apenas o nome original, sem sufixo de coluna
                    # nome_arquivo_coluna = nome_arquivo
                    # Ler o valor do pavimento
                    try:
                        pavimento = sheet[f'{coluna}{self.linhas_grades["pavimento"]}'].value
                        print(f"Valor do pavimento lido: {pavimento}")
                        
                        # Garantir que o pavimento seja uma string válida
                        if pavimento is not None:
                            pavimento = str(pavimento).strip()
                            if not pavimento:  # Se estiver vazio após remover espaços
                                print("AVISO: Pavimento vazio, usando valor padrão '200'")
                                pavimento = "200"
                        else:
                            print("AVISO: Pavimento None, usando valor padrão '200'")
                            pavimento = "200"
                            
                        print(f"Pavimento final: {pavimento}")
                    except Exception as e:
                        print(f"Erro ao ler pavimento: {e}")
                        pavimento = "200"  # Valor padrão para pavimento
                    
                    # Ler distâncias do Grupo 2 (Grade B) do Excel
                    distancia_1_grupo2_valor = "0.00"
                    distancia_2_grupo2_valor = "0.00"
                    
                    print(f"\n[DEBUG DISTÂNCIAS GRUPO 2] Mapeamento das linhas:")
                    print(f"  - distancia_1_grupo2: linha {self.linhas_grades.get('distancia_1_grupo2', 'N/A')} (linha 220 do Excel)")
                    print(f"  - distancia_2_grupo2: linha {self.linhas_grades.get('distancia_2_grupo2', 'N/A')} (linha 222 do Excel)")
                    
                    if 'distancia_1_grupo2' in self.linhas_grades:
                        try:
                            linha_dist1_g2 = self.linhas_grades["distancia_1_grupo2"]
                            celula_dist1_g2 = f'{coluna}{linha_dist1_g2}'
                            formula_dist1_g2 = sheet[celula_dist1_g2].value
                            print(f"[DEBUG GRUPO 2] Lendo distância 1 do Grupo 2 da célula {celula_dist1_g2}: {formula_dist1_g2}")
                            distancia_1_grupo2_valor = self.formatar_valor(self.calcular_valor_formula(sheet, formula_dist1_g2, coluna))
                            print(f"[DEBUG GRUPO 2] Distância 1 do Grupo 2 lida e formatada: {distancia_1_grupo2_valor} cm")
                        except Exception as e:
                            print(f"[DEBUG GRUPO 2] Erro ao ler distancia_1_grupo2 da célula {celula_dist1_g2}: {e}")
                    
                    if 'distancia_2_grupo2' in self.linhas_grades:
                        try:
                            linha_dist2_g2 = self.linhas_grades["distancia_2_grupo2"]
                            celula_dist2_g2 = f'{coluna}{linha_dist2_g2}'
                            formula_dist2_g2 = sheet[celula_dist2_g2].value
                            print(f"[DEBUG GRUPO 2] Lendo distância 2 do Grupo 2 da célula {celula_dist2_g2}: {formula_dist2_g2}")
                            distancia_2_grupo2_valor = self.formatar_valor(self.calcular_valor_formula(sheet, formula_dist2_g2, coluna))
                            print(f"[DEBUG GRUPO 2] Distância 2 do Grupo 2 lida e formatada: {distancia_2_grupo2_valor} cm")
                        except Exception as e:
                            print(f"[DEBUG GRUPO 2] Erro ao ler distancia_2_grupo2 da célula {celula_dist2_g2}: {e}")
                    
                    # Ler distâncias do Grupo 1 do Excel
                    distancia_1_grupo1_valor = "0.00"
                    distancia_2_grupo1_valor = "0.00"
                    
                    print(f"\n[DEBUG DISTÂNCIAS GRUPO 1] Mapeamento das linhas:")
                    print(f"  - distancia_1: linha {self.linhas_grades.get('distancia_1', 'N/A')} (linha 182 do Excel)")
                    print(f"  - distancia_2: linha {self.linhas_grades.get('distancia_2', 'N/A')} (linha 184 do Excel)")
                    
                    if 'distancia_1' in self.linhas_grades:
                        try:
                            linha_dist1 = self.linhas_grades["distancia_1"]
                            celula_dist1 = f'{coluna}{linha_dist1}'
                            formula_dist1_g1 = sheet[celula_dist1].value
                            print(f"[DEBUG GRUPO 1] Lendo distância 1 da célula {celula_dist1}: {formula_dist1_g1}")
                            distancia_1_grupo1_valor = self.formatar_valor(self.calcular_valor_formula(sheet, formula_dist1_g1, coluna))
                            print(f"[DEBUG GRUPO 1] Distância 1 lida e formatada: {distancia_1_grupo1_valor} cm")
                        except Exception as e:
                            print(f"[DEBUG GRUPO 1] Erro ao ler distancia_1 da célula {celula_dist1}: {e}")
                    
                    if 'distancia_2' in self.linhas_grades:
                        try:
                            linha_dist2 = self.linhas_grades["distancia_2"]
                            celula_dist2 = f'{coluna}{linha_dist2}'
                            formula_dist2_g1 = sheet[celula_dist2].value
                            print(f"[DEBUG GRUPO 1] Lendo distância 2 da célula {celula_dist2}: {formula_dist2_g1}")
                            distancia_2_grupo1_valor = self.formatar_valor(self.calcular_valor_formula(sheet, formula_dist2_g1, coluna))
                            print(f"[DEBUG GRUPO 1] Distância 2 lida e formatada: {distancia_2_grupo1_valor} cm")
                        except Exception as e:
                            print(f"[DEBUG GRUPO 1] Erro ao ler distancia_2 da célula {celula_dist2}: {e}")
                    
                    dados_grade = {
                        'pavimento': pavimento,  # Já está como string
                        'nome_arquivo': nome_arquivo,
                        'altura_base': self.formatar_valor(altura_base),
                        # GRUPO 1 (Grade A): distância entre grades
                        'distancia_1': distancia_1_grupo1_valor,  # Grade 1 → Grade 2, Grupo 1 (linha 182)
                        'distancia_2': distancia_2_grupo1_valor,  # Grade 2 → Grade 3, Grupo 1 (linha 184)
                        # Aliases para compatibilidade
                        'distancia_grade1_2': distancia_1_grupo1_valor,  # Grade 1→2 do Grupo 1
                        'distancia_grade2_3': distancia_2_grupo1_valor,  # Grade 2→3 do Grupo 1
                        # GRUPO 2 (Grade B): distância entre grades
                        'distancia_1_grupo2': distancia_1_grupo2_valor,  # Grade 1 → Grade 2, Grupo 2 (linha 220)
                        'distancia_2_grupo2': distancia_2_grupo2_valor,  # Grade 2 → Grade 3, Grupo 2 (linha 222)
                        'x_inicial': robo_module.config_manager.get_config("coordinates", "x_inicial"),
                        'y_inicial': robo_module.config_manager.get_config("coordinates", "y_inicial"),
                        'grades': {}
                    }
                    
                    # Debug das distâncias entre grades armazenadas
                    print(f"\n[DEBUG ARMAZENAMENTO] Distâncias adicionadas ao dicionário dados_grade:")
                    print(f"  [GRUPO 1 - Grade A]")
                    print(f"    - distancia_1 (linha 182): {dados_grade['distancia_1']} cm")
                    print(f"    - distancia_2 (linha 184): {dados_grade['distancia_2']} cm")
                    print(f"    - distancia_grade1_2 (alias): {dados_grade['distancia_grade1_2']} cm")
                    print(f"    - distancia_grade2_3 (alias): {dados_grade['distancia_grade2_3']} cm")
                    print(f"  [GRUPO 2 - Grade B]")
                    print(f"    - distancia_1_grupo2 (linha 220): {dados_grade['distancia_1_grupo2']} cm")
                    print(f"    - distancia_2_grupo2 (linha 222): {dados_grade['distancia_2_grupo2']} cm")

                    # Processar cada grade
                    for grade_num in range(1, 4):
                        grade_key = f'grade{grade_num}'
                        grade_dados = {}
                        
                        # Ler campos da grade
                        for campo, linha in self.linhas_grades[grade_key].items():
                            if linha is not None:  # Ignorar campos desativados
                                try:
                                    formula = sheet[f'{coluna}{linha}'].value
                                    print(f"Lendo {campo} da célula {coluna}{linha}: {formula}")
                                    valor = self.calcular_valor_formula(sheet, formula, coluna)
                                    grade_dados[campo] = self.formatar_valor(valor)
                                    print(f"[DEBUG GRADE {grade_num}] {campo}: {formula} -> {valor} -> {grade_dados[campo]}")
                                except Exception as e:
                                    print(f"Erro ao ler {campo}: {e}")
                                    grade_dados[campo] = "0.00"
                        
                        # Ler distâncias específicas do Excel para esta grade
                        distancias_grade = []
                        for sarr_num in range(1, 4):
                            campo_distancia = f'distancia_{sarr_num}_grade_{grade_num}'
                            if campo_distancia in self.linhas_grades:
                                try:
                                    linha_distancia = self.linhas_grades[campo_distancia]
                                    formula_distancia = sheet[f'{coluna}{linha_distancia}'].value
                                    print(f"Lendo distância {sarr_num} da grade {grade_num} da célula {coluna}{linha_distancia}: {formula_distancia}")
                                    valor_distancia = self.calcular_valor_formula(sheet, formula_distancia, coluna)
                                    distancias_grade.append(self.formatar_valor(valor_distancia))
                                    print(f"Distância {sarr_num} da grade {grade_num}: {valor_distancia}")
                                except Exception as e:
                                    print(f"Erro ao ler distância {sarr_num} da grade {grade_num}: {e}")
                                    distancias_grade.append("0.00")
                            else:
                                print(f"Campo {campo_distancia} não encontrado no mapeamento")
                                distancias_grade.append("0.00")
                        
                        # Se não conseguiu ler as distâncias do Excel, calcular automaticamente
                        if all(float(d) == 0.0 for d in distancias_grade):
                            print(f"Distâncias da grade {grade_num} não encontradas no Excel, calculando automaticamente")
                            largura = float(grade_dados.get('largura', 0))
                            posicoes = self.calcular_posicoes_sarrafos(largura)
                            distancias_grade = [str(posicoes[0]), str(posicoes[1]), str(posicoes[2])]
                            print(f"Distâncias calculadas automaticamente para grade {grade_num}: {distancias_grade}")
                        else:
                            print(f"Distâncias lidas do Excel para grade {grade_num}: {distancias_grade}")
                            
                            # TORNAR AS DISTÂNCIAS CUMULATIVAS
                            print(f"Aplicando cálculo cumulativo para grade {grade_num}")
                            distancias_cumulativas = []
                            acumulador = 0.0
                            
                            for i, dist_str in enumerate(distancias_grade):
                                dist_valor = float(dist_str)
                                if dist_valor > 0:
                                    acumulador += dist_valor
                                    distancias_cumulativas.append(str(acumulador))
                                    print(f"  Sarrafo {i+1}: {dist_valor} + acumulado = {acumulador}")
                                else:
                                    distancias_cumulativas.append("0.00")
                                    print(f"  Sarrafo {i+1}: 0.00 (sem distância)")
                            
                            distancias_grade = distancias_cumulativas
                            print(f"Distâncias cumulativas finais para grade {grade_num}: {distancias_grade}")
                        
                        # Adicionar as distâncias aos dados da grade
                        grade_dados['sarr1_distancia'] = distancias_grade[0]
                        grade_dados['sarr2_distancia'] = distancias_grade[1]
                        grade_dados['sarr3_distancia'] = distancias_grade[2]
                        
                        dados_grade['grades'][grade_key] = grade_dados
                        print(f"Dados da grade {grade_num} processados: {grade_dados}")
                    
                    # Ler larguras dos detalhes das grades
                    print("\nLendo larguras dos detalhes das grades...")
                    
                    # Larguras dos detalhes das grades
                    larguras_detalhes_grades = {}
                    for grade_num in range(1, 4):
                        for detalhe_num in range(1, 6):
                            campo_key = f'detalhe_grade{grade_num}_{detalhe_num}'
                            if campo_key in self.linhas_grades:
                                try:
                                    linha = self.linhas_grades[campo_key]
                                    formula = sheet[f'{coluna}{linha}'].value
                                    print(f"Lendo {campo_key} da célula {coluna}{linha}: {formula}")
                                    valor = self.calcular_valor_formula(sheet, formula, coluna)
                                    larguras_detalhes_grades[campo_key] = self.formatar_valor(valor)
                                    print(f"[DEBUG LARGURA] Grade {grade_num}-{detalhe_num}: {formula} -> {valor} -> {larguras_detalhes_grades[campo_key]}")
                                except Exception as e:
                                    print(f"Erro ao ler {campo_key}: {e}")
                                    larguras_detalhes_grades[campo_key] = "0.00"
                    
                    # Ler alturas dos detalhes das grades
                    print("\nLendo alturas dos detalhes das grades...")
                    
                    # Alturas dos detalhes Grade A
                    alturas_detalhes_grade_a = {}
                    for grade_num in range(1, 4):
                        for detalhe_num in range(6):  # Alterado de range(1, 6) para range(6) para incluir campo 0
                            campo_key = f'altura_detalhe_grade_a_{grade_num}_{detalhe_num}'
                            if campo_key in self.linhas_grades:
                                try:
                                    linha = self.linhas_grades[campo_key]
                                    formula = sheet[f'{coluna}{linha}'].value
                                    print(f"Lendo {campo_key} da célula {coluna}{linha}: {formula}")
                                    valor = self.calcular_valor_formula(sheet, formula, coluna)
                                    alturas_detalhes_grade_a[campo_key] = self.formatar_valor(valor)
                                    print(f"[DEBUG ALTURA A] {campo_key}: {formula} -> {valor} -> {alturas_detalhes_grade_a[campo_key]}")
                                except Exception as e:
                                    print(f"Erro ao ler {campo_key}: {e}")
                                    alturas_detalhes_grade_a[campo_key] = "0.00"
                            else:
                                print(f"❌ Campo {campo_key} NÃO ENCONTRADO em linhas_grades!")
                                alturas_detalhes_grade_a[campo_key] = "0.00"
                    
                    # Alturas dos detalhes Grade B
                    alturas_detalhes_grade_b = {}
                    for grade_num in range(1, 4):
                        for detalhe_num in range(6):  # Alterado de range(1, 6) para range(6) para incluir campo 0
                            campo_key = f'altura_detalhe_grade_b_{grade_num}_{detalhe_num}'
                            if campo_key in self.linhas_grades:
                                try:
                                    linha = self.linhas_grades[campo_key]
                                    formula = sheet[f'{coluna}{linha}'].value
                                    print(f"Lendo {campo_key} da célula {coluna}{linha}: {formula}")
                                    valor = self.calcular_valor_formula(sheet, formula, coluna)
                                    alturas_detalhes_grade_b[campo_key] = self.formatar_valor(valor)
                                    print(f"[DEBUG ALTURA B] {campo_key}: {formula} -> {valor} -> {alturas_detalhes_grade_b[campo_key]}")
                                except Exception as e:
                                    print(f"Erro ao ler {campo_key}: {e}")
                                    alturas_detalhes_grade_b[campo_key] = "0.00"
                    
                    # Ler larguras dos detalhes do Grupo 2
                    print("\nLendo larguras dos detalhes do Grupo 2...")
                    larguras_detalhes_grades_grupo2 = {}
                    for grade_num in range(1, 4):
                        for detalhe_num in range(1, 6):
                            campo_key = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                            if campo_key in self.linhas_grades:
                                try:
                                    linha = self.linhas_grades[campo_key]
                                    celula = f'{coluna}{linha}'
                                    # Ler valor bruto da célula
                                    valor_bruto = sheet[celula].value
                                    print(f"[DEBUG LEITURA GRUPO 2] Lendo {campo_key} da célula {celula}")
                                    print(f"[DEBUG LEITURA GRUPO 2]   - Valor bruto da célula: {valor_bruto} (tipo: {type(valor_bruto)})")
                                    print(f"[DEBUG LEITURA GRUPO 2]   - Célula tem fórmula: {sheet[celula].data_type == 'f'}")
                                    print(f"[DEBUG LEITURA GRUPO 2]   - Célula tem valor: {sheet[celula].data_type == 'n'}")
                                    formula = valor_bruto
                                    print(f"Lendo {campo_key} da célula {celula}: {formula}")
                                    valor = self.calcular_valor_formula(sheet, formula, coluna)
                                    print(f"[DEBUG LEITURA GRUPO 2]   - Valor após calcular_valor_formula: {valor} (tipo: {type(valor)})")
                                    larguras_detalhes_grades_grupo2[campo_key] = self.formatar_valor(valor)
                                    print(f"[DEBUG LARGURA GRUPO 2] Grade {grade_num}-{detalhe_num}: {formula} -> {valor} -> {larguras_detalhes_grades_grupo2[campo_key]}")
                                except Exception as e:
                                    print(f"[DEBUG LEITURA GRUPO 2] ERRO ao ler {campo_key}: {e}")
                                    import traceback
                                    print(f"[DEBUG LEITURA GRUPO 2] Traceback: {traceback.format_exc()}")
                                    larguras_detalhes_grades_grupo2[campo_key] = "0.00"
                    
                    # Adicionar as larguras e alturas dos detalhes aos dados
                    dados_grade['larguras_detalhes_grades'] = larguras_detalhes_grades
                    dados_grade['detalhes_grades_grupo2'] = larguras_detalhes_grades_grupo2  # Adicionar detalhes do Grupo 2
                    dados_grade['alturas_detalhes_grade_a'] = alturas_detalhes_grade_a
                    dados_grade['alturas_detalhes_grade_b'] = alturas_detalhes_grade_b
                    print(f"Alturas dos detalhes Grade A: {alturas_detalhes_grade_a}")
                    print(f"Alturas dos detalhes Grade B: {alturas_detalhes_grade_b}")
                    print(f"Larguras dos detalhes Grupo 2: {larguras_detalhes_grades_grupo2}")
                    
                    # ===== LER GRADES DO GRUPO 2 DO EXCEL =====
                    print("\n[GRUPO 2] Lendo grades do Grupo 2 do Excel...")
                    grades_grupo2 = {}
                    
                    for grade_num in range(1, 4):
                        grade_key = f'grade{grade_num}'
                        grade_dados_grupo2 = {}
                        
                        # Ler largura do Grupo 2
                        largura_key = f'grade{grade_num}_largura_grupo2'
                        if largura_key in self.linhas_grades:
                            try:
                                linha_largura = self.linhas_grades[largura_key]
                                formula_largura = sheet[f'{coluna}{linha_largura}'].value
                                print(f"[GRUPO 2] Lendo largura {grade_key} da célula {coluna}{linha_largura}: {formula_largura}")
                                valor_largura = self.calcular_valor_formula(sheet, formula_largura, coluna)
                                grade_dados_grupo2['largura'] = self.formatar_valor(valor_largura)
                                print(f"[GRUPO 2] Grade {grade_num} largura: {valor_largura}")
                            except Exception as e:
                                print(f"[GRUPO 2] Erro ao ler largura grade {grade_num}: {e}")
                                grade_dados_grupo2['largura'] = "0.00"
                        
                        # Copiar outros campos da grade do Grupo 1 como fallback
                        if grade_key in dados_grade['grades']:
                            # Salvar a largura do Grupo 2 antes de copiar
                            largura_g2 = grade_dados_grupo2.get('largura', '0.00')
                            # Copiar todos os campos do Grupo 1
                            grade_dados_grupo2.update(dados_grade['grades'][grade_key])
                            # Restaurar a largura do Grupo 2
                            grade_dados_grupo2['largura'] = largura_g2
                        
                        grades_grupo2[grade_key] = grade_dados_grupo2
                        print(f"[GRUPO 2] Grade {grade_num} dados: {grade_dados_grupo2}")
                    
                    dados_grade['grades_grupo2_excel'] = grades_grupo2
                    print(f"[GRUPO 2] {len(grades_grupo2)} grades Grupo 2 lidas do Excel")
                    
                    # Adicionar os dados da grade ao dicionário de dados
                    for sufixo in ['A', 'B']:
                        nome_arquivo_suf = f"{nome_arquivo}.{sufixo}"
                        dados_grade_suf = dados_grade.copy()
                        dados_grade_suf['nome_arquivo'] = nome_arquivo_suf
                        # Deepcopy grades para não compartilhar referência
                        import copy
                        
                        # IMPORTANTE: .A usa Grupo 1, .B usa Grupo 2
                        if sufixo == 'A':
                            # Grade A → usar dados do Grupo 1
                            dados_grade_suf['grades'] = copy.deepcopy(dados_grade['grades'])
                            print(f"[SEPARAÇÃO A/B] Arquivo .A → Usando Grades Grupo 1")
                        elif sufixo == 'B' and 'grades_grupo2_excel' in dados_grade:
                            # Grade B → usar dados do Grupo 2
                            dados_grade_suf['grades'] = copy.deepcopy(dados_grade['grades_grupo2_excel'])
                            print(f"[SEPARAÇÃO A/B] Arquivo .B → Usando Grades Grupo 2")
                        else:
                            # Fallback para Grupo 1
                            dados_grade_suf['grades'] = copy.deepcopy(dados_grade['grades'])
                            print(f"[SEPARAÇÃO A/B] Arquivo .{sufixo} → Fallback para Grupo 1")
                        
                        self.dados_grades[nome_arquivo_suf] = dados_grade_suf
                        print(f"[SEPARAÇÃO A/B] Dados da grade '{nome_arquivo_suf}' adicionados")
                else:
                    colunas_vazias += 1
                    if colunas_vazias >= 3:
                        parar_busca = True
                        print(f"Encontradas {colunas_vazias} colunas vazias consecutivas. Parando a busca.")

            # Fechar o Excel
            wb.close()

            print("\nProcessamento da planilha concluído com sucesso!")
            print(f"dados_grades: {self.dados_grades}")
            print(f"Tipo de dados_grades: {type(self.dados_grades)}")
            print(f"Tamanho de dados_grades: {len(self.dados_grades) if isinstance(self.dados_grades, dict) else 'N/A'}")
            return self.dados_grades
            
        except Exception as e:
            print(f"Erro ao processar a planilha: {e}")
            import traceback
            traceback.print_exc()
            return False

    def salvar_arquivo_utf16(self, caminho_arquivo, conteudo):
        """Salva um arquivo em UTF-16 LE."""
        try:
            # Garantir que o conteúdo seja uma string
            if not isinstance(conteudo, str):
                conteudo = str(conteudo)
            
            # Garantir que as linhas terminem com \n
            linhas = conteudo.split('\n')
            conteudo = '\n'.join(linhas)
            
            # Adicionar BOM (Byte Order Mark) para UTF-16 LE
            bom = b'\xFF\xFE'
            
            # Converter o conteúdo para bytes em UTF-16 LE
            conteudo_bytes = conteudo.encode('utf-16-le')
            
            # Salvar o arquivo com BOM
            with open(caminho_arquivo, 'wb') as f:
                f.write(bom)
                f.write(conteudo_bytes)
            return True
        except Exception as e:
            print(f"Erro ao salvar arquivo em UTF-16 LE: {e}")
            return False

    def salvar_script_sql(self, caminho_arquivo):
        """Salva os dados processados em um arquivo SQL."""
        try:
            script = []
            script.append("-- Script gerado automaticamente para inserção de dados das grades")
            script.append("-- Data de geração: " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            script.append("")
            
            for nome_arquivo, dados in self.dados_grades.items():
                script.append(f"-- Dados para o arquivo: {nome_arquivo}")
                script.append(f"INSERT INTO grade (pavimento, nome_arquivo, altura_base, distancia_grade1_2, distancia_grade2_3) VALUES")
                script.append(f"('{dados['pavimento']}', '{dados['nome_arquivo']}', {dados['altura_base']}, {dados['distancia_grade1_2']}, {dados['distancia_grade2_3']});")
                script.append("")
                
                for grade_num in range(1, 4):
                    grade_key = f'grade{grade_num}'
                    grade_dados = dados['grades'][grade_key]
                    
                    # Usar a altura base como altura da grade
                    altura = dados['altura_base']
                    
                    script.append(f"-- Dados da Grade {grade_num} para o arquivo {nome_arquivo}")
                    script.append(f"INSERT INTO grade_detalhes (nome_arquivo, numero_grade, altura, largura) VALUES")
                    script.append(f"('{nome_arquivo}', {grade_num}, {altura}, {grade_dados['largura']});")
                    script.append("")
            
            if self.salvar_arquivo_utf16(caminho_arquivo, "\n".join(script)):
                print(f"Script SQL salvo com sucesso em: {caminho_arquivo}")
                return True
            return False
            
        except Exception as e:
            print(f"Erro ao salvar script SQL: {e}")
            return False

    def ajustar_altura(self, altura):
        """Usa exatamente a altura informada (sem arredondar)."""
        try:
            valor = float(altura)
        except (TypeError, ValueError):
            return 0
        return valor if valor > 0 else 0

    def calcular_num_retangulos_horizontais(self, altura):
        """Calcula o número de retângulos horizontais baseado na altura original"""
        # Mantém exatamente o valor informado
        altura_final = float(altura) if altura and float(altura) > 0 else 0
        
        # Obter as posições configuradas
        positions = robo_module.config_manager.get_config("horizontal_positions", "positions")
        
        # Contar quantos retângulos horizontais são necessários
        # (todos que tiverem valor inferior à altura ajustada - 10)
        num_retangulos = 0
        altura_limite = altura_final - 10
        
        for pos in positions:
            if pos <= altura_limite:
                num_retangulos += 1
        
        return num_retangulos

    def calcular_posicao_horizontal(self, i, altura_base, altura_minima=None):
        """Calcula a posição Y do i-ésimo retângulo horizontal"""
        # Obter as posições configuradas
        positions = robo_module.config_manager.get_config("horizontal_positions", "positions")
        
        # Se o índice estiver dentro do range das posições configuradas
        if i < len(positions):
            posicao = altura_base + positions[i]
            
            # Se houver altura mínima, usar ela como limite
            if altura_minima is not None:
                altura_limite = altura_base + altura_minima - 10
                # Retornar a posição se estiver dentro do limite
                if posicao <= altura_limite:
                    return posicao
        
        # Se o índice estiver fora do range ou a posição estiver fora do limite, retornar None
        return None

    def gerar_comandos_retangulo(self, x1, y1, x2, y2, use_mline=False, is_horizontal=False):
        """
        Gera os comandos para desenhar um retângulo usando LINE ou MLINE
        
        Args:
            x1, y1: Coordenadas do ponto inferior esquerdo
            x2, y2: Coordenadas do ponto superior direito
            use_mline: Se True, usa MLINE; se False, usa LINE
            is_horizontal: Se True, é um retângulo horizontal (sarrafo horizontal)
        """
        if use_mline:
            if is_horizontal:  # Para retângulos horizontais
                # Identificar se é o retângulo base verificando se está no nível do solo (y1 == y_inicial)
                if y1 == robo_module.config_manager.get_config("coordinates", "y_inicial"):
                    mline_size = "2.2"  # Retângulo base
                else:
                    mline_size = "10"   # Sarrafos horizontais normais
                
                return f""";
_MLINE
ST
TRAVA
S
{mline_size}
{x1},{y2}
{x2},{y2}

;
"""
            else:  # Para retângulos verticais
                # Verificar se é central ou lateral
                if abs(x2 - x1) == 3.5:  # Se for central
                    mline_size = "3.5"
                    # Mover 3.5cm para a esquerda
                    x_pos = x2 - 3.5
                    return f""";
_MLINE
ST
MEIOPONT
S
{mline_size}
{x_pos},{y1}
{x_pos},{y2}

;
"""
                else:  # Se for lateral
                    mline_size = "7"
                    # Mover 7cm para a esquerda
                    x_pos = x2 - 7
                    return f""";
_MLINE
ST
PONT1
S
{mline_size}
{x_pos},{y1}
{x_pos},{y2}

;
"""
        else:
            # Para LINE, precisamos desenhar as 4 linhas do retângulo
            return f"""_LINE
{x1},{y1}
{x2},{y1}

;
_LINE
{x2},{y1}
{x2},{y2}

;
_LINE
{x2},{y2}
{x1},{y2}

;
_LINE
{x1},{y2}
{x1},{y1}

;
"""

    def processar_distancias_excel(self, dados_grade):
        """
        Processa as distâncias específicas lidas do Excel e atualiza os dados da grade.
        Esta função é chamada antes de gerar os scripts SCR para garantir que as distâncias
        do Excel sejam usadas em vez das calculadas automaticamente.
        """
        try:
            print(f"\nProcessando distâncias específicas do Excel para {dados_grade['nome_arquivo']}")
            
            for grade_num in range(1, 4):
                grade_key = f'grade{grade_num}'
                if grade_key in dados_grade['grades']:
                    grade_dados = dados_grade['grades'][grade_key]
                    
                    # Verificar se as distâncias foram lidas do Excel
                    sarr1_dist = float(grade_dados.get('sarr1_distancia', 0))
                    sarr2_dist = float(grade_dados.get('sarr2_distancia', 0))
                    sarr3_dist = float(grade_dados.get('sarr3_distancia', 0))
                    
                    print(f"Grade {grade_num} - Distâncias lidas do Excel:")
                    print(f"  Sarrafo 1: {sarr1_dist} cm")
                    print(f"  Sarrafo 2: {sarr2_dist} cm")
                    print(f"  Sarrafo 3: {sarr3_dist} cm")
                    
                    # Se alguma distância foi lida do Excel, usar essas distâncias
                    if sarr1_dist > 0 or sarr2_dist > 0 or sarr3_dist > 0:
                        print(f"Usando distâncias do Excel para grade {grade_num}")
                    else:
                        print(f"Distâncias do Excel não encontradas para grade {grade_num}, mantendo cálculo automático")
            
            return dados_grade
            
        except Exception as e:
            print(f"Erro ao processar distâncias do Excel: {e}")
            return dados_grade

    def gerar_scripts_scr(self, caminho_arquivo_excel):
        """Gera os scripts SCR usando o ROBO_GRADES."""
        print(f"\n>>> [DEBUG GRADE_EXCEL] gerar_scripts_scr INICIADO")
        print(f">>> [DEBUG GRADE_EXCEL]   - caminho_arquivo_excel: {caminho_arquivo_excel}")
        print(f">>> [DEBUG GRADE_EXCEL]   - dados_grades disponíveis: {len(self.dados_grades) if hasattr(self, 'dados_grades') and self.dados_grades else 0}")
        if hasattr(self, 'dados_grades') and self.dados_grades:
            print(f">>> [DEBUG GRADE_EXCEL]   - Itens em dados_grades:")
            for idx, (nome, dados) in enumerate(self.dados_grades.items(), 1):
                print(f">>> [DEBUG GRADE_EXCEL]     [{idx}] {nome}: {dados.get('pavimento', 'N/A')}")
        
        try:
            print("\nGerando scripts SCR...")
            
            # Verificar se pilar especial está ativo (linha 1000, coluna 1 do Excel)
            print(f">>> [DEBUG GRADE_EXCEL] Verificando pilar especial...")
            pilar_especial_ativo = False
            try:
                import openpyxl
                wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
                sheet = wb.active
                valor_checkbox = sheet.cell(row=1000, column=1).value
                print(f"[DEBUG GRADES] Valor lido do Excel (linha 1000, coluna 1): {valor_checkbox} (tipo: {type(valor_checkbox)})")
                
                # Converter valor para booleano considerando vários formatos possíveis
                if valor_checkbox is None:
                    pilar_especial_ativo = False
                elif isinstance(valor_checkbox, bool):
                    pilar_especial_ativo = valor_checkbox
                elif isinstance(valor_checkbox, (int, float)):
                    pilar_especial_ativo = bool(valor_checkbox) and valor_checkbox != 0
                elif isinstance(valor_checkbox, str):
                    valor_str = valor_checkbox.strip().upper()
                    pilar_especial_ativo = valor_str in ('1', 'TRUE', 'SIM', 'S', 'YES', 'Y')
                else:
                    pilar_especial_ativo = bool(valor_checkbox)
                
                wb.close()
                print(f"[DEBUG GRADES] ✅ Pilar especial ativo: {pilar_especial_ativo}")
                if pilar_especial_ativo:
                    print(f"[DEBUG GRADES] ✅ Vai gerar 6 scripts especiais (A, B, E, F, G, H)")
                else:
                    print(f"[DEBUG GRADES] ⚠️ Vai gerar apenas 2 scripts normais (A, B)")
            except Exception as e:
                print(f"[DEBUG GRADES] ❌ Erro ao ler checkbox pilar especial: {e}, assumindo desativado")
                import traceback
                traceback.print_exc()
                pilar_especial_ativo = False
            
            # Se pilar especial ativo, gerar 6 scripts especiais (A, B, E, F, G, H)
            if pilar_especial_ativo:
                return self.gerar_scripts_especiais(caminho_arquivo_excel)
            
            # Usar módulo diretamente (gerar_script_grade é uma função no módulo)
            
            # CORREÇÃO: NÃO limpar pasta aqui - a limpeza já é feita em _dcad_pavimento_grades
            # antes de processar os itens. Se limparmos aqui para cada item, vamos remover
            # os scripts dos itens anteriores (ex: P1 quando processamos P2)
            # 
            # Apenas garantir que a pasta existe
            # Determinar pavimento do primeiro item
            if self.dados_grades:
                primeiro_item = list(self.dados_grades.items())[0]
                pavimento_primeiro = primeiro_item[1]['pavimento']
                
                # Usar path resolver para obter o caminho correto
                import sys
                import os
                sys.path.append(os.path.dirname(os.path.dirname(__file__)))
                from utils.robust_path_resolver import robust_path_resolver
                pasta_pavimento = os.path.join(
                    robust_path_resolver.get_project_root(), 
                    "output/scripts",
                    f"{pavimento_primeiro}_GRADES"
                )
                
                # Apenas garantir que a pasta existe (NÃO limpar - já foi limpo antes)
                print(f">>> [DEBUG GRADE_EXCEL] Verificando se pasta existe: {pasta_pavimento}")
                print(f">>> [DEBUG GRADE_EXCEL]   - Pasta existe: {os.path.exists(pasta_pavimento)}")
                print(f">>> [DEBUG GRADE_EXCEL]   - ⚠️ NÃO limpando pasta aqui (já foi limpa em _dcad_pavimento_grades)")
                
                if not os.path.exists(pasta_pavimento):
                    print(f">>> [DEBUG GRADE_EXCEL]   - Pasta não existe, criando...")
                    os.makedirs(pasta_pavimento)
                    print(f">>> [DEBUG GRADE_EXCEL] ✅ Pasta criada: {pasta_pavimento}")
                else:
                    # Listar arquivos na pasta para debug (mas NÃO remover)
                    print(f">>> [DEBUG GRADE_EXCEL]   - Arquivos na pasta (NÃO removendo):")
                    try:
                        arquivos_na_pasta = os.listdir(pasta_pavimento)
                        for arq in arquivos_na_pasta:
                            caminho_arq = os.path.join(pasta_pavimento, arq)
                            if os.path.isfile(caminho_arq):
                                tamanho = os.path.getsize(caminho_arq)
                                print(f">>> [DEBUG GRADE_EXCEL]     - {arq} ({tamanho} bytes)")
                            else:
                                print(f">>> [DEBUG GRADE_EXCEL]     - {arq} (pasta)")
                    except Exception as e:
                        print(f">>> [DEBUG GRADE_EXCEL]     - Erro ao listar: {e}")
            
            # Processar cada arquivo individualmente
            print(f">>> [DEBUG GRADE_EXCEL] Iniciando loop de processamento de {len(self.dados_grades)} itens...")
            for i, (nome_arquivo, dados) in enumerate(self.dados_grades.items(), 1):
                print(f"\n>>> [DEBUG GRADE_EXCEL] ========================================")
                print(f">>> [DEBUG GRADE_EXCEL] [{i}/{len(self.dados_grades)}] Processando item: {nome_arquivo}")
                print(f">>> [DEBUG GRADE_EXCEL] ========================================")
                print(f"\nPreparando dados para o arquivo: {nome_arquivo}")
                print(f"Dados originais: {dados}")
                print(f">>> [DEBUG GRADE_EXCEL]   - Nome: {nome_arquivo}")
                print(f">>> [DEBUG GRADE_EXCEL]   - Pavimento: {dados.get('pavimento', 'N/A')}")
                
                # Processar distâncias específicas do Excel antes de gerar o script
                dados = self.processar_distancias_excel(dados)
                
                # Determinar pasta do pavimento (usar mesmo código para garantir pasta existe)
                pavimento = dados['pavimento']
                # Usar path resolver para obter o caminho correto
                import sys
                import os
                sys.path.append(os.path.dirname(os.path.dirname(__file__)))
                from utils.robust_path_resolver import robust_path_resolver
                pasta_pavimento = os.path.join(
                    robust_path_resolver.get_project_root(), 
                    "output/scripts",
                    f"{pavimento}_GRADES"
                )
                
                # Converter altura_base para float
                altura_base = float(dados['altura_base'])
                
                # Configurar as grades ativas
                grades_ativas = []
                for grade_key in ["grade1", "grade2", "grade3"]:
                    if grade_key in dados['grades']:
                        grade_data = dados['grades'][grade_key]
                        largura = float(grade_data['largura'])
                        
                        # Verificar se a grade tem larguras dos detalhes preenchidas
                        tem_larguras_detalhes = False
                        grade_index = list(dados['grades'].keys()).index(grade_key)  # Índice da grade atual (0, 1, 2)
                        if 'larguras_detalhes_grades' in dados:
                            for detalhe_num in range(1, 6):
                                campo_key = f'detalhe_grade{grade_index+1}_{detalhe_num}'
                                if campo_key in dados['larguras_detalhes_grades']:
                                    largura_detalhe = float(dados['larguras_detalhes_grades'][campo_key])
                                    if largura_detalhe > 0:
                                        tem_larguras_detalhes = True
                                        break
                        
                        if largura > 0 or tem_larguras_detalhes:  # Ativa se tem largura OU tem detalhes preenchidos
                            # Armazenar alturas originais antes de ajustar
                            # Como sarr_esquerda e sarr_direita não existem no Excel, usar 0 por padrão
                            altura_original_esquerda = float(grade_data.get('sarr_esquerda', 0))
                            altura_original_direita = float(grade_data.get('sarr_direita', 0))
                            alturas_originais = [
                                float(grade_data.get('sarr1_altura', 0)),
                                float(grade_data.get('sarr2_altura', 0)),
                                float(grade_data.get('sarr3_altura', 0))
                            ]
                            
                            # Ajustar alturas dos sarrafos
                            sarr_esquerda = self.ajustar_altura(altura_original_esquerda) if altura_original_esquerda > 0 else 0
                            sarr_direita = self.ajustar_altura(altura_original_direita) if altura_original_direita > 0 else 0
                            alturas = [self.ajustar_altura(altura) if altura > 0 else 0 for altura in alturas_originais]
                            
                            # Processar larguras dos detalhes se disponíveis
                            larguras_detalhes = []
                            grade_num = grade_index + 1  # Número da grade (1, 2, 3)
                            
                            # Verificar se o arquivo é .A ou .B para usar o conjunto correto de larguras
                            nome_original = dados.get('nome_arquivo', '')
                            is_arquivo_b = nome_original.endswith('.B')
                            
                            # Para arquivo .B, usar detalhes_grades_grupo2 se disponível
                            if is_arquivo_b and 'detalhes_grades_grupo2' in dados:
                                # Para arquivo .B, usar larguras dos detalhes do Grupo 2
                                print(f"[DEBUG LARGURAS] Arquivo .B detectado - usando detalhes_grades_grupo2 (Conjunto Grade 2)")
                                for detalhe_num in range(1, 6):
                                    campo_key = f'detalhe_grade{grade_num}_{detalhe_num}_grupo2'
                                    if campo_key in dados['detalhes_grades_grupo2']:
                                        largura_detalhe = float(dados['detalhes_grades_grupo2'][campo_key])
                                        larguras_detalhes.append(largura_detalhe if largura_detalhe > 0 else 0)
                                        print(f"Largura detalhe Grade B {grade_num}-{detalhe_num}: {largura_detalhe}")
                                    else:
                                        larguras_detalhes.append(0)
                                        print(f"Campo {campo_key} não encontrado em detalhes_grades_grupo2, usando 0")
                            elif 'larguras_detalhes_grades' in dados:
                                # Para arquivo .A ou fallback, usar larguras_detalhes_grades (Conjunto Grade 1)
                                if not is_arquivo_b:
                                    print(f"[DEBUG LARGURAS] Arquivo .A detectado - usando larguras_detalhes_grades (Conjunto Grade 1)")
                                for detalhe_num in range(1, 6):
                                    campo_key = f'detalhe_grade{grade_num}_{detalhe_num}'
                                    if campo_key in dados['larguras_detalhes_grades']:
                                        largura_detalhe = float(dados['larguras_detalhes_grades'][campo_key])
                                        larguras_detalhes.append(largura_detalhe if largura_detalhe > 0 else 0)
                                        print(f"Largura detalhe Grade {grade_num}-{detalhe_num}: {largura_detalhe}")
                                    else:
                                        larguras_detalhes.append(0)
                                        print(f"Campo {campo_key} não encontrado, usando 0")
                            else:
                                # Se não houver larguras dos detalhes, usar zeros
                                larguras_detalhes = [0, 0, 0, 0, 0]
                                print(f"Nenhuma largura de detalhe encontrada para grade {grade_num}, usando zeros")
                            
                            # Processar alturas dos detalhes se disponíveis
                            alturas_detalhes = []
                            
                            # Verificar se o arquivo é .A ou .B para usar o conjunto correto de alturas
                            nome_original = dados.get('nome_arquivo', '')
                            is_arquivo_b = nome_original.endswith('.B')
                            
                            if is_arquivo_b and 'alturas_detalhes_grade_b' in dados:
                                # Para arquivo .B, usar alturas_detalhes_grade_b (Conjunto Grade 2)
                                print(f"[DEBUG ALTURAS] Arquivo .B detectado - usando alturas_detalhes_grade_b (Conjunto Grade 2)")
                                print(f"[DEBUG ALTURAS] Chaves disponíveis em alturas_detalhes_grade_b: {list(dados['alturas_detalhes_grade_b'].keys())[:10]}")
                                
                                for detalhe_num in range(6):  # Alterado de range(1, 6) para range(6) para incluir campo 0
                                    campo_key = f'altura_detalhe_grade_b_{grade_num}_{detalhe_num}'
                                    print(f"[DEBUG ALTURAS] Procurando campo: {campo_key}")
                                    if campo_key in dados['alturas_detalhes_grade_b']:
                                        altura_detalhe = float(dados['alturas_detalhes_grade_b'][campo_key])
                                        alturas_detalhes.append(self.ajustar_altura(altura_detalhe) if altura_detalhe > 0 else 0)
                                        print(f"Altura detalhe Grade B {grade_num}-{detalhe_num}: {altura_detalhe} -> {self.ajustar_altura(altura_detalhe) if altura_detalhe > 0 else 0}")
                                    else:
                                        alturas_detalhes.append(0)
                                        print(f"Campo {campo_key} não encontrado, usando 0")
                            elif not is_arquivo_b and 'alturas_detalhes_grade_a' in dados:
                                # Para arquivo .A, usar alturas_detalhes_grade_a (Conjunto Grade 1)
                                print(f"[DEBUG ALTURAS] Arquivo .A detectado - usando alturas_detalhes_grade_a (Conjunto Grade 1)")
                                print(f"[DEBUG ALTURAS] Chaves disponíveis em alturas_detalhes_grade_a: {list(dados['alturas_detalhes_grade_a'].keys())[:10]}")
                                
                                for detalhe_num in range(6):  # Alterado de range(1, 6) para range(6) para incluir campo 0
                                    campo_key = f'altura_detalhe_grade_a_{grade_num}_{detalhe_num}'
                                    print(f"[DEBUG ALTURAS] Procurando campo: {campo_key}")
                                    if campo_key in dados['alturas_detalhes_grade_a']:
                                        altura_detalhe = float(dados['alturas_detalhes_grade_a'][campo_key])
                                        alturas_detalhes.append(self.ajustar_altura(altura_detalhe) if altura_detalhe > 0 else 0)
                                        print(f"Altura detalhe Grade A {grade_num}-{detalhe_num}: {altura_detalhe} -> {self.ajustar_altura(altura_detalhe) if altura_detalhe > 0 else 0}")
                                    else:
                                        alturas_detalhes.append(0)
                                        print(f"Campo {campo_key} não encontrado, usando 0")
                            else:
                                # Se não houver alturas dos detalhes, usar zeros
                                alturas_detalhes = [0, 0, 0, 0, 0, 0]  # Alterado de 5 para 6 zeros para incluir campo 0
                                print(f"Nenhuma altura de detalhe encontrada para grade {grade_num}, usando zeros")
                            
                            # Usar alturas dos detalhes para sarrafos de extremidade
                            # Sarrafo esquerdo: usar altura do campo 0 (altura específica da extremidade esquerda)
                            if alturas_detalhes and len(alturas_detalhes) > 0 and alturas_detalhes[0] > 0:
                                sarr_esquerda = alturas_detalhes[0]
                                print(f"Usando altura do campo 0 para sarrafo esquerdo: {sarr_esquerda}")
                            
                            # Sarrafo direito: usar altura do campo SEGUINTE ao último detalhe preenchido
                            if larguras_detalhes and len(larguras_detalhes) > 0 and alturas_detalhes and len(alturas_detalhes) > 0:
                                # Procurar o último campo de largura preenchido (não altura)
                                ultimo_largura_preenchido = -1
                                for i in range(len(larguras_detalhes) - 1, -1, -1):
                                    if larguras_detalhes[i] > 0:
                                        ultimo_largura_preenchido = i
                                        break
                                
                                # Usar a altura do campo SEGUINTE ao último detalhe preenchido
                                if ultimo_largura_preenchido >= 0 and ultimo_largura_preenchido + 1 < len(alturas_detalhes):
                                    sarr_direita = alturas_detalhes[ultimo_largura_preenchido + 1]
                                    print(f"Usando altura do campo {ultimo_largura_preenchido + 1} (seguinte ao último detalhe {ultimo_largura_preenchido}) para sarrafo direito: {sarr_direita}")
                                else:
                                    # Fallback: usar último campo de altura preenchido
                                    for i in range(len(alturas_detalhes) - 1, -1, -1):
                                        if alturas_detalhes[i] > 0:
                                            sarr_direita = alturas_detalhes[i]
                                            print(f"Fallback: usando altura do campo {i} para sarrafo direito: {sarr_direita}")
                                            break
                            elif alturas_detalhes and len(alturas_detalhes) > 0:
                                # Fallback: usar último campo de altura preenchido
                                for i in range(len(alturas_detalhes) - 1, -1, -1):
                                    if alturas_detalhes[i] > 0:
                                        sarr_direita = alturas_detalhes[i]
                                        print(f"Fallback: usando altura do campo {i} para sarrafo direito: {sarr_direita}")
                                        break
                            
                            grades_ativas.append({
                                'largura': largura,
                                'esquerda': sarr_esquerda,
                                'direita': sarr_direita,
                                'alturas': alturas,
                                'alturas_detalhes': alturas_detalhes,  # Adicionar alturas dos detalhes
                                'larguras_detalhes': larguras_detalhes,  # Adicionar larguras dos detalhes
                                'esquerdo_central': False,
                                'direito_central': False,
                                'distancias': [
                                    float(grade_data['sarr1_distancia']),
                                    float(grade_data['sarr2_distancia']),
                                    float(grade_data['sarr3_distancia'])
                                ]
                            })
                
                # Preparar as distâncias entre grades
                # SEPARAÇÃO: .A usa Grupo 1, .B usa Grupo 2
                nome_original = dados['nome_arquivo']
                
                # Debug: mostrar todas as chaves disponíveis em dados
                print(f"\n[DEBUG DADOS COMPLETOS] Chaves em dados: {list(dados.keys())}")
                print(f"[DEBUG DADOS] Distâncias disponíveis em dados:")
                for key in dados.keys():
                    if 'distancia' in key.lower():
                        print(f"  - {key}: {dados[key]}")
                
                # Preparar distâncias baseadas no tipo de arquivo (.A ou .B)
                if nome_original.endswith('.A'):
                    # Grade A = Grupo 1
                    # DISTÂNCIA 1 = espaço entre Grade 1 e Grade 2 (do início da Grade 1 até início da Grade 2)
                    # DISTÂNCIA 2 = espaço entre Grade 2 e Grade 3 (do início da Grade 2 até início da Grade 3)
                    if 'distancia_1' in dados and 'distancia_2' in dados:
                        print(f"\n[DEBUG GRUPO 1] Usando distâncias do Grupo 1 (linhas 182, 184) para arquivo .A")
                        print(f"  - distancia_1 (linha 182, Grade 1→Grade 2): {dados['distancia_1']} cm")
                        print(f"  - distancia_2 (linha 184, Grade 2→Grade 3): {dados['distancia_2']} cm")
                        print(f"  - VALOR RAW distancia_1: {dados['distancia_1']} (tipo: {type(dados['distancia_1'])})")
                        print(f"  - VALOR RAW distancia_2: {dados['distancia_2']} (tipo: {type(dados['distancia_2'])})")
                        distancias = {
                            'distancia_grade1': float(dados['distancia_1']),  # Grade 1→2
                            'distancia_grade2': float(dados['distancia_2'])   # Grade 2→3
                        }
                        print(f"  - VALOR CONVERTIDO distancia_grade1: {distancias['distancia_grade1']}cm")
                        print(f"  - VALOR CONVERTIDO distancia_grade2: {distancias['distancia_grade2']}cm")
                        print(f"  → Passando para robo: distancia_grade1={distancias['distancia_grade1']}cm, distancia_grade2={distancias['distancia_grade2']}cm")
                    else:
                        print(f"[DEBUG GRUPO 1] ❌ Chaves distancia_1 ou distancia_2 não encontradas!")
                        print(f"  - distancia_1 existe: {'distancia_1' in dados}")
                        print(f"  - distancia_2 existe: {'distancia_2' in dados}")
                        distancias = {
                            'distancia_grade1': float(dados.get('distancia_grade1_2', 50)),
                            'distancia_grade2': float(dados.get('distancia_grade2_3', 50))
                        }
                        print(f"  → Fallback: distancia_grade1={distancias['distancia_grade1']}cm, distancia_grade2={distancias['distancia_grade2']}cm")
                elif nome_original.endswith('.B'):
                    # Grade B = Grupo 2
                    # DISTÂNCIA 1 = espaço entre Grade 1 e Grade 2 (do início da Grade 1 até início da Grade 2)
                    # DISTÂNCIA 2 = espaço entre Grade 2 e Grade 3 (do início da Grade 2 até início da Grade 3)
                    if 'distancia_1_grupo2' in dados and 'distancia_2_grupo2' in dados:
                        print(f"\n[DEBUG GRUPO 2] Usando distâncias do Grupo 2 (linhas 220, 222) para arquivo .B")
                        print(f"  - distancia_1_grupo2 (linha 220, Grade 1→Grade 2): {dados['distancia_1_grupo2']} cm")
                        print(f"  - distancia_2_grupo2 (linha 222, Grade 2→Grade 3): {dados['distancia_2_grupo2']} cm")
                        distancias = {
                            'distancia_grade1': float(dados['distancia_1_grupo2']),  # Grade 1→2
                            'distancia_grade2': float(dados['distancia_2_grupo2'])   # Grade 2→3
                        }
                        print(f"  → Passando para robo: distancia_grade1={distancias['distancia_grade1']}cm, distancia_grade2={distancias['distancia_grade2']}cm")
                    else:
                        print(f"[DEBUG GRUPO 2] ❌ Chaves distancia_1_grupo2 ou distancia_2_grupo2 não encontradas!")
                        print(f"  - distancia_1_grupo2 existe: {'distancia_1_grupo2' in dados}")
                        print(f"  - distancia_2_grupo2 existe: {'distancia_2_grupo2' in dados}")
                        distancias = {
                            'distancia_grade1': float(dados.get('distancia_1', dados.get('distancia_grade1_2', 50))),
                            'distancia_grade2': float(dados.get('distancia_2', dados.get('distancia_grade2_3', 50)))
                        }
                        print(f"  → Fallback: distancia_grade1={distancias['distancia_grade1']}cm, distancia_grade2={distancias['distancia_grade2']}cm")
                else:
                    print(f"[DEBUG] ❌ Arquivo sem sufixo .A ou .B: {nome_original}")
                    distancias = {
                        'distancia_grade1': float(dados.get('distancia_1', dados.get('distancia_grade1_2', 50))),
                        'distancia_grade2': float(dados.get('distancia_2', dados.get('distancia_grade2_3', 50)))
                    }
                
                print(f"Grades ativas: {grades_ativas}")
                print(f"Distâncias: {distancias}")
                
                print(f"\nGerando script para o arquivo: {nome_arquivo}")
                
                # Remover sufixos .A ou .B do nome se existirem para obter nome_base
                nome_original = dados['nome_arquivo']
                if nome_original.endswith('.A') or nome_original.endswith('.B'):
                    nome_base = nome_original[:-2]  # Remove .A ou .B
                    sufixo = nome_original[-2:]  # Pega .A ou .B
                else:
                    nome_base = nome_original
                    sufixo = ""
                
                # Calcular deslocamento se for arquivo .B (para separar de .A)
                x_inicial_custom = None
                if sufixo == '.B':
                    # Ler dados do arquivo .A correspondente
                    nome_arquivo_a = f"{nome_base}.A"
                    if nome_arquivo_a in self.dados_grades:
                        dados_a = self.dados_grades[nome_arquivo_a]
                        # Calcular deslocamento: larguras + distâncias do Grupo 1 + espaço entre A e B
                        deslocamento = 0.0
                        if 'grades' in dados_a:
                            grades_a = dados_a['grades']
                            # Grade 1
                            if 'grade1' in grades_a:
                                deslocamento += float(grades_a['grade1'].get('largura', 0))
                            # Distância 1 (Grupo 1)
                            if 'distancia_1' in dados_a:
                                deslocamento += float(dados_a['distancia_1'])
                            elif 'distancia_grade1_2' in dados_a:
                                deslocamento += float(dados_a['distancia_grade1_2'])
                            # Grade 2
                            if 'grade2' in grades_a:
                                deslocamento += float(grades_a['grade2'].get('largura', 0))
                            # Distância 2 (Grupo 1)
                            if 'distancia_2' in dados_a:
                                deslocamento += float(dados_a['distancia_2'])
                            elif 'distancia_grade2_3' in dados_a:
                                deslocamento += float(dados_a['distancia_grade2_3'])
                            # Grade 3
                            if 'grade3' in grades_a:
                                deslocamento += float(grades_a['grade3'].get('largura', 0))
                        # Adicionar espaço entre A e B
                        deslocamento += 80.0
                        # Adicionar 100cm extra para melhor visualização
                        deslocamento += 100.0
                        x_inicial_custom = dados['x_inicial'] + deslocamento
                        print(f"[DESLOCAMENTO] Arquivo .B deslocado em {deslocamento}cm para separar de .A (incluindo 100cm extra)")
                
                # Usar a função gerar_script_grade do ROBO_GRADES com o sufixo correto
                script = robo_module.gerar_script_grade(
                    dados['pavimento'],      # pavimento
                    nome_original,           # nome - Usa o nome original com sufixo
                    grades_ativas,          # grades_ativas
                    altura_base,            # altura_base
                    x_inicial_custom=x_inicial_custom,  # x_inicial_custom (None para .A, deslocado para .B)
                    y_inicial_custom=None,  # y_inicial_custom
                    distancias=distancias   # distancias entre grades
                )
                
                if script:
                    pavimento = dados['pavimento']
                    
                    # Gerar apenas UM arquivo com o sufixo correto
                    caminho_script = gerar_nome_arquivo(
                        nome=f"{nome_base}",
                        pavimento=pavimento,
                        sufixo=sufixo  # Usa .A ou .B do nome original
                    )
                    if self.salvar_arquivo_utf16(caminho_script, script):
                        print(f"Script gerado com sucesso para: {nome_arquivo}")
                        print(f"Salvo em: {caminho_script}")
                    else:
                        print(f"Erro ao salvar script para: {nome_arquivo}")
                else:
                    print(f"Erro ao gerar script para: {nome_arquivo}")
            
            print(f"\nTodos os scripts SCR foram gerados com sucesso")
            return True
            
        except Exception as e:
            print(f"Erro ao gerar scripts SCR: {e}")
            print("Stack trace completo:")
            traceback.print_exc()
            return False

    def imprimir_informacoes_grades(self):
        """Imprime as informações das grades de forma organizada."""
        print("\n" + "="*80)
        print("INFORMAÇÕES DAS GRADES PROCESSADAS")
        print("="*80)
        
        for nome_arquivo, dados in self.dados_grades.items():
            print(f"\n[ARQUIVO] {nome_arquivo}")
            print(f"[PAVIMENTO] {dados['pavimento']}")
            print(f"[ALTURA BASE] {dados['altura_base']} cm")
            print(f"[DISTÂNCIAS ENTRE GRADES]")
            print(f"   - Grade 1 -> Grade 2: {dados['distancia_grade1_2']} cm")
            print(f"   - Grade 2 -> Grade 3: {dados['distancia_grade2_3']} cm")
            
            for grade_num in range(1, 4):
                grade_key = f'grade{grade_num}'
                grade_dados = dados['grades'][grade_key]
                print(f"\n[GRADE {grade_num}]")
                print(f"   Altura: {grade_dados['altura']} cm")
                print(f"   Largura: {grade_dados['largura']} cm")
            
            print("-"*80)

    def imprimir_resumo_distancias(self):
        """Imprime um resumo das distâncias processadas do Excel."""
        print("\n" + "="*80)
        print("RESUMO DAS DISTÂNCIAS PROCESSADAS DO EXCEL (CUMULATIVAS)")
        print("="*80)
        
        for nome_arquivo, dados in self.dados_grades.items():
            print(f"\n[ARQUIVO] {nome_arquivo}")
            print(f"[PAVIMENTO] {dados['pavimento']}")
            
            for grade_num in range(1, 4):
                grade_key = f'grade{grade_num}'
                if grade_key in dados['grades']:
                    grade_dados = dados['grades'][grade_key]
                    largura = grade_dados.get('largura', '0.00')
                    
                    print(f"\n[GRADE {grade_num}] - Largura: {largura} cm")
                    print(f"   Distâncias dos Sarrafos (CUMULATIVAS):")
                    print(f"     Sarrafo 1: {grade_dados.get('sarr1_distancia', '0.00')} cm")
                    print(f"     Sarrafo 2: {grade_dados.get('sarr2_distancia', '0.00')} cm")
                    print(f"     Sarrafo 3: {grade_dados.get('sarr3_distancia', '0.00')} cm")
                    
                    # Verificar se as distâncias foram lidas do Excel ou calculadas
                    distancias = [
                        float(grade_dados.get('sarr1_distancia', 0)),
                        float(grade_dados.get('sarr2_distancia', 0)),
                        float(grade_dados.get('sarr3_distancia', 0))
                    ]
                    
                    if any(d > 0 for d in distancias):
                        print(f"   Status: Distâncias lidas do Excel e aplicadas cumulativamente ✓")
                    else:
                        print(f"   Status: Distâncias calculadas automaticamente (Excel vazio)")
            
            print("-"*80)

    def imprimir_campos_interface(self):
        """Imprime todos os campos da interface, mostrando quais estão preenchidos e quais estão em branco."""
        print("\n" + "="*100)
        print("CAMPOS DA INTERFACE - VERIFICAÇÃO DE PREENCHIMENTO")
        print("="*100)
        
        for nome_arquivo, dados in self.dados_grades.items():
            print(f"\n[ARQUIVO] {nome_arquivo}")
            print(f"[PAVIMENTO] {dados['pavimento']}")
            print(f"[ALTURA BASE] {dados['altura_base']} cm")
            print(f"[DISTÂNCIAS ENTRE GRADES]")
            print(f"   - Grade 1 -> Grade 2: {dados['distancia_grade1_2']} cm")
            print(f"   - Grade 2 -> Grade 3: {dados['distancia_grade2_3']} cm")
            
            for grade_num in range(1, 4):
                grade_key = f'grade{grade_num}'
                grade_dados = dados['grades'][grade_key]
                print(f"\n[GRADE {grade_num}]")
                print(f"   Largura: {grade_dados.get('largura', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo Esquerdo: {grade_dados.get('sarr_esquerda', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo Direito: {grade_dados.get('sarr_direita', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 1 - Altura: {grade_dados.get('sarr1_altura', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 1 - Distância (Excel Cumulativa): {grade_dados.get('sarr1_distancia', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 2 - Altura: {grade_dados.get('sarr2_altura', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 2 - Distância (Excel Cumulativa): {grade_dados.get('sarr2_distancia', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 3 - Altura: {grade_dados.get('sarr3_altura', 'NÃO PREENCHIDO')} cm")
                print(f"   Sarrafo 3 - Distância (Excel Cumulativa): {grade_dados.get('sarr3_distancia', 'NÃO PREENCHIDO')} cm")
            
            print("-"*100)

    def preparar_dados_grade_especial(self, caminho_arquivo_excel, letra, nome_arquivo_base, pavimento, coluna='E'):
        """
        Prepara dados de uma grade especial específica (A, B, E, F, G, H).
        
        Args:
            caminho_arquivo_excel: Caminho do arquivo Excel
            letra: Letra da grade ('a', 'b', 'e', 'f', 'g', 'h')
            nome_arquivo_base: Nome base do arquivo (sem sufixo)
            pavimento: Nome do pavimento
            coluna: Coluna do Excel (padrão 'E')
        
        Returns:
            dict: Dados estruturados da grade especial, similar a dados_grade normal
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
            sheet = wb.active
            
            # Ler altura base (linha 12)
            try:
                formula_altura = sheet[f'{coluna}12'].value
                altura_base = float(self.calcular_valor_formula(sheet, formula_altura, coluna) or 0)
            except:
                altura_base = 0.0
            
            # Ler grades principais (grade_{letra}_1, grade_{letra}_2, grade_{letra}_3)
            grades = {}
            # Importar EXCEL_MAPPING de forma robusta
            try:
                from ..utils.excel_mapping import EXCEL_MAPPING
            except ImportError:
                # Fallback para import direto quando executado como subprocess
                utils_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'utils')
                if utils_path not in sys.path:
                    sys.path.insert(0, utils_path)
                from excel_mapping import EXCEL_MAPPING
            
            for grade_num in range(1, 4):
                campo_grade = f"grade_{letra}_{grade_num}"
                linha_grade = EXCEL_MAPPING.get(campo_grade)
                if linha_grade:
                    try:
                        valor = sheet.cell(row=linha_grade, column=5).value  # Coluna E
                        largura = float(valor or 0)
                    except:
                        largura = 0.0
                    
                    grades[f'grade{grade_num}'] = {
                        'largura': largura,
                        'sarr_esquerda': 0,  # Não usado em grades especiais
                        'sarr1_altura': altura_base,
                        'sarr2_altura': altura_base,
                        'sarr3_altura': altura_base,
                        'sarr_direita': 0
                    }
            
            # Ler distâncias (dist_{letra}_1, dist_{letra}_2)
            distancia_grade1_2 = 0.0
            distancia_grade2_3 = 0.0
            
            campo_dist1 = f"dist_{letra}_1"
            linha_dist1 = EXCEL_MAPPING.get(campo_dist1)
            if linha_dist1:
                try:
                    distancia_grade1_2 = float(sheet.cell(row=linha_dist1, column=5).value or 0)
                except:
                    distancia_grade1_2 = 0.0
            
            campo_dist2 = f"dist_{letra}_2"
            linha_dist2 = EXCEL_MAPPING.get(campo_dist2)
            if linha_dist2:
                try:
                    distancia_grade2_3 = float(sheet.cell(row=linha_dist2, column=5).value or 0)
                except:
                    distancia_grade2_3 = 0.0
            
            # Ler detalhes (detalhe_{letra}_{grade_num}_{detalhe_num})
            print(f"[DEBUG LEITURA {letra.upper()}] Lendo detalhes da Grade {letra.upper()}...")
            larguras_detalhes_grades = {}
            for grade_num in range(1, 4):
                for detalhe_num in range(1, 6):
                    campo_detalhe = f"detalhe_{letra}_{grade_num}_{detalhe_num}"
                    linha_detalhe = EXCEL_MAPPING.get(campo_detalhe)
                    if linha_detalhe:
                        try:
                            valor = sheet.cell(row=linha_detalhe, column=5).value
                            valor_float = float(valor or 0)
                            larguras_detalhes_grades[campo_detalhe] = valor_float
                            # Mostrar TODOS os detalhes (mesmo zeros) para debug completo
                            status = "✓" if valor_float > 0 else "○"
                            print(f"[DEBUG LEITURA {letra.upper()}] {status} {campo_detalhe} = {valor_float} (linha {linha_detalhe}, Excel valor: {valor})")
                        except Exception as e:
                            larguras_detalhes_grades[campo_detalhe] = 0.0
                            print(f"[DEBUG LEITURA {letra.upper()}] ✗ {campo_detalhe} erro: {e}")
                    else:
                        print(f"[DEBUG LEITURA {letra.upper()}] ✗ {campo_detalhe} → SEM MAPEAMENTO no EXCEL_MAPPING")
            
            # Ler alturas dos detalhes (altura_detalhe_{letra}_{grade_num}_{altura_num})
            print(f"[DEBUG LEITURA {letra.upper()}] Lendo alturas dos detalhes da Grade {letra.upper()}...")
            alturas_detalhes_grade = {}
            for grade_num in range(1, 4):
                for altura_num in range(0, 6):  # 0 a 5
                    campo_altura = f"altura_detalhe_{letra}_{grade_num}_{altura_num}"
                    linha_altura = EXCEL_MAPPING.get(campo_altura)
                    if linha_altura:
                        try:
                            valor = sheet.cell(row=linha_altura, column=5).value
                            # Converter string com vírgula para float (ex: '285,00' -> 285.00)
                            if isinstance(valor, str):
                                # Substituir vírgula por ponto e remover espaços
                                valor = valor.replace(',', '.').strip()
                            # Converter para float
                            valor_float = float(valor) if valor is not None and valor != '' else 0.0
                            valor_ajustado = self.ajustar_altura(valor_float)
                            alturas_detalhes_grade[campo_altura] = valor_ajustado
                            # Mostrar TODAS as alturas (mesmo zeros) para debug completo
                            status = "✓" if valor_ajustado > 0 else "○"
                            print(f"[DEBUG LEITURA {letra.upper()}] {status} {campo_altura} = {valor_ajustado} (linha {linha_altura}, Excel valor: {valor})")
                        except Exception as e:
                            alturas_detalhes_grade[campo_altura] = 0.0
                            print(f"[DEBUG LEITURA {letra.upper()}] ✗ {campo_altura} erro: {e}")
                    else:
                        print(f"[DEBUG LEITURA {letra.upper()}] ✗ {campo_altura} → SEM MAPEAMENTO no EXCEL_MAPPING")
            
            wb.close()
            
            # Estruturar dados similar a dados_grade normal
            dados_grade_especial = {
                'nome_arquivo': f"{nome_arquivo_base}.{letra.upper()}",
                'pavimento': pavimento,
                'altura_base': altura_base,
                'grades': grades,
                'distancia_grade1_2': distancia_grade1_2,
                'distancia_grade2_3': distancia_grade2_3,
                'larguras_detalhes_grades': larguras_detalhes_grades,
                'alturas_detalhes_grade': alturas_detalhes_grade,
                'distancias': {
                    'distancia_grade1': distancia_grade1_2,
                    'distancia_grade2': distancia_grade2_3
                }
            }
            
            # Resumo dos dados lidos
            print(f"\n[DEBUG GRADE ESPECIAL {letra.upper()}] Dados preparados:")
            print(f"  - Altura base: {altura_base}cm")
            print(f"  - Grade 1: {grades.get('grade1', {}).get('largura', 0)}cm")
            print(f"  - Grade 2: {grades.get('grade2', {}).get('largura', 0)}cm")
            print(f"  - Grade 3: {grades.get('grade3', {}).get('largura', 0)}cm")
            print(f"  - Distância 1→2: {distancia_grade1_2}cm")
            print(f"  - Distância 2→3: {distancia_grade2_3}cm")
            
            # Resumo dos detalhes lidos
            detalhes_nao_zero = {k: v for k, v in larguras_detalhes_grades.items() if v > 0}
            print(f"\n[DEBUG GRADE ESPECIAL {letra.upper()}] Detalhes lidos (não-zero): {len(detalhes_nao_zero)} de {len(larguras_detalhes_grades)}")
            if detalhes_nao_zero:
                for campo, valor in sorted(detalhes_nao_zero.items()):
                    print(f"  - {campo} = {valor}cm")
            else:
                print(f"  ⚠️ NENHUM detalhe não-zero encontrado para {letra.upper()}!")
                print(f"  Valores lidos: {list(larguras_detalhes_grades.values())[:10]}...")
            
            # Resumo das alturas lidas
            alturas_nao_zero = {k: v for k, v in alturas_detalhes_grade.items() if v > 0}
            print(f"\n[DEBUG GRADE ESPECIAL {letra.upper()}] Alturas lidas (não-zero): {len(alturas_nao_zero)} de {len(alturas_detalhes_grade)}")
            if alturas_nao_zero:
                for campo, valor in sorted(list(alturas_nao_zero.items())[:10]):  # Primeiros 10
                    print(f"  - {campo} = {valor}cm")
            else:
                print(f"  ⚠️ NENHUMA altura não-zero encontrada para {letra.upper()}!")
                print(f"  Valores lidos: {list(alturas_detalhes_grade.values())[:10]}...")
            print()
            
            return dados_grade_especial
            
        except Exception as e:
            print(f"[ERRO] Falha ao preparar dados da grade especial {letra.upper()}: {e}")
            import traceback
            traceback.print_exc()
            return None

    def gerar_scripts_especiais(self, caminho_arquivo_excel):
        """
        Gera scripts SCR para as 6 grades especiais (A, B, E, F, G, H).
        
        Args:
            caminho_arquivo_excel: Caminho do arquivo Excel
            
        Returns:
            bool: True se sucesso, False caso contrário
        """
        try:
            print("\n[GRADES ESPECIAIS] Gerando scripts para A, B, E, F, G, H...")
            
            # Obter nome do arquivo e pavimento do primeiro item processado (ou ler do Excel)
            import openpyxl
            wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
            sheet = wb.active
            
            # Ler nome do arquivo (linha 4, coluna E) e pavimento (linha 3, coluna E)
            nome_arquivo_base = sheet.cell(row=4, column=5).value or "grade"
            pavimento = sheet.cell(row=3, column=5).value or "200"
            wb.close()
            
            # Usar módulo diretamente (gerar_script_grade é uma função, não método de classe)
            
            # Calcular deslocamento horizontal para cada grade (similar ao .B atual com 460.5cm)
            deslocamentos = {
                'a': 0.0,           # Grade A sem deslocamento
                'b': 460.5,         # Grade B deslocada 460.5cm (como atual)
                'e': 921.0,         # Grade E deslocada 921cm (2x 460.5)
                'f': 1381.5,        # Grade F deslocada 1381.5cm (3x 460.5)
                'g': 1842.0,        # Grade G deslocada 1842cm (4x 460.5)
                'h': 2302.5         # Grade H deslocada 2302.5cm (5x 460.5)
            }
            
            # Gerar script para cada grade especial
            for letra in ['a', 'b', 'e', 'f', 'g', 'h']:
                try:
                    print(f"\n[GRADE ESPECIAL {letra.upper()}] Preparando dados...")
                    
                    # Preparar dados da grade especial
                    dados_grade = self.preparar_dados_grade_especial(
                        caminho_arquivo_excel, letra, nome_arquivo_base, pavimento
                    )
                    
                    if not dados_grade:
                        print(f"[AVISO] Falha ao preparar dados da grade {letra.upper()}, pulando...")
                        continue
                    
                    # Obter deslocamento x_inicial
                    x_inicial_deslocado = deslocamentos[letra]
                    
                    # Preparar grades_ativas (lista de grades similar ao formato atual)
                    grades_ativas = []
                    for grade_key in ['grade1', 'grade2', 'grade3']:
                        if grade_key in dados_grade['grades']:
                            grade_data = dados_grade['grades'][grade_key]
                            if grade_data['largura'] > 0:
                                # Preparar detalhes da grade
                                grade_index = int(grade_key[-1]) - 1  # 0, 1, 2
                                
                                # Ler larguras dos detalhes
                                larguras_detalhes = []
                                for detalhe_num in range(1, 6):
                                    campo_detalhe = f"detalhe_{letra}_{grade_index + 1}_{detalhe_num}"
                                    largura_detalhe = dados_grade['larguras_detalhes_grades'].get(campo_detalhe, 0)
                                    larguras_detalhes.append(float(largura_detalhe) if largura_detalhe and largura_detalhe > 0 else 0)
                                
                                # Debug para Grade A
                                if letra == 'a' and grade_key == 'grade1':
                                    print(f"[DEBUG GRADE A] Larguras detalhes grade 1 lidas: {larguras_detalhes}")
                                    print(f"[DEBUG GRADE A] Total de chaves em larguras_detalhes_grades: {len(dados_grade['larguras_detalhes_grades'])}")
                                    # Mostrar algumas chaves
                                    chaves_relevantes = [k for k in dados_grade['larguras_detalhes_grades'].keys() if f'detalhe_{letra}_1_' in k]
                                    print(f"[DEBUG GRADE A] Chaves relevantes (detalhe_{letra}_1_*): {chaves_relevantes}")
                                    for chave in chaves_relevantes[:5]:
                                        print(f"[DEBUG GRADE A]   {chave} = {dados_grade['larguras_detalhes_grades'].get(chave)}")
                                
                                # Ler alturas dos detalhes
                                alturas_detalhes = []
                                for altura_num in range(6):  # 0 a 5
                                    campo_altura = f"altura_detalhe_{letra}_{grade_index + 1}_{altura_num}"
                                    altura_detalhe = dados_grade['alturas_detalhes_grade'].get(campo_altura, 0)
                                    alturas_detalhes.append(float(altura_detalhe) if altura_detalhe and altura_detalhe > 0 else 0)
                                
                                # Extrair esquerda, direita e alturas centrais a partir das alturas_detalhes
                                # Campo 0 = esquerda, campos 1-3 = centrais, último central válido = direita
                                esquerda = alturas_detalhes[0] if len(alturas_detalhes) > 0 else 0
                                # Direita será calculada pelo robô baseado no último central válido
                                direita = alturas_detalhes[-1] if len(alturas_detalhes) > 0 else 0
                                # Alturas centrais (campos 1-3) - usar os 3 primeiros se disponíveis
                                alturas = [
                                    alturas_detalhes[1] if len(alturas_detalhes) > 1 else 0,
                                    alturas_detalhes[2] if len(alturas_detalhes) > 2 else 0,
                                    alturas_detalhes[3] if len(alturas_detalhes) > 3 else 0
                                ]
                                
                                # Calcular distâncias dos sarrafos baseado nas larguras dos detalhes
                                # Similar ao cálculo das grades normais, mas baseado nas larguras_detalhes
                                distancias = []
                                if larguras_detalhes:
                                    acumulado = 0.0
                                    for i, largura in enumerate(larguras_detalhes[:3]):  # Apenas os 3 primeiros
                                        if largura > 0:
                                            acumulado += largura
                                            distancias.append(float(acumulado))
                                        else:
                                            distancias.append(0.0)
                                else:
                                    distancias = [0.0, 0.0, 0.0]
                                
                                grade_ativa = {
                                    'largura': float(grade_data['largura']),
                                    'esquerda': float(esquerda),
                                    'direita': float(direita),
                                    'alturas': [float(a) for a in alturas],
                                    'larguras_detalhes': larguras_detalhes,
                                    'alturas_detalhes': alturas_detalhes,
                                    'esquerdo_central': False,
                                    'direito_central': False,
                                    'distancias': distancias
                                }
                                grades_ativas.append(grade_ativa)
                    
                    # Verificar se há grades ativas antes de gerar script
                    if not grades_ativas:
                        print(f"[AVISO] Grade {letra.upper()} não tem grades ativas (todas larguras = 0), pulando...")
                        continue
                
                    # Obter coordenadas da configuração (não usar valores hardcoded)
                    x_inicial_base = robo_module.config_manager.get_config("coordinates", "x_inicial")
                    y_inicial_base = robo_module.config_manager.get_config("coordinates", "y_inicial")
                    
                    # Calcular coordenadas finais: base + deslocamento
                    x_inicial_final = x_inicial_base + x_inicial_deslocado
                    y_inicial_final = y_inicial_base
                    
                    print(f"[GRADE ESPECIAL {letra.upper()}] Coordenadas:")
                    print(f"  - Base (config): X={x_inicial_base}, Y={y_inicial_base}")
                    print(f"  - Deslocamento: {x_inicial_deslocado}")
                    print(f"  - Final: X={x_inicial_final}, Y={y_inicial_final}")
                    
                    # Gerar script usando o mesmo método que grades normais
                    script_gerado = robo_module.gerar_script_grade(
                        pavimento=pavimento,
                        nome=dados_grade['nome_arquivo'],
                        grades_ativas=grades_ativas,
                        altura_base=dados_grade['altura_base'],
                        x_inicial_custom=x_inicial_final,  # Usar coordenada base + deslocamento
                        y_inicial_custom=y_inicial_final,  # Usar coordenada da configuração
                        distancias=dados_grade.get('distancias', {})
                    )
                    
                    # Salvar script
                    import os
                    # Importar robust_path_resolver de forma robusta
                    try:
                        from utils.robust_path_resolver import robust_path_resolver
                    except ImportError:
                        # Fallback para import direto quando executado como subprocess
                        utils_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'utils')
                        if utils_path not in sys.path:
                            sys.path.insert(0, utils_path)
                        from robust_path_resolver import robust_path_resolver
                    pasta_pavimento = os.path.join(
                        robust_path_resolver.get_project_root(),
                        "output/scripts",
                        f"{pavimento}_GRADES"
                    )
                    os.makedirs(pasta_pavimento, exist_ok=True)
                    
                    caminho_script = os.path.join(
                        pasta_pavimento,
                        f"{nome_arquivo_base}.{letra.upper()}.scr"
                    )
                    
                    # Salvar com UTF-16 LE (com BOM) para compatibilidade com AutoCAD
                    with open(caminho_script, 'wb') as f:
                        # Adicionar BOM UTF-16 LE
                        f.write(b'\xFF\xFE')
                        # Converter conteúdo para UTF-16 LE
                        f.write(script_gerado.encode('utf-16-le'))
                    
                    print(f"[SUCESSO] Script gerado: {caminho_script}")
                    
                except Exception as e:
                    print(f"[ERRO] Falha ao processar grade {letra.upper()}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue  # Continuar para próxima grade mesmo se esta falhar
            
            print("\n[GRADES ESPECIAIS] Geração concluída!")
            return True
            
        except Exception as e:
            print(f"[ERRO] Falha ao gerar scripts especiais: {e}")
            import traceback
            traceback.print_exc()
            return False

def gerar_nome_arquivo(nome, pavimento, numero_lista=None, sufixo=None):
    """
    Gera o nome do arquivo SCR baseado no nome, pavimento e número da lista.
    
    Args:
        nome (str): Nome base do arquivo
        pavimento (str): Nome do pavimento
        numero_lista (int, optional): Número da lista para adicionar ao nome
        sufixo (str, optional): Sufixo para adicionar ao nome (ex: ".A", ".B")
    
    Returns:
        str: Caminho completo do arquivo SCR
    """
    # Substituir espaços por underscores e remover .0 de números inteiros
    pavimento = str(pavimento).replace(" ", "_")
    if pavimento.endswith('.0'):
        pavimento = pavimento[:-2]
    
    # Criar o diretório do pavimento se não existir
    # Usar path resolver para obter o caminho correto (consistente com o resto do código)
    try:
        from utils.robust_path_resolver import robust_path_resolver
    except ImportError:
        try:
            from src.utils.robust_path_resolver import robust_path_resolver
        except ImportError:
            # Fallback final - calcular manualmente
            script_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(script_dir))
            diretorio_base = os.path.join(project_root, "output", "scripts")
            diretorio_pavimento = os.path.join(diretorio_base, f"{pavimento}_GRADES")
            if not os.path.exists(diretorio_pavimento):
                os.makedirs(diretorio_pavimento)
                print(f"Pasta criada: {diretorio_pavimento}")
            
            # Gerar nome do arquivo
            if numero_lista is not None:
                nome_arquivo = f"{nome}_{numero_lista}.scr"
            elif sufixo is not None:
                nome_arquivo = f"{nome}{sufixo}.scr"
            else:
                nome_arquivo = f"{nome}.scr"
            
            caminho_completo = os.path.join(diretorio_pavimento, nome_arquivo)
            print(f"Gerando arquivo em: {caminho_completo}")
            return caminho_completo
    
    # Usar robust_path_resolver para obter project_root consistente
    project_root = robust_path_resolver.get_project_root()
    diretorio_base = os.path.join(project_root, "output", "scripts")
    diretorio_pavimento = os.path.join(diretorio_base, f"{pavimento}_GRADES")
    
    if not os.path.exists(diretorio_pavimento):
        os.makedirs(diretorio_pavimento)
        print(f"Pasta criada: {diretorio_pavimento}")
    
    # Gerar nome do arquivo
    if numero_lista is not None:
        nome_arquivo = f"{nome}_{numero_lista}.scr"
    elif sufixo is not None:
        nome_arquivo = f"{nome}{sufixo}.scr"
    else:
        nome_arquivo = f"{nome}.scr"
    
    caminho_completo = os.path.join(diretorio_pavimento, nome_arquivo)
    print(f"Gerando arquivo em: {caminho_completo}")
    
    return caminho_completo

def main():
    # Solicitar ao usuário para selecionar o arquivo Excel
    root = tk.Tk()
    root.withdraw()  # Esconder a janela principal
    
    caminho_arquivo_excel = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx;*.xls")]
    )
    
    if not caminho_arquivo_excel:
        print("Nenhum arquivo selecionado. Encerrando o programa.")
        sys.exit(1)
    
    print(f"Arquivo Excel selecionado: {caminho_arquivo_excel}")
    
    # Verificar se o arquivo existe
    if not os.path.exists(caminho_arquivo_excel):
        print(f"Arquivo não encontrado: {caminho_arquivo_excel}")
        sys.exit(1)
    
    # Criar instância da classe de automação e processar a planilha
    automation = GradeAutomation()
    if automation.processar_planilha(caminho_arquivo_excel):
        # Imprimir todos os campos da interface
        automation.imprimir_campos_interface()
        
        # Imprimir resumo das distâncias processadas
        automation.imprimir_resumo_distancias()
        
        # Gerar nome do arquivo SQL baseado no arquivo Excel
        nome_base = os.path.splitext(os.path.basename(caminho_arquivo_excel))[0]
        caminho_script = os.path.join(os.path.dirname(caminho_arquivo_excel), f"{nome_base}_script.sql")
        
        # Salvar o script SQL
        if automation.salvar_script_sql(caminho_script):
            print(f"\nScript SQL gerado com sucesso em: {caminho_script}")
            
            # Gerar scripts SCR
            if automation.gerar_scripts_scr(caminho_arquivo_excel):
                print("\nProcessamento concluído com sucesso!")
            else:
                print("\nErro ao gerar scripts SCR")
        else:
            print("\nErro ao gerar script SQL")
    else:
        print("\nErro ao processar a planilha")

def preencher_campos_e_gerar_scripts(caminho_arquivo_excel, coluna_especifica=None, interface_principal=None):
    """
    Função wrapper para processar planilha Excel e gerar scripts SCR.
    Compatível com a interface esperada pelo Conector_Interface_PainelControle.
    
    Args:
        caminho_arquivo_excel: Caminho para o arquivo Excel.
        coluna_especifica: Coluna específica a ser processada (ex: 'E', 'F', 'G').
        interface_principal: Interface principal (não usado neste módulo, mas mantido para compatibilidade).
    """
    print(f"\n>>> [DEBUG GRADE_EXCEL] preencher_campos_e_gerar_scripts INICIADO")
    print(f">>> [DEBUG GRADE_EXCEL]   - caminho_arquivo_excel: {caminho_arquivo_excel}")
    print(f">>> [DEBUG GRADE_EXCEL]   - coluna_especifica: {coluna_especifica}")
    print(f">>> [DEBUG GRADE_EXCEL]   - interface_principal: {interface_principal}")
    
    try:
        print(f"[GRADE_EXCEL] Iniciando processamento do arquivo: {caminho_arquivo_excel}")
        
        # Criar instância da classe de automação
        print(f">>> [DEBUG GRADE_EXCEL] Criando instância GradeAutomation...")
        automation = GradeAutomation()
        print(f">>> [DEBUG GRADE_EXCEL] ✅ Instância GradeAutomation criada")
        
        # Processar planilha
        print(f">>> [DEBUG GRADE_EXCEL] Chamando processar_planilha...")
        print(f">>> [DEBUG GRADE_EXCEL]   - caminho_arquivo_excel: {caminho_arquivo_excel}")
        print(f">>> [DEBUG GRADE_EXCEL]   - coluna_especifica: {coluna_especifica}")
        resultado_processar = automation.processar_planilha(caminho_arquivo_excel, coluna_especifica=coluna_especifica)
        print(f">>> [DEBUG GRADE_EXCEL] Resultado de processar_planilha: {resultado_processar}")
        
        if resultado_processar:
            print(f"[GRADE_EXCEL] Planilha processada com sucesso")
            
            # Gerar scripts SCR
            print(f">>> [DEBUG GRADE_EXCEL] Chamando gerar_scripts_scr...")
            print(f">>> [DEBUG GRADE_EXCEL]   - caminho_arquivo_excel: {caminho_arquivo_excel}")
            resultado_gerar = automation.gerar_scripts_scr(caminho_arquivo_excel)
            print(f">>> [DEBUG GRADE_EXCEL] Resultado de gerar_scripts_scr: {resultado_gerar}")
            
            if resultado_gerar:
                print(f"[GRADE_EXCEL] Scripts SCR gerados com sucesso")
                return True
            else:
                print(f"[GRADE_EXCEL] Erro ao gerar scripts SCR")
                return False
        else:
            print(f"[GRADE_EXCEL] Erro ao processar planilha")
            return False
            
    except Exception as e:
        print(f"[GRADE_EXCEL] Erro durante processamento: {str(e)}")
        print(f">>> [DEBUG GRADE_EXCEL] ❌ EXCEÇÃO capturada: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Automação de Grades - Geração de Scripts SCR")
    parser.add_argument("excel", nargs="?", help="Caminho do arquivo Excel")
    parser.add_argument("colunas", nargs="?", help="Colunas a serem processadas (ex: E,F,G)")
    parser.add_argument("pavimento", nargs="?", help="Pavimento a ser processado (opcional)")
    args = parser.parse_args()

    if args.excel:
        caminho_arquivo_excel = args.excel
        colunas = args.colunas.split(",") if args.colunas else None
        pavimento = args.pavimento
        print(f"[HEADLESS] Processando arquivo: {caminho_arquivo_excel}, colunas: {colunas}, pavimento: {pavimento}")
        automation = GradeAutomation()
        
        # DEBUG: Verificar se o método existe
        print(f"[DEBUG METODO] gerar_scripts_scr existe na instância? {hasattr(automation, 'gerar_scripts_scr')}")
        print(f"[DEBUG METODO] gerar_scripts_scr existe na classe? {hasattr(GradeAutomation, 'gerar_scripts_scr')}")
        print(f"[DEBUG METODO] Métodos da instância que começam com 'gerar': {[m for m in dir(automation) if m.startswith('gerar')]}")
        
        if colunas:
            for coluna in colunas:
                print(f"Processando coluna {coluna}")
                ok = automation.processar_planilha(caminho_arquivo_excel, coluna_especifica=coluna)
                if ok:
                    # Chamar diretamente o método de geração de scripts
                    try:
                        automation.gerar_scripts_scr(caminho_arquivo_excel)
                    except AttributeError as e:
                        print(f"[ERRO] Falha ao chamar gerar_scripts_scr: {e}")
                    except Exception as e:
                        print(f"[ERRO] Erro durante gerar_scripts_scr: {e}")
        else:
            ok = automation.processar_planilha(caminho_arquivo_excel)
            if ok:
                # Chamar diretamente o método de geração de scripts
                try:
                    automation.gerar_scripts_scr(caminho_arquivo_excel)
                except AttributeError as e:
                    print(f"[ERRO] Falha ao chamar gerar_scripts_scr: {e}")
                except Exception as e:
                    print(f"[ERRO] Erro durante gerar_scripts_scr: {e}")
        print("Processamento concluído com sucesso!")
    else:
        # Modo antigo: interface gráfica
        main()