
# Helper de ofuscação (adicionado automaticamente)
def _get_obf_str(key):
    """Retorna string ofuscada"""
    _obf_map = {
        _get_obf_str("script.google.com"): base64.b64decode("=02bj5SZsd2bvdmL0BXayN2c"[::-1].encode()).decode(),
        _get_obf_str("macros/s/"): base64.b64decode("vM3Lz9mcjFWb"[::-1].encode()).decode(),
        _get_obf_str("AKfycbz"): base64.b64decode("==geiNWemtUQ"[::-1].encode()).decode(),
        _get_obf_str("credit"): base64.b64decode("0lGZlJ3Y"[::-1].encode()).decode(),
        _get_obf_str("saldo"): base64.b64decode("=8GZsF2c"[::-1].encode()).decode(),
        _get_obf_str("consumo"): base64.b64decode("==wbtV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("api_key"): base64.b64decode("==Qelt2XpBXY"[::-1].encode()).decode(),
        _get_obf_str("user_id"): base64.b64decode("==AZp9lclNXd"[::-1].encode()).decode(),
        _get_obf_str("calcular_creditos"): base64.b64decode("=M3b0lGZlJ3YfJXYsV3YsF2Y"[::-1].encode()).decode(),
        _get_obf_str("confirmar_consumo"): base64.b64decode("=8Wb1NnbvN2XyFWbylmZu92Y"[::-1].encode()).decode(),
        _get_obf_str("consultar_saldo"): base64.b64decode("vRGbhN3XyFGdsV3cu92Y"[::-1].encode()).decode(),
        _get_obf_str("debitar_creditos"): base64.b64decode("==wcvRXakVmcj9lchRXaiVGZ"[::-1].encode()).decode(),
        _get_obf_str("CreditManager"): base64.b64decode("==gcldWYuFWT0lGZlJ3Q"[::-1].encode()).decode(),
        _get_obf_str("obter_hwid"): base64.b64decode("==AZpdHafJXZ0J2b"[::-1].encode()).decode(),
        _get_obf_str("generate_signature"): base64.b64decode("lJXd0Fmbnl2cfVGdhJXZuV2Z"[::-1].encode()).decode(),
        _get_obf_str("encrypt_string"): base64.b64decode("=cmbpJHdz9Fdwlncj5WZ"[::-1].encode()).decode(),
        _get_obf_str("decrypt_string"): base64.b64decode("=cmbpJHdz9FdwlncjVGZ"[::-1].encode()).decode(),
        _get_obf_str("integrity_check"): base64.b64decode("rNWZoN2X5RXaydWZ05Wa"[::-1].encode()).decode(),
        _get_obf_str("security_utils"): base64.b64decode("=MHbpRXdflHdpJXdjV2c"[::-1].encode()).decode(),
        _get_obf_str("https://"): base64.b64decode("=8yL6MHc0RHa"[::-1].encode()).decode(),
        _get_obf_str("google.com"): base64.b64decode("==QbvNmLlx2Zv92Z"[::-1].encode()).decode(),
        _get_obf_str("apps.script"): base64.b64decode("=QHcpJ3Yz5ycwBXY"[::-1].encode()).decode(),
    }
    return _obf_map.get(key, key)

import sys
import os

# Configurar paths do projeto de forma robusta
def setup_project_paths():
    """Configura os paths do projeto de forma robusta"""
    try:
        # Tentar usar config_paths se disponível
        from ..config_paths import setup_paths, PROJECT_ROOT
        setup_paths()
        return PROJECT_ROOT
    except ImportError:
        try:
            from config_paths import setup_paths, PROJECT_ROOT
            setup_paths()
            return PROJECT_ROOT
        except ImportError:
            # Fallback manual
            current_file_dir = os.path.dirname(os.path.abspath(__file__))
            project_root_dir = os.path.dirname(os.path.dirname(current_file_dir))
            
            if project_root_dir not in sys.path:
                sys.path.insert(0, project_root_dir)
            
            return project_root_dir

# Configurar paths
project_root = setup_project_paths()
print(f"DEBUG: Projeto configurado em: {project_root}")

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pickle
import openpyxl
import pandas as pd
import win32com.client
import pythoncom
import win32gui
import importlib
import math
import json
from datetime import datetime
import numpy as np
import time
import threading
import keyboard  

# Módulo não encontrado - desabilitado
robo_script_generator = None

class ChanfroDetector:
    def __init__(self):
        self.tolerancia = 1e-3  
    def detectar_por_angulo(self, coords, min_x, max_x, min_y, max_y):
        """Método 1: Detecta chanfros baseado no ângulo entre segmentos."""
        chanfros = {
            'topo_esq': 0,
            'topo_dir': 0,
            'fundo_esq': 0,
            'fundo_dir': 0
        }
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        n = len(pontos)
        for i in range(n):
            p1 = pontos[i]
            p2 = pontos[(i+1)%n]
            p3 = pontos[(i+2)%n]
            v1 = (p2[0] - p1[0], p2[1] - p1[1])
            v2 = (p3[0] - p2[0], p3[1] - p2[1])
            angulo = abs(math.degrees(math.atan2(v2[1], v2[0]) - math.atan2(v1[1], v1[0])))
            if angulo > 180:
                angulo = 360 - angulo
            if 30 <= angulo <= 60:
                dist = math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2)
                if abs(p2[0] - min_x) < self.tolerancia:  
                    if abs(p2[1] - max_y) < self.tolerancia:  
                        chanfros['topo_esq'] = dist
                    elif abs(p2[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_esq'] = dist
                elif abs(p2[0] - max_x) < self.tolerancia:  
                    if abs(p2[1] - max_y) < self.tolerancia:  
                        chanfros['topo_dir'] = dist
                    elif abs(p2[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_dir'] = dist
        return chanfros
    def detectar_por_distancia(self, coords, min_x, max_x, min_y, max_y):
        """Método 2: Detecta chanfros baseado na distância entre pontos."""
        chanfros = {
            'topo_esq': 0,
            'topo_dir': 0,
            'fundo_esq': 0,
            'fundo_dir': 0
        }
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        n = len(pontos)
        for i in range(n):
            p1 = pontos[i]
            p2 = pontos[(i+1)%n]
            dist = math.sqrt((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2)
            if abs(p2[0] - p1[0]) > self.tolerancia and abs(p2[1] - p1[1]) > self.tolerancia:
                if abs(p1[0] - min_x) < self.tolerancia:  
                    if abs(p1[1] - max_y) < self.tolerancia:  
                        chanfros['topo_esq'] = dist
                    elif abs(p1[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_esq'] = dist
                elif abs(p1[0] - max_x) < self.tolerancia:  
                    if abs(p1[1] - max_y) < self.tolerancia:  
                        chanfros['topo_dir'] = dist
                    elif abs(p1[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_dir'] = dist
        return chanfros
    def detectar_por_projecao(self, coords, min_x, max_x, min_y, max_y):
        """Método 3: Detecta chanfros baseado na projeção dos segmentos."""
        chanfros = {
            'topo_esq': 0,
            'topo_dir': 0,
            'fundo_esq': 0,
            'fundo_dir': 0
        }
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        n = len(pontos)
        for i in range(n):
            p1 = pontos[i]
            p2 = pontos[(i+1)%n]
            proj_x = abs(p2[0] - p1[0])
            proj_y = abs(p2[1] - p1[1])
            if proj_x > self.tolerancia and proj_y > self.tolerancia:
                dist = (proj_x + proj_y) / 2
                if abs(p1[0] - min_x) < self.tolerancia:  
                    if abs(p1[1] - max_y) < self.tolerancia:  
                        chanfros['topo_esq'] = dist
                    elif abs(p1[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_esq'] = dist
                elif abs(p1[0] - max_x) < self.tolerancia:  
                    if abs(p1[1] - max_y) < self.tolerancia:  
                        chanfros['topo_dir'] = dist
                    elif abs(p1[1] - min_y) < self.tolerancia:  
                        chanfros['fundo_dir'] = dist
        return chanfros
    def detectar_por_retangulo_perfeito(self, coords, min_x, max_x, min_y, max_y):
        """Método 4: Detecta chanfros comparando com um retângulo perfeito."""
        chanfros = {
            'topo_esq': 0,
            'topo_dir': 0,
            'fundo_esq': 0,
            'fundo_dir': 0
        }
        retangulo_perfeito = [
            (min_x, min_y),  
            (max_x, min_y),  
            (max_x, max_y),  
            (min_x, max_y),  
            (min_x, min_y)   
        ]
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        for i, ponto in enumerate(pontos):
            for j, esquina in enumerate(retangulo_perfeito[:-1]):
                dist = math.sqrt((ponto[0] - esquina[0])**2 + (ponto[1] - esquina[1])**2)
                if dist < self.tolerancia:
                    prox_ponto = pontos[(i+1)%len(pontos)]
                    chanfro = math.sqrt((prox_ponto[0] - ponto[0])**2 + (prox_ponto[1] - ponto[1])**2)
                    if j == 0:  
                        chanfros['fundo_esq'] = chanfro
                    elif j == 1:  
                        chanfros['fundo_dir'] = chanfro
                    elif j == 2:  
                        chanfros['topo_dir'] = chanfro
                    elif j == 3:  
                        chanfros['topo_esq'] = chanfro
        return chanfros
    def detectar_chanfros_robusto(self, coords, min_x, max_x, min_y, max_y, debug=True):
        chanfros = {'topo_esq': 0, 'topo_dir': 0, 'fundo_esq': 0, 'fundo_dir': 0}
        esquinas = {
            'fundo_esq': (min_x, min_y),
            'fundo_dir': (max_x, min_y),
            'topo_dir': (max_x, max_y),
            'topo_esq': (min_x, max_y)
        }
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        tolerancia = 5.0  
        min_ch = 2.0
        max_ch = max(max_x-min_x, max_y-min_y) * 0.7
        for nome, esquina in esquinas.items():
            for idx, (px, py) in enumerate(pontos):
                if abs(px - esquina[0]) < tolerancia and abs(py - esquina[1]) < tolerancia:
                    prox_idx = (idx + 1) % len(pontos)
                    px2, py2 = pontos[prox_idx]
                    dx = px2 - px
                    dy = py2 - py
                    if abs(dx) > min_ch and abs(dy) > min_ch and abs(dx) < max_ch and abs(dy) < max_ch:
                        valor = abs(dx)
                        chanfros[nome] = valor
                        if debug:
                            print(f"[DEBUG] Esquina {nome}: ponto {idx} -> {prox_idx} | dx={dx:.2f} | chanfro={valor:.2f}")
        return chanfros
    def detectar_chanfros_hibrido(self, coords, min_x, max_x, min_y, max_y, debug=True):
        esquinas = {
            'fundo_esq': (min_x, min_y),
            'fundo_dir': (max_x, min_y),
            'topo_dir': (max_x, max_y),
            'topo_esq': (min_x, max_y)
        }
        nomes_ajustados = {
            'fundo_esq': 'topo_esq',
            'fundo_dir': 'topo_dir',
            'topo_dir': 'fundo_dir',
            'topo_esq': 'fundo_esq'
        }
        chanfros = {'topo_esq': 0, 'topo_dir': 0, 'fundo_esq': 0, 'fundo_dir': 0}
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        tolerancia = 5.0
        min_ch = 2.0
        max_ch = 100.0
        for nome, esquina in esquinas.items():
            valor_final = 0
            for idx, (px, py) in enumerate(pontos):
                if abs(px - esquina[0]) < tolerancia and abs(py - esquina[1]) < tolerancia:
                    prox_idx = (idx + 1) % len(pontos)
                    ant_idx = (idx - 1) % len(pontos)
                    px2, py2 = pontos[prox_idx]
                    dx = px2 - px
                    dy = py2 - py
                    if abs(dx) > min_ch and abs(dy) > min_ch and abs(dx) < max_ch and abs(dy) < max_ch:
                        valor_final = abs(dx)
                        if debug:
                            print(f"[HÍBRIDO] Esquina {nome}: segmento diagonal (próximo) dx={dx:.2f} -> {valor_final:.2f}")
                    px2a, py2a = pontos[ant_idx]
                    dx_a = px2a - px
                    dy_a = py2a - py
                    if abs(dx_a) > min_ch and abs(dy_a) > min_ch and abs(dx_a) < max_ch and abs(dy_a) < max_ch:
                        valor_final = max(valor_final, abs(dx_a))
                        if debug:
                            print(f"[HÍBRIDO] Esquina {nome}: segmento diagonal (anterior) dx={dx_a:.2f} -> {abs(dx_a):.2f}")
            if valor_final > min_ch:
                chanfros[nomes_ajustados[nome]] = valor_final
        return chanfros
class BoundaryLearner:
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.examples_file = os.path.join(self.script_dir, "boundary_examples.json")
        self.examples = self._load_examples()
        self.patterns = self._analyze_patterns()
        self.chanfro_detector = ChanfroDetector()
    def _load_examples(self):
        """Carrega exemplos salvos de boundaries."""
        try:
            if os.path.exists(self.examples_file):
                print(f"\nCarregando exemplos de: {self.examples_file}")
                with open(self.examples_file, 'r') as f:
                    exemplos = json.load(f)
                print(f"Exemplos carregados: {len(exemplos)}")
                return exemplos
            print(f"\nArquivo de exemplos não encontrado: {self.examples_file}")
            return []
        except Exception as e:
            print(f"\nErro ao carregar exemplos: {str(e)}")
            return []
    def _save_examples(self):
        """Salva exemplos de boundaries."""
        try:
            print(f"\nSalvando exemplos em: {self.examples_file}")
            with open(self.examples_file, 'w') as f:
                json.dump(self.examples, f, indent=2)
            print(f"Exemplos salvos com sucesso! Total: {len(self.examples)}")
        except Exception as e:
            print(f"\nErro ao salvar exemplos: {str(e)}")
    def _analyze_patterns(self):
        """Analisa padrões nos exemplos salvos."""
        patterns = {
            'chanfros': {
                'min_size': float('inf'),
                'max_size': 0,
                'avg_size': 0
            },
            'aberturas': {
                'min_size': float('inf'),
                'max_size': 0,
                'avg_size': 0
            }
        }
        if not self.examples:
            return patterns
        chanfro_sizes = []
        for example in self.examples:
            for chanfro in example['chanfros'].values():
                if chanfro > 0:
                    chanfro_sizes.append(chanfro)
        if chanfro_sizes:
            patterns['chanfros']['min_size'] = min(chanfro_sizes)
            patterns['chanfros']['max_size'] = max(chanfro_sizes)
            patterns['chanfros']['avg_size'] = sum(chanfro_sizes) / len(chanfro_sizes)
        abertura_sizes = []
        for example in self.examples:
            for abertura in example['aberturas']:
                abertura_sizes.append(abertura['largura'])
                abertura_sizes.append(abertura['altura'])
        if abertura_sizes:
            patterns['aberturas']['min_size'] = min(abertura_sizes)
            patterns['aberturas']['max_size'] = max(abertura_sizes)
            patterns['aberturas']['avg_size'] = sum(abertura_sizes) / len(abertura_sizes)
        return patterns
    def add_example(self, coords, chanfros, aberturas):
        """Adiciona um novo exemplo de boundary."""
        example = {
            'timestamp': datetime.now().isoformat(),
            'coords': coords,
            'chanfros': chanfros,
            'aberturas': aberturas
        }
        self.examples.append(example)
        self._save_examples()
        self.patterns = self._analyze_patterns()
    def get_recommendations(self, coords):
        """Retorna recomendações baseadas nos padrões aprendidos."""
        recommendations = {
            'chanfros': {},
            'aberturas': []
        }
        min_x = min(coords[::2])
        max_x = max(coords[::2])
        min_y = min(coords[1::2])
        max_y = max(coords[1::2])
        largura_total = max_x - min_x
        altura_total = max_y - min_y
        limiar_chanfro = self.patterns['chanfros']['avg_size']
        if limiar_chanfro == 0 or limiar_chanfro == float('inf'):
            limiar_chanfro = min(largura_total, altura_total) * 0.2  
        chanfros_hibrido = self.chanfro_detector.detectar_chanfros_hibrido(coords, min_x, max_x, min_y, max_y)
        print("\nResultado do método híbrido de detecção de chanfros:")
        print("=" * 50)
        for nome, valor in chanfros_hibrido.items():
            if valor > 0:
                print(f"{nome}: {valor:.2f}")
        recommendations['chanfros'] = chanfros_hibrido
        print(f"\nMétodo escolhido: híbrido (robusto e recuo dinâmico)")
        pontos = [(coords[i], coords[i+1]) for i in range(0, len(coords), 2)]
        n = len(pontos)
        aberturas_detectadas = []
        esquinas = [
            (min_x, min_y),  
            (min_x, max_y),  
            (max_x, min_y),  
            (max_x, max_y)   
        ]
        segmentos_em_abertura = set()
        for i in range(n):
            p0 = pontos[i-1]
            p1 = pontos[i]
            p2 = pontos[(i+1)%n]
            dx1 = p1[0] - p0[0]
            dy1 = p1[1] - p0[1]
            dx2 = p2[0] - p1[0]
            dy2 = p2[1] - p1[1]
            if not (p1 in esquinas or p2 in esquinas or p0 in esquinas):
                if abs(dx1) > 2 and abs(dy2) > 2 and abs(dx2) < 1e-3 and abs(dy1) < 1e-3:
                    largura = abs(dx1)
                    altura = abs(dy2)
                    x = min(p0[0], p1[0]) - min_x
                    y = min(p1[1], p2[1]) - min_y
                    if largura > 5 and altura > 5 and largura < largura_total*0.8 and altura < altura_total*0.8:
                        aberturas_detectadas.append({
                            'x': x,
                            'y': y,
                            'largura': largura,
                            'altura': altura
                        })
                        segmentos_em_abertura.add((p0, p1))
                        segmentos_em_abertura.add((p1, p2))
                elif abs(dy1) > 2 and abs(dx2) > 2 and abs(dy2) < 1e-3 and abs(dx1) < 1e-3:
                    largura = abs(dx2)
                    altura = abs(dy1)
                    x = min(p1[0], p2[0]) - min_x
                    y = min(p0[1], p1[1]) - min_y
                    if largura > 5 and altura > 5 and largura < largura_total*0.8 and altura < altura_total*0.8:
                        aberturas_detectadas.append({
                            'x': x,
                            'y': y,
                            'largura': largura,
                            'altura': altura
                        })
                        segmentos_em_abertura.add((p0, p1))
                        segmentos_em_abertura.add((p1, p2))
        aberturas_por_esquina = [None, None, None, None]
        area_max = [0] * 4
        for abertura in aberturas_detectadas:
            centro_x = abertura['x'] + abertura['largura']/2
            centro_y = abertura['y'] + abertura['altura']/2
            area = abertura['largura'] * abertura['altura']
            for idx, (ex, ey) in enumerate([(0, 0), (0, altura_total), (largura_total, 0), (largura_total, altura_total)]):
                dist = ((centro_x - ex)**2 + (centro_y - ey)**2)**0.5
                if dist < 50:  
                    if area > area_max[idx]:
                        area_max[idx] = area
                        aberturas_por_esquina[idx] = abertura
        aberturas_unicas = []
        for abertura in aberturas_por_esquina:
            if abertura is not None:
                existe = False
                for a in aberturas_unicas:
                    if (abs(a['x'] - abertura['x']) < 1e-2 and
                        abs(a['y'] - abertura['y']) < 1e-2 and
                        abs(a['largura'] - abertura['largura']) < 1e-2 and
                        abs(a['altura'] - abertura['altura']) < 1e-2):
                        existe = True
                        break
                if not existe:
                    aberturas_unicas.append(abertura)
        recommendations['aberturas'] = aberturas_unicas
        return recommendations
class FundoProducaoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Fundo Produção - Gerenciador de Fundos")
        self.geometry("900x800")
        self.configure(bg='#f0f0f0')
        # Obter o diretório do script atual
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.fundos_salvos = {}
        self.excel_file = None
        self.ac = None
        self.learner = BoundaryLearner()
        self._preview_cache = {}  
        self._last_update = 0  
        self._update_delay = 100  
        self._update_scheduled = None  
        self._carregando_fundo = False  
        self._ignorar_traces = False    
        self._is_editing_entry = False 
        self._prevent_treeview_load_after_edit = False 
        self._block_treeview_events = False 
        self.tipo_distribuicao_var = tk.StringVar(value="122")  
        self.tipo_painel_inicial_var = tk.StringVar(value="300")  
        self.usar_boundary_duplo = False
        self.tipo_painel2_var = tk.StringVar(value="E/T")  
        self.comprimento_2_var = tk.DoubleVar()
        self.largura_2_var = tk.DoubleVar()
        self.p1_2_var = tk.DoubleVar()
        self.p2_2_var = tk.DoubleVar()
        self.p3_2_var = tk.DoubleVar()
        self.recuos_2_vars = [tk.DoubleVar() for _ in range(4)]
        self.aberturas_2_vars = [[tk.DoubleVar() for _ in range(3)] for _ in range(4)]  
        self.numero_var = tk.StringVar()
        self.nome_var = tk.StringVar()
        self.obs_var = tk.StringVar()
        self.pavimento_var = tk.StringVar()
        self.largura_var = tk.DoubleVar()
        self.altura_var = tk.DoubleVar()
        self.texto_esq_var = tk.StringVar()
        self.texto_dir_var = tk.StringVar()
        self.paineis_vars = [tk.DoubleVar() for _ in range(6)]
        self.recuos_vars = [tk.DoubleVar() for _ in range(4)]
        self.aberturas_vars = [[tk.DoubleVar() for _ in range(3)] for _ in range(4)]  
        self.sarrafo_esq_var = tk.BooleanVar(value=True)
        self.sarrafo_dir_var = tk.BooleanVar(value=True)
        def trace_salvar(*args):
            if not self._carregando_fundo and not self._ignorar_traces:
                self._salvar_fundo()
        all_vars = [
            self.numero_var, self.nome_var, self.obs_var, self.pavimento_var,
            self.largura_var, self.altura_var, self.texto_esq_var, self.texto_dir_var,
            self.tipo_distribuicao_var, self.tipo_painel_inicial_var, self.tipo_painel2_var,
            self.sarrafo_esq_var, self.sarrafo_dir_var,
            *self.paineis_vars, *self.recuos_vars, *sum(self.aberturas_vars, []),
            self.comprimento_2_var, self.largura_2_var, self.p1_2_var, self.p2_2_var, self.p3_2_var,
            *self.recuos_2_vars, *sum(self.aberturas_2_vars, [])
        ]
        for var in all_vars:
            var.trace_add('write', trace_salvar)
        self.largura_var.trace_add('write', self._calcular_area_util)
        self.altura_var.trace_add('write', self._calcular_area_util)
        for linha in self.aberturas_vars:
            for var_abertura in linha:
                var_abertura.trace_add('write', self._calcular_area_util)
        # Adicionar _calcular_area_util para todas as variáveis numéricas que afetam o cálculo
        # Isso garante que a area_util_var seja sempre atualizada antes de ser salva ou exibida.
        for var in [
            *self.paineis_vars, *self.recuos_vars,
            self.comprimento_2_var, self.largura_2_var, self.p1_2_var, self.p2_2_var, self.p3_2_var,
            *self.recuos_2_vars, *sum(self.aberturas_2_vars, [])
        ]:
            var.trace_add('write', self._calcular_area_util)
        self._setup_ui()
        self._carregar_fundos_salvos()
        self._calcular_area_util() 
        def cancelar_selecao_interno():
            print('[ESC] Cancelamento interno de seleção!')
            if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
                try:
                    self._balcao_flutuante.destroy()
                except:
                    pass
                self._balcao_flutuante = None
            if hasattr(self, '_flutuante_after_id') and self._flutuante_after_id is not None:
                try:
                    self.after_cancel(self._flutuante_after_id)
                except:
                    pass
                self._flutuante_after_id = None
            self._flutuante_cancelado = True
            try:
                self.lift()
                self.focus_force()
            except:
                pass
            print('[ESC] Seleção cancelada (NÃO REINICIA)')
        self._cancelar_selecao_interno = cancelar_selecao_interno
    def _setup_ui(self):
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        preview_frame = ttk.LabelFrame(main_frame, text="Visualização", padding=5)
        preview_frame.pack(fill=tk.X, padx=10, pady=5)
        self._criar_preview(preview_frame)
        botoes_frame = ttk.Frame(main_frame)
        botoes_frame.pack(fill=tk.X, padx=10, pady=5)
        self._criar_botoes(botoes_frame)
        content_container = ttk.Frame(main_frame)
        content_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.canvas_scroll = tk.Canvas(content_container)
        scrollbar_y = ttk.Scrollbar(content_container, orient="vertical", command=self.canvas_scroll.yview)
        scrollbar_x = ttk.Scrollbar(content_container, orient="horizontal", command=self.canvas_scroll.xview)
        content_frame = ttk.Frame(self.canvas_scroll)
        self.canvas_scroll.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        self.canvas_scroll.pack(side="left", fill="both", expand=True)
        canvas_frame = self.canvas_scroll.create_window((0, 0), window=content_frame, anchor="nw")
        content_frame.bind("<Configure>", lambda e: self.canvas_scroll.configure(scrollregion=self.canvas_scroll.bbox("all")))
        def _on_mousewheel(event):
            self.canvas_scroll.yview_scroll(int(-1*(event.delta/120)), "units")
        self.canvas_scroll.bind_all("<MouseWheel>", _on_mousewheel)
        self.canvas_scroll.bind_all("<Button-4>", lambda e: self.canvas_scroll.yview_scroll(-1, "units"))
        self.canvas_scroll.bind_all("<Button-5>", lambda e: self.canvas_scroll.yview_scroll(1, "units"))
        content_frame.columnconfigure(1, weight=1)
        content_frame.rowconfigure(0, weight=1)
        self._criar_lista_fundos(content_frame)
        self._criar_aba_dados_gerais(content_frame)
    def _schedule_preview_update(self, *args):
        """Agenda uma atualização do preview com debounce."""
        if self._update_scheduled is not None:
            self.after_cancel(self._update_scheduled)
        self._update_scheduled = self.after(100, self._atualizar_preview)
    def vincular_eventos_alteracao(self):
        """Vincula os eventos de alteração aos campos."""
        for var in [self.largura_var, self.altura_var, *self.paineis_vars, 
                    *self.recuos_vars, *sum(self.aberturas_vars, []),
                    self.sarrafo_esq_var, self.sarrafo_dir_var]:
            var.trace_add('write', self._schedule_preview_update)
    def _criar_preview(self, parent):
        self.preview_frame = ttk.LabelFrame(parent, text="Visualização", padding=5)
        self.preview_frame.pack(fill=tk.X, expand=True)
        self.cores = {
            'fundo': "#FFFFFF",
            'painel': "#E3F2FD",
            'sarrafo': "#FF9800",
            'abertura': "#FFEB3B",
            'chanfro': "#4CAF50",
            'divisao': "#000000",
            'cota': "#2196F3",
            'texto': "#000000"
        }
        self.canvas_frame = ttk.Frame(self.preview_frame)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(self.canvas_frame, width=450, height=220, 
                              bg=self.cores['fundo'],
                              scrollregion=(0, 0, 900, 400))
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.zoom_level = 1.0
        self.pan_start = None
        self.canvas.bind("<ButtonPress-1>", self._start_pan)
        self.canvas.bind("<B1-Motion>", self._pan)
        self.canvas.bind("<ButtonRelease-1>", self._stop_pan)
        self.canvas.bind("<MouseWheel>", self._mouse_wheel)
        self.canvas.bind("<Button-4>", self._mouse_wheel)
        self.canvas.bind("<Button-5>", self._mouse_wheel)
        self.preview_frame.update_idletasks()
        self.canvas.update_idletasks()
        self._schedule_preview_update()
    def _mouse_wheel(self, event):
        if event.num == 5 or event.delta < 0:
            self._zoom_out()
        else:
            self._zoom_in()
    def _zoom_in(self):
        new_zoom = min(self.zoom_level + 0.1, 2.0)
        self._update_zoom(new_zoom)
    def _zoom_out(self):
        new_zoom = max(self.zoom_level - 0.1, 0.5)
        self._update_zoom(new_zoom)
    def _update_zoom(self, value):
        try:
            self.zoom_level = float(value)
            self._atualizar_preview()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        except ValueError:
            pass
    def _start_pan(self, event):
        self.canvas.scan_mark(event.x, event.y)
        self.pan_start = (event.x, event.y)
    def _pan(self, event):
        if self.pan_start:
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            self.pan_start = (event.x, event.y)
    def _stop_pan(self, event):
        self.pan_start = None
    def _atualizar_preview(self, *args):
        """Atualiza o preview do painel."""
        self._update_scheduled = None  
        if not hasattr(self, 'canvas') or not self.canvas.winfo_exists():
            return
        try:
            self.canvas.delete("all")
            try:
                largura = float(self.largura_var.get() or 0)
                altura = float(self.altura_var.get() or 0)
            except (tk.TclError, ValueError):
                return
            if largura <= 0 or altura <= 0:
                return
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            if canvas_width <= 1 or canvas_height <= 1:
                self.canvas.update_idletasks()
                canvas_width = self.canvas.winfo_width()
                canvas_height = self.canvas.winfo_height()
                if canvas_width <= 1 or canvas_height <= 1:
                    self._schedule_preview_update()
                    return
            escala_x = 400 / (largura * 1.2) * self.zoom_level
            escala_y = 200 / (altura * 1.2) * self.zoom_level
            escala = min(escala_x, escala_y)
            x_inicial = 50 * self.zoom_level
            y_inicial = 250 * self.zoom_level
            self._desenhar_elementos(x_inicial, y_inicial, largura, altura, escala)
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
            if not self._ignorar_traces:
                self.after(50, self._reset_load_prevention_flag)
        except Exception as e:
            print(f"Erro ao atualizar preview: {str(e)}")
            self._schedule_preview_update()
    def _desenhar_elementos(self, x_inicial, y_inicial, largura, altura, escala):
        """Desenha todos os elementos do preview."""
        self.canvas.create_line(
            x_inicial, y_inicial,
            x_inicial + largura * escala, y_inicial,
            fill=self.cores['cota'], dash=(4, 4)
        )
        self.canvas.create_line(
            x_inicial, y_inicial - altura * escala,
            x_inicial + largura * escala, y_inicial - altura * escala,
            fill=self.cores['cota'], dash=(4, 4)
        )
        self._desenhar_cotas(x_inicial, y_inicial, largura, altura, escala)
        self._desenhar_painel_principal(x_inicial, y_inicial, largura, altura, escala)
        if self.sarrafo_esq_var.get():
            self._desenhar_sarrafo(x_inicial, y_inicial, largura, altura, escala, 'esquerda')
        if self.sarrafo_dir_var.get():
            self._desenhar_sarrafo(x_inicial, y_inicial, largura, altura, escala, 'direita')
        self._desenhar_aberturas(x_inicial, y_inicial, largura, altura, escala)
        self._desenhar_textos(x_inicial, y_inicial, largura, altura, escala)
        try:
            comprimento_2 = float(self.comprimento_2_var.get() or 0)
            largura_2 = float(self.largura_2_var.get() or 0)
            if comprimento_2 > 0 and largura_2 > 0:
                self._desenhar_painel_l(x_inicial, y_inicial, largura, altura, 
                                      comprimento_2, largura_2, escala)
        except (ValueError, tk.TclError):
            pass
    def _desenhar_painel_l(self, x_inicial, y_inicial, largura_principal, altura_principal, 
                          comprimento, largura, escala):
        """Desenha o painel em L baseado no tipo selecionado, colado na lateral do painel principal."""
        tipo_l = self.tipo_painel2_var.get()
        if comprimento < largura:
            l_final = comprimento
            h_final = largura
        else:
            l_final = largura
            h_final = comprimento
        if tipo_l.startswith("E"):  
            x = x_inicial - l_final * escala  
        else:  
            x = x_inicial + largura_principal * escala  
        if tipo_l.endswith("T"):  
            y = y_inicial  
            y_final = y - h_final * escala
        else:  
            y = y_inicial - altura_principal * escala + h_final * escala  
            y_final = y - h_final * escala
        self.canvas.create_rectangle(
            x, y,
            x + l_final * escala, y_final,
            fill=self.cores['painel'],
            outline=self.cores['divisao']
        )
        # Desenhar triângulos vermelhos para os chanfros do painel em L
        chanfro_color = '#FF3333'
        recuos = [float(v.get() or 0) for v in self.recuos_2_vars]
        if h_final > l_final:  
            if recuos[0] > 0:
                self.canvas.create_polygon(
                    [x, y,  
                     x + recuos[0] * escala, y,  
                     x, y - recuos[0] * escala if y_final < y else y + recuos[0] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[3] > 0:
                self.canvas.create_polygon(
                    [x + l_final * escala, y_final,  
                     x + (l_final - recuos[3]) * escala, y_final,  
                     x + l_final * escala, y_final + recuos[3] * escala if y_final < y else y_final - recuos[3] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[2] > 0:
                self.canvas.create_polygon(
                    [x, y_final,  
                     x + recuos[2] * escala, y_final,  
                     x, y_final + recuos[2] * escala if y_final < y else y_final - recuos[2] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[1] > 0:
                self.canvas.create_polygon(
                    [x + l_final * escala, y,  
                     x + (l_final - recuos[1]) * escala, y,  
                     x + l_final * escala, y - recuos[1] * escala if y_final < y else y + recuos[1] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
        else:  
            if recuos[0] > 0:
                self.canvas.create_polygon(
                    [x, y_final,  
                     x + recuos[0] * escala, y_final,  
                     x, y_final + recuos[0] * escala if y_final < y else y_final - recuos[0] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[1] > 0:
                self.canvas.create_polygon(
                    [x, y,  
                     x + recuos[1] * escala, y,  
                     x, y - recuos[1] * escala if y_final < y else y + recuos[1] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[2] > 0:
                self.canvas.create_polygon(
                    [x + l_final * escala, y_final,  
                     x + (l_final - recuos[2]) * escala, y_final,  
                     x + l_final * escala, y_final + recuos[2] * escala if y_final < y else y_final - recuos[2] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[3] > 0:
                self.canvas.create_polygon(
                    [x + l_final * escala, y,  
                     x + (l_final - recuos[3]) * escala, y,  
                     x + l_final * escala, y - recuos[3] * escala if y_final < y else y + recuos[3] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
        self._desenhar_cotas_painel_l(x, y, l_final, h_final, escala, tipo_l, y_final)
        self._desenhar_aberturas_painel_l(x, y, l_final, h_final, escala, tipo_l, y_final)
    def _desenhar_cotas_painel_l(self, x, y, largura, altura, escala, tipo_l, y_final):
        """Desenha as cotas do painel em L."""
        y_cota = y + 20 if tipo_l.endswith("F") else y - 20
        self.canvas.create_line(
            x, y_cota,
            x + largura * escala, y_cota,
            fill=self.cores['cota'], arrow="both"
        )
        self.canvas.create_text(
            x + (largura * escala) / 2, y_cota + (15 if tipo_l.endswith("F") else -15),
            text=f"Larg.: {largura:.1f}",
            fill=self.cores['texto']
        )
        x_cota = x + largura * escala + 20
        self.canvas.create_line(
            x_cota, y,
            x_cota, y_final,
            fill=self.cores['cota'], arrow="both"
        )
        self.canvas.create_text(
            x_cota + 15, (y + y_final) / 2,
            text=f"Alt.: {altura:.1f}",
            fill=self.cores['texto'],
            angle=90
        )
    def _desenhar_aberturas_painel_l(self, x, y, largura, altura, escala, tipo_l, y_final):
        """Desenha as aberturas do painel em L SEM campo de distância (sempre encostadas na parede), invertendo profundidade/largura (prof em X, larg em Y)."""
        for i, linha in enumerate(self.aberturas_2_vars):
            valores = [float(v.get() or 0) for v in linha]
            if valores[0] > 0 and valores[1] > 0:
                prof, larg = valores  
                if altura > largura:
                    if i == 0:  
                        x_pos = x
                        y_pos = y - larg * escala
                    elif i == 1:  
                        x_pos = x + largura * escala - prof * escala
                        y_pos = y - larg * escala
                    elif i == 2:  
                        x_pos = x
                        y_pos = y_final
                    elif i == 3:  
                        x_pos = x + largura * escala - prof * escala
                        y_pos = y_final
                    else:
                        continue
                else:  
                    if i == 0:  
                        x_pos = x
                        y_pos = y_final
                    elif i == 1:  
                        x_pos = x + largura * escala - prof * escala
                        y_pos = y - larg * escala
                    elif i == 2:  
                        x_pos = x + largura * escala - prof * escala
                        y_pos = y_final
                    elif i == 3:  
                        x_pos = x
                        y_pos = y - larg * escala
                    else:
                        continue
                self.canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + prof * escala, y_pos + larg * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura(x_pos, y_pos, x_pos + prof * escala, y_pos + larg * escala)
    def _desenhar_hachura_abertura(self, x1, y1, x2, y2):
        largura = x2 - x1
        altura = y2 - y1
        if largura < 5 or altura < 5:
            return
        espacamento = 5
        for i in range(0, int(largura + altura), espacamento):
            x_inicio = x1 + min(i, largura) if i < largura else x2
            y_inicio = y1 if i < largura else y1 + (i - largura)
            x_fim = x1 if i < altura else x1 + (i - altura)
            y_fim = y1 + min(i, altura) if i < altura else y2
            if all(coord <= limite for coord, limite in [(x_inicio, x2), (y_inicio, y2), (x_fim, x2), (y_fim, y2)]):
                self.canvas.create_line(
                    x_inicio, y_inicio,
                    x_fim, y_fim,
                    fill=self.cores['divisao'],
                    width=1
                )
    def _desenhar_textos(self, x_inicial, y_inicial, largura, altura, escala):
        self.canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial - altura * escala - 40, 
            text=f"Fundo: {self.numero_var.get() or ''} - {self.nome_var.get() or ''}",
            fill=self.cores['texto'],
            font=("Arial", 12, "bold"), 
            anchor='s'
        )
        self.canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial + 50,
            text=f"{self.nome_var.get() or ''} - {self.obs_var.get() or ''}",
            fill=self.cores['texto']
        )
        self.canvas.create_text(
            x_inicial - 50, y_inicial - (altura * escala) / 2,
            text=self.texto_esq_var.get() or "",
            fill=self.cores['texto'],
            angle=90
        )
        self.canvas.create_text(
            x_inicial + largura * escala + 50, y_inicial - (altura * escala) / 2,
            text=self.texto_dir_var.get() or "",
            fill=self.cores['texto'],
            angle=90
        )
    def _criar_botoes(self, parent):
        botoes_container = ttk.Frame(parent)
        botoes_container.pack(fill=tk.X, pady=5)
        botoes_frame = ttk.Frame(botoes_container)
        botoes_frame.pack(expand=True)
        ttk.Button(botoes_frame, text="Novo", command=self._limpar_campos).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Salvar", command=self._salvar_fundo).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Selecionar Excel", command=self._selecionar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Criar Excel", command=self._criar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Importar Excel", command=self._importar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Exportar Excel", command=self._exportar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Excluir Selecionados", command=self._excluir_selecionados).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Salvar Teste", command=self._salvar_teste_e_gerar_script).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Salvar Pavimento", command=self._salvar_pavimento_e_gerar_multiplos_scripts).pack(side=tk.LEFT, padx=5)
    def _reiniciar_selecao(self):
        self._selecionar_textos_consecutivos()
    def _criar_lista_fundos(self, parent):
        self.lista_frame = ttk.Frame(parent)
        self.lista_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.pavimento_filtro_var = tk.StringVar(value="Todos")
        self.combo_pavimento = ttk.Combobox(self.lista_frame, textvariable=self.pavimento_filtro_var, state="readonly", width=12)
        self.combo_pavimento.pack(anchor="w", padx=2, pady=(2, 1), fill=tk.X)
        self.combo_pavimento.bind("<<ComboboxSelected>>", self._atualizar_lista)
        self.total_m2_por_pavimento_var = tk.StringVar(value="Total M²: 0.00")
        self.total_m2_por_pavimento_label = ttk.Label(self.lista_frame, textvariable=self.total_m2_por_pavimento_var, font=("Arial", 10, "bold"), foreground="#1976d2")
        self.total_m2_por_pavimento_label.pack(anchor="w", padx=2, pady=(0, 2))
        ttk.Label(self.lista_frame, text="Fundos Salvos").pack(anchor="w")
        self.lista_fundos = ttk.Treeview(self.lista_frame, columns=("Número", "Nome", "Pavimento", "M² Individual"), show="headings", height=25, selectmode='extended')
        self.lista_fundos.heading("Número", text="Número", command=lambda: self._ordenar_lista_fundos('Número'))
        self.lista_fundos.heading("Nome", text="Nome", command=lambda: self._ordenar_lista_fundos('Nome'))
        self.lista_fundos.heading("Pavimento", text="Pavimento", command=lambda: self._ordenar_lista_fundos('Pavimento'))
        self.lista_fundos.heading("M² Individual", text="M² Individual", command=lambda: self._ordenar_lista_fundos('M² Individual')) 
        self.lista_fundos.column("Número", width=80)
        self.lista_fundos.column("Nome", width=120)
        self.lista_fundos.column("Pavimento", width=80)
        self.lista_fundos.column("M² Individual", width=100, anchor="e") 
        self.lista_fundos.pack(fill=tk.BOTH, expand=True)
        self.lista_fundos.bind('<<TreeviewSelect>>', self._carregar_fundo_selecionado)
        self.menu_contexto = tk.Menu(self, tearoff=0)
        self.menu_contexto.add_command(label="Excluir", command=self._excluir_fundo)
        self.lista_fundos.bind("<Button-3>", self._mostrar_menu_contexto)
        self._coluna_ordenada = None
        self._ordem_crescente = True
        btn_excluir_mult = ttk.Button(self.lista_frame, text="Excluir Selecionados", command=self._excluir_selecionados)
        btn_excluir_mult.pack(fill=tk.X, pady=(4, 2))
    def _ordenar_lista_fundos(self, coluna):
        col_map = {'Número': 0, 'Nome': 1, 'Pavimento': 2, 'M² Individual': 3} 
        idx = col_map[coluna]
        itens = [(self.lista_fundos.item(child)['values'], child) for child in self.lista_fundos.get_children()]
        if self._coluna_ordenada == coluna:
            self._ordem_crescente = not self._ordem_crescente
        else:
            self._ordem_crescente = True
        self._coluna_ordenada = coluna
        def try_float(val):
            try:
                return float(val)
            except:
                return val
        if coluna in ['Número', 'M² Individual']: 
            itens.sort(key=lambda x: try_float(x[0][idx]), reverse=not self._ordem_crescente)
        else:
            itens.sort(key=lambda x: str(x[0][idx]).lower(), reverse=not self._ordem_crescente)
        self.lista_fundos.delete(*self.lista_fundos.get_children())
        for values, _ in itens:
            self.lista_fundos.insert('', 'end', values=values)
    def _criar_aba_dados_gerais(self, parent):
        self.dados_frame = ttk.LabelFrame(parent, text="Dados do Fundo", padding=10)
        self.dados_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        self.dados_frame.columnconfigure(0, weight=1)
        self.dados_frame.columnconfigure(1, weight=1)
        botoes_topo_frame = ttk.Frame(self.dados_frame)
        botoes_topo_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 2))
        def abrir_producao_manual():
            largura = float(self.largura_var.get() or 0)
            altura = float(self.altura_var.get() or 0)
            chanfros = {
                'topo_esq': float(self.recuos_vars[0].get() or 0),
                'fundo_esq': float(self.recuos_vars[1].get() or 0),
                'topo_dir': float(self.recuos_vars[2].get() or 0),
                'fundo_dir': float(self.recuos_vars[3].get() or 0)
            }
            aberturas = []
            for i, linha in enumerate(self.aberturas_vars):
                valores = [float(v.get() or 0) for v in linha]
                if all(v > 0 for v in valores):
                    aberturas.append({'x': 0, 'altura': valores[0], 'largura': valores[1]})
            msg = f"Dimensões: {largura:.2f} x {altura:.2f}\n\nChanfros:\n"
            for nome, valor in chanfros.items():
                msg += f"{nome}: {valor:.2f}\n"
            msg += "\nAberturas:\n"
            if not aberturas:
                msg += "Nenhuma abertura identificada\n"
            else:
                for i, abertura in enumerate(aberturas):
                    msg += (f"Abertura {i+1}:\n  Posição: ({abertura['x']:.2f}, ?)\n"
                            f"  Largura: {abertura['largura']:.2f}\n  Profundidade: {abertura['altura']:.2f}\n")
            self._janela_confirmacao_salvamento(msg, largura, altura, aberturas, lambda _: None)
        ttk.Button(botoes_topo_frame, text="Iniciar Produção", command=abrir_producao_manual).pack(side="left", padx=5)
        ttk.Button(botoes_topo_frame, text="Boundary", command=self._realizar_boundary).pack(side="left", padx=5)
        ttk.Button(botoes_topo_frame, text="Select", command=self._selecionar_textos_consecutivos).pack(side="left", padx=5)
        ttk.Button(botoes_topo_frame, text="Boundary L", command=self._executar_boundary_l).pack(side="left", padx=5)
        info_frame = ttk.LabelFrame(self.dados_frame, text="Informações Básicas", padding=5)
        info_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=5)
        self.numero_var = tk.StringVar()
        self.nome_var = tk.StringVar()
        self.obs_var = tk.StringVar()
        self.pavimento_var = tk.StringVar()
        self.largura_var = tk.DoubleVar()
        self.altura_var = tk.DoubleVar()
        self.texto_esq_var = tk.StringVar()
        self.texto_dir_var = tk.StringVar()
        self.paineis_vars = [tk.DoubleVar() for _ in range(6)]
        self.recuos_vars = [tk.DoubleVar() for _ in range(4)]
        self.aberturas_vars = [[tk.DoubleVar() for _ in range(3)] for _ in range(4)]  
        self.sarrafo_esq_var = tk.BooleanVar(value=True)
        self.sarrafo_dir_var = tk.BooleanVar(value=True)
        self.tipo_distribuicao_var = tk.StringVar(value="122")  
        for i in range(2):
            info_frame.columnconfigure(i, weight=1)
        numero_frame = ttk.Frame(info_frame)
        numero_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=5)
        ttk.Label(numero_frame, text="Número:").pack(side="left")
        ttk.Entry(numero_frame, textvariable=self.numero_var, width=8).pack(side="left", fill="x", expand=True, padx=5)
        nome_frame = ttk.Frame(info_frame)
        nome_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=5)
        ttk.Label(nome_frame, text="Nome:").pack(side="left")
        ttk.Entry(nome_frame, textvariable=self.nome_var).pack(side="left", fill="x", expand=True, padx=5)
        obs_frame = ttk.Frame(info_frame)
        obs_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=5)
        ttk.Label(obs_frame, text="Observações:").pack(side="left")
        ttk.Entry(obs_frame, textvariable=self.obs_var).pack(side="left", fill="x", expand=True, padx=5)
        ttk.Label(info_frame, text="Pavimento:").grid(row=3, column=0, sticky="w", padx=5)
        ttk.Entry(info_frame, textvariable=self.pavimento_var).grid(row=3, column=1, sticky="ew", padx=5)
        ttk.Label(info_frame, text="Largura:").grid(row=0, column=2, sticky="w", padx=5)
        largura_entry = ttk.Entry(info_frame, textvariable=self.largura_var)
        largura_entry.grid(row=0, column=3, sticky="ew", padx=5)
        ttk.Label(info_frame, text="Altura:").grid(row=1, column=2, sticky="w", padx=5)
        ttk.Entry(info_frame, textvariable=self.altura_var).grid(row=1, column=3, sticky="ew", padx=5)
        ttk.Label(info_frame, text="Texto Esquerda:").grid(row=2, column=2, sticky="w", padx=5)
        ttk.Entry(info_frame, textvariable=self.texto_esq_var).grid(row=2, column=3, sticky="ew", padx=5)
        ttk.Label(info_frame, text="Texto Direita:").grid(row=3, column=2, sticky="w", padx=5)
        ttk.Entry(info_frame, textvariable=self.texto_dir_var).grid(row=3, column=3, sticky="ew", padx=5)
        for var in [self.largura_var, self.altura_var, *self.paineis_vars, *self.recuos_vars, *sum(self.aberturas_vars, [])]:
            var.trace_add('write', lambda *a: self._atualizar_preview())
        largura_entry.bind('<KeyRelease>', self._calcular_paineis_automatico)
        largura_entry.bind('<FocusOut>', self._calcular_paineis_automatico)
        paineis_frame = ttk.LabelFrame(self.dados_frame, text="Painéis", padding=10)
        paineis_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=10)
        radio_frame = ttk.Frame(paineis_frame)
        radio_frame.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 8))
        ttk.Label(radio_frame, text="Distribuição do último painel:").pack(anchor="w", padx=5, pady=(0,2))
        ttk.Radiobutton(radio_frame, text="122 (se último <60, penúltimo vira 122 e o restante vai para o último)", variable=self.tipo_distribuicao_var, value="122", command=self._calcular_paineis_automatico).pack(anchor="w", padx=20, pady=2)
        ttk.Radiobutton(radio_frame, text="307 (se último <63, soma ao anterior até 307 máx)", variable=self.tipo_distribuicao_var, value="307", command=self._calcular_paineis_automatico).pack(anchor="w", padx=20, pady=2)
        radio_frame_inicial = ttk.Frame(paineis_frame)
        radio_frame_inicial.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(0, 8))
        ttk.Label(radio_frame_inicial, text="Medida máx para o primeiro painel: 244 (padrão), 300, 307").pack(side="left", padx=5, pady=(0,2))
        ttk.Radiobutton(radio_frame_inicial, text="244", variable=self.tipo_painel_inicial_var, value="244", command=self._calcular_paineis_automatico).pack(side="left", padx=5, pady=2)
        ttk.Radiobutton(radio_frame_inicial, text="300", variable=self.tipo_painel_inicial_var, value="300", command=self._calcular_paineis_automatico).pack(side="left", padx=5, pady=2)
        ttk.Radiobutton(radio_frame_inicial, text="307", variable=self.tipo_painel_inicial_var, value="307", command=self._calcular_paineis_automatico).pack(side="left", padx=5, pady=2)
        divisao_frame = ttk.Frame(paineis_frame)
        divisao_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(0, 8))
        self.tipo_divisao_paineis_var = tk.StringVar(value="244")
        ttk.Label(divisao_frame, text="Divisão dos painéis intermediários:").pack(anchor="w", padx=5, pady=(0,2))
        ttk.Radiobutton(divisao_frame, text="Divisão de 244 em 244 (padrão)", variable=self.tipo_divisao_paineis_var, value="244", command=self._calcular_paineis_automatico).pack(anchor="w", padx=20, pady=2)
        ttk.Radiobutton(divisao_frame, text="Divisão de 307 em 307", variable=self.tipo_divisao_paineis_var, value="307", command=self._calcular_paineis_automatico).pack(anchor="w", padx=20, pady=2)
        self.paineis_entries = []
        painel_labels_frame = ttk.Frame(paineis_frame)
        painel_labels_frame.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(0, 2))
        for i in range(6):
            ttk.Label(painel_labels_frame, text=f"P{i+1}:").grid(row=0, column=i, sticky="ew", padx=(8,8), pady=(0,2))
        painel_inputs_frame = ttk.Frame(paineis_frame)
        painel_inputs_frame.grid(row=4, column=0, columnspan=4, sticky="ew", pady=(0, 8))
        for i in range(6):
            entry = ttk.Entry(painel_inputs_frame, textvariable=self.paineis_vars[i], width=8)
            entry.grid(row=0, column=i, padx=(8,8), pady=(0,2))
            self.paineis_entries.append(entry)
        opcoes_frame = ttk.LabelFrame(self.dados_frame, text="Opções", padding=5)
        opcoes_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo vertical esquerda", variable=self.sarrafo_esq_var, command=self._atualizar_preview).grid(row=0, column=0, padx=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo vertical direita", variable=self.sarrafo_dir_var, command=self._atualizar_preview).grid(row=0, column=1, padx=5)
        chanfros_aberturas_frame = ttk.LabelFrame(self.dados_frame, text="Chanfros e Aberturas", padding=5)
        chanfros_aberturas_frame.grid(row=4, column=0, columnspan=2, sticky="ew", pady=5)
        chanfros_frame = ttk.Frame(chanfros_aberturas_frame)
        chanfros_frame.grid(row=0, column=0, sticky="nsew", padx=5)
        labels = ["T/E:", "F/E:", "T/D:", "F/D:"]
        for i, label in enumerate(labels):
            ttk.Label(chanfros_frame, text=label).grid(row=i, column=0, sticky="w", padx=5)
            ttk.Entry(chanfros_frame, textvariable=self.recuos_vars[i]).grid(row=i, column=1, sticky="ew", padx=5)
        aberturas_frame = ttk.Frame(chanfros_aberturas_frame)
        aberturas_frame.grid(row=0, column=1, sticky="nsew", padx=5)
        aberturas_labels = ["Dist", "Prof", "Larg"]
        lados = ["T/E", "F/E", "T/D", "F/D"]
        ttk.Label(aberturas_frame, text="").grid(row=0, column=0, sticky="w", padx=5)
        for j, lab in enumerate(aberturas_labels):
            ttk.Label(aberturas_frame, text=lab, font=("Arial", 8, "bold")).grid(row=0, column=j+1, sticky="w", padx=2)
        for i, lado in enumerate(lados):
            ttk.Label(aberturas_frame, text=lado+":").grid(row=i+1, column=0, sticky="w", padx=5)
            for j, lab in enumerate(aberturas_labels):
                ttk.Entry(aberturas_frame, textvariable=self.aberturas_vars[i][j], width=8).grid(row=i+1, column=j+1, padx=2)
        painel2_frame = ttk.LabelFrame(self.dados_frame, text="Painel 2 em L", padding=10)
        painel2_frame.grid(row=5, column=0, columnspan=2, sticky="ew", pady=5)
        colunas_frame = ttk.Frame(painel2_frame)
        colunas_frame.pack(fill=tk.X, pady=5)
        header_frame = ttk.Frame(colunas_frame)
        header_frame.pack(fill=tk.X)
        ttk.Label(header_frame, text="Tipo e Dimensões", font=("Arial", 8, "bold")).pack(side=tk.LEFT, padx=35)
        ttk.Label(header_frame, text="Chanfro", font=("Arial", 8, "bold")).pack(side=tk.LEFT, padx=60)
        aberturas_labels = ["Dist", "Prof", "Larg"]
        aberturas_header = ttk.Frame(header_frame)
        aberturas_header.pack(side=tk.LEFT, padx=30)
        for lab in aberturas_labels:
            ttk.Label(aberturas_header, text=lab, font=("Arial", 8, "bold")).pack(side=tk.LEFT, padx=2)
        coluna1_frame = ttk.Frame(colunas_frame)
        coluna1_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        tipo_frame = ttk.Frame(coluna1_frame)
        tipo_frame.pack(fill=tk.X, pady=5)
        tipos = ["E/T", "E/F", "D/T", "D/F"]
        for i, tipo in enumerate(tipos):
            ttk.Radiobutton(tipo_frame, text=tipo, variable=self.tipo_painel2_var, value=tipo).pack(side=tk.LEFT, padx=5)
        campos = [
            ("Comprimento:", self.comprimento_2_var),
            ("Largura:", self.largura_2_var),
            ("P1:", self.p1_2_var),
            ("P2:", self.p2_2_var),
            ("P3:", self.p3_2_var)
        ]
        for label, var in campos:
            frame = ttk.Frame(coluna1_frame)
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=5)
            ttk.Entry(frame, textvariable=var, width=8).pack(side=tk.LEFT, padx=2)
        coluna2_frame = ttk.Frame(colunas_frame)
        coluna2_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        labels = ["T/E:", "F/E:", "T/D:", "F/D:"]
        for i, label in enumerate(labels):
            frame = ttk.Frame(coluna2_frame)
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=5)
            ttk.Entry(frame, textvariable=self.recuos_2_vars[i], width=8).pack(side=tk.LEFT, padx=2)
        coluna3_frame = ttk.Frame(colunas_frame)
        coluna3_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        lados = ["T/E", "F/E", "T/D", "F/D"]
        for i, lado in enumerate(lados):
            frame = ttk.Frame(coluna3_frame)
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=lado+":").pack(side=tk.LEFT, padx=5)
            for j in range(3):
                ttk.Entry(frame, textvariable=self.aberturas_2_vars[i][j], width=8).pack(side=tk.LEFT, padx=2)
        if not hasattr(self, 'area_util_var'):
            self.area_util_var = tk.StringVar(value="M² = 0.00")
        area_util_label = ttk.Label(info_frame, textvariable=self.area_util_var, font=("Arial", 11, "bold"), foreground="#1976d2")
        area_util_label.grid(row=6, column=0, columnspan=4, sticky="w", padx=5, pady=(2, 8))
        for widget in info_frame.winfo_children():
            if isinstance(widget, ttk.Entry):
                widget.bind('<FocusOut>', self._on_entry_focus_out)
                widget.bind('<FocusIn>', self._on_entry_focus_in)
        for entry in self.paineis_entries:
            entry.bind('<FocusOut>', self._on_entry_focus_out)
            entry.bind('<FocusIn>', self._on_entry_focus_in)
        for child in chanfros_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind('<FocusOut>', self._on_entry_focus_out)
                child.bind('<FocusIn>', self._on_entry_focus_in)
        for child in aberturas_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind('<FocusOut>', self._on_entry_focus_out)
                child.bind('<FocusIn>', self._on_entry_focus_in)
        for child in coluna1_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind('<FocusOut>', self._on_entry_focus_out)
                child.bind('<FocusIn>', self._on_entry_focus_in)
        for child in coluna2_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind('<FocusOut>', self._on_entry_focus_out)
                child.bind('<FocusIn>', self._on_entry_focus_in)
        for child in coluna3_frame.winfo_children():
            if isinstance(child, ttk.Entry):
                child.bind('<FocusOut>', self._on_entry_focus_out)
                child.bind('<FocusIn>', self._on_entry_focus_in)
        def bind_focusout_recursivo(widget):
            for child in widget.winfo_children():
                if isinstance(child, ttk.Entry) or isinstance(child, tk.Entry):
                    child.bind('<FocusOut>', self._on_entry_focus_out)
                    child.bind('<FocusIn>', self._on_entry_focus_in)
                else:
                    bind_focusout_recursivo(child)
        bind_focusout_recursivo(coluna1_frame)
        bind_focusout_recursivo(coluna2_frame)
        bind_focusout_recursivo(coluna3_frame)
    def _selecionar_textos_consecutivos(self):
        self._flutuante_cancelado = False  
        def erro_flutuante(msg):
            self._mostrar_comentario_flutuante(msg, tempo=2)
            def reiniciar():
                if not getattr(self, '_flutuante_cancelado', False):
                    self._selecionar_textos_consecutivos()
                else:
                    print('[DEBUG] Reinício de seleção abortado pois _flutuante_cancelado=True')
                    self._flutuante_after_id = self.after(2100, reiniciar)
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                erro_flutuante("Nenhum documento AutoCAD ativo encontrado!")
                return
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
            except:
                pass
            self.iconify()
            self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.nome_var.set(entity.TextString)
                else:
                    erro_flutuante("O primeiro objeto selecionado não é um texto!")
                    self.deiconify()
                    return
            else:
                erro_flutuante("Nenhum objeto selecionado!")
                self.deiconify()
                return
            self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_esq_var.set(entity.TextString)
                else:
                    erro_flutuante("O segundo objeto selecionado não é um texto!")
                    self.deiconify()
                    return
            else:
                erro_flutuante("Nenhum objeto selecionado!")
                self.deiconify()
                return
            self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_dir_var.set(entity.TextString)
                else:
                    erro_flutuante("O terceiro objeto selecionado não é um texto!")
                    self.deiconify()
                    return
            else:
                erro_flutuante("Nenhum objeto selecionado!")
                self.deiconify()
                return
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            if self.usar_boundary_duplo:
                self.usar_boundary_duplo = False
                self.after(100, self._realizar_boundary_duplo)
            else:
                self.after(100, self._realizar_boundary)
        except Exception as e:
            erro_flutuante(f"Erro ao selecionar textos: {str(e)}")
            self.deiconify()
        finally:
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _salvar_fundo(self):
        if self._carregando_fundo:
            return
        try:
            dados = {
                'numero': self.numero_var.get(),
                'nome': self.nome_var.get(),
                'observacoes': self.obs_var.get(),
                'pavimento': self.pavimento_var.get(),
                'largura': self.largura_var.get(),
                'altura': self.altura_var.get(),
                'texto_esquerda': self.texto_esq_var.get(),
                'texto_direita': self.texto_dir_var.get(),
                'paineis': [v.get() for v in self.paineis_vars],
                'sarrafo_esq': self.sarrafo_esq_var.get(),
                'sarrafo_dir': self.sarrafo_dir_var.get(),
                'recuos': [v.get() for v in self.recuos_vars],
                'aberturas': [[v.get() for v in linha] for linha in self.aberturas_vars],  
                'tipo_distribuicao': self.tipo_distribuicao_var.get(),
                'tipo_painel_inicial': self.tipo_painel_inicial_var.get(),
                'tipo_painel2': self.tipo_painel2_var.get(),
                'comprimento_2': self.comprimento_2_var.get(),
                'largura_2': self.largura_2_var.get(),
                'p1_2': self.p1_2_var.get(),
                'p2_2': self.p2_2_var.get(),
                'p3_2': self.p3_2_var.get(),
                'recuos_2': [v.get() for v in self.recuos_2_vars],
                'aberturas_2': [[v.get() for v in linha] for linha in self.aberturas_2_vars]  
            }
            numero = self.numero_var.get().strip()
            if not numero:
                return
            self.fundos_salvos[numero] = dados.copy()
            self._atualizar_lista()
            self._salvar_fundos_salvos()
        except Exception as e:
            print(f"Erro ao salvar fundo: {str(e)}")
    def _coletar_dados(self):
        return {
            "numero": self.numero_var.get(),
            "nome": self.nome_var.get(),
            "observacoes": self.obs_var.get(),
            "pavimento": self.pavimento_var.get(),
            "largura": self._float_safe(self.largura_var),
            "altura": self._float_safe(self.altura_var),
            "texto_esquerda": self.texto_esq_var.get(),
            "texto_direita": self.texto_dir_var.get(),
            "paineis": [self._float_safe(v) for v in self.paineis_vars],
            "sarrafo_esq": self.sarrafo_esq_var.get(),
            "sarrafo_dir": self.sarrafo_dir_var.get(),
            "recuos": [self._float_safe(v) for v in self.recuos_vars],
            "aberturas": [[self._float_safe(v) for v in linha] for linha in self.aberturas_vars],
            "tipo_painel2": self.tipo_painel2_var.get(),
            "comprimento_2": self._float_safe(self.comprimento_2_var),
            "largura_2": self._float_safe(self.largura_2_var),
            "p1_2": self._float_safe(self.p1_2_var),
            "p2_2": self._float_safe(self.p2_2_var),
            "p3_2": self._float_safe(self.p3_2_var),
            "recuos_2": [self._float_safe(v) for v in self.recuos_2_vars],
            "aberturas_2": [[self._float_safe(v) for v in linha] for linha in self.aberturas_2_vars],
            "tipo_distribuicao": self.tipo_distribuicao_var.get(),
            "tipo_painel_inicial": self.tipo_painel_inicial_var.get(),
            "area_util_m2": self._float_from_m2_string(self.area_util_var.get())
        }
    def _atualizar_lista(self, event=None):
        pavimentos = set(dados.get("pavimento", "") for dados in self.fundos_salvos.values())
        pavimentos = sorted(p for p in pavimentos if p)
        self.combo_pavimento['values'] = ["Todos"] + pavimentos
        pav_filtro = self.pavimento_filtro_var.get()
        if pav_filtro not in pavimentos and pav_filtro != "Todos":
            self.pavimento_filtro_var.set("Todos")
        sel = self.lista_fundos.selection()
        numero_sel = None
        if sel:
            item = self.lista_fundos.item(sel[0])
            values = item.get('values', [])
            if values:
                numero_sel = str(values[0])
        self.lista_fundos.delete(*self.lista_fundos.get_children())
        total_m2 = 0
        novo_iid = None
        for numero, dados in self.fundos_salvos.items():
            if pav_filtro == "Todos" or dados.get("pavimento", "") == pav_filtro:
                m2 = dados.get("area_util_m2", 0)
                total_m2 += m2
                iid = self.lista_fundos.insert('', 'end', values=(numero, dados.get("nome", ""), dados.get("pavimento", ""), f"{m2:.5g}"))
                if numero_sel and str(numero) == numero_sel:
                    novo_iid = iid
        self.total_m2_por_pavimento_var.set(f"Total M²: {total_m2:.2f}")
        if novo_iid:
            self.lista_fundos.selection_set(novo_iid)
    def _carregar_fundo_selecionado(self, event=None):
        print(f"[DEBUG] _carregar_fundo_selecionado chamado. prevent_load={self._prevent_treeview_load_after_edit}, is_editing={self._is_editing_entry}, carregando={self._carregando_fundo}")
        self._is_editing_entry = False 
        if self._block_treeview_events: 
            print("[DEBUG] Carregamento do fundo bloqueado pela flag _block_treeview_events.")
            return
        if self._prevent_treeview_load_after_edit: 
            print("[DEBUG] Carregamento do fundo impedido devido a edição recente.")
            return
        if self._carregando_fundo: 
            return
        try:
            self._carregando_fundo = True 
            selected_item = self.lista_fundos.focus()
            if not selected_item:
                return
            if selected_item not in self.lista_fundos.get_children():
                return
            item = self.lista_fundos.item(selected_item)
            values = item.get('values', [])
            if len(values) < 1:
                return
            numero = str(values[0])
            dados = self.fundos_salvos.get(numero, {})
            self._ignorar_traces = True
            self.numero_var.set(dados.get("numero", numero))
            self.nome_var.set(dados.get("nome", ""))
            self.obs_var.set(dados.get("observacoes", ""))
            self.pavimento_var.set(dados.get("pavimento", ""))
            try:
                self.largura_var.set(float(dados.get("largura", 0)))
            except Exception:
                self.largura_var.set(0)
            try:
                self.altura_var.set(float(dados.get("altura", 0)))
            except Exception:
                self.altura_var.set(0)
            self.texto_esq_var.set(dados.get("texto_esquerda", ""))
            self.texto_dir_var.set(dados.get("texto_direita", ""))
            paineis = dados.get("paineis", [])
            for i in range(6):
                try:
                    self.paineis_vars[i].set(float(paineis[i]))
                except Exception:
                    self.paineis_vars[i].set(0)
            self.sarrafo_esq_var.set(dados.get("sarrafo_esq", True))
            self.sarrafo_dir_var.set(dados.get("sarrafo_dir", True))
            recuos = dados.get("recuos", [])
            for i in range(4):
                try:
                    self.recuos_vars[i].set(float(recuos[i]))
                except Exception:
                    self.recuos_vars[i].set(0)
            aberturas = dados.get("aberturas", [])
            for i in range(4):
                for j in range(3):
                    try:
                        self.aberturas_vars[i][j].set(float(aberturas[i][j]))
                    except Exception:
                        self.aberturas_vars[i][j].set(0)
            self.tipo_painel2_var.set(dados.get('tipo_painel2', 'E/T'))
            self.comprimento_2_var.set(dados.get('comprimento_2', 0))
            self.largura_2_var.set(dados.get('largura_2', 0))
            self.p1_2_var.set(dados.get('p1_2', 0))
            self.p2_2_var.set(dados.get('p2_2', 0))
            self.p3_2_var.set(dados.get('p3_2', 0))
            recuos_2 = dados.get('recuos_2', [0]*4)
            for i in range(4):
                try:
                    self.recuos_2_vars[i].set(float(recuos_2[i]))
                except Exception:
                    self.recuos_2_vars[i].set(0)
            aberturas_2 = dados.get('aberturas_2', [[0,0,0] for _ in range(4)])
            for i in range(4):
                for j in range(3):
                    try:
                        self.aberturas_2_vars[i][j].set(float(aberturas_2[i][j]))
                    except Exception:
                        self.aberturas_2_vars[i][j].set(0)
            self.tipo_distribuicao_var.set(dados.get('tipo_distribuicao', '122'))
            self.tipo_painel_inicial_var.set(dados.get('tipo_painel_inicial', '300'))
            self._carregando_fundo = False
            self._ignorar_traces = False
            self._atualizar_preview()
            self._atualizando_area_util = True
            self._calcular_area_util()
            self._atualizando_area_util = False
        except Exception as e:
            print(f"Erro ao carregar fundo: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar fundo: {str(e)}")
        finally: 
            self._carregando_fundo = False 
    def _excluir_fundo(self):
        sel = self.lista_fundos.selection()
        if not sel:
            messagebox.showinfo("Aviso", "Selecione um item para excluir")
            return
        numero = str(self.lista_fundos.item(sel[0])['values'][0]).strip()
        print(f"[DEBUG] Chaves antes da exclusão: {list(self.fundos_salvos.keys())}")
        if messagebox.askyesno("Excluir", f"Excluir fundo '{numero}'?"):
            self.lista_fundos.delete(sel[0])
            if numero in self.fundos_salvos:
                del self.fundos_salvos[numero]
            print(f"[DEBUG] Chaves após exclusão: {list(self.fundos_salvos.keys())}")
            print(f"[DEBUG] Estado do fundos_salvos após exclusão: {self.fundos_salvos}")
            self._salvar_fundos_salvos()  
            self._limpar_campos()  
            self._atualizar_lista()
    def _excluir_selecionados(self):
        selecoes = self.lista_fundos.selection()
        if not selecoes:
            messagebox.showinfo("Aviso", "Selecione pelo menos um item para excluir")
            return
        numeros = [str(self.lista_fundos.item(item)['values'][0]).strip() for item in selecoes]
        if messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir os {len(numeros)} itens selecionados?"):
            for item, numero in zip(selecoes, numeros):
                self.lista_fundos.delete(item)
                if numero in self.fundos_salvos:
                    del self.fundos_salvos[numero]
            self._salvar_fundos_salvos()
            self._limpar_campos()
            self._atualizar_lista()
            messagebox.showinfo("Sucesso", f"{len(numeros)} itens foram excluídos com sucesso")
    def _mostrar_menu_contexto(self, event):
        item = self.lista_fundos.identify_row(event.y)
        if item:
            self.lista_fundos.selection_set(item)
            self.menu_contexto.post(event.x_root, event.y_root)
    def _limpar_campos(self):
        self.numero_var.set("")
        self.nome_var.set("")
        self.obs_var.set("")
        self.largura_var.set(0)
        self.altura_var.set(0)
        self.texto_esq_var.set("")
        self.texto_dir_var.set("")
        for v in self.paineis_vars:
            v.set(0)
        self.sarrafo_esq_var.set(True)
        self.sarrafo_dir_var.set(True)
        for v in self.recuos_vars:
            v.set(0)
        for linha in self.aberturas_vars:
            for v in linha:
                v.set(0)
        self.tipo_painel2_var.set("E/T")
        self.comprimento_2_var.set(0)
        self.largura_2_var.set(0)
        self.p1_2_var.set(0)
        self.p2_2_var.set(0)
        self.p3_2_var.set(0)
        for v in self.recuos_2_vars:
            v.set(0)
        for linha in self.aberturas_2_vars:
            for v in linha:
                v.set(0)
        self._atualizar_preview()
    def _salvar_fundos_salvos(self):
        try:
            print("Salvando no pickle...")  
            print(f"Dados a serem salvos: {self.fundos_salvos}")  
            pickle_path = os.path.join(self.script_dir, "fundos_salvos.pkl")
            with open(pickle_path, "wb") as f:
                pickle.dump(self.fundos_salvos, f)
            print("Pickle salvo com sucesso!")  
        except Exception as e:
            print(f"Erro ao salvar pickle: {str(e)}")  
    def _carregar_fundos_salvos(self):
        try:
            pickle_path = os.path.join(self.script_dir, "fundos_salvos.pkl")
            if os.path.exists(pickle_path):
                print("Carregando arquivo pickle...")  
                with open(pickle_path, "rb") as f:
                    self.fundos_salvos = pickle.load(f)
                print(f"Dados carregados: {self.fundos_salvos}")  
                self._atualizar_lista()
            else:
                print("Arquivo pickle não encontrado")  
        except Exception as e:
            print(f"Erro ao carregar pickle: {str(e)}")  
            self.fundos_salvos = {}
    def _selecionar_excel(self):
        arquivo = filedialog.askopenfilename(title="Selecionar Excel", filetypes=[("Excel", "*.xlsx *.xls")])
        if arquivo:
            self.excel_file = arquivo
            messagebox.showinfo("Sucesso", f"Arquivo selecionado: {os.path.basename(arquivo)}")
    def _criar_excel(self):
        arquivo = filedialog.asksaveasfilename(title="Salvar Novo Excel", filetypes=[("Excel", "*.xlsx")], defaultextension=".xlsx")
        if not arquivo:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        mapeamento = {
            'numero': 52,
            'nome': 53,
            'observacoes': 54,
            'pavimento': 55,
            'largura': 56,
            'altura': 57,
            'texto_esquerda': 58,
            'texto_direita': 59,
            'tipo_distribuicao': 60,
            'tipo_painel_inicial': 61,
            'sarrafo_esq': 62,
            'sarrafo_dir': 63,
            'painel1': 64,
            'painel2': 65,
            'painel3': 66,
            'painel4': 67,
            'painel5': 68,
            'painel6': 69,
            'chanfro_te': 70,
            'chanfro_fe': 71,
            'chanfro_td': 72,
            'chanfro_fd': 73,
            'abertura_te_dist': 74,
            'abertura_te_prof': 75,
            'abertura_te_larg': 76,
            'abertura_fe_dist': 77,
            'abertura_fe_prof': 78,
            'abertura_fe_larg': 79,
            'abertura_td_dist': 80,
            'abertura_td_prof': 81,
            'abertura_td_larg': 82,
            'abertura_fd_dist': 83,
            'abertura_fd_prof': 84,
            'abertura_fd_larg': 85,
            'tipo_painel2': 86,
            'comprimento_2': 87,
            'largura_2': 88,
            'p1_2': 89,
            'p2_2': 90,
            'p3_2': 91,
            'chanfro2_te': 92,
            'chanfro2_fe': 93,
            'chanfro2_td': 94,
            'chanfro2_fd': 95,
            'abertura2_te_dist': 96,
            'abertura2_te_prof': 97,
            'abertura2_te_larg': 98,
            'abertura2_fe_dist': 99,
            'abertura2_fe_prof': 100,
            'abertura2_fe_larg': 101,
            'abertura2_td_dist': 102,
            'abertura2_td_prof': 103,
            'abertura2_td_larg': 104,
            'abertura2_fd_dist': 105,
            'abertura2_fd_prof': 106,
            'abertura2_fd_larg': 107
        }
        coluna = 5
        for numero, dados in self.fundos_salvos.items():
            for campo, linha in mapeamento.items():
                if campo == 'painel1':
                    valor = dados.get('paineis', [0]*6)[0]
                elif campo == 'painel2':
                    valor = dados.get('paineis', [0]*6)[1]
                elif campo == 'painel3':
                    valor = dados.get('paineis', [0]*6)[2]
                elif campo == 'painel4':
                    valor = dados.get('paineis', [0]*6)[3]
                elif campo == 'painel5':
                    valor = dados.get('paineis', [0]*6)[4]
                elif campo == 'painel6':
                    valor = dados.get('paineis', [0]*6)[5]
                elif campo == 'chanfro_te':
                    valor = dados.get('recuos', [0]*4)[0]
                elif campo == 'chanfro_fe':
                    valor = dados.get('recuos', [0]*4)[1]
                elif campo == 'chanfro_td':
                    valor = dados.get('recuos', [0]*4)[2]
                elif campo == 'chanfro_fd':
                    valor = dados.get('recuos', [0]*4)[3]
                elif campo.startswith('abertura_'):
                    idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.split('_')[1]]
                    if campo.endswith('_dist'):
                        valor = 0.0
                    elif campo.endswith('_prof'):
                        valor = dados.get('aberturas', [[0.0, 0.0]] * 4)[idx_lado][0]
                    elif campo.endswith('_larg'):
                        valor = dados.get('aberturas', [[0.0, 0.0]] * 4)[idx_lado][1]
                elif campo == 'chanfro2_te':
                    valor = dados.get('recuos_2', [0]*4)[0]
                elif campo == 'chanfro2_fe':
                    valor = dados.get('recuos_2', [0]*4)[1]
                elif campo == 'chanfro2_td':
                    valor = dados.get('recuos_2', [0]*4)[2]
                elif campo == 'chanfro2_fd':
                    valor = dados.get('recuos_2', [0]*4)[3]
                elif campo.startswith('abertura2_'):
                    idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.split('_')[1]]
                    if campo.endswith('_dist'):
                        valor = 0.0
                    elif campo.endswith('_prof'):
                        valor = dados.get('aberturas_2', [[0.0, 0.0]] * 4)[idx_lado][0]
                    elif campo.endswith('_larg'):
                        valor = dados.get('aberturas_2', [[0.0, 0.0]] * 4)[idx_lado][1]
                else:
                    valor = dados.get(campo, '')
                ws.cell(row=linha, column=coluna, value=valor)
            coluna += 1
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Excel criado em {arquivo}")
    def _importar_excel(self):
        arquivo = filedialog.askopenfilename(title="Importar Excel", filetypes=[("Excel", "*.xlsx *.xls")])
        if not arquivo:
            return
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        ws = wb.active
        mapeamento = {
            'numero': 52,
            'nome': 53,
            'observacoes': 54,
            'pavimento': 55,
            'largura': 56,
            'altura': 57,
            'texto_esquerda': 58,
            'texto_direita': 59,
            'tipo_distribuicao': 60,
            'tipo_painel_inicial': 61,
            'sarrafo_esq': 62,
            'sarrafo_dir': 63,
            'painel1': 64,
            'painel2': 65,
            'painel3': 66,
            'painel4': 67,
            'painel5': 68,
            'painel6': 69,
            'chanfro_te': 70,
            'chanfro_fe': 71,
            'chanfro_td': 72,
            'chanfro_fd': 73,
            'abertura_te_dist': 74,
            'abertura_te_prof': 75,
            'abertura_te_larg': 76,
            'abertura_fe_dist': 77,
            'abertura_fe_prof': 78,
            'abertura_fe_larg': 79,
            'abertura_td_dist': 80,
            'abertura_td_prof': 81,
            'abertura_td_larg': 82,
            'abertura_fd_dist': 83,
            'abertura_fd_prof': 84,
            'abertura_fd_larg': 85,
            'tipo_painel2': 86,
            'comprimento_2': 87,
            'largura_2': 88,
            'p1_2': 89,
            'p2_2': 90,
            'p3_2': 91,
            'chanfro2_te': 92,
            'chanfro2_fe': 93,
            'chanfro2_td': 94,
            'chanfro2_fd': 95,
            'abertura2_te_dist': 96,
            'abertura2_te_prof': 97,
            'abertura2_te_larg': 98,
            'abertura2_fe_dist': 99,
            'abertura2_fe_prof': 100,
            'abertura2_fe_larg': 101,
            'abertura2_td_dist': 102,
            'abertura2_td_prof': 103,
            'abertura2_td_larg': 104,
            'abertura2_fd_dist': 105,
            'abertura2_fd_prof': 106,
            'abertura2_fd_larg': 107
        }
        coluna = 5
        while True:
            numero = ws.cell(row=mapeamento['numero'], column=coluna).value
            if not numero:
                break
            dados = {}
            for campo, linha in mapeamento.items():
                valor = ws.cell(row=linha, column=coluna).value
                dados[campo] = valor
            fundo = {
                'numero': str(dados.get('numero', '')).strip(),
                'nome': dados.get('nome', ''),
                'observacoes': dados.get('observacoes', ''),
                'pavimento': dados.get('pavimento', ''),
                'largura': float(dados.get('largura', 0)),
                'altura': float(dados.get('altura', 0)),
                'texto_esquerda': dados.get('texto_esquerda', ''),
                'texto_direita': dados.get('texto_direita', ''),
                'tipo_distribuicao': dados.get('tipo_distribuicao', '122'),
                'tipo_painel_inicial': dados.get('tipo_painel_inicial', '300'),
                'sarrafo_esq': bool(dados.get('sarrafo_esq', True)),
                'sarrafo_dir': bool(dados.get('sarrafo_dir', True)),
                'paineis': [float(dados.get(f'painel{i+1}', 0)) for i in range(6)],
                'recuos': [float(dados.get(f'chanfro_{k}', 0)) for k in ['te','fe','td','fd']],
                'aberturas': [
                    [0.0, float(dados.get(f'abertura_{lado}_prof', 0)), float(dados.get(f'abertura_{lado}_larg', 0))] # Distância agora é sempre 0.0
                    for lado in ['te','fe','td','fd']
                ],
                'tipo_painel2': dados.get('tipo_painel2', 'E/T'),
                'comprimento_2': float(dados.get('comprimento_2', 0)),
                'largura_2': float(dados.get('largura_2', 0)),
                'p1_2': float(dados.get('p1_2', 0)),
                'p2_2': float(dados.get('p2_2', 0)),
                'p3_2': float(dados.get('p3_2', 0)),
                'recuos_2': [float(dados.get(f'chanfro2_{k}', 0)) for k in ['te','fe','td','fd']],
                'aberturas_2': [
                    [0.0, float(dados.get(f'abertura2_{lado}_prof', 0)), float(dados.get(f'abertura2_{lado}_larg', 0))] # Distância agora é sempre 0.0
                    for lado in ['te','fe','td','fd']
                ]
            }
            self.fundos_salvos[fundo['numero']] = fundo
            coluna += 1
        self._atualizar_lista()
        self._salvar_fundos_salvos()
        messagebox.showinfo("Sucesso", "Fundos importados do Excel!")
    def _exportar_excel(self):
        if not self.fundos_salvos:
            messagebox.showwarning("Aviso", "Nenhum fundo salvo para exportar!")
            return
        arquivo = filedialog.asksaveasfilename(title="Exportar Excel", filetypes=[("Excel", "*.xlsx")], defaultextension=".xlsx")
        if not arquivo:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        mapeamento = {
            'nome': 4,
            'observacoes': 5,
            'pavimento': 2,
            'largura': 7,
            'altura': 8,
            'texto_esquerda': 10,
            'texto_direita': 11,
            'linha1': 13,
            'linha2': 14,
            'linha3': 15,
            'linha4': 16,
            'linha5': 17,
            'linha6': 18,
            'recuo_topo_esq': 33,
            'recuo_fundo_esq': 34,
            'recuo_topo_dir': 35,
            'recuo_fundo_dir': 36,
            'abertura_esq_topo_dist': 20,
            'abertura_esq_topo_prof': 21,
            'abertura_esq_topo_larg': 22,
            'abertura_esq_fundo_dist': 23,
            'abertura_esq_fundo_prof': 24,
            'abertura_esq_fundo_larg': 25,
            'abertura_dir_topo_dist': 26,
            'abertura_dir_topo_prof': 27,
            'abertura_dir_topo_larg': 28,
            'abertura_dir_fundo_dist': 29,
            'abertura_dir_fundo_prof': 30,
            'abertura_dir_fundo_larg': 31
        }
        coluna = 5
        for numero, dados in self.fundos_salvos.items():
            for campo, linha in mapeamento.items():
                if campo == 'nome':
                    valor = numero  
                elif campo == 'linha1':
                    valor = dados.get('paineis', [0]*6)[0]
                elif campo == 'linha2':
                    valor = dados.get('paineis', [0]*6)[1]
                elif campo == 'linha3':
                    valor = dados.get('paineis', [0]*6)[2]
                elif campo == 'linha4':
                    valor = dados.get('paineis', [0]*6)[3]
                elif campo == 'linha5':
                    valor = dados.get('paineis', [0]*6)[4]
                elif campo == 'linha6':
                    valor = dados.get('paineis', [0]*6)[5]
                elif campo == 'recuo_topo_esq':
                    valor = dados.get('recuos', [0]*4)[0]
                elif campo == 'recuo_fundo_esq':
                    valor = dados.get('recuos', [0]*4)[1]
                elif campo == 'recuo_topo_dir':
                    valor = dados.get('recuos', [0]*4)[2]
                elif campo == 'recuo_fundo_dir':
                    valor = dados.get('recuos', [0]*4)[3]
                elif campo == 'abertura_esq_topo_dist':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[0][0]
                elif campo == 'abertura_esq_topo_prof':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[0][1]
                elif campo == 'abertura_esq_topo_larg':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[0][2]
                elif campo == 'abertura_esq_fundo_dist':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[1][0]
                elif campo == 'abertura_esq_fundo_prof':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[1][1]
                elif campo == 'abertura_esq_fundo_larg':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[1][2]
                elif campo == 'abertura_dir_topo_dist':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[2][0]
                elif campo == 'abertura_dir_topo_prof':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[2][1]
                elif campo == 'abertura_dir_topo_larg':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[2][2]
                elif campo == 'abertura_dir_fundo_dist':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[3][0]
                elif campo == 'abertura_dir_fundo_prof':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[3][1]
                elif campo == 'abertura_dir_fundo_larg':
                    valor = dados.get('aberturas', [[0,0,0]]*4)[3][2]
                else:
                    valor = dados.get(campo, '')
                ws.cell(row=linha, column=coluna, value=valor)
            coluna += 1
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Excel exportado para {arquivo}")
    def _on_scroll(self, event):
        pass
    def _on_press(self, event):
        pass
    def _on_release(self, event):
        pass
    def _on_motion(self, event):
        pass
    def _calcular_paineis_automatico(self, event=None):
        try:
            largura = float(self.largura_var.get())
        except (tk.TclError, ValueError):
            for v in self.paineis_vars:
                v.set(0)
            return
        try:
            for v in self.paineis_vars:
                v.set(0)
            tipo_distribuicao = self.tipo_distribuicao_var.get()
            base_divisao = int(self.tipo_divisao_paineis_var.get())
            if base_divisao == 307:
                paineis = []
                largura_restante = largura
                if largura_restante > 307:
                    paineis.append(307)
                    largura_restante -= 307
                else:
                    paineis.append(largura_restante)
                    largura_restante = 0
                while largura_restante > 0:
                    if largura_restante > 300:
                        paineis.append(300)
                        largura_restante -= 300
                    else:
                        paineis.append(largura_restante)
                        largura_restante = 0
                if len(paineis) > 1:
                    ultimo = paineis[-1]
                    penultimo = paineis[-2]
                    if ultimo < 67:
                        paineis[-2] = 122
                        paineis[-1] = penultimo + ultimo - 122
                    elif ultimo >= 67 and ultimo < 300:
                        diff = 307 - ultimo
                        paineis[-1] = 307
                        paineis[-2] = penultimo - diff
                for i, valor in enumerate(paineis[:6]):
                    self.paineis_vars[i].set(valor)
                return
            tamanho_inicial = float(self.tipo_painel_inicial_var.get())
            if largura <= tamanho_inicial:
                self.paineis_vars[0].set(largura)
                return
            if tamanho_inicial in (300, 307) and largura > tamanho_inicial:
                painel_inicial = 244
            else:
                painel_inicial = tamanho_inicial
            num_paineis = int((largura + base_divisao - 1) // base_divisao)
            largura_restante = largura
            paineis = []
            for i in range(num_paineis):
                if i == 0:
                    if largura_restante >= painel_inicial:
                        paineis.append(painel_inicial)
                        largura_restante -= painel_inicial
                    else:
                        paineis.append(largura_restante)
                        largura_restante = 0
                        break
                elif i == num_paineis - 1:
                    if tipo_distribuicao == "122":
                        if largura_restante < 60 and len(paineis) > 0:
                            paineis[-1] = 122
                            paineis.append(largura_restante + 122)
                        else:
                            paineis.append(largura_restante)
                    elif tipo_distribuicao == "307":
                        if largura_restante < 63 and len(paineis) > 0:
                            soma = paineis[-1] + largura_restante
                            if soma > 307:
                                paineis.append(largura_restante)
                            else:
                                paineis[-1] = soma
                        else:
                            paineis.append(largura_restante)
                else:
                    paineis.append(base_divisao)
                    largura_restante -= base_divisao
            for i, valor in enumerate(paineis[:6]):
                self.paineis_vars[i].set(valor)
        except Exception:
            pass
    def _analisar_coordenadas(self, coords):
        """Analisa as coordenadas do boundary para identificar chanfros e aberturas."""
        print("\nIniciando análise de coordenadas...")
        min_x = min(coords[::2])
        max_x = max(coords[::2])
        min_y = min(coords[1::2])
        max_y = max(coords[1::2])
        print(f"Limites do boundary:")
        print(f"X: {min_x:.2f} a {max_x:.2f}")
        print(f"Y: {min_y:.2f} a {max_y:.2f}")
        largura_x = max_x - min_x
        altura_y = max_y - min_y
        print(f"Dimensões iniciais: {largura_x:.2f} x {altura_y:.2f}")
        if altura_y > largura_x:
            print("Trocando dimensões para garantir que largura seja a maior...")
            largura_total = altura_y
            altura_total = largura_x
            coords = [y if i % 2 == 0 else x for i, (x, y) in enumerate(zip(coords[::2], coords[1::2]))]
            min_x = min(coords[::2])
            max_x = max(coords[::2])
            min_y = min(coords[1::2])
            max_y = max(coords[1::2])
            print(f"Novas dimensões: {largura_total:.2f} x {altura_total:.2f}")
        else:
            largura_total = largura_x
            altura_total = altura_y
        recommendations = self.learner.get_recommendations(coords)
        chanfros = {
            'topo_esq': 0,
            'topo_dir': 0,
            'fundo_esq': 0,
            'fundo_dir': 0
        }
        for nome, valor in recommendations['chanfros'].items():
            chanfros[nome] = valor
        aberturas = []
        for abertura in recommendations['aberturas']:
            x = abertura['x']
            y = abertura['y']
            largura_ab = abertura['largura']
            altura_ab = abertura['altura']
            limiar = max(1.0, altura_total * 0.05)
            if abs(x) < 1e-2:
                lado = 'esquerda'
                pos_x = 'parede'
            elif abs((x + largura_ab) - largura_total) < 1e-2:
                lado = 'direita'
                pos_x = 'parede'
            elif x < largura_total/2:
                lado = 'esquerda'
                pos_x = 'central'
            else:
                lado = 'direita'
                pos_x = 'central'
            if (y + altura_ab) >= (altura_total - limiar):
                topo_fundo = 'topo'
                pos_y = 'parede'
                profundidade = altura_ab
            elif y <= limiar:
                topo_fundo = 'fundo'
                pos_y = 'parede'
                profundidade = altura_ab
            elif y < (altura_total - (y + altura_ab)):
                topo_fundo = 'fundo'
                pos_y = 'central'
                profundidade = min(y, altura_total - (y + altura_ab))
            else:
                topo_fundo = 'topo'
                pos_y = 'central'
                profundidade = min(y, altura_total - (y + altura_ab))
            aberturas.append({
                'x': x,
                'y': y,
                'largura': largura_ab,
                'altura': altura_ab,
                'lado': lado,
                'topo_fundo': topo_fundo,
                'pos_x': pos_x,
                'pos_y': pos_y,
                'profundidade': profundidade
            })
        print("\nResultado da análise:")
        print(f"Dimensões: {largura_total:.2f} x {altura_total:.2f}")
        print("\nChanfros:")
        for nome, valor in chanfros.items():
            print(f"{nome}: {valor:.2f}")
        print("\nAberturas identificadas:")
        if not aberturas:
            print("Nenhuma abertura identificada")
        else:
            for i, abertura in enumerate(aberturas):
                print(f"\nAbertura {i+1}:")
                print(f"  Posição: ({abertura['x']:.2f}, {abertura['y']:.2f})")
                print(f"  Largura: {abertura['largura']:.2f}")
                print(f"  Profundidade: {abertura['profundidade']:.2f}")
                print(f"  Classificação: {abertura['lado']} {abertura['topo_fundo']} ({abertura['pos_x']}, {abertura['pos_y']})")
        msg = f"Dimensões: {largura_total:.2f} x {altura_total:.2f}\n\nChanfros:\n"
        for nome, valor in chanfros.items():
            msg += f"{nome}: {valor:.2f}\n"
        msg += "\nAberturas:\n"
        if not aberturas:
            msg += "Nenhuma abertura identificada\n"
        else:
            for i, abertura in enumerate(aberturas):
                msg += (f"Abertura {i+1}:\n  Posição: ({abertura['x']:.2f}, {abertura['y']:.2f})\n"
                        f"  Largura: {abertura['largura']:.2f}\n  Profundidade: {abertura['profundidade']:.2f}\n"
                        f"  Classificação: {abertura['lado']} {abertura['topo_fundo']} ({abertura['pos_x']}, {abertura['pos_y']})\n")
        try:
            msg += "\nPreview:\n"
            ascii_preview = self._desenhar_ascii_retangulo(largura_total, altura_total, aberturas, resol_x=30, resol_y=8)
            msg += ascii_preview
        except Exception as e:
            msg += f"\n[Erro ao gerar preview ASCII: {e}]\n"
        return {
            'dimensoes': {
                'largura': largura_total,
                'altura': altura_total
            },
            'chanfros': chanfros,
            'aberturas': aberturas
        }, msg
    def _desenhar_ascii_retangulo(self, largura, altura, aberturas, resol_x=30, resol_y=8):
        """Desenha um retângulo em ASCII com aberturas."""
        escala_x = resol_x / largura
        escala_y = resol_y / altura
        matriz = [[' ' for _ in range(resol_x)] for _ in range(resol_y)]
        for i in range(resol_x):
            matriz[0][i] = '-'
            matriz[-1][i] = '-'
        for i in range(resol_y):
            matriz[i][0] = '|'
            matriz[i][-1] = '|'
        for abertura in aberturas:
            x = int(abertura['x'] * escala_x)
            y = int(abertura['y'] * escala_y)
            larg = int(abertura['largura'] * escala_x)
            prof = int(abertura['profundidade'] * escala_y)
            for i in range(max(0, y), min(resol_y, y + prof)):
                for j in range(max(0, x), min(resol_x, x + larg)):
                    matriz[i][j] = ' '
        return '\n'.join([''.join(linha) for linha in matriz])
    def _preencher_campos(self, analise):
        if analise is None:
            return
        try:
            self._ignorar_traces = True
            self.largura_var.set(analise['dimensoes']['largura'])
            self.altura_var.set(analise['dimensoes']['altura'])
            self.recuos_vars[0].set(analise['chanfros']['topo_esq'])
            self.recuos_vars[1].set(analise['chanfros']['fundo_esq'])
            self.recuos_vars[2].set(analise['chanfros']['topo_dir'])
            self.recuos_vars[3].set(analise['chanfros']['fundo_dir'])
            for i in range(4):
                for j in range(3):
                    self.aberturas_vars[i][j].set(0)
            for i, abertura in enumerate(analise['aberturas']):
                if i < 4:  
                    self.aberturas_vars[i][0].set(abertura['profundidade'])  
                    self.aberturas_vars[i][1].set(abertura['largura'])     
                    self.aberturas_vars[i][2].set(abertura['altura'])
            for v in self.paineis_vars:
                v.set(0)
            self._calcular_paineis_automatico() 
            self.tipo_painel2_var.set("E/T")
            self.comprimento_2_var.set(0)
            self.largura_2_var.set(0)
            self.p1_2_var.set(0)
            self.p2_2_var.set(0)
            self.p3_2_var.set(0)
            for v in self.recuos_2_vars:
                v.set(0)
            for linha in self.aberturas_2_vars:
                for v in linha:
                    v.set(0)
            self._ignorar_traces = False
            self._atualizar_preview()
            self._calcular_area_util()
        except Exception as e:
            import tkinter.messagebox as messagebox
            messagebox.showerror("Erro", f"Erro ao preencher campos: {e}")
    def _realizar_boundary(self):
        try:
            print("[DEBUG] Iniciando _realizar_boundary")
            pythoncom.CoInitialize()
            if self.ac is None:
                print("[DEBUG] Criando conexão com AutoCAD")
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Nenhum documento AutoCAD ativo encontrado!", tempo=2)
                return
            print("[DEBUG] Documento AutoCAD encontrado")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.ShowWindow(acad_window, 3)
                    win32gui.SetForegroundWindow(acad_window)
                    self.update()
                    self.after(100)
            except Exception as e:
                print(f"[DEBUG] Erro ao trazer janela para frente: {str(e)}")
                pass
            self.iconify()
            self._mostrar_comentario_flutuante("Selecione a área para criar o hatch e pressione Enter")
            self.update()
            try:
                if hasattr(doc, 'SetVariable'):
                    print("[DEBUG] Configurando OSMODE")
                    doc.SetVariable("OSMODE", 35)
            except Exception as e:
                print(f"[DEBUG] Erro ao configurar OSMODE: {str(e)}")
                pass
            while True:
                print("[DEBUG] Enviando comandos para o AutoCAD (BOUNDARY)")
                try:
                    count_antes = doc.ModelSpace.Count
                except Exception as e:
                    print(f"[DEBUG] Erro ao acessar ModelSpace antes: {str(e)}")
                    count_antes = None
                try:
                    doc.SendCommand("._BOUNDARY\n")
                    print("[DEBUG] Comando BOUNDARY enviado")
                except Exception as e:
                    print(f"[DEBUG] Erro ao enviar comando BOUNDARY: {str(e)}")
                    raise
                print("[DEBUG] Aguardando seleção do usuário")
                time.sleep(1)
                self._fechar_comentario_flutuante()
                try:
                    count_depois = doc.ModelSpace.Count
                except Exception as e:
                    print(f"[DEBUG] Erro ao acessar ModelSpace depois: {str(e)}")
                    count_depois = None
                boundary_criado = False
                if count_antes is not None and count_depois is not None and count_depois > count_antes:
                    boundary_criado = True
                if boundary_criado:
                    print("[DEBUG] Boundary criado com sucesso!")
                    break
                print("[DEBUG] Nenhum boundary criado. Emitindo comando EXTEND para fechar retângulo. Pressione ENTER para prosseguir.")
                self._mostrar_comentario_flutuante(
                    "Boundary não selecionado, emitido comando EXTEND para fechar retângulo. Pressione ENTER para prosseguir.", tempo=2)
                try:
                    doc.SendCommand("EX\n")
                except Exception as e:
                    print(f"[DEBUG] Erro ao enviar comando EX: {str(e)}")
                time.sleep(2)  
                self._fechar_comentario_flutuante()
                try:
                    temp_sel = doc.SelectionSets.Item("TempSelectionExtend")
                    temp_sel.Delete()
                except Exception:
                    pass
                selection = doc.SelectionSets.Add("TempSelectionExtend")
                try:
                    selection.Clear()
                except:
                    pass
                try:
                    selection.Clear()
                    selection.Delete()
                except:
                    pass
            try:
                last_entity = doc.ModelSpace.Item(count_depois - 1)
                if hasattr(last_entity, 'Coordinates'):
                    print("[DEBUG] Obtendo coordenadas do boundary")
                    coords = last_entity.Coordinates
                    num_points = len(coords) // 2
                    print(f"[DEBUG] Número de pontos obtidos: {num_points}")
                    print("\nCoordenadas do Boundary (original):")
                    print("=" * 50)
                    for i in range(0, len(coords), 2):
                        x, y = coords[i], coords[i+1]
                        print(f"Ponto {i//2 + 1}: ({x:.2f}, {y:.2f})")
                    print("[DEBUG] Alinhando coordenadas")
                    coords_alinhadas = alinhar_boundary_horizontal(coords)
                    print("\n[INFO] Boundary foi alinhado horizontalmente para análise!")
                    print("Coordenadas após alinhamento:")
                    for i in range(0, len(coords_alinhadas), 2):
                        x, y = coords_alinhadas[i], coords_alinhadas[i+1]
                        print(f"Ponto {i//2 + 1}: ({x:.2f}, {y:.2f})")
                    print("[DEBUG] Analisando coordenadas")
                    analise, msg_confirmacao = self._analisar_coordenadas(coords_alinhadas)
                    if analise is None:
                        return
                    print("[DEBUG] Preenchendo campos imediatamente")
                    self._preencher_campos(analise)
                    def on_confirmar_boundary(resposta):
                        if resposta:
                            self.learner.add_example(coords_alinhadas, analise['chanfros'], analise['aberturas'])
                            print("\nExemplo salvo para aprendizado!")
                        else:
                            print("\nExemplo não foi salvo.")
                    self.deiconify() 
                    self._janela_confirmacao_salvamento(msg_confirmacao, analise['dimensoes']['largura'],
                                                        analise['dimensoes']['altura'], analise['aberturas'],
                                                        on_confirmar_boundary)
                else:
                    print("[DEBUG] Objeto não possui coordenadas")
                    self._mostrar_comentario_flutuante("Não foi possível obter as coordenadas do boundary!", tempo=2)
            except Exception as e:
                print(f"[DEBUG] Erro ao processar o boundary: {str(e)}")
                self._mostrar_comentario_flutuante(f"Erro ao processar o boundary: {str(e)}", tempo=2)
            self.deiconify()
        except Exception as e:
            print(f"[DEBUG] Erro geral: {str(e)}")
            self._mostrar_comentario_flutuante(f"Erro ao criar boundary: {str(e)}", tempo=2)
            self.deiconify()
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            self.deiconify()
    def _janela_confirmacao_salvamento(self, msg, largura, altura, aberturas, confirmar_callback):
        win = tk.Toplevel(self)
        win.title("Confirmar Salvamento")
        win.grab_set()
        win.geometry("500x600")  
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame_texto = ttk.Frame(win)
        frame_texto.pack(fill="both", expand=True, padx=5, pady=5)
        preview_frame = ttk.LabelFrame(frame_texto, text="Visualização", padding=5)
        preview_frame.pack(fill=tk.X, expand=True, pady=(0, 8))
        canvas_preview = tk.Canvas(preview_frame, width=450, height=120, bg=self.cores['fundo'], scrollregion=(0, 0, 900, 240))  
        canvas_preview.pack(fill=tk.BOTH, expand=True)
        zoom_level = 1.0
        pan_start = [None]
        def start_pan(event):
            canvas_preview.scan_mark(event.x, event.y)
            pan_start[0] = (event.x, event.y)
        def pan(event):
            if pan_start[0]:
                canvas_preview.scan_dragto(event.x, event.y, gain=1)
                pan_start[0] = (event.x, event.y)
        def stop_pan(event):
            pan_start[0] = None
        def mouse_wheel(event):
            nonlocal zoom_level
            if hasattr(event, 'num') and event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
                zoom_level = max(zoom_level - 0.1, 0.5)
            else:
                zoom_level = min(zoom_level + 0.1, 2.0)
            atualizar_preview_local()
        canvas_preview.bind("<ButtonPress-1>", start_pan)
        canvas_preview.bind("<B1-Motion>", pan)
        canvas_preview.bind("<ButtonRelease-1>", stop_pan)
        canvas_preview.bind("<MouseWheel>", mouse_wheel)
        canvas_preview.bind("<Button-4>", mouse_wheel)
        canvas_preview.bind("<Button-5>", mouse_wheel)
        def atualizar_preview_local():
            if not canvas_preview.winfo_exists():
                return
            canvas_preview.delete("all")
            try:
                largura_val = float(self.largura_var.get() or 0)
                altura_val = float(self.altura_var.get() or 0)
            except (tk.TclError, ValueError):
                return
            if largura_val <= 0 or altura_val <= 0:
                return
            canvas_width = canvas_preview.winfo_width()
            canvas_height = canvas_preview.winfo_height()
            if canvas_width <= 1 or canvas_height <= 1:
                canvas_preview.update_idletasks()
                canvas_width = canvas_preview.winfo_width()
                canvas_height = canvas_preview.winfo_height()
                if canvas_width <= 1 or canvas_height <= 1:
                    return
            escala_x = 400 / (largura_val * 1.2) * zoom_level
            escala_y = 100 / (altura_val * 1.2) * zoom_level  
            escala = min(escala_x, escala_y)
            x_inicial = 50 * zoom_level
            y_inicial = 130 * zoom_level  
            self._desenhar_elementos_preview(canvas_preview, x_inicial, y_inicial, largura_val, altura_val, escala)
            canvas_preview.configure(scrollregion=canvas_preview.bbox("all"))
        preview_frame.update_idletasks()
        canvas_preview.update_idletasks()
        atualizar_preview_local()
        self.largura_var.trace_add('write', lambda *a: atualizar_preview_local())
        self.altura_var.trace_add('write', lambda *a: atualizar_preview_local())
        campos_frame = ttk.Frame(frame_texto)
        campos_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(campos_frame, text="Pavimento:").grid(row=0, column=0, sticky="w", padx=2)
        pavimento_entry = ttk.Entry(campos_frame, textvariable=self.pavimento_var)
        pavimento_entry.grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Label(campos_frame, text="Observações:").grid(row=1, column=0, sticky="w", padx=2)
        obs_entry = ttk.Entry(campos_frame, textvariable=self.obs_var)
        obs_entry.grid(row=1, column=1, sticky="ew", padx=2)
        campos_frame.columnconfigure(1, weight=1)
        msg = msg.split("Preview:")[0].strip()
        txt = tk.Text(frame_texto, height=5, width=35, wrap="word")  
        txt.insert("1.0", msg)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True, pady=(0, 0), padx=0)
        frame_botoes = ttk.Frame(frame_texto)
        frame_botoes.pack(fill="x", pady=5)
        def confirmar_e_retornar():
            win.destroy()
            confirmar_callback(True)
        def continuacao_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima_fracao = numero_base + 0.1
            while proxima_fracao in numeros_existentes:
                proxima_fracao += 0.1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(f"{proxima_fracao:.1f}")
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def nova_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numeros_existentes = []
                for numero in self.fundos_salvos.keys():
                    try:
                        num = float(numero)
                        if num.is_integer():
                            numeros_existentes.append(int(num))
                    except (ValueError, TypeError):
                        continue
                if numeros_existentes:
                    numero_base = min(numeros_existentes) - 1
                else:
                    numero_base = 0
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo_numero = numero_base + 1
            while proximo_numero in numeros_existentes:
                proximo_numero += 1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(str(proximo_numero))
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def extensao_l():
            win.destroy()
            self._salvar_fundo() 
            self._executar_boundary_l()
        def cancelar():
            win.destroy()
            confirmar_callback(False)
        def reiniciar_selecao():
            win.destroy()
            self._reiniciar_selecao()
        style = ttk.Style()
        style.configure('Custom.TButton', padding=5)
        btn_confirmar = ttk.Button(frame_botoes, text="1: Confirmar e Retornar", 
                                 command=confirmar_e_retornar, style='Custom.TButton')
        btn_continuacao = ttk.Button(frame_botoes, text="2: Continuação da Viga", 
                                   command=continuacao_viga, style='Custom.TButton')
        btn_nova = ttk.Button(frame_botoes, text="3: Nova Viga", 
                            command=nova_viga, style='Custom.TButton')
        btn_extensao = ttk.Button(frame_botoes, text="4: Extensão em L", 
                                command=extensao_l, style='Custom.TButton')
        btn_cancelar = ttk.Button(frame_botoes, text="5: Cancelar", 
                                command=cancelar, style='Custom.TButton')
        btn_reiniciar = ttk.Button(frame_botoes, text="6: Reiniciar Seleção", 
                                command=reiniciar_selecao, style='Custom.TButton')
        btn_confirmar.grid(row=0, column=0, padx=5, pady=2, sticky="ew")
        btn_continuacao.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        btn_nova.grid(row=2, column=0, padx=5, pady=2, sticky="ew")
        btn_extensao.grid(row=3, column=0, padx=5, pady=2, sticky="ew")
        btn_cancelar.grid(row=4, column=0, padx=5, pady=2, sticky="ew")
        btn_reiniciar.grid(row=5, column=0, padx=5, pady=2, sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)
    def _desenhar_elementos_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        canvas.create_line(
            x_inicial, y_inicial,
            x_inicial + largura * escala, y_inicial,
            fill=self.cores['cota'], dash=(4, 4)
        )
        canvas.create_line(
            x_inicial, y_inicial - altura * escala,
            x_inicial + largura * escala, y_inicial - altura * escala,
            fill=self.cores['cota'], dash=(4, 4)
        )
        self._desenhar_cotas_preview(canvas, x_inicial, y_inicial, largura, altura, escala)
        self._desenhar_painel_principal_preview(canvas, x_inicial, y_inicial, largura, altura, escala)
        if self.sarrafo_esq_var.get():
            self._desenhar_sarrafo_preview(canvas, x_inicial, y_inicial, largura, altura, escala, 'esquerda')
        if self.sarrafo_dir_var.get():
            self._desenhar_sarrafo_preview(canvas, x_inicial, y_inicial, largura, altura, escala, 'direita')
        self._desenhar_aberturas_preview(canvas, x_inicial, y_inicial, largura, altura, escala)
        self._desenhar_textos_preview(canvas, x_inicial, y_inicial, largura, altura, escala)
        try:
            comprimento_2 = float(self.comprimento_2_var.get() or 0)
            largura_2 = float(self.largura_2_var.get() or 0)
            if comprimento_2 > 0 and largura_2 > 0:
                self._desenhar_painel_l_preview(canvas, x_inicial, y_inicial, largura, altura, comprimento_2, largura_2, escala)
        except (ValueError, tk.TclError):
            pass
    def _janela_teste_interativo(self, nome, resultado, on_finalizar):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        win = tk.Toplevel(self)
        win.title(f"Teste: {nome}")
        win.grab_set()
        win.geometry("650x440")
        frame_preview = ttk.Frame(win)
        frame_preview.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        fig, ax = plt.subplots(figsize=(4, 3.5))
        largura = resultado['dimensoes']['largura']
        altura = resultado['dimensoes']['altura']
        aberturas = resultado['aberturas']
        chanfros = resultado['chanfros']
        ce = chanfros.get('fundo_esq', 0)
        cd = chanfros.get('fundo_dir', 0)
        te = chanfros.get('topo_esq', 0)
        td = chanfros.get('topo_dir', 0)
        pontos = [
            (0, 0),
            (ce, 0),
            (largura-cd, 0),
            (largura, 0),
            (largura, altura),
            (largura-td, altura),
            (te, altura),
            (0, altura),
            (0, 0)
        ]
        ax.plot([p[0] for p in pontos], [p[1] for p in pontos], color='black', lw=2)
        for (x, y) in [(ce, 0), (largura-cd, 0), (largura-td, altura), (te, altura)]:
            ax.plot(x, y, 'o', color='orange', markersize=8)
        for ab in aberturas:
            ax.add_patch(plt.Rectangle((ab['x'], ab['y']), ab['largura'], ab['altura'], fill=True, color='red', alpha=0.5))
            x0 = ab['x']
            x1 = ab['x'] + ab['largura']
            y_cota = ab['y'] + ab['altura'] + altura*0.03
            ax.annotate('', xy=(x0, y_cota), xytext=(x1, y_cota), arrowprops=dict(arrowstyle='<->', color='blue'))
            ax.text((x0+x1)/2, y_cota+altura*0.02, f"L={ab['largura']:.1f}", color='blue', ha='center', fontsize=8)
            y0 = ab['y']
            y1 = ab['y'] + ab['altura']
            x_cota = ab['x'] + ab['largura'] + largura*0.03
            ax.annotate('', xy=(x_cota, y0), xytext=(x_cota, y1), arrowprops=dict(arrowstyle='<->', color='green'))
            ax.text(x_cota+largura*0.02, (y0+y1)/2, f"P={ab['altura']:.1f}", color='green', va='center', fontsize=8, rotation=90)
        ax.set_xlim(-10, largura+30)
        ax.set_ylim(-10, altura+30)
        ax.set_aspect('equal')
        ax.axis('off')
        canvas = FigureCanvasTkAgg(fig, master=frame_preview)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)
        canvas.draw()
        frame_texto = ttk.Frame(win)
        frame_texto.pack(side="right", fill="both", expand=True, padx=5, pady=5)
        txt = tk.Text(frame_texto, height=13, width=38, wrap="word")
        resumo = f"Dimensões: {largura:.2f} x {altura:.2f}\n\nChanfros:\n"
        for nome, valor in chanfros.items():
            resumo += f"{nome}: {valor:.2f}\n"
        resumo += "\nAberturas:\n"
        if not aberturas:
            resumo += "Nenhuma abertura identificada\n"
        else:
            for i, ab in enumerate(aberturas):
                resumo += (f"Abertura {i+1}:\n  Posição: ({ab['x']:.2f}, {ab['y']:.2f})\n"
                           f"  Largura: {ab['largura']:.2f}\n  Profundidade: {ab['profundidade']:.2f}\n"
                           f"  Classificação: {ab['lado']} {ab['topo_fundo']} ({ab['pos_x']}, {ab['pos_y']})\n")
        txt.insert("1.0", resumo)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True)
        label_coment = ttk.Label(frame_texto, text="Comentário (opcional):")
        label_coment.pack(anchor="w", pady=(8,0))
        entry_coment = tk.Text(frame_texto, height=2, width=35)
        entry_coment.pack(fill="x", pady=(0,8))
        frame_botoes = ttk.Frame(frame_texto)
        frame_botoes.pack(fill="x", pady=5)
        def confirmar(correto):
            comentario = entry_coment.get("1.0", "end").strip()
            win.destroy()
            on_finalizar(correto, comentario)
        ttk.Button(frame_botoes, text="Correto", command=lambda: confirmar(True)).pack(side="left", expand=True, fill="x", padx=5)
        ttk.Button(frame_botoes, text="Errado", command=lambda: confirmar(False)).pack(side="right", expand=True, fill="x", padx=5)
        win.wait_window()
    def rodar_teste_sintetico_interativo(self):
        print("\n===== TESTE SINTÉTICO INTERATIVO DE BOUNDARY =====\n")
        testes = [
            ("Chanfro esquerdo topo", [0,0, 0,100, 10,90, 200,90, 200,0, 0,0]),
            ("Chanfro esquerdo fundo", [0,0, 0,10, 10,0, 10,100, 200,100, 200,0, 0,0]),
            ("Chanfro direito topo", [0,0, 0,100, 200,100, 200,90, 190,100, 190,0, 0,0]),
            ("Chanfro direito fundo", [0,0, 0,100, 200,100, 200,10, 190,0, 0,0]),
            ("Abertura esquerda fundo", [0,0, 0,100, 200,100, 200,0, 0,0, 0,0, 20,0, 20,20, 0,20, 0,0]),
            ("Abertura esquerda topo", [0,0, 0,100, 200,100, 200,0, 0,0, 0,80, 20,80, 20,100, 0,100, 0,80]),
            ("Abertura direita fundo", [0,0, 0,100, 200,100, 200,0, 0,0, 180,0, 200,0, 200,20, 180,20, 180,0]),
            ("Abertura direita topo", [0,0, 0,100, 200,100, 200,0, 0,0, 180,80, 200,80, 200,100, 180,100, 180,80]),
            ("Chanfro esquerdo topo + abertura direita fundo", [0,0, 0,100, 10,90, 200,90, 200,0, 0,0, 180,0, 200,0, 200,20, 180,20, 180,0]),
            ("Chanfro direito fundo + abertura esquerda topo", [0,0, 0,100, 200,100, 200,10, 190,0, 0,0, 0,80, 20,80, 20,100, 0,100, 0,80]),
            ("Aberturas esquerda topo + direita fundo", [0,0, 0,100, 200,100, 200,0, 0,0, 0,80, 20,80, 20,100, 0,100, 0,80, 180,0, 200,0, 200,20, 180,20, 180,0]),
            ("Aberturas em todas as esquinas", [0,0, 0,100, 200,100, 200,0, 0,0, 0,0, 20,0, 20,20, 0,20, 0,0, 0,80, 20,80, 20,100, 0,100, 0,80, 180,0, 200,0, 200,20, 180,20, 180,0, 180,80, 200,80, 200,100, 180,100, 180,80])
        ]
        self.resultados_teste = []
        def proximo_teste(idx):
            if idx >= len(testes):
                print("\n===== FIM DOS TESTES INTERATIVOS =====\n")
                for i, (nome, _, correto, comentario) in enumerate(self.resultados_teste, 1):
                    print(f"{i}. {nome}: {'OK' if correto else 'ERRO'}" + (f" | Comentário: {comentario}" if comentario else ""))
                return
            nome, coords = testes[idx]
            resultado = self._analisar_coordenadas(coords, confirmar=False)
            def on_finalizar(correto, comentario):
                self.resultados_teste.append((nome, coords, correto, comentario))
                if correto:
                    self.learner.add_example(coords, resultado['chanfros'], resultado['aberturas'])
                    print(f"\nExemplo '{nome}' salvo para aprendizado!")
                proximo_teste(idx+1)
            self._janela_teste_interativo(nome, resultado, on_finalizar)
        proximo_teste(0)
    def _executar_boundary_l(self):
        """Executa a mesma função do botão 4 (Extensão em L) da janela de confirmação"""
        self._salvar_fundo()
        try:
            numero_atual = float(self.numero_var.get())
            numero_base = int(numero_atual)
        except (ValueError, TypeError):
            messagebox.showerror("Erro", "O campo número está vazio ou inválido")
            return
        proximo_numero = numero_base + 1
        while str(proximo_numero) in self.fundos_salvos:
            proximo_numero += 1
        pavimento_atual = self.pavimento_var.get()
        self._limpar_campos()
        self.numero_var.set(str(proximo_numero))
        self.pavimento_var.set(pavimento_atual)
        self.usar_boundary_duplo = True
        self._selecionar_textos_consecutivos()
    def _desenhar_cotas(self, x_inicial, y_inicial, largura, altura, escala):
        """Desenha as cotas do painel principal."""
        self.canvas.create_line(
            x_inicial, y_inicial + 20,
            x_inicial + largura * escala, y_inicial + 20,
            fill=self.cores['cota'], arrow="both"
        )
        self.canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial + 35,
            text=f"Largura: {largura:.1f}",
            fill=self.cores['texto']
        )
        self.canvas.create_line(
            x_inicial - 20, y_inicial,
            x_inicial - 20, y_inicial - altura * escala,
            fill=self.cores['cota'], arrow="both"
        )
        self.canvas.create_text(
            x_inicial - 35, y_inicial - (altura * escala) / 2,
            text=f"Altura: {altura:.1f}",
            fill=self.cores['texto'],
            angle=90
        )
    def _desenhar_painel_principal(self, x_inicial, y_inicial, largura, altura, escala):
        """Desenha o painel principal com chanfros."""
        recuos = [float(v.get() or 0) for v in self.recuos_vars]
        pontos = [
            x_inicial + recuos[1] * escala, y_inicial,  
            x_inicial + (largura - recuos[3]) * escala, y_inicial,  
            x_inicial + largura * escala, y_inicial,  
            x_inicial + (largura - recuos[2]) * escala, y_inicial - altura * escala,  
            x_inicial + recuos[0] * escala, y_inicial - altura * escala,  
            x_inicial, y_inicial - altura * escala,  
            x_inicial + recuos[1] * escala, y_inicial  
        ]
        self.canvas.create_polygon(
            pontos,
            fill=self.cores['painel'],
            outline=self.cores['divisao']
        )
        # Desenhar triângulos vermelhos para os chanfros
        chanfro_color = '#FF3333'
        # T/E (topo esquerda)
        if recuos[0] > 0:
            self.canvas.create_polygon(
                [x_inicial, y_inicial - altura * escala,
                 x_inicial + recuos[0] * escala, y_inicial - altura * escala,
                 x_inicial, y_inicial - (altura - recuos[0]) * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[1] > 0:
            self.canvas.create_polygon(
                [x_inicial, y_inicial,
                 x_inicial + recuos[1] * escala, y_inicial,
                 x_inicial, y_inicial - recuos[1] * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[2] > 0:
            self.canvas.create_polygon(
                [x_inicial + largura * escala, y_inicial - altura * escala,
                 x_inicial + (largura - recuos[2]) * escala, y_inicial - altura * escala,
                 x_inicial + largura * escala, y_inicial - (altura - recuos[2]) * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[3] > 0:
            self.canvas.create_polygon(
                [x_inicial + largura * escala, y_inicial,
                 x_inicial + (largura - recuos[3]) * escala, y_inicial,
                 x_inicial + largura * escala, y_inicial - recuos[3] * escala],
                fill=chanfro_color, outline=chanfro_color
            )
    def _desenhar_sarrafo(self, x_inicial, y_inicial, largura, altura, escala, lado):
        """Desenha os sarrafos verticais."""
        recuos = [float(v.get() or 0) for v in self.recuos_vars]
        if lado == 'esquerda':
            x = x_inicial + (recuos[1] + 7) * escala
        else:
            x = x_inicial + (largura - recuos[2] - 7) * escala
        self.canvas.create_line(
            x, y_inicial,
            x, y_inicial - altura * escala,
            fill=self.cores['sarrafo'],
            width=2
        )
    def _desenhar_aberturas(self, x_inicial, y_inicial, largura, altura, escala):
        """Desenha as aberturas do painel principal SEM campo de distância (sempre encostadas na parede)."""
        for i, linha in enumerate(self.aberturas_vars):
            valores = [self._float_safe(v) for v in linha] # Usar _float_safe
            print(f"[DEBUG] Abertura {i}: valores = {valores}")
            if len(valores) < 3: # Compatibilidade com dados antigos
                dist, prof, larg = 0, valores[0], valores[1]
            else:
                dist, prof, larg = valores[0], valores[1], valores[2]

            if prof > 0 and larg > 0:  
                # A lógica de posicionamento atual já assume posições fixas, 'dist' seria um offset.
                # Se 'dist' for 0, o comportamento é o mesmo. Se 'dist' != 0, a abertura se move.
                # Para manter "sempre encostadas na parede" e permitir 'dist' no modelo de dados,
                # vou interpretar 'dist' como uma distância ao longo da aresta, não um offset flutuante.
                # No entanto, a lógica de desenho abaixo está fixada nos cantos.
                # Para simplificar e focar no M2, e dada a premissa de "encostadas na parede",
                # vou manter o posicionamento baseado em 'i' e usar 'dist' apenas se o usuário
                # realmente quiser aberturas "flutuantes" no futuro.
                # Por ora, 'dist' será apenas um dado para o Excel.
                if i == 0: # Topo Esquerda
                    x_pos = x_inicial + dist * escala
                    y_pos = y_inicial - altura * escala
                elif i == 1: # Fundo Esquerda
                    x_pos = x_inicial + dist * escala
                    y_pos = y_inicial - prof * escala
                elif i == 2: # Topo Direita
                    x_pos = x_inicial + (largura - larg - dist) * escala
                    y_pos = y_inicial - altura * escala
                elif i == 3: # Fundo Direita
                    x_pos = x_inicial + (largura - larg - dist) * escala
                    y_pos = y_inicial - prof * escala
                else:
                    continue
                print(f"[DEBUG] Desenhando abertura {i} em x={x_pos}, y={y_pos}, larg={larg}, prof={prof}")
                self.canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + larg * escala, y_pos + prof * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura(x_pos, y_pos, x_pos + larg * escala, y_pos + prof * escala)
    def _desenhar_aberturas_painel_l(self, x, y, largura, altura, escala, tipo_l, y_final):
        """Desenha as aberturas do painel em L SEM campo de distância (sempre encostadas na parede), invertendo profundidade/largura (prof em X, larg em Y)."""
        for i, linha in enumerate(self.aberturas_2_vars):
            valores = [self._float_safe(v) for v in linha] # Usar _float_safe
            if len(valores) < 3: # Compatibilidade com dados antigos
                dist, prof, larg = 0, valores[0], valores[1]
            else:
                dist, prof, larg = valores[0], valores[1], valores[2]

            if prof > 0 and larg > 0:
                if altura > largura: # Se o painel L está "alto"
                    if i == 0: # Topo Esquerda (no painel principal é TE, no L pode ser outro canto relativo)
                        x_pos = x + dist * escala
                        y_pos = y - larg * escala
                    elif i == 1: # Fundo Esquerda
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y - larg * escala
                    elif i == 2: # Topo Direita
                        x_pos = x + dist * escala
                        y_pos = y_final
                    elif i == 3: # Fundo Direita
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y_final
                    else:
                        continue
                else: # Se o painel L está "largo"
                    if i == 0: # Topo Esquerda
                        x_pos = x + dist * escala
                        y_pos = y_final
                    elif i == 1: # Fundo Esquerda
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y - larg * escala
                    elif i == 2: # Topo Direita
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y_final
                    elif i == 3: # Fundo Direita
                        x_pos = x + dist * escala
                        y_pos = y - larg * escala
                    else:
                        continue
                self.canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + prof * escala, y_pos + larg * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura(x_pos, y_pos, x_pos + prof * escala, y_pos + larg * escala)
    def _desenhar_cotas_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        canvas.create_line(
            x_inicial, y_inicial + 20,
            x_inicial + largura * escala, y_inicial + 20,
            fill=self.cores['cota'], arrow="both"
        )
        canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial + 35,
            text=f"Largura: {largura:.1f}",
            fill=self.cores['texto']
        )
        canvas.create_line(
            x_inicial - 20, y_inicial,
            x_inicial - 20, y_inicial - altura * escala,
            fill=self.cores['cota'], arrow="both"
        )
        canvas.create_text(
            x_inicial - 35, y_inicial - (altura * escala) / 2,
            text=f"Altura: {altura:.1f}",
            fill=self.cores['texto'],
            angle=90
        )
    def _desenhar_painel_principal_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        recuos = [float(v.get() or 0) for v in self.recuos_vars]
        pontos = [
            x_inicial + recuos[1] * escala, y_inicial,  
            x_inicial + (largura - recuos[3]) * escala, y_inicial,  
            x_inicial + largura * escala, y_inicial,  
            x_inicial + (largura - recuos[2]) * escala, y_inicial - altura * escala,  
            x_inicial + recuos[0] * escala, y_inicial - altura * escala,  
            x_inicial, y_inicial - altura * escala,  
            x_inicial + recuos[1] * escala, y_inicial  
        ]
        canvas.create_polygon(
            pontos,
            fill=self.cores['painel'],
            outline=self.cores['divisao']
        )
        chanfro_color = '#FF3333'
        if recuos[0] > 0:
            canvas.create_polygon(
                [x_inicial, y_inicial - altura * escala,
                 x_inicial + recuos[0] * escala, y_inicial - altura * escala,
                 x_inicial, y_inicial - (altura - recuos[0]) * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[1] > 0:
            canvas.create_polygon(
                [x_inicial, y_inicial,
                 x_inicial + recuos[1] * escala, y_inicial,
                 x_inicial, y_inicial - recuos[1] * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[2] > 0:
            canvas.create_polygon(
                [x_inicial + largura * escala, y_inicial - altura * escala,
                 x_inicial + (largura - recuos[2]) * escala, y_inicial - altura * escala,
                 x_inicial + largura * escala, y_inicial - (altura - recuos[2]) * escala],
                fill=chanfro_color, outline=chanfro_color
            )
        if recuos[3] > 0:
            canvas.create_polygon(
                [x_inicial + largura * escala, y_inicial,
                 x_inicial + (largura - recuos[3]) * escala, y_inicial,
                 x_inicial + largura * escala, y_inicial - recuos[3] * escala],
                fill=chanfro_color, outline=chanfro_color
            )
    def _desenhar_sarrafo_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala, lado):
        recuos = [float(v.get() or 0) for v in self.recuos_vars]
        if lado == 'esquerda':
            x = x_inicial + (recuos[1] + 7) * escala
        else:
            x = x_inicial + (largura - recuos[2] - 7) * escala
        canvas.create_line(
            x, y_inicial,
            x, y_inicial - altura * escala,
            fill=self.cores['sarrafo'],
            width=2
        )
    def _desenhar_aberturas_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        for i, linha in enumerate(self.aberturas_vars):
            valores = [self._float_safe(v) for v in linha] # Usar _float_safe
            if len(valores) < 3:
                dist, prof, larg = 0, valores[0], valores[1]
            else:
                dist, prof, larg = valores[0], valores[1], valores[2]

            if prof > 0 and larg > 0:
                if i == 0:
                    x_pos = x_inicial + dist * escala
                    y_pos = y_inicial - altura * escala
                elif i == 1:
                    x_pos = x_inicial + dist * escala
                    y_pos = y_inicial - prof * escala
                elif i == 2:
                    x_pos = x_inicial + (largura - larg - dist) * escala
                    y_pos = y_inicial - altura * escala
                elif i == 3:
                    x_pos = x_inicial + (largura - larg - dist) * escala
                    y_pos = y_inicial - prof * escala
                else:
                    continue
                canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + larg * escala, y_pos + prof * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura_preview(canvas, x_pos, y_pos, x_pos + larg * escala, y_pos + prof * escala)
    def _desenhar_hachura_abertura_preview(self, canvas, x1, y1, x2, y2):
        largura = x2 - x1
        altura = y2 - y1
        if largura < 5 or altura < 5:
            return
        espacamento = 5
        for i in range(0, int(largura + altura), espacamento):
            x_inicio = x1 + min(i, largura) if i < largura else x2
            y_inicio = y1 if i < largura else y1 + (i - largura)
            x_fim = x1 if i < altura else x1 + (i - altura)
            y_fim = y1 + min(i, altura) if i < altura else y2
            if all(coord <= limite for coord, limite in [(x_inicio, x2), (y_inicio, y2), (x_fim, x2), (y_fim, y2)]):
                canvas.create_line(
                    x_inicio, y_inicio,
                    x_fim, y_fim,
                    fill=self.cores['divisao'],
                    width=1
                )
    def _desenhar_textos_preview(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial - altura * escala - 40,
            text=f"Fundo: {self.numero_var.get() or ''} - {self.nome_var.get() or ''}",
            fill=self.cores['texto'],
            font=("Arial", 12, "bold"),
            anchor='s'
        )
        canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial + 50,
            text=f"{self.nome_var.get() or ''} - {self.obs_var.get() or ''}",
            fill=self.cores['texto']
        )
        canvas.create_text(
            x_inicial - 50, y_inicial - (altura * escala) / 2,
            text=self.texto_esq_var.get() or "",
            fill=self.cores['texto'],
            angle=90
        )
        canvas.create_text(
            x_inicial + largura * escala + 50, y_inicial - (altura * escala) / 2,
            text=self.texto_dir_var.get() or "",
            fill=self.cores['texto'],
            angle=90
        )
    def _desenhar_painel_l_preview(self, canvas, x_inicial, y_inicial, largura_principal, altura_principal, comprimento, largura, escala):
        tipo_l = self.tipo_painel2_var.get()
        if comprimento < largura:
            l_final = comprimento
            h_final = largura
        else:
            l_final = largura
            h_final = comprimento
        if tipo_l.startswith("E"):
            x = x_inicial - l_final * escala
        else:
            x = x_inicial + largura_principal * escala
        if tipo_l.endswith("T"):
            y = y_inicial
            y_final = y - h_final * escala
        else:
            y = y_inicial - altura_principal * escala + h_final * escala
            y_final = y - h_final * escala
        canvas.create_rectangle(
            x, y,
            x + l_final * escala, y_final,
            fill=self.cores['painel'],
            outline=self.cores['divisao']
        )
        chanfro_color = '#FF3333'
        recuos = [float(v.get() or 0) for v in self.recuos_2_vars]
        if h_final > l_final:
            if recuos[0] > 0:
                canvas.create_polygon(
                    [x, y,
                     x + recuos[0] * escala, y,
                     x, y - recuos[0] * escala if y_final < y else y + recuos[0] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[3] > 0:
                canvas.create_polygon(
                    [x + l_final * escala, y_final,
                     x + (l_final - recuos[3]) * escala, y_final,
                     x + l_final * escala, y_final + recuos[3] * escala if y_final < y else y_final - recuos[3] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[2] > 0:
                canvas.create_polygon(
                    [x, y_final,
                     x + recuos[2] * escala, y_final,
                     x, y_final + recuos[2] * escala if y_final < y else y_final - recuos[2] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[1] > 0:
                canvas.create_polygon(
                    [x + l_final * escala, y,
                     x + (l_final - recuos[1]) * escala, y,
                     x + l_final * escala, y - recuos[1] * escala if y_final < y else y + recuos[1] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
        else:
            if recuos[0] > 0:
                canvas.create_polygon(
                    [x, y_final,
                     x + recuos[0] * escala, y_final,
                     x, y_final + recuos[0] * escala if y_final < y else y_final - recuos[0] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[1] > 0:
                canvas.create_polygon(
                    [x, y,
                     x + recuos[1] * escala, y,
                     x, y - recuos[1] * escala if y_final < y else y + recuos[1] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[2] > 0:
                canvas.create_polygon(
                    [x + l_final * escala, y_final,
                     x + (l_final - recuos[2]) * escala, y_final,
                     x + l_final * escala, y_final + recuos[2] * escala if y_final < y else y_final - recuos[2] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
            if recuos[3] > 0:
                canvas.create_polygon(
                    [x + l_final * escala, y,
                     x + (l_final - recuos[3]) * escala, y,
                     x + l_final * escala, y - recuos[3] * escala if y_final < y else y + recuos[3] * escala],
                    fill=chanfro_color, outline=chanfro_color
                )
        self._desenhar_cotas_painel_l(x, y, l_final, h_final, escala, tipo_l, y_final)
        self._desenhar_aberturas_painel_l_preview(canvas, x, y, l_final, h_final, escala, tipo_l, y_final)
    def _desenhar_aberturas_painel_l_preview(self, canvas, x, y, largura, altura, escala, tipo_l, y_final):
        for i, linha in enumerate(self.aberturas_2_vars):
            valores = [self._float_safe(v) for v in linha] # Usar _float_safe
            if len(valores) < 3:
                dist, prof, larg = 0, valores[0], valores[1]
            else:
                dist, prof, larg = valores[0], valores[1], valores[2]

            if prof > 0 and larg > 0:
                if altura > largura:
                    if i == 0:
                        x_pos = x + dist * escala
                        y_pos = y - larg * escala
                    elif i == 1:
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y - larg * escala
                    elif i == 2:
                        x_pos = x + dist * escala
                        y_pos = y_final
                    elif i == 3:
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y_final
                    else:
                        continue
                else:
                    if i == 0:
                        x_pos = x + dist * escala
                        y_pos = y_final
                    elif i == 1:
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y - larg * escala
                    elif i == 2:
                        x_pos = x + (largura - prof - dist) * escala
                        y_pos = y_final
                    elif i == 3:
                        x_pos = x + dist * escala
                        y_pos = y - larg * escala
                    else:
                        continue
                canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + prof * escala, y_pos + larg * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura_preview(canvas, x_pos, y_pos, x_pos + prof * escala, y_pos + larg * escala)
    def _mostrar_comentario_flutuante(self, mensagem, tempo=None):
        import tkinter as tk
        if getattr(self, '_flutuante_cancelado', False):
            print('[DEBUG] Balão NÃO criado pois _flutuante_cancelado=True')
            return None
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
            self._balcao_flutuante = None
        if hasattr(self, '_flutuante_after_id') and self._flutuante_after_id is not None:
            try:
                self.after_cancel(self._flutuante_after_id)
            except:
                pass
            self._flutuante_after_id = None
        self._flutuante_cancelado = False
        balcao = tk.Toplevel()
        balcao.attributes('-topmost', True)
        balcao.configure(bg='#fff8b0')
        # Centralizar na tela
        largura = 480
        altura = 120  
        x = (balcao.winfo_screenwidth() - largura) // 2
        y = 40
        balcao.geometry(f"{largura}x{altura}+{x}+{y}")
        balcao.lift()
        balcao.focus_force()  
        balcao.grab_set()  
        try:
            balcao.bind_all('<Escape>', lambda e: self.event_generate('<<GlobalEscape>>'))
        except:
            pass
        label = tk.Label(balcao, text=mensagem, font=("Arial", 16, "bold"), bg='#fff8b0', fg='#333', wraplength=460)
        label.pack(expand=True, fill=tk.BOTH, padx=10, pady=(10,0))
        def fechar_balao_e_trazer_producao():
            print('[DEBUG] Fechar balão flutuante chamado')
            if hasattr(self, '_flutuante_after_id') and self._flutuante_after_id is not None:
                try:
                    self.after_cancel(self._flutuante_after_id)
                except:
                    pass
                self._flutuante_after_id = None
            self._flutuante_cancelado = True
            try:
                print('[DEBUG] Chamando balcao.quit()')
                balcao.quit()  
                print('[DEBUG] Chamando balcao.destroy()')
                balcao.destroy()
            except Exception as e:
                print(f'[DEBUG] Erro ao fechar balcao: {e}')
            self._balcao_flutuante = None
        btn_cancelar = tk.Button(balcao, text="Cancelar Seleção", font=("Arial", 12, "bold"), bg="#ffb0b0", command=fechar_balao_e_trazer_producao)
        btn_cancelar.pack(pady=(8, 10))
        btn_cancelar.focus_set()
        if tempo is None:
            def on_enter(event=None):
                print('[DEBUG] Evento ENTER/ESC capturado no balcao')
                fechar_balao_e_trazer_producao()
            balcao.bind_all('<Return>', on_enter)
            balcao.bind_all('<KP_Enter>', on_enter)
            balcao.bind_all('<Escape>', on_enter)
        print('[DEBUG] Balcao flutuante criado e aguardando interação')
        self._balcao_flutuante = balcao
        balcao.update()
        if tempo is not None:
            def after_callback():
                if not getattr(self, '_flutuante_cancelado', False):
                    try:
                        balcao.destroy()
                    except:
                        pass
                    self._balcao_flutuante = None
            self._flutuante_after_id = self.after(int(tempo*1000), after_callback)
        else:
            self._flutuante_after_id = None
        return balcao
    def _fechar_comentario_flutuante(self):
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
            self._balcao_flutuante = None
    def _area_intersecao(self, rx1, ry1, rx2, ry2, ax1, ay1, ax2, ay2):
        """Calcula a área de interseção entre dois retângulos."""
        inter_x1 = max(rx1, ax1)
        inter_y1 = max(ry1, ay1)
        inter_x2 = min(rx2, ax2)
        inter_y2 = min(ry2, ay2)
        if inter_x2 > inter_x1 and inter_y2 > inter_y1:
            return (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
        return 0
    def _calcular_area_util(self, *args):
        try:
            largura_fundo = self._float_safe(self.largura_var)
            altura_fundo = self._float_safe(self.altura_var)
            print(f"[DEBUG CALC] Largura Fundo: {largura_fundo}, Altura Fundo: {altura_fundo}")
            area_total = largura_fundo * altura_fundo
            print(f"[DEBUG CALC] Área Total do Fundo: {area_total}")
            area_sobreposta = 0.0
            rx1_fundo, ry1_fundo, rx2_fundo, ry2_fundo = 0, 0, largura_fundo, altura_fundo
            for i, linha in enumerate(self.aberturas_vars):
                print(f"[DEBUG CALC] Processando Abertura {i}: {linha}")
                try:
                    dist = self._float_safe(linha[0]) # Distância (não usada no cálculo da área)
                    prof = self._float_safe(linha[1]) # Profundidade
                    larg = self._float_safe(linha[2]) # Largura
                    print(f"[DEBUG CALC] Abertura {i} - Distância: {dist}, Profundidade: {prof}, Largura: {larg}")
                    if prof > 0 and larg > 0:
                        ax1 = 0
                        ay1 = 0
                        ax2 = ax1 + larg
                        ay2 = ay1 + prof
                        ax1 = max(0, ax1)
                        ay1 = max(0, ay1)
                        ax2 = min(largura_fundo, ax2)
                        ay2 = min(altura_fundo, ay2)
                        print(f"[DEBUG CALC] Abertura {i} - Coords (x1,y1,x2,y2): ({ax1}, {ay1}, {ax2}, {ay2})")
                        if ax2 > ax1 and ay2 > ay1:
                            area_intersecao = self._area_intersecao(rx1_fundo, ry1_fundo, rx2_fundo, ry2_fundo, ax1, ay1, ax2, ay2)
                            area_sobreposta += area_intersecao
                            print(f"[DEBUG CALC] Abertura {i} - Área Interseção: {area_intersecao}, Área Sobreposta Acumulada: {area_sobreposta}")
                        else:
                            print(f"[DEBUG] Abertura inválida ou fora dos limites para cálculo: {linha}")
                except Exception as e:
                    print(f"Erro ao processar dados de abertura para cálculo: {e} - Dados: {linha}")
                    continue
            recuos = [self._float_safe(v) for v in self.recuos_vars]
            area_chanfros = 0.0
            if recuos[0] > 0:
                area_chanfros += (recuos[0] * recuos[0]) / 2
            if recuos[1] > 0:
                area_chanfros += (recuos[1] * recuos[1]) / 2
            if recuos[2] > 0:
                area_chanfros += (recuos[2] * recuos[2]) / 2
            if recuos[3] > 0:
                area_chanfros += (recuos[3] * recuos[3]) / 2
            largura_l = self._float_safe(self.largura_2_var)
            comprimento_l = self._float_safe(self.comprimento_2_var)
            area_l = 0.0
            area_sobreposta_l = 0.0
            area_chanfros_l = 0.0
            if largura_l > 0 and comprimento_l > 0:
                area_l = largura_l * comprimento_l
                rx1_l, ry1_l, rx2_l, ry2_l = 0, 0, largura_l, comprimento_l
                for i, linha in enumerate(self.aberturas_2_vars):
                    try:
                        dist = self._float_safe(linha[0]) # Distância (não usada no cálculo da área)
                        prof = self._float_safe(linha[1]) # Profundidade
                        larg = self._float_safe(linha[2]) # Largura
                        if prof > 0 and larg > 0:
                            ax1 = 0
                            ay1 = 0
                            ax2 = ax1 + larg
                            ay2 = ay1 + prof
                            ax1 = max(0, ax1)
                            ay1 = max(0, ay1)
                            ax2 = min(largura_l, ax2)
                            ay2 = min(comprimento_l, ay2)
                            if ax2 > ax1 and ay2 > ay1:
                                area_intersecao = self._area_intersecao(rx1_l, ry1_l, rx2_l, ry2_l, ax1, ay1, ax2, ay2)
                                area_sobreposta_l += area_intersecao
                    except Exception as e:
                        print(f"Erro ao processar dados de abertura L para cálculo: {e} - Dados: {linha}")
                        continue
                recuos_l = [self._float_safe(v) for v in self.recuos_2_vars]
                if recuos_l[0] > 0:
                    area_chanfros_l += (recuos_l[0] * recuos_l[0]) / 2
                if recuos_l[1] > 0:
                    area_chanfros_l += (recuos_l[1] * recuos_l[1]) / 2
                if recuos_l[2] > 0:
                    area_chanfros_l += (recuos_l[2] * recuos_l[2]) / 2
                if recuos_l[3] > 0:
                    area_chanfros_l += (recuos_l[3] * recuos_l[3]) / 2
            area_util_m2 = max(0, (area_total - area_sobreposta - area_chanfros + area_l - area_sobreposta_l - area_chanfros_l) / 10000)
            print(f"[DEBUG CALC] Área Útil Final (M²): {area_util_m2}")
            self.area_util_var.set(f"M² = {area_util_m2:.2f}")
            if hasattr(self, '_atualizando_area_util') and self._atualizando_area_util:
                return
            numero_fundo = self.numero_var.get()
            if numero_fundo:
                if numero_fundo not in self.fundos_salvos:
                    self.fundos_salvos[numero_fundo] = self._coletar_dados()
                self.fundos_salvos[numero_fundo]['area_util_m2'] = area_util_m2
                self._salvar_fundos_salvos()
            self._atualizar_lista()
        except Exception as e:
            print(f"Erro no cálculo da área útil principal: {e}")
            self.area_util_var.set("M² = ERRO")
    def _float_safe(self, var):
        """Tenta converter a variável Tkinter (StringVar, DoubleVar) para float, retornando 0.0 se houver erro."""
        try:
            if isinstance(var, (tk.StringVar, tk.DoubleVar)):
                return float(var.get())
            else:
                return float(var)
        except (ValueError, AttributeError):
            return 0.0
    def _float_from_m2_string(self, m2_string):
        """Converte a string de M² (ex: 'M² = 123.45') para float."""
        try:
            if m2_string.startswith("M² = "):
                return float(m2_string.replace("M² = ", ""))
            return float(m2_string)
        except (ValueError, AttributeError):
            return 0.0
    def _reset_load_prevention_flag(self):
        print(f"[DEBUG] Flag _prevent_treeview_load_after_edit resetada para False. Estado atual: {self._prevent_treeview_load_after_edit}")
        self._prevent_treeview_load_after_edit = False
    def _on_entry_focus_in(self, event):
        self._is_editing_entry = True
        print(f"[DEBUG] _is_editing_entry set to True on FocusIn.")
    def _on_entry_focus_out(self, event):
        print(f"[DEBUG] _is_editing_entry set to False on FocusOut.")
        self._is_editing_entry = False
        self._calcular_area_util()
    def _preencher_campos_duplos(self, analise_combinada):
        self._ignorar_traces = True
        try:
            principal = analise_combinada['principal']
            self.largura_var.set(principal['dimensoes']['largura'])
            self.altura_var.set(principal['dimensoes']['altura'])
            self.recuos_vars[0].set(principal['chanfros']['topo_esq'])
            self.recuos_vars[1].set(principal['chanfros']['fundo_esq'])
            self.recuos_vars[2].set(principal['chanfros']['topo_dir'])
            self.recuos_vars[3].set(principal['chanfros']['fundo_dir'])
            for i, abertura in enumerate(principal['aberturas']):
                if i < len(self.aberturas_vars):
                    self.aberturas_vars[i][0].set(0) # Distância (assumindo 0 para aberturas encostadas)
                    self.aberturas_vars[i][1].set(abertura['altura'])  # Profundidade
                    self.aberturas_vars[i][2].set(abertura['largura']) # Largura
            painel_l = analise_combinada['painel_l']
            self.comprimento_2_var.set(painel_l['dimensoes']['largura'])
            self.largura_2_var.set(painel_l['dimensoes']['altura'])
            self.recuos_2_vars[0].set(painel_l['chanfros']['topo_esq'])
            self.recuos_2_vars[1].set(painel_l['chanfros']['fundo_esq'])
            self.recuos_2_vars[2].set(painel_l['chanfros']['topo_dir'])
            self.recuos_2_vars[3].set(painel_l['chanfros']['fundo_dir'])
            for i, abertura in enumerate(painel_l['aberturas']):
                if i < len(self.aberturas_2_vars):
                    self.aberturas_2_vars[i][0].set(0) # Distância (assumindo 0 para aberturas encostadas)
                    self.aberturas_2_vars[i][1].set(abertura['altura'])  # Profundidade
                    self.aberturas_2_vars[i][2].set(abertura['largura']) # Largura
            if 'tipo_l' in analise_combinada and analise_combinada['tipo_l'] is not None:
                self.tipo_painel2_var.set(analise_combinada['tipo_l'])
            else:
                self.tipo_painel2_var.set("E/T") 
            self._atualizar_preview()
        finally:
            self._ignorar_traces = False
    def _realizar_boundary_duplo(self):
        try:
            print("[DEBUG] Iniciando _realizar_boundary_duplo")
            pythoncom.CoInitialize()
            if self.ac is None:
                print("[DEBUG] Criando conexão com AutoCAD")
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Nenhum documento AutoCAD ativo encontrado!", tempo=2)
                return
            print("[DEBUG] Documento AutoCAD encontrado")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.ShowWindow(acad_window, 3)  
                    win32gui.SetForegroundWindow(acad_window)
                    self.update()
                    self.after(100)
            except Exception as e:
                print(f"[DEBUG] Erro ao trazer janela para frente: {str(e)}")
                pass
            self.iconify()
            self._mostrar_comentario_flutuante("Selecione as duas áreas (retângulo principal e em L) e pressione Enter")
            self.update()
            try:
                if hasattr(doc, 'SetVariable'):
                    print("[DEBUG] Configurando OSMODE")
                    doc.SetVariable("OSMODE", 35)
            except Exception as e:
                print(f"[DEBUG] Erro ao configurar OSMODE: {str(e)}")
                pass
            try:
                doc.SendCommand("._BOUNDARY\n")
                print("[DEBUG] Comando BOUNDARY enviado")
                time.sleep(0.5)
                print("[DEBUG] Modo de seleção ativado")
            except Exception as e:
                print(f"[DEBUG] Erro ao enviar comando: {str(e)}")
                raise
            time.sleep(1)
            self._fechar_comentario_flutuante()
            try:
                count = doc.ModelSpace.Count
                if count < 2:
                    raise Exception("Não foram encontradas duas áreas selecionadas")
                last_entity = doc.ModelSpace.Item(count - 1)
                prev_entity = doc.ModelSpace.Item(count - 2)
                coords_1 = last_entity.Coordinates
                coords_2 = prev_entity.Coordinates
                coords_1 = alinhar_boundary_horizontal(coords_1)
                coords_2 = alinhar_boundary_horizontal(coords_2)
            except Exception as e:
                print(f"[DEBUG] Erro ao obter boundaries: {str(e)}")
                self._mostrar_comentario_flutuante(f"Erro ao obter boundaries: {str(e)}", tempo=2)
                return
            self.deiconify()
            analise_1, _ = self._analisar_coordenadas(coords_1)
            analise_2, _ = self._analisar_coordenadas(coords_2)
            if analise_1 is None or analise_2 is None:
                self._mostrar_comentario_flutuante("Não foi possível analisar um ou ambos os boundaries.", tempo=2)
                return
            if analise_1['dimensoes']['largura'] >= analise_2['dimensoes']['largura']:
                retangulo_principal = analise_1
                retangulo_l = analise_2
            else:
                retangulo_principal = analise_2
                retangulo_l = analise_1
            tipo_l_detectado = self._detectar_tipo_l(retangulo_principal, retangulo_l)
            analise_combinada = {
                'principal': retangulo_principal,
                'painel_l': retangulo_l,
                'tipo_l': tipo_l_detectado 
            }
            print("[DEBUG] Preenchendo campos duplos imediatamente")
            self._preencher_campos_duplos(analise_combinada)
            def on_confirmar_boundary_duplo(resposta):
                if resposta:
                    self.learner.add_example(coords_1, analise_1['chanfros'], analise_1['aberturas'])
                    self.learner.add_example(coords_2, analise_2['chanfros'], analise_2['aberturas'])
                    print("\nExemplos de boundary duplo salvos para aprendizado!")
                else:
                    print("\nExemplos de boundary duplo não foram salvos.")
            self.deiconify() 
            self._janela_confirmacao_salvamento_duplo(
                retangulo_principal['dimensoes']['largura'],
                retangulo_principal['dimensoes']['altura'],
                retangulo_l['dimensoes']['largura'],
                retangulo_l['dimensoes']['altura'],
                retangulo_principal['aberturas'],
                retangulo_l['aberturas'],
                analise_combinada,
                on_confirmar_boundary_duplo
            )
        except Exception as e:
            print(f"[DEBUG] Erro geral em _realizar_boundary_duplo: {str(e)}")
            self._mostrar_comentario_flutuante(f"Erro ao criar boundary duplo: {str(e)}", tempo=2)
            self.deiconify()
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
            self.deiconify()
    def _janela_confirmacao_salvamento_duplo(self, largura1, altura1, largura2, altura2, aberturas1, aberturas2, analise_combinada, confirmar_callback):
        win = tk.Toplevel(self)
        win.title("Confirmar Salvamento Duplo")
        win.grab_set()
        win.geometry("600x600")
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame_texto = ttk.Frame(win)
        frame_texto.pack(fill="both", expand=True, padx=5, pady=5)
        preview_frame = ttk.LabelFrame(frame_texto, text="Visualização", padding=5)
        preview_frame.pack(fill=tk.X, expand=True, pady=(0, 8))
        canvas_preview = tk.Canvas(preview_frame, width=450, height=120, bg=self.cores['fundo'], scrollregion=(0, 0, 900, 240))  
        canvas_preview.pack(fill=tk.BOTH, expand=True)
        zoom_level = 1.0
        pan_start = [None]
        def start_pan(event):
            canvas_preview.scan_mark(event.x, event.y)
            pan_start[0] = (event.x, event.y)
        def pan(event):
            if pan_start[0]:
                canvas_preview.scan_dragto(event.x, event.y, gain=1)
                pan_start[0] = (event.x, event.y)
        def stop_pan(event):
            pan_start[0] = None
        def mouse_wheel(event):
            nonlocal zoom_level
            if hasattr(event, 'num') and event.num == 5 or (hasattr(event, 'delta') and event.delta < 0):
                zoom_level = max(zoom_level - 0.1, 0.5)
            else:
                zoom_level = min(zoom_level + 0.1, 2.0)
            atualizar_preview_local()
        canvas_preview.bind("<ButtonPress-1>", start_pan)
        canvas_preview.bind("<B1-Motion>", pan)
        canvas_preview.bind("<ButtonRelease-1>", stop_pan)
        canvas_preview.bind("<MouseWheel>", mouse_wheel)
        canvas_preview.bind("<Button-4>", mouse_wheel)
        canvas_preview.bind("<Button-5>", mouse_wheel)
        def atualizar_preview_local():
            if not canvas_preview.winfo_exists():
                return
            canvas_preview.delete("all")
            try:
                largura_val = float(self.largura_var.get() or 0)
                altura_val = float(self.altura_var.get() or 0)
            except (tk.TclError, ValueError):
                return
            if largura_val <= 0 or altura_val <= 0:
                return
            canvas_width = canvas_preview.winfo_width()
            canvas_height = canvas_preview.winfo_height()
            if canvas_width <= 1 or canvas_height <= 1:
                canvas_preview.update_idletasks()
                canvas_width = canvas_preview.winfo_width()
                canvas_height = canvas_preview.winfo_height()
                if canvas_width <= 1 or canvas_height <= 1:
                    return
            escala_x = 400 / (largura_val * 1.2) * zoom_level
            escala_y = 100 / (altura_val * 1.2) * zoom_level  
            escala = min(escala_x, escala_y)
            x_inicial = 50 * zoom_level
            y_inicial = 130 * zoom_level  
            self._desenhar_elementos_preview(canvas_preview, x_inicial, y_inicial, largura_val, altura_val, escala)
            canvas_preview.configure(scrollregion=canvas_preview.bbox("all"))
        preview_frame.update_idletasks()
        canvas_preview.update_idletasks()
        atualizar_preview_local()
        self.largura_var.trace_add('write', lambda *a: atualizar_preview_local())
        self.altura_var.trace_add('write', lambda *a: atualizar_preview_local())
        campos_frame = ttk.Frame(frame_texto)
        campos_frame.pack(fill="x", pady=(0, 8))
        ttk.Label(campos_frame, text="Pavimento:").grid(row=0, column=0, sticky="w", padx=2)
        pavimento_entry = ttk.Entry(campos_frame, textvariable=self.pavimento_var)
        pavimento_entry.grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Label(campos_frame, text="Observações:").grid(row=1, column=0, sticky="w", padx=2)
        obs_entry = ttk.Entry(campos_frame, textvariable=self.obs_var)
        obs_entry.grid(row=1, column=1, sticky="ew", padx=2)
        campos_frame.columnconfigure(1, weight=1)
        msg_display = f"Retângulo Principal: {largura1:.2f} x {altura1:.2f}\n"
        msg_display += f"Retângulo em L: {largura2:.2f} x {altura2:.2f}\n\n"
        msg_display += "Chanfros Principal:\n"
        for nome, valor in analise_combinada['principal']['chanfros'].items():
            msg_display += f"{nome}: {valor:.2f}\n"
        msg_display += "\nAberturas Principal:\n"
        if not analise_combinada['principal']['aberturas']:
            msg_display += "Nenhuma abertura identificada\n"
        else:
            for i, ab in enumerate(analise_combinada['principal']['aberturas'], 1):
                msg_display += f"Abertura {i}:\n"
                msg_display += f"  Profundidade: {ab['altura']:.2f}\n"
                msg_display += f"  Largura: {ab['largura']:.2f}\n"
        msg_display += "\nChanfros L:\n"
        for nome, valor in analise_combinada['painel_l']['chanfros'].items():
            msg_display += f"{nome}: {valor:.2f}\n"
        msg_display += "\nAberturas L:\n"
        if not analise_combinada['painel_l']['aberturas']:
            msg_display += "Nenhuma abertura identificada\n"
        else:
            for i, ab in enumerate(analise_combinada['painel_l']['aberturas'], 1):
                msg_display += f"Abertura {i}:\n"
                msg_display += f"  Profundidade: {ab['altura']:.2f}\n"
                msg_display += f"  Largura: {ab['largura']:.2f}\n"
        txt = tk.Text(frame_texto, height=5, width=35, wrap="word")  
        txt.insert("1.0", msg_display)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True, pady=(0, 0), padx=0)
        frame_botoes = ttk.Frame(frame_texto)
        frame_botoes.pack(fill="x", pady=5)
        def confirmar_e_retornar():
            win.destroy()
            confirmar_callback(True)
        def continuacao_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima_fracao = numero_base + 0.1
            while proxima_fracao in numeros_existentes:
                proxima_fracao += 0.1
            confirmar_callback(True) 
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(f"{proxima_fracao:.1f}")
            self.pavimento_var.set(pavimento_atual)
            self.usar_boundary_duplo = False 
            self._selecionar_textos_consecutivos() 
        def nova_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numeros_existentes = []
                for numero in self.fundos_salvos.keys():
                    try:
                        num = float(numero)
                        if num.is_integer():
                            numeros_existentes.append(int(num))
                    except (ValueError, TypeError):
                        continue
                if numeros_existentes:
                    numero_base = max(numeros_existentes) 
                else:
                    numero_base = 0 
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo_numero = numero_base + 1
            while proximo_numero in numeros_existentes:
                proximo_numero += 1
            confirmar_callback(True) 
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(str(proximo_numero))
            self.pavimento_var.set(pavimento_atual)
            self.usar_boundary_duplo = False 
            self._selecionar_textos_consecutivos() 
        def extensao_l():
            win.destroy()
            confirmar_callback(True) 
            self._salvar_fundo()
            numero_str = self.numero_var.get().strip()
            if not numero_str:
                messagebox.showwarning("Aviso", "O número do fundo deve ser preenchido antes de criar uma extensão em L.")
                self.deiconify() 
                return
            try:
                numero_base = int(float(numero_str))
            except ValueError:
                messagebox.showwarning("Aviso", "O número do fundo não é um valor numérico válido.")
                self.deiconify() 
                return
            proximo_numero = numero_base + 1
            while str(proximo_numero) in self.fundos_salvos:
                proximo_numero += 1
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(str(proximo_numero))
            self.pavimento_var.set(pavimento_atual)
            self.usar_boundary_duplo = True 
            self._executar_boundary_l() 
        def cancelar():
            win.destroy()
            confirmar_callback(False)
        def reiniciar_selecao():
            win.destroy()
            confirmar_callback(False) 
            self._reiniciar_selecao() 
        style = ttk.Style()
        style.configure('Custom.TButton', padding=5)
        btn_confirmar = ttk.Button(frame_botoes, text="1: Confirmar e Retornar", 
                                 command=confirmar_e_retornar, style='Custom.TButton')
        btn_continuacao = ttk.Button(frame_botoes, text="2: Continuação da Viga", 
                                   command=continuacao_viga, style='Custom.TButton')
        btn_nova = ttk.Button(frame_botoes, text="3: Nova Viga", 
                            command=nova_viga, style='Custom.TButton')
        btn_extensao = ttk.Button(frame_botoes, text="4: Extensão em L", 
                                command=extensao_l, style='Custom.TButton')
        btn_cancelar = ttk.Button(frame_botoes, text="5: Cancelar", 
                                 command=cancelar, style='Custom.TButton')
        btn_reiniciar = ttk.Button(frame_botoes, text="6: Reiniciar Seleção", 
                                 command=reiniciar_selecao, style='Custom.TButton')
        btn_confirmar.grid(row=0, column=0, padx=5, pady=2, sticky="ew")
        btn_continuacao.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        btn_nova.grid(row=2, column=0, padx=5, pady=2, sticky="ew")
        btn_extensao.grid(row=3, column=0, padx=5, pady=2, sticky="ew")
        btn_cancelar.grid(row=4, column=0, padx=5, pady=2, sticky="ew")
        btn_reiniciar.grid(row=5, column=0, padx=5, pady=2, sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)
        # Adicionar Radiobuttons para tipo de painel L
        tipo_l_frame = ttk.Frame(frame_texto)
        tipo_l_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(tipo_l_frame, text="Tipo de Painel L:").pack(side=tk.LEFT, padx=5)
        tipos = ["E/T", "E/F", "D/T", "D/F"]
        tipo_painel2_confirm_var = tk.StringVar(value=self.tipo_painel2_var.get())
        def on_tipo_l_change(*args):
            self.tipo_painel2_var.set(tipo_painel2_confirm_var.get())
            atualizar_preview_local()
        tipo_painel2_confirm_var.trace_add('write', on_tipo_l_change)
        self.tipo_painel2_var.trace_add('write', lambda *a: tipo_painel2_confirm_var.set(self.tipo_painel2_var.get()))
        for i, tipo in enumerate(tipos):
            ttk.Radiobutton(tipo_l_frame, text=tipo, variable=tipo_painel2_confirm_var, value=tipo).pack(side=tk.LEFT, padx=5)

        preview_frame = ttk.LabelFrame(frame_texto, text="Visualização", padding=5)
        preview_frame.pack(fill=tk.X, expand=True, pady=(0, 8))
    def _detectar_tipo_l(self, retangulo_principal, retangulo_l):
        if retangulo_principal['dimensoes']['largura'] > retangulo_l['dimensoes']['largura']:
            return "E/T"
        elif retangulo_principal['dimensoes']['largura'] < retangulo_l['dimensoes']['largura']:
            return "E/F"
        else:
            return "E/T"
    def _salvar_teste_e_gerar_script(self):
        try:
            import Robo.Robo_fundos_TASF_limpo_copy_22 as robo_mod
            import importlib
            import tempfile
            import openpyxl
            import os
            importlib.reload(robo_mod)

            # 1. Extrair dados do item selecionado
            dados_fundo = self._coletar_dados()
            print("\n===== DADOS DO FUNDO ENVIADOS PARA O ROBO =====\n")
            for key, value in dados_fundo.items():
                print(f"  {key}: {value}")
            print("\n===== FIM DOS DADOS DO FUNDO =====\n")

            # 2. Mapeamento de linhas do Robo (referência)
            mapeamento_robo = {
                'nome': 53,
                'observacoes': 54,
                'pavimento': 55,
                'largura': 56,
                'altura': 57,
                'texto_esquerda': 58,
                'texto_direita': 59,
                'linha1': 64,
                'linha2': 65,
                'linha3': 66,
                'linha4': 67,
                'linha5': 68,
                'linha6': 69,
                'recuo_topo_esq': 70,
                'recuo_fundo_esq': 71,
                'recuo_topo_dir': 72,
                'recuo_fundo_dir': 73,
                'abertura_esq_topo_dist': 74,
                'abertura_esq_topo_prof': 75,
                'abertura_esq_topo_larg': 76,
                'abertura_esq_fundo_dist': 77,
                'abertura_esq_fundo_prof': 78,
                'abertura_esq_fundo_larg': 79,
                'abertura_dir_topo_dist': 80,
                'abertura_dir_topo_prof': 81,
                'abertura_dir_topo_larg': 82,
                'abertura_dir_fundo_dist': 83,
                'abertura_dir_fundo_prof': 84,
                'abertura_dir_fundo_larg': 85,
                'tipo_painel2': 86,
                'comprimento_2': 87,
                'largura_2': 88,
                'p1_2': 89,
                'p2_2': 90,
                'p3_2': 91,
                'p4_2': 92,
                'p5_2': 93,
                'p6_2': 94,
                'texto_nome_2': 95,
                'texto_esquerda_2': 96,
                'texto_direita_2': 97,
                'recuo_topo_esq_2': 98,
                'recuo_fundo_esq_2': 99,
                'recuo_topo_dir_2': 100,
                'recuo_fundo_dir_2': 101,
                'abertura_esq_topo_dist_2': 102,
                'abertura_esq_topo_prof_2': 103,
                'abertura_esq_topo_larg_2': 104,
                'abertura_esq_fundo_dist_2': 105,
                'abertura_esq_fundo_prof_2': 106,
                'abertura_esq_fundo_larg_2': 107,
                'abertura_dir_topo_dist_2': 108,
                'abertura_dir_topo_prof_2': 109,
                'abertura_dir_topo_larg_2': 110,
                'abertura_dir_fundo_dist_2': 111,
                'abertura_dir_fundo_prof_2': 112,
                'abertura_dir_fundo_larg_2': 113,
                'linha_esquerda_2': 114,
                'linha_direita_2': 115
            }

            # 3. Gerar Excel temporário apenas com os campos em comum (por linha)
            wb = openpyxl.Workbook()
            ws = wb.active
            coluna = 5  # conforme padrão do Robo
            # Campos diretos
            diretos = ['nome', 'observacoes', 'pavimento', 'largura', 'altura', 'texto_esquerda', 'texto_direita',
                        'comprimento_2', 'largura_2', 'texto_nome_2', 'texto_esquerda_2', 'texto_direita_2']
            for campo in diretos:
                if campo in dados_fundo:
                    ws.cell(row=mapeamento_robo[campo], column=coluna, value=dados_fundo[campo])
                elif campo in ['comprimento_2', 'largura_2']:
                    ws.cell(row=mapeamento_robo[campo], column=coluna, value=0.0) # Valor padrão numérico
                elif campo in ['texto_nome_2', 'texto_esquerda_2', 'texto_direita_2']:
                    ws.cell(row=mapeamento_robo[campo], column=coluna, value='') # Valor padrão string

                # Mapeamento de tipo_painel2
                tipo_painel2_map = {'E/T': 1, 'P/T': 2, 'L/T': 3, 'L/L': 4}
                tipo_painel2_valor = tipo_painel2_map.get(dados_fundo.get('tipo_painel2'), 1)
                ws.cell(row=mapeamento_robo['tipo_painel2'], column=coluna, value=tipo_painel2_valor)

                # Mapeamento de paineis (linhas)
                for i in range(6):
                    campo_robo = f'linha{i+1}'
                    if 'paineis' in dados_fundo and i < len(dados_fundo['paineis']):
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=dados_fundo['paineis'][i])
                    else:
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=0.0)

                # Mapeamento de recuos
                recuo_campos_robo = ['recuo_topo_esq', 'recuo_fundo_esq', 'recuo_topo_dir', 'recuo_fundo_dir']
                for i, campo_robo in enumerate(recuo_campos_robo):
                    if 'recuos' in dados_fundo and i < len(dados_fundo['recuos']):
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=dados_fundo['recuos'][i])
                    else:
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=0.0)

                # Mapeamento de aberturas
                abertura_campos_robo = [
                    ('abertura_esq_topo_dist', 'abertura_esq_topo_prof', 'abertura_esq_topo_larg'),
                    ('abertura_esq_fundo_dist', 'abertura_esq_fundo_prof', 'abertura_esq_fundo_larg'),
                    ('abertura_dir_topo_dist', 'abertura_dir_topo_prof', 'abertura_dir_topo_larg'),
                    ('abertura_dir_fundo_dist', 'abertura_dir_fundo_prof', 'abertura_dir_fundo_larg')
                ]
                for i, (dist_c, prof_c, larg_c) in enumerate(abertura_campos_robo):
                    if 'aberturas' in dados_fundo and i < len(dados_fundo['aberturas']):
                        abertura_data = dados_fundo['aberturas'][i]
                        ws.cell(row=mapeamento_robo[dist_c], column=coluna, value=abertura_data[0] if len(abertura_data) > 0 else 0.0)
                        ws.cell(row=mapeamento_robo[prof_c], column=coluna, value=abertura_data[1] if len(abertura_data) > 1 else 0.0)
                        ws.cell(row=mapeamento_robo[larg_c], column=coluna, value=abertura_data[2] if len(abertura_data) > 2 else 0.0)
                    else:
                        ws.cell(row=mapeamento_robo[dist_c], column=coluna, value=0.0)
                        ws.cell(row=mapeamento_robo[prof_c], column=coluna, value=0.0)
                        ws.cell(row=mapeamento_robo[larg_c], column=coluna, value=0.0)

                # Mapeamento de pX_2
                for i in range(1, 7):
                    campo_robo = f'p{i}_2'
                    ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=dados_fundo.get(campo_robo, 0.0))

                # Mapeamento de recuos_2
                recuo_2_campos_robo = ['recuo_topo_esq_2', 'recuo_fundo_esq_2', 'recuo_topo_dir_2', 'recuo_fundo_dir_2']
                for i, campo_robo in enumerate(recuo_2_campos_robo):
                    if 'recuos_2' in dados_fundo and i < len(dados_fundo['recuos_2']):
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=dados_fundo['recuos_2'][i])
                    else:
                        ws.cell(row=mapeamento_robo[campo_robo], column=coluna, value=0.0)

                # Mapeamento de aberturas_2
                abertura_2_campos_robo = [
                    ('abertura_esq_topo_dist_2', 'abertura_esq_topo_prof_2', 'abertura_esq_topo_larg_2'),
                    ('abertura_esq_fundo_dist_2', 'abertura_esq_fundo_prof_2', 'abertura_esq_fundo_larg_2'),
                    ('abertura_dir_topo_dist_2', 'abertura_dir_topo_prof_2', 'abertura_dir_topo_larg_2'),
                    ('abertura_dir_fundo_dist_2', 'abertura_dir_fundo_prof_2', 'abertura_dir_fundo_larg_2')
                ]
                for i, (dist_c, prof_c, larg_c) in enumerate(abertura_2_campos_robo):
                    if 'aberturas_2' in dados_fundo and i < len(dados_fundo['aberturas_2']):
                        abertura_data = dados_fundo['aberturas_2'][i]
                        ws.cell(row=mapeamento_robo[dist_c], column=coluna, value=abertura_data[0] if len(abertura_data) > 0 else 0.0)
                        ws.cell(row=mapeamento_robo[prof_c], column=coluna, value=abertura_data[1] if len(abertura_data) > 1 else 0.0)
                        ws.cell(row=mapeamento_robo[larg_c], column=coluna, value=abertura_data[2] if len(abertura_data) > 2 else 0.0)
                    else:
                        ws.cell(row=mapeamento_robo[dist_c], column=coluna, value=0.0)
                        ws.cell(row=mapeamento_robo[prof_c], column=coluna, value=0.0)
                        ws.cell(row=mapeamento_robo[larg_c], column=coluna, value=0.0)

                # Mapeamento de sarrafo_esq_2 e sarrafo_dir_2 para linha_esquerda_2 e linha_direita_2
                ws.cell(row=mapeamento_robo['linha_esquerda_2'], column=coluna, value=int(dados_fundo.get('sarrafo_esq_2', True)))
                ws.cell(row=mapeamento_robo['linha_direita_2'], column=coluna, value=int(dados_fundo.get('sarrafo_dir_2', True)))
            # Salvar Excel temporário
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                wb.save(tmp.name)
                tmp_excel_path = tmp.name

            # 4. Chamar função do Robo para processar esse Excel sem abrir diálogo
            from tkinter import filedialog
            original_askopenfilename = filedialog.askopenfilename
            filedialog.askopenfilename = lambda **kwargs: tmp_excel_path

            # --- MOCK DE VARIÁVEIS GLOBAIS DO ROBO ---
            class MockEntry:
                def __init__(self): self.value = ""
                def get(self): return self.value
                def delete(self, *args, **kwargs): self.value = ""
                def insert(self, index, value): self.value = str(value)
            class MockVar:
                def __init__(self, value=None): self._value = value
                def get(self): return self._value
                def set(self, value): self._value = value
            class MockLabel:
                def __init__(self): self.text = ""
                def config(self, text): self.text = text

            class MockTk:
                END = 'end'
                BooleanVar = MockVar
                DoubleVar = MockVar
                IntVar = MockVar
                StringVar = MockVar
                
            class MockTkinterTtk:
                # Mockar classes ttk que são instanciadas ou usadas diretamente
                Entry = MockEntry
                Label = MockLabel # Adicionado para caso resultado_label seja do ttk.Label

            # Garantir que tk e ttk estejam mockados no robo_mod.__dict__
            if 'tk' not in robo_mod.__dict__ or robo_mod.__dict__['tk'] is None:
                robo_mod.__dict__['tk'] = MockTk()
            # Se o Robo usa 'ttk' diretamente (e não 'tkinter.ttk')
            if 'ttk' not in robo_mod.__dict__ or robo_mod.__dict__['ttk'] is None:
                robo_mod.__dict__['ttk'] = MockTkinterTtk()

            # Assegurar que tk.END está acessível no módulo Robo
            if not hasattr(robo_mod, 'tk') or not hasattr(robo_mod.tk, 'END'):
                if 'tk' not in robo_mod.__dict__ or robo_mod.__dict__['tk'] is None: # Se tk ainda não foi mockado
                    robo_mod.__dict__['tk'] = MockTk()
                robo_mod.tk.END = 'end'

            # ... (restante dos mocks como antes)

            # Lista dos campos de entrada esperados pelo Robo
            campos_entry = [
                'nome_entry', 'observacoes_entry', 'pavimento_entry', 'largura_entry', 'altura_entry',
                'texto_esquerda_entry', 'texto_direita_entry', 'linha1_entry', 'linha2_entry', 'linha3_entry',
                'linha4_entry', 'linha5_entry', 'linha6_entry', 'recuo_topo_esq_entry', 'recuo_fundo_esq_entry',
                'recuo_topo_dir_entry', 'recuo_fundo_dir_entry', 'abertura_esq_topo_dist_entry', 'abertura_esq_topo_prof_entry',
                'abertura_esq_topo_larg_entry', 'abertura_esq_fundo_dist_entry', 'abertura_esq_fundo_prof_entry',
                'abertura_esq_fundo_larg_entry', 'abertura_dir_topo_dist_entry', 'abertura_dir_topo_prof_entry',
                'abertura_dir_topo_larg_entry', 'abertura_dir_fundo_dist_entry', 'abertura_dir_fundo_prof_entry',
                'abertura_dir_fundo_larg_entry'
            ]
            campos_var = [
                'tipo_painel2_var', 'comprimento_2_var', 'largura_2_var', 'p1_2_var', 'p2_2_var', 'p3_2_var',
                'p4_2_var', 'p5_2_var', 'p6_2_var', 'linha_esquerda_var', 'linha_direita_var',
                'linha_esquerda_2_var', 'linha_direita_2_var',
                'texto_nome_2_var', 'texto_esquerda_2_var', 'texto_direita_2_var',
                'recuo_topo_esq_2_var', 'recuo_fundo_esq_2_var', 'recuo_topo_dir_2_var', 'recuo_fundo_dir_2_var',
                'abertura_esq_topo_dist_2_var', 'abertura_esq_topo_prof_2_var', 'abertura_esq_topo_larg_2_var',
                'abertura_esq_fundo_dist_2_var', 'abertura_esq_fundo_prof_2_var', 'abertura_esq_fundo_larg_2_var',
                'abertura_dir_topo_dist_2_var', 'abertura_dir_topo_prof_2_var', 'abertura_dir_topo_larg_2_var',
                'abertura_dir_fundo_dist_2_var', 'abertura_dir_fundo_prof_2_var', 'abertura_dir_fundo_larg_2_var',
                'recuos_2_vars', 'aberturas_2_vars'
            ]

            # Mockar entradas (garantir que sejam MockEntry instanciadas)
            for var_name in campos_entry:
                if var_name not in robo_mod.__dict__ or not isinstance(robo_mod.__dict__[var_name], MockEntry):
                    robo_mod.__dict__[var_name] = MockEntry()
            # Mockar variáveis (garantir que sejam MockVar instanciadas)
            for var_name in campos_var:
                if var_name not in robo_mod.__dict__ or not isinstance(robo_mod.__dict__[var_name], MockVar):
                    robo_mod.__dict__[var_name] = MockVar()
            
            # Mockar listas especiais, garantindo que sejam listas de MockVar
            if 'recuos_2_vars' not in robo_mod.__dict__ or not isinstance(robo_mod.__dict__['recuos_2_vars'], list) or len(robo_mod.__dict__['recuos_2_vars']) < 4 or not all(isinstance(v, MockVar) for v in robo_mod.__dict__['recuos_2_vars']):
                robo_mod.__dict__['recuos_2_vars'] = [MockVar() for _ in range(4)]
            if 'aberturas_2_vars' not in robo_mod.__dict__ or not isinstance(robo_mod.__dict__['aberturas_2_vars'], list) or len(robo_mod.__dict__['aberturas_2_vars']) < 4 or not all(isinstance(row, list) and len(row) >=3 and all(isinstance(v, MockVar) for v in row) for row in robo_mod.__dict__['aberturas_2_vars']):
                robo_mod.__dict__['aberturas_2_vars'] = [[MockVar() for _ in range(3)] for _ in range(4)]
            
            # Mockar resultado_label e passá-lo explicitamente
            if 'resultado_label' not in robo_mod.__dict__ or not isinstance(robo_mod.__dict__['resultado_label'], MockLabel):
                robo_mod.__dict__['resultado_label'] = MockLabel() # Garante que o global esteja mockado
            mocked_resultado_label = robo_mod.resultado_label # Usar a instância que está no globals do robo_mod

            # Mockar config_manager se necessário (como antes)
            if 'config_manager' not in robo_mod.__dict__ or robo_mod.__dict__['config_manager'] is None:
                if hasattr(robo_mod, 'ConfigManager'):
                    robo_mod.__dict__['config_manager'] = robo_mod.ConfigManager()
                else:
                    class MockConfigManager:
                        def get_layer(self, name): return "0"
                        def get_comando(self, name): return ""
                        def get_tipo_linha(self): return "PLINE"
                        def save_config(self): return True
                        def load_templates(self, *a, **kw): return {}
                        def save_template(self, *a, **kw): return True
                        def delete_template(self, *a, **kw): return True
                    robo_mod.__dict__['config_manager'] = MockConfigManager()
            # --- FIM DO MOCK ---

            try:
                # Passar o resultado_label mockado explicitamente para ler_excel_e_preencher
                robo_mod.ler_excel_e_preencher(resultado_label=mocked_resultado_label)
            finally:
                filedialog.askopenfilename = original_askopenfilename
                os.unlink(tmp_excel_path)  # Remove o arquivo temporário

            # --- AÇÃO ADICIONAL: Copiar script gerado para comando_FD.scr ---
            # 1. Obter o caminho do script que foi salvo pelo Robo
            # As variáveis nome_entry, observacoes_entry, DIRETORIO_ATUAL foram populadas pelo ler_excel_e_preencher.
            nome_robo = robo_mod.nome_entry.get() if hasattr(robo_mod, 'nome_entry') and robo_mod.nome_entry else "SemNome"
            observacoes_robo = robo_mod.observacoes_entry.get() if hasattr(robo_mod, 'observacoes_entry') and robo_mod.observacoes_entry else ""
            diretorio_robo = robo_mod.DIRETORIO_ATUAL if hasattr(robo_mod, 'DIRETORIO_ATUAL') and robo_mod.DIRETORIO_ATUAL else None

            # Define um diretório padrão se o Robo não conseguiu inicializar DIRETORIO_ATUAL
            if not diretorio_robo:
                # Assumindo que a base é Robo/PASTAS-FUNDOS. Ajuste se necessário.
                base_dir_fallback = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Robo', 'PASTAS-FUNDOS')
                diretorio_robo = os.path.join(base_dir_fallback, dados_fundo.get('pavimento', 'SEM_PAVIMENTO')) # Usa o pavimento do fundo atual como subpasta

            # Garante que o diretório de destino do Robo exista
            os.makedirs(diretorio_robo, exist_ok=True)

            # Constrói o nome do arquivo que o Robo salvou (com base nas entradas)
            source_script_name = f"{nome_robo}{'_' + observacoes_robo if observacoes_robo else ''}.scr"
            source_script_path = os.path.join(diretorio_robo, source_script_name)

            # Caminho do arquivo de destino (comando_FD.scr)
            target_script_path = r"C:\Users\rvene\Desktop\Automacao_cad\Fundos\Robo\comando_FD.scr"

            script_content_to_copy = ""
            if os.path.exists(source_script_path):
                with open(source_script_path, "r", encoding="utf-16") as f:
                    script_content_to_copy = f.read()

                # Sobreescrever o arquivo de destino
                with open(target_script_path, "w", encoding="utf-16") as f:
                    f.write(script_content_to_copy)
                print(f"Script copiado com sucesso para: {target_script_path}")
            else:
                print(f"[AVISO] Script original não encontrado em '{source_script_path}'. Não foi possível copiar para '{target_script_path}'.")
            # --- FIM DA AÇÃO ADICIONAL ---

            messagebox.showinfo("Sucesso", "Script gerado, copiado para comando_FD.scr e executado via Robo com sucesso!")
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            messagebox.showerror("Erro", f"Erro ao gerar script via Robo: {str(e)}")

    def _salvar_pavimento_e_gerar_multiplos_scripts(self):
        import Robo.Robo_fundos_TASF_limpo_copy_22 as robo_mod
        import importlib
        import tempfile
        import openpyxl
        import os

        pavimento_selecionado = self.pavimento_filtro_var.get()
        if not pavimento_selecionado or pavimento_selecionado == "Todos":
            messagebox.showwarning("Aviso", "Selecione um pavimento específico para salvar.")
            return

        output_dir = filedialog.askdirectory(title=f"Selecionar Pasta para Salvar Scripts do Pavimento {pavimento_selecionado}")
        if not output_dir:
            return

        fundos_do_pavimento = {
            num: dados for num, dados in self.fundos_salvos.items()
            if dados.get("pavimento", "") == pavimento_selecionado
        }

        if not fundos_do_pavimento:
            messagebox.showinfo("Info", f"Nenhum fundo encontrado para o pavimento {pavimento_selecionado}.")
            return

        messagebox.showinfo("Salvar Pavimento", "Etapa 1: Imports e validação de pavimento concluídos.")

    def _reiniciar_selecao(self):
        self._selecionar_textos_consecutivos()

def alinhar_boundary_horizontal(coords):
    """
    Alinha o boundary pelo lado maior do retângulo de menor área (minimum area rectangle),
    garantindo que o lado do fundo (base real, mesmo com aberturas ou chanfros) fique para baixo após a rotação.
    """
    import numpy as np
    pontos = np.array([(coords[i], coords[i+1]) for i in range(0, len(coords), 2)])
    cx, cy = pontos.mean(axis=0)
    pontos_c = pontos - [cx, cy]
    segmentos = np.diff(np.vstack([pontos_c, pontos_c[0]]), axis=0)
    angulos = np.arctan2(segmentos[:,1], segmentos[:,0])
    angulos = np.unique(np.mod(angulos, np.pi/2))  
    melhor_area = None
    melhor_ang = 0
    for ang in angulos:
        rot = np.array([[np.cos(-ang), -np.sin(-ang)], [np.sin(-ang), np.cos(-ang)]])
        rot_pontos = pontos_c @ rot.T
        min_x, min_y = rot_pontos.min(axis=0)
        max_x, max_y = rot_pontos.max(axis=0)
        area = (max_x - min_x) * (max_y - min_y)
        if (melhor_area is None) or (area < melhor_area):
            melhor_area = area
            melhor_ang = ang
    rot = np.array([[np.cos(-melhor_ang), -np.sin(-melhor_ang)], [np.sin(-melhor_ang), np.cos(-melhor_ang)]])
    pontos_rot = pontos_c @ rot.T
    min_x, min_y = pontos_rot.min(axis=0)
    max_x, max_y = pontos_rot.max(axis=0)
    largura = max_x - min_x
    altura = max_y - min_y
    if altura > largura:
        rot90 = np.array([[0, -1], [1, 0]])
        pontos_rot = pontos_rot @ rot90.T
        melhor_ang += np.pi/2
    y_min = pontos[:,1].min()
    pontos_fundo_idx = np.where(np.abs(pontos[:,1] - y_min) < 1e-3)[0]
    pontos_fundo_rot = pontos_rot[pontos_fundo_idx]
    if len(pontos_fundo_rot) > 2:
        indices = pontos_fundo_idx
        indices = np.sort(indices)
        grupos = []
        grupo = [indices[0]]
        for i in range(1, len(indices)):
            if indices[i] == indices[i-1] + 1:
                grupo.append(indices[i])
            else:
                grupos.append(grupo)
                grupo = [indices[i]]
        grupos.append(grupo)
        segmentos_rot = np.diff(np.vstack([pontos_rot, pontos_rot[0]]), axis=0)
        soma_grupos = []
        for grupo in grupos:
            soma = 0
            for i in range(len(grupo)-1):
                idx1 = grupo[i]
                idx2 = grupo[i+1]
                soma += np.linalg.norm(pontos_rot[idx2] - pontos_rot[idx1])
            soma_grupos.append(soma)
        idx_base = np.argmax(soma_grupos)
        grupo_base = grupos[idx_base]
        y_base = pontos_rot[grupo_base,1].mean()
        if y_base > pontos_rot[:,1].mean():
            print('[INFO] Invertendo eixo Y para manter a base (abertura ou múltiplos segmentos) para baixo.')
            pontos_rot[:,1] *= -1
    else:
        segmentos_rot = np.diff(np.vstack([pontos_rot, pontos_rot[0]]), axis=0)
        comprimentos = np.linalg.norm(segmentos_rot, axis=1)
        idxs_maiores = np.argsort(comprimentos)[-2:]  
        medias_y = []
        for idx in idxs_maiores:
            y1 = pontos_rot[idx][1]
            y2 = pontos_rot[(idx+1)%len(pontos_rot)][1]
            medias_y.append((y1 + y2) / 2)
        idx_fundo = idxs_maiores[np.argmin(medias_y)]
        y_fundo = medias_y[np.argmin(medias_y)]
        if y_fundo > pontos_rot[:,1].mean():
            print('[INFO] Invertendo eixo Y para garantir maior segmento mais abaixo na base (caso chanfro).')
            pontos_rot[:,1] *= -1
    pontos_final = pontos_rot + [cx, cy]
    coords_final = []
    for x, y in pontos_final:
        coords_final.extend([float(x), float(y)])
    return coords_final
if __name__ == "__main__":
    app = FundoProducaoApp()
    def cancelar_tudo(event=None):
        print('[ESC] Cancelamento global solicitado!')
        if hasattr(app, '_balcao_flutuante') and app._balcao_flutuante is not None:
            try:
                app._balcao_flutuante.destroy()
            except:
                pass
            app._balcao_flutuante = None
        for w in app.winfo_children():
            if isinstance(w, tk.Toplevel):
                try:
                    w.destroy()
                except:
                    pass
        if hasattr(app, '_cancelar_selecao_interno'):
            try:
                app._cancelar_selecao_interno()
            except:
                pass
        try:
            app.lift()
            app.focus_force()
        except:
            pass
        print('[ESC] Seleção cancelada, interface mantida.')
    def esc_listener():
        while True:
            keyboard.wait('esc')
            app.after(0, cancelar_tudo)
    esc_thread = threading.Thread(target=esc_listener, daemon=True)
    esc_thread.start()
    app.bind_all('<Escape>', cancelar_tudo)
    app.mainloop()