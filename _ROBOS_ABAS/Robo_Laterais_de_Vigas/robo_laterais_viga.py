import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import os
import pickle
import openpyxl
import win32com.client
import pythoncom
import win32gui
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import time
import re
import math
import pyautogui
import keyboard
import sys
from gerador_script_viga import gerar_script_viga
class FundoProducaoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Robô das Laterais de Viga - Gerenciador")
        self.geometry("1050x800")
        self.configure(bg='#f0f0f0')
        self.fundos_salvos = {}
        self.combinacoes = []
        self.excel_file = None
        self.ac = None
        self._preview_cache = {}  
        self._last_update = 0  
        self._update_delay = 100  
        self._update_scheduled = None  
        self._carregando_fundo = False  
        self.protocol("WM_DELETE_WINDOW", self._ao_fechar)
        self.tipo_distribuicao_var = tk.StringVar(value="122")  
        self.tipo_painel_inicial_var = tk.StringVar(value="300")  
        self.numero_var = tk.StringVar()
        self.nome_var = tk.StringVar()
        self.obs_var = tk.StringVar()
        self.pavimento_var = tk.StringVar()
        self.largura_var = tk.DoubleVar()
        self.altura_geral_var = tk.StringVar()
        self.texto_esq_var = tk.StringVar()
        self.texto_dir_var = tk.StringVar()
        self.paineis_larguras_vars = [tk.DoubleVar() for _ in range(6)]
        self.paineis_alturas_vars = [tk.StringVar() for _ in range(6)]
        self.paineis_alturas2_vars = [tk.StringVar() for _ in range(6)]
        self.aberturas_vars = [[tk.DoubleVar() for _ in range(3)] for _ in range(4)]
        self.sarrafo_esq_var = tk.BooleanVar(value=True)
        self.sarrafo_dir_var = tk.BooleanVar(value=True)
        self.sarrafo_alt2_esq_var = tk.BooleanVar(value=False)
        self.sarrafo_alt2_dir_var = tk.BooleanVar(value=False)
        self.lajes_sup_vars = [tk.DoubleVar() for _ in range(6)]  
        self.lajes_inf_vars = [tk.DoubleVar() for _ in range(6)]  
        self.lajes_central_alt_vars = [tk.DoubleVar() for _ in range(6)]  
        self.detalhe_pilar_esq_vars = [tk.DoubleVar(), tk.DoubleVar()]
        self.detalhe_pilar_dir_vars = [tk.DoubleVar(), tk.DoubleVar()]
        self.cores = {
            'fundo': "#FFFFFF",
            'painel': "#E3F2FD",
            'sarrafo': "#FF9800",
            'abertura': "#FFEB3B",
            'divisao': "#000000",
            'cota': "#2196F3",
            'texto': "#000000"
        }
        self.altura_geral_var.trace_add('write', self._atualizar_alturas_paineis)
        self.altura_2_geral_var = tk.StringVar()
        self.altura_2_geral_var.trace_add('write', self._atualizar_alturas2_paineis)
        self.laje_sup_universal = tk.DoubleVar()
        self.laje_inf_universal = tk.DoubleVar()
        self.laje_central_alt_universal = tk.DoubleVar()
        self.nivel_oposto_var = tk.StringVar()
        self.nivel_viga_var = tk.StringVar()
        self.nivel_pe_direito_var = tk.StringVar()
        self.ajuste_var = tk.StringVar()
        self.paineis_tipo1_vars = [tk.StringVar(value="Sarrafeado") for _ in range(6)]  
        self.paineis_tipo2_vars = [tk.StringVar(value="Sarrafeado") for _ in range(6)]  
        self.tipo1_universal_var = tk.StringVar(value="Sarrafeado")
        self.tipo2_universal_var = tk.StringVar(value="Sarrafeado")
        self.paineis_grade_altura1_vars = [tk.DoubleVar(value=7.0) for _ in range(6)]
        self.paineis_grade_altura2_vars = [tk.DoubleVar(value=7.0) for _ in range(6)]
        self.grade_altura1_universal = tk.DoubleVar(value=7.0)
        self.grade_altura2_universal = tk.DoubleVar(value=7.0)
        def sync_grade_altura1(*args):
            try:
                valor = float(self.grade_altura1_universal.get())
                for v in self.paineis_grade_altura1_vars:
                    v.set(valor)
            except ValueError:
                pass
        self.grade_altura1_universal.trace_add('write', sync_grade_altura1)
        self.nivel_oposto_var.trace_add('write', self._atualizar_altura2_universal_por_niveis)
        self.nivel_viga_var.trace_add('write', self._atualizar_altura2_universal_por_niveis)
        self.continuacao_var = tk.StringVar(value="Proxima Parte")
        self.combinacoes = []  
        self.combinacao_selecionada = None
        self.checkbox_vars = {}  
        self.modo_combinacao = False
        self.modo_reciclagem = False  
        self.reciclagem_vinculos = {}  
        self.reciclagem_reverse = {}   
        self._setup_ui()
        self._carregar_fundos_salvos()
        self._carregar_combinacoes()  
        self._f8_cancelando = False
        self._sequencia_finalizada = False  
        self._registrar_f8_global()
        self._segmentos_temp_corte = []
        self._sequencia_finalizada = False
        self.nivel_oposto_var.trace_add('write', self._calcular_ajuste)
        self.nivel_viga_var.trace_add('write', self._calcular_ajuste)
        self.nivel_pe_direito_var.trace_add('write', self._calcular_ajuste)
    def _setup_ui(self):
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        preview_frame = ttk.LabelFrame(main_frame, text="Visualização", padding=2)
        preview_frame.pack(fill=tk.X, padx=5, pady=2)
        self._criar_preview(preview_frame)
        botoes_frame = ttk.Frame(main_frame)
        botoes_frame.pack(fill=tk.X, padx=5, pady=2)
        self._criar_botoes(botoes_frame)
        self.content_container = ttk.Frame(main_frame)
        self.content_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)
        self.canvas_scroll = tk.Canvas(self.content_container)
        scrollbar_y = ttk.Scrollbar(self.content_container, orient="vertical", command=self.canvas_scroll.yview)
        scrollbar_x = ttk.Scrollbar(self.content_container, orient="horizontal", command=self.canvas_scroll.xview)
        self.content_frame = ttk.Frame(self.canvas_scroll)
        self.canvas_scroll.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        self.canvas_scroll.pack(side="left", fill="both", expand=True)
        self.canvas_frame = self.canvas_scroll.create_window((0, 0), window=self.content_frame, anchor="nw")
        self.content_frame.bind("<Configure>", lambda e: self.canvas_scroll.configure(scrollregion=self.canvas_scroll.bbox("all")))
        def _on_mousewheel(event):
            if event.state & 0x4:  
                return
            self.canvas_scroll.yview_scroll(int(-1*(event.delta/120)), "units")
        self.canvas_scroll.bind_all("<MouseWheel>", _on_mousewheel)
        self.canvas_scroll.bind_all("<Button-4>", lambda e: self.canvas_scroll.yview_scroll(-1, "units"))
        self.canvas_scroll.bind_all("<Button-5>", lambda e: self.canvas_scroll.yview_scroll(1, "units"))
        self.content_frame.columnconfigure(0, weight=0, minsize=110)  
        self.content_frame.columnconfigure(1, weight=0, minsize=170)  
        self.content_frame.columnconfigure(2, weight=1, minsize=320)  
        self.content_frame.rowconfigure(0, weight=1)
        self._criar_lista_combinacoes(self.content_frame)
        self._criar_lista_fundos(self.content_frame)
        self._criar_aba_dados_gerais(self.content_frame)
        self.comb_frame.grid_remove()
        self.listas_duplas_frame.grid_remove()
        self._atualizar_layout_listas()
    def _atualizar_layout_listas(self):
        col = 0
        if self.modo_combinacao:
            self.comb_frame.grid(row=0, column=col, sticky="nsew", padx=(0, 2))
            col += 1
        else:
            self.comb_frame.grid_remove()
        self.lista1_frame.pack_forget()
        self.lista1_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 2))
        self.lista_fundos1.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.listas_duplas_frame.grid(row=0, column=col, sticky="nsew", padx=(0, 2))
        if self.modo_reciclagem:
            self.lista2_frame.pack_forget()
            self.lista2_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(2, 0))
            self.lista_fundos2.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        else:
            self.lista2_frame.pack_forget()
        col += 1
        self.dados_frame.grid(row=0, column=col, sticky="nsew", padx=(8, 0))
        for i in range(col+1, 3):
            self.content_frame.grid_columnconfigure(i, minsize=0)
        for i in range(col+1):
            self.content_frame.grid_columnconfigure(i, minsize=110 if i==0 else 170 if i==1 else 320)
        self.content_frame.update_idletasks()
        self.canvas_scroll.configure(scrollregion=self.canvas_scroll.bbox("all"))
    def _schedule_preview_update(self, *args):
        """Agenda uma atualização do preview com debounce."""
        if self._update_scheduled is not None:
            self.after_cancel(self._update_scheduled)
        self._update_scheduled = self.after(100, self._atualizar_preview)
    def vincular_eventos_alteracao(self):
        """Vincula os eventos de alteração aos campos."""
        for var in [self.largura_var, self.altura_geral_var, *self.paineis_larguras_vars, *self.paineis_alturas_vars, 
                    self.sarrafo_esq_var, self.sarrafo_dir_var]:
            var.trace_add('write', self._schedule_preview_update)
        self.altura_geral_var.trace_add('write', self._atualizar_alturas_paineis)
        for var in self.lajes_sup_vars + self.lajes_inf_vars + self.lajes_central_alt_vars:
            var.trace_add('write', self._schedule_preview_update)
    def _criar_preview(self, parent):
        self.preview_frame = ttk.LabelFrame(parent, text="Visualização", padding=5)
        self.preview_frame.pack(fill=tk.X, expand=True)
        self.label_nome_preview = tk.Label(self.preview_frame, text=self.nome_var.get(), font=("Arial", 16, "bold"))
        self.label_nome_preview.pack(side=tk.TOP, pady=(0, 2))
        def atualizar_nome_preview(*args):
            self.label_nome_preview.config(text=self.nome_var.get())
        self.nome_var.trace_add('write', lambda *a: atualizar_nome_preview())
        self.cores = {
            'fundo': "#FFFFFF",
            'painel': "#E3F2FD",
            'sarrafo': "#FF9800",
            'abertura': "#FFEB3B",
            'chanfro': "#4CAF50",
            'divisao': "#000000",
            'cota': "#2196F3",
            'texto': "#000000"
        }
        self.canvas_frame = ttk.Frame(self.preview_frame)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(self.canvas_frame, width=450, height=220, 
                              bg=self.cores['fundo'],
                              scrollregion=(0, 0, 900, 400))
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.zoom_level = 1.0
        self.pan_start = None
        self.canvas.bind("<ButtonPress-1>", self._start_pan)
        self.canvas.bind("<B1-Motion>", self._pan)
        self.canvas.bind("<ButtonRelease-1>", self._stop_pan)
        self.canvas.bind("<MouseWheel>", self._mouse_wheel)
        self.canvas.bind("<Button-4>", self._mouse_wheel)
        self.canvas.bind("<Button-5>", self._mouse_wheel)
        self.preview_frame.update_idletasks()
        self.canvas.update_idletasks()
        self._schedule_preview_update()
    def _mouse_wheel(self, event):
        if event.num == 5 or event.delta < 0:
            self._zoom_out()
        else:
            self._zoom_in()
    def _zoom_in(self):
        new_zoom = min(self.zoom_level + 0.1, 2.0)
        self._update_zoom(new_zoom)
    def _zoom_out(self):
        new_zoom = max(self.zoom_level - 0.1, 0.5)
        self._update_zoom(new_zoom)
    def _update_zoom(self, value):
        try:
            self.zoom_level = float(value)
            self._atualizar_preview()
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        except ValueError:
            pass
    def _start_pan(self, event):
        self.canvas.scan_mark(event.x, event.y)
        self.pan_start = (event.x, event.y)
    def _pan(self, event):
        if self.pan_start:
            self.canvas.scan_dragto(event.x, event.y, gain=1)
            self.pan_start = (event.x, event.y)
    def _stop_pan(self, event):
        self.pan_start = None
    def _atualizar_preview(self, *args):
        """Atualiza o preview do painel."""
        self._update_scheduled = None  
        if not hasattr(self, 'canvas') or not self.canvas.winfo_exists():
            return
        try:
            self.canvas.delete("all")
            try:
                largura = self._float_safe(self.largura_var)
                altura_geral = self._float_safe(self.altura_geral_var)
            except Exception:
                return
            if largura <= 0 or altura_geral <= 0:
                return
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            if canvas_width <= 1 or canvas_height <= 1:
                self.canvas.update_idletasks()
                canvas_width = self.canvas.winfo_width()
                canvas_height = self.canvas.winfo_height()
                if canvas_width <= 1 or canvas_height <= 1:
                    self._schedule_preview_update()
                    return
            escala_x = 400 / (largura * 1.2) * self.zoom_level
            escala_y = 200 / (altura_geral * 1.2) * self.zoom_level
            escala = min(escala_x, escala_y)
            x_inicial = 50 * self.zoom_level
            y_inicial = 250 * self.zoom_level
            self._desenhar_elementos(self.canvas, x_inicial, y_inicial, largura, altura_geral, escala)
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        except Exception as e:
            print(f"Erro ao atualizar preview: {str(e)}")
            self._schedule_preview_update()
    def _desenhar_elementos(self, canvas, x_inicial, y_inicial, largura, altura_geral, escala):
        """Desenha todos os elementos do preview no canvas especificado."""
        canvas.create_line(
            x_inicial, y_inicial,
            x_inicial + largura * escala, y_inicial,
            fill=self.cores['cota'], dash=(4, 4)
        )
        x_atual = x_inicial
        segmentos = []  
        primeiro_painel = True
        ultimo_painel_idx = -1
        paineis_validos = []
        for i in range(6):
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            if largura_painel > 0:
                ultimo_painel_idx = i
                paineis_validos.append(i)
        if not paineis_validos:
            return
        for i in range(6):
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            altura1_str = self.paineis_alturas_vars[i].get()
            if largura_painel > 0 and altura1_str:
                if '+' in altura1_str:
                    partes = altura1_str.split('+')
                    try:
                        h1 = float(partes[0].strip())
                        y_tracejada = y_inicial - h1 * escala
                        canvas.create_line(
                            x_atual, y_tracejada,
                            x_atual + largura_painel * escala, y_tracejada,
                            fill='black', dash=(4, 2), width=2
                        )
                    except Exception:
                        pass
            x_atual += largura_painel * escala
        x_atual = x_inicial
        for i in range(6):
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            altura1 = self._float_safe(self.paineis_alturas_vars[i])
            altura2 = self._float_safe(self.paineis_alturas2_vars[i])
            laje_sup = self._float_safe(self.lajes_sup_vars[i])
            laje_inf = self._float_safe(self.lajes_inf_vars[i])
            laje_c_alt = self._float_safe(self.lajes_central_alt_vars[i])
            altura_painel = altura1 + laje_c_alt + altura2 + laje_sup + laje_inf
            is_primeiro = (i == paineis_validos[0])
            is_ultimo = (i == ultimo_painel_idx)
            if largura_painel > 0 and (altura1 + altura2) > 0:
                if i == paineis_validos[0]:
                    segmentos = [
                        ("Laje Inf.", laje_inf),
                        ("Altura 1", altura1),
                        ("Laje Central", laje_c_alt),
                        ("Altura 2", altura2),
                        ("Laje Sup.", laje_sup)
                    ]
                    x_cota = x_inicial - 35
                    self._desenhar_cotas_verticais_segmentos(
                        canvas, x_cota, y_inicial, segmentos, escala, lado='esq'
                    )
                if i == ultimo_painel_idx:
                    segmentos = [
                        ("Laje Inf.", laje_inf),
                        ("Altura 1", altura1),
                        ("Laje Central", laje_c_alt),
                        ("Altura 2", altura2),
                        ("Laje Sup.", laje_sup)
                    ]
                    x_cota = x_atual + largura_painel * escala + 35
                    self._desenhar_cotas_verticais_segmentos(
                        canvas, x_cota, y_inicial, segmentos, escala, lado='dir'
                    )
                    x_cota_total = x_cota + 35
                    segmentos_total = [("Altura Total", altura_painel)]
                    self._desenhar_cotas_verticais_segmentos(
                        canvas, x_cota_total, y_inicial, segmentos_total, escala, lado='dir'
                    )
                if laje_sup > 0:
                    self._desenhar_laje_suave(
                        canvas,
                        x_atual,
                        y_inicial - altura_painel * escala,
                        largura_painel * escala, laje_sup * escala
                    )
                laje_inf = self._float_safe(self.lajes_inf_vars[i])
                if laje_inf > 0:
                    self._desenhar_laje_suave(
                        canvas, x_atual, y_inicial,
                        largura_painel * escala, laje_inf * escala
                    )
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[i])
                if laje_c_alt > 0:
                    y_laje_central = y_inicial - altura1 * escala
                    self._desenhar_laje_suave(
                        canvas,
                        x_atual,
                        y_laje_central - laje_c_alt * escala,
                        largura_painel * escala,
                        laje_c_alt * escala,
                        fill="#888888"
                    )
                if altura1 > 0:
                    y_topo_alt1 = y_inicial - altura1 * escala
                    canvas.create_rectangle(
                        x_atual, y_inicial,
                        x_atual + largura_painel * escala, y_topo_alt1,
                        outline=self.cores['painel'], width=2, fill="#E3F2FD"
                    )
                    if self.paineis_tipo1_vars[i].get() == "Sarrafeado":
                        altura_sarr = 7 * escala
                        y_sarr_top = y_topo_alt1
                        y_sarr_bot = y_topo_alt1 + altura_sarr
                        canvas.create_rectangle(
                            x_atual, y_sarr_top,
                            x_atual + largura_painel * escala, y_sarr_bot,
                            fill="#FF9800", outline="#FF9800"
                        )
                    elif self.paineis_tipo1_vars[i].get() == "Grade":
                        altura_grade = 2.2 * escala
                        y_grade_top = y_topo_alt1
                        y_grade_bot = y_topo_alt1 + altura_grade
                        x_grade_ini = x_atual + (15 * escala if is_primeiro else 0)
                        x_grade_fim = x_atual + largura_painel * escala - (15 * escala if is_ultimo else 0)
                        canvas.create_rectangle(
                            x_grade_ini, y_grade_top,
                            x_grade_fim, y_grade_bot,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                        largura_vert = 3.5 * escala
                        altura_vert = self.paineis_grade_altura1_vars[i].get() * escala
                        x_vert_esq = x_atual + (15 * escala if is_primeiro else 0)
                        canvas.create_rectangle(
                            x_vert_esq, y_grade_bot,
                            x_vert_esq + largura_vert, y_grade_bot + altura_vert,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                        x_vert_dir = x_atual + largura_painel * escala - largura_vert
                        if is_ultimo:
                            x_vert_dir = x_atual + largura_painel * escala - 15 * escala - largura_vert
                        canvas.create_rectangle(
                            x_vert_dir, y_grade_bot,
                            x_vert_dir + largura_vert, y_grade_bot + altura_vert,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                if altura2 > 0:
                    y_base_alt2 = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_base_alt2 -= laje_c_alt * escala
                    y_topo_alt2 = y_base_alt2 - altura2 * escala
                    canvas.create_rectangle(
                        x_atual, y_base_alt2,
                        x_atual + largura_painel * escala, y_topo_alt2,
                        outline=self.cores['painel'], width=2, fill="#E3F2FD"
                    )
                    if self.paineis_tipo2_vars[i].get() == "Sarrafeado":
                        altura_sarr2 = 7 * escala
                        y_sarr2_top = y_topo_alt2
                        y_sarr2_bot = y_topo_alt2 + altura_sarr2
                        canvas.create_rectangle(
                            x_atual, y_sarr2_top,
                            x_atual + largura_painel * escala, y_sarr2_bot,
                            fill="#FF9800", outline="#FF9800"
                        )
                    elif self.paineis_tipo2_vars[i].get() == "Grade":
                        altura_grade2 = 2.2 * escala
                        y_grade2_top = y_topo_alt2
                        y_grade2_bot = y_topo_alt2 + altura_grade2
                        x_grade2_ini = x_atual + (15 * escala if is_primeiro else 0)
                        x_grade2_fim = x_atual + largura_painel * escala - (15 * escala if is_ultimo else 0)
                        canvas.create_rectangle(
                            x_grade2_ini, y_grade2_top,
                            x_grade2_fim, y_grade2_bot,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                        largura_vert2 = 3.5 * escala
                        altura_vert2 = self.paineis_grade_altura2_vars[i].get() * escala
                        x_vert2_esq = x_atual + (15 * escala if is_primeiro else 0)
                        canvas.create_rectangle(
                            x_vert2_esq, y_grade2_bot,
                            x_vert2_esq + largura_vert2, y_grade2_bot + altura_vert2,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                        x_vert2_dir = x_atual + largura_painel * escala - largura_vert2
                        if is_ultimo:
                            x_vert2_dir = x_atual + largura_painel * escala - 15 * escala - largura_vert2
                        canvas.create_rectangle(
                            x_vert2_dir, y_grade2_bot,
                            x_vert2_dir + largura_vert2, y_grade2_bot + altura_vert2,
                            fill="#4CAF50", outline="#4CAF50"
                        )
                if altura2 > 0 and altura1 > 0:
                    y_divisao = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_divisao -= laje_c_alt * escala
                    canvas.create_line(
                        x_atual, y_divisao,
                        x_atual + largura_painel * escala, y_divisao,
                        fill="#000", width=2
                    )
                if i > 0:
                    y_topo = y_inicial - (altura1 + laje_c_alt + altura2 + laje_sup) * escala
                    y_base = y_inicial + laje_inf * escala
                    canvas.create_line(
                        x_atual, y_base,
                        x_atual, y_topo,
                        fill=self.cores['divisao'], width=2
                    )
                x_atual += largura_painel * escala
        if self.sarrafo_esq_var.get():
            idx_primeiro = next((i for i in range(6) if self._float_safe(self.paineis_larguras_vars[i]) > 0), None)
            if idx_primeiro is not None:
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_primeiro])
                altura2 = self._float_safe(self.paineis_alturas2_vars[idx_primeiro])
                laje_sup = self._float_safe(self.lajes_sup_vars[idx_primeiro])
                laje_inf = self._float_safe(self.lajes_inf_vars[idx_primeiro])
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[idx_primeiro])
                y_topo_alt1 = y_inicial - altura1 * escala
                y_base_alt1 = y_inicial
                y_topo_alt2 = y_base_alt2 = None
                if altura2 > 0:
                    y_base_alt2 = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_base_alt2 -= laje_c_alt * escala
                    y_topo_alt2 = y_base_alt2 - altura2 * escala
                largura_sarrafo = 7 * escala
                x_sarrafo1 = x_inicial
                if self.paineis_tipo1_vars[idx_primeiro].get() == "Grade":
                    x_sarrafo1 = x_inicial + 7 * escala
                canvas.create_rectangle(
                    x_sarrafo1, y_base_alt1,
                    x_sarrafo1 + largura_sarrafo, y_topo_alt1,
                    fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                )
                if altura2 > 0 and y_base_alt2 is not None and y_topo_alt2 is not None:
                    x_sarrafo2 = x_inicial
                    if self.paineis_tipo2_vars[idx_primeiro].get() == "Grade":
                        x_sarrafo2 = x_inicial + 7 * escala
                    canvas.create_rectangle(
                        x_sarrafo2, y_base_alt2,
                        x_sarrafo2 + largura_sarrafo, y_topo_alt2,
                        fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                    )
        if self.sarrafo_alt2_esq_var.get():
            idx_primeiro = next((i for i in range(6) if self._float_safe(self.paineis_larguras_vars[i]) > 0), None)
            if idx_primeiro is not None:
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_primeiro])
                altura2 = self._float_safe(self.paineis_alturas2_vars[idx_primeiro])
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[idx_primeiro])
                if altura2 > 0:
                    y_base_alt2 = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_base_alt2 -= laje_c_alt * escala
                    y_topo_alt2 = y_base_alt2 - altura2 * escala
                    largura_sarrafo = 7 * escala
                    x_sarrafo2 = x_inicial
                    canvas.create_rectangle(
                        x_sarrafo2, y_base_alt2,
                        x_sarrafo2 + largura_sarrafo, y_topo_alt2,
                        fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                    )
        if self.sarrafo_dir_var.get():
            idx_ultimo = None
            x_dir = x_inicial
            for i in range(6):
                if self._float_safe(self.paineis_larguras_vars[i]) > 0:
                    idx_ultimo = i
            if idx_ultimo is not None:
                for i in range(idx_ultimo+1):
                    x_dir += self._float_safe(self.paineis_larguras_vars[i]) * escala
                x_dir -= self._float_safe(self.paineis_larguras_vars[idx_ultimo]) * escala
                largura_painel = self._float_safe(self.paineis_larguras_vars[idx_ultimo])
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_ultimo])
                altura2 = self._float_safe(self.paineis_alturas2_vars[idx_ultimo])
                laje_sup = self._float_safe(self.lajes_sup_vars[idx_ultimo])
                laje_inf = self._float_safe(self.lajes_inf_vars[idx_ultimo])
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[idx_ultimo])
                y_topo_alt1 = y_inicial - altura1 * escala
                y_base_alt1 = y_inicial
                y_topo_alt2 = y_base_alt2 = None
                if altura2 > 0:
                    y_base_alt2 = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_base_alt2 -= laje_c_alt * escala
                    y_topo_alt2 = y_base_alt2 - altura2 * escala
                largura_sarrafo = 7 * escala
                x_sarrafo1 = x_dir + largura_painel * escala - largura_sarrafo
                if self.paineis_tipo1_vars[idx_ultimo].get() == "Grade":
                    x_sarrafo1 = x_dir + largura_painel * escala - 7 * escala - largura_sarrafo
                canvas.create_rectangle(
                    x_sarrafo1, y_base_alt1,
                    x_sarrafo1 + largura_sarrafo, y_topo_alt1,
                    fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                )
                if altura2 > 0 and y_base_alt2 is not None and y_topo_alt2 is not None:
                    x_sarrafo2 = x_dir + largura_painel * escala - largura_sarrafo
                    if self.paineis_tipo2_vars[idx_ultimo].get() == "Grade":
                        x_sarrafo2 = x_dir + largura_painel * escala - 7 * escala - largura_sarrafo
                    canvas.create_rectangle(
                        x_sarrafo2, y_base_alt2,
                        x_sarrafo2 + largura_sarrafo, y_topo_alt2,
                        fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                    )
        if self.sarrafo_alt2_dir_var.get():
            idx_ultimo = None
            x_dir = x_inicial
            for i in range(6):
                if self._float_safe(self.paineis_larguras_vars[i]) > 0:
                    idx_ultimo = i
            if idx_ultimo is not None:
                for i in range(idx_ultimo+1):
                    x_dir += self._float_safe(self.paineis_larguras_vars[i]) * escala
                x_dir -= self._float_safe(self.paineis_larguras_vars[idx_ultimo]) * escala
                largura_painel = self._float_safe(self.paineis_larguras_vars[idx_ultimo])
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_ultimo])
                altura2 = self._float_safe(self.paineis_alturas2_vars[idx_ultimo])
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[idx_ultimo])
                if altura2 > 0:
                    y_base_alt2 = y_inicial - altura1 * escala
                    if laje_c_alt > 0:
                        y_base_alt2 -= laje_c_alt * escala
                    y_topo_alt2 = y_base_alt2 - altura2 * escala
                    largura_sarrafo = 7 * escala
                    x_sarrafo2 = x_dir + largura_painel * escala - largura_sarrafo
                    canvas.create_rectangle(
                        x_sarrafo2, y_base_alt2,
                        x_sarrafo2 + largura_sarrafo, y_topo_alt2,
                        fill=self.cores['sarrafo'], outline=self.cores['sarrafo']
                    )
        self._desenhar_aberturas(canvas, x_inicial, y_inicial, largura, altura_geral, escala)
        self._desenhar_textos(canvas, x_inicial, y_inicial, largura, altura_geral, escala)
        dist_pilar_esq = self._float_safe(self.detalhe_pilar_esq_vars[0])
        larg_pilar_esq = self._float_safe(self.detalhe_pilar_esq_vars[1])
        if larg_pilar_esq > 0:  
            idx_primeiro = -1
            for i in range(6):
                if self._float_safe(self.paineis_larguras_vars[i]) > 0:
                    idx_primeiro = i
                    break
            if idx_primeiro != -1:
                x_painel_esq = x_inicial
                for i in range(idx_primeiro):
                    x_painel_esq += self._float_safe(self.paineis_larguras_vars[i]) * escala
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_primeiro])
                y_base = y_inicial
                y_topo = y_inicial - altura1 * escala
                canvas.create_rectangle(
                    x_painel_esq + dist_pilar_esq * escala, y_topo,
                    x_painel_esq + (dist_pilar_esq + larg_pilar_esq) * escala, y_base,
                    fill='#FF69B4', outline='#FF69B4', stipple='gray25'
                )
        dist_pilar_dir = self._float_safe(self.detalhe_pilar_dir_vars[0])
        larg_pilar_dir = self._float_safe(self.detalhe_pilar_dir_vars[1])
        if larg_pilar_dir > 0:  
            idx_ultimo = -1
            x_painel_dir = x_inicial
            for i in range(6):
                if self._float_safe(self.paineis_larguras_vars[i]) > 0:
                    idx_ultimo = i
            if idx_ultimo != -1:
                for i in range(idx_ultimo+1):
                    x_painel_dir += self._float_safe(self.paineis_larguras_vars[i]) * escala
                altura1 = self._float_safe(self.paineis_alturas_vars[idx_ultimo])
                y_base = y_inicial
                y_topo = y_inicial - altura1 * escala
                canvas.create_rectangle(
                    x_painel_dir - dist_pilar_dir * escala - larg_pilar_dir * escala, y_topo,
                    x_painel_dir - dist_pilar_dir * escala, y_base,
                    fill='#FF69B4', outline='#FF69B4', stipple='gray25'
                )
        if getattr(self, 'continuacao_var', None) and self.continuacao_var.get() == "Obstaculo":
            if ultimo_painel_idx != -1:
                x_obs_ini = x_inicial
                for i in range(ultimo_painel_idx+1):
                    x_obs_ini += self._float_safe(self.paineis_larguras_vars[i]) * escala
                largura_obs = 30 * escala
                altura1 = self._float_safe(self.paineis_alturas_vars[ultimo_painel_idx])
                altura2 = self._float_safe(self.paineis_alturas2_vars[ultimo_painel_idx])
                laje_sup = self._float_safe(self.lajes_sup_vars[ultimo_painel_idx])
                laje_inf = self._float_safe(self.lajes_inf_vars[ultimo_painel_idx])
                laje_c_alt = self._float_safe(self.lajes_central_alt_vars[ultimo_painel_idx])
                altura_total = altura1 + altura2 + laje_sup + laje_inf + laje_c_alt
                y_obs_top = y_inicial - (altura1 + laje_c_alt + altura2 + laje_sup) * escala
                y_obs_base = y_inicial + laje_inf * escala
                canvas.create_rectangle(
                    x_obs_ini, y_obs_base,
                    x_obs_ini + largura_obs, y_obs_top,
                    fill="#888888", outline="#888888"
                )
        x_atual = x_inicial
        y_cota_paineis = y_inicial + 30  
        paineis_larguras = [self._float_safe(self.paineis_larguras_vars[i]) for i in range(6)]
        for i, largura_painel in enumerate(paineis_larguras):
            if largura_painel > 0:
                x_fim = x_atual + largura_painel * escala
                canvas.create_line(x_atual, y_cota_paineis, x_fim, y_cota_paineis, fill=self.cores['cota'], arrow="both")
                canvas.create_text((x_atual + x_fim) / 2, y_cota_paineis + 12, text=f"{largura_painel:.1f}", fill=self.cores['texto'])
                x_atual = x_fim
        if sum(paineis_larguras) > 0:
            x_total_ini = x_inicial
            x_total_fim = x_inicial + sum(paineis_larguras) * escala
            y_cota_total = y_cota_paineis + 28
            canvas.create_line(x_total_ini, y_cota_total, x_total_fim, y_cota_total, fill=self.cores['cota'], arrow="both", width=2)
            canvas.create_text((x_total_ini + x_total_fim) / 2, y_cota_total + 12, text=f"Total: {sum(paineis_larguras):.1f}", fill=self.cores['texto'], font=("Arial", 10, "bold"))
    def _desenhar_cotas(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        """Desenha as cotas do painel principal."""
        canvas.create_line(
            x_inicial, y_inicial + 20,
            x_inicial + largura * escala, y_inicial + 20,
            fill=self.cores['cota'], arrow="both"
        )
        canvas.create_text(
            x_inicial + (largura * escala) / 2, y_inicial + 35,
            text=f"Largura: {largura:.1f}",
            fill=self.cores['texto']
        )
        canvas.create_line(
            x_inicial - 20, y_inicial,
            x_inicial - 20, y_inicial - altura * escala,
            fill=self.cores['cota'], arrow="both"
        )
        canvas.create_text(
            x_inicial - 35, y_inicial - (altura * escala) / 2,
            text=f"Altura: {altura:.1f}",
            fill=self.cores['texto'],
            angle=90
        )
    def _desenhar_painel_principal(self, x_inicial, y_inicial, largura, altura, escala):
        """Desenha o painel principal."""
        self.canvas.create_rectangle(
            x_inicial, y_inicial,
            x_inicial + largura * escala, y_inicial - altura * escala,
            outline=self.cores['painel'],
            width=2
        )
    def _desenhar_aberturas(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        """Desenha as aberturas do painel principal nas esquinas corretas dos painéis, considerando lajes e alturas."""
        x_painel_esq = x_inicial
        x_painel_dir = x_inicial + largura * escala
        altura1_esq = self._float_safe(self.paineis_alturas_vars[0])
        altura2_esq = self._float_safe(self.paineis_alturas2_vars[0])
        laje_central_esq = self._float_safe(self.lajes_central_alt_vars[0])
        laje_sup_esq = self._float_safe(self.lajes_sup_vars[0])
        altura1_dir = self._float_safe(self.paineis_alturas_vars[2])
        altura2_dir = self._float_safe(self.paineis_alturas2_vars[2])
        laje_central_dir = self._float_safe(self.lajes_central_alt_vars[2])
        laje_sup_dir = self._float_safe(self.lajes_sup_vars[2])
        altura_total_esq = altura1_esq + altura2_esq + laje_central_esq + laje_sup_esq
        altura_total_dir = altura1_dir + altura2_dir + laje_central_dir + laje_sup_dir
        y_topo_esq = y_inicial - altura_total_esq * escala
        y_fundo_esq = y_inicial
        y_topo_dir = y_inicial - altura_total_dir * escala
        y_fundo_dir = y_inicial
        for i, linha in enumerate(self.aberturas_vars):
            try:
                valores = [self._float_safe(v) for v in linha]
                print(f"[DEBUG] Abertura {i}: valores = {valores}")
                dist, prof, larg = valores
                if prof <= 0 or larg <= 0:
                    print(f"[DEBUG] Abertura {i} ignorada (profundidade ou largura <= 0)")
                    continue
                forcar_altura1_ativo = self.forcar_altura1_vars[i].get() if hasattr(self, 'forcar_altura1_vars') and i < len(self.forcar_altura1_vars) else False
                if i == 0:  
                    x_pos = x_painel_esq
                    if forcar_altura1_ativo and laje_central_esq > 0:
                        y_pos = y_inicial - altura1_esq * escala - laje_central_esq * escala + dist * escala
                    else:
                        y_pos = y_topo_esq + dist * escala
                elif i == 1:  
                    x_pos = x_painel_esq
                    y_pos = y_fundo_esq - prof * escala - dist * escala
                elif i == 2:  
                    x_pos = x_painel_dir - larg * escala
                    if forcar_altura1_ativo and laje_central_dir > 0:
                        y_pos = y_inicial - altura1_dir * escala - laje_central_dir * escala + dist * escala
                    else:
                        y_pos = y_topo_dir + dist * escala
                elif i == 3:  
                    x_pos = x_painel_dir - larg * escala
                    y_pos = y_fundo_dir - prof * escala - dist * escala
                else:
                    continue
                print(f"[DEBUG] Desenhando abertura {i} em x={x_pos}, y={y_pos}, larg={larg}, prof={prof}")
                canvas.create_rectangle(
                    x_pos, y_pos,
                    x_pos + larg * escala, y_pos + prof * escala,
                    fill=self.cores['abertura'],
                    outline=self.cores['divisao']
                )
                self._desenhar_hachura_abertura(x_pos, y_pos, x_pos + larg * escala, y_pos + prof * escala)
            except Exception as e:
                print(f"[DEBUG] Erro ao desenhar abertura {i}: {e}")
                continue
    def _desenhar_sarrafo(self, x_inicial, y_inicial, largura, altura, escala, lado):
        """Desenha o sarrafo no lado especificado."""
        if lado == 'esquerda':
            x = x_inicial
        else:
            x = x_inicial + largura * escala
        self.canvas.create_line(
            x, y_inicial,
            x, y_inicial - altura * escala,
            fill=self.cores['sarrafo'],
            width=2
        )
    def _desenhar_textos(self, canvas, x_inicial, y_inicial, largura, altura, escala):
        """Desenha os textos do painel, incluindo a área útil em m² ao lado do nome."""
        area_total = 0
        for i in range(6):  
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            altura1 = self._float_safe(self.paineis_alturas_vars[i])
            altura2 = self._float_safe(self.paineis_alturas2_vars[i])
            if largura_painel > 0:
                if altura1 > 0:
                    area_total += largura_painel * altura1
                if altura2 > 0:
                    area_total += largura_painel * altura2
        area_aberturas = 0
        area_sobreposta = 0
        rx1 = x_inicial
        ry1 = y_inicial - altura * escala
        rx2 = x_inicial + largura * escala
        ry2 = y_inicial
        x_painel_esq = x_inicial
        x_painel_dir = x_inicial + largura * escala
        for i, linha in enumerate(self.aberturas_vars):
            try:
                dist, prof, larg = [self._float_safe(v) for v in linha]
                if prof > 0 and larg > 0:
                    area_abertura = prof * larg
                    area_aberturas += area_abertura
                    if i < 2:  
                        x_pos = x_painel_esq
                        y_pos = y_inicial - prof * escala - dist * escala
                    else:  
                        x_pos = x_painel_dir - larg * escala
                        y_pos = y_inicial - prof * escala - dist * escala
                    ax1 = x_pos
                    ay1 = y_pos
                    ax2 = x_pos + larg * escala
                    ay2 = y_pos + prof * escala
                    area_sobrep = self._area_intersecao(rx1, ry1, rx2, ry2, ax1, ay1, ax2, ay2) / (escala * escala)
                    area_sobreposta += area_sobrep
                    print(f"\nAbertura {i+1}:")
                    print(f"Posição: x={x_pos-x_inicial:.1f}, y={y_inicial-y_pos:.1f}")
                    print(f"Dimensões: {larg:.1f}x{prof:.1f} cm")
                    print(f"Área total: {area_abertura/10000:.4f} m²")
                    print(f"Área sobreposta: {area_sobrep/10000:.4f} m²")
            except Exception as e:
                print(f"Erro ao calcular abertura {i+1}: {str(e)}")
                continue
        area_util_m2 = max(0, (area_total - area_sobreposta) / 10000)  
        texto_area = f"M² = {area_util_m2:.2f}"
        canvas.create_text(
            x_inicial + largura * escala / 2,
            y_inicial - altura * escala - 20,
            text=texto_area,
            fill=self.cores['texto'],
            anchor='s',
            font=("Arial", 12, "bold")
        )
        canvas.create_text(
            x_inicial - 80 * escala,
            y_inicial - altura * escala / 2,
            text=self.texto_esq_var.get(),
            fill=self.cores['texto'],
            anchor='e'
        )
        canvas.create_text(
            x_inicial + largura * escala + 110 * escala,
            y_inicial - altura * escala / 2,
            text=self.texto_dir_var.get(),
            fill=self.cores['texto'],
            anchor='w'
        )
        print("\nÁreas individuais dos painéis por altura:")
        print("-" * 50)
        for i in range(6):
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            altura1 = self._float_safe(self.paineis_alturas_vars[i])
            altura2 = self._float_safe(self.paineis_alturas2_vars[i])
            if largura_painel > 0:
                if altura1 > 0:
                    area1 = largura_painel * altura1 / 10000  
                    print(f"P{i+1} Altura 1: {largura_painel:.1f} x {altura1:.1f} cm = {area1:.2f} m²")
                if altura2 > 0:
                    area2 = largura_painel * altura2 / 10000  
                    print(f"P{i+1} Altura 2: {largura_painel:.1f} x {altura2:.1f} cm = {area2:.2f} m²")
        print("-" * 50)
        print("\nResumo das áreas totais:")
        print(f"Área total dos painéis: {area_total/10000:.2f} m²")
        print(f"Área total das aberturas: {area_aberturas/10000:.2f} m²")
        print(f"Área de abertura sobreposta ao painel: {area_sobreposta/10000:.2f} m²")
        print(f"Área útil final: {area_util_m2:.2f} m²")
    def _criar_botoes(self, parent):
        """Cria os botões da interface."""
        botoes_container = ttk.Frame(parent)
        botoes_container.pack(fill=tk.X, pady=5)
        botoes_frame = ttk.Frame(botoes_container)
        botoes_frame.pack(expand=True)
        
        # Botões principais
        ttk.Button(botoes_frame, text="Novo", command=self._limpar_campos).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Salvar", command=self._salvar_fundo).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Selecionar Excel", command=self._selecionar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Importar Excel", command=self._importar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Criar Excel", command=self._criar_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(botoes_frame, text="Exportar Excel", command=self._exportar_excel).pack(side=tk.LEFT, padx=5)
        
        # NOTA: Os seguintes botões foram temporariamente ocultados e podem ser restaurados posteriormente:
        # - Gerar Script
        # - Gerar Todos Scripts
        # - Salvar Teste
        # - Combinar Scripts
        # - Ordenar Scripts
        
        # Frame para botões de modo
        frame_modais = ttk.Frame(parent)
        frame_modais.pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_modais, text="Modo Reciclagem", command=self._toggle_modo_reciclagem).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_modais, text="Modo Combinar", command=self._toggle_modo_combinacao).pack(side=tk.LEFT, padx=2)
    def _toggle_modo_reciclagem(self):
        self.modo_reciclagem = not self.modo_reciclagem
        self.checkbox_reciclagem_selecionados = set()
        self._atualizar_lista()
        if self.modo_reciclagem:
            from tkinter import messagebox
            messagebox.showinfo("Modo Reciclagem", "Selecione um item na lista 2 (direita) e depois vincule a um item da lista 1 (central) usando os checkboxes.")
        self._atualizar_layout_listas()
    def _criar_lista_fundos(self, parent):
        self.listas_duplas_frame = ttk.Frame(parent)
        self.listas_duplas_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 2))
        
        # Variáveis para controle de ordenação
        self.ordem_atual = {'coluna': 'Número', 'reverso': False}
        
        self.lista1_frame = ttk.Frame(self.listas_duplas_frame)
        self.lista1_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 2))
        
        self.pavimento_filtro_var1 = tk.StringVar(value="Todos")
        self.combo_pavimento1 = ttk.Combobox(self.lista1_frame, textvariable=self.pavimento_filtro_var1, state="readonly", width=12)
        self.combo_pavimento1.pack(anchor="w", padx=2, pady=(2, 1), fill=tk.X)
        self.combo_pavimento1.bind("<<ComboboxSelected>>", lambda e: self._atualizar_lista(1))
        self.total_m2_var = tk.StringVar(value="Total m²: 0.00")
        self.total_m2_label = ttk.Label(self.lista1_frame, textvariable=self.total_m2_var, font=("Arial", 11, "bold"), foreground="#1976d2")
        self.total_m2_label.pack(anchor="w", padx=5, pady=(2, 4))
        ttk.Label(self.lista1_frame, text="Fundos Salvos 1", font=("Arial", 10, "bold")).pack(anchor="w", padx=2, pady=1)
        self.lista2_frame = ttk.Frame(self.listas_duplas_frame)
        self.lista2_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(2, 0))
        self.pavimento_filtro_var2 = tk.StringVar(value="Todos")
        self.combo_pavimento2 = ttk.Combobox(self.lista2_frame, textvariable=self.pavimento_filtro_var2, state="readonly", width=12)
        self.combo_pavimento2.pack(anchor="w", padx=2, pady=(2, 1), fill=tk.X)
        self.combo_pavimento2.bind("<<ComboboxSelected>>", lambda e: self._atualizar_lista(2))
        ttk.Label(self.lista2_frame, text="Fundos Salvos 2", font=("Arial", 10, "bold")).pack(anchor="w", padx=2, pady=1)
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Treeview",
                        background="#ffffff",
                        foreground="#222",
                        rowheight=22,
                        fieldbackground="#ffffff",
                        font=("Arial", 10),
                        borderwidth=0,
                        relief="flat")
        style.map("Treeview",
                  background=[('selected', '#a5d6a7')],
                  foreground=[('selected', '#000')])
        style.configure("Treeview.Heading",
                        background="#388e3c",
                        foreground="#fff",
                        font=("Arial", 10, "bold"),
                        borderwidth=0,
                        relief="flat")
        style.layout("Treeview.Heading", [
            ('Treeheading.cell', {'sticky': 'nswe'}),
            ('Treeheading.border', {'sticky': 'nswe', 'children': [
                ('Treeheading.padding', {'sticky': 'nswe', 'children': [
                    ('Treeheading.image', {'side': 'right', 'sticky': ''}),
                    ('Treeheading.text', {'sticky': 'we'})
                ]})
            ]})
        ])
        # Linhas alternadas em verde claro
        zebra1 = {'odd': '#e8f5e9', 'even': '#ffffff'}
        zebra2 = {'odd': '#c8e6c9', 'even': '#ffffff'}
        # --- PRIMEIRA TREEVIEW ---
        self.lista_fundos1 = ttk.Treeview(self.lista1_frame, columns=("Sel", "Rec", "Número", "Nome", "Pavimento"), show="headings", height=14, style="Treeview")
        self.lista_fundos1.heading("Sel", text="Sel")
        self.lista_fundos1.heading("Rec", text="Rec")
        self.lista_fundos1.heading("Número", text="Nº", command=lambda: self._ordenar_lista("Número"))
        self.lista_fundos1.heading("Nome", text="Nome", command=lambda: self._ordenar_lista("Nome"))
        self.lista_fundos1.heading("Pavimento", text="Pav", command=lambda: self._ordenar_lista("Pavimento"))
        self.lista_fundos1.column("Sel", anchor="center", width=38, stretch=False)
        self.lista_fundos1.column("Rec", anchor="center", width=38, stretch=False)
        self.lista_fundos1.column("Número", anchor="center", width=38, stretch=False)
        self.lista_fundos1.column("Nome", anchor="w", width=90, stretch=False)
        self.lista_fundos1.column("Pavimento", anchor="w", width=90, stretch=False)
        self.lista_fundos1.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.lista_fundos1.bind('<<TreeviewSelect>>', self._carregar_fundo_selecionado)
        self.lista_fundos1.bind("<Button-1>", self._on_click_checkbox_col)
        self.lista_fundos1.tag_configure('odd', background=zebra1['odd'])
        self.lista_fundos1.tag_configure('even', background=zebra1['even'])
        self.lista_fundos2 = ttk.Treeview(self.lista2_frame, columns=("Sel", "Número", "Nome", "Pavimento"), show="headings", height=14, style="Treeview")
        self.lista_fundos2.heading("Sel", text="Sel")
        self.lista_fundos2.heading("Número", text="Nº", command=lambda: self._ordenar_lista("Número", lista2=True))
        self.lista_fundos2.heading("Nome", text="Nome", command=lambda: self._ordenar_lista("Nome", lista2=True))
        self.lista_fundos2.heading("Pavimento", text="Pav", command=lambda: self._ordenar_lista("Pavimento", lista2=True))
        self.lista_fundos2.column("Sel", anchor="center", width=38, stretch=False)
        self.lista_fundos2.column("Número", anchor="center", width=38, stretch=False)
        self.lista_fundos2.column("Nome", anchor="w", width=90, stretch=False)
        self.lista_fundos2.column("Pavimento", anchor="w", width=90, stretch=False)
        self.lista_fundos2.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.lista_fundos2.bind('<<TreeviewSelect>>', self._carregar_fundo_selecionado)
        self.lista_fundos2.bind("<Button-1>", self._on_click_checkbox_col)
        self.lista_fundos2.tag_configure('odd', background=zebra2['odd'])
        self.lista_fundos2.tag_configure('even', background=zebra2['even'])
        self.menu_contexto = tk.Menu(self, tearoff=0)
        self.menu_contexto.add_command(label="Excluir", command=self._excluir_fundo)
        self.lista_fundos1.bind("<Button-3>", self._mostrar_menu_contexto)
        self.lista_fundos2.bind("<Button-3>", self._mostrar_menu_contexto)
        self._atualizar_lista(1)
        self._atualizar_lista(2)
        
        # Configurar seleção múltipla para as listas
        self.lista_fundos1.configure(selectmode='extended')
        self.lista_fundos2.configure(selectmode='extended')
        
        # Adicionar botão de exclusão múltipla
        botoes_frame = ttk.Frame(self.lista1_frame)
        botoes_frame.pack(fill=tk.X, pady=(2, 4))
        
        btn_excluir_mult = ttk.Button(botoes_frame, text="Excluir Selecionados", command=self._excluir_multiplos)
        btn_excluir_mult.pack(side=tk.LEFT, padx=2)
    def _excluir_multiplos(self):
        """Exclui múltiplos itens selecionados das listas."""
        # Coleta seleções de ambas as listas
        selection1 = self.lista_fundos1.selection()
        selection2 = self.lista_fundos2.selection()
        
        if not selection1 and not selection2:
            messagebox.showinfo("Aviso", "Selecione pelo menos um item para excluir")
            return
            
        # Coleta números dos itens selecionados
        numeros = []
        for item in selection1:
            numeros.append(str(self.lista_fundos1.item(item)['values'][2]))  # índice 2 para lista1
        for item in selection2:
            numeros.append(str(self.lista_fundos2.item(item)['values'][1]))  # índice 1 para lista2
            
        # Confirma exclusão
        if messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir os {len(numeros)} itens selecionados?"):
            # Remove os itens do dicionário e das listas
            for numero in numeros:
                if numero in self.fundos_salvos:
                    del self.fundos_salvos[numero]
            
            # Limpa seleções e atualiza interface
            self.lista_fundos1.selection_remove(*selection1)
            self.lista_fundos2.selection_remove(*selection2)
            self._salvar_fundos_salvos()
            self._limpar_campos()
            self._atualizar_lista()
            
            messagebox.showinfo("Sucesso", f"{len(numeros)} itens foram excluídos com sucesso")
    def _atualizar_lista(self, qual=None):
        pavimentos = set()
        for dados in self.fundos_salvos.values():
            pav = str(dados.get('pavimento', '')).strip()
            if pav:
                pavimentos.add(pav)
        pavimentos = sorted(pavimentos)
        opcoes = ["Todos"] + pavimentos
        if hasattr(self, 'combo_pavimento1'):
            self.combo_pavimento1['values'] = opcoes
            if self.pavimento_filtro_var1.get() not in opcoes:
                self.pavimento_filtro_var1.set("Todos")
        if hasattr(self, 'combo_pavimento2'):
            self.combo_pavimento2['values'] = opcoes
            if self.pavimento_filtro_var2.get() not in opcoes:
                self.pavimento_filtro_var2.set("Todos")
        if qual is None:
            self._atualizar_lista(1)
            self._atualizar_lista(2)
            return
        if qual == 2:
            tree = self.lista_fundos2
            filtro = self.pavimento_filtro_var2.get()
        else:
            tree = self.lista_fundos1
            filtro = self.pavimento_filtro_var1.get()
        for item in tree.get_children():
            tree.delete(item)
        i = 0
        total_m2 = 0.0
        for numero, dados in sorted(self.fundos_salvos.items()):
            pav = str(dados.get('pavimento', '')).strip()
            if filtro != "Todos" and pav != filtro:
                continue
            area = float(dados.get('area_util_m2', 0) or 0)
            if qual == 1:
                total_m2 += area
            sel = self._get_checkbox_state(numero) if self.modo_combinacao else ''
            rec = ''
            if self.modo_reciclagem and qual == 1 and hasattr(self, 'reciclagem_item_lista2_selecionado') and self.reciclagem_item_lista2_selecionado:
                if self.reciclagem_vinculos.get(self.reciclagem_item_lista2_selecionado) == numero:
                    rec = '☑'
                else:
                    rec = '☐'
            tag = 'odd' if i % 2 == 0 else 'even'
            color_tag = tag
            if self.modo_reciclagem:
                if qual == 1:
                    if numero in self.reciclagem_reverse:
                        color_tag = 'rec_red'
                    else:
                        color_tag = 'rec_green'
                elif qual == 2:
                    if numero in self.reciclagem_vinculos:
                        color_tag = 'rec_blue'
                    else:
                        color_tag = 'rec_yellow'
            lado = dados.get('lado', 'A')
            nome = dados.get('nome', '')
            if not nome.endswith(f'.{lado}'):
                if nome.endswith('.A') or nome.endswith('.B'):
                    nome = nome[:-2]
                nome = f"{nome}.{lado}"
            if qual == 1 and self.modo_reciclagem:
                tree.insert('', 'end', iid=numero, values=(sel, rec, numero, nome, pav), tags=(color_tag,))
            elif qual == 1:
                tree.insert('', 'end', iid=numero, values=(sel, '', numero, nome, pav), tags=(color_tag,))
            elif qual == 2:
                tree.insert('', 'end', iid=numero, values=(sel, numero, nome, pav), tags=(color_tag,))
            i += 1
        if self.modo_combinacao:
            tree.heading("Sel", text="Sel")
            tree.column("Sel", width=38, anchor="center")
        else:
            tree.heading("Sel", text="")
            tree.column("Sel", width=0)
        if qual == 1:
            if self.modo_reciclagem:
                tree.heading("Rec", text="Rec")
                tree.column("Rec", width=38, anchor="center")
            else:
                tree.heading("Rec", text="")
                tree.column("Rec", width=0)
        if self.modo_reciclagem:
            tree.tag_configure('rec_red', background='#ffcccc')    # Vermelho claro
            tree.tag_configure('rec_green', background='#ccffcc')  # Verde claro
            tree.tag_configure('rec_blue', background='#cce0ff')   # Azul claro
            tree.tag_configure('rec_yellow', background='#fff9cc') # Amarelo claro
        if qual == 1 and hasattr(self, 'total_m2_var'):
            self.total_m2_var.set(f"Total m²: {total_m2:.2f}")
    def _criar_lista_combinacoes(self, parent):
        self.comb_frame = ttk.Frame(parent)
        self.comb_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 2))
        self.comb_pavimento_filtro_var = tk.StringVar(value="Todos")
        self.comb_combo_pavimento = ttk.Combobox(self.comb_frame, textvariable=self.comb_pavimento_filtro_var, state="readonly", width=12)
        self.comb_combo_pavimento.pack(anchor="w", padx=2, pady=(2, 1), fill=tk.X)
        self.comb_combo_pavimento.bind("<<ComboboxSelected>>", lambda e: self._atualizar_lista_combinacoes())
        ttk.Label(self.comb_frame, text="Combinações").pack(anchor="w", padx=2, pady=1)
        botoes = ttk.Frame(self.comb_frame)
        botoes.pack(fill=tk.X, pady=(0, 2))
        btn_criar = ttk.Button(botoes, text="Criar", command=self._modo_criar_combinacao, width=6)
        btn_criar.pack(side=tk.LEFT, padx=0)
        btn_salvar = ttk.Button(botoes, text="Salvar", command=self._salvar_combinacao, width=6)
        btn_salvar.pack(side=tk.LEFT, padx=0)
        btn_excluir = ttk.Button(botoes, text="Excluir", command=self._excluir_combinacao, width=6)
        btn_excluir.pack(side=tk.LEFT, padx=0)
        self.lista_combinacoes = ttk.Treeview(self.comb_frame, columns=("Nome",), show="headings", height=18)
        self.lista_combinacoes.heading("Nome", text="Nome")
        self.lista_combinacoes.column("Nome", width=90)
        self.lista_combinacoes.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        self.lista_combinacoes.bind('<<TreeviewSelect>>', self._on_select_combinacao)
        self._atualizar_lista_combinacoes()
    def _atualizar_lista_combinacoes(self):
        pavimentos = set()
        for comb in self.combinacoes:
            if 'pavimento' in comb:
                pav = str(comb['pavimento']).strip()
                if pav:
                    pavimentos.add(pav)
        pavimentos = sorted(pavimentos)
        opcoes = ["Todos"] + pavimentos
        if hasattr(self, 'comb_combo_pavimento'):
            self.comb_combo_pavimento['values'] = opcoes
            if self.comb_pavimento_filtro_var.get() not in opcoes:
                self.comb_pavimento_filtro_var.set("Todos")
        filtro = self.comb_pavimento_filtro_var.get() if hasattr(self, 'comb_pavimento_filtro_var') else "Todos"
        for item in self.lista_combinacoes.get_children():
            self.lista_combinacoes.delete(item)
        for idx, comb in enumerate(self.combinacoes):
            pav = comb.get('pavimento', '')
            if filtro != "Todos" and pav != filtro:
                continue
            self.lista_combinacoes.insert('', 'end', iid=str(idx), values=(comb['nome'],))
    def _modo_criar_combinacao(self):
        self.modo_combinacao = True
        self.checkbox_selecionados = set()
        self._atualizar_lista()
    def _salvar_combinacao(self):
        if not self.modo_combinacao:
            return
        selecionados = list(self.checkbox_selecionados)
        if len(selecionados) != 2:
            messagebox.showerror("Erro", "Selecione exatamente 2 itens para combinar.")
            return
        nomes = []
        pavimentos = set()
        for iid in selecionados:
            valores = None
            if hasattr(self, 'lista_fundos1') and self.lista_fundos1.exists(iid):
                valores = self.lista_fundos1.item(iid)['values']
            elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.exists(iid):
                valores = self.lista_fundos2.item(iid)['values']
            if valores:
                nomes.append(str(valores[2]))  
                pavimentos.add(str(valores[3]))
        if len(pavimentos) != 1:
            messagebox.showerror("Erro", "Só é possível combinar fundos do MESMO pavimento!")
            return
        pavimento = pavimentos.pop()
        nome_comb = f"{nomes[0]} + {nomes[1]}"
        self.combinacoes.append({'nome': nome_comb, 'ids': selecionados, 'pavimento': pavimento})
        self._atualizar_lista_combinacoes()
        self._salvar_combinacoes()  
        self.modo_combinacao = False
        self.checkbox_selecionados = set()
        self._atualizar_lista()
    def _on_select_combinacao(self, event=None):
        sel = self.lista_combinacoes.selection()
        if not sel:
            return
        idx = int(sel[0])
        comb = self.combinacoes[idx]
        self.modo_combinacao = True
        self.checkbox_selecionados = set(comb['ids'])
        self._atualizar_lista()
    def _criar_aba_dados_gerais(self, parent):
        self.paineis_entries = []
        self.dados_frame = ttk.LabelFrame(parent, text="Dados Lateral de Viga", padding=4)
        self.dados_frame.grid(row=0, column=2, sticky="nsew", padx=(8, 0))
        self.dados_frame.columnconfigure(0, weight=1)
        self.dados_frame.columnconfigure(1, weight=1)
        frame_botoes_acoes = ttk.Frame(self.dados_frame)
        frame_botoes_acoes.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        ttk.Button(frame_botoes_acoes, text="Iniciar Produção", command=self._abrir_janela_producao).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Nomes", command=self._select_nomes).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Linha", command=self._select_linha).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Níveis", command=self._select_niveis).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Níveis 2", command=self._select_niveis2).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Pilar", command=self._select_pilar).pack(side=tk.LEFT, padx=2)
        ttk.Button(frame_botoes_acoes, text="Select Abertura", command=self._select_abertura).pack(side=tk.LEFT, padx=2)
        info_frame = ttk.LabelFrame(self.dados_frame, text="Informações Básicas", padding=5)
        info_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=5)
        for i in range(2):
            info_frame.columnconfigure(i, weight=1)
        ttk.Label(info_frame, text="Número:").grid(row=0, column=0, sticky="w", padx=5)
        numero_entry = ttk.Entry(info_frame, textvariable=self.numero_var, width=8)
        numero_entry.grid(row=0, column=0, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Pavimento:").grid(row=0, column=1, sticky="w", padx=5)
        pavimento_entry = ttk.Entry(info_frame, textvariable=self.pavimento_var)
        pavimento_entry.grid(row=0, column=1, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Nome:").grid(row=1, column=0, sticky="w", padx=5)
        nome_entry = ttk.Entry(info_frame, textvariable=self.nome_var)
        nome_entry.grid(row=1, column=0, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Observação:").grid(row=1, column=1, sticky="w", padx=5)
        obs_entry = ttk.Entry(info_frame, textvariable=self.obs_var)
        obs_entry.grid(row=1, column=1, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Nível Viga:").grid(row=2, column=0, sticky="w", padx=5)
        nivel_viga_entry = ttk.Entry(info_frame, textvariable=self.nivel_viga_var)
        nivel_viga_entry.grid(row=2, column=0, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Nível Oposto:").grid(row=2, column=1, sticky="w", padx=5)
        nivel_oposto_entry = ttk.Entry(info_frame, textvariable=self.nivel_oposto_var)
        nivel_oposto_entry.grid(row=2, column=1, sticky="e", padx=(70,5))
        self.nivel_pe_direito_var = tk.StringVar()
        ttk.Label(info_frame, text="Nível Pé Direito:").grid(row=3, column=0, sticky="w", padx=5)
        nivel_pe_direito_entry = ttk.Entry(info_frame, textvariable=self.nivel_pe_direito_var)
        nivel_pe_direito_entry.grid(row=3, column=0, sticky="e", padx=(70,5))
        self.ajuste_var = tk.StringVar()
        ttk.Label(info_frame, text="Ajuste:").grid(row=3, column=1, sticky="w", padx=5)
        ajuste_entry = ttk.Entry(info_frame, textvariable=self.ajuste_var)
        ajuste_entry.grid(row=3, column=1, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Texto Esquerdo:").grid(row=4, column=0, sticky="w", padx=5)
        texto_esq_entry = ttk.Entry(info_frame, textvariable=self.texto_esq_var)
        texto_esq_entry.grid(row=4, column=0, sticky="e", padx=(70,5))
        ttk.Label(info_frame, text="Texto Direito:").grid(row=4, column=1, sticky="w", padx=5)
        texto_dir_entry = ttk.Entry(info_frame, textvariable=self.texto_dir_var)
        texto_dir_entry.grid(row=4, column=1, sticky="e", padx=(70,5))
        frame_continuacao = ttk.Frame(info_frame)
        frame_continuacao.grid(row=5, column=0, columnspan=2, sticky="w", pady=2)
        ttk.Label(frame_continuacao, text="Continuação:").pack(side=tk.LEFT, padx=(0,5))
        tk.Radiobutton(frame_continuacao, text="Obstaculo", variable=self.continuacao_var, value="Obstaculo").pack(side=tk.LEFT)
        tk.Radiobutton(frame_continuacao, text="Proxima Parte", variable=self.continuacao_var, value="Proxima Parte").pack(side=tk.LEFT)
        tk.Radiobutton(frame_continuacao, text="Viga Continuacao", variable=self.continuacao_var, value="Viga Continuacao").pack(side=tk.LEFT)
        self.lado_var = tk.StringVar(value="A")
        def atualizar_sufixo_nome(*args):
            nome_atual = self.nome_var.get()
            lado = self.lado_var.get()
            if not nome_atual:
                return
            if not nome_atual.endswith(f'.{lado}'):
                if nome_atual.endswith('.A') or nome_atual.endswith('.B'):
                    nome_atual = nome_atual[:-2]
                self.nome_var.set(f"{nome_atual}.{lado}")
        self.lado_var.trace_add('write', atualizar_sufixo_nome)
        frame_lado = ttk.Frame(frame_continuacao)
        frame_lado.pack(side=tk.LEFT, padx=(18,0))
        ttk.Label(frame_lado, text="Lado:").pack(side=tk.LEFT, padx=(0,4))
        tk.Radiobutton(frame_lado, text="A", variable=self.lado_var, value="A").pack(side=tk.LEFT)
        tk.Radiobutton(frame_lado, text="B", variable=self.lado_var, value="B").pack(side=tk.LEFT)
        self.area_util_var = tk.StringVar(value="M² = 0.00")
        area_util_label = ttk.Label(info_frame, textvariable=self.area_util_var, font=("Arial", 11, "bold"), foreground="#1976d2")
        area_util_label.grid(row=6, column=0, columnspan=2, sticky="w", padx=5, pady=(2, 8))
        def atualizar_area_util(*args):
            area_total = 0
            for i in range(6):
                largura_painel = self._float_safe(self.paineis_larguras_vars[i])
                altura1 = self._float_safe(self.paineis_alturas_vars[i])
                altura2 = self._float_safe(self.paineis_alturas2_vars[i])
                if largura_painel > 0:
                    if altura1 > 0:
                        area_total += largura_painel * altura1
                    if altura2 > 0:
                        area_total += largura_painel * altura2
            area_sobreposta = 0
            largura = self._float_safe(self.largura_var)
            altura = self._float_safe(self.altura_geral_var)
            rx1, ry1, rx2, ry2 = 0, -altura, largura, 0
            for i, linha in enumerate(self.aberturas_vars):
                try:
                    dist, prof, larg = [self._float_safe(v) for v in linha]
                    if prof > 0 and larg > 0:
                        if i < 2:
                            x_pos = 0
                            y_pos = -prof - dist
                        else:
                            x_pos = largura - larg
                            y_pos = -prof - dist
                        ax1 = x_pos
                        ay1 = y_pos
                        ax2 = x_pos + larg
                        ay2 = y_pos + prof
                        area_sobrep = self._area_intersecao(rx1, ry1, rx2, ry2, ax1, ay1, ax2, ay2)
                        area_sobreposta += area_sobrep
                except Exception:
                    continue
            area_util_m2 = max(0, (area_total - area_sobreposta) / 10000)
            self.area_util_var.set(f"M² = {area_util_m2:.2f}")
            if hasattr(self, 'numero_var') and self.numero_var.get():
                numero = self.numero_var.get()
                if numero in self.fundos_salvos:
                    self.fundos_salvos[numero]['area_util_m2'] = area_util_m2
                    self._salvar_fundos_salvos()
        for var in [self.largura_var, self.altura_geral_var, *self.paineis_larguras_vars, *self.paineis_alturas_vars, *self.paineis_alturas2_vars]:
            var.trace_add('write', atualizar_area_util)
        for linha in self.aberturas_vars:
            for var in linha:
                var.trace_add('write', atualizar_area_util)
        atualizar_area_util()
        for var in [self.largura_var, self.altura_geral_var, *self.paineis_larguras_vars, *self.paineis_alturas_vars, self.sarrafo_esq_var, self.sarrafo_dir_var]:
            var.trace_add('write', self._atualizar_preview)
        paineis_frame = ttk.LabelFrame(self.dados_frame, text="Painéis", padding=5)
        paineis_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=5)
        ttk.Label(paineis_frame, text="Painel").grid(row=0, column=0, padx=5, pady=2)
        ttk.Label(paineis_frame, text="Largura").grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(paineis_frame, text="Altura 1").grid(row=0, column=2, padx=5, pady=2)
        ttk.Label(paineis_frame, text="SARR/GRA").grid(row=0, column=3, padx=5, pady=2)
        ttk.Label(paineis_frame, text="Altura 2").grid(row=0, column=4, padx=5, pady=2)
        ttk.Label(paineis_frame, text="SARR/GRA").grid(row=0, column=5, padx=5, pady=2)
        ttk.Label(paineis_frame, text="Laje Sup.").grid(row=0, column=6, padx=2, pady=2)
        ttk.Label(paineis_frame, text="Laje Inf.").grid(row=0, column=7, padx=2, pady=2)
        ttk.Label(paineis_frame, text="Central Alt.").grid(row=0, column=8, padx=2, pady=2)
        ttk.Label(paineis_frame, text="Universal:").grid(row=1, column=0, padx=5, pady=2)
        largura_universal_entry = ttk.Entry(paineis_frame, textvariable=self.largura_var, width=8)
        largura_universal_entry.grid(row=1, column=1, padx=5, pady=2)
        self.largura_var.trace_add('write', lambda *a: self._calcular_paineis_automatico())
        altura_geral_universal_entry = ttk.Entry(paineis_frame, textvariable=self.altura_geral_var, width=8)
        altura_geral_universal_entry.grid(row=1, column=2, padx=5, pady=2)
        frame_tipo1_universal = ttk.Frame(paineis_frame)
        frame_tipo1_universal.grid(row=1, column=3, padx=0, pady=2)
        tk.Radiobutton(frame_tipo1_universal, variable=self.tipo1_universal_var, value="Sarrafeado", width=1, padx=0, pady=0, relief='flat', command=lambda: [v.set("Sarrafeado") for v in self.paineis_tipo1_vars]).pack(side=tk.LEFT, padx=0, pady=0)
        tk.Radiobutton(frame_tipo1_universal, variable=self.tipo1_universal_var, value="Grade", width=1, padx=0, pady=0, relief='flat', command=lambda: [v.set("Grade") for v in self.paineis_tipo1_vars]).pack(side=tk.LEFT, padx=0, pady=0)
        ttk.Entry(frame_tipo1_universal, textvariable=self.grade_altura1_universal, width=4).pack(side=tk.LEFT, padx=2)
        altura_2_geral_universal_entry = ttk.Entry(paineis_frame, textvariable=self.altura_2_geral_var, width=8)
        altura_2_geral_universal_entry.grid(row=1, column=4, padx=5, pady=2)
        frame_tipo2_universal = ttk.Frame(paineis_frame)
        frame_tipo2_universal.grid(row=1, column=5, padx=0, pady=2)
        tk.Radiobutton(frame_tipo2_universal, variable=self.tipo2_universal_var, value="Sarrafeado", width=1, padx=0, pady=0, relief='flat', command=lambda: [v.set("Sarrafeado") for v in self.paineis_tipo2_vars]).pack(side=tk.LEFT, padx=0, pady=0)
        tk.Radiobutton(frame_tipo2_universal, variable=self.tipo2_universal_var, value="Grade", width=1, padx=0, pady=0, relief='flat', command=lambda: [v.set("Grade") for v in self.paineis_tipo2_vars]).pack(side=tk.LEFT, padx=0, pady=0)
        ttk.Entry(frame_tipo2_universal, textvariable=self.grade_altura2_universal, width=4).pack(side=tk.LEFT, padx=2)
        self.grade_altura2_universal.trace_add('write', lambda *a: [v.set(self.grade_altura2_universal.get()) for v in self.paineis_grade_altura2_vars])
        laje_sup_universal_entry = ttk.Entry(paineis_frame, textvariable=self.laje_sup_universal, width=8)
        laje_sup_universal_entry.grid(row=1, column=6, padx=2, pady=2)
        laje_inf_universal_entry = ttk.Entry(paineis_frame, textvariable=self.laje_inf_universal, width=8)
        laje_inf_universal_entry.grid(row=1, column=7, padx=2, pady=2)
        laje_central_universal_entry = ttk.Entry(paineis_frame, textvariable=self.laje_central_alt_universal, width=8)
        laje_central_universal_entry.grid(row=1, column=8, padx=2, pady=2)
        for i in range(6):
            ttk.Label(paineis_frame, text=f"P{i+1}:").grid(row=i+2, column=0, sticky="w", padx=5, pady=2)
            entry_larg = ttk.Entry(paineis_frame, textvariable=self.paineis_larguras_vars[i], width=8)
            entry_larg.grid(row=i+2, column=1, sticky="ew", padx=5, pady=2)
            entry_alt1 = ttk.Entry(paineis_frame, textvariable=self.paineis_alturas_vars[i], width=8)
            entry_alt1.grid(row=i+2, column=2, sticky="ew", padx=5, pady=2)
            frame_tipo1 = ttk.Frame(paineis_frame)
            frame_tipo1.grid(row=i+2, column=3, padx=0, pady=2)
            tk.Radiobutton(frame_tipo1, variable=self.paineis_tipo1_vars[i], value="Sarrafeado", width=1, padx=0, pady=0, relief='flat').pack(side=tk.LEFT, padx=0, pady=0)
            tk.Radiobutton(frame_tipo1, variable=self.paineis_tipo1_vars[i], value="Grade", width=1, padx=0, pady=0, relief='flat').pack(side=tk.LEFT, padx=0, pady=0)
            ttk.Entry(frame_tipo1, textvariable=self.paineis_grade_altura1_vars[i], width=4).pack(side=tk.LEFT, padx=2)
            entry_alt2 = ttk.Entry(paineis_frame, textvariable=self.paineis_alturas2_vars[i], width=8)
            entry_alt2.grid(row=i+2, column=4, sticky="ew", padx=5, pady=2)
            frame_tipo2 = ttk.Frame(paineis_frame)
            frame_tipo2.grid(row=i+2, column=5, padx=0, pady=2)
            tk.Radiobutton(frame_tipo2, variable=self.paineis_tipo2_vars[i], value="Sarrafeado", width=1, padx=0, pady=0, relief='flat').pack(side=tk.LEFT, padx=0, pady=0)
            tk.Radiobutton(frame_tipo2, variable=self.paineis_tipo2_vars[i], value="Grade", width=1, padx=0, pady=0, relief='flat').pack(side=tk.LEFT, padx=0, pady=0)
            ttk.Entry(frame_tipo2, textvariable=self.paineis_grade_altura2_vars[i], width=4).pack(side=tk.LEFT, padx=2)
            entry_laje_sup = ttk.Entry(paineis_frame, textvariable=self.lajes_sup_vars[i], width=6)
            entry_laje_sup.grid(row=i+2, column=6, padx=2, pady=2)
            entry_laje_inf = ttk.Entry(paineis_frame, textvariable=self.lajes_inf_vars[i], width=6)
            entry_laje_inf.grid(row=i+2, column=7, padx=2, pady=2)
            entry_laje_c_alt = ttk.Entry(paineis_frame, textvariable=self.lajes_central_alt_vars[i], width=6)
            entry_laje_c_alt.grid(row=i+2, column=8, padx=2, pady=2)
            self.paineis_entries.append((entry_larg, entry_alt1, entry_alt2, entry_laje_sup, entry_laje_inf, entry_laje_c_alt))
        self.laje_sup_universal.trace_add('write', lambda *a: [v.set(self._float_safe(self.laje_sup_universal)) for v in self.lajes_sup_vars])
        self.laje_inf_universal.trace_add('write', lambda *a: [v.set(self._float_safe(self.laje_inf_universal)) for v in self.lajes_inf_vars])
        self.laje_central_alt_universal.trace_add('write', lambda *a: [v.set(self._float_safe(self.laje_central_alt_universal)) for v in self.lajes_central_alt_vars])
        opcoes_frame = ttk.LabelFrame(self.dados_frame, text="Opções", padding=5)
        opcoes_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo vertical esquerda", variable=self.sarrafo_esq_var, command=self._atualizar_preview).grid(row=0, column=0, padx=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo vertical direita", variable=self.sarrafo_dir_var, command=self._atualizar_preview).grid(row=0, column=1, padx=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo Altura 2 Esquerda", variable=self.sarrafo_alt2_esq_var, command=self._atualizar_preview).grid(row=1, column=0, padx=5)
        ttk.Checkbutton(opcoes_frame, text="Sarrafo Altura 2 Direita", variable=self.sarrafo_alt2_dir_var, command=self._atualizar_preview).grid(row=1, column=1, padx=5)
        aberturas_frame = ttk.LabelFrame(self.dados_frame, text="Aberturas", padding=5)
        aberturas_frame.grid(row=4, column=0, columnspan=2, sticky="ew", pady=5)
        frame_aberturas = ttk.Frame(aberturas_frame)
        frame_aberturas.grid(row=0, column=0, sticky="nw")
        frame_pilar = ttk.Frame(aberturas_frame)
        frame_pilar.grid(row=0, column=1, sticky="ne", padx=(30,0))
        aberturas_labels = ["Dist", "Prof", "Larg"]
        lados = ["T/E", "F/E", "T/D", "F/D"]
        self.forcar_altura1_vars = [tk.BooleanVar(value=False) for _ in range(4)]
        ttk.Label(frame_aberturas, text="").grid(row=0, column=0, sticky="w", padx=5)
        for j, lab in enumerate(aberturas_labels):
            ttk.Label(frame_aberturas, text=lab, font=("Arial", 8, "bold")).grid(row=0, column=j+1, sticky="w", padx=2)
        ttk.Label(frame_aberturas, text="Forçar Alt.1", font=("Arial", 8, "bold")).grid(row=0, column=4, sticky="w", padx=2)
        for i, lado in enumerate(lados):
            ttk.Label(frame_aberturas, text=lado+":").grid(row=i+1, column=0, sticky="w", padx=5)
            for j, lab in enumerate(aberturas_labels):
                ab_entry = ttk.Entry(frame_aberturas, textvariable=self.aberturas_vars[i][j], width=8)
                ab_entry.grid(row=i+1, column=j+1, padx=2)
                ab_entry.bind('<FocusOut>', lambda e: self._salvar_fundo())
            ttk.Checkbutton(frame_aberturas, variable=self.forcar_altura1_vars[i]).grid(row=i+1, column=4, padx=2)
        ttk.Label(frame_pilar, text="").grid(row=0, column=0, padx=5)
        ttk.Label(frame_pilar, text="Detalhes").grid(row=0, column=1, padx=2)
        ttk.Label(frame_pilar, text="Distancia").grid(row=0, column=2, padx=2)
        ttk.Label(frame_pilar, text="Largura").grid(row=0, column=3, padx=2)
        ttk.Label(frame_pilar, text="Det. Pilar Esq.").grid(row=1, column=1, sticky="w", padx=5)
        ttk.Entry(frame_pilar, textvariable=self.detalhe_pilar_esq_vars[0], width=8).grid(row=1, column=2, padx=2)
        ttk.Entry(frame_pilar, textvariable=self.detalhe_pilar_esq_vars[1], width=8).grid(row=1, column=3, padx=2)
        ttk.Label(frame_pilar, text="Det. Pilar Dir.").grid(row=2, column=1, sticky="w", padx=5)
        ttk.Entry(frame_pilar, textvariable=self.detalhe_pilar_dir_vars[0], width=8).grid(row=2, column=2, padx=2)
        ttk.Entry(frame_pilar, textvariable=self.detalhe_pilar_dir_vars[1], width=8).grid(row=2, column=3, padx=2)
    def _excluir_combinacao(self):
        sel = self.lista_combinacoes.selection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.combinacoes):
            return
        import tkinter.messagebox as mb
        if not mb.askyesno("Excluir", f"Deseja realmente excluir a combinação '{self.combinacoes[idx]['nome']}'?"):
            return
        del self.combinacoes[idx]
        self._atualizar_lista_combinacoes()
        self._salvar_combinacoes()
    def _selecionar_textos_consecutivos(self):
        print('[DEBUG] _selecionar_textos_consecutivos chamado')
        if getattr(self, '_f8_cancelando', False):
            print('[DEBUG] _selecionar_textos_consecutivos abortado por F8')
            self._f8_cancelando = False
            return
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                print(f"[DEBUG] messagebox.showerror: Nenhum documento AutoCAD ativo encontrado!")
                messagebox.showerror("Erro", "Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            info_label = tk.Label(self, 
                                text="Selecione o primeiro texto (Nome) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.nome_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O primeiro objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O primeiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O primeiro objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            info_label = tk.Label(self, 
                                text="Selecione o segundo texto (Texto Esquerda) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_esq_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O segundo objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O segundo objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O segundo objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            info_label = tk.Label(self, 
                                text="Selecione o terceiro texto (Texto Direita) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_dir_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O terceiro objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O terceiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O terceiro objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar textos: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar textos: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _salvar_fundo(self):
        """Salva os dados do fundo atual."""
        nome_atual = self.nome_var.get()
        lado = None
        if hasattr(self, 'lado_var'):
            lado = self.lado_var.get()
        elif hasattr(self, '_prod_lado'):
            lado = self._prod_lado.get()
        if nome_atual and lado and not nome_atual.endswith(f'.{lado}'):
            if nome_atual.endswith('.A') or nome_atual.endswith('.B'):
                nome_atual = nome_atual[:-2]
            self.nome_var.set(f"{nome_atual}.{lado}")
        if getattr(self, '_carregando_fundo', False):
            return
        numero = self.numero_var.get()
        if not numero:
            from tkinter import messagebox
            messagebox.showerror("Erro", "O campo 'Número' está vazio! Preencha para salvar.")
            return
        try:
            dados = {
                'numero': numero,
                'pavimento': self.pavimento_var.get(),
                'nome': self.nome_var.get(),
                'texto_esq': self.texto_esq_var.get(),
                'texto_dir': self.texto_dir_var.get(),
                'obs': self.obs_var.get(),
                'largura': self._float_safe(self.largura_var),
                'altura_geral': self.altura_geral_var.get(),
                'paineis_larguras': [self._float_safe(v) for v in self.paineis_larguras_vars],
                'paineis_alturas': [v.get() for v in self.paineis_alturas_vars],
                'paineis_alturas2': [v.get() for v in self.paineis_alturas2_vars],
                'sarrafo_esq': self.sarrafo_esq_var.get(),
                'sarrafo_dir': self.sarrafo_dir_var.get(),
                'sarrafo_alt2_esq': self.sarrafo_alt2_esq_var.get(),
                'sarrafo_alt2_dir': self.sarrafo_alt2_dir_var.get(),
                'aberturas': [[self._float_safe(var) for var in linha] for linha in self.aberturas_vars],
                'tipo_painel_inicial': self.tipo_painel_inicial_var.get(),
                'tipo_distribuicao': self.tipo_distribuicao_var.get(),
                'lajes_sup': [self._float_safe(v) for v in self.lajes_sup_vars],
                'lajes_inf': [self._float_safe(v) for v in self.lajes_inf_vars],
                'lajes_central_alt': [self._float_safe(v) for v in self.lajes_central_alt_vars],
                'detalhe_pilar_esq': [self._float_safe(v) for v in self.detalhe_pilar_esq_vars],
                'detalhe_pilar_dir': [self._float_safe(v) for v in self.detalhe_pilar_dir_vars],
                'nivel_oposto': self.nivel_oposto_var.get(),
                'nivel_viga': self.nivel_viga_var.get(),
                'nivel_pe_direito': self.nivel_pe_direito_var.get(),
                'ajuste': self.ajuste_var.get(),
                'paineis_tipo1': [v.get() for v in self.paineis_tipo1_vars],
                'paineis_tipo2': [v.get() for v in self.paineis_tipo2_vars],
                'paineis_grade_altura1': [v.get() for v in self.paineis_grade_altura1_vars],
                'paineis_grade_altura2': [v.get() for v in self.paineis_grade_altura2_vars],
                'continuacao': self.continuacao_var.get(),
                'lado': self.lado_var.get(),
            }
            self.fundos_salvos[numero] = dados
            self._salvar_fundos_salvos()
            print(f"Fundo '{numero}' salvo com sucesso.")
            self._atualizar_lista()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro ao salvar fundo", str(e))
    def _carregar_fundo_selecionado(self, event=None):
        if self.modo_reciclagem and event and hasattr(event, 'widget') and event.widget == self.lista_fundos2:
            sel = self.lista_fundos2.selection()
            if sel:
                self.reciclagem_item_lista2_selecionado = sel[0]
                self._atualizar_lista(1)
        tree = None
        if event and hasattr(event, 'widget'):
            if event.widget == self.lista_fundos1:
                tree = self.lista_fundos1
            elif event.widget == self.lista_fundos2:
                tree = self.lista_fundos2
        if tree is None:
            if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
                tree = self.lista_fundos1
            elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
                tree = self.lista_fundos2
        if tree is None:
            return
        selection = tree.selection()
        if not selection:
            self._atualizar_preview()  
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        self.numero_var.set(dados.get('numero', numero))
        self.nome_var.set(dados.get('nome', ''))
        self.obs_var.set(dados.get('obs', ''))
        self.pavimento_var.set(dados.get('pavimento', ''))
        try:
            self.largura_var.set(float(dados.get('largura', 0)))
        except Exception:
            self.largura_var.set(0)
        try:
            self.altura_geral_var.set(str(dados.get('altura_geral', '0')))
        except Exception:
            self.altura_geral_var.set('0')
        self.texto_esq_var.set(dados.get('texto_esq', ''))
        self.texto_dir_var.set(dados.get('texto_dir', ''))
        self.tipo_distribuicao_var.set(dados.get('tipo_distribuicao', '122'))
        self.tipo_painel_inicial_var.set(dados.get('tipo_painel_inicial', '300'))
        self.sarrafo_esq_var.set(dados.get('sarrafo_esq', True))
        self.sarrafo_dir_var.set(dados.get('sarrafo_dir', True))
        self.sarrafo_alt2_esq_var.set(dados.get('sarrafo_alt2_esq', False))
        self.sarrafo_alt2_dir_var.set(dados.get('sarrafo_alt2_dir', False))
        paineis_larguras = dados.get('paineis_larguras', [0]*6)
        paineis_alturas = dados.get('paineis_alturas', [self.altura_geral_var.get()]*6)
        paineis_alturas2 = dados.get('paineis_alturas2', ["0"]*6)
        for i in range(6):
            try:
                self.paineis_larguras_vars[i].set(float(paineis_larguras[i]))
            except Exception:
                self.paineis_larguras_vars[i].set(0)
            try:
                self.paineis_alturas_vars[i].set(str(paineis_alturas[i]))
            except Exception:
                self.paineis_alturas_vars[i].set('')
            try:
                self.paineis_alturas2_vars[i].set(str(paineis_alturas2[i]))
            except Exception:
                self.paineis_alturas2_vars[i].set('')
        aberturas = dados.get('aberturas', [[0,0,0] for _ in range(4)])
        for i, abertura in enumerate(aberturas):
            for j in range(3):
                try:
                    self.aberturas_vars[i][j].set(float(abertura[j]))
                except Exception:
                    self.aberturas_vars[i][j].set(0)
        lajes_sup = dados.get('lajes_sup', [0]*6)
        lajes_inf = dados.get('lajes_inf', [0]*6)
        lajes_central_alt = dados.get('lajes_central_alt', [0]*6)
        for i in range(6):
            try:
                self.lajes_sup_vars[i].set(float(lajes_sup[i] or 0))
            except Exception:
                self.lajes_sup_vars[i].set(0)
            try:
                self.lajes_inf_vars[i].set(float(lajes_inf[i] or 0))
            except Exception:
                self.lajes_inf_vars[i].set(0)
            try:
                self.lajes_central_alt_vars[i].set(float(lajes_central_alt[i] or 0))
            except Exception:
                self.lajes_central_alt_vars[i].set(0)
        detalhe_pilar_esq = dados.get('detalhe_pilar_esq', [0, 0])
        detalhe_pilar_dir = dados.get('detalhe_pilar_dir', [0, 0])
        for i in range(2):
            try:
                self.detalhe_pilar_esq_vars[i].set(float(detalhe_pilar_esq[i]))
            except Exception:
                self.detalhe_pilar_esq_vars[i].set(0)
            try:
                self.detalhe_pilar_dir_vars[i].set(float(detalhe_pilar_dir[i]))
            except Exception:
                self.detalhe_pilar_dir_vars[i].set(0)
        self.nivel_oposto_var.set(dados.get('nivel_oposto', ''))
        self.nivel_viga_var.set(dados.get('nivel_viga', ''))
        self.nivel_pe_direito_var.set(dados.get('nivel_pe_direito', ''))
        self.ajuste_var.set(dados.get('ajuste', ''))
        paineis_tipo1 = dados.get('paineis_tipo1', [False]*6)
        paineis_tipo2 = dados.get('paineis_tipo2', [False]*6)
        paineis_grade_altura1 = dados.get('paineis_grade_altura1', [False]*6)
        paineis_grade_altura2 = dados.get('paineis_grade_altura2', [False]*6)
        for i in range(6):
            self.paineis_tipo1_vars[i].set(paineis_tipo1[i])
            self.paineis_tipo2_vars[i].set(paineis_tipo2[i])
            self.paineis_grade_altura1_vars[i].set(paineis_grade_altura1[i])
            self.paineis_grade_altura2_vars[i].set(paineis_grade_altura2[i])
        self.continuacao_var.set(dados.get('continuacao', False))
        self.lado_var.set(dados.get('lado', 'A'))
        self._atualizar_preview()
    def _limpar_campos(self):
        """Limpa todos os campos do formulário."""
        self.numero_var.set('')
        self.pavimento_var.set('')
        self.nome_var.set('')
        self.texto_esq_var.set('')
        self.texto_dir_var.set('')
        self.obs_var.set('')
        self.largura_var.set('')
        self.altura_geral_var.set('')
        self.sarrafo_esq_var.set(False)
        self.sarrafo_dir_var.set(False)
        for linha in self.aberturas_vars:
            for var in linha:
                var.set('')
        for var in self.paineis_larguras_vars:
            var.set('')
        for var in self.paineis_alturas_vars:
            var.set('')
        for var in self.paineis_alturas2_vars:
            var.set('')
        for var in self.lajes_sup_vars:
            var.set('')
        for var in self.lajes_inf_vars:
            var.set('')
        for var in self.lajes_central_alt_vars:
                var.set('')
        self.tipo_painel_inicial_var.set('300')
        self._atualizar_preview()
    def _exportar_excel(self):
        """Exporta os dados para um arquivo Excel."""
        if not self.fundos_salvos:
            messagebox.showerror("Erro", "Não há dados para exportar!")
            return
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Salvar como"
        )
        if not arquivo:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fundos"
            cabecalhos = [
                'Número', 'Pavimento', 'Nome', 'Texto Esquerda', 'Texto Direita',
                'Observação', 'Comprimento', 'Largura', 'P1', 'P2', 'P3',
                'Sarrafo Esquerda', 'Sarrafo Direita', 'Tipo Painel Inicial',
                'Nível Oposto', 'Nível Viga'
            ]
            for i, cabecalho in enumerate(cabecalhos, 1):
                ws.cell(row=1, column=i, value=cabecalho)
            for i, abertura in enumerate(['Abertura 1', 'Abertura 2', 'Abertura 3', 'Abertura 4'], 1):
                ws.cell(row=1, column=len(cabecalhos) + i, value=abertura)
            row = 2
            for numero, dados in sorted(self.fundos_salvos.items()):
                ws.cell(row=row, column=1, value=numero)
                ws.cell(row=row, column=2, value=dados.get('pavimento', ''))
                ws.cell(row=row, column=3, value=dados.get('nome', ''))
                ws.cell(row=row, column=4, value=dados.get('texto_esq', ''))
                ws.cell(row=row, column=5, value=dados.get('texto_dir', ''))
                ws.cell(row=row, column=6, value=dados.get('obs', ''))
                ws.cell(row=row, column=7, value=dados.get('comprimento', ''))
                ws.cell(row=row, column=8, value=dados.get('largura', ''))
                ws.cell(row=row, column=9, value=dados.get('p1', ''))
                ws.cell(row=row, column=10, value=dados.get('p2', ''))
                ws.cell(row=row, column=11, value=dados.get('p3', ''))
                ws.cell(row=row, column=12, value=dados.get('sarrafo_esq', False))
                ws.cell(row=row, column=13, value=dados.get('sarrafo_dir', False))
                ws.cell(row=row, column=14, value=dados.get('tipo_painel_inicial', '300'))
                ws.cell(row=row, column=15, value=dados.get('nivel_oposto', ''))
                ws.cell(row=row, column=16, value=dados.get('nivel_viga', ''))
                aberturas = dados.get('aberturas', [[0, 0, 0] for _ in range(4)])
                for i, abertura in enumerate(aberturas, 1):
                    ws.cell(row=row, column=len(cabecalhos) + i, value=str(abertura))
                row += 1
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar dados: {str(e)}")
    def _criar_excel(self):
        arquivo = filedialog.asksaveasfilename(title="Salvar Novo Excel", filetypes=[("Excel", "*.xlsx")], defaultextension=".xlsx")
        if not arquivo:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        mapeamento = {
            'numero': 52,
            'nome': 53,
            'observacoes': 54,
            'pavimento': 55,
            'largura': 56,
            'altura': 57,
            'texto_esquerda': 58,
            'texto_direita': 59,
            'tipo_distribuicao': 60,
            'tipo_painel_inicial': 61,
            'sarrafo_esq': 62,
            'sarrafo_dir': 63,
            'painel1': 64,
            'painel2': 65,
            'painel3': 66,
            'painel4': 67,
            'painel5': 68,
            'painel6': 69,
            'abertura_te_dist': 70,
            'abertura_te_prof': 71,
            'abertura_te_larg': 72,
            'abertura_fe_dist': 73,
            'abertura_fe_prof': 74,
            'abertura_fe_larg': 75,
            'abertura_td_dist': 76,
            'abertura_td_prof': 77,
            'abertura_td_larg': 78,
            'abertura_fd_dist': 79,
            'abertura_fd_prof': 80,
            'abertura_fd_larg': 81
        }
        coluna = 5
        for numero, dados in self.fundos_salvos.items():
            for campo, linha in mapeamento.items():
                if campo == 'painel1':
                    valor = dados.get('paineis', [0]*6)[0]
                elif campo == 'painel2':
                    valor = dados.get('paineis', [0]*6)[1]
                elif campo == 'painel3':
                    valor = dados.get('paineis', [0]*6)[2]
                elif campo == 'painel4':
                    valor = dados.get('paineis', [0]*6)[3]
                elif campo == 'painel5':
                    valor = dados.get('paineis', [0]*6)[4]
                elif campo == 'painel6':
                    valor = dados.get('paineis', [0]*6)[5]
                elif campo.startswith('abertura_'):
                    idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.split('_')[1]]
                    idx_tipo = {'dist':0, 'prof':1, 'larg':2}[campo.split('_')[2]]
                    valor = dados.get('aberturas', [[0,0,0]]*4)[idx_lado][idx_tipo]
                else:
                    valor = dados.get(campo, '')
                ws.cell(row=linha, column=coluna, value=valor)
            coluna += 1
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Excel criado em {arquivo}")
    def _calcular_paineis_automatico(self, event=None):
        try:
            largura = float(self.largura_var.get())
        except (tk.TclError, ValueError):
            for v in self.paineis_larguras_vars:
                v.set(0)
            return
        try:
            for v in self.paineis_larguras_vars:
                v.set(0)
            tamanho_inicial = float(self.tipo_painel_inicial_var.get())
            if largura <= tamanho_inicial:
                self.paineis_larguras_vars[0].set(largura)
                return
            tipo_distribuicao = self.tipo_distribuicao_var.get()
            num_paineis = int((largura + 243) // 244)
            largura_restante = largura
            paineis = []
            for i in range(num_paineis):
                if i == num_paineis - 1:  
                    if tipo_distribuicao == "122":
                        if largura_restante < 60 and len(paineis) > 0:
                            paineis[-1] = 122
                            paineis.append(122 + largura_restante)
                        else:
                            paineis.append(largura_restante)
                    elif tipo_distribuicao == "307":
                        if largura_restante < 63 and len(paineis) > 0:
                            paineis[-1] = paineis[-1] + largura_restante
                        else:
                            paineis.append(largura_restante)
                    else:  
                        if largura_restante < 70 and len(paineis) > 0:
                            paineis[-1] = paineis[-1] + largura_restante
                        else:
                            paineis.append(largura_restante)
                else:
                    paineis.append(244)
                    largura_restante -= 244
            for i, valor in enumerate(paineis[:6]):
                self.paineis_larguras_vars[i].set(valor)
        except Exception:
            pass
    def _janela_confirmacao_salvamento(self, msg, largura, altura, aberturas, confirmar_callback):
        win = tk.Toplevel(self)
        win.title("Confirmar Salvamento")
        win.grab_set()
        win.geometry("500x450")
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame_preview = ttk.Frame(win)
        frame_preview.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        fig, ax = plt.subplots(figsize=(3, 3))
        ax.add_patch(plt.Rectangle((0, 0), largura, altura, fill=None, edgecolor='black', lw=2))
        for ab in aberturas:
            ax.add_patch(plt.Rectangle((ab['x'], ab['y']), ab['largura'], ab['altura'], fill=True, color='red', alpha=0.5))
        ax.set_xlim(-10, largura+10)
        ax.set_ylim(-10, altura+10)
        ax.set_aspect('equal')
        ax.axis('off')
        canvas = FigureCanvasTkAgg(fig, master=frame_preview)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)
        canvas.draw()
        frame_texto = ttk.Frame(win)
        frame_texto.pack(side="right", fill="both", expand=True, padx=5, pady=5)
        msg = msg.split("Preview:")[0].strip()
        txt = tk.Text(frame_texto, height=15, width=35, wrap="word")
        txt.insert("1.0", msg)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True)
        frame_botoes = ttk.Frame(frame_texto)
        frame_botoes.pack(fill="x", pady=5)
        def confirmar_e_retornar():
            win.destroy()
            confirmar_callback(True)
        def continuacao_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)  
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima_fracao = numero_base + 0.1  
            while proxima_fracao in numeros_existentes:
                proxima_fracao += 0.1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(f"{proxima_fracao:.1f}")
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def nova_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)  
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():  
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo_numero = numero_base + 1  
            while proximo_numero in numeros_existentes:
                proximo_numero += 1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(str(proximo_numero))
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def cancelar():
            win.destroy()
            confirmar_callback(False)
        style = ttk.Style()
        style.configure('Custom.TButton', padding=5)
        btn_confirmar = ttk.Button(frame_botoes, text="1: Confirmar e Retornar", 
                                 command=confirmar_e_retornar, style='Custom.TButton')
        btn_continuacao = ttk.Button(frame_botoes, text="2: Continuação da Viga", 
                                   command=continuacao_viga, style='Custom.TButton')
        btn_nova = ttk.Button(frame_botoes, text="3: Nova Viga", 
                            command=nova_viga, style='Custom.TButton')
        btn_cancelar = ttk.Button(frame_botoes, text="4: Cancelar", 
                                command=cancelar, style='Custom.TButton')
        btn_confirmar.grid(row=0, column=0, padx=5, pady=2, sticky="ew")
        btn_continuacao.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        btn_nova.grid(row=2, column=0, padx=5, pady=2, sticky="ew")
        btn_cancelar.grid(row=3, column=0, padx=5, pady=2, sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)
    def _importar_excel(self):
        """Importa dados de um arquivo Excel."""
        arquivo = filedialog.askopenfilename(
            title="Selecionar Excel",
            filetypes=[("Excel files", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if not arquivo:
            return
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            mapeamento = {
                'numero': 52,
                'nome': 53,
                'observacoes': 54,
                'pavimento': 55,
                'largura': 56,
                'altura': 57,
                'texto_esquerda': 58,
                'texto_direita': 59,
                'tipo_distribuicao': 60,
                'tipo_painel_inicial': 61,
                'sarrafo_esq': 62,
                'sarrafo_dir': 63,
                'painel1': 64,
                'painel2': 65,
                'painel3': 66,
                'painel4': 67,
                'painel5': 68,
                'painel6': 69,
                'abertura_te_dist': 70,
                'abertura_te_prof': 71,
                'abertura_te_larg': 72,
                'abertura_fe_dist': 73,
                'abertura_fe_prof': 74,
                'abertura_fe_larg': 75,
                'abertura_td_dist': 76,
                'abertura_td_prof': 77,
                'abertura_td_larg': 78,
                'abertura_fd_dist': 79,
                'abertura_fd_prof': 80,
                'abertura_fd_larg': 81
            }
            coluna = 5
            while True:
                numero = ws.cell(row=mapeamento['numero'], column=coluna).value
                if not numero:
                    break
                dados = {}
                for campo, linha in mapeamento.items():
                    valor = ws.cell(row=linha, column=coluna).value
                    if campo.startswith('painel'):
                        if 'paineis' not in dados:
                            dados['paineis'] = [0] * 6
                        idx = int(campo[-1]) - 1
                        dados['paineis'][idx] = float(valor) if valor else 0
                    elif campo.startswith('abertura_'):
                        if 'aberturas' not in dados:
                            dados['aberturas'] = [[0, 0, 0] for _ in range(4)]
                        lado = {'te': 0, 'fe': 1, 'td': 2, 'fd': 3}[campo.split('_')[1]]
                        tipo = {'dist': 0, 'prof': 1, 'larg': 2}[campo.split('_')[2]]
                        dados['aberturas'][lado][tipo] = float(valor) if valor else 0
                    else:
                        dados[campo] = valor
                self.fundos_salvos[str(numero)] = dados
                coluna += 1
            self._atualizar_lista()
            self._salvar_fundos_salvos()
            messagebox.showinfo("Sucesso", "Dados importados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar dados: {str(e)}")
    def _selecionar_excel(self):
        """Seleciona um arquivo Excel para uso posterior."""
        arquivo = filedialog.askopenfilename(
            title="Selecionar Excel",
            filetypes=[("Excel files", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if arquivo:
            self.excel_file = arquivo
            messagebox.showinfo("Sucesso", f"Excel selecionado: {arquivo}")
    def _excluir_fundo(self):
        """Exclui o fundo selecionado da lista."""
        # Verifica seleção na lista 1
        selection1 = self.lista_fundos1.selection()
        selection2 = self.lista_fundos2.selection()
        
        if not selection1 and not selection2:
            return
            
        if selection1:
            item = selection1[0]
            lista = self.lista_fundos1
            indice_numero = 2  # índice da coluna número na lista1
        else:
            item = selection2[0]
            lista = self.lista_fundos2
            indice_numero = 1  # índice da coluna número na lista2
            
        numero = str(lista.item(item)['values'][indice_numero])
        
        if messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir o fundo {numero}?"):
            if numero in self.fundos_salvos:
                del self.fundos_salvos[numero]
                lista.delete(item)
                self._salvar_fundos_salvos()
                self._limpar_campos()
                self._atualizar_lista()  # Atualiza ambas as listas
    def _mostrar_menu_contexto(self, event):
        """Mostra o menu de contexto ao clicar com o botão direito."""
        try:
            self.menu_contexto.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu_contexto.grab_release()
    def _carregar_fundos_salvos(self):
        """Carrega os fundos salvos do arquivo."""
        try:
            if os.path.exists('fundos_salvos.pkl'):
                with open('fundos_salvos.pkl', 'rb') as f:
                    self.fundos_salvos = pickle.load(f)
                self._atualizar_lista()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar fundos salvos: {str(e)}")
    def _salvar_fundos_salvos(self):
        """Salva os fundos no arquivo."""
        try:
            with open('fundos_salvos.pkl', 'wb') as f:
                pickle.dump(self.fundos_salvos, f)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar fundos: {str(e)}")
    def _get_checkbox_state(self, numero):
        if hasattr(self, 'checkbox_selecionados') and numero in self.checkbox_selecionados:
            return '☑'
        return '☐'
    def _on_click_checkbox_col(self, event):
        tree = None
        if event and hasattr(event, 'widget'):
            if event.widget == self.lista_fundos1:
                tree = self.lista_fundos1
            elif event.widget == self.lista_fundos2:
                tree = self.lista_fundos2
        if tree is None:
            return
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            return
        # --- Lógica de reciclagem ---
        if self.modo_reciclagem and tree == self.lista_fundos1 and col == '#2':
            # Só permite se um item da lista 2 estiver selecionado
            if not hasattr(self, 'reciclagem_item_lista2_selecionado') or not self.reciclagem_item_lista2_selecionado:
                from tkinter import messagebox
                messagebox.showerror("Erro", "Selecione primeiro um item na lista 2 para vincular.")
                return 'break'
            id_lista2 = self.reciclagem_item_lista2_selecionado
            id_lista1 = row
            if id_lista1 in self.reciclagem_reverse and self.reciclagem_reverse[id_lista1] != id_lista2:
                from tkinter import messagebox
                nome_vinc = self.fundos_salvos[id_lista1]['nome']
                nome_direita = self.fundos_salvos[self.reciclagem_reverse[id_lista1]]['nome']
                messagebox.showerror("Erro", f"Item já selecionado por '{nome_direita}'!")
                return 'break'
            if id_lista2 in self.reciclagem_vinculos:
                antigo = self.reciclagem_vinculos[id_lista2]
                if antigo in self.reciclagem_reverse:
                    del self.reciclagem_reverse[antigo]
            self.reciclagem_vinculos[id_lista2] = id_lista1
            self.reciclagem_reverse[id_lista1] = id_lista2
            self._atualizar_lista(1)
            return 'break'
        # --- Lógica padrão ---
        if col != '#1':  # Coluna 'Sel'
            return
        if not hasattr(self, 'checkbox_selecionados'):
            self.checkbox_selecionados = set()
        if row in self.checkbox_selecionados:
            self.checkbox_selecionados.remove(row)
        else:
            if len(self.checkbox_selecionados) >= 2:
                return
            self.checkbox_selecionados.add(row)
        self._atualizar_lista()
        return 'break'  
    def _atualizar_alturas_paineis(self, *args):
        try:
            valor = self.altura_geral_var.get()
            for var in self.paineis_alturas_vars:
                var.set(valor)
        except Exception:
            pass
    def _atualizar_alturas2_paineis(self, *args):
        try:
            valor = self.altura_2_geral_var.get()
            for var in self.paineis_alturas2_vars:
                var.set(str(valor))
        except Exception:
            pass
    def _desenhar_laje_suave(self, canvas, x, y, largura, altura, fill=None):
        """Desenha um retângulo de laje com cor cinza escuro e borda cinza clara."""
        canvas.create_rectangle(
            x, y, x + largura, y + altura,
            outline="#222", width=1, fill="#444444"
        )
        if fill:
            canvas.create_rectangle(
                x, y, x + largura, y + altura,
                fill=fill,
                outline="#222"
            )
    def _float_safe(self, var):
        try:
            val = str(var.get()).replace(',', '.')
            if '+' in val:
                partes = val.split('+')
                return sum(float(p.strip()) for p in partes if p.strip())
            return float(val)
        except Exception:
            return 0.0
    def _desenhar_cotas_verticais_segmentos(self, canvas, x_cota, y_base, segmentos, escala, lado='esq'):
        """Desenha cotas verticais para os segmentos (lajes e alturas) na lateral do painel, apenas números, em horizontal."""
        y_atual = y_base
        offset_texto = -20 if lado == 'esq' else 20
        for nome, altura in segmentos:
            if altura > 0:
                y_prox = y_atual - altura * escala
                canvas.create_line(
                    x_cota, y_atual,
                    x_cota, y_prox,
                    fill=self.cores['cota'], arrow="both"
                )
                canvas.create_text(
                    x_cota + offset_texto, (y_atual + y_prox) / 2,
                    text=f"{altura:.1f}",
                    fill=self.cores['texto'], angle=0
                )
                y_atual = y_prox
        print("\nÁreas individuais dos painéis por altura:")
        print("-" * 50)
        for i in range(6):
            largura_painel = self._float_safe(self.paineis_larguras_vars[i])
            altura1 = self._float_safe(self.paineis_alturas_vars[i])
            altura2 = self._float_safe(self.paineis_alturas2_vars[i])
            if largura_painel > 0:
                if altura1 > 0:
                    area1 = largura_painel * altura1 / 10000  
                    print(f"P{i+1} Altura 1: {largura_painel:.1f} x {altura1:.1f} cm = {area1:.2f} m²")
                if altura2 > 0:
                    area2 = largura_painel * altura2 / 10000  
                    print(f"P{i+1} Altura 2: {largura_painel:.1f} x {altura2:.1f} cm = {area2:.2f} m²")
        print("-" * 50)
    def _atualizar_altura2_universal_por_niveis(self, *args):
        try:
            nivel_oposto = float(self.nivel_oposto_var.get())
            nivel_viga = float(self.nivel_viga_var.get())
            altura2 = (nivel_oposto - nivel_viga) * 100
            self.altura_2_geral_var.set(altura2)
            self.grade_altura2_universal.set(round(altura2, 2))
        except Exception:
            pass
    def _cad_safe(func):
        def wrapper(self, *args, **kwargs):
            import tkinter as tk
            from tkinter import messagebox
            try:
                print(f"[DEBUG] Iniciando função '{func.__name__}'")
                resultado = func(self, *args, **kwargs)
            except Exception as e:
                msg = f"Erro ao executar função '{func.__name__}': {str(e)}"
                print(f"[DEBUG] {msg}")
                if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                    print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                    raise
                raise
            else:
                print(f"[DEBUG] Função '{func.__name__}' concluída com sucesso.")
                return resultado
        return wrapper
    @_cad_safe
    def _select_nomes(self):
        import pythoncom
        import win32com.client
        import win32gui
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            balcao = self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.nome_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="O primeiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O primeiro objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_esq_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter", mensagem_erro="O segundo objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O segundo objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_dir_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter", mensagem_erro="O terceiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O terceiro objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_nomes()
            return
        finally:
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_linha(self):
        import pythoncom
        import win32com.client
        import win32gui
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            balcao = self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter")
            self.update()
            try:
                doc.SendCommand("._LINE ")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro=f"Erro ao enviar comando LINE: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao enviar comando LINE: {str(e)}")
            import time
            time.sleep(1.5)
            try:
                last_entity = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
                if last_entity.ObjectName != "AcDbLine":
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro="A última entidade criada não é uma linha!")
                    self.deiconify()
                    raise Exception("A última entidade criada não é uma linha!")
                x1, y1 = last_entity.StartPoint[0], last_entity.StartPoint[1]
                x2, y2 = last_entity.EndPoint[0], last_entity.EndPoint[1]
                comprimento = ((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5
                self.largura_var.set(round(comprimento, 2))
                self._calcular_paineis_automatico()
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro=f"Erro ao obter a linha: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao obter a linha: {str(e)}")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_linha()
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_niveis(self):
        import pythoncom
        import win32com.client
        import win32gui
        from tkinter import messagebox
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o texto da Laje e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            nivel_viga = 0.0
            nivel_oposto = 0.0
            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
            self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
            self.altura_2_geral_var.set(0)
            self.grade_altura2_universal.set(0)
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                print(f"[DEBUG] Falha ao enviar F8 para o AutoCAD: {e}")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto ou linha da Altura da Laje e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            altura_laje = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                except Exception as e:
                    print(f"[DEBUG] messagebox.showwarning: Erro ao acessar ObjectName: {e}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", f"Erro ao acessar o tipo do objeto selecionado: {e}\nSelecione novamente!")
                    self.deiconify()
                    raise Exception(f"Erro ao acessar o tipo do objeto selecionado: {e}")
                if objname == "AcDbLine":
                    altura_laje = 0.0
                elif objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        altura_laje = float(match.group(1).replace(",", ".")) + 2
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        for attrib in entity.GetAttributes():
                            try:
                                texto = attrib.TextString
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    valor_mais_alto = max(valor_mais_alto, valor)
                            except:
                                continue
                        if valor_mais_alto > 0:
                            altura_laje = valor_mais_alto + 2
                        else:
                            print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do bloco!")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] messagebox.showwarning: Erro ao acessar atributos do bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao acessar atributos do bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao acessar atributos do bloco: {e}")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é texto nem linha!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é texto nem linha!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é texto nem linha!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            self.laje_sup_universal.set(altura_laje)
            self.laje_central_alt_universal.set(0)
            balcao = self._mostrar_comentario_flutuante("Selecione o texto da Altura 1 (ex: 20/60 ou 20x60) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            valor2 = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"[x/](\s*[0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        valor2 = float(match.group(1).replace(",", ".")) + 4
                    else:
                        nums = re.findall(r"[0-9]+[\.,]?[0-9]*", texto)
                        if len(nums) >= 2:
                            valor2 = float(nums[1].replace(",", ".")) + 4
                        else:
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o segundo valor do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o segundo valor do texto!")
                else:
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            altura_1 = valor2 - altura_laje
            self.altura_geral_var.set(round(altura_1, 2))
            self.grade_altura1_universal.set(round(altura_1, 2))
            self.deiconify()
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar níveis: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar níveis: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_pilar(self):
        import pythoncom
        import win32com.client
        import win32gui
        import math
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            balcao = self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter")
            self.update()
            try:
                doc.SendCommand("._LINE ")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro=f"Erro ao enviar comando LINE: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao enviar comando LINE: {str(e)}")
            import time
            time.sleep(1.5)
            try:
                last_entity = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
                pontos = []
                if last_entity.ObjectName == "AcDbLine":
                    p1 = last_entity.StartPoint
                    p2 = last_entity.EndPoint
                    pontos = [p1, p2]
                elif last_entity.ObjectName == "AcDbPolyline":
                    n = last_entity.Coordinates.Count // 2
                    for i in range(n):
                        x = last_entity.Coordinates[2*i]
                        y = last_entity.Coordinates[2*i+1]
                        pontos.append((x, y, 0))
                else:
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="A última entidade criada não é uma linha ou polilinha!")
                    self.deiconify()
                    raise Exception("A última entidade criada não é uma linha ou polilinha!")
                if len(pontos) == 2:
                    x1, y1, _ = pontos[0]
                    x2, y2, _ = pontos[1]
                    larg = math.hypot(x2 - x1, y2 - y1)
                    self.detalhe_pilar_esq_vars[0].set(0.0)
                    self.detalhe_pilar_esq_vars[1].set(round(larg, 2))
                elif len(pontos) >= 3:
                    x1, y1, _ = pontos[0]
                    x2, y2, _ = pontos[1]
                    x3, y3, _ = pontos[2]
                    dist = math.hypot(x2 - x1, y2 - y1)
                    larg = math.hypot(x3 - x2, y3 - y2)
                    self.detalhe_pilar_esq_vars[0].set(round(dist, 2))
                    self.detalhe_pilar_esq_vars[1].set(round(larg, 2))
                else:
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="Desenhe uma linha ou polilinha válida!")
                    self.deiconify()
                    raise Exception("Desenhe uma linha ou polilinha válida!")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro=f"Erro ao obter a linha do pilar: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao obter a linha do pilar: {str(e)}")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_pilar()
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_abertura(self, idx=0, instrucao=None):
        import pythoncom
        import win32com.client
        import win32gui
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            instrucao_texto = (instrucao + " (abertura)") if instrucao else "Selecione o texto da abertura (ex: 20/60 ou 20x60) (abertura) e pressione Enter"
            balcao = self._mostrar_comentario_flutuante(instrucao_texto)
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)\s*[x/]\s*([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        largura = float(match.group(1).replace(",", "."))
                        profundidade = float(match.group(2).replace(",", "."))
                        self.aberturas_vars[idx][0].set(0)
                        self.aberturas_vars[idx][1].set(profundidade + 4)
                        self.aberturas_vars[idx][2].set(largura + 8)
                    else:
                        self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Não foi possível extrair os valores do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair os valores do texto!")
                else:
                    self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_abertura(idx, instrucao)
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _abrir_janela_producao(self):
        import tkinter as tk
        from tkinter import ttk
        win = tk.Toplevel(self)
        win.title("Produção Dinâmica - Lateral de Viga")
        win.geometry("520x900")
        win.resizable(True, True)
        win.minsize(400, 600)
        win.grab_set()
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        self._janela_producao = win
        main_canvas = tk.Canvas(win)
        main_scrollbar = ttk.Scrollbar(win, orient="vertical", command=main_canvas.yview)
        frame = ttk.Frame(main_canvas)
        def _on_frame_configure(event):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        def _on_mousewheel(event):
            if event.state & 0x4:  
                return
            try:
                if event.widget == canvas_producao:
                    return  
                main_canvas.yview_scroll(int(-1 * (event.delta/120)), "units")
            except:
                pass
        frame.bind("<Configure>", _on_frame_configure)
        frame_window = main_canvas.create_window((0, 0), window=frame, anchor="nw")
        main_canvas.pack(side="left", fill="both", expand=True)
        main_scrollbar.pack(side="right", fill="y")
        main_canvas.configure(yscrollcommand=main_scrollbar.set)
        win.bind("<MouseWheel>", _on_mousewheel)
        preview_temp_frame = ttk.LabelFrame(frame, text="Visualização", padding=5)
        preview_temp_frame.pack(fill=tk.X, padx=2, pady=2)
        label_nome_producao = tk.Label(preview_temp_frame, text=self.nome_var.get(), font=("Arial", 16, "bold"))
        label_nome_producao.pack(side=tk.TOP, pady=(0, 2))
        def atualizar_nome_producao(*args):
            label_nome_producao.config(text=self.nome_var.get())
        self.nome_var.trace_add('write', atualizar_nome_producao)
        canvas_frame = ttk.Frame(preview_temp_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        canvas_producao = tk.Canvas(canvas_frame, width=900, height=220, bg=self.cores['fundo'])
        scrollbar_y = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas_producao.yview)
        scrollbar_x = ttk.Scrollbar(canvas_frame, orient="horizontal", command=canvas_producao.xview)
        canvas_producao.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        canvas_producao.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)
        global zoom_level_producao, pan_start_producao, pan_x_producao, pan_y_producao
        zoom_level_producao = 1.0
        pan_start_producao = None
        pan_x_producao = 0
        pan_y_producao = 0
        def _on_preview_mousewheel(event):
            widget = event.widget
            if widget == canvas_producao:
                global zoom_level_producao
                if event.delta < 0:
                    zoom_level_producao = max(0.5, zoom_level_producao - 0.1)
                else:
                    zoom_level_producao = min(2.0, zoom_level_producao + 0.1)
                atualizar_preview_producao()
                return "break"  
        canvas_producao.bind("<MouseWheel>", _on_preview_mousewheel)
        def on_window_resize(event=None):
            if event and event.widget == win:
                atualizar_preview_producao()
        win.bind("<Configure>", on_window_resize)
        def _start_pan_producao(event):
            global pan_start_producao
            pan_start_producao = (event.x, event.y)
        def _pan_producao(event):
            global pan_x_producao, pan_y_producao, pan_start_producao
            if pan_start_producao:
                dx = event.x - pan_start_producao[0]
                dy = event.y - pan_start_producao[1]
                pan_x_producao += dx
                pan_y_producao += dy
                pan_start_producao = (event.x, event.y)
                atualizar_preview_producao()
        def _stop_pan_producao(event):
            global pan_start_producao
            pan_start_producao = None
        canvas_producao.bind("<ButtonPress-1>", _start_pan_producao)
        canvas_producao.bind("<B1-Motion>", _pan_producao)
        canvas_producao.bind("<ButtonRelease-1>", _stop_pan_producao)
        def atualizar_preview_producao():
            canvas_producao.delete("all")
            try:
                largura = self._float_safe(self.largura_var)
                altura_geral = self._float_safe(self.altura_geral_var)
            except Exception:
                return
            if largura <= 0 or altura_geral <= 0:
                return
            escala_x = (800 / (largura * 1.2)) * zoom_level_producao
            escala_y = (200 / (altura_geral * 1.2)) * zoom_level_producao
            escala = min(escala_x, escala_y)
            x_inicial = 50 + pan_x_producao
            y_inicial = 200 + pan_y_producao
            self._desenhar_elementos(canvas_producao, x_inicial, y_inicial, largura, altura_geral, escala)
            canvas_producao.configure(scrollregion=canvas_producao.bbox("all"))
        atualizar_preview_producao()
        win.bind_all("<MouseWheel>", _on_mousewheel)
        botoes_frame = ttk.Frame(frame)
        botoes_frame.pack(fill=tk.X, pady=(8, 2))
        btn_finalizar = ttk.Button(botoes_frame, text="Finalizar e Voltar")
        btn_finalizar.pack(side=tk.LEFT, padx=2)
        btn_proxima_viga = ttk.Button(botoes_frame, text="Próxima Viga")
        btn_proxima_viga.pack(side=tk.LEFT, padx=2)
        btn_prox_seg = ttk.Button(botoes_frame, text="Próximo Segmento de Viga")
        btn_prox_seg.pack(side=tk.LEFT, padx=2)
        btn_reiniciar = ttk.Button(botoes_frame, text="Reiniciar Seleção")
        btn_reiniciar.pack(side=tk.LEFT, padx=2)
        def _salvar_e_abrir_combinacoes_corte():
            self._salvar_fundo()
            dados_atual = self._coletar_dados_atual()
            if not self._segmentos_temp_corte or (dados_atual.get('numero') != self._segmentos_temp_corte[-1].get('numero')):
                self._segmentos_temp_corte.append(dados_atual)
            self._abrir_combinacoes_corte()
        btn_combinacoes_corte = ttk.Button(botoes_frame, text="Combinações P/ Corte", command=_salvar_e_abrir_combinacoes_corte)
        btn_combinacoes_corte.pack(side=tk.LEFT, padx=2)
        self._prod_abr_esq_topo = tk.BooleanVar(value=True)
        self._prod_abr_esq_fundo = tk.BooleanVar(value=False)
        self._prod_abr_dir_topo = tk.BooleanVar(value=False)
        self._prod_abr_dir_fundo = tk.BooleanVar(value=False)
        self._prod_pilar_esq = tk.BooleanVar(value=False)
        self._prod_pilar_dir = tk.BooleanVar(value=False)
        self._prod_laje_abaixo = tk.BooleanVar(value=False)
        self._prod_continuacao = tk.StringVar(value="Proxima Parte")
        self._prod_viga_com_nivel = tk.BooleanVar(value=False)
        opcoes_frame = ttk.LabelFrame(frame, text="Opções de Sequência", padding=4)
        opcoes_frame.pack(fill=tk.X, padx=2, pady=4)
        print("[DEBUG] Criando frame de edição de largura")
        frame_edicao_largura = ttk.Frame(frame)
        frame_edicao_largura.pack(fill=tk.X, padx=2, pady=(4, 8))
        ttk.Label(frame_edicao_largura, text="Edição Largura:").pack(side=tk.LEFT, padx=(0,4))
        self.edicao_largura_var = tk.DoubleVar(value=0)
        entry_ed_larg = ttk.Entry(frame_edicao_largura, textvariable=self.edicao_largura_var, width=8)
        entry_ed_larg.pack(side=tk.LEFT, padx=2)
        def atualizar_largura_universal(*args):
            try:
                base = float(self.largura_var.get())
            except Exception:
                base = 0
            try:
                ed = float(self.edicao_largura_var.get())
            except Exception:
                ed = 0
            self.largura_var.set(base + ed)
        self.edicao_largura_var.trace_add('write', atualizar_largura_universal)
        def preencher_rapido(valor):
            self.edicao_largura_var.set(valor)
        botoes = [(-4, "-4"), (-8, "-8"), (11, "+11"), (7, "+7"), (22, "+22")]
        for val, txt in botoes:
            ttk.Button(frame_edicao_largura, text=txt, width=4, command=lambda v=val: preencher_rapido(v)).pack(side=tk.LEFT, padx=1)
        frame.update_idletasks()
        main_canvas.config(scrollregion=main_canvas.bbox("all"))
        grupo1_frame = ttk.LabelFrame(opcoes_frame, text="1 - Opções para próxima Seleção", padding=4)
        grupo1_frame.pack(fill=tk.X, padx=2, pady=(0, 6))
        aberturas_frame = ttk.Frame(grupo1_frame)
        aberturas_frame.pack(anchor="w", fill="x")
        ttk.Label(aberturas_frame, text="Aberturas:").grid(row=0, column=0, columnspan=3, sticky="w", padx=(0,4))
        lados = ["Lado Esquerdo / Topo", "Lado Esquerdo / Fundo", "Lado Direita / Topo", "Lado Direita / Fundo"]
        self._prod_forcar_altura1_vars = [tk.BooleanVar(value=self.forcar_altura1_vars[i].get()) for i in range(4)]
        for i, lado in enumerate(lados):
            ttk.Checkbutton(aberturas_frame, text=lado, variable=getattr(self, f'_prod_abr_{["esq_topo","esq_fundo","dir_topo","dir_fundo"][i]}')).grid(row=i+1, column=0, sticky="w", padx=12, pady=1)
            chk = ttk.Checkbutton(aberturas_frame, text="Forçar Alt.1", variable=self._prod_forcar_altura1_vars[i])
            chk.grid(row=i+1, column=2, sticky="w", padx=2)
            def make_sync(idx):
                def sync_main(*args):
                    if self.forcar_altura1_vars[idx].get() != self._prod_forcar_altura1_vars[idx].get():
                        self.forcar_altura1_vars[idx].set(self._prod_forcar_altura1_vars[idx].get())
                def sync_prod(*args):
                    if self._prod_forcar_altura1_vars[idx].get() != self.forcar_altura1_vars[idx].get():
                        self._prod_forcar_altura1_vars[idx].set(self.forcar_altura1_vars[idx].get())
                self._prod_forcar_altura1_vars[idx].trace_add('write', sync_main)
                self.forcar_altura1_vars[idx].trace_add('write', sync_prod)
            make_sync(i)
        detalhe_pilar_frame = ttk.Frame(grupo1_frame)
        detalhe_pilar_frame.pack(anchor="w", fill="x", pady=(6,0))
        ttk.Label(detalhe_pilar_frame, text="Detalhe Pilar:").grid(row=0, column=0, padx=(0,4))
        ttk.Checkbutton(detalhe_pilar_frame, text="Esquerda", variable=self._prod_pilar_esq).grid(row=0, column=1, padx=2)
        ttk.Checkbutton(detalhe_pilar_frame, text="Direita", variable=self._prod_pilar_dir).grid(row=0, column=2, padx=2)
        viga_nivel_frame = ttk.Frame(grupo1_frame)
        viga_nivel_frame.pack(anchor="w", fill="x", pady=(6,0))
        ttk.Checkbutton(viga_nivel_frame, text="Viga com Nível", variable=self._prod_viga_com_nivel).pack(side=tk.LEFT, padx=2)
        grupo2_frame = ttk.LabelFrame(opcoes_frame, text="2 - Detalhes de Salvamento desta Seleção", padding=4)
        grupo2_frame.pack(fill=tk.X, padx=2, pady=(0, 2))
        cont_frame = ttk.Frame(grupo2_frame)
        cont_frame.pack(anchor="w", pady=(0,0))
        ttk.Label(cont_frame, text="Continuação:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(cont_frame, text="Obstáculo", variable=self._prod_continuacao, value="Obstaculo").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(cont_frame, text="Próxima Parte", variable=self._prod_continuacao, value="Proxima Parte").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(cont_frame, text="Viga Continuação", variable=self._prod_continuacao, value="Viga Continuacao").pack(side=tk.LEFT, padx=2)
        frame_laje_abaixo = ttk.Frame(grupo2_frame)
        frame_laje_abaixo.pack(anchor="w", pady=(6,0), fill="x")
        ttk.Label(frame_laje_abaixo, text="Laje Abaixo:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Checkbutton(frame_laje_abaixo, variable=self._prod_laje_abaixo).pack(side=tk.LEFT)
        frame_tipo1_universal = ttk.Frame(grupo2_frame)
        frame_tipo1_universal.pack(anchor="w", pady=(6,0))
        tipo1_universal_var = tk.StringVar(value=self.tipo1_universal_var.get())
        def set_tipo1_universal():
            self.tipo1_universal_var.set(tipo1_universal_var.get())
            for v in self.paineis_tipo1_vars:
                v.set(tipo1_universal_var.get())
            atualizar_preview_e_canvas()
        ttk.Label(frame_tipo1_universal, text="Paineis Altura 1:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_tipo1_universal, text="Sarrafeado", variable=tipo1_universal_var, value="Sarrafeado", command=set_tipo1_universal).pack(side=tk.LEFT)
        ttk.Radiobutton(frame_tipo1_universal, text="Grade", variable=tipo1_universal_var, value="Grade", command=set_tipo1_universal).pack(side=tk.LEFT)
        entry_grade1 = ttk.Entry(frame_tipo1_universal, textvariable=self.grade_altura1_universal, width=4)
        entry_grade1.pack(side=tk.LEFT, padx=2)
        self.extensao_painel1_var = tk.DoubleVar(value=0)
        ttk.Label(frame_tipo1_universal, text="Extensão de Painel altura 1:").pack(side=tk.LEFT, padx=(8,2))
        entry_ext1 = ttk.Entry(frame_tipo1_universal, textvariable=self.extensao_painel1_var, width=6)
        entry_ext1.pack(side=tk.LEFT, padx=2)
        def atualizar_altura1_e_grade(*args):
            try:
                valor_base = float(self.altura_geral_var.get().split('+')[0].replace(',', '.')) if '+' in self.altura_geral_var.get() else float(self.altura_geral_var.get().replace(',', '.'))
            except Exception:
                valor_base = 0
            try:
                ext = float(self.extensao_painel1_var.get())
            except Exception:
                ext = 0
            if ext > 0:
                self.altura_geral_var.set(f"{valor_base:.2f}+{ext:.2f}")
            else:
                self.altura_geral_var.set(f"{valor_base:.2f}")
            try:
                grade_base = float(self.grade_altura1_universal.get())
            except Exception:
                grade_base = 0
            self.grade_altura1_universal.set(grade_base + ext)
        self.extensao_painel1_var.trace_add('write', atualizar_altura1_e_grade)
        frame_tipo2_universal = ttk.Frame(grupo2_frame)
        frame_tipo2_universal.pack(anchor="w", pady=(2,0))
        tipo2_universal_var = tk.StringVar(value=self.tipo2_universal_var.get())
        def set_tipo2_universal():
            self.tipo2_universal_var.set(tipo2_universal_var.get())
            for v in self.paineis_tipo2_vars:
                v.set(tipo2_universal_var.get())
            atualizar_preview_e_canvas()
        ttk.Label(frame_tipo2_universal, text="Paineis Altura 2:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_tipo2_universal, text="Sarrafeado", variable=tipo2_universal_var, value="Sarrafeado", command=set_tipo2_universal).pack(side=tk.LEFT)
        ttk.Radiobutton(frame_tipo2_universal, text="Grade", variable=tipo2_universal_var, value="Grade", command=set_tipo2_universal).pack(side=tk.LEFT)
        entry_grade2 = ttk.Entry(frame_tipo2_universal, textvariable=self.grade_altura2_universal, width=4)
        entry_grade2.pack(side=tk.LEFT, padx=2)
        self.extensao_painel2_var = tk.DoubleVar(value=0)
        ttk.Label(frame_tipo2_universal, text="Extensão de Painel altura 2:").pack(side=tk.LEFT, padx=(8,2))
        entry_ext2 = ttk.Entry(frame_tipo2_universal, textvariable=self.extensao_painel2_var, width=6)
        entry_ext2.pack(side=tk.LEFT, padx=2)
        def atualizar_altura2_e_grade(*args):
            try:
                valor_base = float(self.altura_2_geral_var.get().split('+')[0].replace(',', '.')) if '+' in self.altura_2_geral_var.get() else float(self.altura_2_geral_var.get().replace(',', '.'))
            except Exception:
                valor_base = 0
            try:
                ext = float(self.extensao_painel2_var.get())
            except Exception:
                ext = 0
            if ext > 0:
                self.altura_2_geral_var.set(f"{valor_base:.2f}+{ext:.2f}")
            else:
                self.altura_2_geral_var.set(f"{valor_base:.2f}")
            try:
                grade_base = float(self.grade_altura2_universal.get())
            except Exception:
                grade_base = 0
            self.grade_altura2_universal.set(grade_base + ext)
        self.extensao_painel2_var.trace_add('write', atualizar_altura2_e_grade)
        frame_lado = ttk.Frame(grupo2_frame)
        frame_lado.pack(anchor="w", pady=(6,0))
        self._prod_lado = tk.StringVar(value="A")
        ttk.Label(frame_lado, text="Lado:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_lado, text="A", variable=self._prod_lado, value="A").pack(side=tk.LEFT)
        ttk.Radiobutton(frame_lado, text="B", variable=self._prod_lado, value="B").pack(side=tk.LEFT)
        frame_pav_obs = ttk.Frame(grupo2_frame)
        frame_pav_obs.pack(anchor="w", pady=(6,0))
        self._prod_pavimento = tk.StringVar(value=self.pavimento_var.get())
        self._prod_obs = tk.StringVar(value=self.obs_var.get())
        self._prod_pe_direito = tk.StringVar(value=self.nivel_pe_direito_var.get())
        ttk.Label(frame_pav_obs, text="Pavimento:").pack(side=tk.LEFT, padx=(0,4))
        entry_pav = ttk.Entry(frame_pav_obs, textvariable=self._prod_pavimento, width=10)
        entry_pav.pack(side=tk.LEFT, padx=(0,8))
        ttk.Label(frame_pav_obs, text="Observações:").pack(side=tk.LEFT, padx=(0,4))
        entry_obs = ttk.Entry(frame_pav_obs, textvariable=self._prod_obs, width=18)
        entry_obs.pack(side=tk.LEFT, padx=(0,8))
        ttk.Label(frame_pav_obs, text="Pé Direito:").pack(side=tk.LEFT, padx=(0,4))
        entry_pe_direito = ttk.Entry(frame_pav_obs, textvariable=self._prod_pe_direito, width=10)
        entry_pe_direito.pack(side=tk.LEFT)
        def sync_pavimento_var(*args):
            if hasattr(self, 'pavimento_var'):
                self.pavimento_var.set(self._prod_pavimento.get())
        def sync_obs_var(*args):
            if hasattr(self, 'obs_var'):
                self.obs_var.set(self._prod_obs.get())
        def sync_pe_direito_var(*args):
            if hasattr(self, 'nivel_pe_direito_var'):
                self.nivel_pe_direito_var.set(self._prod_pe_direito.get())
        self._prod_pavimento.trace_add('write', sync_pavimento_var)
        self._prod_obs.trace_add('write', sync_obs_var)
        self._prod_pe_direito.trace_add('write', sync_pe_direito_var)
        def sync_lado_var(*args):
            if hasattr(self, 'lado_var'):
                self.lado_var.set(self._prod_lado.get())
        self._prod_lado.trace_add('write', sync_lado_var)
        def sync_continuacao_var(*args):
            if hasattr(self, 'continuacao_var'):
                self.continuacao_var.set(self._prod_continuacao.get())
        self._prod_continuacao.trace_add('write', sync_continuacao_var)
        def sync_fundo_var(*args):
            if hasattr(self, 'sarrafo_esq_var'):
                self.sarrafo_esq_var.set(self._prod_laje_abaixo.get())
        self._prod_laje_abaixo.trace_add('write', sync_fundo_var)
        def atualizar_preview_e_canvas():
            self._atualizar_preview()
            atualizar_preview_producao()
        def finalizar_e_voltar():
            nome_atual = self.nome_var.get()
            lado = self._prod_lado.get()
            if not nome_atual.endswith(f'.{lado}'):
                if nome_atual.endswith('.A') or nome_atual.endswith('.B'):
                    nome_atual = nome_atual[:-2]
                self.nome_var.set(f"{nome_atual}.{lado}")
            if self._prod_laje_abaixo.get():
                valor_laje_sup = self.laje_sup_universal.get()
                self.laje_inf_universal.set(valor_laje_sup)
                self.laje_sup_universal.set(0)
            self._salvar_fundo()
            self._atualizar_preview()
            win.destroy()
            self.deiconify()
            try:
                self.bring_to_front()
            except:
                pass
        btn_finalizar.config(command=finalizar_e_voltar)
        def proximo_numero_inteiro():
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numero_base = 1
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo = numero_base + 1
            while proximo in numeros_existentes:
                proximo += 1
            return proximo
        def proxima_fracao():
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numero_base = 1
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima = numero_base + 0.1
            while proxima in numeros_existentes:
                proxima += 0.1
            return round(proxima, 1)
        def limpar_campos_producao():
            numero = self.numero_var.get()
            self._limpar_campos()
            self.numero_var.set(numero)
        def sequencia_producao():
            etapas = []
            etapas.append((self._select_nomes, "Seleção de Nome"))
            etapas.append((self._select_linha, "Seleção de Linha"))
            if self._prod_viga_com_nivel.get():
                etapas.append((self._select_niveis2, "Seleção de Níveis 2"))
            else:
                etapas.append((self._select_niveis, "Seleção de Níveis"))
            aberturas_checks = [
                (self._prod_abr_esq_topo, 0, "Lado Esquerdo / Topo"),
                (self._prod_abr_esq_fundo, 1, "Lado Esquerdo / Fundo"),
                (self._prod_abr_dir_topo, 2, "Lado Direita / Topo"),
                (self._prod_abr_dir_fundo, 3, "Lado Direita / Fundo")
            ]
            for var, idx, label in aberturas_checks:
                if var.get():
                    def make_select_abertura(idx, label):
                        def func():
                            return self._select_abertura(idx, label)
                        return func
                    etapas.append((make_select_abertura(idx, label), f"Seleção de Abertura: {label}"))
            if self._prod_pilar_esq.get():
                etapas.append((self._select_pilar, "Seleção Pilar Esquerda"))
            if self._prod_pilar_dir.get():
                etapas.append((self._select_pilar, "Seleção Pilar Direita"))
            def executar_etapa(idx=0):
                if self._sequencia_finalizada:
                    print("[DEBUG] Sequência interrompida por F8")
                    return
                if idx >= len(etapas):
                    print(f"[DEBUG] Sequência de produção finalizada.")
                    self.continuacao_var.set(self._prod_continuacao.get())
                    self._atualizar_preview_producao()
                    try:
                        win.lift()
                        win.focus_force()
                        win.attributes('-topmost', True)
                        win.after_idle(win.attributes, '-topmost', False)
                    except Exception as e:
                        print(f"[DEBUG] Falha ao trazer janela de produção à frente: {e}")
                    return
                func, nome = etapas[idx]
                print(f"[DEBUG] Iniciando etapa {idx+1}/{len(etapas)}: {nome}")
                try:
                    func()
                    if self._sequencia_finalizada:
                        print("[DEBUG] Sequência interrompida por F8 durante execução")
                        return
                    print(f"[DEBUG] Etapa '{nome}' concluída com sucesso.")
                    self._atualizar_preview_producao()
                    executar_etapa(idx+1)
                except Exception as e:
                    print(f"[DEBUG] Erro na etapa '{nome}': {e}. Repetindo etapa.")
                    if not self._sequencia_finalizada:  
                        executar_etapa(idx)
            executar_etapa(0)
        def proxima_viga():
            self._sequencia_finalizada = False  
            self._salvar_fundo()
            self._segmentos_temp_corte = []
            proximo = proximo_numero_inteiro()
            self.numero_var.set(str(proximo))
            limpar_campos_producao()
            self._atualizar_preview_producao()
            win.lift()
            win.focus_force()
            sequencia_producao()
        def proximo_segmento():
            self._sequencia_finalizada = False  
            self._salvar_fundo()
            dados_segmento = self._coletar_dados_atual()
            self._segmentos_temp_corte.append(dados_segmento)
            prox_frac = proxima_fracao()
            self.numero_var.set(f"{prox_frac:.1f}")
            limpar_campos_producao()
            self._atualizar_preview_producao()
            win.lift()
            win.focus_force()
            sequencia_producao()
        def reiniciar_selecao():
            self._sequencia_finalizada = False  
            sequencia_producao()
        btn_proxima_viga.config(command=proxima_viga)
        btn_prox_seg.config(command=proximo_segmento)
        btn_reiniciar.config(command=reiniciar_selecao)
        btn_finalizar.config(command=finalizar_e_voltar)
        def _salvar_e_abrir_combinacoes_corte():
            self._salvar_fundo()
            dados_atual = self._coletar_dados_atual()
        win.reiniciar_selecao = reiniciar_selecao
        win.sequencia_producao = sequencia_producao
    def _atualizar_preview_producao(self):
        if not hasattr(self, '_producao_canvas'):
            return
        canvas = self._producao_canvas
        canvas.delete("all")
        try:
            largura = self._float_safe(self.largura_var)
            altura_geral = self._float_safe(self.altura_geral_var)
        except Exception:
            return
        if largura <= 0 or altura_geral <= 0:
            return
        escala_x = 300 / (largura * 1.2)
        escala_y = 100 / (altura_geral * 1.2)
        escala = min(escala_x, escala_y)
        x_inicial = 10
        y_inicial = 110
        self._desenhar_elementos(canvas, x_inicial, y_inicial, largura, altura_geral, escala)
    def _mostrar_comentario_flutuante(self, mensagem, mensagem_erro=None):
        import tkinter as tk
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
        balcao = tk.Toplevel()
        balcao.overrideredirect(True)
        balcao.attributes('-topmost', True)
        balcao.configure(bg='#fff8b0')
        # Centralizar na tela
        largura = 480
        altura = 120 if not mensagem_erro else 160
        x = (balcao.winfo_screenwidth() - largura) // 2
        y = 40
        balcao.geometry(f"{largura}x{altura}+{x}+{y}")
        label = tk.Label(balcao, text=mensagem, font=("Arial", 16, "bold"), bg='#fff8b0', fg='#333', wraplength=460)
        label.pack(expand=True, fill=tk.BOTH, padx=10, pady=(10,0))
        if mensagem_erro:
            erro_label = tk.Label(balcao, text=mensagem_erro, font=("Arial", 12, "bold"), bg='#fff8b0', fg='red', wraplength=460)
            erro_label.pack(expand=True, fill=tk.BOTH, padx=10, pady=(0,10))
        def fechar_balao_e_trazer_producao():
            try:
                balcao.destroy()
            except:
                pass
            self._balcao_flutuante = None
            self.bring_to_front()
        btn_cancelar = tk.Button(balcao, text="Cancelar Seleção F8 e Esc", font=("Arial", 12, "bold"), bg="#ffb0b0", command=fechar_balao_e_trazer_producao)
        btn_cancelar.pack(pady=(8, 10))
        self._balcao_flutuante = balcao
        balcao.update()
        return balcao
    def _fechar_comentario_flutuante(self):
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
            self._balcao_flutuante = None
    def _desenhar_hachura_abertura(self, x1, y1, x2, y2, espacamento=6):
        """
        Desenha linhas diagonais (hachura) dentro do retângulo da abertura.
        x1, y1: canto superior esquerdo
        x2, y2: canto inferior direito
        espacamento: distância entre as linhas da hachura
        """
        x1, x2 = min(x1, x2), max(x1, x2)
        y1, y2 = min(y1, y2), max(y1, y2)
        largura = x2 - x1
        altura = y2 - y1
        for i in range(-int(altura), int(largura), espacamento):
            xi = x1 + max(i, 0)
            yi = y2 - max(-i, 0)
            xf = x1 + min(i + altura, largura)
            yf = y2 - min(i + altura, altura)
            self.canvas.create_line(xi, yi, xf, yf, fill="#BDB76B", width=1)

    def bring_to_front(self):
        self._fechar_comentario_flutuante()
        try:
            self.deiconify()
            self.lift()
            self.attributes('-topmost', True)
            self.focus_force()
            self.after_idle(self.attributes, '-topmost', False)
            try:
                import win32gui
                import win32con
                hwnd = self.winfo_id()
                win32gui.SetForegroundWindow(hwnd)
                win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
            except Exception as e:
                print(f'[DEBUG] win32gui bring_to_front falhou: {e}')
            self.after(100, lambda: self.focus_force())
        except Exception as e:
            print(f'[DEBUG] Falha ao trazer janela para frente: {e}')
    def _coletar_dados_atual(self):
        def safe_get(var):
            try:
                return float(var.get())
            except Exception:
                return 0.0
        def safe_get_str(var):
            try:
                return str(var.get())
            except Exception:
                return ''
        return {
            'numero': safe_get_str(self.numero_var),
            'pavimento': safe_get_str(self.pavimento_var),
            'nome': safe_get_str(self.nome_var),
            'texto_esq': safe_get_str(self.texto_esq_var),
            'texto_dir': safe_get_str(self.texto_dir_var),
            'obs': safe_get_str(self.obs_var),
            'largura': safe_get(self.largura_var),
            'altura_geral': safe_get(self.altura_geral_var),
            'paineis_larguras': [safe_get(v) for v in self.paineis_larguras_vars],
            'paineis_alturas': [safe_get(v) for v in self.paineis_alturas_vars],
            'paineis_alturas2': [safe_get(v) for v in self.paineis_alturas2_vars],
            'aberturas': [[safe_get(v) for v in linha] for linha in self.aberturas_vars],
            'sarrafo_esq': bool(self.sarrafo_esq_var.get()),
            'sarrafo_dir': bool(self.sarrafo_dir_var.get()),
            'lajes_sup': [safe_get(v) for v in self.lajes_sup_vars],
            'lajes_inf': [safe_get(v) for v in self.lajes_inf_vars],
            'lajes_central_alt': [safe_get(v) for v in self.lajes_central_alt_vars],
            'detalhe_pilar_esq': [safe_get(v) for v in self.detalhe_pilar_esq_vars],
            'detalhe_pilar_dir': [safe_get(v) for v in self.detalhe_pilar_dir_vars],
            'nivel_oposto': safe_get_str(self.nivel_oposto_var),
            'nivel_viga': safe_get_str(self.nivel_viga_var),
            'nivel_pe_direito': safe_get_str(self.nivel_pe_direito_var),
            'ajuste': safe_get_str(self.ajuste_var),
            'paineis_tipo1': [safe_get_str(v) for v in self.paineis_tipo1_vars],
            'paineis_tipo2': [safe_get_str(v) for v in self.paineis_tipo2_vars],
            'paineis_grade_altura1': [safe_get(v) for v in self.paineis_grade_altura1_vars],
            'paineis_grade_altura2': [safe_get(v) for v in self.paineis_grade_altura2_vars],
            'continuacao': safe_get_str(self.continuacao_var),
            'lado': safe_get_str(self.lado_var),
        }
    def _salvar_combinacoes(self):
        import pickle
        try:
            with open('combinacoes_salvas.pkl', 'wb') as f:
                pickle.dump(self.combinacoes, f)
        except Exception as e:
            print(f"Erro ao salvar combinações: {e}")
    def _carregar_combinacoes(self):
        import os
        import pickle
        if os.path.exists('combinacoes_salvas.pkl'):
            try:
                with open('combinacoes_salvas.pkl', 'rb') as f:
                    self.combinacoes = pickle.load(f)
                self._atualizar_lista_combinacoes()
            except Exception as e:
                print(f"Erro ao carregar combinações: {e}")
    def _toggle_modo_combinacao(self):
        self.modo_combinacao = not self.modo_combinacao
        self.checkbox_selecionados = set()
        self._atualizar_lista()
        self._atualizar_layout_listas()
    def _abrir_combinacoes_corte(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        win = tk.Toplevel(self)
        win.title("Combinações para Corte")
        win.geometry("600x400")
        win.grab_set()
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        frame_listas = ttk.Frame(frame)
        frame_listas.pack(fill=tk.BOTH, expand=True)
        lbl_segmentos = ttk.Label(frame_listas, text="Segmentos")
        lbl_segmentos.grid(row=0, column=0, padx=5, pady=2)
        self.segmentos_corte_vars = []
        self.tree_segmentos_corte = ttk.Treeview(frame_listas, columns=("Sel", "Desc"), show="headings", height=12)
        self.tree_segmentos_corte.heading("Sel", text="Sel")
        self.tree_segmentos_corte.heading("Desc", text="Segmento")
        self.tree_segmentos_corte.column("Sel", width=38, anchor="center")
        self.tree_segmentos_corte.column("Desc", width=180, anchor="w")
        self.tree_segmentos_corte.grid(row=1, column=0, padx=5, pady=2, sticky="nsew")
        self.tree_segmentos_corte.delete(*self.tree_segmentos_corte.get_children())
        self.segmentos_corte_vars.clear()
        for idx, seg in enumerate(self._segmentos_temp_corte):
            var = tk.BooleanVar(value=False)
            self.segmentos_corte_vars.append(var)
            desc = f"{seg.get('numero', '')} - {seg.get('nome', '')}"
            self.tree_segmentos_corte.insert('', 'end', iid=str(idx), values=('☐', desc))
        def on_click_checkbox(event):
            region = self.tree_segmentos_corte.identify("region", event.x, event.y)
            col = self.tree_segmentos_corte.identify_column(event.x)
            row = self.tree_segmentos_corte.identify_row(event.y)
            if region == "cell" and col == '#1' and row:
                idx = int(row)
                atual = self.segmentos_corte_vars[idx].get()
                if not atual and sum(v.get() for v in self.segmentos_corte_vars) >= 2:
                    return
                self.segmentos_corte_vars[idx].set(not atual)
                self.tree_segmentos_corte.set(row, "Sel", '☑' if self.segmentos_corte_vars[idx].get() else '☐')
        self.tree_segmentos_corte.bind("<Button-1>", on_click_checkbox)
        lbl_combinacoes = ttk.Label(frame_listas, text="Combinações")
        lbl_combinacoes.grid(row=0, column=1, padx=5, pady=2)
        self.lista_combinacoes_corte = tk.Listbox(frame_listas, height=12, width=30)
        self.lista_combinacoes_corte.grid(row=1, column=1, padx=5, pady=2, sticky="nsew")
        frame_listas.columnconfigure(0, weight=1)
        frame_listas.columnconfigure(1, weight=1)
        self.lista_combinacoes_corte.delete(0, tk.END)
        for comb in getattr(self, '_combinacoes_temp_corte', []):
            self.lista_combinacoes_corte.insert(tk.END, comb['nome'])
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill=tk.X, pady=8)
        def criar_combinacao():
            selecionados = [i for i, v in enumerate(self.segmentos_corte_vars) if v.get()]
            if len(selecionados) != 2:
                messagebox.showerror("Erro", "Selecione exatamente 2 segmentos para combinar.")
                return
            idx1, idx2 = selecionados
            seg1 = self._segmentos_temp_corte[idx1]
            seg2 = self._segmentos_temp_corte[idx2]
            nome_comb = f"{seg1.get('nome','')} + {seg2.get('nome','')}"
            ids = [seg1.get('numero',''), seg2.get('numero','')]
            pavimento = seg1.get('pavimento','')
            if seg1.get('pavimento','') != seg2.get('pavimento',''):
                messagebox.showerror("Erro", "Só é possível combinar segmentos do MESMO pavimento!")
                return
            comb = {'nome': nome_comb, 'ids': ids, 'pavimento': pavimento}
            if not hasattr(self, '_combinacoes_temp_corte'):
                self._combinacoes_temp_corte = []
            self._combinacoes_temp_corte.append(comb)
            self.lista_combinacoes_corte.insert(tk.END, nome_comb)
            for v in self.segmentos_corte_vars:
                v.set(False)
            for iid in self.tree_segmentos_corte.get_children():
                self.tree_segmentos_corte.set(iid, "Sel", '☐')
        def salvar_combinacoes():
            messagebox.showinfo("Salvo", "Combinações temporárias salvas. Pressione Finalizar para enviar para a lista principal.")
        def finalizar_combinacoes():
            if hasattr(self, '_combinacoes_temp_corte') and self._combinacoes_temp_corte:
                self.combinacoes.extend(self._combinacoes_temp_corte)
                self._salvar_combinacoes()
                self._atualizar_lista_combinacoes()
                self._combinacoes_temp_corte = []
                messagebox.showinfo("Finalizado", "Combinações enviadas para a lista principal.")
            win.destroy()
        btn_criar = ttk.Button(frame_botoes, text="Criar", command=criar_combinacao)
        btn_criar.pack(side=tk.LEFT, padx=5)
        btn_salvar = ttk.Button(frame_botoes, text="Salvar", command=salvar_combinacoes)
        btn_salvar.pack(side=tk.LEFT, padx=5)
        btn_finalizar = ttk.Button(frame_botoes, text="Finalizar", command=finalizar_combinacoes)
        btn_finalizar.pack(side=tk.LEFT, padx=5)
    @_cad_safe
    def _select_niveis2(self):
        import pythoncom
        import win32com.client
        import win32gui
        from tkinter import messagebox
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                print(f"[DEBUG] messagebox.showerror: Nenhum documento AutoCAD ativo encontrado!")
                messagebox.showerror("Erro", "Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                print(f"[DEBUG] Falha ao enviar F8 para o AutoCAD: {e}")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto do Nível Viga e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            nivel_viga = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                    print(f"[DEBUG] Tipo do objeto selecionado: {objname}")
                except Exception as e:
                    print(f"[DEBUG] Erro ao obter ObjectName: {e}")
                    print(f"[DEBUG] Tentando obter mais informações do objeto...")
                    try:
                        print(f"[DEBUG] Handle do objeto: {entity.Handle}")
                    except:
                        pass
                    try:
                        print(f"[DEBUG] Layer do objeto: {entity.Layer}")
                    except:
                        pass
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "Erro ao identificar tipo do objeto. Tente selecionar novamente.")
                    self.deiconify()
                    raise Exception(f"Erro ao identificar tipo do objeto: {e}")
                if objname in ["AcDbText", "AcDbMText"]:
                    try:
                        texto = entity.TextString
                        print(f"[DEBUG] Texto encontrado: {texto}")
                        match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                        if match:
                            nivel_viga = float(match.group(1).replace(",", "."))
                            print(f"[DEBUG] Valor extraído: {nivel_viga}")
                            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
                        else:
                            print(f"[DEBUG] Nenhum número encontrado no texto: {texto}")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do texto!")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao processar texto: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao processar texto: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao processar texto: {e}")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        print(f"[DEBUG] Processando bloco...")
                        try:
                            print(f"[DEBUG] Nome do bloco: {entity.Name}")
                        except Exception as e:
                            print(f"[DEBUG] Não foi possível obter nome do bloco: {e}")
                        try:
                            print(f"[DEBUG] Layer do bloco: {entity.Layer}")
                        except Exception as e:
                            print(f"[DEBUG] Não foi possível obter layer do bloco: {e}")
                        atributos = entity.GetAttributes()
                        print(f"[DEBUG] Número de atributos encontrados: {len(atributos)}")
                        for attrib in atributos:
                            try:
                                tag = attrib.TagString
                                texto = attrib.TextString
                                print(f"[DEBUG] Atributo encontrado - Tag: {tag}, Valor: {texto}")
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    print(f"[DEBUG] Valor numérico extraído: {valor}")
                                    valor_mais_alto = max(valor_mais_alto, valor)
                                else:
                                    print(f"[DEBUG] Nenhum número encontrado no atributo")
                            except Exception as e:
                                print(f"[DEBUG] Erro ao processar atributo: {e}")
                                continue
                        if valor_mais_alto > 0:
                            nivel_viga = valor_mais_alto
                            print(f"[DEBUG] Usando valor mais alto encontrado: {nivel_viga}")
                            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
                        else:
                            print(f"[DEBUG] Nenhum valor numérico encontrado em nenhum atributo")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao processar bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao processar bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao processar bloco: {e}")
                else:
                    print(f"[DEBUG] Objeto não suportado: {objname}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto ou bloco!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto ou bloco!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto do Nível Oposto e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            nivel_oposto = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                objname = entity.ObjectName
                if objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        nivel_oposto = float(match.group(1).replace(",", "."))
                        self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
                        if nivel_oposto >= nivel_viga:
                            altura2 = (nivel_oposto - nivel_viga) * 100
                        else:
                            altura2 = 0
                        self.altura_2_geral_var.set(round(altura2, 2))
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        for attrib in entity.GetAttributes():
                            try:
                                texto = attrib.TextString
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    valor_mais_alto = max(valor_mais_alto, valor)
                            except:
                                continue
                        if valor_mais_alto > 0:
                            nivel_oposto = valor_mais_alto
                            self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
                            if nivel_oposto >= nivel_viga:
                                altura2 = (nivel_oposto - nivel_viga) * 100
                            else:
                                altura2 = 0
                            self.altura_2_geral_var.set(round(altura2, 2))
                        else:
                            print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do bloco!")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] messagebox.showwarning: Erro ao acessar atributos do bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao acessar atributos do bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao acessar atributos do bloco: {e}")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é um texto ou bloco!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto ou bloco!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto ou bloco!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto ou linha da Altura da Laje e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            altura_laje = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                except Exception as e:
                    print(f"[DEBUG] messagebox.showwarning: Erro ao acessar ObjectName: {e}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", f"Erro ao acessar o tipo do objeto selecionado: {e}\nSelecione novamente!")
                    self.deiconify()
                    raise Exception(f"Erro ao acessar o tipo do objeto selecionado: {e}")
                if objname == "AcDbLine":
                    altura_laje = 0.0
                elif objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        altura_laje = float(match.group(1).replace(",", ".")) + 2
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é texto nem linha!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é texto nem linha!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é texto nem linha!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto da Altura 1 (ex: 20/60 ou 20x60) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            valor2 = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"[x/](\s*[0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        valor2 = float(match.group(1).replace(",", ".")) + 4
                    else:
                        nums = re.findall(r"[0-9]+[\.,]?[0-9]*", texto)
                        if len(nums) >= 2:
                            valor2 = float(nums[1].replace(",", ".")) + 4
                        else:
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o segundo valor do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o segundo valor do texto!")
                else:
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            diferenca_niveis = nivel_oposto - nivel_viga
            if diferenca_niveis < 0:
                diferenca_niveis = 0
            altura_1 = valor2 - altura_laje - (diferenca_niveis * 100)
            self.altura_geral_var.set(round(altura_1, 2))
            self.grade_altura1_universal.set(round(altura_1, 2))
            self.deiconify()
            if nivel_oposto <= nivel_viga:
                self.laje_sup_universal.set(altura_laje)
                self.laje_central_alt_universal.set(0)
            else:
                self.laje_central_alt_universal.set(altura_laje)
                self.laje_sup_universal.set(0)
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar níveis: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar níveis: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _gerar_script(self):
        """Gera o script SCR para o fundo selecionado na lista."""
        tree = None
        if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
            tree = self.lista_fundos1
        elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
            tree = self.lista_fundos2
        if not tree:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        selection = tree.selection()
        if not selection:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        if not dados:
            messagebox.showerror("Erro", f"Dados do item {numero} não encontrados.")
            return
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")
        pavimento = str(dados.get('pavimento', 'SEM_PAVIMENTO')).strip() or 'SEM_PAVIMENTO'
        diretorio_saida = os.path.join(diretorio_base, pavimento)
        if not os.path.exists(diretorio_saida):
            os.makedirs(diretorio_saida)
        try:
            script, caminho_arquivo = gerar_script_viga(dados, diretorio_saida)
            if script and caminho_arquivo:
                messagebox.showinfo("Sucesso", f"Script gerado com sucesso em: {caminho_arquivo}")
            elif script and not caminho_arquivo:
                nome = dados.get('nome', 'SemNome')
                observacoes = dados.get('obs', '')
                arquivo_nome = f"{nome}{'_' + observacoes if observacoes else ''}.scr"
                caminho_arquivo = os.path.join(diretorio_saida, arquivo_nome)
                resposta = messagebox.askyesno("Arquivo Existente", 
                                            f"O arquivo {arquivo_nome} já existe. Deseja sobrescrever?")
                if resposta:
                    with open(caminho_arquivo, "w", encoding="utf-16") as f:
                        f.write(script)
                    messagebox.showinfo("Sucesso", f"Script gerado com sucesso em: {caminho_arquivo}")
            else:
                messagebox.showerror("Erro", "Não foi possível gerar o script.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar o script: {str(e)}")
    def _ao_fechar(self):
        """Método chamado quando a aplicação é fechada para salvar os dados automaticamente."""
        try:
            self._salvar_fundos_salvos()
            self._salvar_combinacoes()
            print("Dados salvos com sucesso ao fechar a aplicação.")
        except Exception as e:
            print(f"Erro ao salvar dados durante o fechamento: {str(e)}")
        self.destroy()
    def _salvar_teste(self):
        """Gera o script SCR para o item selecionado e salva em um arquivo de teste fixo."""
        tree = None
        if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
            tree = self.lista_fundos1
        elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
            tree = self.lista_fundos2
        if not tree:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        selection = tree.selection()
        if not selection:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        if not dados:
            messagebox.showerror("Erro", f"Dados do item {numero} não encontrados.")
            return
        caminho_arquivo = r"C:\Users\rvene\Desktop\Automacao_cad\Vigas\A_B\Ferramentas\TESTE_VIGA_TV.scr"
        pasta_ferramentas = os.path.dirname(caminho_arquivo)
        if not os.path.exists(pasta_ferramentas):
            os.makedirs(pasta_ferramentas)
        try:
            script, _ = gerar_script_viga(dados, "temp")
            if script:
                with open(caminho_arquivo, "w", encoding="utf-16") as f:
                    f.write(script)
                messagebox.showinfo("Sucesso", f"Script de teste salvo com sucesso em:\n{caminho_arquivo}")
            else:
                messagebox.showerror("Erro", "Não foi possível gerar o script.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar o script de teste: {str(e)}")
    def _registrar_f8_global(self):
        def f8_handler(e=None):
            print('[DEBUG] F8 pressionado: iniciando cancelamento global')
            if not self._f8_cancelando:
                self._f8_cancelando = True
                self._sequencia_finalizada = True  
                self._cancelar_selecao_global()
        keyboard.on_press_key('f8', lambda e: f8_handler())
    def _cancelar_selecao_global(self):
        print('[DEBUG] _cancelar_selecao_global executado')
        self._fechar_comentario_flutuante()
        self.deiconify()
        self.bring_to_front()
        self._f8_cancelando = False
        if hasattr(self, '_janela_producao') and self._janela_producao:
            try:
                if self._sequencia_finalizada:
                    self._janela_producao.destroy()
                    self._janela_producao = None
                else:
                    self.after(100, lambda: self._janela_producao.reiniciar_selecao())
            except:
                pass
    def _gerar_todos_scripts(self):
        import os
        from tkinter import messagebox, Toplevel, StringVar
        import subprocess
        from tkinter import ttk
        pavimentos = set()
        for dados in self.fundos_salvos.values():
            pav = str(dados.get('pavimento', '')).strip()
            if pav:
                pavimentos.add(pav)
        pavimentos = sorted(pavimentos)
        if not pavimentos:
            messagebox.showerror("Erro", "Nenhum pavimento encontrado nos fundos salvos!")
            return
        selected_pavimento = StringVar()
        dialog = Toplevel(self)
        dialog.title("Selecionar Pavimento")
        dialog.geometry("300x120")
        dialog.grab_set()
        ttk.Label(dialog, text="Escolha o pavimento para gerar os scripts:").pack(pady=10)
        combo = ttk.Combobox(dialog, values=pavimentos, textvariable=selected_pavimento, state="readonly")
        combo.pack(pady=5)
        combo.current(0)
        confirmed = {'ok': False}
        def confirmar():
            confirmed['ok'] = True
            dialog.destroy()
        def cancelar():
            dialog.destroy()
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=confirmar).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=cancelar).pack(side=tk.LEFT, padx=10)
        dialog.wait_window()
        pavimento = selected_pavimento.get()
        if not confirmed['ok'] or not pavimento or pavimento not in pavimentos:
            messagebox.showwarning("Cancelado", "Operação cancelada ou pavimento inválido.")
            return
        total = 0
        erros = []
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")
        diretorio_saida = os.path.join(diretorio_base, pavimento)
        if not os.path.exists(diretorio_saida):
            os.makedirs(diretorio_saida)
        from gerador_script_viga import gerar_script_viga
        for numero, dados in self.fundos_salvos.items():
            try:
                pav = str(dados.get('pavimento', '')).strip()
                if pav != pavimento:
                    continue
                script, caminho_arquivo = gerar_script_viga(dados, diretorio_saida)
                if script and caminho_arquivo:
                    total += 1
            except Exception as e:
                erros.append(f"{numero}: {e}")
        if not erros:
            messagebox.showinfo("Sucesso", f"Todos os scripts ({total}) do pavimento '{pavimento}' foram gerados!")
        else:
            messagebox.showwarning("Concluído com Erros", f"{total} scripts gerados. Erros:\n" + '\n'.join(erros))
        try:
            combinador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Combinador_VIGA.py")
            subprocess.run(["python", combinador_path, diretorio_saida], check=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao executar o Combinador: {e}")
            return
        try:
            pasta_combinados = os.path.join(diretorio_saida, "Combinados")
            ordenador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Ordenador_VIGA.py")
            subprocess.run(["python", ordenador_path, pasta_combinados], check=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao executar o Ordenador: {e}")
            return
    def _combinar_scripts(self):
        import subprocess
        from tkinter import filedialog, messagebox
        import os
        import sys
        diretorio = filedialog.askdirectory(title="Selecione a pasta SCRIPTS para combinar")
        if not diretorio:
            return
        script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'Combinador_VIGA.py'))
        try:
            subprocess.run([sys.executable, script_path], check=True, cwd=os.path.dirname(script_path))
            messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao combinar scripts: {e}")
    def _ordenar_scripts(self):
        import subprocess
        from tkinter import filedialog, messagebox
        import os
        import sys
        diretorio = filedialog.askdirectory(title="Selecione a pasta Combinados para ordenar")
        if not diretorio:
            return
        script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'Ordenador_VIGA.py'))
        try:
            subprocess.run([sys.executable, script_path], check=True, cwd=os.path.dirname(script_path))
            messagebox.showinfo("Sucesso", "Scripts ordenados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ordenar scripts: {e}")
    def _calcular_ajuste(self, *args):
        """Calcula o ajuste baseado nos níveis e pé direito"""
        try:
            nivel_viga = float(self.nivel_viga_var.get() or 0)
            nivel_oposto = float(self.nivel_oposto_var.get() or 0)
            pe_direito = float(self.nivel_pe_direito_var.get() or 0)
            if nivel_viga > 0 and nivel_oposto > 0 and pe_direito > 0:
                nivel_maior = max(nivel_viga, nivel_oposto)
                ajuste = (pe_direito - nivel_maior) * 100
                self.ajuste_var.set(f"{ajuste:.2f}")
            else:
                self.ajuste_var.set("")
        except (ValueError, TypeError):
            self.ajuste_var.set("")
    def _area_intersecao(self, rx1, ry1, rx2, ry2, ax1, ay1, ax2, ay2):
        """Calcula a área de interseção entre dois retângulos."""
        inter_x1 = max(rx1, ax1)
        inter_y1 = max(ry1, ay1)
        inter_x2 = min(rx2, ax2)
        inter_y2 = min(ry2, ay2)
        if inter_x2 > inter_x1 and inter_y2 > inter_y1:
            return (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
        return 0
    def _ordenar_lista(self, coluna, lista2=False):
        """Ordena a lista de acordo com a coluna clicada."""
        # Inverte a ordem se clicar na mesma coluna
        reverso = False
        if self.ordem_atual['coluna'] == coluna:
            reverso = not self.ordem_atual['reverso']
        
        # Atualiza a ordem atual
        self.ordem_atual = {'coluna': coluna, 'reverso': reverso}
        
        # Seleciona a lista correta
        lista = self.lista_fundos2 if lista2 else self.lista_fundos1
        
        # Obtém todos os itens da lista
        itens = [(lista.set(item, coluna), item) for item in lista.get_children('')]
        
        # Função para converter número para ordenação numérica
        def converter_numero(valor):
            try:
                return float(valor)
            except:
                return valor
        
        # Ordena os itens
        if coluna == "Número":
            itens.sort(key=lambda x: converter_numero(x[0]), reverse=reverso)
        else:
            itens.sort(key=lambda x: x[0].lower(), reverse=reverso)
        
        # Reordena a lista
        for index, (val, item) in enumerate(itens):
            lista.move(item, '', index)
            # Atualiza as cores zebradas
            lista.tag_configure('odd', background='#e8f5e9')
            lista.tag_configure('even', background='#ffffff')
            lista.item(item, tags=('odd' if index % 2 == 0 else 'even',))
if __name__ == "__main__":
    app = FundoProducaoApp()
    app.mainloop()