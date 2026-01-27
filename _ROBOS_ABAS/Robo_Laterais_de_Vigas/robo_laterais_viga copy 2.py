import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import os
import json
import openpyxl
import win32com.client
import pythoncom
import win32gui
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import time
import re
import math
import pyautogui
import keyboard
import sys
import shutil
from gerador_script_viga import gerar_script_viga
from preview_combinacao import PreviewCombinacao
from gerador_script_combinados import gerar_script_teste_combinados, GeradorScriptCombinados
from A_B.preview_principal import PreviewPrincipal

    def _excluir_combinacao(self):
        sel = self.lista_combinacoes.selection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.combinacoes):
            return
        import tkinter.messagebox as mb
        if not mb.askyesno("Excluir", f"Deseja realmente excluir a combinação '{self.combinacoes[idx]['nome']}'?"):
            return
        del self.combinacoes[idx]
        self._atualizar_lista_combinacoes()
        self._salvar_combinacoes()
    def _selecionar_textos_consecutivos(self):
        print('[DEBUG] _selecionar_textos_consecutivos chamado')
        if getattr(self, '_f8_cancelando', False):
            print('[DEBUG] _selecionar_textos_consecutivos abortado por F8')
            self._f8_cancelando = False
            return
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                print(f"[DEBUG] messagebox.showerror: Nenhum documento AutoCAD ativo encontrado!")
                messagebox.showerror("Erro", "Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            info_label = tk.Label(self, 
                                text="Selecione o primeiro texto (Nome) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.nome_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O primeiro objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O primeiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O primeiro objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            info_label = tk.Label(self, 
                                text="Selecione o segundo texto (Texto Esquerda) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_esq_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O segundo objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O segundo objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O segundo objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            info_label = tk.Label(self, 
                                text="Selecione o terceiro texto (Texto Direita) e pressione Enter",
                                font=("Arial", 10, "bold"),
                                fg="blue",
                                bg="light yellow")
            info_label.place(relx=0.5, rely=0.1, anchor="center")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            info_label.destroy()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_dir_var.set(entity.TextString)
                else:
                    print(f"[DEBUG] messagebox.showwarning: O terceiro objeto selecionado não é um texto!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O terceiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O terceiro objeto selecionado não é um texto!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar textos: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar textos: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _salvar_fundo(self):
        """Salva os dados do fundo atual."""
        nome_atual = self.nome_var.get()
        lado = None
        if hasattr(self, 'lado_var'):
            lado = self.lado_var.get()
        elif hasattr(self, '_prod_lado'):
            lado = self._prod_lado.get()
        if nome_atual and lado and not nome_atual.endswith(f'.{lado}'):
            if nome_atual.endswith('.A') or nome_atual.endswith('.B'):
                nome_atual = nome_atual[:-2]
            self.nome_var.set(f"{nome_atual}.{lado}")
        if getattr(self, '_carregando_fundo', False):
            return
        numero = self.numero_var.get()
        if not numero:
            from tkinter import messagebox
            messagebox.showerror("Erro", "O campo 'Número' está vazio! Preencha para salvar.")
            return
        try:
            dados = {
                'numero': numero,
                'pavimento': self.pavimento_var.get(),
                'nome': self.nome_var.get(),
                'texto_esq': self.texto_esq_var.get(),
                'fundo': self.fundo_var.get(),
                'texto_dir': self.texto_dir_var.get(),
                'obs': self.obs_var.get(),
                'largura': self._float_safe(self.largura_var),
                'altura_geral': self.altura_geral_var.get(),
                'paineis_larguras': [self._float_safe(v) for v in self.paineis_larguras_vars],
                'paineis_alturas': [v.get() for v in self.paineis_alturas_vars],
                'paineis_alturas2': [v.get() for v in self.paineis_alturas2_vars],
                'sarrafo_esq': self.sarrafo_esq_var.get(),
                'sarrafo_dir': self.sarrafo_dir_var.get(),
                'sarrafo_alt2_esq': self.sarrafo_alt2_esq_var.get(),
                'sarrafo_alt2_dir': self.sarrafo_alt2_dir_var.get(),
                'aberturas': [[self._float_safe(var) for var in linha] for linha in self.aberturas_vars],
                'tipo_painel_inicial': self.tipo_painel_inicial_var.get(),
                'tipo_distribuicao': self.tipo_distribuicao_var.get(),
                'lajes_sup': [self._float_safe(v) for v in self.lajes_sup_vars],
                'lajes_inf': [self._float_safe(v) for v in self.lajes_inf_vars],
                'lajes_central_alt': [self._float_safe(v) for v in self.lajes_central_alt_vars],
                'detalhe_pilar_esq': [self._float_safe(v) for v in self.detalhe_pilar_esq_vars],
                'detalhe_pilar_dir': [self._float_safe(v) for v in self.detalhe_pilar_dir_vars],
                'nivel_oposto': self.nivel_oposto_var.get(),
                'nivel_viga': self.nivel_viga_var.get(),
                'nivel_pe_direito': self.nivel_pe_direito_var.get(),
                'ajuste': self.ajuste_var.get(),
                'paineis_tipo1': [v.get() for v in self.paineis_tipo1_vars],
                'paineis_tipo2': [v.get() for v in self.paineis_tipo2_vars],
                'paineis_grade_altura1': [v.get() for v in self.paineis_grade_altura1_vars],
                'paineis_grade_altura2': [v.get() for v in self.paineis_grade_altura2_vars],
                'continuacao': self.continuacao_var.get(),
                'lado': self.lado_var.get(),
            }
            self.fundos_salvos[numero] = dados
            self._salvar_fundos_salvos()
            print(f"Fundo '{numero}' salvo com sucesso.")
            self._atualizar_lista()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Erro ao salvar fundo", str(e))
    def _carregar_fundo_selecionado(self, event=None):
        if self.modo_reciclagem and event and hasattr(event, 'widget') and event.widget == self.lista_fundos2:
            sel = self.lista_fundos2.selection()
            if sel:
                self.reciclagem_item_lista2_selecionado = sel[0]
                self._atualizar_lista(1)
        tree = None
        if event and hasattr(event, 'widget'):
            if event.widget == self.lista_fundos1:
                tree = self.lista_fundos1
            elif event.widget == self.lista_fundos2:
                tree = self.lista_fundos2
        if tree is None:
            if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
                tree = self.lista_fundos1
            elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
                tree = self.lista_fundos2
        if tree is None:
            return
        selection = tree.selection()
        if not selection:
            self._atualizar_preview()  
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        self.numero_var.set(dados.get('numero', numero))
        self.nome_var.set(dados.get('nome', ''))
        self.obs_var.set(dados.get('obs', ''))
        self.pavimento_var.set(dados.get('pavimento', ''))
        try:
            self.largura_var.set(float(dados.get('largura', 0)))
        except Exception:
            self.largura_var.set(0)
        try:
            self.altura_geral_var.set(str(dados.get('altura_geral', '0')))
        except Exception:
            self.altura_geral_var.set('0')
        self.texto_esq_var.set(dados.get('texto_esq', ''))
        self.fundo_var.set(dados.get('fundo', ''))
        self.texto_dir_var.set(dados.get('texto_dir', ''))
        self.tipo_distribuicao_var.set(dados.get('tipo_distribuicao', '122'))
        self.tipo_painel_inicial_var.set(dados.get('tipo_painel_inicial', '300'))
        self.sarrafo_esq_var.set(dados.get('sarrafo_esq', True))
        self.sarrafo_dir_var.set(dados.get('sarrafo_dir', True))
        self.sarrafo_alt2_esq_var.set(dados.get('sarrafo_alt2_esq', False))
        self.sarrafo_alt2_dir_var.set(dados.get('sarrafo_alt2_dir', False))
        paineis_larguras = dados.get('paineis_larguras', [0]*6)
        paineis_alturas = dados.get('paineis_alturas', [self.altura_geral_var.get()]*6)
        paineis_alturas2 = dados.get('paineis_alturas2', ["0"]*6)
        for i in range(6):
            try:
                self.paineis_larguras_vars[i].set(float(paineis_larguras[i]))
            except Exception:
                self.paineis_larguras_vars[i].set(0)
            try:
                self.paineis_alturas_vars[i].set(str(paineis_alturas[i]))
            except Exception:
                self.paineis_alturas_vars[i].set('')
            try:
                self.paineis_alturas2_vars[i].set(str(paineis_alturas2[i]))
            except Exception:
                self.paineis_alturas2_vars[i].set('')
        aberturas = dados.get('aberturas', [[0,0,0] for _ in range(4)])
        for i, abertura in enumerate(aberturas):
            for j in range(3):
                try:
                    self.aberturas_vars[i][j].set(float(abertura[j]))
                except Exception:
                    self.aberturas_vars[i][j].set(0)
        lajes_sup = dados.get('lajes_sup', [0]*6)
        lajes_inf = dados.get('lajes_inf', [0]*6)
        lajes_central_alt = dados.get('lajes_central_alt', [0]*6)
        for i in range(6):
            try:
                self.lajes_sup_vars[i].set(float(lajes_sup[i] or 0))
            except Exception:
                self.lajes_sup_vars[i].set(0)
            try:
                self.lajes_inf_vars[i].set(float(lajes_inf[i] or 0))
            except Exception:
                self.lajes_inf_vars[i].set(0)
            try:
                self.lajes_central_alt_vars[i].set(float(lajes_central_alt[i] or 0))
            except Exception:
                self.lajes_central_alt_vars[i].set(0)
        detalhe_pilar_esq = dados.get('detalhe_pilar_esq', [0, 0])
        detalhe_pilar_dir = dados.get('detalhe_pilar_dir', [0, 0])
        for i in range(2):
            try:
                self.detalhe_pilar_esq_vars[i].set(float(detalhe_pilar_esq[i]))
            except Exception:
                self.detalhe_pilar_esq_vars[i].set(0)
            try:
                self.detalhe_pilar_dir_vars[i].set(float(detalhe_pilar_dir[i]))
            except Exception:
                self.detalhe_pilar_dir_vars[i].set(0)
        self.nivel_oposto_var.set(dados.get('nivel_oposto', ''))
        self.nivel_viga_var.set(dados.get('nivel_viga', ''))
        self.nivel_pe_direito_var.set(dados.get('nivel_pe_direito', ''))
        self.ajuste_var.set(dados.get('ajuste', ''))
        paineis_tipo1 = dados.get('paineis_tipo1', [False]*6)
        paineis_tipo2 = dados.get('paineis_tipo2', [False]*6)
        paineis_grade_altura1 = dados.get('paineis_grade_altura1', [False]*6)
        paineis_grade_altura2 = dados.get('paineis_grade_altura2', [False]*6)
        for i in range(6):
            self.paineis_tipo1_vars[i].set(paineis_tipo1[i])
            self.paineis_tipo2_vars[i].set(paineis_tipo2[i])
            self.paineis_grade_altura1_vars[i].set(paineis_grade_altura1[i])
            self.paineis_grade_altura2_vars[i].set(paineis_grade_altura2[i])
        self.continuacao_var.set(dados.get('continuacao', False))
        self.lado_var.set(dados.get('lado', 'A'))
        self._atualizar_preview()
    def _limpar_campos(self):
        """Limpa todos os campos do formulário."""
        self.numero_var.set('')
        self.pavimento_var.set('')
        self.nome_var.set('')
        self.texto_esq_var.set('')
        self.fundo_var.set('')
        self.texto_dir_var.set('')
        self.obs_var.set('')
        self.largura_var.set('')
        self.altura_geral_var.set('')
        self.sarrafo_esq_var.set(False)
        self.sarrafo_dir_var.set(False)
        for linha in self.aberturas_vars:
            for var in linha:
                var.set('')
        for var in self.paineis_larguras_vars:
            var.set('')
        for var in self.paineis_alturas_vars:
            var.set('')
        for var in self.paineis_alturas2_vars:
            var.set('')
        for var in self.lajes_sup_vars:
            var.set('')
        for var in self.lajes_inf_vars:
            var.set('')
        for var in self.lajes_central_alt_vars:
                var.set('')
        self.tipo_painel_inicial_var.set('300')
        self._atualizar_preview()
    def _exportar_excel(self):
        """Exporta os dados para um arquivo Excel."""
        if not self.fundos_salvos:
            messagebox.showerror("Erro", "Não há dados para exportar!")
            return
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Salvar como"
        )
        if not arquivo:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fundos"
            cabecalhos = [
                'Número', 'Pavimento', 'Nome', 'Texto Esquerda', 'Texto Direita',
                'Observação', 'Comprimento', 'Largura', 'P1', 'P2', 'P3',
                'Sarrafo Esquerda', 'Sarrafo Direita', 'Tipo Painel Inicial',
                'Nível Oposto', 'Nível Viga'
            ]
            for i, cabecalho in enumerate(cabecalhos, 1):
                ws.cell(row=1, column=i, value=cabecalho)
            for i, abertura in enumerate(['Abertura 1', 'Abertura 2', 'Abertura 3', 'Abertura 4'], 1):
                ws.cell(row=1, column=len(cabecalhos) + i, value=abertura)
            row = 2
            for numero, dados in sorted(self.fundos_salvos.items()):
                ws.cell(row=row, column=1, value=numero)
                ws.cell(row=row, column=2, value=dados.get('pavimento', ''))
                ws.cell(row=row, column=3, value=dados.get('nome', ''))
                ws.cell(row=row, column=4, value=dados.get('texto_esq', ''))
                ws.cell(row=row, column=5, value=dados.get('texto_dir', ''))
                ws.cell(row=row, column=6, value=dados.get('obs', ''))
                ws.cell(row=row, column=7, value=dados.get('comprimento', ''))
                ws.cell(row=row, column=8, value=dados.get('largura', ''))
                ws.cell(row=row, column=9, value=dados.get('p1', ''))
                ws.cell(row=row, column=10, value=dados.get('p2', ''))
                ws.cell(row=row, column=11, value=dados.get('p3', ''))
                ws.cell(row=row, column=12, value=dados.get('sarrafo_esq', False))
                ws.cell(row=row, column=13, value=dados.get('sarrafo_dir', False))
                ws.cell(row=row, column=14, value=dados.get('tipo_painel_inicial', '300'))
                ws.cell(row=row, column=15, value=dados.get('nivel_oposto', ''))
                ws.cell(row=row, column=16, value=dados.get('nivel_viga', ''))
                aberturas = dados.get('aberturas', [[0, 0, 0] for _ in range(4)])
                for i, abertura in enumerate(aberturas, 1):
                    ws.cell(row=row, column=len(cabecalhos) + i, value=str(abertura))
                row += 1
            wb.save(arquivo)
            messagebox.showinfo("Sucesso", "Dados exportados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar dados: {str(e)}")
    def _criar_excel(self):
        arquivo = filedialog.asksaveasfilename(title="Salvar Novo Excel", filetypes=[("Excel", "*.xlsx")], defaultextension=".xlsx")
        if not arquivo:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        mapeamento = {
            'numero': 52,
            'nome': 53,
            'observacoes': 54,
            'pavimento': 55,
            'largura': 56,
            'altura': 57,
            'texto_esquerda': 58,
            'texto_direita': 59,
            'tipo_distribuicao': 60,
            'tipo_painel_inicial': 61,
            'sarrafo_esq': 62,
            'sarrafo_dir': 63,
            'painel1': 64,
            'painel2': 65,
            'painel3': 66,
            'painel4': 67,
            'painel5': 68,
            'painel6': 69,
            'abertura_te_dist': 70,
            'abertura_te_prof': 71,
            'abertura_te_larg': 72,
            'abertura_fe_dist': 73,
            'abertura_fe_prof': 74,
            'abertura_fe_larg': 75,
            'abertura_td_dist': 76,
            'abertura_td_prof': 77,
            'abertura_td_larg': 78,
            'abertura_fd_dist': 79,
            'abertura_fd_prof': 80,
            'abertura_fd_larg': 81
        }
        coluna = 5
        for numero, dados in self.fundos_salvos.items():
            for campo, linha in mapeamento.items():
                if campo == 'painel1':
                    valor = dados.get('paineis', [0]*6)[0]
                elif campo == 'painel2':
                    valor = dados.get('paineis', [0]*6)[1]
                elif campo == 'painel3':
                    valor = dados.get('paineis', [0]*6)[2]
                elif campo == 'painel4':
                    valor = dados.get('paineis', [0]*6)[3]
                elif campo == 'painel5':
                    valor = dados.get('paineis', [0]*6)[4]
                elif campo == 'painel6':
                    valor = dados.get('paineis', [0]*6)[5]
                elif campo.startswith('abertura_'):
                    idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.split('_')[1]]
                    idx_tipo = {'dist':0, 'prof':1, 'larg':2}[campo.split('_')[2]]
                    valor = dados.get('aberturas', [[0,0,0]]*4)[idx_lado][idx_tipo]
                else:
                    valor = dados.get(campo, '')
                ws.cell(row=linha, column=coluna, value=valor)
            coluna += 1
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Excel criado em {arquivo}")
    def _calcular_paineis_automatico(self, event=None):
        try:
            largura = float(self.largura_var.get())
        except (tk.TclError, ValueError):
            for v in self.paineis_larguras_vars:
                v.set(0)
            return
        try:
            for v in self.paineis_larguras_vars:
                v.set(0)
            tamanho_inicial = float(self.tipo_painel_inicial_var.get())
            if largura <= tamanho_inicial:
                self.paineis_larguras_vars[0].set(largura)
                return
            tipo_distribuicao = self.tipo_distribuicao_var.get()
            num_paineis = int((largura + 243) // 244)
            largura_restante = largura
            paineis = []
            for i in range(num_paineis):
                if i == num_paineis - 1:  
                    if tipo_distribuicao == "122":
                        if largura_restante < 60 and len(paineis) > 0:
                            paineis[-1] = 122
                            paineis.append(122 + largura_restante)
                        else:
                            paineis.append(largura_restante)
                    elif tipo_distribuicao == "307":
                        if largura_restante < 63 and len(paineis) > 0:
                            paineis[-1] = paineis[-1] + largura_restante
                        else:
                            paineis.append(largura_restante)
                    else:  
                        if largura_restante < 70 and len(paineis) > 0:
                            paineis[-1] = paineis[-1] + largura_restante
                        else:
                            paineis.append(largura_restante)
                else:
                    paineis.append(244)
                    largura_restante -= 244
            for i, valor in enumerate(paineis[:6]):
                self.paineis_larguras_vars[i].set(valor)
        except Exception:
            pass
    def _janela_confirmacao_salvamento(self, msg, largura, altura, aberturas, confirmar_callback):
        win = tk.Toplevel(self)
        win.title("Confirmar Salvamento")
        win.grab_set()
        win.geometry("500x450")
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame_preview = ttk.Frame(win)
        frame_preview.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        fig, ax = plt.subplots(figsize=(3, 3))
        ax.add_patch(plt.Rectangle((0, 0), largura, altura, fill=None, edgecolor='black', lw=2))
        for ab in aberturas:
            ax.add_patch(plt.Rectangle((ab['x'], ab['y']), ab['largura'], ab['altura'], fill=True, color='red', alpha=0.5))
        ax.set_xlim(-10, largura+10)
        ax.set_ylim(-10, altura+10)
        ax.set_aspect('equal')
        ax.axis('off')
        canvas = FigureCanvasTkAgg(fig, master=frame_preview)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill="both", expand=True)
        canvas.draw()
        frame_texto = ttk.Frame(win)
        frame_texto.pack(side="right", fill="both", expand=True, padx=5, pady=5)
        msg = msg.split("Preview:")[0].strip()
        txt = tk.Text(frame_texto, height=15, width=35, wrap="word")
        txt.insert("1.0", msg)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True)
        frame_botoes = ttk.Frame(frame_texto)
        frame_botoes.pack(fill="x", pady=5)
        def confirmar_e_retornar():
            win.destroy()
            confirmar_callback(True)
        def continuacao_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)  
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima_fracao = numero_base + 0.1  
            while proxima_fracao in numeros_existentes:
                proxima_fracao += 0.1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(f"{proxima_fracao:.1f}")
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def nova_viga():
            win.destroy()
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)  
            except (ValueError, TypeError):
                messagebox.showerror("Erro", "O campo número está vazio ou inválido")
                return
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():  
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo_numero = numero_base + 1  
            while proximo_numero in numeros_existentes:
                proximo_numero += 1
            confirmar_callback(True)
            self._salvar_fundo()
            pavimento_atual = self.pavimento_var.get()
            self._limpar_campos()
            self.numero_var.set(str(proximo_numero))
            self.pavimento_var.set(pavimento_atual)
            self._selecionar_textos_consecutivos()
        def cancelar():
            win.destroy()
            confirmar_callback(False)
        style = ttk.Style()
        style.configure('Custom.TButton', padding=5)
        btn_confirmar = ttk.Button(frame_botoes, text="1: Confirmar e Retornar", 
                                 command=confirmar_e_retornar, style='Custom.TButton')
        btn_continuacao = ttk.Button(frame_botoes, text="2: Continuação da Viga", 
                                   command=continuacao_viga, style='Custom.TButton')
        btn_nova = ttk.Button(frame_botoes, text="3: Nova Viga", 
                            command=nova_viga, style='Custom.TButton')
        btn_cancelar = ttk.Button(frame_botoes, text="4: Cancelar", 
                                command=cancelar, style='Custom.TButton')
        btn_confirmar.grid(row=0, column=0, padx=5, pady=2, sticky="ew")
        btn_continuacao.grid(row=1, column=0, padx=5, pady=2, sticky="ew")
        btn_nova.grid(row=2, column=0, padx=5, pady=2, sticky="ew")
        btn_cancelar.grid(row=3, column=0, padx=5, pady=2, sticky="ew")
        frame_botoes.grid_columnconfigure(0, weight=1)
    def _importar_excel(self):
        """Importa dados de um arquivo Excel."""
        arquivo = filedialog.askopenfilename(
            title="Selecionar Excel",
            filetypes=[("Excel files", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if not arquivo:
            return
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            mapeamento = {
                'numero': 52,
                'nome': 53,
                'observacoes': 54,
                'pavimento': 55,
                'largura': 56,
                'altura': 57,
                'texto_esquerda': 58,
                'texto_direita': 59,
                'tipo_distribuicao': 60,
                'tipo_painel_inicial': 61,
                'sarrafo_esq': 62,
                'sarrafo_dir': 63,
                'painel1': 64,
                'painel2': 65,
                'painel3': 66,
                'painel4': 67,
                'painel5': 68,
                'painel6': 69,
                'abertura_te_dist': 70,
                'abertura_te_prof': 71,
                'abertura_te_larg': 72,
                'abertura_fe_dist': 73,
                'abertura_fe_prof': 74,
                'abertura_fe_larg': 75,
                'abertura_td_dist': 76,
                'abertura_td_prof': 77,
                'abertura_td_larg': 78,
                'abertura_fd_dist': 79,
                'abertura_fd_prof': 80,
                'abertura_fd_larg': 81
            }
            coluna = 5
            while True:
                numero = ws.cell(row=mapeamento['numero'], column=coluna).value
                if not numero:
                    break
                dados = {}
                for campo, linha in mapeamento.items():
                    valor = ws.cell(row=linha, column=coluna).value
                    if campo.startswith('painel'):
                        if 'paineis' not in dados:
                            dados['paineis'] = [0] * 6
                        idx = int(campo[-1]) - 1
                        dados['paineis'][idx] = float(valor) if valor else 0
                    elif campo.startswith('abertura_'):
                        if 'aberturas' not in dados:
                            dados['aberturas'] = [[0, 0, 0] for _ in range(4)]
                        lado = {'te': 0, 'fe': 1, 'td': 2, 'fd': 3}[campo.split('_')[1]]
                        tipo = {'dist': 0, 'prof': 1, 'larg': 2}[campo.split('_')[2]]
                        dados['aberturas'][lado][tipo] = float(valor) if valor else 0
                    else:
                        dados[campo] = valor
                self.fundos_salvos[str(numero)] = dados
                coluna += 1
            self._atualizar_lista()
            self._salvar_fundos_salvos()
            messagebox.showinfo("Sucesso", "Dados importados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar dados: {str(e)}")
    def _selecionar_excel(self):
        """Seleciona um arquivo Excel para uso posterior."""
        arquivo = filedialog.askopenfilename(
            title="Selecionar Excel",
            filetypes=[("Excel files", "*.xlsx")],
            defaultextension=".xlsx"
        )
        if arquivo:
            self.excel_file = arquivo
            messagebox.showinfo("Sucesso", f"Excel selecionado: {arquivo}")
    def _excluir_fundo(self):
        """Exclui o fundo selecionado da lista."""
        # Verifica seleção na lista 1
        selection1 = self.lista_fundos1.selection()
        selection2 = self.lista_fundos2.selection()
        
        if not selection1 and not selection2:
            return
            
        if selection1:
            item = selection1[0]
            lista = self.lista_fundos1
            indice_numero = 2  # índice da coluna número na lista1
        else:
            item = selection2[0]
            lista = self.lista_fundos2
            indice_numero = 1  # índice da coluna número na lista2
            
        numero = str(lista.item(item)['values'][indice_numero])
        
        if messagebox.askyesno("Confirmar Exclusão", f"Deseja realmente excluir o fundo {numero}?"):
            if numero in self.fundos_salvos:
                del self.fundos_salvos[numero]
                lista.delete(item)
                self._salvar_fundos_salvos()
                self._limpar_campos()
                self._atualizar_lista()  # Atualiza ambas as listas
    def _mostrar_menu_contexto(self, event):
        """Mostra o menu de contexto ao clicar com o botão direito."""
        try:
            self.menu_contexto.tk_popup(event.x_root, event.y_root)
        finally:
            self.menu_contexto.grab_release()
    def _carregar_fundos_salvos(self):
        """Carrega os fundos salvos do arquivo."""
        try:
            if os.path.exists('fundos_salvos.json'):
                with open('fundos_salvos.json', 'r', encoding='utf-8') as f:

                    self.fundos_salvos  = json.load(f)
                self._atualizar_lista()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar fundos salvos: {str(e)}")
    def _salvar_fundos_salvos(self):
        """Salva os fundos no arquivo."""
        try:
            with open('fundos_salvos.json', 'w', encoding='utf-8') as f:

                json.dump(self.fundos_salvos, f, indent=4, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar fundos: {str(e)}")
    def _get_checkbox_state(self, numero):
        if hasattr(self, 'checkbox_selecionados') and numero in self.checkbox_selecionados:
            return '☑'
        return '☐'
    def _on_click_checkbox_col(self, event):
        tree = None
        if event and hasattr(event, 'widget'):
            if event.widget == self.lista_fundos1:
                tree = self.lista_fundos1
            elif event.widget == self.lista_fundos2:
                tree = self.lista_fundos2
        if tree is None:
            return
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        if not row:
            return
        # --- Lógica de reciclagem ---
        if self.modo_reciclagem and tree == self.lista_fundos1 and col == '#2':
            # Só permite se um item da lista 2 estiver selecionado
            if not hasattr(self, 'reciclagem_item_lista2_selecionado') or not self.reciclagem_item_lista2_selecionado:
                from tkinter import messagebox
                messagebox.showerror("Erro", "Selecione primeiro um item na lista 2 para vincular.")
                return 'break'
            id_lista2 = self.reciclagem_item_lista2_selecionado
            id_lista1 = row
            if id_lista1 in self.reciclagem_reverse and self.reciclagem_reverse[id_lista1] != id_lista2:
                from tkinter import messagebox
                nome_vinc = self.fundos_salvos[id_lista1]['nome']
                nome_direita = self.fundos_salvos[self.reciclagem_reverse[id_lista1]]['nome']
                messagebox.showerror("Erro", f"Item já selecionado por '{nome_direita}'!")
                return 'break'
            if id_lista2 in self.reciclagem_vinculos:
                antigo = self.reciclagem_vinculos[id_lista2]
                if antigo in self.reciclagem_reverse:
                    del self.reciclagem_reverse[antigo]
            self.reciclagem_vinculos[id_lista2] = id_lista1
            self.reciclagem_reverse[id_lista1] = id_lista2
            self._atualizar_lista(1)
            return 'break'
        # --- Lógica padrão ---
        if col != '#1':  # Coluna 'Sel'
            return
        if not hasattr(self, 'checkbox_selecionados'):
            self.checkbox_selecionados = set()
        if row in self.checkbox_selecionados:
            self.checkbox_selecionados.remove(row)
        else:
            if len(self.checkbox_selecionados) >= 2:
                return
            self.checkbox_selecionados.add(row)
        self._atualizar_lista()
        return 'break'  
    def _atualizar_alturas_paineis(self, *args):
        try:
            valor = self.altura_geral_var.get()
            for var in self.paineis_alturas_vars:
                var.set(valor)
        except Exception:
            pass
    def _atualizar_alturas2_paineis(self, *args):
        try:
            valor = self.altura_2_geral_var.get()
            for var in self.paineis_alturas2_vars:
                var.set(str(valor))
        except Exception:
            pass
    def _float_safe(self, var):
        try:
            val = str(var.get()).replace(',', '.')
            if '+' in val:
                partes = val.split('+')
                return sum(float(p.strip()) for p in partes if p.strip())
            return float(val)
        except Exception:
            return 0.0
    def _atualizar_altura2_universal_por_niveis(self, *args):
        try:
            nivel_oposto = float(self.nivel_oposto_var.get())
            nivel_viga = float(self.nivel_viga_var.get())
            altura2 = (nivel_oposto - nivel_viga) * 100
            self.altura_2_geral_var.set(altura2)
            self.grade_altura2_universal.set(round(altura2, 2))
        except Exception:
            pass
    def _cad_safe(func):
        def wrapper(self, *args, **kwargs):
            import tkinter as tk
            from tkinter import messagebox
            try:
                print(f"[DEBUG] Iniciando função '{func.__name__}'")
                resultado = func(self, *args, **kwargs)
            except Exception as e:
                msg = f"Erro ao executar função '{func.__name__}': {str(e)}"
                print(f"[DEBUG] {msg}")
                if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                    print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                    raise
                raise
            else:
                print(f"[DEBUG] Função '{func.__name__}' concluída com sucesso.")
                return resultado
        return wrapper
    @_cad_safe
    def _select_nomes(self):
        import pythoncom
        import win32com.client
        import win32gui
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            balcao = self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.nome_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="O primeiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O primeiro objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o primeiro texto (Nome) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_esq_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter", mensagem_erro="O segundo objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O segundo objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o segundo texto (Texto Esquerda) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    self.texto_dir_var.set(entity.TextString)
                else:
                    self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter", mensagem_erro="O terceiro objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O terceiro objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o terceiro texto (Texto Direita) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_nomes()
            return
        finally:
            try:
                selection.Clear()
                selection.Delete()
            except:
                pass
            self.deiconify()
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_linha(self):
        import pythoncom
        import win32com.client
        import win32gui
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            balcao = self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter")
            self.update()
            try:
                doc.SendCommand("._LINE ")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro=f"Erro ao enviar comando LINE: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao enviar comando LINE: {str(e)}")
            import time
            time.sleep(1.5)
            try:
                last_entity = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
                if last_entity.ObjectName != "AcDbLine":
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro="A última entidade criada não é uma linha!")
                    self.deiconify()
                    raise Exception("A última entidade criada não é uma linha!")
                x1, y1 = last_entity.StartPoint[0], last_entity.StartPoint[1]
                x2, y2 = last_entity.EndPoint[0], last_entity.EndPoint[1]
                comprimento = ((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5
                self.largura_var.set(round(comprimento, 2))
                self._calcular_paineis_automatico()
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) e pressione Enter", mensagem_erro=f"Erro ao obter a linha: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao obter a linha: {str(e)}")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_linha()
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_niveis(self):
        import pythoncom
        import win32com.client
        import win32gui
        from tkinter import messagebox
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o texto da Laje e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            nivel_viga = 0.0
            nivel_oposto = 0.0
            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
            self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
            self.altura_2_geral_var.set(0)
            self.grade_altura2_universal.set(0)
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                print(f"[DEBUG] Falha ao enviar F8 para o AutoCAD: {e}")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto ou linha da Altura da Laje e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            altura_laje = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                except Exception as e:
                    print(f"[DEBUG] messagebox.showwarning: Erro ao acessar ObjectName: {e}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", f"Erro ao acessar o tipo do objeto selecionado: {e}\nSelecione novamente!")
                    self.deiconify()
                    raise Exception(f"Erro ao acessar o tipo do objeto selecionado: {e}")
                if objname == "AcDbLine":
                    altura_laje = 0.0
                elif objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        altura_laje = float(match.group(1).replace(",", ".")) + 2
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        for attrib in entity.GetAttributes():
                            try:
                                texto = attrib.TextString
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    valor_mais_alto = max(valor_mais_alto, valor)
                            except:
                                continue
                        if valor_mais_alto > 0:
                            altura_laje = valor_mais_alto + 2
                        else:
                            print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do bloco!")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] messagebox.showwarning: Erro ao acessar atributos do bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao acessar atributos do bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao acessar atributos do bloco: {e}")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é texto nem linha!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é texto nem linha!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é texto nem linha!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            self.laje_sup_universal.set(altura_laje)
            self.laje_central_alt_universal.set(0)
            balcao = self._mostrar_comentario_flutuante("Selecione o texto da Altura 1 (ex: 20/60 ou 20x60) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            valor2 = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"[x/](\s*[0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        valor2 = float(match.group(1).replace(",", ".")) + 4
                    else:
                        nums = re.findall(r"[0-9]+[\.,]?[0-9]*", texto)
                        if len(nums) >= 2:
                            valor2 = float(nums[1].replace(",", ".")) + 4
                        else:
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o segundo valor do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o segundo valor do texto!")
                else:
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            altura_1 = valor2 - altura_laje
            self.altura_geral_var.set(round(altura_1, 2))
            self.grade_altura1_universal.set(round(altura_1, 2))
            self.deiconify()
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar níveis: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar níveis: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_pilar(self):
        import pythoncom
        import win32com.client
        import win32gui
        import math
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            balcao = self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter")
            self.update()
            try:
                doc.SendCommand("._LINE ")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro=f"Erro ao enviar comando LINE: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao enviar comando LINE: {str(e)}")
            import time
            time.sleep(1.5)
            try:
                last_entity = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
                pontos = []
                if last_entity.ObjectName == "AcDbLine":
                    p1 = last_entity.StartPoint
                    p2 = last_entity.EndPoint
                    pontos = [p1, p2]
                elif last_entity.ObjectName == "AcDbPolyline":
                    n = last_entity.Coordinates.Count // 2
                    for i in range(n):
                        x = last_entity.Coordinates[2*i]
                        y = last_entity.Coordinates[2*i+1]
                        pontos.append((x, y, 0))
                else:
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="A última entidade criada não é uma linha ou polilinha!")
                    self.deiconify()
                    raise Exception("A última entidade criada não é uma linha ou polilinha!")
                if len(pontos) == 2:
                    x1, y1, _ = pontos[0]
                    x2, y2, _ = pontos[1]
                    larg = math.hypot(x2 - x1, y2 - y1)
                    self.detalhe_pilar_esq_vars[0].set(0.0)
                    self.detalhe_pilar_esq_vars[1].set(round(larg, 2))
                elif len(pontos) >= 3:
                    x1, y1, _ = pontos[0]
                    x2, y2, _ = pontos[1]
                    x3, y3, _ = pontos[2]
                    dist = math.hypot(x2 - x1, y2 - y1)
                    larg = math.hypot(x3 - x2, y3 - y2)
                    self.detalhe_pilar_esq_vars[0].set(round(dist, 2))
                    self.detalhe_pilar_esq_vars[1].set(round(larg, 2))
                else:
                    self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro="Desenhe uma linha ou polilinha válida!")
                    self.deiconify()
                    raise Exception("Desenhe uma linha ou polilinha válida!")
            except Exception as e:
                self._mostrar_comentario_flutuante("Desenhe uma linha (2 pontos) ou polilinha (3 pontos) para o pilar e pressione Enter", mensagem_erro=f"Erro ao obter a linha do pilar: {str(e)}")
                self.deiconify()
                raise Exception(f"Erro ao obter a linha do pilar: {str(e)}")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_pilar()
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    @_cad_safe
    def _select_abertura(self, idx=0, instrucao=None):
        import pythoncom
        import win32com.client
        import win32gui
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Nenhum documento AutoCAD ativo encontrado!")
                self.deiconify()
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        pass
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                pass
            instrucao_texto = (instrucao + " (abertura)") if instrucao else "Selecione o texto da abertura (ex: 20/60 ou 20x60) (abertura) e pressione Enter"
            balcao = self._mostrar_comentario_flutuante(instrucao_texto)
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)\s*[x/]\s*([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        largura = float(match.group(1).replace(",", "."))
                        profundidade = float(match.group(2).replace(",", "."))
                        self.aberturas_vars[idx][0].set(0)
                        self.aberturas_vars[idx][1].set(profundidade + 4)
                        self.aberturas_vars[idx][2].set(largura + 8)
                    else:
                        self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Não foi possível extrair os valores do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair os valores do texto!")
                else:
                    self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self._mostrar_comentario_flutuante("Selecione o texto da abertura (ex: 20/60 ou 20x60) e pressione Enter", mensagem_erro="Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            self.deiconify()
        except Exception as e:
            import time
            time.sleep(1.2)
            if hasattr(self, '_sequencia_finalizada') and self._sequencia_finalizada:
                print("[DEBUG] Sequência finalizada por F8, não tentando novamente")
                raise
            self._select_abertura(idx, instrucao)
            return
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _abrir_janela_producao(self):
        import tkinter as tk
        from tkinter import ttk
        win = tk.Toplevel(self)
        win.title("Produção Dinâmica - Lateral de Viga")
        win.geometry("520x900")
        win.resizable(True, True)
        win.minsize(400, 600)
        win.grab_set()
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        self._janela_producao = win
        main_canvas = tk.Canvas(win)
        main_scrollbar = ttk.Scrollbar(win, orient="vertical", command=main_canvas.yview)
        frame = ttk.Frame(main_canvas)
        def _on_frame_configure(event):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        def _on_mousewheel(event):
            if event.state & 0x4:  
                return
            try:
                if event.widget == canvas_producao:
                    return  
                main_canvas.yview_scroll(int(-1 * (event.delta/120)), "units")
            except:
                pass
        frame.bind("<Configure>", _on_frame_configure)
        frame_window = main_canvas.create_window((0, 0), window=frame, anchor="nw")
        main_canvas.pack(side="left", fill="both", expand=True)
        main_scrollbar.pack(side="right", fill="y")
        main_canvas.configure(yscrollcommand=main_scrollbar.set)
        win.bind("<MouseWheel>", _on_mousewheel)
        preview_temp_frame = ttk.LabelFrame(frame, text="Visualização", padding=5)
        preview_temp_frame.pack(fill=tk.X, padx=2, pady=2)
        label_nome_producao = tk.Label(preview_temp_frame, text=self.nome_var.get(), font=("Arial", 16, "bold"))
        label_nome_producao.pack(side=tk.TOP, pady=(0, 2))
        def atualizar_nome_producao(*args):
            label_nome_producao.config(text=self.nome_var.get())
        self.nome_var.trace_add('write', atualizar_nome_producao)
        canvas_frame = ttk.Frame(preview_temp_frame)
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        canvas_producao = tk.Canvas(canvas_frame, width=900, height=220, bg=self.cores['fundo'])
        scrollbar_y = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas_producao.yview)
        scrollbar_x = ttk.Scrollbar(canvas_frame, orient="horizontal", command=canvas_producao.xview)
        canvas_producao.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        canvas_producao.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)
        global zoom_level_producao, pan_start_producao, pan_x_producao, pan_y_producao
        zoom_level_producao = 1.0
        pan_start_producao = None
        pan_x_producao = 0
        pan_y_producao = 0
        def _on_preview_mousewheel(event):
            widget = event.widget
            if widget == canvas_producao:
                global zoom_level_producao
                if event.delta < 0:
                    zoom_level_producao = max(0.5, zoom_level_producao - 0.1)
                else:
                    zoom_level_producao = min(2.0, zoom_level_producao + 0.1)
                atualizar_preview_producao()
                return "break"  
        canvas_producao.bind("<MouseWheel>", _on_preview_mousewheel)
        def on_window_resize(event=None):
            if event and event.widget == win:
                atualizar_preview_producao()
        win.bind("<Configure>", on_window_resize)
        def _start_pan_producao(event):
            global pan_start_producao
            pan_start_producao = (event.x, event.y)
        def _pan_producao(event):
            global pan_x_producao, pan_y_producao, pan_start_producao
            if pan_start_producao:
                dx = event.x - pan_start_producao[0]
                dy = event.y - pan_start_producao[1]
                pan_x_producao += dx
                pan_y_producao += dy
                pan_start_producao = (event.x, event.y)
                atualizar_preview_producao()
        def _stop_pan_producao(event):
            global pan_start_producao
            pan_start_producao = None
        canvas_producao.bind("<ButtonPress-1>", _start_pan_producao)
        canvas_producao.bind("<B1-Motion>", _pan_producao)
        canvas_producao.bind("<ButtonRelease-1>", _stop_pan_producao)
        def atualizar_preview_producao():
            canvas_producao.delete("all")
            try:
                largura = self._float_safe(self.largura_var)
                altura_geral = self._float_safe(self.altura_geral_var)
            except Exception:
                return
            if largura <= 0 or altura_geral <= 0:
                return
            escala_x = (800 / (largura * 1.2)) * zoom_level_producao
            escala_y = (200 / (altura_geral * 1.2)) * zoom_level_producao
            escala = min(escala_x, escala_y)
            x_inicial = 50 + pan_x_producao
            y_inicial = 200 + pan_y_producao
            self._desenhar_elementos(canvas_producao, x_inicial, y_inicial, largura, altura_geral, escala)
            canvas_producao.configure(scrollregion=canvas_producao.bbox("all"))
        atualizar_preview_producao()
        win.bind_all("<MouseWheel>", _on_mousewheel)
        botoes_frame = ttk.Frame(frame)
        botoes_frame.pack(fill=tk.X, pady=(8, 2))
        btn_finalizar = ttk.Button(botoes_frame, text="Finalizar e Voltar")
        btn_finalizar.pack(side=tk.LEFT, padx=2)
        btn_proxima_viga = ttk.Button(botoes_frame, text="Próxima Viga")
        btn_proxima_viga.pack(side=tk.LEFT, padx=2)
        btn_prox_seg = ttk.Button(botoes_frame, text="Próximo Segmento de Viga")
        btn_prox_seg.pack(side=tk.LEFT, padx=2)
        btn_reiniciar = ttk.Button(botoes_frame, text="Reiniciar Seleção")
        btn_reiniciar.pack(side=tk.LEFT, padx=2)
        def _salvar_e_abrir_combinacoes_corte():
            self._salvar_fundo()
            dados_atual = self._coletar_dados_atual()
            if not self._segmentos_temp_corte or (dados_atual.get('numero') != self._segmentos_temp_corte[-1].get('numero')):
                self._segmentos_temp_corte.append(dados_atual)
            self._abrir_combinacoes_corte()
        btn_combinacoes_corte = ttk.Button(botoes_frame, text="Combinações P/ Corte", command=_salvar_e_abrir_combinacoes_corte)
        btn_combinacoes_corte.pack(side=tk.LEFT, padx=2)
        self._prod_abr_esq_topo = tk.BooleanVar(value=True)
        self._prod_abr_esq_fundo = tk.BooleanVar(value=False)
        self._prod_abr_dir_topo = tk.BooleanVar(value=False)
        self._prod_abr_dir_fundo = tk.BooleanVar(value=False)
        self._prod_pilar_esq = tk.BooleanVar(value=False)
        self._prod_pilar_dir = tk.BooleanVar(value=False)
        self._prod_laje_abaixo = tk.BooleanVar(value=False)
        self._prod_continuacao = tk.StringVar(value="Proxima Parte")
        self._prod_viga_com_nivel = tk.BooleanVar(value=False)
        opcoes_frame = ttk.LabelFrame(frame, text="Opções de Sequência", padding=4)
        opcoes_frame.pack(fill=tk.X, padx=2, pady=4)
        print("[DEBUG] Criando frame de edição de largura")
        frame_edicao_largura = ttk.Frame(frame)
        frame_edicao_largura.pack(fill=tk.X, padx=2, pady=(4, 8))
        ttk.Label(frame_edicao_largura, text="Edição Largura:").pack(side=tk.LEFT, padx=(0,4))
        self.edicao_largura_var = tk.DoubleVar(value=0)
        entry_ed_larg = ttk.Entry(frame_edicao_largura, textvariable=self.edicao_largura_var, width=8)
        entry_ed_larg.pack(side=tk.LEFT, padx=2)
        def atualizar_largura_universal(*args):
            try:
                base = float(self.largura_var.get())
            except Exception:
                base = 0
            try:
                ed = float(self.edicao_largura_var.get())
            except Exception:
                ed = 0
            self.largura_var.set(base + ed)
        self.edicao_largura_var.trace_add('write', atualizar_largura_universal)
        def preencher_rapido(valor):
            self.edicao_largura_var.set(valor)
        botoes = [(-4, "-4"), (-8, "-8"), (11, "+11"), (7, "+7"), (22, "+22")]
        for val, txt in botoes:
            ttk.Button(frame_edicao_largura, text=txt, width=4, command=lambda v=val: preencher_rapido(v)).pack(side=tk.LEFT, padx=1)
        frame.update_idletasks()
        main_canvas.config(scrollregion=main_canvas.bbox("all"))
        grupo1_frame = ttk.LabelFrame(opcoes_frame, text="1 - Opções para próxima Seleção", padding=4)
        grupo1_frame.pack(fill=tk.X, padx=2, pady=(0, 6))
        aberturas_frame = ttk.Frame(grupo1_frame)
        aberturas_frame.pack(anchor="w", fill="x")
        ttk.Label(aberturas_frame, text="Aberturas:").grid(row=0, column=0, columnspan=3, sticky="w", padx=(0,4))
        lados = ["Lado Esquerdo / Topo", "Lado Esquerdo / Fundo", "Lado Direita / Topo", "Lado Direita / Fundo"]
        self._prod_forcar_altura1_vars = [tk.BooleanVar(value=self.forcar_altura1_vars[i].get()) for i in range(4)]
        for i, lado in enumerate(lados):
            ttk.Checkbutton(aberturas_frame, text=lado, variable=getattr(self, f'_prod_abr_{["esq_topo","esq_fundo","dir_topo","dir_fundo"][i]}')).grid(row=i+1, column=0, sticky="w", padx=12, pady=1)
            chk = ttk.Checkbutton(aberturas_frame, text="Forçar Alt.1", variable=self._prod_forcar_altura1_vars[i])
            chk.grid(row=i+1, column=2, sticky="w", padx=2)
            def make_sync(idx):
                def sync_main(*args):
                    if self.forcar_altura1_vars[idx].get() != self._prod_forcar_altura1_vars[idx].get():
                        self.forcar_altura1_vars[idx].set(self._prod_forcar_altura1_vars[idx].get())
                def sync_prod(*args):
                    if self._prod_forcar_altura1_vars[idx].get() != self.forcar_altura1_vars[idx].get():
                        self._prod_forcar_altura1_vars[idx].set(self.forcar_altura1_vars[idx].get())
                self._prod_forcar_altura1_vars[idx].trace_add('write', sync_main)
                self.forcar_altura1_vars[idx].trace_add('write', sync_prod)
            make_sync(i)
        detalhe_pilar_frame = ttk.Frame(grupo1_frame)
        detalhe_pilar_frame.pack(anchor="w", fill="x", pady=(6,0))
        ttk.Label(detalhe_pilar_frame, text="Detalhe Pilar:").grid(row=0, column=0, padx=(0,4))
        ttk.Checkbutton(detalhe_pilar_frame, text="Esquerda", variable=self._prod_pilar_esq).grid(row=0, column=1, padx=2)
        ttk.Checkbutton(detalhe_pilar_frame, text="Direita", variable=self._prod_pilar_dir).grid(row=0, column=2, padx=2)
        viga_nivel_frame = ttk.Frame(grupo1_frame)
        viga_nivel_frame.pack(anchor="w", fill="x", pady=(6,0))
        ttk.Checkbutton(viga_nivel_frame, text="Viga com Nível", variable=self._prod_viga_com_nivel).pack(side=tk.LEFT, padx=2)
        grupo2_frame = ttk.LabelFrame(opcoes_frame, text="2 - Detalhes de Salvamento desta Seleção", padding=4)
        grupo2_frame.pack(fill=tk.X, padx=2, pady=(0, 2))
        cont_frame = ttk.Frame(grupo2_frame)
        cont_frame.pack(anchor="w", pady=(0,0))
        ttk.Label(cont_frame, text="Continuação:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(cont_frame, text="Obstáculo", variable=self._prod_continuacao, value="Obstaculo").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(cont_frame, text="Próxima Parte", variable=self._prod_continuacao, value="Proxima Parte").pack(side=tk.LEFT, padx=2)
        ttk.Radiobutton(cont_frame, text="Viga Continuação", variable=self._prod_continuacao, value="Viga Continuacao").pack(side=tk.LEFT, padx=2)
        frame_laje_abaixo = ttk.Frame(grupo2_frame)
        frame_laje_abaixo.pack(anchor="w", pady=(6,0), fill="x")
        ttk.Label(frame_laje_abaixo, text="Laje Abaixo:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Checkbutton(frame_laje_abaixo, variable=self._prod_laje_abaixo).pack(side=tk.LEFT)
        frame_tipo1_universal = ttk.Frame(grupo2_frame)
        frame_tipo1_universal.pack(anchor="w", pady=(6,0))
        tipo1_universal_var = tk.StringVar(value=self.tipo1_universal_var.get())
        def set_tipo1_universal():
            self.tipo1_universal_var.set(tipo1_universal_var.get())
            for v in self.paineis_tipo1_vars:
                v.set(tipo1_universal_var.get())
            atualizar_preview_e_canvas()
        ttk.Label(frame_tipo1_universal, text="Paineis Altura 1:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_tipo1_universal, text="Sarrafeado", variable=tipo1_universal_var, value="Sarrafeado", command=set_tipo1_universal).pack(side=tk.LEFT)
        ttk.Radiobutton(frame_tipo1_universal, text="Grade", variable=tipo1_universal_var, value="Grade", command=set_tipo1_universal).pack(side=tk.LEFT)
        entry_grade1 = ttk.Entry(frame_tipo1_universal, textvariable=self.grade_altura1_universal, width=4)
        entry_grade1.pack(side=tk.LEFT, padx=2)
        self.extensao_painel1_var = tk.DoubleVar(value=0)
        ttk.Label(frame_tipo1_universal, text="Extensão de Painel altura 1:").pack(side=tk.LEFT, padx=(8,2))
        entry_ext1 = ttk.Entry(frame_tipo1_universal, textvariable=self.extensao_painel1_var, width=6)
        entry_ext1.pack(side=tk.LEFT, padx=2)
        def atualizar_altura1_e_grade(*args):
            try:
                valor_base = float(self.altura_geral_var.get().split('+')[0].replace(',', '.')) if '+' in self.altura_geral_var.get() else float(self.altura_geral_var.get().replace(',', '.'))
            except Exception:
                valor_base = 0
            try:
                ext = float(self.extensao_painel1_var.get())
            except Exception:
                ext = 0
            if ext > 0:
                self.altura_geral_var.set(f"{valor_base:.2f}+{ext:.2f}")
            else:
                self.altura_geral_var.set(f"{valor_base:.2f}")
            try:
                grade_base = float(self.grade_altura1_universal.get())
            except Exception:
                grade_base = 0
            self.grade_altura1_universal.set(grade_base + ext)
        self.extensao_painel1_var.trace_add('write', atualizar_altura1_e_grade)
        frame_tipo2_universal = ttk.Frame(grupo2_frame)
        frame_tipo2_universal.pack(anchor="w", pady=(2,0))
        tipo2_universal_var = tk.StringVar(value=self.tipo2_universal_var.get())
        def set_tipo2_universal():
            self.tipo2_universal_var.set(tipo2_universal_var.get())
            for v in self.paineis_tipo2_vars:
                v.set(tipo2_universal_var.get())
            atualizar_preview_e_canvas()
        ttk.Label(frame_tipo2_universal, text="Paineis Altura 2:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_tipo2_universal, text="Sarrafeado", variable=tipo2_universal_var, value="Sarrafeado", command=set_tipo2_universal).pack(side=tk.LEFT)
        ttk.Radiobutton(frame_tipo2_universal, text="Grade", variable=tipo2_universal_var, value="Grade", command=set_tipo2_universal).pack(side=tk.LEFT)
        entry_grade2 = ttk.Entry(frame_tipo2_universal, textvariable=self.grade_altura2_universal, width=4)
        entry_grade2.pack(side=tk.LEFT, padx=2)
        self.extensao_painel2_var = tk.DoubleVar(value=0)
        ttk.Label(frame_tipo2_universal, text="Extensão de Painel altura 2:").pack(side=tk.LEFT, padx=(8,2))
        entry_ext2 = ttk.Entry(frame_tipo2_universal, textvariable=self.extensao_painel2_var, width=6)
        entry_ext2.pack(side=tk.LEFT, padx=2)
        def atualizar_altura2_e_grade(*args):
            try:
                valor_base = float(self.altura_2_geral_var.get().split('+')[0].replace(',', '.')) if '+' in self.altura_2_geral_var.get() else float(self.altura_2_geral_var.get().replace(',', '.'))
            except Exception:
                valor_base = 0
            try:
                ext = float(self.extensao_painel2_var.get())
            except Exception:
                ext = 0
            if ext > 0:
                self.altura_2_geral_var.set(f"{valor_base:.2f}+{ext:.2f}")
            else:
                self.altura_2_geral_var.set(f"{valor_base:.2f}")
            try:
                grade_base = float(self.grade_altura2_universal.get())
            except Exception:
                grade_base = 0
            self.grade_altura2_universal.set(grade_base + ext)
        self.extensao_painel2_var.trace_add('write', atualizar_altura2_e_grade)
        frame_lado = ttk.Frame(grupo2_frame)
        frame_lado.pack(anchor="w", pady=(6,0))
        self._prod_lado = tk.StringVar(value="A")
        ttk.Label(frame_lado, text="Lado:").pack(side=tk.LEFT, padx=(0,4))
        ttk.Radiobutton(frame_lado, text="A", variable=self._prod_lado, value="A").pack(side=tk.LEFT)
        ttk.Radiobutton(frame_lado, text="B", variable=self._prod_lado, value="B").pack(side=tk.LEFT)
        frame_pav_obs = ttk.Frame(grupo2_frame)
        frame_pav_obs.pack(anchor="w", pady=(6,0))
        self._prod_pavimento = tk.StringVar(value=self.pavimento_var.get())
        self._prod_obs = tk.StringVar(value=self.obs_var.get())
        self._prod_pe_direito = tk.StringVar(value=self.nivel_pe_direito_var.get())
        ttk.Label(frame_pav_obs, text="Pavimento:").pack(side=tk.LEFT, padx=(0,4))
        entry_pav = ttk.Entry(frame_pav_obs, textvariable=self._prod_pavimento, width=10)
        entry_pav.pack(side=tk.LEFT, padx=(0,8))
        ttk.Label(frame_pav_obs, text="Observações:").pack(side=tk.LEFT, padx=(0,4))
        entry_obs = ttk.Entry(frame_pav_obs, textvariable=self._prod_obs, width=18)
        entry_obs.pack(side=tk.LEFT, padx=(0,8))
        ttk.Label(frame_pav_obs, text="Pé Direito:").pack(side=tk.LEFT, padx=(0,4))
        entry_pe_direito = ttk.Entry(frame_pav_obs, textvariable=self._prod_pe_direito, width=10)
        entry_pe_direito.pack(side=tk.LEFT)
        def sync_pavimento_var(*args):
            if hasattr(self, 'pavimento_var'):
                self.pavimento_var.set(self._prod_pavimento.get())
        def sync_obs_var(*args):
            if hasattr(self, 'obs_var'):
                self.obs_var.set(self._prod_obs.get())
        def sync_pe_direito_var(*args):
            if hasattr(self, 'nivel_pe_direito_var'):
                self.nivel_pe_direito_var.set(self._prod_pe_direito.get())
        self._prod_pavimento.trace_add('write', sync_pavimento_var)
        self._prod_obs.trace_add('write', sync_obs_var)
        self._prod_pe_direito.trace_add('write', sync_pe_direito_var)
        def sync_lado_var(*args):
            if hasattr(self, 'lado_var'):
                self.lado_var.set(self._prod_lado.get())
        self._prod_lado.trace_add('write', sync_lado_var)
        def sync_continuacao_var(*args):
            if hasattr(self, 'continuacao_var'):
                self.continuacao_var.set(self._prod_continuacao.get())
        self._prod_continuacao.trace_add('write', sync_continuacao_var)
        def sync_fundo_var(*args):
            if hasattr(self, 'sarrafo_esq_var'):
                self.sarrafo_esq_var.set(self._prod_laje_abaixo.get())
        self._prod_laje_abaixo.trace_add('write', sync_fundo_var)
        def atualizar_preview_e_canvas():
            self._atualizar_preview()
            atualizar_preview_producao()
        def finalizar_e_voltar():
            nome_atual = self.nome_var.get()
            lado = self._prod_lado.get()
            if not nome_atual.endswith(f'.{lado}'):
                if nome_atual.endswith('.A') or nome_atual.endswith('.B'):
                    nome_atual = nome_atual[:-2]
                self.nome_var.set(f"{nome_atual}.{lado}")
            if self._prod_laje_abaixo.get():
                valor_laje_sup = self.laje_sup_universal.get()
                self.laje_inf_universal.set(valor_laje_sup)
                self.laje_sup_universal.set(0)
            self._salvar_fundo()
            self._atualizar_preview()
            win.destroy()
            self.deiconify()
            try:
                self.bring_to_front()
            except:
                pass
        btn_finalizar.config(command=finalizar_e_voltar)
        def proximo_numero_inteiro():
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numero_base = 1
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if num.is_integer():
                        numeros_existentes.append(int(num))
                except (ValueError, TypeError):
                    continue
            proximo = numero_base + 1
            while proximo in numeros_existentes:
                proximo += 1
            return proximo
        def proxima_fracao():
            try:
                numero_atual = float(self.numero_var.get())
                numero_base = int(numero_atual)
            except (ValueError, TypeError):
                numero_base = 1
            numeros_existentes = []
            for numero in self.fundos_salvos.keys():
                try:
                    num = float(numero)
                    if int(num) == numero_base:
                        numeros_existentes.append(num)
                except (ValueError, TypeError):
                    continue
            proxima = numero_base + 0.1
            while proxima in numeros_existentes:
                proxima += 0.1
            return round(proxima, 1)
        def limpar_campos_producao():
            numero = self.numero_var.get()
            self._limpar_campos()
            self.numero_var.set(numero)
        def sequencia_producao():
            etapas = []
            etapas.append((self._select_nomes, "Seleção de Nome"))
            etapas.append((self._select_linha, "Seleção de Linha"))
            if self._prod_viga_com_nivel.get():
                etapas.append((self._select_niveis2, "Seleção de Níveis 2"))
            else:
                etapas.append((self._select_niveis, "Seleção de Níveis"))
            aberturas_checks = [
                (self._prod_abr_esq_topo, 0, "Lado Esquerdo / Topo"),
                (self._prod_abr_esq_fundo, 1, "Lado Esquerdo / Fundo"),
                (self._prod_abr_dir_topo, 2, "Lado Direita / Topo"),
                (self._prod_abr_dir_fundo, 3, "Lado Direita / Fundo")
            ]
            for var, idx, label in aberturas_checks:
                if var.get():
                    def make_select_abertura(idx, label):
                        def func():
                            return self._select_abertura(idx, label)
                        return func
                    etapas.append((make_select_abertura(idx, label), f"Seleção de Abertura: {label}"))
            if self._prod_pilar_esq.get():
                etapas.append((self._select_pilar, "Seleção Pilar Esquerda"))
            if self._prod_pilar_dir.get():
                etapas.append((self._select_pilar, "Seleção Pilar Direita"))
            def executar_etapa(idx=0):
                if self._sequencia_finalizada:
                    print("[DEBUG] Sequência interrompida por F8")
                    return
                if idx >= len(etapas):
                    print(f"[DEBUG] Sequência de produção finalizada.")
                    self.continuacao_var.set(self._prod_continuacao.get())
                    self._atualizar_preview_producao()
                    try:
                        win.lift()
                        win.focus_force()
                        win.attributes('-topmost', True)
                        win.after_idle(win.attributes, '-topmost', False)
                    except Exception as e:
                        print(f"[DEBUG] Falha ao trazer janela de produção à frente: {e}")
                    return
                func, nome = etapas[idx]
                print(f"[DEBUG] Iniciando etapa {idx+1}/{len(etapas)}: {nome}")
                try:
                    func()
                    if self._sequencia_finalizada:
                        print("[DEBUG] Sequência interrompida por F8 durante execução")
                        return
                    print(f"[DEBUG] Etapa '{nome}' concluída com sucesso.")
                    self._atualizar_preview_producao()
                    executar_etapa(idx+1)
                except Exception as e:
                    print(f"[DEBUG] Erro na etapa '{nome}': {e}. Repetindo etapa.")
                    if not self._sequencia_finalizada:  
                        executar_etapa(idx)
            executar_etapa(0)
        def proxima_viga():
            self._sequencia_finalizada = False  
            self._salvar_fundo()
            self._segmentos_temp_corte = []
            proximo = proximo_numero_inteiro()
            self.numero_var.set(str(proximo))
            limpar_campos_producao()
            self._atualizar_preview_producao()
            win.lift()
            win.focus_force()
            sequencia_producao()
        def proximo_segmento():
            self._sequencia_finalizada = False  
            self._salvar_fundo()
            dados_segmento = self._coletar_dados_atual()
            self._segmentos_temp_corte.append(dados_segmento)
            prox_frac = proxima_fracao()
            self.numero_var.set(f"{prox_frac:.1f}")
            limpar_campos_producao()
            self._atualizar_preview_producao()
            win.lift()
            win.focus_force()
            sequencia_producao()
        def reiniciar_selecao():
            self._sequencia_finalizada = False  
            sequencia_producao()
        btn_proxima_viga.config(command=proxima_viga)
        btn_prox_seg.config(command=proximo_segmento)
        btn_reiniciar.config(command=reiniciar_selecao)
        btn_finalizar.config(command=finalizar_e_voltar)
        def _salvar_e_abrir_combinacoes_corte():
            self._salvar_fundo()
            dados_atual = self._coletar_dados_atual()
        win.reiniciar_selecao = reiniciar_selecao
        win.sequencia_producao = sequencia_producao
    def _atualizar_preview_producao(self):
        if not hasattr(self, '_producao_canvas'):
            return
        canvas = self._producao_canvas
        canvas.delete("all")
        try:
            largura = self._float_safe(self.largura_var)
            altura_geral = self._float_safe(self.altura_geral_var)
        except Exception:
            return
        if largura <= 0 or altura_geral <= 0:
            return
        escala_x = 300 / (largura * 1.2)
        escala_y = 100 / (altura_geral * 1.2)
        escala = min(escala_x, escala_y)
        x_inicial = 10
        y_inicial = 110
        self._desenhar_elementos(canvas, x_inicial, y_inicial, largura, altura_geral, escala)
    def _mostrar_comentario_flutuante(self, mensagem, mensagem_erro=None):
        import tkinter as tk
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
        balcao = tk.Toplevel()
        balcao.overrideredirect(True)
        balcao.attributes('-topmost', True)
        balcao.configure(bg='#fff8b0')
        # Centralizar na tela
        largura = 480
        altura = 120 if not mensagem_erro else 160
        x = (balcao.winfo_screenwidth() - largura) // 2
        y = 40
        balcao.geometry(f"{largura}x{altura}+{x}+{y}")
        label = tk.Label(balcao, text=mensagem, font=("Arial", 16, "bold"), bg='#fff8b0', fg='#333', wraplength=460)
        label.pack(expand=True, fill=tk.BOTH, padx=10, pady=(10,0))
        if mensagem_erro:
            erro_label = tk.Label(balcao, text=mensagem_erro, font=("Arial", 12, "bold"), bg='#fff8b0', fg='red', wraplength=460)
            erro_label.pack(expand=True, fill=tk.BOTH, padx=10, pady=(0,10))
        def fechar_balao_e_trazer_producao():
            try:
                balcao.destroy()
            except:
                pass
            self._balcao_flutuante = None
            self.bring_to_front()
        btn_cancelar = tk.Button(balcao, text="Cancelar Seleção F8 e Esc", font=("Arial", 12, "bold"), bg="#ffb0b0", command=fechar_balao_e_trazer_producao)
        btn_cancelar.pack(pady=(8, 10))
        self._balcao_flutuante = balcao
        balcao.update()
        return balcao
    def _fechar_comentario_flutuante(self):
        if hasattr(self, '_balcao_flutuante') and self._balcao_flutuante is not None:
            try:
                self._balcao_flutuante.destroy()
            except:
                pass
            self._balcao_flutuante = None
    def _desenhar_hachura_abertura(self, x1, y1, x2, y2, espacamento=6):
        """
        Desenha linhas diagonais (hachura) dentro do retângulo da abertura.
        x1, y1: canto superior esquerdo
        x2, y2: canto inferior direito
        espacamento: distância entre as linhas da hachura
        """
        x1, x2 = min(x1, x2), max(x1, x2)
        y1, y2 = min(y1, y2), max(y1, y2)
        largura = x2 - x1
        altura = y2 - y1
        for i in range(-int(altura), int(largura), espacamento):
            xi = x1 + max(i, 0)
            yi = y2 - max(-i, 0)
            xf = x1 + min(i + altura, largura)
            yf = y2 - min(i + altura, altura)
            self.canvas.create_line(xi, yi, xf, yf, fill="#BDB76B", width=1)

    def bring_to_front(self):
        self._fechar_comentario_flutuante()
        try:
            self.deiconify()
            self.lift()
            self.attributes('-topmost', True)
            self.focus_force()
            self.after_idle(self.attributes, '-topmost', False)
            try:
                import win32gui
                import win32con
                hwnd = self.winfo_id()
                win32gui.SetForegroundWindow(hwnd)
                win32gui.ShowWindow(hwnd, win32con.SW_SHOWNORMAL)
            except Exception as e:
                print(f'[DEBUG] win32gui bring_to_front falhou: {e}')
            self.after(100, lambda: self.focus_force())
        except Exception as e:
            print(f'[DEBUG] Falha ao trazer janela para frente: {e}')
    def _coletar_dados_atual(self):
        def safe_get(var):
            try:
                return float(var.get())
            except Exception:
                return 0.0
        def safe_get_str(var):
            try:
                return str(var.get())
            except Exception:
                return ''
        return {
            'numero': safe_get_str(self.numero_var),
            'pavimento': safe_get_str(self.pavimento_var),
            'nome': safe_get_str(self.nome_var),
            'texto_esq': safe_get_str(self.texto_esq_var),
            'fundo': safe_get_str(self.fundo_var),
            'texto_dir': safe_get_str(self.texto_dir_var),
            'obs': safe_get_str(self.obs_var),
            'largura': safe_get(self.largura_var),
            'altura_geral': safe_get(self.altura_geral_var),
            'paineis_larguras': [safe_get(v) for v in self.paineis_larguras_vars],
            'paineis_alturas': [safe_get(v) for v in self.paineis_alturas_vars],
            'paineis_alturas2': [safe_get(v) for v in self.paineis_alturas2_vars],
            'aberturas': [[safe_get(v) for v in linha] for linha in self.aberturas_vars],
            'sarrafo_esq': bool(self.sarrafo_esq_var.get()),
            'sarrafo_dir': bool(self.sarrafo_dir_var.get()),
            'lajes_sup': [safe_get(v) for v in self.lajes_sup_vars],
            'lajes_inf': [safe_get(v) for v in self.lajes_inf_vars],
            'lajes_central_alt': [safe_get(v) for v in self.lajes_central_alt_vars],
            'detalhe_pilar_esq': [safe_get(v) for v in self.detalhe_pilar_esq_vars],
            'detalhe_pilar_dir': [safe_get(v) for v in self.detalhe_pilar_dir_vars],
            'nivel_oposto': safe_get_str(self.nivel_oposto_var),
            'nivel_viga': safe_get_str(self.nivel_viga_var),
            'nivel_pe_direito': safe_get_str(self.nivel_pe_direito_var),
            'ajuste': safe_get_str(self.ajuste_var),
            'paineis_tipo1': [safe_get_str(v) for v in self.paineis_tipo1_vars],
            'paineis_tipo2': [safe_get_str(v) for v in self.paineis_tipo2_vars],
            'paineis_grade_altura1': [safe_get(v) for v in self.paineis_grade_altura1_vars],
            'paineis_grade_altura2': [safe_get(v) for v in self.paineis_grade_altura2_vars],
            'continuacao': safe_get_str(self.continuacao_var),
            'lado': safe_get_str(self.lado_var),
        }
    def _salvar_combinacoes(self):
        import json
        try:
            with open('combinacoes_salvas.json', 'w', encoding='utf-8') as f:

                json.dump(self.combinacoes, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Erro ao salvar combinações: {e}")
    def _carregar_combinacoes(self):
        import os
        import json
        if os.path.exists('combinacoes_salvas.json'):
            try:
                with open('combinacoes_salvas.json', 'r', encoding='utf-8') as f:

                    self.combinacoes  = json.load(f)
                self._atualizar_lista_combinacoes()
            except Exception as e:
                print(f"Erro ao carregar combinações: {e}")
    def _toggle_modo_combinacao(self):
        self.modo_combinacao = not self.modo_combinacao
        self.checkbox_selecionados = set()
        self._atualizar_lista()
        self._atualizar_layout_listas()
    def _abrir_combinacoes_corte(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        win = tk.Toplevel(self)
        win.title("Combinações para Corte")
        win.geometry("600x400")
        win.grab_set()
        win.lift()
        win.attributes('-topmost', True)
        win.after_idle(win.attributes, '-topmost', False)
        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        frame_listas = ttk.Frame(frame)
        frame_listas.pack(fill=tk.BOTH, expand=True)
        lbl_segmentos = ttk.Label(frame_listas, text="Segmentos")
        lbl_segmentos.grid(row=0, column=0, padx=5, pady=2)
        self.segmentos_corte_vars = []
        self.tree_segmentos_corte = ttk.Treeview(frame_listas, columns=("Sel", "Desc"), show="headings", height=12)
        self.tree_segmentos_corte.heading("Sel", text="Sel")
        self.tree_segmentos_corte.heading("Desc", text="Segmento")
        self.tree_segmentos_corte.column("Sel", width=38, anchor="center")
        self.tree_segmentos_corte.column("Desc", width=180, anchor="w")
        self.tree_segmentos_corte.grid(row=1, column=0, padx=5, pady=2, sticky="nsew")
        self.tree_segmentos_corte.delete(*self.tree_segmentos_corte.get_children())
        self.segmentos_corte_vars.clear()
        for idx, seg in enumerate(self._segmentos_temp_corte):
            var = tk.BooleanVar(value=False)
            self.segmentos_corte_vars.append(var)
            desc = f"{seg.get('numero', '')} - {seg.get('nome', '')}"
            self.tree_segmentos_corte.insert('', 'end', iid=str(idx), values=('☐', desc))
        def on_click_checkbox(event):
            region = self.tree_segmentos_corte.identify("region", event.x, event.y)
            col = self.tree_segmentos_corte.identify_column(event.x)
            row = self.tree_segmentos_corte.identify_row(event.y)
            if region == "cell" and col == '#1' and row:
                idx = int(row)
                atual = self.segmentos_corte_vars[idx].get()
                if not atual and sum(v.get() for v in self.segmentos_corte_vars) >= 2:
                    return
                self.segmentos_corte_vars[idx].set(not atual)
                self.tree_segmentos_corte.set(row, "Sel", '☑' if self.segmentos_corte_vars[idx].get() else '☐')
        self.tree_segmentos_corte.bind("<Button-1>", on_click_checkbox)
        lbl_combinacoes = ttk.Label(frame_listas, text="Combinações")
        lbl_combinacoes.grid(row=0, column=1, padx=5, pady=2)
        self.lista_combinacoes_corte = tk.Listbox(frame_listas, height=12, width=30)
        self.lista_combinacoes_corte.grid(row=1, column=1, padx=5, pady=2, sticky="nsew")
        frame_listas.columnconfigure(0, weight=1)
        frame_listas.columnconfigure(1, weight=1)
        self.lista_combinacoes_corte.delete(0, tk.END)
        for comb in getattr(self, '_combinacoes_temp_corte', []):
            self.lista_combinacoes_corte.insert(tk.END, comb['nome'])
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill=tk.X, pady=8)
        def criar_combinacao():
            selecionados = [i for i, v in enumerate(self.segmentos_corte_vars) if v.get()]
            if len(selecionados) != 2:
                messagebox.showerror("Erro", "Selecione exatamente 2 segmentos para combinar.")
                return
            idx1, idx2 = selecionados
            seg1 = self._segmentos_temp_corte[idx1]
            seg2 = self._segmentos_temp_corte[idx2]
            nome_comb = f"{seg1.get('nome','')} + {seg2.get('nome','')}"
            ids = [seg1.get('numero',''), seg2.get('numero','')]
            pavimento = seg1.get('pavimento','')
            if seg1.get('pavimento','') != seg2.get('pavimento',''):
                messagebox.showerror("Erro", "Só é possível combinar segmentos do MESMO pavimento!")
                return
            comb = {'nome': nome_comb, 'ids': ids, 'pavimento': pavimento}
            if not hasattr(self, '_combinacoes_temp_corte'):
                self._combinacoes_temp_corte = []
            self._combinacoes_temp_corte.append(comb)
            self.lista_combinacoes_corte.insert(tk.END, nome_comb)
            for v in self.segmentos_corte_vars:
                v.set(False)
            for iid in self.tree_segmentos_corte.get_children():
                self.tree_segmentos_corte.set(iid, "Sel", '☐')
        def salvar_combinacoes():
            messagebox.showinfo("Salvo", "Combinações temporárias salvas. Pressione Finalizar para enviar para a lista principal.")
        def finalizar_combinacoes():
            if hasattr(self, '_combinacoes_temp_corte') and self._combinacoes_temp_corte:
                self.combinacoes.extend(self._combinacoes_temp_corte)
                self._salvar_combinacoes()
                self._atualizar_lista_combinacoes()
                self._combinacoes_temp_corte = []
                messagebox.showinfo("Finalizado", "Combinações enviadas para a lista principal.")
            win.destroy()
        btn_criar = ttk.Button(frame_botoes, text="Criar", command=criar_combinacao)
        btn_criar.pack(side=tk.LEFT, padx=5)
        btn_salvar = ttk.Button(frame_botoes, text="Salvar", command=salvar_combinacoes)
        btn_salvar.pack(side=tk.LEFT, padx=5)
        btn_finalizar = ttk.Button(frame_botoes, text="Finalizar", command=finalizar_combinacoes)
        btn_finalizar.pack(side=tk.LEFT, padx=5)
    @_cad_safe
    def _select_niveis2(self):
        import pythoncom
        import win32com.client
        import win32gui
        from tkinter import messagebox
        import re
        try:
            pythoncom.CoInitialize()
            if self.ac is None:
                self.ac = win32com.client.Dispatch("AutoCAD.Application")
            doc = self.ac.ActiveDocument
            if doc is None:
                print(f"[DEBUG] messagebox.showerror: Nenhum documento AutoCAD ativo encontrado!")
                messagebox.showerror("Erro", "Nenhum documento AutoCAD ativo encontrado!")
                raise Exception("Nenhum documento AutoCAD ativo encontrado!")
            try:
                acad_window = win32gui.FindWindow(None, self.ac.Caption)
                if acad_window:
                    win32gui.SetForegroundWindow(acad_window)
                    win32gui.ShowWindow(acad_window, 5)
                    try:
                        import pyautogui
                        pyautogui.press('f8')
                    except Exception as e:
                        print(f"[DEBUG] Falha ao pressionar F8 com pyautogui: {e}")
            except Exception:
                pass
            self.iconify()
            try:
                doc.SendCommand("F8 ")
            except Exception as e:
                print(f"[DEBUG] Falha ao enviar F8 para o AutoCAD: {e}")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto do Nível Viga e pressione Enter")
            self.update()
            try:
                temp_sel = doc.SelectionSets.Item("TempSelection")
                temp_sel.Delete()
            except Exception:
                pass
            selection = doc.SelectionSets.Add("TempSelection")
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            nivel_viga = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                    print(f"[DEBUG] Tipo do objeto selecionado: {objname}")
                except Exception as e:
                    print(f"[DEBUG] Erro ao obter ObjectName: {e}")
                    print(f"[DEBUG] Tentando obter mais informações do objeto...")
                    try:
                        print(f"[DEBUG] Handle do objeto: {entity.Handle}")
                    except:
                        pass
                    try:
                        print(f"[DEBUG] Layer do objeto: {entity.Layer}")
                    except:
                        pass
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "Erro ao identificar tipo do objeto. Tente selecionar novamente.")
                    self.deiconify()
                    raise Exception(f"Erro ao identificar tipo do objeto: {e}")
                if objname in ["AcDbText", "AcDbMText"]:
                    try:
                        texto = entity.TextString
                        print(f"[DEBUG] Texto encontrado: {texto}")
                        match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                        if match:
                            nivel_viga = float(match.group(1).replace(",", "."))
                            print(f"[DEBUG] Valor extraído: {nivel_viga}")
                            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
                        else:
                            print(f"[DEBUG] Nenhum número encontrado no texto: {texto}")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do texto!")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao processar texto: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao processar texto: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao processar texto: {e}")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        print(f"[DEBUG] Processando bloco...")
                        try:
                            print(f"[DEBUG] Nome do bloco: {entity.Name}")
                        except Exception as e:
                            print(f"[DEBUG] Não foi possível obter nome do bloco: {e}")
                        try:
                            print(f"[DEBUG] Layer do bloco: {entity.Layer}")
                        except Exception as e:
                            print(f"[DEBUG] Não foi possível obter layer do bloco: {e}")
                        atributos = entity.GetAttributes()
                        print(f"[DEBUG] Número de atributos encontrados: {len(atributos)}")
                        for attrib in atributos:
                            try:
                                tag = attrib.TagString
                                texto = attrib.TextString
                                print(f"[DEBUG] Atributo encontrado - Tag: {tag}, Valor: {texto}")
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    print(f"[DEBUG] Valor numérico extraído: {valor}")
                                    valor_mais_alto = max(valor_mais_alto, valor)
                                else:
                                    print(f"[DEBUG] Nenhum número encontrado no atributo")
                            except Exception as e:
                                print(f"[DEBUG] Erro ao processar atributo: {e}")
                                continue
                        if valor_mais_alto > 0:
                            nivel_viga = valor_mais_alto
                            print(f"[DEBUG] Usando valor mais alto encontrado: {nivel_viga}")
                            self.nivel_viga_var.set(f"{nivel_viga:.2f}")
                        else:
                            print(f"[DEBUG] Nenhum valor numérico encontrado em nenhum atributo")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] Erro ao processar bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao processar bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao processar bloco: {e}")
                else:
                    print(f"[DEBUG] Objeto não suportado: {objname}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto ou bloco!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto ou bloco!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto do Nível Oposto e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            nivel_oposto = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                objname = entity.ObjectName
                if objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        nivel_oposto = float(match.group(1).replace(",", "."))
                        self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
                        if nivel_oposto >= nivel_viga:
                            altura2 = (nivel_oposto - nivel_viga) * 100
                        else:
                            altura2 = 0
                        self.altura_2_geral_var.set(round(altura2, 2))
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                elif objname == "AcDbBlockReference":
                    valor_mais_alto = 0.0
                    try:
                        for attrib in entity.GetAttributes():
                            try:
                                texto = attrib.TextString
                                match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                                if match:
                                    valor = float(match.group(1).replace(",", "."))
                                    valor_mais_alto = max(valor_mais_alto, valor)
                            except:
                                continue
                        if valor_mais_alto > 0:
                            nivel_oposto = valor_mais_alto
                            self.nivel_oposto_var.set(f"{nivel_oposto:.2f}")
                            if nivel_oposto >= nivel_viga:
                                altura2 = (nivel_oposto - nivel_viga) * 100
                            else:
                                altura2 = 0
                            self.altura_2_geral_var.set(round(altura2, 2))
                        else:
                            print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do bloco!")
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o número do bloco!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o número do bloco!")
                    except Exception as e:
                        print(f"[DEBUG] messagebox.showwarning: Erro ao acessar atributos do bloco: {e}")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", f"Erro ao acessar atributos do bloco: {e}")
                        self.deiconify()
                        raise Exception(f"Erro ao acessar atributos do bloco: {e}")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é um texto ou bloco!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto ou bloco!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto ou bloco!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto ou linha da Altura da Laje e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            altura_laje = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                try:
                    objname = entity.ObjectName
                except Exception as e:
                    print(f"[DEBUG] messagebox.showwarning: Erro ao acessar ObjectName: {e}")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", f"Erro ao acessar o tipo do objeto selecionado: {e}\nSelecione novamente!")
                    self.deiconify()
                    raise Exception(f"Erro ao acessar o tipo do objeto selecionado: {e}")
                if objname == "AcDbLine":
                    altura_laje = 0.0
                elif objname in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"([0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        altura_laje = float(match.group(1).replace(",", ".")) + 2
                    else:
                        print(f"[DEBUG] messagebox.showwarning: Não foi possível extrair o número do texto!")
                        self.bring_to_front()
                        messagebox.showwarning("Aviso", "Não foi possível extrair o número do texto!")
                        self.deiconify()
                        raise Exception("Não foi possível extrair o número do texto!")
                else:
                    print(f"[DEBUG] messagebox.showwarning: O objeto selecionado não é texto nem linha!")
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é texto nem linha!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é texto nem linha!")
            else:
                print(f"[DEBUG] messagebox.showwarning: Nenhum objeto selecionado!")
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            balcao = self._mostrar_comentario_flutuante("Selecione o texto da Altura 1 (ex: 20/60 ou 20x60) e pressione Enter")
            self.update()
            try:
                selection.Clear()
            except:
                pass
            try:
                selection.SelectOnScreen()
            except:
                try:
                    selection.Select(5)
                except:
                    pass
            self._fechar_comentario_flutuante()
            valor2 = 0.0
            if selection.Count > 0:
                entity = selection.Item(0)
                if entity.ObjectName in ["AcDbText", "AcDbMText"]:
                    texto = entity.TextString
                    match = re.search(r"[x/](\s*[0-9]+[\.,]?[0-9]*)", texto)
                    if match:
                        valor2 = float(match.group(1).replace(",", ".")) + 4
                    else:
                        nums = re.findall(r"[0-9]+[\.,]?[0-9]*", texto)
                        if len(nums) >= 2:
                            valor2 = float(nums[1].replace(",", ".")) + 4
                        else:
                            self.bring_to_front()
                            messagebox.showwarning("Aviso", "Não foi possível extrair o segundo valor do texto!")
                            self.deiconify()
                            raise Exception("Não foi possível extrair o segundo valor do texto!")
                else:
                    self.bring_to_front()
                    messagebox.showwarning("Aviso", "O objeto selecionado não é um texto!")
                    self.deiconify()
                    raise Exception("O objeto selecionado não é um texto!")
            else:
                self.bring_to_front()
                messagebox.showwarning("Aviso", "Nenhum objeto selecionado!")
                self.deiconify()
                raise Exception("Nenhum objeto selecionado!")
            diferenca_niveis = nivel_oposto - nivel_viga
            if diferenca_niveis < 0:
                diferenca_niveis = 0
            altura_1 = valor2 - altura_laje - (diferenca_niveis * 100)
            self.altura_geral_var.set(round(altura_1, 2))
            self.grade_altura1_universal.set(round(altura_1, 2))
            self.deiconify()
            if nivel_oposto <= nivel_viga:
                self.laje_sup_universal.set(altura_laje)
                self.laje_central_alt_universal.set(0)
            else:
                self.laje_central_alt_universal.set(altura_laje)
                self.laje_sup_universal.set(0)
        except Exception as e:
            print(f"[DEBUG] messagebox.showerror: Erro ao selecionar níveis: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao selecionar níveis: {str(e)}")
            self.deiconify()
            raise
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    def _gerar_script(self):
        """Gera o script SCR para o fundo selecionado na lista."""
        tree = None
        if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
            tree = self.lista_fundos1
        elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
            tree = self.lista_fundos2
        if not tree:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        selection = tree.selection()
        if not selection:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        if not dados:
            messagebox.showerror("Erro", f"Dados do item {numero} não encontrados.")
            return
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")
        pavimento = str(dados.get('pavimento', 'SEM_PAVIMENTO')).strip() or 'SEM_PAVIMENTO'
        diretorio_saida = os.path.join(diretorio_base, pavimento)
        if not os.path.exists(diretorio_saida):
            os.makedirs(diretorio_saida)
        try:
            script, caminho_arquivo = gerar_script_viga(dados, diretorio_saida)
            if script and caminho_arquivo:
                messagebox.showinfo("Sucesso", f"Script gerado com sucesso em: {caminho_arquivo}")
            elif script and not caminho_arquivo:
                nome = dados.get('nome', 'SemNome')
                observacoes = dados.get('obs', '')
                arquivo_nome = f"{nome}{'_' + observacoes if observacoes else ''}.scr"
                caminho_arquivo = os.path.join(diretorio_saida, arquivo_nome)
                resposta = messagebox.askyesno("Arquivo Existente", 
                                            f"O arquivo {arquivo_nome} já existe. Deseja sobrescrever?")
                if resposta:
                    with open(caminho_arquivo, "w", encoding="utf-16") as f:
                        f.write(script)
                    messagebox.showinfo("Sucesso", f"Script gerado com sucesso em: {caminho_arquivo}")
            else:
                messagebox.showerror("Erro", "Não foi possível gerar o script.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar o script: {str(e)}")
    def _ao_fechar(self):
        """Método chamado quando a aplicação é fechada para salvar os dados automaticamente."""
        try:
            self._salvar_fundos_salvos()
            self._salvar_combinacoes()
            print("Dados salvos com sucesso ao fechar a aplicação.")
        except Exception as e:
            print(f"Erro ao salvar dados durante o fechamento: {str(e)}")
        self.destroy()
    def _salvar_teste(self):
        """Gera o script SCR para o item selecionado e salva em um arquivo de teste fixo."""
        tree = None
        if hasattr(self, 'lista_fundos1') and self.lista_fundos1.focus():
            tree = self.lista_fundos1
        elif hasattr(self, 'lista_fundos2') and self.lista_fundos2.focus():
            tree = self.lista_fundos2
        if not tree:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        selection = tree.selection()
        if not selection:
            messagebox.showerror("Erro", "Nenhum item selecionado na lista.")
            return
        item = selection[0]
        if tree == self.lista_fundos1:
            values = tree.item(item)['values']
            numero = str(values[2])
        else:
            values = tree.item(item)['values']
            numero = str(values[1])
        dados = self.fundos_salvos.get(numero, {})
        if not dados:
            messagebox.showerror("Erro", f"Dados do item {numero} não encontrados.")
            return
        caminho_arquivo = r"C:\Users\rvene\Desktop\Automacao_cad\Vigas\A_B\Ferramentas\TESTE_VIGA_TV.scr"
        pasta_ferramentas = os.path.dirname(caminho_arquivo)
        if not os.path.exists(pasta_ferramentas):
            os.makedirs(pasta_ferramentas)
        try:
            script, _ = gerar_script_viga(dados, "temp")
            if script:
                with open(caminho_arquivo, "w", encoding="utf-16") as f:
                    f.write(script)
                messagebox.showinfo("Sucesso", f"Script de teste salvo com sucesso em:\n{caminho_arquivo}")
            else:
                messagebox.showerror("Erro", "Não foi possível gerar o script.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar o script de teste: {str(e)}")
    def _registrar_f8_global(self):
        def f8_handler(e=None):
            print('[DEBUG] F8 pressionado: iniciando cancelamento global')
            if not self._f8_cancelando:
                self._f8_cancelando = True
                self._sequencia_finalizada = True  
                self._cancelar_selecao_global()
        keyboard.on_press_key('f8', lambda e: f8_handler())
    def _cancelar_selecao_global(self):
        print('[DEBUG] _cancelar_selecao_global executado')
        self._fechar_comentario_flutuante()
        self.deiconify()
        self.bring_to_front()
        self._f8_cancelando = False
        if hasattr(self, '_janela_producao') and self._janela_producao:
            try:
                if self._sequencia_finalizada:
                    self._janela_producao.destroy()
                    self._janela_producao = None
                else:
                    self.after(100, lambda: self._janela_producao.reiniciar_selecao())
            except:
                pass
    def _gerar_todos_scripts(self):
        import os
        from tkinter import messagebox, Toplevel, StringVar
        import subprocess
        from tkinter import ttk
        pavimentos = set()
        for dados in self.fundos_salvos.values():
            pav = str(dados.get('pavimento', '')).strip()
            if pav:
                pavimentos.add(pav)
        pavimentos = sorted(pavimentos)
        if not pavimentos:
            messagebox.showerror("Erro", "Nenhum pavimento encontrado nos fundos salvos!")
            return
        selected_pavimento = StringVar()
        dialog = Toplevel(self)
        dialog.title("Selecionar Pavimento")
        dialog.geometry("300x120")
        dialog.grab_set()
        ttk.Label(dialog, text="Escolha o pavimento para gerar os scripts:").pack(pady=10)
        combo = ttk.Combobox(dialog, values=pavimentos, textvariable=selected_pavimento, state="readonly")
        combo.pack(pady=5)
        combo.current(0)
        confirmed = {'ok': False}
        def confirmar():
            confirmed['ok'] = True
            dialog.destroy()
        def cancelar():
            dialog.destroy()
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=confirmar).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=cancelar).pack(side=tk.LEFT, padx=10)
        dialog.wait_window()
        pavimento = selected_pavimento.get()
        if not confirmed['ok'] or not pavimento or pavimento not in pavimentos:
            messagebox.showwarning("Cancelado", "Operação cancelada ou pavimento inválido.")
            return
        total = 0
        erros = []
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")
        diretorio_saida = os.path.join(diretorio_base, pavimento)
        # Remove a pasta se já existir para evitar duplicidade
        if os.path.exists(diretorio_saida):
            try:
                shutil.rmtree(diretorio_saida)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao remover a pasta existente do pavimento: {e}")
                return
        os.makedirs(diretorio_saida)
        from gerador_script_viga import gerar_script_viga
        for numero, dados in self.fundos_salvos.items():
            try:
                pav = str(dados.get('pavimento', '')).strip()
                if pav != pavimento:
                    continue
                script, caminho_arquivo = gerar_script_viga(dados, diretorio_saida)
                if script and caminho_arquivo:
                    total += 1
            except Exception as e:
                erros.append(f"{numero}: {e}")
        if not erros:
            messagebox.showinfo("Sucesso", f"Todos os scripts ({total}) do pavimento '{pavimento}' foram gerados!")
        else:
            messagebox.showwarning("Concluído com Erros", f"{total} scripts gerados. Erros:\n" + '\n'.join(erros))
        try:
            combinador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Combinador_VIGA.py")
            subprocess.run(["python", combinador_path, diretorio_saida], check=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao executar o Combinador: {e}")
            return

        try:
            pasta_combinados = os.path.join(diretorio_saida, "Combinados")
            ordenador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Ordenador_VIGA.py")
            subprocess.run(["python", ordenador_path, pasta_combinados], check=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao executar o Ordenador: {e}")
            return
    def _combinar_scripts(self):
        import subprocess
        from tkinter import filedialog, messagebox
        import os
        import sys
        diretorio = filedialog.askdirectory(title="Selecione a pasta SCRIPTS para combinar")
        if not diretorio:
            return
        script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'Combinador_VIGA.py'))
        try:
            subprocess.run([sys.executable, script_path], check=True, cwd=os.path.dirname(script_path))
            messagebox.showinfo("Sucesso", "Scripts combinados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao combinar scripts: {e}")
    def _ordenar_scripts(self):
        import subprocess
        from tkinter import filedialog, messagebox
        import os
        import sys
        diretorio = filedialog.askdirectory(title="Selecione a pasta Combinados para ordenar")
        if not diretorio:
            return
        script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'Ordenador_VIGA.py'))
        try:
            subprocess.run([sys.executable, script_path], check=True, cwd=os.path.dirname(script_path))
            messagebox.showinfo("Sucesso", "Scripts ordenados com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ordenar scripts: {e}")
    def _calcular_ajuste(self, *args):
        """Calcula o ajuste baseado nos níveis e pé direito"""
        try:
            nivel_viga = float(self.nivel_viga_var.get() or 0)
            nivel_oposto = float(self.nivel_oposto_var.get() or 0)
            pe_direito = float(self.nivel_pe_direito_var.get() or 0)
            if nivel_viga > 0 and nivel_oposto > 0 and pe_direito > 0:
                nivel_maior = max(nivel_viga, nivel_oposto)
                ajuste = (pe_direito - nivel_maior) * 100
                self.ajuste_var.set(f"{ajuste:.2f}")
            else:
                self.ajuste_var.set("")
        except (ValueError, TypeError):
            self.ajuste_var.set("")
    def _area_intersecao(self, rx1, ry1, rx2, ry2, ax1, ay1, ax2, ay2):
        """Calcula a área de interseção entre dois retângulos."""
        inter_x1 = max(rx1, ax1)
        inter_y1 = max(ry1, ay1)
        inter_x2 = min(rx2, ax2)
        inter_y2 = min(ry2, ay2)
        if inter_x2 > inter_x1 and inter_y2 > inter_y1:
            return (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
        return 0
    def _ordenar_lista(self, coluna, lista2=False):
        """Ordena a lista de acordo com a coluna clicada."""
        # Inverte a ordem se clicar na mesma coluna
        reverso = False
        if self.ordem_atual['coluna'] == coluna:
            reverso = not self.ordem_atual['reverso']
        
        # Atualiza a ordem atual
        self.ordem_atual = {'coluna': coluna, 'reverso': reverso}
        
        # Seleciona a lista correta
        lista = self.lista_fundos2 if lista2 else self.lista_fundos1
        
        # Obtém todos os itens da lista
        itens = [(lista.set(item, coluna), item) for item in lista.get_children('')]
        
        # Função para converter número para ordenação numérica
        def converter_numero(valor):
            try:
                return float(valor)
            except:
                return valor
        
        # Ordena os itens
        if coluna == "Número":
            itens.sort(key=lambda x: converter_numero(x[0]), reverse=reverso)
        else:
            itens.sort(key=lambda x: x[0].lower(), reverse=reverso)
        
        # Reordena a lista
        for index, (val, item) in enumerate(itens):
            lista.move(item, '', index)
            # Atualiza as cores zebradas
            lista.tag_configure('odd', background='#e8f5e9')
            lista.tag_configure('even', background='#ffffff')
            lista.item(item, tags=('odd' if index % 2 == 0 else 'even',))

    def _gerar_teste_combinados(self):
        """Gera o script da combinação selecionada e salva no arquivo de teste."""
        # 1. Verificar se alguma combinação está selecionada
        if not self.lista_combinacoes.selection():
            messagebox.showerror("Erro", "Nenhuma combinação selecionada para gerar o script de teste!")
            return

        # 2. Obter a combinação selecionada
        item_id = self.lista_combinacoes.selection()[0]
        idx = int(item_id)
        combinacao_selecionada = self.combinacoes[idx]

        nome_combinacao = combinacao_selecionada.get('nome', 'SemNome')
        fundos_ids = combinacao_selecionada.get('ids', [])

        if not fundos_ids:
            messagebox.showerror("Erro", f"Combinação '{nome_combinacao}' não possui fundos associados para gerar o script de teste!")
            return

        # 3. Definir o caminho do arquivo de saída específico para o teste
        caminho_base = os.path.dirname(os.path.abspath(__file__))
        # O caminho para A_B/Ferramentas/TESTE_VIGA_TVC.scr
        test_output_filepath = os.path.join(caminho_base, "Ferramentas", "TESTE_VIGA_TVC.scr")

        # 4. Chamar gerar_script_combinado da instância gerador_combinados
        # Passar None para diretorio_saida, pois estamos usando output_filepath
        sucesso, caminho_arquivo_gerado = self.gerador_combinados.gerar_script_combinado(
            combinacao_selecionada,
            self.fundos_salvos,
            None, # diretorio_saida será ignorado
            output_filepath=test_output_filepath
        )

        if sucesso:
            messagebox.showinfo("Script de Teste Gerado", f"Script de teste da combinação '{nome_combinacao}' gerado com sucesso em: {caminho_arquivo_gerado}")
        else:
            messagebox.showerror("Erro ao Gerar Script de Teste", f"Erro ao gerar script de teste da combinação '{nome_combinacao}'.")
    
    def _gerar_script_um_combinado(self):
        """Gera o script para a combinação selecionada e salva na pasta SCRIPTS."""
        if not self.lista_combinacoes.selection():
            messagebox.showerror("Erro", "Nenhuma combinação selecionada!")
            return

        item_id = self.lista_combinacoes.selection()[0]
        idx = int(item_id) # O Treeview usa 0-indexed para IDs numéricos se não houver um ID customizado
        combinacao_selecionada = self.combinacoes[idx]
        
        nome_combinacao = combinacao_selecionada.get('nome', 'SemNome')
        fundos_ids = combinacao_selecionada.get('ids', [])

        if not fundos_ids:
            messagebox.showerror("Erro", f"Combinação '{nome_combinacao}' não possui fundos associados!")
            return

        # Para o "Gerar 1", usaremos a função gerar_script_combinado que gera um único script para a combinação.
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")

        # Tenta inferir o pavimento do primeiro fundo na combinação ou usa um padrão
        primeiro_fundo_id = fundos_ids[0]
        primeiro_fundo_dados = self.fundos_salvos.get(primeiro_fundo_id)
        pavimento = 'SEM_PAVIMENTO'
        if primeiro_fundo_dados and primeiro_fundo_dados.get('pavimento'):
            pavimento = str(primeiro_fundo_dados['pavimento']).strip() or 'SEM_PAVIMENTO'
        
        diretorio_saida = os.path.join(diretorio_base, pavimento, "Combinados")
        
        sucesso, caminho_arquivo = self.gerador_combinados.gerar_script_combinado(
            combinacao_selecionada, 
            self.fundos_salvos, 
            diretorio_saida
        )

        if sucesso:
            messagebox.showinfo("Script Gerado", f"Script da combinação '{nome_combinacao}' gerado com sucesso em: {caminho_arquivo}")
        else:
            messagebox.showerror("Erro ao Gerar Script", f"Erro ao gerar script da combinação '{nome_combinacao}'.")

    def _gerar_todos_scripts_combinados(self):
        """Gera todos os scripts para as combinações do pavimento selecionado."""
        import os
        from tkinter import messagebox, Toplevel, StringVar
        from tkinter import ttk
        import subprocess

        # 1. Coletar todos os pavimentos únicos das combinações
        pavimentos = set()
        for combinacao in self.combinacoes:
            # Pega o pavimento do primeiro fundo na combinação como referência para o pavimento da combinação
            if combinacao['ids']:
                primeiro_fundo_id = combinacao['ids'][0]
                primeiro_fundo_dados = self.fundos_salvos.get(primeiro_fundo_id)
                if primeiro_fundo_dados and primeiro_fundo_dados.get('pavimento'):
                    pav = str(primeiro_fundo_dados['pavimento']).strip()
                    if pav:
                        pavimentos.add(pav)
        
        pavimentos = sorted(list(pavimentos))
        if not pavimentos:
            messagebox.showerror("Erro", "Nenhum pavimento encontrado nas combinações salvas!")
            return

        # 2. Apresentar um diálogo para o usuário selecionar um pavimento
        selected_pavimento = StringVar()
        dialog = Toplevel(self)
        dialog.title("Selecionar Pavimento para Combinações")
        dialog.geometry("350x150")
        dialog.grab_set()
        
        ttk.Label(dialog, text="Escolha o pavimento para gerar os scripts de combinações:").pack(pady=10)
        combo = ttk.Combobox(dialog, values=pavimentos, textvariable=selected_pavimento, state="readonly")
        combo.pack(pady=5)
        if pavimentos: # Define a primeira opção como padrão se houver pavimentos
            combo.current(0)

        confirmed = {'ok': False}
        def confirmar():
            confirmed['ok'] = True
            dialog.destroy()
        def cancelar():
            dialog.destroy()
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="OK", command=confirmar).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=cancelar).pack(side=tk.LEFT, padx=10)
        
        self.wait_window(dialog) # Espera a janela de diálogo ser fechada

        pavimento_selecionado = selected_pavimento.get()
        if not confirmed['ok'] or not pavimento_selecionado or pavimento_selecionado not in pavimentos:
            messagebox.showerror("Operação Cancelada", "Operação cancelada ou pavimento inválido.")
            return
        
        total_gerados = 0
        erros = []
        
        diretorio_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SCRIPTS")
        diretorio_saida_combinados = os.path.join(diretorio_base, pavimento_selecionado, "Combinados")
        
        if not os.path.exists(diretorio_saida_combinados):
            os.makedirs(diretorio_saida_combinados)

        # 3. Iterar sobre as combinações e gerar scripts
        for combinacao in self.combinacoes:
            # Verifica se o primeiro fundo da combinação pertence ao pavimento selecionado
            if combinacao['ids']:
                primeiro_fundo_id = combinacao['ids'][0]
                primeiro_fundo_dados = self.fundos_salvos.get(primeiro_fundo_id)
                if primeiro_fundo_dados and str(primeiro_fundo_dados.get('pavimento', '')).strip() == pavimento_selecionado:
                    try:
                        sucesso, caminho_arquivo = self.gerador_combinados.gerar_script_combinado(
                            combinacao, 
                            self.fundos_salvos, 
                            diretorio_saida_combinados
                        )
                        if sucesso:
                            total_gerados += 1
                        else:
                            erros.append(f"Combinação '{combinacao.get('nome', combinacao['ids'])}': Não foi possível gerar o script.")
                    except Exception as e:
                        erros.append(f"Combinação '{combinacao.get('nome', combinacao['ids'])}': Erro: {str(e)}")
        
        if not erros:
            messagebox.showinfo("Scripts Gerados", f"Todos os scripts de combinações ({total_gerados}) do pavimento '{pavimento_selecionado}' foram gerados com sucesso!")
        else:
            messagebox.showerror("Erros na Geração de Scripts", f"{total_gerados} scripts de combinações gerados. Erros:\n" + '\n'.join(erros))

        # 6. Executar o Combinador_VIGA.py e o Ordenador_VIGA.py
        try:
            combinador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Combinador_VIGA.py")
            subprocess.run(["python", combinador_path, diretorio_saida_combinados], check=True)
            messagebox.showinfo("Execução de Script", f"Combinador_VIGA.py executado com sucesso em: {diretorio_saida_combinados}")
        except Exception as e:
            messagebox.showerror("Erro de Execução", f"Erro ao executar o Combinador_VIGA.py: {e}")
            return

        try:
            ordenador_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Ordenador_VIGA.py")
            subprocess.run(["python", ordenador_path, diretorio_saida_combinados], check=True)
            messagebox.showinfo("Execução de Script", f"Ordenador_VIGA.py executado com sucesso em: {diretorio_saida_combinados}")
        except Exception as e:
            messagebox.showerror("Erro de Execução", f"Erro ao executar o Ordenador_VIGA.py: {e}")
            return

if __name__ == "__main__":
    app = FundoProducaoApp()
    app.mainloop()