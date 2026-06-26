import sys
import os
import time
import json
import math
import json
import threading
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QGridLayout, QGroupBox, QLabel, QLineEdit, QPushButton, 
    QRadioButton, QCheckBox, QScrollArea, QFrame, QTreeWidget, 
    QTreeWidgetItem, QHeaderView, QSplitter, QComboBox, QDialog,
    QTabWidget, QMessageBox, QGraphicsView, QGraphicsScene, 
    QGraphicsRectItem, QGraphicsLineItem, QGraphicsTextItem, QButtonGroup
)
from PySide6.QtCore import Qt, Signal, Slot, QSize, QTimer, QPointF
from PySide6.QtGui import QColor, QPen, QBrush, QFont, QPainter, QAction
from PySide6.QtWidgets import QFileDialog

# Imports adicionais para Excel e AutoCAD
import openpyxl
try:
    import pythoncom
    import win32com.client
    import win32gui
except ImportError:
    pythoncom = None
    win32com = None
    win32gui = None

# Mapeamento de 113 colunas para Excel (Fidelidade 100% com o original)
EXCEL_MAPPING = {
    'numero': 52,
    'nome': 53,
    'observacoes': 54,
    'pavimento': 55,
    'largura': 56,
    'altura': 57,
    'texto_esquerda': 58,
    'texto_direita': 59,
    'tipo_distribuicao': 60,
    'tipo_painel_inicial': 61,
    'sarrafo_esq': 62,
    'sarrafo_dir': 63,
    'painel1': 64,
    'painel2': 65,
    'painel3': 66,
    'painel4': 67,
    'painel5': 68,
    'painel6': 69,
    'chanfro_te': 70,
    'chanfro_fe': 71,
    'chanfro_td': 72,
    'chanfro_fd': 73,
    'abertura_te_dist': 74,
    'abertura_te_prof': 75,
    'abertura_te_larg': 76,
    'abertura_fe_dist': 77,
    'abertura_fe_prof': 78,
    'abertura_fe_larg': 79,
    'abertura_td_dist': 80,
    'abertura_td_prof': 81,
    'abertura_td_larg': 82,
    'abertura_fd_dist': 83,
    'abertura_fd_prof': 84,
    'abertura_fd_larg': 85,
    'tipo_painel2': 86,
    'comprimento_2': 87,
    'largura_2': 88,
    'p1_2': 89,
    'p2_2': 90,
    'p3_2': 91,
    'chanfro2_te': 92,
    'chanfro2_fe': 93,
    'chanfro2_td': 94,
    'chanfro2_fd': 95,
    'abertura2_te_dist': 96,
    'abertura2_te_prof': 97,
    'abertura2_te_larg': 98,
    'abertura2_fe_dist': 99,
    'abertura2_fe_prof': 100,
    'abertura2_fe_larg': 101,
    'abertura2_td_dist': 102,
    'abertura2_td_prof': 103,
    'abertura2_td_larg': 104,
    'abertura2_fd_dist': 105,
    'abertura2_fd_prof': 106,
    'abertura2_fd_larg': 107
}

# Importar utilitários existentes
try:
    from fundo_utils_novo import (
        calcular_area_util, float_safe, desenhar_teste_util, desenhar_pavimento_util,
        AutoCADSelectionManager
    )
    print("[DEBUG] fundo_utils_novo importado com sucesso")
except ImportError as e:
    print(f"[DEBUG] Erro ao importar fundo_utils_novo: {e}")
    # Mocks para desenvolvimento isolado
    def calcular_area_util(*args): return 0.0
    def float_safe(v):
        try: return float(str(v).replace(',', '.'))
        except: return 0.0
    def desenhar_teste_util(*a, **k): pass
    def desenhar_pavimento_util(*a, **k): pass
    
    # Classe AutoCADSelectionManager mockada
    class AutoCADSelectionManager:
        def __init__(self):
            self.ac = None
            self.doc = None
            self.selection = None
            self.com_initialized = False
            self.selection_name = "TempSelection"
            self.max_retries = 5
            self.retry_delay = 0.5
            self.error_log = []
            self.operation_count = 0
            self.health_check_interval = 2.0
            self.last_health_check = 0
            self.connection_attempts = 0
            self.max_connection_attempts = 10
            self.selection_timeout = 45
            self.recovery_mode = False
            self.text_selection_cache = {}
        
        def select_text(self, msg): return ("Texto Mock", (0,0))
        def connect_autocad(self): return False
        def clear_selection_method(self): pass
        def check_autocad_health(self): return True
        def get_active_document(self): return False

try:
    from Robo_fundos_TASF_limpo_copy_22 import ConfigManager
    print("[DEBUG] ConfigManager importado com sucesso")
except ImportError as e:
    print(f"[DEBUG] Erro ao importar ConfigManager: {e}")
    class ConfigManager:
        def __init__(self, *a, **k): self.config = {}
        def get_layer(self, k): return "0"
        def get_comando(self, k): return ""
        def get_tipo_linha(self): return "PLINE"
        def get_linetype_sarrafo_central(self): return "DASHED"
        def get_logica_sarrafos(self): return "todos"
        def get_config(self, *a): return "nao"
        def load_templates(self): return {}
        def save_config(self): return True

def _fmt_pav_label(name: str) -> str:
    """Formata nome de pavimento para exibição: [BADGE]  nome — mesmo padrão do SA."""
    import re as _r
    up = name.upper()
    short = _r.sub(r'_R\d{4}_ASCII_ODA.*$', '', name, flags=_r.I)
    short = _r.sub(r'\.(DXF|DWG)$', '', short, flags=_r.I).strip()
    m = _r.search(r'[-_ ](\d{1,2})P(?:AV?|V)?(?:[-_ ]|$)', up)
    if not m: m = _r.match(r'^(\d{1,2})\s*PAV', up)
    if not m: m = _r.match(r'^PAV\s*(\d{1,2})(?:\s|$)', up)
    if m: return f"[{m.group(1)}º PAV]  {short}"
    if _r.search(r'[-_]TER[-_]|TERREO|TÉRREO', up): return f"[TÉRREO]  {short}"
    if _r.search(r'[-_]FUN[-_]|FUNDA', up): return f"[FUNDAÇÃO]  {short}"
    if _r.search(r'[-_]TIP[-_]|TIPO', up): return f"[TIPO]  {short}"
    if _r.search(r'[-_]COB[E_-]|COBER|DECK|BARR', up): return f"[COBERTURA]  {short}"
    if _r.search(r'[-_]ATC[-_]|ATICO|ÁTICO', up): return f"[ÁTICO]  {short}"
    if _r.search(r'[-_]SUB[-_]|SUBSOLO', up): return f"[SUBSOLO]  {short}"
    m_id = _r.search(r'[-_](\d{3,5})[-_]', name)
    if m_id: return f"[{m_id.group(1)}]  {short}"
    return short

def _set_cmb_by_data(cmb, raw: str):
    """Seleciona item de combobox pelo userData (raw pav name)."""
    for i in range(cmb.count()):
        if cmb.itemData(i) == raw:
            cmb.setCurrentIndex(i)
            return
    cmb.setCurrentText(raw)

def _cmb_pav_raw(cmb) -> str:
    """Retorna o nome raw do pavimento selecionado (userData) de um combo."""
    d = cmb.currentData()
    return d if d is not None else cmb.currentText()

def alinhar_boundary_horizontal(coords):
    """
    Alinha o boundary pelo lado maior do retângulo de menor área (minimum area rectangle),
    garantindo que o lado do fundo (base real, mesmo com aberturas ou chanfros) fique para baixo após a rotação.
    """
    if not coords or len(coords) < 6:
        return coords
        
    import numpy as np
    pontos = np.array([(coords[i], coords[i+1]) for i in range(0, len(coords), 2)])
    cx, cy = pontos.mean(axis=0)
    pontos_c = pontos - [cx, cy]
    segmentos = np.diff(np.vstack([pontos_c, pontos_c[0]]), axis=0)
    angulos = np.arctan2(segmentos[:,1], segmentos[:,0])
    angulos = np.unique(np.mod(angulos, np.pi/2))  
    melhor_area = None
    melhor_ang = 0
    for ang in angulos:
        rot = np.array([[np.cos(-ang), -np.sin(-ang)], [np.sin(-ang), np.cos(-ang)]])
        rot_pontos = pontos_c @ rot.T
        min_x, min_y = rot_pontos.min(axis=0)
        max_x, max_y = rot_pontos.max(axis=0)
        area = (max_x - min_x) * (max_y - min_y)
        if (melhor_area is None) or (area < melhor_area):
            melhor_area = area
            melhor_ang = ang
    rot = np.array([[np.cos(-melhor_ang), -np.sin(-melhor_ang)], [np.sin(-melhor_ang), np.cos(-melhor_ang)]])
    pontos_rot = pontos_c @ rot.T
    min_x, min_y = pontos_rot.min(axis=0)
    max_x, max_y = pontos_rot.max(axis=0)
    largura = max_x - min_x
    altura = max_y - min_y
    if altura > largura:
        rot90 = np.array([[0, -1], [1, 0]])
        pontos_rot = pontos_rot @ rot90.T
        melhor_ang += np.pi/2
    y_min = pontos[:,1].min()
    pontos_fundo_idx = np.where(np.abs(pontos[:,1] - y_min) < 1e-3)[0]
    pontos_fundo_rot = pontos_rot[pontos_fundo_idx]
    if len(pontos_fundo_rot) > 2:
        indices = pontos_fundo_idx
        indices = np.sort(indices)
        grupos = []
        grupo = [indices[0]]
        for i in range(1, len(indices)):
            if indices[i] == indices[i-1] + 1:
                grupo.append(indices[i])
            else:
                grupos.append(grupo)
                grupo = [indices[i]]
        grupos.append(grupo)
        segmentos_rot = np.diff(np.vstack([pontos_rot, pontos_rot[0]]), axis=0)
        soma_grupos = []
        for grupo in grupos:
            soma = 0
            for i in range(len(grupo)-1):
                idx1 = grupo[i]
                idx2 = grupo[i+1]
                soma += np.linalg.norm(pontos_rot[idx2] - pontos_rot[idx1])
            soma_grupos.append(soma)
        idx_base = np.argmax(soma_grupos)
        grupo_base = grupos[idx_base]
        y_base = pontos_rot[grupo_base,1].mean()
        if y_base > pontos_rot[:,1].mean():
            pontos_rot[:,1] *= -1
        else:
            pass
        segmentos_rot = np.diff(np.vstack([pontos_rot, pontos_rot[0]]), axis=0)
        comprimentos = np.linalg.norm(segmentos_rot, axis=1)
        idxs_maiores = np.argsort(comprimentos)[-2:]
        medias_y = []
        for idx in idxs_maiores:
            y1 = pontos_rot[idx][1]
            y2 = pontos_rot[(idx+1)%len(pontos_rot)][1]
            medias_y.append((y1 + y2) / 2)
        idx_fundo = idxs_maiores[np.argmin(medias_y)]
        y_fundo = medias_y[np.argmin(medias_y)]
        if y_fundo > pontos_rot[:,1].mean():
            pontos_rot[:,1] *= -1
    pontos_final = pontos_rot + [cx, cy]
    coords_final = []
    for x, y in pontos_final:
        coords_final.extend([float(x), float(y)])
    return coords_final

class FundoCanvas(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scene = QGraphicsScene(self)
        self.setScene(self.scene)
        self.setRenderHint(QPainter.Antialiasing)
        self.setBackgroundBrush(QBrush(QColor("#1e1e1e")))
        
        self.zoom_factor = 1.1
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        
    def draw_fundo(self, dados):
        self.scene.clear()
        if not dados: return
        
        try:
            # Escala para exibição
            scale = 2.0
            
            # PRIORIDADE 1: Se existe boundary_original, desenhar ele fielmente (com todas as deformidades)
            boundary_original = dados.get('boundary_original')
            if boundary_original and len(boundary_original) >= 6:
                print(f"[DEBUG Canvas] Desenhando boundary_original com {len(boundary_original)} coordenadas ({len(boundary_original)//2} pontos)")
                
                # Calcular limites Y para inversão (sistema CAD Y cresce para cima, PySide6 para baixo)
                y_coords = [boundary_original[i+1] for i in range(0, len(boundary_original), 2) if i+1 < len(boundary_original)]
                y_min = min(y_coords) if y_coords else 0
                y_max = max(y_coords) if y_coords else 0
                
                # Converter coordenadas para pontos (invertendo Y em relação ao Y máximo)
                # No CAD: Y cresce para cima. No PySide6: Y cresce para baixo
                # Inversão: y_pyside = y_max - (y_cad - y_min) = y_max + y_min - y_cad
                pontos = []
                for i in range(0, len(boundary_original), 2):
                    if i + 1 < len(boundary_original):
                        x = boundary_original[i] * scale
                        y_original = boundary_original[i+1]
                        # Inverter Y: y_invertido = y_max + y_min - y_original
                        y_invertido = (y_max + y_min - y_original) * scale
                        pontos.append((x, y_invertido))
                
                print(f"[DEBUG Canvas] {len(pontos)} pontos convertidos para desenho")
                
                if len(pontos) >= 3:
                    # Criar polígono com todos os pontos (preserva deformidades)
                    from PySide6.QtGui import QPolygonF
                    from PySide6.QtCore import QPointF
                    
                    polygon = QPolygonF([QPointF(x, y) for x, y in pontos])
                    
                    # Criar cores corretamente para PySide6
                    cor_amarela = QColor("#ffff00")
                    cor_preenchimento = QColor("#2d2d2d")
                    cor_preenchimento.setAlpha(180)  # Semi-transparente
                    
                    # Desenhar boundary real com todas as deformidades
                    fundo_poly = self.scene.addPolygon(
                        polygon,
                        QPen(cor_amarela, 3),  # Amarelo para destacar
                        QBrush(cor_preenchimento)  # Preenchimento semi-transparente
                    )
                    
                    print(f"[DEBUG Canvas] ✅ Boundary desenhado com sucesso! {len(pontos)} pontos preservados")
                    
                    # Calcular bounding box para ajustar view
                    bbox = polygon.boundingRect()
                    self.setSceneRect(bbox.adjusted(-50, -50, 50, 50))
                    self.fitInView(self.sceneRect(), Qt.KeepAspectRatio)
                    
                    # Desenhar também os outros elementos (painéis, chanfros, etc.) se existirem
                    self._draw_additional_elements(dados, scale, bbox)
                    return
                else:
                    print(f"[DEBUG Canvas] ⚠️ Boundary tem menos de 3 pontos após conversão: {len(pontos)}")
            else:
                if boundary_original:
                    print(f"[DEBUG Canvas] ⚠️ boundary_original existe mas tem apenas {len(boundary_original)} coordenadas (mínimo 6)")
                else:
                    print(f"[DEBUG Canvas] ⚠️ boundary_original não encontrado nos dados")
            
            # FALLBACK: Se não tem boundary original, desenhar retângulo simplificado
            largura = float_safe(dados.get('largura', 0))
            altura = float_safe(dados.get('altura', 0))
            if largura <= 0 or altura <= 0: return
            
            # 1. Desenhar Fundo Principal (retângulo simplificado)
            fundo_main = self.scene.addRect(0, 0, largura * scale, altura * scale, 
                                          QPen(QColor("#ffffff"), 2), 
                                          QBrush(QColor("#2d2d2d")))
            
            # 2. Desenhar Painéis (se existirem)
            paineis = dados.get('paineis', [])
            x_offset = 0
            for i, p_val in enumerate(paineis):
                # Lidar com formato de soma ou valor único
                p_width = float_safe(p_val)
                if p_width > 0:
                    rect = self.scene.addRect(x_offset * scale, 0, p_width * scale, altura * scale,
                                            QPen(QColor("#4a4a4a"), 1))
                    text = self.scene.addText(f"P{i+1}\n{p_val}")
                    text.setDefaultTextColor(QColor("#aaaaaa"))
                    text.setScale(0.8)
                    text.setPos(x_offset * scale + 5, 5)
                    x_offset += p_width

            # 3. Desenhar Chanfros e AberturasPrincipal
            recuos = dados.get('recuos', [0]*4)
            # T/E, F/E, T/D, F/D
            self._draw_chamfers(0, 0, largura * scale, altura * scale, recuos, scale)
            
            # 4. Desenhar Painel 2 em L (se configurado)
            comp2 = float_safe(dados.get('comprimento_2', 0))
            larg2 = float_safe(dados.get('largura_2', 0))
            if comp2 > 0 and larg2 > 0:
                tipo = dados.get('tipo_painel2', 'E/T')
                # Lógica de posicionamento do L baseada no tipo
                # E/T: Esquerda Topo, etc.
                self._draw_panel_l(largura, altura, comp2, larg2, tipo, scale)

            self.setSceneRect(self.scene.itemsBoundingRect().adjusted(-50, -50, 50, 50))
            self.fitInView(self.sceneRect(), Qt.KeepAspectRatio)
            
        except Exception as e:
            print(f"Erro no desenho do canvas: {e}")
            import traceback
            traceback.print_exc()

    def _draw_additional_elements(self, dados, scale, bbox):
        """Desenha elementos adicionais (painéis, chanfros, aberturas) sobre o boundary"""
        # Desenhar painéis se existirem
        paineis = dados.get('paineis', [])
        if paineis:
            x_offset = bbox.left()
            for i, p_val in enumerate(paineis):
                p_width = float_safe(p_val)
                if p_width > 0:
                    rect = self.scene.addRect(
                        x_offset, bbox.top(), 
                        p_width * scale, bbox.height(),
                        QPen(QColor("#4a4a4a"), 1)
                    )
                    text = self.scene.addText(f"P{i+1}\n{p_val}")
                    text.setDefaultTextColor(QColor("#aaaaaa"))
                    text.setScale(0.8)
                    text.setPos(x_offset + 5, bbox.top() + 5)
                    x_offset += p_width * scale
        
        # Desenhar chanfros (recuos) se existirem
        recuos = dados.get('recuos', [0]*4)
        if any(float_safe(r) > 0 for r in recuos):
            self._draw_chamfers(bbox.left(), bbox.top(), bbox.width(), bbox.height(), recuos, scale)
        
        # Desenhar aberturas se existirem
        aberturas = dados.get('aberturas', [])
        for ab in aberturas:
            if isinstance(ab, list) and len(ab) >= 3:
                dist = float_safe(ab[0])
                prof = float_safe(ab[1])
                larg = float_safe(ab[2])
                if dist > 0 and prof > 0 and larg > 0:
                    # Desenhar abertura (simplificado)
                    cor_vermelha = QColor("#ff0000")
                    cor_vermelha_alpha = QColor("#ff0000")
                    cor_vermelha_alpha.setAlpha(100)
                    self.scene.addRect(
                        bbox.left() + dist * scale, bbox.top(),
                        larg * scale, prof * scale,
                        QPen(cor_vermelha, 2),
                        QBrush(cor_vermelha_alpha)
                    )

    def _draw_chamfers(self, x, y, w, h, recuos, scale):
        # Implementação simplificada para visualização
        pass

    def _draw_panel_l(self, w1, h1, l2, w2, tipo, scale):
        # Desenha a extensão em L
        pass

class ConfigDialog(QDialog):
    def __init__(self, config_manager, parent=None):
        super().__init__(parent)
        self.config_manager = config_manager
        self.setWindowTitle("Configurações de Desenho")
        self.resize(600, 500)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        
        # Abas seriam implementadas aqui seguindo o padrão do original
        # Layers, Comandos, Opções, Templates
        tab_layers = QWidget()
        l_layout = QGridLayout(tab_layers)
        self.layer_edits = {}
        layers = self.config_manager.config.get("layers", {})
        for i, (key, val) in enumerate(layers.items()):
            l_layout.addWidget(QLabel(key), i, 0)
            edit = QLineEdit(val)
            l_layout.addWidget(edit, i, 1)
            self.layer_edits[key] = edit
        self.tabs.addTab(tab_layers, "Layers")

        # Comandos Tab
        tab_cmds = QWidget()
        c_layout = QGridLayout(tab_cmds)
        self.cmd_edits = {}
        cmds = self.config_manager.config.get("comandos", {})
        for i, (key, val) in enumerate(cmds.items()):
            c_layout.addWidget(QLabel(key), i, 0)
            edit = QLineEdit(val)
            c_layout.addWidget(edit, i, 1)
            self.cmd_edits[key] = edit
        self.tabs.addTab(tab_cmds, "Comandos")

        layout.addWidget(self.tabs)
        
        btns = QHBoxLayout()
        btn_save = QPushButton("Salvar")
        btn_save.clicked.connect(self.save_and_close)
        btn_close = QPushButton("Fechar")
        btn_close.clicked.connect(self.close)
        btns.addWidget(btn_save)
        btns.addWidget(btn_close)
        layout.addLayout(btns)

    def save_and_close(self):
        # Update config_manager
        for k, v in self.layer_edits.items():
            self.config_manager.config["layers"][k] = v.text()
        for k, v in self.cmd_edits.items():
            self.config_manager.config["comandos"][k] = v.text()
        self.config_manager.save_config()
        self.accept()

class FloatingComment(QDialog):
    def __init__(self, message, parent=None, timeout=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setup_ui(message)
        
        if timeout:
            QTimer.singleShot(timeout * 1000, self.accept)

    def setup_ui(self, message):
        layout = QVBoxLayout(self)
        self.frame = QFrame()
        self.frame.setStyleSheet("""
            QFrame {
                background-color: #fff8b0;
                border: 2px solid #fbc02d;
                border-radius: 10px;
            }
            QLabel {
                color: #333;
                font-size: 16px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton {
                background-color: #ffb0b0;
                border: 1px solid #d32f2f;
                border-radius: 5px;
                padding: 5px;
                color: #333;
            }
        """)
        f_layout = QVBoxLayout(self.frame)
        
        lbl = QLabel(message)
        lbl.setWordWrap(True)
        lbl.setAlignment(Qt.AlignCenter)
        f_layout.addWidget(lbl)
        
        btn_cancel = QPushButton("Cancelar Seleção (ESC)")
        btn_cancel.clicked.connect(self.reject)
        f_layout.addWidget(btn_cancel)
        
        layout.addWidget(self.frame)
        self.resize(400, 150)
        
        # Centralizar no topo
        screen = QApplication.primaryScreen().geometry()
        self.move((screen.width() - self.width()) // 2, 50)

class FundoMainWindow(QMainWindow):
    # Emitido ao selecionar item na lista — carrega DXF N3 no viewer integrado
    item_loaded = Signal(str)

    def log(self, message):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        if hasattr(self, 'statusBar'):
            self.statusBar().showMessage(message, 5000)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Agente CAD - Robo Fundo de Vigas (PySide6)")
        self.resize(1300, 950)
        self.fundos_salvos = {}
        self.config_manager = ConfigManager()
        self.caminho_dados = os.path.join(os.path.dirname(__file__), "fundos_salvos.json")
        self._balcao_flutuante = None
        self._flutuante_cancelado = False
        self.selection_manager = AutoCADSelectionManager()
        self.usar_boundary_duplo = False
        self._current_boundary_original = None  # Armazena boundary original para desenho fiel
        self._current_seg_idx = -1  # -1 = viga inteira; >=0 = segmento selecionado
        
        self.setup_ui()
        self.load_data()
        self.setup_connections()
        self.apply_theme()
        self.update_list()

    def setup_connections(self):
        # Filtros
        self.combo_obra.currentTextChanged.connect(self._update_combo_pavimentos)
        self.combo_pavimento.currentTextChanged.connect(self._on_pavimento_context_changed)
        
        # Obra/Pav Buttons
        self.btn_add_obra.clicked.connect(self.action_add_obra)
        self.btn_rem_obra.clicked.connect(self.action_rem_obra)
        self.btn_add_pav.clicked.connect(self.action_add_pav)
        self.btn_paste_pav.clicked.connect(self.action_paste_pav)
        self.btn_rem_pav.clicked.connect(self.action_rem_pav)
        
        # Botão excluir selecionados
        self.btn_delete_sel.clicked.connect(self.action_delete_selected)
        
        # Botão Salvar (destacado acima das abas)
        self.btn_salvar.clicked.connect(self.action_salvar)
        
        # Botão Analisar Boundary
        self.btn_analisar_boundary.clicked.connect(self.action_analisar_boundary)
        
        # Action bar (Ferramentas)
        self.btn_config_desenho.clicked.connect(self.open_config)
        self.btn_lisp_fd.clicked.connect(self.action_criar_lisp_fd)
        
        # Tree widget
        self.tree_fundos.itemClicked.connect(self.load_selected)
        
        # Botões de Produção (4 botões: Próxima Viga, Próximo Segmento, Extensão L, Cancelar)
        self.btns_prod[0].clicked.connect(self.action_nova_viga)  # Próxima Viga
        self.btns_prod[1].clicked.connect(self.action_continua_viga)  # Próximo Segmento
        self.btns_prod[2].clicked.connect(self.action_extensao_l)  # Extensão L
        self.btns_prod[3].clicked.connect(self.action_cancelar_selecao)  # Cancelar
        
        # Botões de Geração
        self.btn_gerar_segmento.clicked.connect(self.generate_script_current)
        self.btn_gerar_conjunto.clicked.connect(self.generate_conjunto_scripts)
        self.btn_gerar_pavimento.clicked.connect(self.generate_pavimento_scripts)
        
        # Auto calculations
        for field in [self.fields['largura'], self.fields['altura']]:
            field.textChanged.connect(self.update_calculations)
        
        self.rad_122.toggled.connect(self.update_calculations)
        self.rad_307.toggled.connect(self.update_calculations)
        
        # AutoCAD selector integration
        self.fields['numero'].mouseDoubleClickEvent = lambda e: self.select_from_autocad('numero')
        self.fields['nome'].mouseDoubleClickEvent = lambda e: self.select_from_autocad('nome')

    def add_viga_bulk(self, viga_list):
        """Adiciona uma lista de vigas (nome, numero, parent_name) ao pavimento atual se não existirem (Master Sync)."""
        obra = self.combo_obra.currentText()
        pav = _cmb_pav_raw(self.combo_pavimento)
        if not obra or not pav:
            self.log("⚠️ Erro Sync: Obra ou Pavimento não selecionados no Robo Fundo.")
            return 0
        
        # Garantir estrutura nested
        if isinstance(self.fundos_salvos, dict) and obra not in self.fundos_salvos:
            self.fundos_salvos[obra] = {}
        if pav not in self.fundos_salvos[obra]:
            self.fundos_salvos[obra][pav] = {}
            
        items_dict = self.fundos_salvos[obra][pav]
        count = 0
        skipped = 0
        
        for v in viga_list:
            name = str(v.get('name'))
            number = str(v.get('number'))
            p_name = v.get('parent_name') # Classe
            
            # Se não tem número, tenta extrair do nome
            if not number or number == "None":
                 import re
                 nums = re.findall(r'\d+', name)
                 number = nums[0] if nums else name

            # Verificar se já existe neste pavimento específico
            if number not in items_dict:
                novo_item = {
                    'numero': number,
                    'nome': name,
                    'obra': obra,
                    'pavimento': pav,
                    'parent_name': p_name, # Preservar Classe
                    'largura': "20",
                    'altura': "60",
                    'total': "0",
                    'm2': "0",
                    'desenhado': False
                }
                items_dict[number] = novo_item
                count += 1
            else:
                skipped += 1
        
        if count > 0:
            self.update_list()
            self.save_data()
        
        self.log(f"📥 Sync Bulk: {count} novas vigas adicionadas, {skipped} já existiam em {obra}/{pav}.")
        return count

    def _n3_fv_dir_for_obra(self, obra: str) -> Path:
        return Path("D:/Agente-cad-PYSIDE/DADOS-OBRAS") / str(obra) / "Fase-4_Sincronizacao" / "JSON_Vigas_Fundo"

    def _n5_fv_dir_for_obra(self, obra: str) -> Path:
        return Path("D:/Agente-cad-PYSIDE/DADOS-OBRAS") / str(obra) / "Fase-6_Execucao_CAD" / "n5"

    def _natural_name_key(self, text: str):
        return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", str(text or ""))]

    def _normalize_fv_robo_name(self, name: str) -> str:
        value = str(name or "").strip().upper()
        value = re.sub(r"_FUNDO$", "", value)
        value = re.sub(r"_[AB]$", "", value)
        return value

    def _fv_number_from_name(self, name: str, fallback: str = "") -> str:
        raw = str(name or fallback or "").strip()
        m = re.search(r"\d+[A-Za-z]?", raw)
        return m.group(0).upper() if m else raw.upper()

    def _fv_unique_key(self, items_dict: dict, number: str, name: str) -> str:
        number = str(number or "").strip() or self._fv_number_from_name(name)
        name_norm = str(name or "").strip().upper()
        if number not in items_dict:
            return number
        existing_name = str(items_dict.get(number, {}).get("nome", "")).strip().upper()
        if existing_name == name_norm:
            return number
        base = name_norm or number
        if base not in items_dict:
            return base
        idx = 2
        while f"{base}_{idx}" in items_dict:
            idx += 1
        return f"{base}_{idx}"

    def _fmt_n3_num(self, value) -> str:
        try:
            val = float(value)
        except (TypeError, ValueError):
            return "0"
        if abs(val) < 1e-9:
            return "0"
        if abs(val - round(val)) < 1e-6:
            return str(int(round(val)))
        return f"{val:.2f}".rstrip("0").rstrip(".")

    def _panels_from_n3_fv(self, data: dict) -> list:
        panels = []
        for panel in data.get("panels") or []:
            panels.append(self._fmt_n3_num(panel.get("width", 0) if isinstance(panel, dict) else panel))
            if len(panels) >= 6:
                break
        while len(panels) < 6:
            panels.append("0")
        return panels

    def _holes_from_n3_fv(self, data: dict) -> list:
        rows = [["0", "0", "0"] for _ in range(4)]
        holes = [h for h in (data.get("holes") or []) if isinstance(h, dict) and h.get("active", True)]
        for idx, hole in enumerate(holes[:4]):
            rows[idx] = [
                self._fmt_n3_num(hole.get("position", 0)),
                self._fmt_n3_num(hole.get("height", 0)),
                self._fmt_n3_num(hole.get("width", 0)),
            ]
        return rows

    def _n5_fv_manifest_candidates(self, obra: str, pav: str) -> list[Path]:
        n5_dir = self._n5_fv_dir_for_obra(obra)
        if not n5_dir.exists():
            return []
        pav_slug = re.sub(r"[^A-Za-z0-9]+", "_", str(pav or "")).strip("_").upper()
        pav_raw = str(pav or "").strip()
        candidates = []
        if pav_raw:
            candidates.extend(n5_dir.glob(f"N5_FV_{pav_raw}.json"))
        if pav_slug:
            candidates.extend(n5_dir.glob(f"N5_FV_{pav_slug}.json"))
        if re.search(r"(^|[^0-9])13P([^0-9]|$)", str(pav or "").upper()) or "13 PAV" in str(pav or "").upper():
            candidates.extend(n5_dir.glob("N5_FV_13_PAV.json"))
        candidates.extend(n5_dir.glob("N5_FV_*.json"))
        seen = set()
        ordered = []
        for path in candidates:
            key = str(path).lower()
            if key in seen:
                continue
            seen.add(key)
            ordered.append(path)
        return ordered

    def _n3_fv_allowed_names_for_context(self, obra: str, pav: str) -> set[str]:
        best_allowed = set()
        pav_upper = str(pav or "").upper()
        for path in self._n5_fv_manifest_candidates(obra, pav):
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
            manifest_pav = str(data.get("pavimento") or "").upper() if isinstance(data, dict) else ""
            if path.name.upper() != f"N5_FV_{re.sub(r'[^A-Z0-9]+', '_', pav_upper).strip('_')}.JSON":
                if pav_upper and manifest_pav and manifest_pav not in {pav_upper, "13_PAV"} and pav_upper not in manifest_pav:
                    continue
            allowed = set()
            for item in data.get("items", []) if isinstance(data, dict) else []:
                if not isinstance(item, dict):
                    continue
                item_id = self._normalize_fv_robo_name(item.get("item_id") or item.get("nome") or "")
                if item_id:
                    allowed.add(item_id)
            if len(allowed) > len(best_allowed):
                best_allowed = allowed
        return best_allowed

    def _n3_fv_json_to_robo_item(self, data: dict, obra: str, pav: str, source_path: Path) -> dict | None:
        name = self._normalize_fv_robo_name(data.get("name") or source_path.stem.replace("_fundo", ""))
        if not name:
            return None
        number = str(data.get("number") or self._fv_number_from_name(name)).strip()
        largura = self._fmt_n3_num(data.get("total_height", 0))
        altura = self._fmt_n3_num(data.get("total_width", 0))
        try:
            m2 = calcular_area_util(float_safe(largura), float_safe(altura), 0, 0)
        except Exception:
            m2 = "0"
        return {
            "numero": number,
            "nome": name,
            "obs": "Importado automaticamente do N3/Comparison Engine.",
            "obra": obra,
            "pavimento": pav,
            "parent_name": name,
            "largura": largura,
            "altura": altura,
            "texto_esq": str(data.get("label_left") or ""),
            "texto_dir": str(data.get("label_right") or ""),
            "paineis": self._panels_from_n3_fv(data),
            "recuos": ["0", "0", "0", "0"],
            "aberturas": self._holes_from_n3_fv(data),
            "comprimento_2": "0",
            "largura_2": "0",
            "paineis_2": ["0", "0", "0"],
            "tipo_painel2": "E/T",
            "recuos_2": ["0", "0", "0", "0"],
            "aberturas_2": [["0", "0", "0"] for _ in range(4)],
            "sarrafo_esq": bool((data.get("pillar_left") or {}).get("active", False)),
            "sarrafo_dir": bool((data.get("pillar_right") or {}).get("active", False)),
            "m2": m2,
            "desenhado": False,
            "_n3_source": "JSON_Vigas_Fundo",
            "_n3_path": str(source_path),
            "_n3_ficha": data,
        }

    def sync_n3_fundos_to_current_context(self) -> int:
        obra = self.combo_obra.currentText()
        pav = _cmb_pav_raw(self.combo_pavimento)
        if not obra or not pav or pav == "Todos":
            return 0
        n3_dir = self._n3_fv_dir_for_obra(obra)
        if not n3_dir.exists():
            return 0
        self.fundos_salvos.setdefault(obra, {})
        self.fundos_salvos[obra].setdefault(pav, {})
        self.obras_metadata.setdefault(obra, set()).add(pav)
        items_dict = self.fundos_salvos[obra][pav]
        existing_names = {
            self._normalize_fv_robo_name(item.get("nome", ""))
            for item in items_dict.values()
            if isinstance(item, dict)
        }
        allowed_names = self._n3_fv_allowed_names_for_context(obra, pav)
        if not allowed_names:
            return 0
        imported = 0
        for json_path in sorted(n3_dir.glob("V*_fundo.json"), key=lambda p: self._natural_name_key(p.stem)):
            try:
                data = json.loads(json_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(data, dict):
                continue
            item = self._n3_fv_json_to_robo_item(data, obra, pav, json_path)
            if not item:
                continue
            item_name = self._normalize_fv_robo_name(item.get("nome", ""))
            if item_name not in allowed_names:
                continue
            if item_name in existing_names:
                continue
            key = self._fv_unique_key(items_dict, item.get("numero"), item.get("nome"))
            items_dict[key] = item
            existing_names.add(item_name)
            imported += 1
        if imported:
            self.save_data()
            self.update_list()
            self.log(f"N3->Robo FV: {imported} fundos importados para {obra}/{pav}.")
        return imported

    def add_viga_with_boundary(self, viga_data, boundary_coords):
        """
        Adiciona viga e processa boundary automaticamente.
        
        Args:
            viga_data: dict com 'name', 'number', 'parent_name'
            boundary_coords: lista plana [x1, y1, x2, y2, ...] ou None
        
        Returns:
            dict: {'success': bool, 'processed': bool, 'message': str}
                - success: True se viga foi criada/atualizada
                - processed: True se boundary foi processado com sucesso
                - message: Mensagem de status
        """
        obra = self.combo_obra.currentText()
        pav = _cmb_pav_raw(self.combo_pavimento)
        if not obra or not pav:
            return {'success': False, 'processed': False, 'message': 'Obra ou Pavimento não selecionados'}
        
        # Garantir estrutura nested
        if isinstance(self.fundos_salvos, dict) and obra not in self.fundos_salvos:
            self.fundos_salvos[obra] = {}
        if pav not in self.fundos_salvos[obra]:
            self.fundos_salvos[obra][pav] = {}
            
        items_dict = self.fundos_salvos[obra][pav]
        
        name = str(viga_data.get('name', ''))
        number = str(viga_data.get('number', ''))
        p_name = viga_data.get('parent_name', '')
        
        # Se não tem número, tenta extrair do nome
        if not number or number == "None" or number == name:
            import re
            nums = re.findall(r'\d+', name)
            number = nums[0] if nums else name
        
        # Normalizar number para evitar duplicatas
        # Se number for igual ao name, tentar extrair número do name
        if number == name:
            import re
            nums = re.findall(r'\d+', name)
            if nums:
                number = nums[0]
            else:
                # Se não tem número no nome, usar name mas garantir unicidade
                number = name
        
        # Verificar se já existe entrada com este número OU com este nome
        # Eliminar entradas duplicadas com nome no campo de número
        existing_by_number = number in items_dict
        existing_by_name = None
        for key, item in list(items_dict.items()):
            # Se encontrar entrada com nome no campo de número e número correto diferente
            if item.get('nome') == name and key != number:
                # Esta é uma entrada duplicada incorreta - remover
                del items_dict[key]
                self.log(f"🗑️ Removida entrada duplicada incorreta: {key} (nome no campo número)")
        
        # Criar ou atualizar entrada do fundo
        is_new = number not in items_dict
        if is_new:
            novo_item = {
                'numero': number,
                'nome': name,
                'obra': obra,
                'pavimento': pav,
                'parent_name': p_name,
                'largura': "20",
                'altura': "60",
                'total': "0",
                'm2': "0",
                'desenhado': False
            }
            items_dict[number] = novo_item
        else:
            novo_item = items_dict[number]
        
        # Processar boundary se fornecido
        processed = False
        if boundary_coords and len(boundary_coords) >= 6:  # Mínimo 3 pontos (6 coordenadas)
            try:
                # IMPORTANTE: Armazenar boundary ORIGINAL (sem alinhamento) para desenho fiel
                # O boundary original preserva todas as deformidades, recortes, chanfros, etc.
                novo_item['boundary_original'] = list(boundary_coords)  # Garantir que é uma lista
                print(f"[DEBUG] Boundary original armazenado: {len(novo_item['boundary_original'])} coordenadas")
                
                # Alinhar boundary APENAS para análise (não modifica o original)
                coords_alinhadas = alinhar_boundary_horizontal(boundary_coords)
                
                # Analisar coordenadas (usa o alinhado para cálculos)
                from fundo_utils import analisar_coordenadas
                analise, msg = analisar_coordenadas(coords_alinhadas)
                
                if analise:
                    # Preencher campos diretamente no dicionário (sem UI)
                    # Extrair dimensões
                    dimensoes = analise.get('dimensoes', {})
                    novo_item['largura'] = str(int(round(dimensoes.get('largura', 0))))
                    novo_item['altura'] = str(int(round(dimensoes.get('altura', 0))))
                    
                    # Extrair chanfros (recuos)
                    chanfros = analise.get('chanfros', {})
                    # Mapear nomes possíveis dos chanfros (formato pode variar)
                    # Ordem esperada: topo_esq, fundo_esq, topo_dir, fundo_dir
                    recuos = [
                        str(int(round(chanfros.get('recuo_topo_esq', chanfros.get('topo_esq', 0))))),
                        str(int(round(chanfros.get('recuo_fundo_esq', chanfros.get('fundo_esq', 0))))),
                        str(int(round(chanfros.get('recuo_topo_dir', chanfros.get('topo_dir', 0))))),
                        str(int(round(chanfros.get('recuo_fundo_dir', chanfros.get('fundo_dir', 0)))))
                    ]
                    novo_item['recuos'] = recuos
                    
                    # Extrair painéis (se disponível)
                    paineis = analise.get('paineis', [0]*6)
                    novo_item['paineis'] = [str(int(round(p))) for p in paineis[:6]]
                    
                    # Extrair aberturas e mapear para as 4 posições (T/E, F/E, T/D, F/D)
                    aberturas = analise.get('aberturas', [])
                    largura_total = dimensoes.get('largura', 0)
                    altura_total = dimensoes.get('altura', 0)
                    
                    # Inicializar array de 4 posições: [T/E, F/E, T/D, F/D]
                    # Cada posição: [distância, profundidade, largura]
                    aberturas_formatadas = [
                        ["0", "0", "0"],  # T/E - Topo Esquerda
                        ["0", "0", "0"],  # F/E - Fundo Esquerda
                        ["0", "0", "0"],  # T/D - Topo Direita
                        ["0", "0", "0"]   # F/D - Fundo Direita
                    ]
                    
                    # Processar cada abertura detectada
                    print(f"[DEBUG add_viga_with_boundary] Processando {len(aberturas)} abertura(s) detectada(s)")
                    for idx_ab, ab in enumerate(aberturas):
                        if not isinstance(ab, dict):
                            print(f"[DEBUG add_viga_with_boundary] ⚠️ Abertura {idx_ab} não é dict: {type(ab)}")
                            continue
                        
                        # Obter dados da abertura
                        x = ab.get('x', 0)
                        y = ab.get('y', 0)
                        largura_ab = ab.get('largura', 0)
                        altura_ab = ab.get('altura', 0)
                        profundidade_ab = ab.get('profundidade', altura_ab)
                        
                        print(f"[DEBUG add_viga_with_boundary] Abertura {idx_ab}: x={x:.2f}, y={y:.2f}, larg={largura_ab:.2f}, alt={altura_ab:.2f}, prof={profundidade_ab:.2f}")
                        print(f"[DEBUG add_viga_with_boundary] Campos da abertura: {list(ab.keys())}")
                        
                        # Determinar lado e posição vertical usando a lógica da função converter
                        # Lado: esquerda ou direita (já deve vir de converter_aberturas_deteccao_para_interface)
                        lado = ab.get('lado', '')
                        print(f"[DEBUG add_viga_with_boundary] lado obtido da abertura: {lado!r}")
                        if not lado:
                            # Determinar lado baseado em x
                            if abs(x) < 1e-2:
                                lado = 'esquerda'
                            elif abs((x + largura_ab) - largura_total) < 1e-2:
                                lado = 'direita'
                            elif x < largura_total / 2:
                                lado = 'esquerda'
                            else:
                                lado = 'direita'
                        
                        # Topo/Fundo
                        topo_fundo = ab.get('topo_fundo', '')
                        if not topo_fundo:
                            # Determinar topo/fundo baseado em y
                            # y=0 significa fundo, y próximo de altura_total significa topo
                            limiar = max(1.0, altura_total * 0.05)
                            if y <= limiar:
                                # Está no fundo (y próximo de 0)
                                topo_fundo = 'fundo'
                            elif (y + altura_ab) >= (altura_total - limiar):
                                # Está no topo (y + altura próximo de altura_total)
                                topo_fundo = 'topo'
                            else:
                                # Está no meio - determinar pela posição relativa
                                dist_fundo = y
                                dist_topo = altura_total - (y + altura_ab)
                                if dist_fundo < dist_topo:
                                    topo_fundo = 'fundo'
                                else:
                                    topo_fundo = 'topo'
                        
                        print(f"[DEBUG add_viga_with_boundary] Abertura detectada: x={x:.2f}, y={y:.2f}, larg={largura_ab:.2f}, alt={altura_ab:.2f}, lado={lado}, topo_fundo={topo_fundo}, largura_total={largura_total:.2f}, altura_total={altura_total:.2f}")
                        
                        # Determinar se está na extremidade (distância = 0)
                        pos_x = ab.get('pos_x', '')
                        if not pos_x:
                            if abs(x) < 1e-2 or abs((x + largura_ab) - largura_total) < 1e-2:
                                pos_x = 'parede'
                            else:
                                pos_x = 'central'
                        
                        # Calcular distância (0 se na extremidade, senão distância até a extremidade)
                        if pos_x == 'parede':
                            distancia = 0
                        else:
                            # Distância até a extremidade mais próxima
                            if lado == 'esquerda':
                                distancia = int(round(x))
                            else:  # direita
                                distancia = int(round(largura_total - (x + largura_ab)))
                        
                        # Determinar índice da posição (T/E=0, F/E=1, T/D=2, F/D=3)
                        idx_pos = -1
                        if lado == 'esquerda' and topo_fundo == 'topo':
                            idx_pos = 0  # T/E
                        elif lado == 'esquerda' and topo_fundo == 'fundo':
                            idx_pos = 1  # F/E
                        elif lado == 'direita' and topo_fundo == 'topo':
                            idx_pos = 2  # T/D
                        elif lado == 'direita' and topo_fundo == 'fundo':
                            idx_pos = 3  # F/D
                        
                        # Preencher a posição se válida
                        if 0 <= idx_pos < 4:
                            aberturas_formatadas[idx_pos] = [
                                str(distancia),  # distância (0 se na extremidade)
                                str(int(round(profundidade_ab))),  # profundidade
                                str(int(round(largura_ab)))  # largura
                            ]
                            print(f"[DEBUG] Abertura mapeada: {lado} {topo_fundo} -> posição {idx_pos} (T/E=0, F/E=1, T/D=2, F/D=3): dist={distancia}, prof={profundidade_ab:.2f}, larg={largura_ab:.2f}")
                    
                    # IMPORTANTE: Garantir que aberturas_formatadas é uma lista de listas de strings
                    # e que está sendo atribuída corretamente ao item
                    aberturas_para_salvar = []
                    for ab in aberturas_formatadas:
                        if isinstance(ab, (list, tuple)):
                            aberturas_para_salvar.append([str(v) for v in ab[:3]])
                        else:
                            aberturas_para_salvar.append(["0", "0", "0"])
                    
                    # Garantir que tem 4 posições
                    while len(aberturas_para_salvar) < 4:
                        aberturas_para_salvar.append(["0", "0", "0"])
                    
                    # Atualizar no item (que é uma referência ao items_dict[number])
                    novo_item['aberturas'] = aberturas_para_salvar
                    
                    # IMPORTANTE: Garantir que items_dict[number] também está atualizado
                    if number in items_dict:
                        items_dict[number]['aberturas'] = aberturas_para_salvar.copy()
                    
                    # DEBUG: Verificar o que foi salvo
                    print(f"[DEBUG add_viga_with_boundary] aberturas_formatadas salvas no novo_item: {aberturas_para_salvar}")
                    print(f"[DEBUG add_viga_with_boundary] Tipo: {type(aberturas_para_salvar)}, Tamanho: {len(aberturas_para_salvar)} posições")
                    for idx, ab in enumerate(aberturas_para_salvar):
                        if isinstance(ab, (list, tuple)) and len(ab) >= 3:
                            if any(float_safe(v) > 0 for v in ab):
                                posicoes = ['T/E', 'F/E', 'T/D', 'F/D']
                                print(f"[DEBUG add_viga_with_boundary] ✅ {posicoes[idx]} preenchido: dist={ab[0]}, prof={ab[1]}, larg={ab[2]}")
                        else:
                            print(f"[DEBUG add_viga_with_boundary] ⚠️ Posição {idx} tem formato inválido: {ab}")
                    
                    # Verificar se está realmente no dicionário
                    print(f"[DEBUG add_viga_with_boundary] Verificando novo_item['aberturas']: {novo_item.get('aberturas')}")
                    if number in items_dict:
                        print(f"[DEBUG add_viga_with_boundary] Verificando items_dict[{number}]['aberturas']: {items_dict[number].get('aberturas')}")
                    
                    processed = True
                    self.log(f"✅ Boundary processado para viga {name}: {msg}")
                    # Contar aberturas não vazias
                    aberturas_nao_vazias = sum(1 for ab in aberturas_formatadas if any(float_safe(v) > 0 for v in ab))
                    print(f"[DEBUG] ✅ Campos preenchidos automaticamente: largura={novo_item['largura']}, altura={novo_item['altura']}, recuos={novo_item['recuos']}, aberturas={aberturas_nao_vazias} detectadas")
                else:
                    self.log(f"⚠️ Falha na análise de boundary para viga {name}")
            except Exception as e:
                self.log(f"⚠️ Erro ao processar boundary para viga {name}: {e}")
                import traceback
                traceback.print_exc()
        
        # Salvar dados automaticamente após processamento
        if processed:
            try:
                self.save_data()
                print(f"[DEBUG] ✅ Dados salvos automaticamente após processamento do boundary")
                
                # IMPORTANTE: Se o item atual está selecionado, recarregar para atualizar a UI
                selected_items = self.tree_fundos.selectedItems()
                if selected_items:
                    current_item = selected_items[0]
                    current_num = current_item.data(0, Qt.UserRole) or current_item.text(0)
                    if str(current_num) == str(number):
                        print(f"[DEBUG] ✅ Item {number} está selecionado, recarregando para atualizar campos da UI...")
                        self.load_selected(current_item)
            except Exception as e:
                print(f"[DEBUG] ⚠️ Erro ao salvar dados automaticamente: {e}")
        
        # Salvar dados (mantido para compatibilidade)
        try:
            # Verificar se aberturas foram salvas antes de salvar
            if not is_new and number in items_dict:
                item_salvo = items_dict[number]
                aberturas_salvas = item_salvo.get('aberturas', [])
                print(f"[DEBUG add_viga_with_boundary] Antes de salvar - aberturas no item: {aberturas_salvas}")
            
            # IMPORTANTE: Garantir que novo_item['aberturas'] está realmente no items_dict
            if number in items_dict:
                item_atual = items_dict[number]
                # Se processou boundary, garantir que aberturas estão no item
                if processed and 'aberturas' in novo_item:
                    item_atual['aberturas'] = novo_item['aberturas'].copy() if isinstance(novo_item['aberturas'], list) else novo_item['aberturas']
                    print(f"[DEBUG add_viga_with_boundary] ✅ Forçando atualização de aberturas no item: {item_atual['aberturas']}")
            
            self.save_data()
            
            # Verificar após salvar - ler do arquivo para confirmar
            if number in items_dict:
                item_salvo = items_dict[number]
                aberturas_salvas = item_salvo.get('aberturas', [])
                print(f"[DEBUG add_viga_with_boundary] Após salvar - aberturas no item: {aberturas_salvas}")
                
                # Verificar também nos fundos_salvos
                if isinstance(self.fundos_salvos, dict):
                    dados_salvos = self.fundos_salvos.get(obra, {}).get(pav, {}).get(str(number))
                    if dados_salvos:
                        aberturas_fundos = dados_salvos.get('aberturas', [])
                        print(f"[DEBUG add_viga_with_boundary] Aberturas em fundos_salvos: {aberturas_fundos}")
                        if aberturas_fundos != aberturas_salvas:
                            print(f"[DEBUG add_viga_with_boundary] ⚠️ INCONSISTÊNCIA: aberturas em items_dict != fundos_salvos!")
            
            if is_new:
                self.update_list()
            else:
                # Se não é novo, atualizar a lista também para garantir que está sincronizado
                self.update_list()
        except Exception as e:
            self.log(f"⚠️ Erro ao salvar dados: {e}")
        
        return {
            'success': True,
            'processed': processed,
            'message': f"Viga {name} {'criada' if is_new else 'atualizada'}" + (f" e boundary processado" if processed else "")
        }

    def action_nova_viga(self):
        """Cria nova viga (reseta sufixo, volta para base sem sufixo)"""
        self.usar_boundary_duplo = False
        try:
            num_text = self.fields['numero'].text().strip()
            numero_atual = float_safe(num_text)
            numero_base = int(numero_atual)
        except (ValueError, TypeError):
            numeros_existentes = []
            obra = self.combo_obra.currentText()
            pav = _cmb_pav_raw(self.combo_pavimento)
            if obra and pav and obra in self.fundos_salvos and pav in self.fundos_salvos[obra]:
                for numero in self.fundos_salvos[obra][pav].keys():
                    try:
                        import re
                        match = re.match(r'^[Vv]?(\d+)', str(numero))
                        if match:
                            numeros_existentes.append(int(match.group(1)))
                    except (ValueError, TypeError): continue
            numero_base = max(numeros_existentes) if numeros_existentes else 0
        
        # Encontrar próximo disponível
        numeros_existentes = []
        obra = self.combo_obra.currentText()
        pav = _cmb_pav_raw(self.combo_pavimento)
        if obra and pav and obra in self.fundos_salvos and pav in self.fundos_salvos[obra]:
            for numero in self.fundos_salvos[obra][pav].keys():
                try:
                    import re
                    match = re.match(r'^[Vv]?(\d+)', str(numero))
                    if match:
                        numeros_existentes.append(int(match.group(1)))
                except: continue
        
        proximo_numero = numero_base + 1
        while proximo_numero in numeros_existentes:
            proximo_numero += 1
            
        self.action_salvar()
        pavimento_atual = self.fields['pavimento'].text()
        self.action_novo()
        # Resetar para base sem sufixo
        self.fields['numero'].setText(str(proximo_numero))
        # Nome será gerado automaticamente ou pode ser V{numero}
        nome_base = f"V{proximo_numero}"
        self.fields['nome'].setText(nome_base)
        self.fields['pavimento'].setText(pavimento_atual)
        self._selecionar_textos_consecutivos()

    def action_continua_viga(self):
        """Cria próximo segmento sequencial (V1a -> V1b -> V1c)"""
        self.usar_boundary_duplo = False
        try:
            num_text = self.fields['numero'].text().strip()
            nome_text = self.fields['nome'].text().strip()
            
            # Usar nome se disponível, senão usar número
            base_name = nome_text if nome_text else num_text
            
            # Obter próximo sufixo sequencial
            proximo_nome = self.get_next_segment_suffix(base_name)
            
            self.action_salvar()
            pavimento_atual = self.fields['pavimento'].text()
            self.action_novo()
            
            # Atualizar nome com próximo sufixo
            self.fields['nome'].setText(proximo_nome)
            # Manter número base se for diferente do nome
            if num_text and not nome_text:
                # Extrair número base
                import re
                match = re.match(r'^([Vv]?\d+)', num_text)
                if match:
                    self.fields['numero'].setText(match.group(1))
                else:
                    self.fields['numero'].setText(num_text)
            else:
                # Extrair número do nome
                import re
                match = re.match(r'^[Vv]?(\d+)', proximo_nome)
                if match:
                    self.fields['numero'].setText(match.group(1))
            
            self.fields['pavimento'].setText(pavimento_atual)
            self._selecionar_textos_consecutivos()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro na continuação: {e}")

    def action_extensao_l(self):
        self.action_salvar()
        # Incrementar se já tiver um número
        num_text = self.fields['numero'].text().strip()
        if num_text:
            self.action_continua_viga() # Cria o ".1" ou incrementa decimal
        
        self.usar_boundary_duplo = True
        self._selecionar_textos_consecutivos()

    def action_reiniciar_sel(self):
        self._selecionar_textos_consecutivos()

    def get_continuation_mode(self):
        """Retorna o modo de continuação selecionado"""
        if self.rb_cont_obs.isChecked():
            return "Obstaculo"
        elif self.rb_cont_abertura.isChecked():
            return "Abertura"
        else:  # rb_cont_last (padrão)
            return "UltimoSegmento"
    
    def apply_continuation_rules(self, data):
        """Aplica regras de continuação aos dados"""
        continuacao = self.get_continuation_mode()
        data['continuacao'] = continuacao
        
        # Aplicar regras específicas baseadas no modo
        if continuacao == "Obstaculo":
            # Adicionar layer especial e marcar para cotas de obstáculo
            data['layer_obstaculo'] = True
            data['cota_obstaculo'] = True
        elif continuacao == "UltimoSegmento":
            # Aplicar cotas finais
            data['cota_final'] = True
        elif continuacao == "Abertura":
            # Aplicar cotas de continuação
            data['cota_continuacao'] = True
        
        return data
    
    def get_next_segment_suffix(self, current_name):
        """Retorna o próximo sufixo sequencial (a, b, c, ...) para o nome base"""
        import re
        # Extrair base (ex: "V1" de "V1a" ou "V1")
        match = re.match(r'^([Vv]?\d+[a-zA-Z]*)([a-z]?)$', str(current_name))
        if match:
            base = match.group(1)
            current_suffix = match.group(2) if match.group(2) else ''
        else:
            # Fallback: assumir que o nome inteiro é a base
            base = str(current_name)
            current_suffix = ''
        
        # Se não tem sufixo, começar com 'a'
        if not current_suffix:
            return base + 'a'
        
        # Incrementar sufixo: a->b->c->...->z->aa->ab...
        if current_suffix == 'z':
            return base + 'aa'
        else:
            next_char = chr(ord(current_suffix) + 1)
            return base + next_char

    def sync_context(self, obra_name, pav_name=None):
        """Sincroniza o robô com o contexto mestre (Master Sync)"""
        if not obra_name: return
        
        self.combo_obra.blockSignals(True)
        self.combo_pavimento.blockSignals(True)
        
        try:
            # 1. Garantir Obra
            if obra_name not in self.obras_metadata:
                self.obras_metadata[obra_name] = set()
                self.combo_obra.addItem(obra_name)
            
            self.combo_obra.setCurrentText(obra_name)
            
            # Atualizar pavimentos disponíveis para esta obra (sem disparar update_list ainda)
            pavs = sorted(list(self.obras_metadata.get(obra_name, [])))
            # No update_combo_pavimentos manual logic here to keep signals blocked
            self.combo_pavimento.clear()
            if not pavs and not pav_name:
                self.combo_pavimento.addItem("1", "1")
            elif pavs:
                for _p in pavs: self.combo_pavimento.addItem(_fmt_pav_label(_p), _p)

            # 2. Garantir Pavimento
            if pav_name:
                if pav_name not in self.obras_metadata[obra_name]:
                    self.obras_metadata[obra_name].add(pav_name)
                    # Re-adicionar agora que o novo existe
                    pavs = sorted(list(self.obras_metadata.get(obra_name, [])))
                    self.combo_pavimento.clear()
                    for _p in pavs: self.combo_pavimento.addItem(_fmt_pav_label(_p), _p)

                _set_cmb_by_data(self.combo_pavimento, pav_name)
                
                # Atualizar campo de texto legacy em Informações Básicas
                if 'pavimento' in self.fields:
                    self.fields['pavimento'].setText(pav_name)
            
            # Importar N3 antes de atualizar a lista filtrada.
            self.sync_n3_fundos_to_current_context()
            self.update_list()
            
            self.statusBar().showMessage(f"🔄 Sincronizado: {obra_name} | {pav_name}", 3000)
            print(f"2026-01-20 18:36:46,428 - INFO - 🔄 Sincronizando Robôs -> Obra: {obra_name}, Pav: {pav_name}")
            
        finally:
            self.combo_obra.blockSignals(False)
            self.combo_pavimento.blockSignals(False)

    def _selecionar_textos_consecutivos(self):
        self._flutuante_cancelado = False
        
        passos = [
            ("Selecione o NÚMERO da viga", 'numero'),
            ("Selecione o NOME da viga", 'nome'),
            ("Selecione a LARGURA da viga", 'largura'),
            ("Selecione a ALTURA da viga", 'altura'),
            ("Selecione o TEXTO ESQUERDA", 'texto_esq'),
            ("Selecione o TEXTO DIREITA", 'texto_dir'),
            ("Selecione o PAVIMENTO", 'pavimento')
        ]

        def process_step(idx):
            if idx >= len(passos) or self._flutuante_cancelado:
                if not self._flutuante_cancelado:
                    # Ao final dos textos, dispara o Boundary
                    if self.usar_boundary_duplo:
                        self._realizar_boundary_duplo()
                    else:
                        self._realizar_boundary()
                return

            msg, key = passos[idx]
            # Mostrar balão e aguardar
            res = self._mostrar_comentario_flutuante(msg)
            # Como FloatingComment agora aguarda o usuário ou timeout, precisamos disparar
            # a seleção do AutoCAD.
            
            try:
                texto, ponto = self.selection_manager.select_text(msg)
                if texto:
                    self.fields[key].setText(str(texto))
                    if key in ['largura', 'altura']:
                        self.update_calculations()
                
                if not self._flutuante_cancelado:
                    QTimer.singleShot(300, lambda: process_step(idx + 1))
            except Exception as e:
                print(f"Erro na seleção: {e}")
                QTimer.singleShot(1000, lambda: process_step(idx))

        process_step(0)

    def action_cancelar_selecao(self):
        self._flutuante_cancelado = True
        if self._balcao_flutuante:
            self._balcao_flutuante.reject()
        QMessageBox.information(self, "Cancelado", "Seleção de produção cancelada.")

    def action_desenhar_teste(self):
        print("[DEBUG] Iniciando action_desenhar_teste")

        # Primeiro salvar os dados atuais
        self.action_salvar()

        # Obter dados após salvar
        dados = self.get_current_data()
        print(f"[DEBUG] Dados obtidos: {dados}")

        try:
            pavimento = dados.get('pavimento')
            numero = dados.get('numero')
            print(f"[DEBUG] Pavimento: {pavimento}, Numero: {numero}")

            if not numero or not pavimento:
                QMessageBox.warning(self, "Erro", "Número e Pavimento são obrigatórios!")
                return

            # Obter dados do fundo atual
            dados_fundo = self.get_current_data()
            dados_fundo['obra'] = self.combo_obra.currentText()
            
            # Criar callback que sempre retorna True (sucesso)
            def salvar_callback():
                print("[DEBUG] salvar_callback chamado")
                self.action_salvar()  # Salvar novamente para gerar script
                return True

            # Usar utilitário original com dados do fundo
            print("[DEBUG] Chamando desenhar_teste_util com dados do fundo")
            desenhar_teste_util(pavimento, numero, salvar_callback, dados_fundo=dados_fundo, config_manager=self.config_manager)
            print("[DEBUG] desenhar_teste_util executado com sucesso")
            QMessageBox.information(self, "Sucesso", f"Script gerado e enviado para o AutoCAD para a viga {numero}!")
        except Exception as e:
            print(f"[DEBUG] Erro em action_desenhar_teste: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erro", f"Erro ao gerar script: {e}")

    def action_desenhar_pavimento(self):
        print("[DEBUG] Iniciando action_desenhar_pavimento")
        pav = self.fields['pavimento'].text()
        print(f"[DEBUG] Pavimento selecionado: '{pav}'")

        if not pav or pav == "Todos":
            print("[DEBUG] Pavimento inválido ou 'Todos'")
            QMessageBox.warning(self, "Aviso", "Selecione um pavimento específico!")
            return

        # Filtrar fundos corretamente da estrutura nested
        obra_atual = self.combo_obra.currentText()
        print(f"[DEBUG] Obra atual: '{obra_atual}'")
        print(f"[DEBUG] Pavimento selecionado: '{pav}'")
        print(f"[DEBUG] Estrutura fundos_salvos: {list(self.fundos_salvos.keys())}")

        fundos = []
        if obra_atual in self.fundos_salvos and pav in self.fundos_salvos[obra_atual]:
            for numero, dados in self.fundos_salvos[obra_atual][pav].items():
                fundos.append(dados)
                print(f"[DEBUG] Fundo encontrado: numero='{numero}', nome='{dados.get('nome')}'")

        print(f"[DEBUG] Fundos encontrados para o pavimento: {len(fundos)}")

        if not fundos:
            print("[DEBUG] Nenhum fundo encontrado para o pavimento")
            QMessageBox.information(self, "Info", "Nenhum fundo encontrado para este pavimento.")
            return

        try:
            print(f"[DEBUG] Chamando desenhar_pavimento_util com {len(fundos)} fundos")

            # Preparar dados no formato esperado pela função
            fundos_formatados = {}
            obra_atual = self.combo_obra.currentText()
            for dados_fundo in fundos:
                numero = dados_fundo.get('numero')
                if numero:
                    fundos_formatados[numero] = dados_fundo

            print(f"[DEBUG] Fundos formatados para pavimento {pav}: {list(fundos_formatados.keys())}")

            def salvar_pavimento_callback():
                print("[DEBUG] salvar_pavimento_callback chamado")
                return True

            # Sistema de créditos removido - geração direta com config_manager
            desenhar_pavimento_util(pav, fundos_formatados, None, salvar_pavimento_callback, config_manager=self.config_manager)
            print("[DEBUG] desenhar_pavimento_util executado com sucesso")
            QMessageBox.information(self, "Sucesso", f"Script do pavimento {pav} gerado com sucesso!")
        except Exception as e:
            print(f"[DEBUG] Erro em action_desenhar_pavimento: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erro", f"Erro ao gerar script do pavimento: {e}")

    def generate_script_current(self):
        """Gera o script para o fundo atual (equivalente a Gerar Segmento)"""
        try:
            # Salvar dados atuais
            self.action_salvar()
            
            # Obter dados
            dados = self.get_current_data()
            pavimento = dados.get('pavimento', '').strip()
            nome = dados.get('nome', dados.get('numero', 'SemNome')).strip()
            
            if not pavimento:
                QMessageBox.warning(self, "Erro", "Pavimento é obrigatório!")
                return
            
            # Usar action_desenhar_teste que já gera o script
            self.action_desenhar_teste()
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar script: {e}")
            import traceback
            traceback.print_exc()
    
    def generate_conjunto_scripts(self):
        """Gera scripts para o conjunto de segmentos sequenciais (ex: V1a, V1b, V1c)"""
        try:
            obra = self.combo_obra.currentText()
            pav = _cmb_pav_raw(self.combo_pavimento)
            
            if not obra or not pav:
                QMessageBox.warning(self, "Aviso", "Selecione uma Obra e Pavimento.")
                return
            
            # Obter fundo atual para identificar prefixo
            dados_atual = self.get_current_data()
            nome_atual = dados_atual.get('nome', '')
            
            # Extrair prefixo (ex: "V1" de "V1a")
            import re
            match = re.match(r'^([Vv]?\d+[a-zA-Z]*)', nome_atual)
            if not match:
                QMessageBox.warning(self, "Aviso", "Não foi possível identificar o conjunto. Use um nome como V1a, V1b, etc.")
                return
            
            prefixo_base = match.group(1)
            
            # Filtrar fundos do mesmo conjunto (mesmo prefixo)
            fundos_conjunto = {}
            if obra in self.fundos_salvos and pav in self.fundos_salvos[obra]:
                for numero, dados in self.fundos_salvos[obra][pav].items():
                    nome_fundo = dados.get('nome', str(numero))
                    # Verificar se começa com o mesmo prefixo
                    if re.match(rf'^{re.escape(prefixo_base)}', nome_fundo, re.IGNORECASE):
                        fundos_conjunto[numero] = dados
            
            if not fundos_conjunto:
                QMessageBox.warning(self, "Aviso", f"Nenhum fundo encontrado para o conjunto '{prefixo_base}'.")
                return
            
            # Usar função comum de geração bulk
            self._generate_bulk_scripts(fundos_conjunto, pav, f"Deseja gerar scripts para o conjunto '{prefixo_base}' ({len(fundos_conjunto)} fundos)?")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar conjunto: {e}")
            import traceback
            traceback.print_exc()
    
    def generate_pavimento_scripts(self):
        """Gera scripts para todos os fundos do pavimento atual"""
        try:
            obra = self.combo_obra.currentText()
            pav = _cmb_pav_raw(self.combo_pavimento)
            
            if not obra or not pav:
                QMessageBox.warning(self, "Aviso", "Selecione uma Obra e Pavimento.")
                return
            
            # Obter todos os fundos do pavimento
            fundos_pavimento = {}
            if obra in self.fundos_salvos and pav in self.fundos_salvos[obra]:
                fundos_pavimento = self.fundos_salvos[obra][pav].copy()
            
            if not fundos_pavimento:
                QMessageBox.warning(self, "Aviso", "Nenhum fundo encontrado no pavimento atual.")
                return
            
            # Usar função comum de geração bulk
            self._generate_bulk_scripts(fundos_pavimento, pav, f"Deseja gerar scripts para TODOS ({len(fundos_pavimento)}) os fundos do pavimento {pav}?")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar scripts do pavimento: {e}")
            import traceback
            traceback.print_exc()
    
    def _agrupar_segmentos_sequenciais(self, fundos_dict):
        """Agrupa fundos por prefixo base (ex: V1a, V1b, V1c -> grupo V1)"""
        import re
        grupos = {}
        
        for numero, dados in fundos_dict.items():
            nome = dados.get('nome', str(numero))
            # Extrair prefixo base (ex: "V1" de "V1a")
            match = re.match(r'^([Vv]?\d+[a-zA-Z]*)', nome)
            if match:
                prefixo = match.group(1)
            else:
                # Fallback: usar número como prefixo
                prefixo = str(numero)
            
            if prefixo not in grupos:
                grupos[prefixo] = []
            grupos[prefixo].append((numero, dados))
        
        # Ordenar segmentos dentro de cada grupo
        for prefixo in grupos:
            grupos[prefixo].sort(key=lambda x: x[1].get('nome', str(x[0])))
        
        return grupos
    
    def _generate_bulk_scripts(self, fundos_dict, folder_name, question_text):
        """Lógica comum para gerar múltiplos scripts em uma pasta, ordenar e combinar."""
        print(f"📊 [ROBO FUNDOS] Iniciando geração bulk - {len(fundos_dict)} fundos encontrados, Pasta: '{folder_name}'")
        
        # Confirmar geração
        confirm = QMessageBox.question(self, "Confirmar", question_text, QMessageBox.Yes | QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        
        # Preparar diretório de saída
        from pathlib import Path
        import os
        import shutil
        import re
        
        # Normalizar nome do pavimento: substituir espaços por underscore
        folder_name_normalized = folder_name.replace(' ', '_')
        
        project_root = Path(__file__).parent.parent.parent.parent
        base_dir = project_root / "SCRIPTS_ROBOS"
        out_dir = base_dir / folder_name_normalized
        
        # Limpeza da pasta
        if out_dir.exists():
            try:
                shutil.rmtree(str(out_dir))
            except Exception as e:
                print(f"Erro ao limpar pasta {out_dir}: {e}")
        
        out_dir.mkdir(parents=True, exist_ok=True)
        
        count = 0
        scripts_gerados = []
        
        # Ordenar fundos por nome (ordem natural)
        # Garantir que a comparação seja sempre entre tipos compatíveis
        def natural_sort_key(item):
            nome = str(item[1].get('nome', item[0]))
            parts = re.split(r'(\d+)', nome)
            key_list = []
            for t in parts:
                if t:  # Ignorar strings vazias
                    if t.isdigit():
                        key_list.append((0, int(t)))  # Tipo 0 = número
                    else:
                        key_list.append((1, t.lower()))  # Tipo 1 = string
            return key_list
        
        fundos_items = sorted(fundos_dict.items(), key=natural_sort_key)
        
        for numero, dados_fundo in fundos_items:
            nome_fundo = dados_fundo.get('nome', str(numero))
            pavimento_real = dados_fundo.get('pavimento', folder_name)
            # Normalizar nome do pavimento: substituir espaços por underscore
            pavimento_real_normalized = pavimento_real.replace(' ', '_')
            print(f"🔧 [ROBO FUNDOS] Processando fundo '{nome_fundo}'")
            
            try:
                # Garantir que painéis sejam strings numéricas válidas (sem .0 desnecessário)
                paineis_raw = dados_fundo.get('paineis', ['100', '100', '0', '0', '0', '0'])
                paineis_formatados = []
                for p in paineis_raw:
                    if isinstance(p, str):
                        p_stripped = p.strip()
                        if not p_stripped or p_stripped == '0' or p_stripped == '':
                            paineis_formatados.append('0')
                        elif '+' in p_stripped:
                            # Manter formato de soma como está
                            paineis_formatados.append(p_stripped)
                        else:
                            try:
                                p_float = float(p_stripped)
                                # Remover .0 se for inteiro
                                if p_float.is_integer():
                                    paineis_formatados.append(str(int(p_float)))
                                else:
                                    paineis_formatados.append(str(p_float))
                            except (ValueError, TypeError):
                                paineis_formatados.append('0')
                    else:
                        try:
                            p_float = float(p) if p is not None else 0.0
                            if p_float.is_integer():
                                paineis_formatados.append(str(int(p_float)))
                            else:
                                paineis_formatados.append(str(p_float))
                        except (ValueError, TypeError):
                            paineis_formatados.append('0')
                
                # Garantir que largura e altura sejam strings numéricas válidas
                largura_raw = dados_fundo.get('largura', '200')
                altura_raw = dados_fundo.get('altura', '20')
                
                try:
                    largura_float = float(str(largura_raw).strip() or '200')
                    largura_str = str(int(largura_float)) if largura_float.is_integer() else str(largura_float)
                except:
                    largura_str = '200'
                
                try:
                    altura_float = float(str(altura_raw).strip() or '20')
                    altura_str = str(int(altura_float)) if altura_float.is_integer() else str(altura_float)
                except:
                    altura_str = '20'
                
                # Preparar dados no formato correto
                dados_formatados = {
                    'numero': str(dados_fundo.get('numero', numero)),
                    'nome': nome_fundo,
                    'observacoes': dados_fundo.get('obs', ''),
                    'pavimento': pavimento_real_normalized,  # Usar pavimento normalizado (sem espaços)
                    'largura': largura_str,
                    'altura': altura_str,
                    'texto_esquerda': str(dados_fundo.get('texto_esq', '')),
                    'texto_direita': str(dados_fundo.get('texto_dir', '')),
                    'paineis': paineis_formatados,
                    'recuos': [str(r) for r in dados_fundo.get('recuos', ['0', '0', '0', '0'])],
                    'aberturas': [[str(a) for a in ab] for ab in dados_fundo.get('aberturas', [['0', '0', '0']] * 4)],
                    'comprimento_2': str(dados_fundo.get('comprimento_2', '0')),
                    'largura_2': str(dados_fundo.get('largura_2', '0')),
                    'paineis_2': [str(p) for p in dados_fundo.get('paineis_2', ['0', '0', '0'])],
                    'tipo_painel2': dados_fundo.get('tipo_painel2', 'E/T'),
                    'recuos_2': [str(r) for r in dados_fundo.get('recuos_2', ['0', '0', '0', '0'])],
                    'aberturas_2': [[str(a) for a in ab] for ab in dados_fundo.get('aberturas_2', [['0', '0', '0']] * 4)],
                    'sarrafo_esq': dados_fundo.get('sarrafo_esq', True),
                    'sarrafo_dir': dados_fundo.get('sarrafo_dir', True),
                    'continuacao': dados_fundo.get('continuacao', 'UltimoSegmento')
                }
                
                # Gerar script usando gerar_script_robo_util
                from fundo_utils import gerar_script_robo_util
                import time
                
                sucesso = False
                try:
                    sucesso = gerar_script_robo_util(dados_formatados, self.config_manager)
                    # Aguardar um pouco para garantir que o arquivo foi salvo
                    time.sleep(0.5)
                except Exception as e:
                    print(f"❌ Erro ao gerar script para {nome_fundo}: {e}")
                    import traceback
                    traceback.print_exc()
                    # Mesmo com erro, tentar buscar script (pode ter sido gerado parcialmente)
                    time.sleep(0.5)
                
                # SEMPRE tentar buscar script gerado, mesmo se sucesso for False
                robo_dir = Path(__file__).parent
                # Normalizar nome do pavimento para pasta (substituir espaços por underscore)
                pavimento_dir = robo_dir / "PASTAS-FUNDOS" / pavimento_real_normalized
                
                # Buscar script mais recente gerado para este fundo
                script_orig_path = None
                if pavimento_dir.exists():
                    # Normalizar nome para busca (remover caracteres especiais problemáticos)
                    nome_busca = nome_fundo.replace('.', '_').replace(' ', '_')
                    
                    # 1. Buscar por {nome_fundo}.scr (formato direto)
                    script_path_direto = pavimento_dir / f"{nome_fundo}.scr"
                    if script_path_direto.exists():
                        script_orig_path = script_path_direto
                        print(f"✅ Script encontrado (direto): {script_path_direto}")
                    else:
                        # 2. Buscar por {nome_fundo}.{numero}.scr (padrão numerado)
                        padrao_numerado = re.compile(rf"^{re.escape(nome_fundo)}\.(\d+)\.scr$")
                        arquivos_numerados = []
                        for f in pavimento_dir.glob("*.scr"):
                            m = padrao_numerado.match(f.name)
                            if m:
                                arquivos_numerados.append((int(m.group(1)), f))
                        
                        if arquivos_numerados:
                            arquivos_numerados.sort()
                            script_orig_path = arquivos_numerados[-1][1]
                            print(f"✅ Script encontrado (numerado): {script_orig_path}")
                        else:
                            # 3. Buscar por padrão que começa com nome_fundo (inclui V1.b.scr, V1.5.scr, etc)
                            # Extrair base do nome (ex: "V1" de "V1.b")
                            nome_base = re.match(r'^([Vv]?\d+)', nome_fundo)
                            if nome_base:
                                base_match = nome_base.group(1)
                                padrao_prefixo = re.compile(rf"^{re.escape(base_match)}[\.\-_].*\.scr$", re.IGNORECASE)
                            else:
                                padrao_prefixo = re.compile(rf"^{re.escape(nome_fundo)}[\.\-_].*\.scr$", re.IGNORECASE)
                            
                            arquivos_prefixo = []
                            for f in pavimento_dir.glob("*.scr"):
                                if padrao_prefixo.match(f.name):
                                    try:
                                        tempo_modificacao = f.stat().st_mtime
                                        arquivos_prefixo.append((tempo_modificacao, f))
                                    except:
                                        pass
                            
                            if arquivos_prefixo:
                                arquivos_prefixo.sort(reverse=True)  # Mais recente primeiro
                                script_orig_path = arquivos_prefixo[0][1]
                                print(f"✅ Script encontrado (prefixo): {script_orig_path}")
                            else:
                                # 4. Buscar qualquer arquivo .scr recente (últimos 15 segundos) que contenha o nome
                                tempo_atual = time.time()
                                arquivos_recentes = []
                                for f in pavimento_dir.glob("*.scr"):
                                    try:
                                        tempo_modificacao = f.stat().st_mtime
                                        if tempo_atual - tempo_modificacao < 15:
                                            # Verificar se o nome do arquivo contém partes do nome do fundo
                                            nome_limpo = nome_fundo.replace('.', '').replace(' ', '').lower()
                                            fname_limpo = f.name.replace('.scr', '').replace('.', '').replace(' ', '').lower()
                                            if nome_limpo in fname_limpo or str(numero) in f.name:
                                                arquivos_recentes.append((tempo_modificacao, f))
                                    except Exception as e:
                                        print(f"[DEBUG] Erro ao verificar arquivo {f.name}: {e}")
                                        pass
                                
                                if arquivos_recentes:
                                    arquivos_recentes.sort(reverse=True)
                                    script_orig_path = arquivos_recentes[0][1]
                                    print(f"✅ Script encontrado (recente): {script_orig_path}")
                                else:
                                    # 5. Última tentativa: buscar o arquivo .scr mais recente na pasta
                                    todos_arquivos = []
                                    for f in pavimento_dir.glob("*.scr"):
                                        try:
                                            tempo_modificacao = f.stat().st_mtime
                                            todos_arquivos.append((tempo_modificacao, f))
                                        except:
                                            pass
                                    
                                    if todos_arquivos:
                                        todos_arquivos.sort(reverse=True)
                                        script_orig_path = todos_arquivos[0][1]
                                        print(f"⚠️ Usando script mais recente da pasta: {script_orig_path}")
                
                # Primeiro, salvar script individual em SCRIPTS_ROBOS/{pavimento}/
                if script_orig_path and script_orig_path.exists():
                    # Criar pasta do pavimento em SCRIPTS_ROBOS se não existir
                    scripts_individual_dir = base_dir / pavimento_real_normalized
                    scripts_individual_dir.mkdir(parents=True, exist_ok=True)
                    
                    # Sanitizar nome do arquivo (remover caracteres problemáticos)
                    nome_arquivo = nome_fundo.replace('/', '_').replace('\\', '_').replace(':', '_').replace(' ', '_')
                    
                    # Salvar script individual em SCRIPTS_ROBOS/{pavimento}/
                    script_individual_path = scripts_individual_dir / f"{nome_arquivo}.scr"
                    try:
                        shutil.copy2(str(script_orig_path), str(script_individual_path))
                        print(f"✅ [{count+1}/{len(fundos_items)}] Script individual salvo: {script_individual_path}")
                    except Exception as e:
                        print(f"⚠️ Erro ao salvar script individual: {e}")
                        import traceback
                        traceback.print_exc()
                    
                    # Depois, copiar para pasta de saída (para ordenação e combinação)
                    script_dest_path = out_dir / f"{nome_arquivo}.scr"
                    try:
                        shutil.copy2(str(script_orig_path), str(script_dest_path))
                        scripts_gerados.append(nome_fundo)
                        count += 1
                        print(f"✅ [{count}/{len(fundos_items)}] Script copiado para pasta de saída: {script_dest_path}")
                    except Exception as e:
                        print(f"❌ Erro ao copiar script para {script_dest_path}: {e}")
                        import traceback
                        traceback.print_exc()
                else:
                    print(f"⚠️ [{count+1}/{len(fundos_items)}] Script não encontrado para fundo {nome_fundo} em {pavimento_dir}")
                # Listar arquivos disponíveis para debug se script não foi encontrado
                if not (script_orig_path and script_orig_path.exists()):
                    if pavimento_dir.exists():
                        arquivos_disponiveis = list(pavimento_dir.glob("*.scr"))
                        if arquivos_disponiveis:
                            print(f"   Arquivos .scr disponíveis na pasta: {[f.name for f in arquivos_disponiveis[:10]]}")
                        else:
                            print(f"   Nenhum arquivo .scr encontrado na pasta {pavimento_dir}")
                    
            except Exception as e:
                print(f"❌ Erro ao processar fundo {nome_fundo}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        # Executar ordenador e combinador
        pasta_ordenar = str(out_dir)
        try:
            # 1. Ordenar scripts
            ordenador_path = os.path.join(os.path.dirname(__file__), "Ordenador Cordenadas Fundos.py")
            if os.path.exists(ordenador_path):
                import importlib.util
                spec = importlib.util.spec_from_file_location("ordenador", ordenador_path)
                ordenador = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(ordenador)
                
                if hasattr(ordenador, 'processar_pasta_automatico'):
                    ordenador.processar_pasta_automatico(pasta_ordenar)
                    print(f"[INFO] Ordenador executado em {pasta_ordenar}")
            
            # 2. Combinar scripts
            combinador_path = os.path.join(os.path.dirname(__file__), "Combinador_de_SCR_Fundos.py")
            if os.path.exists(combinador_path):
                import importlib.util
                spec = importlib.util.spec_from_file_location("combinador", combinador_path)
                combinador = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(combinador)
                
                if hasattr(combinador, 'processar_arquivos'):
                    combinador.processar_arquivos(pasta_ordenar)
                    print(f"[INFO] Combinador executado em {pasta_ordenar}")
                    
                    # 3. Atualizar FD.scr com o primeiro arquivo da pasta Combinados
                    pasta_combinados = out_dir / "Combinados"
                    if pasta_combinados.exists():
                        # Encontrar o primeiro arquivo .scr na pasta Combinados (ordenado numericamente)
                        arquivos_scr = sorted(
                            [f for f in pasta_combinados.glob("*.scr")],
                            key=lambda x: int(x.stem) if x.stem.isdigit() else float('inf')
                        )
                        
                        if arquivos_scr:
                            primeiro_arquivo = arquivos_scr[0]
                            print(f"[INFO] Atualizando FD.scr com conteúdo de: {primeiro_arquivo.name}")
                            
                            try:
                                # Ler conteúdo do primeiro arquivo combinado de forma robusta (bytes -> texto)
                                # Motivo: muitos .scr combinados saem como UTF-16LE, o que aparece como "NUL" (\x00)
                                from pathlib import Path
                                
                                raw = Path(primeiro_arquivo).read_bytes()
                                
                                def _looks_like_utf16le(data: bytes) -> bool:
                                    # BOM UTF-16LE
                                    if data.startswith(b"\xff\xfe"):
                                        return True
                                    if not data:
                                        return False
                                    sample = data[: min(400, len(data))]
                                    # UTF-16LE costuma ter muitos bytes 0 em texto ASCII
                                    return (sample.count(0) / len(sample)) > 0.15
                                
                                encoding_usado = None
                                if _looks_like_utf16le(raw):
                                    # Preferir utf-16 (respeita BOM se existir); fallback explícito
                                    try:
                                        conteudo_script = raw.decode("utf-16", errors="strict")
                                        encoding_usado = "utf-16"
                                    except UnicodeDecodeError:
                                        conteudo_script = raw.decode("utf-16-le", errors="replace")
                                        encoding_usado = "utf-16-le"
                                else:
                                    # Tentar UTF-8, com fallback
                                    try:
                                        conteudo_script = raw.decode("utf-8", errors="strict")
                                        encoding_usado = "utf-8"
                                    except UnicodeDecodeError:
                                        conteudo_script = raw.decode("latin-1", errors="replace")
                                        encoding_usado = "latin-1"
                                
                                # Remover NULs no texto (e qualquer resquício de bytes 0 "virando" \x00)
                                if "\x00" in conteudo_script:
                                    conteudo_script = conteudo_script.replace("\x00", "")
                                
                                # Normalizar quebras de linha (evita \r\n misturado com \n)
                                conteudo_script = conteudo_script.replace("\r\n", "\n").replace("\r", "\n")
                                
                                print(f"[INFO] Arquivo lido com encoding: {encoding_usado} (modo bytes->texto)")
                                
                                # Atualizar FD.scr na pasta ferramentas_LOAD_LISP
                                lisp_dir = project_root / "ferramentas_LOAD_LISP"
                                lisp_dir.mkdir(exist_ok=True)
                                fd_scr_path = lisp_dir / "FD.scr"
                                
                                # Processar o conteúdo:
                                # - remover linha "ferramentas_LOAD_LISP"
                                # - ajustar linhas vazias ENTRE ';' e comandos iniciados por '_' (ex: _TEXT, _SCRIPT)
                                linhas = conteudo_script.split('\n')
                                linhas_filtradas = []
                                
                                for i, linha in enumerate(linhas):
                                    linha_stripped = linha.strip()
                                    
                                    # Remover linha "ferramentas_LOAD_LISP" se existir
                                    if linha_stripped.lower() == "ferramentas_load_lisp":
                                        continue
                                    
                                    # Se a linha está vazia, verificar se está entre ; e comando que começa com _
                                    if not linha_stripped:
                                        # Verificar se a linha anterior termina com ; e a próxima começa com _
                                        linha_anterior_termina_com_ponto_virgula = False
                                        proxima_linha_comando_underscore = False
                                        
                                        if i > 0:
                                            anterior_stripped = linhas[i - 1].strip()
                                            if anterior_stripped == ';' or anterior_stripped.endswith(';'):
                                                linha_anterior_termina_com_ponto_virgula = True
                                        
                                        if i + 1 < len(linhas):
                                            prox_stripped = linhas[i + 1].strip()
                                            if prox_stripped.startswith('_') and not prox_stripped.startswith('_SCRIPT'):
                                                proxima_linha_comando_underscore = True
                                        
                                        # Se está entre ; e comando _, remover a linha vazia
                                        if linha_anterior_termina_com_ponto_virgula and proxima_linha_comando_underscore:
                                            continue
                                    
                                    linhas_filtradas.append(linha)
                                
                                # Remover linhas vazias no final
                                while linhas_filtradas and not linhas_filtradas[-1].strip():
                                    linhas_filtradas.pop()
                                
                                conteudo_final = '\n'.join(linhas_filtradas)
                                
                                # Garantir que o conteúdo final está em UTF-8 válido
                                try:
                                    conteudo_final.encode('utf-8')
                                except UnicodeEncodeError:
                                    # Se houver caracteres problemáticos, substituir
                                    conteudo_final = conteudo_final.encode('utf-8', errors='replace').decode('utf-8')
                                
                                # Adicionar cabeçalho com timestamp
                                from datetime import datetime
                                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                cabecalho = f";; Script FD - Fundo Desenhar\n;; Atualizado automaticamente em {timestamp}\n;; Este script é atualizado quando scripts de teste ou pavimento são gerados\n"
                                
                                # Adicionar cabeçalho no início do conteúdo
                                conteudo_final = cabecalho + conteudo_final
                                
                                # Escrever no FD.scr em UTF-8 (binário) e validar que não sobrou nenhum byte NUL
                                fd_bytes = conteudo_final.encode("utf-8", errors="replace")
                                if b"\x00" in fd_bytes:
                                    # Fail-safe extremo: remover qualquer byte 0 (não deveria ocorrer)
                                    fd_bytes = fd_bytes.replace(b"\x00", b"")
                                Path(fd_scr_path).write_bytes(fd_bytes)
                                
                                print(f"[INFO] ✅ FD.scr atualizado com sucesso: {fd_scr_path}")
                            except Exception as e:
                                print(f"[AVISO] Erro ao atualizar FD.scr: {e}")
                                import traceback
                                traceback.print_exc()
                        else:
                            print(f"[AVISO] Nenhum arquivo .scr encontrado na pasta Combinados: {pasta_combinados}")
                    else:
                        print(f"[AVISO] Pasta Combinados não encontrada: {pasta_combinados}")
                    
        except Exception as e:
            print(f"[AVISO] Erro ao executar ordenador/combinador: {e}")
            import traceback
            traceback.print_exc()
        
        if count > 0:
            QMessageBox.information(
                self, 
                "Sucesso", 
                f"{count} de {len(fundos_dict)} scripts gerados com sucesso!\n\n"
                f"Scripts individuais salvos em: {base_dir / folder_name_normalized}\n"
                f"Pasta de processamento: {out_dir}\n\n"
                f"O ordenador e combinador foram executados.\n"
                f"Arquivos combinados estão em: {out_dir / 'Combinados'}"
            )
        else:
            QMessageBox.warning(
                self,
                "Aviso",
                f"Nenhum script foi gerado!\n\n"
                f"Verifique se os fundos têm dados válidos e se o Robo está funcionando corretamente."
            )

    def action_criar_lisp_fd(self):
        """Cria comando lisp FD e script FD na pasta SCRIPTS_ROBOS"""
        print("[DEBUG] Iniciando action_criar_lisp_fd")

        try:
            from pathlib import Path
            import datetime

            # Obter diretórios
            project_root = Path(__file__).parent.parent.parent.parent
            lisp_dir = project_root / "ferramentas_LOAD_LISP"
            scripts_dir = project_root / "SCRIPTS_ROBOS"

            # Criar diretórios se não existirem
            lisp_dir.mkdir(exist_ok=True)
            scripts_dir.mkdir(exist_ok=True)

            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            # 1. Criar script FD na pasta ferramentas_LOAD_LISP
            script_fd_path = lisp_dir / "FD.scr"
            
            # Gerar script FD com conteúdo útil
            script_content = f""";; Script FD - Fundo Desenhar
;; Gerado em {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
;; Comando para desenhar fundos de vigas

ferramentas_LOAD_LISP

;; Mensagem de inicio
(princ "\\n=== SISTEMA DE FUNDOS DE VIGAS ===\\n")
(princ "Comando FD executado com sucesso!\\n")
(princ "Use este comando para desenhar fundos de vigas.\\n")

;; Fim do script
(princ "\\nScript FD concluido.\\n")
"""

            with open(script_fd_path, 'w', encoding='utf-8') as f:
                f.write(script_content)

            print(f"[DEBUG] Script FD criado: {script_fd_path}")

            # 2. Criar comando lisp FD (apontando para ferramentas_LOAD_LISP)
            lisp_fd_path = lisp_dir / "FD.LSP"
            script_path_str = str(script_fd_path).replace('\\', '/')  # Converter para formato Unix para lisp

            lisp_content = f""";; Comando FD - Fundo Desenhar
;; Carrega e executa o script FD
;; Gerado em {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

(defun c:FD ()
  (princ "\\nExecutando comando FD - Sistema de Fundos...\\n")

  ;; Carregar script FD
  (command "_.SCRIPT" "{script_path_str}")

  (princ "\\nComando FD concluido.\\n")
  (princ)
)
"""

            with open(lisp_fd_path, 'w', encoding='utf-8') as f:
                f.write(lisp_content)

            print(f"[DEBUG] Lisp FD criado: {lisp_fd_path}")
            print(f"[DEBUG] Tentando exibir mensagem de sucesso...")

            # Criar cópia em SCRIPTS_ROBOS como backup
            script_fd_backup = scripts_dir / "FD.scr"
            try:
                with open(script_fd_backup, 'w', encoding='utf-8') as f:
                    f.write(script_content)
                print(f"[DEBUG] Backup FD.scr criado em: {script_fd_backup}")
            except Exception as e:
                print(f"[DEBUG] Erro ao criar backup: {e}")

            # Garantir que a mensagem seja exibida
            try:
                mensagem = (
                    f"Comando lisp FD criado com sucesso!\n\n"
                    f"Lisp: {lisp_fd_path}\n"
                    f"Script: {script_fd_path}\n"
                    f"Backup: {script_fd_backup}\n\n"
                    "Use o comando 'FD' no AutoCAD para executar."
                )
                print(f"[DEBUG] Mensagem preparada: {mensagem[:50]}...")
                
                QMessageBox.information(
                    self, 
                    "Sucesso", 
                    mensagem
                )
                print("[DEBUG] Mensagem de sucesso exibida")
                
            except Exception as msg_error:
                print(f"[DEBUG] Erro ao exibir mensagem: {msg_error}")
                import traceback
                traceback.print_exc()
                # Tentar exibir mensagem simples como fallback
                try:
                    QMessageBox.information(self, "Sucesso", "Comando lisp FD criado com sucesso!")
                except:
                    print("[DEBUG] Falha ao exibir mensagem de fallback")

        except Exception as e:
            print(f"[DEBUG] Erro em action_criar_lisp_fd: {e}")
            import traceback
            traceback.print_exc()
            try:
                QMessageBox.critical(self, "Erro", f"Erro ao criar comando lisp FD: {e}")
            except Exception as msg_error:
                print(f"[DEBUG] Erro ao exibir mensagem de erro: {msg_error}")

    def action_criar_excel(self):
        if not self.fundos_salvos:
            QMessageBox.warning(self, "Aviso", "Não há fundos salvos para criar o Excel!")
            return
            
        arquivo, _ = QFileDialog.getSaveFileName(
            self, "Criar Novo Arquivo Excel", 
            os.path.expanduser(f"~/Documents/fundos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx)"
        )
        if not arquivo: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            coluna = 5
            for numero, dados in self.fundos_salvos.items():
                self._escrever_coluna_excel(ws, coluna, dados)
                coluna += 1
            wb.save(arquivo)
            QMessageBox.information(self, "Sucesso", f"Excel criado em {arquivo}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao criar Excel: {e}")

    def action_excel_export(self):
        if not self.fundos_salvos:
            QMessageBox.warning(self, "Aviso", "Não há fundos salvos para exportar!")
            return
            
        arquivo, _ = QFileDialog.getSaveFileName(self, "Exportar para Excel", "", "Excel Files (*.xlsx)")
        if not arquivo: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            coluna = 5
            for numero, dados in self.fundos_salvos.items():
                self._escrever_coluna_excel(ws, coluna, dados)
                coluna += 1
            wb.save(arquivo)
            QMessageBox.information(self, "Sucesso", f"Excel exportado para {arquivo}")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {e}")

    def action_excel_import(self):
        arquivo, _ = QFileDialog.getOpenFileName(self, "Importar Excel", "", "Excel Files (*.xlsx *.xls)")
        if not arquivo: return
        
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active
            
            itens_importados = 0
            coluna = 5
            novos_fundos = {}
            
            while True:
                numero = ws.cell(row=EXCEL_MAPPING['numero'], column=coluna).value
                if not numero: break
                
                # Reconstruir dados do fundo
                dados = {}
                for campo, linha in EXCEL_MAPPING.items():
                    val = ws.cell(row=linha, column=coluna).value
                    dados[campo] = val if val is not None else ""
                
                fundo_num = str(dados['numero']).strip()
                if not fundo_num: break
                
                # Converter estrutura chapada do Excel de volta para o dict do sistema
                fundo_dict = self._converter_excel_para_dict(dados)
                novos_fundos[fundo_num] = fundo_dict
                itens_importados += 1
                coluna += 1
                
            if novos_fundos:
                self.fundos_salvos.update(novos_fundos)
                self.save_data()
                self.update_list()
                QMessageBox.information(self, "Sucesso", f"Importação concluída! {itens_importados} fundos importados.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro na importação: {e}")

    def _escrever_coluna_excel(self, ws, col, dados):
        # Mapeia os dados do dict para as linhas do Excel
        for campo, linha in EXCEL_MAPPING.items():
            valor = ""
            if campo == 'numero': valor = dados.get('numero', '')
            elif campo == 'nome': valor = dados.get('nome', '')
            elif campo == 'observacoes': valor = dados.get('obs', '')
            elif campo == 'pavimento': valor = dados.get('pavimento', '')
            elif campo == 'largura': valor = float_safe(dados.get('largura', 0))
            elif campo == 'altura': valor = float_safe(dados.get('altura', 0))
            elif campo == 'texto_esquerda': valor = dados.get('texto_esq', '')
            elif campo == 'texto_direita': valor = dados.get('texto_dir', '')
            elif campo == 'tipo_distribuicao': valor = dados.get('distribuicao', '122')
            elif campo == 'tipo_painel_inicial': valor = "P1"
            elif campo == 'sarrafo_esq': valor = "S" if dados.get('sarrafo_esq') else "N"
            elif campo == 'sarrafo_dir': valor = "S" if dados.get('sarrafo_dir') else "N"
            elif campo.startswith('painel'):
                idx = int(campo.replace('painel', '')) - 1
                paineis = dados.get('paineis', ["0"]*6)
                valor = paineis[idx] if idx < len(paineis) else "0"
            elif campo.startswith('chanfro_'):
                idx = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.replace('chanfro_', '')]
                recuos = dados.get('recuos', ["0"]*4)
                valor = recuos[idx] if idx < len(recuos) else "0"
            elif campo.startswith('abertura_'):
                partes = campo.split('_') # abertura_te_dist
                idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[partes[1]]
                idx_tipo = {'dist':0, 'prof':1, 'larg':2}[partes[2]]
                aberturas = dados.get('aberturas', [["0"]*3 for _ in range(4)])
                valor = aberturas[idx_lado][idx_tipo]
            elif campo == 'tipo_painel2': valor = dados.get('tipo_painel2', 'E/T')
            elif campo == 'comprimento_2': valor = float_safe(dados.get('comprimento_2', 0))
            elif campo == 'largura_2': valor = float_safe(dados.get('largura_2', 0))
            elif campo.startswith('p') and '_2' in campo:
                idx = int(campo.replace('p', '').replace('_2', '')) - 1
                paineis2 = dados.get('paineis_2', ["0"]*3)
                valor = paineis2[idx] if idx < len(paineis2) else "0"
            elif campo.startswith('chanfro2_'):
                idx = {'te':0, 'fe':1, 'td':2, 'fd':3}[campo.replace('chanfro2_', '')]
                recuos2 = dados.get('recuos_2', ["0"]*4)
                valor = recuos2[idx] if idx < len(recuos2) else "0"
            elif campo.startswith('abertura2_'):
                partes = campo.split('_')
                idx_lado = {'te':0, 'fe':1, 'td':2, 'fd':3}[partes[1].replace('2', '')]
                idx_tipo = {'dist':0, 'prof':1, 'larg':2}[partes[2]]
                aberturas2 = dados.get('aberturas_2', [["0"]*3 for _ in range(4)])
                valor = aberturas2[idx_lado][idx_tipo]
            
            ws.cell(row=linha, column=col, value=valor)

    def _converter_excel_para_dict(self, dados_chapados):
        # Converte o mapeamento plano do Excel para a estrutura aninhada do dict
        paineis = [dados_chapados.get(f'painel{i}', "0") for i in range(1, 7)]
        recuos = [dados_chapados.get(f'chanfro_{l}', "0") for l in ['te', 'fe', 'td', 'fd']]
        aberturas = []
        for l in ['te', 'fe', 'td', 'fd']:
            aberturas.append([
                dados_chapados.get(f'abertura_{l}_dist', "0"),
                dados_chapados.get(f'abertura_{l}_prof', "0"),
                dados_chapados.get(f'abertura_{l}_larg', "0")
            ])
            
        paineis2 = [dados_chapados.get(f'p{i}_2', "0") for i in range(1, 4)]
        recuos2 = [dados_chapados.get(f'chanfro2_{l}', "0") for l in ['te', 'fe', 'td', 'fd']]
        aberturas2 = []
        for l in ['te', 'fe', 'td', 'fd']:
            aberturas2.append([
                dados_chapados.get(f'abertura2_{l}_dist', "0"),
                dados_chapados.get(f'abertura2_{l}_prof', "0"),
                dados_chapados.get(f'abertura2_{l}_larg', "0")
            ])

        return {
            'numero': str(dados_chapados.get('numero', '')),
            'nome': dados_chapados.get('nome', ''),
            'obs': dados_chapados.get('observacoes', ''),
            'pavimento': dados_chapados.get('pavimento', ''),
            'largura': str(dados_chapados.get('largura', '0')),
            'altura': str(dados_chapados.get('altura', '0')),
            'texto_esq': dados_chapados.get('texto_esquerda', ''),
            'texto_dir': dados_chapados.get('texto_direita', ''),
            'paineis': paineis,
            'recuos': recuos,
            'aberturas': aberturas,
            'comprimento_2': str(dados_chapados.get('comprimento_2', '0')),
            'largura_2': str(dados_chapados.get('largura_2', '0')),
            'paineis_2': paineis2,
            'tipo_painel2': dados_chapados.get('tipo_painel2', 'E/T'),
            'recuos_2': recuos2,
            'aberturas_2': aberturas2,
            'sarrafo_esq': dados_chapados.get('sarrafo_esq') == 'S',
            'sarrafo_dir': dados_chapados.get('sarrafo_dir') == 'S'
        }

    def open_config(self):
        diag = ConfigDialog(self.config_manager, self)
        diag.exec()

    def update_list(self):
        """Atualiza a lista de fundos filtrando por obra e pavimento atuais"""
        self.tree_fundos.clear()
        total_m2 = 0
        
        obra_atual = self.combo_obra.currentText()
        pav_atual = _cmb_pav_raw(self.combo_pavimento)
        
        if obra_atual and pav_atual and obra_atual in self.fundos_salvos and pav_atual in self.fundos_salvos[obra_atual]:
            items_dict = self.fundos_salvos[obra_atual][pav_atual]
            for num, dados in items_dict.items():
                m2 = float_safe(dados.get('m2', 0))
                total_m2 += m2
                item = QTreeWidgetItem([str(num), dados.get('nome', ''), dados.get('pavimento', ''), f"{m2:.4f}"])
                # Armazenar número e dados no item para facilitar exclusão
                item.setData(0, Qt.UserRole, num)
                item.setData(0, Qt.UserRole + 1, obra_atual)
                item.setData(0, Qt.UserRole + 2, pav_atual)
                item.setData(0, Qt.UserRole + 3, -1)  # -1 = viga inteira
                self.tree_fundos.addTopLevelItem(item)
                # Filhos: um por segmento em segments_rich
                segs = dados.get('segments_rich', [])
                for seg_i, seg in enumerate(segs):
                    sw = float_safe(seg.get('total_width', seg.get('width', 0)))
                    np_seg = len(seg.get('panels', []))
                    seg_item = QTreeWidgetItem([
                        f"  S{seg_i + 1}",
                        f"{dados.get('nome', 'V' + str(num))} Seg {seg_i + 1}",
                        '',
                        f"{sw:.1f}cm / {np_seg}P",
                    ])
                    seg_item.setData(0, Qt.UserRole,     num)
                    seg_item.setData(0, Qt.UserRole + 1, obra_atual)
                    seg_item.setData(0, Qt.UserRole + 2, pav_atual)
                    seg_item.setData(0, Qt.UserRole + 3, seg_i)
                    item.addChild(seg_item)
                if segs:
                    item.setExpanded(True)
        
        self.label_total_m2.setText(f"Total M²: {total_m2:.2f}")

    def get_current_data(self):
        """Coleta todos os dados atuais da interface para um dicionário."""
        paineis = [f.text() for f in self.paineis_fields]
        recuos = [f.text() for f in self.chanfros_fields]
        aberturas = []
        for row in self.aberturas_fields:
            aberturas.append([f.text() for f in row])

        l_type = "E/T"
        for rad in self.l_type_rads:
            if rad.isChecked():
                l_type = rad.text()
                break

        paineis2 = [self.l_fields['p1_2'].text(), self.l_fields['p2_2'].text(), self.l_fields['p3_2'].text()]
        recuos2 = [f.text() for f in self.l_chanfros_fields]
        aberturas2 = []
        for row in self.l_aberturas_fields:
            aberturas2.append([f.text() for f in row])

        return {
            'numero': self.fields['numero'].text(),
            'nome': self.fields['nome'].text(),
            'obs': self.fields['obs'].text(),
            'pavimento': self.fields['pavimento'].text(),
            'largura': self.fields['largura'].text(),
            'altura': self.fields['altura'].text(),
            'texto_esq': self.fields['texto_esq'].text(),
            'texto_dir': self.fields['texto_dir'].text(),
            'distribuicao': "122" if self.rad_122.isChecked() else "307",
            'paineis': paineis,
            'recuos': recuos,
            'aberturas': aberturas,
            'tipo_painel2': l_type,
            'comprimento_2': self.l_fields['comp2'].text(),
            'largura_2': self.l_fields['larg2'].text(),
            'paineis_2': paineis2,
            'recuos_2': recuos2,
            'aberturas_2': aberturas2,
            'sarrafo_esq': self.chk_sarrafo_esq.isChecked(),
            'sarrafo_dir': self.chk_sarrafo_dir.isChecked(),
            'continuacao': self.get_continuation_mode()
        }

    def action_novo(self):
        for f in self.fields.values(): f.clear()
        for f in self.paineis_fields: f.setText("0")
        for f in self.chanfros_fields: f.setText("0")
        for row in self.aberturas_fields:
            for f in row: f.setText("0")
        # Resetar modo de continuação para padrão
        self.rb_cont_last.setChecked(True)
        self.update_canvas()

    def add_global_pavimento(self, obra_name, pav_name):
        """Pre-fills pavement field from global sync."""
        if not pav_name: return
        
        # Fundo does not have strict hierarchy, but we can suggest/fill the field
        # if the user is in 'New' mode (fields empty)
        if hasattr(self, 'fields') and 'pavimento' in self.fields:
             if not self.fields['pavimento'].text():
                 self.fields['pavimento'].setText(pav_name)
                 print(f"Global Sync: Fundo pav field auto-filled with '{pav_name}'")

    def action_salvar(self):
        print("[DEBUG] Iniciando action_salvar")
        num = self.fields['numero'].text().strip()
        print(f"[DEBUG] Numero a salvar: '{num}'")
        if not num:
            QMessageBox.warning(self, "Erro", "Número do fundo é obrigatório!")
            return

        obra = self.combo_obra.currentText()
        pav = self.fields['pavimento'].text().strip() # Usar o pavimento do dado, não necessariamente do combo
        print(f"[DEBUG] Obra: '{obra}', Pavimento: '{pav}'")
        if not obra or not pav:
            QMessageBox.warning(self, "Erro", "Obra e Pavimento são necessários para salvar!")
            return

        dados = self.get_current_data()
        dados['numero'] = num
        dados['obra'] = obra
        dados['pavimento'] = pav
        print(f"[DEBUG] Dados a salvar: numero='{dados['numero']}', obra='{dados['obra']}', pavimento='{dados['pavimento']}'")
        
        # Aplicar regras de continuação
        dados = self.apply_continuation_rules(dados)
        
        # Calcular M2 antes de salvar
        dados['m2'] = calcular_area_util(
            float_safe(dados['largura']), 
            float_safe(dados['altura']),
            float_safe(dados['largura_2']),
            float_safe(dados['comprimento_2'])
        )
        
        # Garantir estrutura
        if obra not in self.fundos_salvos: self.fundos_salvos[obra] = {}
        if pav not in self.fundos_salvos[obra]: self.fundos_salvos[obra][pav] = {}

        seg_idx = getattr(self, '_current_seg_idx', -1)
        if seg_idx is None:
            seg_idx = -1

        if seg_idx >= 0:
            # Salvar apenas o segmento específico dentro de segments_rich
            viga_dados = self.fundos_salvos[obra][pav].get(str(num), {})
            segs = viga_dados.setdefault('segments_rich', [])
            while len(segs) <= seg_idx:
                segs.append({})
            paineis_vals = [float_safe(p) for p in dados.get('paineis', [])]
            segs[seg_idx]['total_width'] = float_safe(dados.get('largura', 0))
            segs[seg_idx]['panels'] = [
                {'width': w} for w in paineis_vals if w > 0
            ]
            self.fundos_salvos[obra][pav][str(num)] = viga_dados
            self.save_data()
            self.update_list()
            QMessageBox.information(self, "Sucesso", f"Segmento {seg_idx + 1} da viga {num} salvo!")
        else:
            self.fundos_salvos[obra][pav][num] = dados
            self.save_data()
            self.update_list()
            QMessageBox.information(self, "Sucesso", f"Fundo {num} salvo em {obra}/{pav}!")

    def action_analisar_boundary(self):
        """
        Função manual de análise de boundary com prints detalhados para debug.
        Analisa o boundary do item selecionado e gera logs completos de todo o processo.
        """
        print("\n" + ("=" * 100))
        print("[ANÁLISE BOUNDARY] ===== INÍCIO DA ANÁLISE MANUAL =====")
        print("=" * 100)
        
        # 1. Verificar se há item selecionado
        selected_items = self.tree_fundos.selectedItems()
        if not selected_items:
            print("[ANÁLISE BOUNDARY] ⚠️ ERRO: Nenhum item selecionado na lista!")
            QMessageBox.warning(self, "Aviso", "Selecione um item na lista para analisar o boundary.")
            print("=" * 100 + "\n")
            return
        
        item = selected_items[0]
        num = item.data(0, Qt.UserRole)
        obra = item.data(0, Qt.UserRole + 1)
        pav = item.data(0, Qt.UserRole + 2)
        
        if not num:
            num = item.text(0)
        if not obra:
            obra = self.combo_obra.currentText()
        if not pav:
            pav = item.text(2)
        
        print(f"[ANÁLISE BOUNDARY] Item selecionado: número={num!r}, obra={obra!r}, pavimento={pav!r}")
        
        # 2. Obter boundary_original
        boundary_original = None
        
        # Tentar obter do item selecionado (dados salvos)
        if isinstance(self.fundos_salvos, dict):
            dados = self.fundos_salvos.get(obra, {}).get(pav, {}).get(str(num))
            if dados and 'boundary_original' in dados:
                boundary_original = dados.get('boundary_original')
                print(f"[ANÁLISE BOUNDARY] ✅ Boundary encontrado nos dados salvos: {len(boundary_original)} coordenadas")
        
        # Fallback: tentar do _current_boundary_original (carregado no canvas)
        if not boundary_original and hasattr(self, '_current_boundary_original') and self._current_boundary_original:
            boundary_original = self._current_boundary_original
            print(f"[ANÁLISE BOUNDARY] ✅ Boundary encontrado em _current_boundary_original: {len(boundary_original)} coordenadas")
        
        if not boundary_original or len(boundary_original) < 6:
            print(f"[ANÁLISE BOUNDARY] ⚠️ ERRO: Boundary não encontrado ou inválido!")
            print(f"[ANÁLISE BOUNDARY]   - boundary_original existe: {boundary_original is not None}")
            if boundary_original:
                print(f"[ANÁLISE BOUNDARY]   - Tamanho: {len(boundary_original)} coordenadas (mínimo necessário: 6)")
            QMessageBox.warning(self, "Aviso", f"Boundary não encontrado ou inválido para o item {num}.\n\nVerifique se o boundary foi sincronizado corretamente.")
            print("=" * 100 + "\n")
            return
        
        # 3. PRINT DETALHADO DO BOUNDARY ORIGINAL
        print("\n" + ("-" * 100))
        print("[ANÁLISE BOUNDARY] ETAPA 1: BOUNDARY ORIGINAL")
        print("-" * 100)
        print(f"[ANÁLISE BOUNDARY] Total de coordenadas: {len(boundary_original)} ({len(boundary_original)//2} pontos)")
        
        # Calcular limites
        min_x = min(boundary_original[::2])
        max_x = max(boundary_original[::2])
        min_y = min(boundary_original[1::2])
        max_y = max(boundary_original[1::2])
        largura_original = max_x - min_x
        altura_original = max_y - min_y
        
        print(f"[ANÁLISE BOUNDARY] Limites:")
        print(f"[ANÁLISE BOUNDARY]   X: {min_x:.4f} a {max_x:.4f} (largura: {largura_original:.4f})")
        print(f"[ANÁLISE BOUNDARY]   Y: {min_y:.4f} a {max_y:.4f} (altura: {altura_original:.4f})")
        
        # Mostrar TODOS os pontos (importante para debug de recortes)
        n_pontos = len(boundary_original) // 2
        print(f"[ANÁLISE BOUNDARY] Todos os {n_pontos} pontos do boundary:")
        for i in range(n_pontos):
            idx = i * 2
            x = boundary_original[idx]
            y = boundary_original[idx+1]
            # Calcular distância até as bordas
            dist_x_min = abs(x - min_x)
            dist_x_max = abs(x - max_x)
            dist_y_min = abs(y - min_y)
            dist_y_max = abs(y - max_y)
            min_dist_borda = min(dist_x_min, dist_x_max, dist_y_min, dist_y_max)
            tipo = "INTERNO" if min_dist_borda > 5 else "BORDA"
            print(f"[ANÁLISE BOUNDARY]   P{i+1}: ({x:.4f}, {y:.4f}) - {tipo} (dist_borda={min_dist_borda:.2f})")
        
        # 4. ALINHAR BOUNDARY
        print("\n" + ("-" * 100))
        print("[ANÁLISE BOUNDARY] ETAPA 2: ALINHAMENTO HORIZONTAL")
        print("-" * 100)
        
        # alinhar_boundary_horizontal está definida neste mesmo arquivo (linha 153)
        coords_alinhadas = alinhar_boundary_horizontal(list(boundary_original))  # Cópia para não modificar original
        
        min_x_al = min(coords_alinhadas[::2])
        max_x_al = max(coords_alinhadas[::2])
        min_y_al = min(coords_alinhadas[1::2])
        max_y_al = max(coords_alinhadas[1::2])
        largura_al = max_x_al - min_x_al
        altura_al = max_y_al - min_y_al
        
        print(f"[ANÁLISE BOUNDARY] Após alinhamento:")
        print(f"[ANÁLISE BOUNDARY]   X: {min_x_al:.4f} a {max_x_al:.4f} (largura: {largura_al:.4f})")
        print(f"[ANÁLISE BOUNDARY]   Y: {min_y_al:.4f} a {max_y_al:.4f} (altura: {altura_al:.4f})")
        print(f"[ANÁLISE BOUNDARY]   Diferença de dimensões: largura={abs(largura_original - largura_al):.4f}, altura={abs(altura_original - altura_al):.4f}")
        
        # 5. ANÁLISE DE COORDENADAS
        print("\n" + ("-" * 100))
        print("[ANÁLISE BOUNDARY] ETAPA 3: ANÁLISE DE COORDENADAS (CHANFROS E ABERTURAS)")
        print("-" * 100)
        
        from fundo_utils import analisar_coordenadas
        try:
            analise, msg = analisar_coordenadas(coords_alinhadas)
            
            if not analise:
                print("[ANÁLISE BOUNDARY] ⚠️ ERRO: analisar_coordenadas retornou None!")
                QMessageBox.warning(self, "Erro", "Falha na análise de coordenadas. Verifique o console para detalhes.")
                print("=" * 100 + "\n")
                return
            
            print(f"[ANÁLISE BOUNDARY] ✅ Análise concluída: {msg}")
            
            # 5.1. DIMENSÕES
            dimensoes = analise.get('dimensoes', {})
            print(f"\n[ANÁLISE BOUNDARY] Dimensões detectadas:")
            print(f"[ANÁLISE BOUNDARY]   Largura: {dimensoes.get('largura', 0):.4f}")
            print(f"[ANÁLISE BOUNDARY]   Altura: {dimensoes.get('altura', 0):.4f}")
            print(f"[ANÁLISE BOUNDARY]   min_x: {dimensoes.get('min_x', 0):.4f}")
            print(f"[ANÁLISE BOUNDARY]   max_x: {dimensoes.get('max_x', 0):.4f}")
            print(f"[ANÁLISE BOUNDARY]   min_y: {dimensoes.get('min_y', 0):.4f}")
            print(f"[ANÁLISE BOUNDARY]   max_y: {dimensoes.get('max_y', 0):.4f}")
            
            # 5.2. CHANFROS
            chanfros = analise.get('chanfros', {})
            print(f"\n[ANÁLISE BOUNDARY] Chanfros detectados:")
            chanfros_nao_zero = {k: v for k, v in chanfros.items() if v and float_safe(v) > 0}
            if chanfros_nao_zero:
                for nome, valor in chanfros_nao_zero.items():
                    print(f"[ANÁLISE BOUNDARY]   {nome}: {float_safe(valor):.4f}")
            else:
                print(f"[ANÁLISE BOUNDARY]   Nenhum chanfro detectado (todos zero)")
            print(f"[ANÁLISE BOUNDARY]   Dicionário completo: {chanfros}")
            
            # 5.3. ABERTURAS (ANTES DA CONVERSÃO)
            aberturas_raw = analise.get('aberturas', [])
            print(f"\n[ANÁLISE BOUNDARY] Aberturas detectadas (formato raw): {len(aberturas_raw)}")
            for i, ab in enumerate(aberturas_raw):
                print(f"[ANÁLISE BOUNDARY]   Abertura {i+1}:")
                print(f"[ANÁLISE BOUNDARY]     x: {ab.get('x', 0):.4f}")
                print(f"[ANÁLISE BOUNDARY]     y: {ab.get('y', 0):.4f}")
                print(f"[ANÁLISE BOUNDARY]     largura: {ab.get('largura', 0):.4f}")
                print(f"[ANÁLISE BOUNDARY]     altura: {ab.get('altura', 0):.4f}")
                print(f"[ANÁLISE BOUNDARY]     lado: {ab.get('lado', 'N/A')}")
                print(f"[ANÁLISE BOUNDARY]     topo_fundo: {ab.get('topo_fundo', 'N/A')}")
                print(f"[ANÁLISE BOUNDARY]     profundidade: {ab.get('profundidade', 0):.4f}")
            
            # 6. MAPEAMENTO PARA UI (SIMULAR O QUE add_viga_with_boundary FAZ)
            print("\n" + ("-" * 100))
            print("[ANÁLISE BOUNDARY] ETAPA 4: MAPEAMENTO PARA INTERFACE (T/E, F/E, T/D, F/D)")
            print("-" * 100)
            
            largura_total = dimensoes.get('largura', 0)
            altura_total = dimensoes.get('altura', 0)
            
            aberturas_formatadas = [
                ["0", "0", "0"],  # T/E
                ["0", "0", "0"],  # F/E
                ["0", "0", "0"],  # T/D
                ["0", "0", "0"]   # F/D
            ]
            
            for ab in aberturas_raw:
                if not isinstance(ab, dict):
                    continue
                
                x = ab.get('x', 0)
                y = ab.get('y', 0)
                largura_ab = ab.get('largura', 0)
                altura_ab = ab.get('altura', 0)
                profundidade_ab = ab.get('profundidade', altura_ab)
                lado = ab.get('lado', '')
                topo_fundo = ab.get('topo_fundo', '')
                
                # Determinar lado se não estiver presente
                if not lado:
                    if abs(x) < 1e-2:
                        lado = 'esquerda'
                    elif abs((x + largura_ab) - largura_total) < 1e-2:
                        lado = 'direita'
                    elif x < largura_total / 2:
                        lado = 'esquerda'
                    else:
                        lado = 'direita'
                
                # Determinar topo/fundo se não estiver presente
                if not topo_fundo:
                    limiar = max(1.0, altura_total * 0.05)
                    if (y + altura_ab) >= (altura_total - limiar):
                        topo_fundo = 'topo'
                    elif y <= limiar:
                        topo_fundo = 'fundo'
                    elif y < (altura_total - (y + altura_ab)):
                        topo_fundo = 'fundo'
                    else:
                        topo_fundo = 'topo'
                
                # Calcular distância
                pos_x = ab.get('pos_x', '')
                if not pos_x:
                    if abs(x) < 1e-2 or abs((x + largura_ab) - largura_total) < 1e-2:
                        pos_x = 'parede'
                    else:
                        pos_x = 'central'
                
                if pos_x == 'parede':
                    distancia = 0
                else:
                    if lado == 'esquerda':
                        distancia = int(round(x))
                    else:
                        distancia = int(round(largura_total - (x + largura_ab)))
                
                # Determinar índice
                idx_pos = -1
                if lado == 'esquerda' and topo_fundo == 'topo':
                    idx_pos = 0  # T/E
                elif lado == 'esquerda' and topo_fundo == 'fundo':
                    idx_pos = 1  # F/E
                elif lado == 'direita' and topo_fundo == 'topo':
                    idx_pos = 2  # T/D
                elif lado == 'direita' and topo_fundo == 'fundo':
                    idx_pos = 3  # F/D
                
                if 0 <= idx_pos < 4:
                    aberturas_formatadas[idx_pos] = [
                        str(distancia),
                        str(int(round(profundidade_ab))),
                        str(int(round(largura_ab)))
                    ]
                    print(f"[ANÁLISE BOUNDARY] ✅ Abertura mapeada para posição {idx_pos}:")
                    print(f"[ANÁLISE BOUNDARY]   Posição: {['T/E', 'F/E', 'T/D', 'F/D'][idx_pos]}")
                    print(f"[ANÁLISE BOUNDARY]   Lado: {lado}, Topo/Fundo: {topo_fundo}")
                    print(f"[ANÁLISE BOUNDARY]   Distância: {distancia}, Profundidade: {profundidade_ab:.2f}, Largura: {largura_ab:.2f}")
            
            print(f"\n[ANÁLISE BOUNDARY] Array final de aberturas (formato UI):")
            posicoes = ['T/E (Topo Esquerda)', 'F/E (Fundo Esquerda)', 'T/D (Topo Direita)', 'F/D (Fundo Direita)']
            for i, (pos, ab_array) in enumerate(zip(posicoes, aberturas_formatadas)):
                dist, prof, larg = ab_array
                if float_safe(dist) > 0 or float_safe(prof) > 0 or float_safe(larg) > 0:
                    print(f"[ANÁLISE BOUNDARY]   {pos}: dist={dist}, prof={prof}, larg={larg}")
                else:
                    print(f"[ANÁLISE BOUNDARY]   {pos}: vazio")
            
            # 7. RESUMO FINAL
            print("\n" + ("-" * 100))
            print("[ANÁLISE BOUNDARY] RESUMO FINAL")
            print("-" * 100)
            aberturas_nao_vazias = sum(1 for ab in aberturas_formatadas if any(float_safe(v) > 0 for v in ab))
            chanfros_nao_zero_count = len(chanfros_nao_zero)
            
            print(f"[ANÁLISE BOUNDARY] ✅ Dimensões: {largura_total:.2f} x {altura_total:.2f}")
            print(f"[ANÁLISE BOUNDARY] ✅ Chanfros detectados: {chanfros_nao_zero_count}")
            print(f"[ANÁLISE BOUNDARY] ✅ Aberturas detectadas (raw): {len(aberturas_raw)}")
            print(f"[ANÁLISE BOUNDARY] ✅ Aberturas mapeadas para UI: {aberturas_nao_vazias}/4")
            
            if len(aberturas_raw) > 0 and aberturas_nao_vazias == 0:
                print(f"[ANÁLISE BOUNDARY] ⚠️ ATENÇÃO: {len(aberturas_raw)} abertura(s) detectada(s) mas nenhuma foi mapeada para UI!")
                print(f"[ANÁLISE BOUNDARY]   Possível problema no mapeamento lado/topo_fundo.")
            
            print("\n" + ("=" * 100))
            print("[ANÁLISE BOUNDARY] ===== FIM DA ANÁLISE MANUAL =====")
            print("=" * 100 + "\n")
            
            # Verificar se os dados estão salvos no item atual
            dados_salvos = None
            aberturas_salvas = None
            if isinstance(self.fundos_salvos, dict):
                dados_salvos = self.fundos_salvos.get(obra, {}).get(pav, {}).get(str(num))
                if dados_salvos:
                    aberturas_salvas = dados_salvos.get('aberturas', [])
            
            # IMPORTANTE: Se detectou aberturas mas não estão salvas, salvar agora!
            if aberturas_nao_vazias > 0:
                if not dados_salvos or not aberturas_salvas or all(all(float_safe(v) == 0 for v in ab[:3]) for ab in aberturas_salvas):
                    print(f"[ANÁLISE BOUNDARY] ⚠️ Aberturas detectadas mas não salvas! Salvando agora...")
                    # Garantir estrutura
                    if obra not in self.fundos_salvos:
                        self.fundos_salvos[obra] = {}
                    if pav not in self.fundos_salvos[obra]:
                        self.fundos_salvos[obra][pav] = {}
                    if str(num) not in self.fundos_salvos[obra][pav]:
                        self.fundos_salvos[obra][pav][str(num)] = {}
                    
                    # Atualizar aberturas no item salvo
                    self.fundos_salvos[obra][pav][str(num)]['aberturas'] = aberturas_formatadas.copy()
                    self.save_data()
                    print(f"[ANÁLISE BOUNDARY] ✅ Aberturas salvas: {aberturas_formatadas}")
                    
                    # Recarregar item se estiver selecionado
                    selected_items = self.tree_fundos.selectedItems()
                    if selected_items and (selected_items[0].data(0, Qt.UserRole) == num or selected_items[0].text(0) == str(num)):
                        print(f"[ANÁLISE BOUNDARY] ✅ Recarregando item para atualizar UI...")
                        self.load_selected(selected_items[0])
            
            # Mostrar resumo detalhado em MessageBox
            resumo = f"""Análise de Boundary - Item {num}

Dimensões: {largura_total:.2f} x {altura_total:.2f} mm
Chanfros detectados: {chanfros_nao_zero_count}
Aberturas detectadas: {len(aberturas_raw)} (mapeadas: {aberturas_nao_vazias}/4)

DETALHES DAS ABERTURAS MAPEADAS:"""
            
            # Adicionar detalhes de cada abertura mapeada
            posicoes_nomes = ['T/E', 'F/E', 'T/D', 'F/D']
            for i, (pos, ab_array) in enumerate(zip(posicoes_nomes, aberturas_formatadas)):
                dist, prof, larg = ab_array
                if float_safe(dist) > 0 or float_safe(prof) > 0 or float_safe(larg) > 0:
                    resumo += f"\n  ✅ {pos}: Dist={dist}, Prof={prof}, Larg={larg}"
                else:
                    resumo += f"\n  {pos}: vazio"
            
            # Adicionar status de salvamento
            resumo += f"\n\nSTATUS DE SALVAMENTO:"
            if dados_salvos:
                resumo += f"\n  ✅ Item encontrado nos dados salvos"
                if aberturas_salvas:
                    aberturas_salvas_count = sum(1 for ab in aberturas_salvas if isinstance(ab, (list, tuple)) and any(float_safe(v) > 0 for v in ab[:3]))
                    resumo += f"\n  Aberturas salvas: {aberturas_salvas_count}/4"
                    if aberturas_salvas_count == 0 and aberturas_nao_vazias > 0:
                        resumo += f"\n  ⚠️ PROBLEMA: Aberturas detectadas mas NÃO salvas!"
                        resumo += f"\n  → Tentando salvar agora..."
                    elif aberturas_salvas_count > 0:
                        resumo += f"\n  ✅ Aberturas salvas corretamente"
                else:
                    resumo += f"\n  ⚠️ Campo 'aberturas' não existe"
                    if aberturas_nao_vazias > 0:
                        resumo += f"\n  → Tentando salvar agora..."
            else:
                resumo += f"\n  ⚠️ Item não encontrado nos dados salvos!"
                if aberturas_nao_vazias > 0:
                    resumo += f"\n  → Tentando salvar agora..."
            
            resumo += f"\n\nVerifique o console para detalhes completos."""
            
            QMessageBox.information(self, "Análise de Boundary", resumo)
            
        except Exception as e:
            print(f"[ANÁLISE BOUNDARY] ⚠️ ERRO durante análise: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erro", f"Erro durante análise:\n\n{e}\n\nVerifique o console para detalhes.")
            print("=" * 100 + "\n")

    def load_selected(self, item):
        # Usar dados armazenados no UserRole (mais confiável)
        num = item.data(0, Qt.UserRole)
        obra = item.data(0, Qt.UserRole + 1)
        pav = item.data(0, Qt.UserRole + 2)
        seg_idx = item.data(0, Qt.UserRole + 3)  # -1 = viga inteira; >=0 = segmento
        if seg_idx is None:
            seg_idx = -1
        self._current_seg_idx = seg_idx  # para uso em action_salvar e item_loaded

        # Fallback para texto se não tiver dados armazenados
        if not num:
            num = item.text(0)
        if not obra:
            obra = self.combo_obra.currentText()
        if not pav:
            pav = item.text(2)  # Coluna Pavimento na Tree

        # Tentar buscar na estrutura nested
        dados = None
        if isinstance(self.fundos_salvos, dict):
            dados = self.fundos_salvos.get(obra, {}).get(pav, {}).get(str(num))
            # Lazy-load: se viga não tem segments_rich, tentar DB
            if dados and not dados.get('segments_rich'):
                segs_db = self._try_load_segments_from_db(obra, dados.get('nome', ''))
                if segs_db:
                    dados['segments_rich'] = segs_db
                    self.fundos_salvos[obra][pav][str(num)]['segments_rich'] = segs_db
                    self.save_data()
                    self._refresh_viga_children(item, segs_db, num, obra, pav)
            
            # IMPORTANTE: Se dados têm boundary_original, ele será usado no canvas
            # para desenhar a geometria real com todas as deformidades
            
            if dados:
                # Armazenar boundary_original se existir para uso no canvas
                if 'boundary_original' in dados and dados['boundary_original']:
                    self._current_boundary_original = dados['boundary_original']
                    print(f"[DEBUG] Boundary original carregado: {len(self._current_boundary_original)} coordenadas")
                else:
                    self._current_boundary_original = None
                    print(f"[DEBUG] ⚠️ Boundary original não encontrado nos dados carregados")

                # --- Modo segmento: preencher campos com dados deste segmento específico ---
                if seg_idx >= 0:
                    segs = dados.get('segments_rich', [])
                    if seg_idx < len(segs):
                        seg = segs[seg_idx]
                        nome_viga = dados.get('nome', 'V' + str(num))
                        seg_nome = f"{nome_viga}_seg{seg_idx}"
                        self.fields['numero'].setText(str(num))
                        self.fields['nome'].setText(seg_nome)
                        self.fields['obs'].setText(dados.get('obs', ''))
                        self.fields['pavimento'].setText(dados.get('pavimento', ''))
                        # largura deste segmento
                        seg_w = seg.get('total_width', seg.get('width', 0))
                        self.fields['largura'].setText(str(seg_w))
                        self.fields['altura'].setText(str(dados.get('altura', dados.get('total_width', '0'))))
                        self.fields['texto_esq'].setText('')
                        self.fields['texto_dir'].setText('')
                        # Painéis deste segmento
                        panels = seg.get('panels', [])
                        for i, pf in enumerate(self.paineis_fields):
                            if i < len(panels):
                                pf.setText(str(panels[i].get('width', panels[i].get('w', 0))))
                            else:
                                pf.setText('0')
                        # Recuos/aberturas: zerar (por segmento não tem)
                        for f in self.chanfros_fields: f.setText('0')
                        for row in self.aberturas_fields:
                            for f in row: f.setText('0')
                        self.update_canvas()
                        # Emitir formato "V301|seg0" para o viewer em main.py
                        self.item_loaded.emit(f"{nome_viga}|seg{seg_idx}")
                    return  # não continuar com o carregamento de viga inteira

                # --- Modo viga inteira (comportamento original) ---
                self.fields['numero'].setText(str(num))
                self.fields['nome'].setText(dados.get('nome', ''))
                self.fields['obs'].setText(dados.get('obs', ''))
                self.fields['pavimento'].setText(dados.get('pavimento', ''))
                self.fields['largura'].setText(str(dados.get('largura', '0')))
                self.fields['altura'].setText(str(dados.get('altura', '0')))
                self.fields['texto_esq'].setText(dados.get('texto_esq', ''))
                self.fields['texto_dir'].setText(dados.get('texto_dir', ''))

                
                paineis = dados.get('paineis', ["0"]*6)
                for i, p in enumerate(paineis):
                    if i < len(self.paineis_fields): self.paineis_fields[i].setText(str(p))
                
                recuos = dados.get('recuos', ["0"]*4)
                for i, r in enumerate(recuos):
                    if i < len(self.chanfros_fields): self.chanfros_fields[i].setText(str(r))
                
                # Carregar aberturas
                aberturas = dados.get('aberturas', [["0"]*3 for _ in range(4)])
                print(f"[DEBUG load_selected] Carregando aberturas do item {num}: {aberturas}")
                
                # Garantir que aberturas tem formato correto (4 posições, cada uma com 3 valores)
                if not aberturas or len(aberturas) != 4:
                    print(f"[DEBUG load_selected] ⚠️ Formato de aberturas inválido, usando padrão")
                    aberturas = [["0"]*3 for _ in range(4)]
                
                for i, ab_row in enumerate(aberturas):
                    if i < len(self.aberturas_fields):
                        # Garantir que ab_row tem 3 elementos
                        if not ab_row or len(ab_row) < 3:
                            ab_row = ["0", "0", "0"]
                        
                        for j, valor in enumerate(ab_row[:3]):
                            if j < len(self.aberturas_fields[i]):
                                valor_str = str(valor).strip()
                                self.aberturas_fields[i][j].setText(valor_str)
                                if float_safe(valor) > 0:
                                    posicoes = ['T/E', 'F/E', 'T/D', 'F/D']
                                    print(f"[DEBUG load_selected] ✅ Campo {posicoes[i]}[{j}] preenchido com: {valor_str}")
                
                # Verificar se algum campo foi preenchido
                aberturas_preenchidas = sum(1 for i, ab_row in enumerate(aberturas) if i < len(self.aberturas_fields) and any(float_safe(v) > 0 for v in ab_row[:3]))
                if aberturas_preenchidas > 0:
                    print(f"[DEBUG load_selected] ✅ Total de {aberturas_preenchidas} posição(ões) de abertura preenchida(s)")
                else:
                    print(f"[DEBUG load_selected] ⚠️ Nenhuma abertura preenchida nos campos da UI")
                
                # Carregar modo de continuação
                continuacao = dados.get('continuacao', 'UltimoSegmento')
                if continuacao == "Obstaculo":
                    self.rb_cont_obs.setChecked(True)
                elif continuacao == "Abertura":
                    self.rb_cont_abertura.setChecked(True)
                else:  # UltimoSegmento (padrão)
                    self.rb_cont_last.setChecked(True)
                
                self.update_canvas()
                self.item_loaded.emit(str(self.fields['nome'].text() or num))

    def _try_load_segments_from_db(self, obra: str, elemento_id: str) -> list:
        """Busca segments_rich do DB SQLite para uma viga FV."""
        if not elemento_id:
            return []
        try:
            import sqlite3 as _sqlite3, json as _json2
            _db_candidates = [
                'D:/Agente-cad-PYSIDE/project_data.vision',
                os.path.join(os.path.dirname(__file__), '..', '..', '..', 'project_data.vision'),
            ]
            for _db in _db_candidates:
                if not os.path.exists(_db):
                    continue
                conn = _sqlite3.connect(_db)
                cur = conn.cursor()
                cur.execute(
                    "SELECT campos_json FROM reverse_eng_fichas "
                    "WHERE classe='FV' AND elemento_id=?",
                    (elemento_id,)
                )
                row = cur.fetchone()
                conn.close()
                if row and row[0]:
                    ficha = _json2.loads(row[0])
                    segs = ficha.get('segments_rich', [])
                    if segs:
                        print(f"[FV] Segmentos carregados do DB: {elemento_id} → {len(segs)} segmentos")
                        return segs
        except Exception as e:
            print(f"[FV] Erro ao carregar segmentos do DB para {elemento_id}: {e}")
        return []

    def _refresh_viga_children(self, viga_item, segs: list, num, obra: str, pav: str):
        """Atualiza filhos do item de viga na árvore com os segmentos carregados."""
        viga_item.takeChildren()
        for seg_i, seg in enumerate(segs):
            sw = float_safe(seg.get('total_width', seg.get('width', 0)))
            np_seg = len(seg.get('panels', []))
            seg_item = QTreeWidgetItem([
                f"  S{seg_i + 1}",
                f"{viga_item.text(1)} Seg {seg_i + 1}",
                '',
                f"{sw:.1f}cm / {np_seg}P",
            ])
            seg_item.setData(0, Qt.UserRole,     num)
            seg_item.setData(0, Qt.UserRole + 1, obra)
            seg_item.setData(0, Qt.UserRole + 2, pav)
            seg_item.setData(0, Qt.UserRole + 3, seg_i)
            viga_item.addChild(seg_item)
        viga_item.setExpanded(True)

    def save_data(self):
        try:
            with open(self.caminho_dados, 'w', encoding='utf-8') as f:

                json.dump(self.fundos_salvos, f, indent=4, ensure_ascii=False)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo: {e}")

    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # 1. Content Splitter (Sem Top Bar)
        splitter = QSplitter(Qt.Horizontal)
        
        # --- COLUNA 1: LISTA (Esquerda) ---
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(5, 5, 5, 5)
        
        lbl_list_header = QLabel("LISTA DE VIGAS")
        lbl_list_header.setStyleSheet("font-size: 14px; font-weight: bold; color: #4fc3f7; margin-bottom: 5px;")
        left_layout.addWidget(lbl_list_header)
        
        # Container de Filtros (Estilo "Dark Box" opcional ou apenas Layout)
        filter_container = QWidget()
        filter_container.setStyleSheet("background-color: #252526; border-radius: 5px; border: 1px solid #333;")
        filter_layout_main = QVBoxLayout(filter_container)
        filter_layout_main.setSpacing(5)
        filter_layout_main.setContentsMargins(5, 5, 5, 5)

        # Linha 1: Obra
        row_obra = QHBoxLayout()
        lbl_obra = QLabel("Obra:")
        lbl_obra.setFixedWidth(35)
        self.combo_obra = QComboBox()
        self.combo_obra.addItems(["Obra 1"]) # Default
        self.btn_add_obra = QPushButton("➕")
        self.btn_add_obra.setFixedWidth(35)
        self.btn_add_obra.setFixedHeight(25)
        self.btn_add_obra.setToolTip("Adicionar Nova Obra")
        self.btn_add_obra.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 3px;
                border: 1px solid #388e3c;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
        """)
        self.btn_rem_obra = QPushButton("🗑️")
        self.btn_rem_obra.setFixedWidth(35)
        self.btn_rem_obra.setFixedHeight(25)
        self.btn_rem_obra.setToolTip("Remover Obra (exclui todos os pavimentos e itens)")
        self.btn_rem_obra.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-size: 12px;
                font-weight: bold;
                border-radius: 3px;
                border: 1px solid #d32f2f;
            }
            QPushButton:hover {
                background-color: #e53935;
            }
            QPushButton:pressed {
                background-color: #d32f2f;
            }
        """)
        
        row_obra.addWidget(lbl_obra)
        row_obra.addWidget(self.combo_obra)
        # btn_add_obra e btn_rem_obra criados mas NÃO exibidos (sync gerencia obras)
        filter_layout_main.addLayout(row_obra)

        # Linha 2: Pavimento
        row_pav = QHBoxLayout()
        lbl_pav = QLabel("Pav:")
        lbl_pav.setFixedWidth(35)
        self.combo_pavimento = QComboBox() # Será populado dinamicamente
        self.combo_pavimento.addItem("1", "1")  # Default inicial
        self.btn_add_pav = QPushButton("➕")
        self.btn_add_pav.setFixedWidth(35)
        self.btn_add_pav.setFixedHeight(25)
        self.btn_add_pav.setToolTip("Adicionar Novo Pavimento")
        self.btn_add_pav.setStyleSheet("""
            QPushButton {
                background-color: #4caf50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 3px;
                border: 1px solid #388e3c;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
        """)
        self.btn_paste_pav = QPushButton("📋") # Paste icon
        self.btn_paste_pav.setToolTip("Colar/Editar")
        self.btn_paste_pav.setFixedWidth(35)
        self.btn_paste_pav.setFixedHeight(25)
        self.btn_paste_pav.setStyleSheet("""
            QPushButton {
                background-color: #2196f3;
                color: white;
                font-size: 12px;
                border-radius: 3px;
                border: 1px solid #1976d2;
            }
            QPushButton:hover {
                background-color: #1e88e5;
            }
            QPushButton:pressed {
                background-color: #1976d2;
            }
        """)
        self.btn_rem_pav = QPushButton("🗑️")
        self.btn_rem_pav.setFixedWidth(35)
        self.btn_rem_pav.setFixedHeight(25)
        self.btn_rem_pav.setToolTip("Remover Pavimento (exclui todos os itens do pavimento)")
        self.btn_rem_pav.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-size: 12px;
                font-weight: bold;
                border-radius: 3px;
                border: 1px solid #d32f2f;
            }
            QPushButton:hover {
                background-color: #e53935;
            }
            QPushButton:pressed {
                background-color: #d32f2f;
            }
        """)
        
        row_pav.addWidget(lbl_pav)
        row_pav.addWidget(self.combo_pavimento)
        # btn_add_pav, btn_paste_pav, btn_rem_pav criados mas NÃO exibidos (sync gerencia pavimentos)
        filter_layout_main.addLayout(row_pav)

        # Total M2 dentro do container ou logo abaixo
        self.label_total_m2 = QLabel("Total m²: 0.00")
        self.label_total_m2.setStyleSheet("font-weight: bold; color: #4fc3f7; margin-top: 5px;")
        filter_layout_main.addWidget(self.label_total_m2)

        left_layout.addWidget(filter_container)
        
        # Lista (TreeWidget) - Configurado para seleção múltipla
        self.tree_fundos = QTreeWidget()
        self.tree_fundos.setHeaderLabels(["Nº", "Nome", "Pav", "Face/Comb"]) # Ajustado headers conforme img
        self.tree_fundos.header().setSectionResizeMode(QHeaderView.ResizeToContents)
        # Habilitar seleção múltipla com Shift/Ctrl
        self.tree_fundos.setSelectionMode(QTreeWidget.ExtendedSelection)
        self.tree_fundos.setStyleSheet("QTreeWidget { background-color: #1e1e1e; border: 1px solid #333; } QHeaderView::section { background-color: #2d2d2d; }")
        left_layout.addWidget(self.tree_fundos)

        # Botões de ação secundária no footer da esquerda
        left_actions = QHBoxLayout()
        self.btn_delete_sel = QPushButton("🗑️ Excluir Selecionados")
        self.btn_delete_sel.setStyleSheet("""
            QPushButton {
                background-color: #d32f2f; 
                color: white; 
                font-weight: bold;
                padding: 8px 15px;
                border-radius: 4px;
                border: 1px solid #b71c1c;
            }
            QPushButton:hover {
                background-color: #c62828;
            }
            QPushButton:pressed {
                background-color: #b71c1c;
            }
        """)
        left_actions.addWidget(self.btn_delete_sel)
        left_layout.addLayout(left_actions)
        
        splitter.addWidget(left_widget)

        # --- COLUNA 2: VISUALIZAÇÃO + FICHA (Direita, layout vertical) ---
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(5, 5, 5, 5)
        right_layout.setSpacing(4)

        # Canvas (viewer) horizontal no topo — ocupa a maior parte vertical
        self.canvas = FundoCanvas()
        self.canvas.setMinimumHeight(300)
        right_layout.addWidget(self.canvas, stretch=3)

        # ScrollArea para comandos se a tela for pequena
        scroll_cmds = QScrollArea()
        scroll_cmds.setWidgetResizable(True)
        scroll_cmds.setFrameShape(QFrame.NoFrame)
        cmd_widget = QWidget()
        cmd_layout = QVBoxLayout(cmd_widget)
        cmd_layout.setContentsMargins(0, 0, 0, 0)
        cmd_layout.setSpacing(10)

        # GRUPO 1: PRINCIPAL (REMOVIDO)
        # grp_main = QGroupBox("Principal")
        # ... (Código removido a pedido do usuário)

        # Botões de Produção/Geração/Ferramentas: criados mas NÃO exibidos no painel.
        # Os QGroupBox são self.* para evitar GC dos filhos ("Signal source has been deleted").
        self.grp_prod = QGroupBox("Produção")
        prod_grid = QGridLayout(self.grp_prod)
        self.btns_prod = [
            QPushButton("Próxima Viga"), QPushButton("Próximo Segmento"),
            QPushButton("Extensão L"), QPushButton("Cancelar")
        ]
        colors = ["#2E7D32", "#2E7D32", "#ff9800", "#f44336"]
        for i, b in enumerate(self.btns_prod):
            b.setStyleSheet(f"background-color: {colors[i]}; font-weight: bold; color: white; padding: 4px; font-size: 11px;")
            prod_grid.addWidget(b, i // 2, i % 2)
        # self.grp_prod NÃO adicionado ao layout

        self.grp_geracao = QGroupBox("Geração")
        geracao_grid = QGridLayout(self.grp_geracao)
        self.btn_gerar_segmento = QPushButton("Gerar Segmento")
        self.btn_gerar_conjunto = QPushButton("Gerar conjunto de Viga")
        self.btn_gerar_pavimento = QPushButton("Gerar Pavimento")
        self.btn_gerar_segmento.setStyleSheet("background-color: #FFEB3B; color: black; font-weight: bold; padding: 4px; font-size: 11px;")
        self.btn_gerar_conjunto.setStyleSheet("background-color: #1976D2; color: white; font-weight: bold; padding: 4px; font-size: 11px;")
        self.btn_gerar_pavimento.setStyleSheet("background-color: #C62828; color: white; font-weight: bold; padding: 4px; font-size: 11px;")
        geracao_grid.addWidget(self.btn_gerar_segmento, 0, 0)
        geracao_grid.addWidget(self.btn_gerar_conjunto, 0, 1)
        geracao_grid.addWidget(self.btn_gerar_pavimento, 1, 0, 1, 2)
        # self.grp_geracao NÃO adicionado ao layout

        self.grp_tools = QGroupBox("Ferramentas")
        tools_grid = QGridLayout(self.grp_tools)
        self.btn_config_desenho = QPushButton("Configuração Desenho")
        self.btn_config_ordenador = QPushButton("Configuração Ordenador")
        self.btn_lisp_fd = QPushButton("Criar comando lisp FD")
        for btn in [self.btn_config_desenho, self.btn_config_ordenador, self.btn_lisp_fd]:
            btn.setStyleSheet("background-color: #BA68C8; color: white; padding: 4px; font-size: 11px;")
        tools_grid.addWidget(self.btn_config_desenho, 0, 0)
        tools_grid.addWidget(self.btn_config_ordenador, 0, 1)
        tools_grid.addWidget(self.btn_lisp_fd, 1, 0, 1, 2)
        # self.grp_tools NÃO adicionado ao layout

        # btn_analisar_boundary: criado mas NÃO exibido
        self.btn_analisar_boundary = QPushButton("Analisar Boundary")
        self.btn_analisar_boundary.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold; font-size: 12px; padding: 8px; min-height: 30px;")

        # Somente o botão Salvar é exibido
        self.btn_salvar = QPushButton("Salvar")
        self.btn_salvar.setStyleSheet(
            "background-color: #1976D2; color: white; font-weight: bold; "
            "font-size: 13px; padding: 8px; min-height: 32px; border-radius: 4px;")
        cmd_layout.addWidget(self.btn_salvar)

        # GRUPO 4: DETALHES (Tabs)
        # O TabWidget fica como último elemento do scroll
        self.tabs_details = QTabWidget()
        self.tabs_details.setMinimumHeight(400) # Garantir que tenha altura razoável
        
        # ABA 1: GERAL
        tab_geral = QWidget()
        layout_geral = QVBoxLayout(tab_geral)
        
        info_group = QGroupBox("Informações Básicas")
        info_grid = QGridLayout(info_group)
        self.fields = {}
        basic_fields = [
            ("Número:", "numero", 0, 0), ("Nome:", "nome", 1, 0),
            ("Observações:", "obs", 2, 0), ("Pavimento:", "pavimento", 3, 0),
            ("Largura:", "largura", 0, 2), ("Altura:", "altura", 1, 2),
            ("Texto Esq:", "texto_esq", 2, 2), ("Texto Dir:", "texto_dir", 3, 2)
        ]
        for label, key, r, c in basic_fields:
            info_grid.addWidget(QLabel(label), r, c)
            edit = QLineEdit()
            info_grid.addWidget(edit, r, c + 1)
            self.fields[key] = edit
            if key in ["largura", "altura"]:
                edit.textChanged.connect(self.update_calculations)
        layout_geral.addWidget(info_group)
        
        # 2 - Detalhes de Salvamento (Modo de Continuação)
        cont_group = QGroupBox("2 - Detalhes de Salvamento")
        cont_layout = QVBoxLayout(cont_group)
        
        cont_label = QLabel("Deseja continuar a viga?")
        cont_layout.addWidget(cont_label)
        
        h_radios_layout = QHBoxLayout()
        self.rb_cont_obs = QRadioButton("Viga continua após\nObstáculo")
        self.rb_cont_last = QRadioButton("Último Segmento\ndeste lado da Viga")
        self.rb_cont_last.setChecked(True)  # Padrão
        self.rb_cont_abertura = QRadioButton("Próximo segmento continua\nabertura de viga")
        
        h_radios_layout.addWidget(self.rb_cont_obs)
        h_radios_layout.addWidget(self.rb_cont_last)
        h_radios_layout.addWidget(self.rb_cont_abertura)
        
        # Grouping for Isolation
        self.bg_continuation = QButtonGroup(self)
        self.bg_continuation.addButton(self.rb_cont_obs)
        self.bg_continuation.addButton(self.rb_cont_last)
        self.bg_continuation.addButton(self.rb_cont_abertura)
        
        cont_layout.addLayout(h_radios_layout)
        layout_geral.addWidget(cont_group)
        
        layout_geral.addStretch()
        self.tabs_details.addTab(tab_geral, "Geral")

        # ABA 2: PAINÉIS
        tab_paineis = QWidget()
        layout_paineis = QVBoxLayout(tab_paineis)
        
        paineis_group = QGroupBox("Configuração de Painéis")
        paineis_vbox = QVBoxLayout(paineis_group)
        
        dist_layout = QHBoxLayout()
        dist_layout.addWidget(QLabel("Distribuição:"))
        self.rad_122 = QRadioButton("122")
        self.rad_307 = QRadioButton("307")
        self.rad_122.setChecked(True)
        dist_layout.addWidget(self.rad_122)
        dist_layout.addWidget(self.rad_307)
        paineis_vbox.addLayout(dist_layout)
        
        p_inputs_layout = QGridLayout()
        self.paineis_fields = []
        for i in range(6):
            p_inputs_layout.addWidget(QLabel(f"P{i+1}:"), 0, i)
            edit = QLineEdit("0")
            edit.setFixedWidth(50)
            p_inputs_layout.addWidget(edit, 1, i)
            self.paineis_fields.append(edit)
        paineis_vbox.addLayout(p_inputs_layout)
        layout_paineis.addWidget(paineis_group)

        # Opções (Sarrafos)
        opts_group = QGroupBox("Opções")
        opts_layout = QVBoxLayout(opts_group)
        self.chk_sarrafo_esq = QCheckBox("Sarrafo vertical esquerda")
        self.chk_sarrafo_dir = QCheckBox("Sarrafo vertical direita")
        self.chk_sarrafo_esq.setChecked(True)
        self.chk_sarrafo_dir.setChecked(True)
        opts_layout.addWidget(self.chk_sarrafo_esq)
        opts_layout.addWidget(self.chk_sarrafo_dir)
        layout_paineis.addWidget(opts_group)
        layout_paineis.addStretch()
        self.tabs_details.addTab(tab_paineis, "Painéis")

        # ABA 3: RECUOS/ABERTURAS
        tab_recuos = QWidget()
        layout_recuos = QVBoxLayout(tab_recuos)
        
        ca_group = QGroupBox("Chanfros e Aberturas")
        ca_layout = QVBoxLayout(ca_group)
        
        # Chanfros
        chanfros_grid = QGridLayout()
        chanfros_grid.addWidget(QLabel("Chanfros:"), 0, 0, 1, 2)
        self.chanfros_fields = []
        for i, lab in enumerate(["T/E", "F/E", "T/D", "F/D"]):
            chanfros_grid.addWidget(QLabel(lab), 1, i)
            edit = QLineEdit("0")
            edit.setFixedWidth(50)
            chanfros_grid.addWidget(edit, 2, i)
            self.chanfros_fields.append(edit)
        ca_layout.addLayout(chanfros_grid)
        
        # Aberturas
        ca_layout.addWidget(QLabel("Aberturas (Dist/Prof/Larg):"))
        abert_grid = QGridLayout()
        self.aberturas_fields = []
        for i, lab in enumerate(["T/E", "F/E", "T/D", "F/D"]):
            abert_grid.addWidget(QLabel(lab + ":"), i, 0)
            row_fields = []
            for j in range(3):
                edit = QLineEdit("0")
                edit.setFixedWidth(45)
                abert_grid.addWidget(edit, i, j + 1)
                row_fields.append(edit)
            self.aberturas_fields.append(row_fields)
        ca_layout.addLayout(abert_grid)
        
        layout_recuos.addWidget(ca_group)
        layout_recuos.addStretch()
        self.tabs_details.addTab(tab_recuos, "Chanfros/Aberturas")

        # ABA 4: PAINEL L
        tab_l = QWidget()
        layout_l = QVBoxLayout(tab_l)
        
        l_group = QGroupBox("Painel 2 em L")
        l_layout = QVBoxLayout(l_group)
        
        l_types = QHBoxLayout()
        self.l_type_rads = []
        for t in ["E/T", "E/F", "D/T", "D/F"]:
            rad = QRadioButton(t)
            l_types.addWidget(rad)
            self.l_type_rads.append(rad)
        self.l_type_rads[0].setChecked(True)
        l_layout.addLayout(l_types)
        
        l_basic_grid = QGridLayout()
        self.l_fields = {}
        l_basic_configs = [
            ("Comp:", "comp2", 0, 0), ("Larg:", "larg2", 0, 2),
            ("P1:", "p1_2", 1, 0), ("P2:", "p2_2", 1, 2), ("P3:", "p3_2", 1, 4)
        ]
        for label, key, r, c in l_basic_configs:
            l_basic_grid.addWidget(QLabel(label), r, c)
            edit = QLineEdit("0")
            edit.setFixedWidth(50)
            l_basic_grid.addWidget(edit, r, c+1)
            self.l_fields[key] = edit
        l_layout.addLayout(l_basic_grid)

        l_utils_tabs = QTabWidget()
        # L Chanfros
        tab_l_chan = QWidget()
        lay_l_chan = QHBoxLayout(tab_l_chan)
        self.l_chanfros_fields = []
        for lab in ["T/E", "F/E", "T/D", "F/D"]:
            v = QVBoxLayout()
            v.addWidget(QLabel(lab))
            edit = QLineEdit("0")
            edit.setFixedWidth(40)
            v.addWidget(edit)
            lay_l_chan.addLayout(v)
            self.l_chanfros_fields.append(edit)
        l_utils_tabs.addTab(tab_l_chan, "Chanfros L")

        # L Aberturas
        tab_l_ab = QWidget()
        lay_l_ab = QGridLayout(tab_l_ab)
        self.l_aberturas_fields = []
        for i, lab in enumerate(["T/E", "F/E", "T/D", "F/D"]):
            lay_l_ab.addWidget(QLabel(lab), i, 0)
            row_fields = []
            for j in range(3):
                edit = QLineEdit("0")
                edit.setFixedWidth(35)
                lay_l_ab.addWidget(edit, i, j+1)
                row_fields.append(edit)
            self.l_aberturas_fields.append(row_fields)
        l_utils_tabs.addTab(tab_l_ab, "Aberturas L")
        
        l_layout.addWidget(l_utils_tabs)
        layout_l.addWidget(l_group)
        layout_l.addStretch()
        self.tabs_details.addTab(tab_l, "Painel L")

        cmd_layout.addWidget(self.tabs_details)
        cmd_layout.addStretch() # Empurrar tudo pra cima

        scroll_cmds.setWidget(cmd_widget)
        # Scroll de campos abaixo do canvas — stretch menor
        right_layout.addWidget(scroll_cmds, stretch=2)

        splitter.addWidget(right_container)

        # Splitter 2 colunas: Lista (1) | Viewer+Ficha (4)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 4)
        splitter.setSizes([260, 1040])
        
        main_layout.addWidget(splitter)

    def apply_theme(self):
        self.setStyleSheet("""
            QMainWindow, QWidget { background-color: #1e1e1e; color: #e0e0e0; font-family: 'Segoe UI', Arial; }
            QGroupBox { border: 1px solid #333; border-radius: 5px; margin-top: 10px; font-weight: bold; color: #4fc3f7; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px 0 3px; }
            QLineEdit { background-color: #252526; border: 1px solid #333; color: #ffffff; padding: 4px; border-radius: 3px; }
            QPushButton { background-color: #333; border: 1px solid #444; color: #ffffff; padding: 6px 12px; border-radius: 4px; }
            QPushButton:hover { background-color: #444; border: 1px solid #4fc3f7; }
            QTreeWidget { background-color: #252526; border: 1px solid #333; alternate-background-color: #2d2d2d; }
            QHeaderView::section { background-color: #333; color: white; border: 1px solid #222; }
            QScrollBar:vertical { border: none; background: #2d2d2d; width: 10px; }
            QScrollBar::handle:vertical { background: #444; min-height: 20px; }
        """)

    def load_data(self):
        self.obras_metadata = {"Obra 1": set()} # Obra -> Set de Pavimentos
        
        raw_data = {}
        if os.path.exists(self.caminho_dados):
            try:
                with open(self.caminho_dados, 'r', encoding='utf-8') as f:

                    raw_data  = json.load(f)
            except:
                raw_data = {}
        
        # Estrutura Alvo: Nested [Obra][Pavimento][Numero]
        self.fundos_salvos = {} # {obra: {pav: {num: dados}}}
        
        # Migração de dados legados (Flat ou parciamente nested)
        if isinstance(raw_data, dict):
            for key, dados in raw_data.items():
                if isinstance(dados, dict) and ('numero' in dados or 'nome' in dados):
                    # É um item individual (Legado Flat)
                    obra = dados.get('obra', 'Obra 1')
                    pav = str(dados.get('pavimento', 'Indefinido'))
                    num = str(key)
                    
                    if obra not in self.fundos_salvos: self.fundos_salvos[obra] = {}
                    if pav not in self.fundos_salvos[obra]: self.fundos_salvos[obra][pav] = {}
                    self.fundos_salvos[obra][pav][num] = dados
                elif isinstance(dados, dict):
                    # Assume que já é nested {pav: {num: dados}}
                    self.fundos_salvos[key] = dados
        
        # Sincronizar Metadados
        for obra, pavs_dict in self.fundos_salvos.items():
            if obra not in self.obras_metadata:
                self.obras_metadata[obra] = set()
            for pav_name in pavs_dict.keys():
                self.obras_metadata[obra].add(pav_name)
                
        # Atualizar Combos
        self.combo_obra.blockSignals(True)
        self.combo_obra.clear()
        obras_sorted = sorted(list(self.obras_metadata.keys()))
        self.combo_obra.addItems(obras_sorted)
        self.combo_obra.blockSignals(False)
        
        self._update_combo_pavimentos()


    def _update_combo_pavimentos(self):
        current_obra = self.combo_obra.currentText()
        if not current_obra: return
        
        pavs = sorted(list(self.obras_metadata.get(current_obra, [])))
        if not pavs: pavs = ["1"]

        self.combo_pavimento.blockSignals(True)
        self.combo_pavimento.clear()
        for _p in pavs: self.combo_pavimento.addItem(_fmt_pav_label(_p), _p)
        self.combo_pavimento.blockSignals(False)
        
        self.update_list()

    def _on_pavimento_context_changed(self):
        self.sync_n3_fundos_to_current_context()
        self.update_list()

    def update_list(self):
        self.tree_fundos.clear()
        total_m2 = 0
        
        current_obra = self.combo_obra.currentText()
        current_pav = _cmb_pav_raw(self.combo_pavimento)

        # Filtros
        if not current_obra or current_obra not in self.fundos_salvos: return
        
        # Coletar itens baseado no pavimento selecionado
        items_to_process = []
        if current_pav == "Todos":
             # Todos os pavimentos da obra atual
             for pav_name, items_dict in self.fundos_salvos[current_obra].items():
                 for num, dados in items_dict.items():
                     items_to_process.append((num, dados))
        else:
             # Pavimento específico
             items_pav = self.fundos_salvos[current_obra].get(current_pav, {})
             for num, dados in items_pav.items():
                 items_to_process.append((num, dados))

        # Agrupamento (Clusters)
        clusters = {} # Chave -> [Items]
        
        for num, dados in items_to_process:
            m2 = float_safe(dados.get('m2', 0))
            total_m2 += m2
            
            # Lógica de Cluster / Classes (suporta formato simplificado V1a, V1b, V1c)
            parent_name = dados.get('parent_name')
            nome = str(dados.get('nome', ''))
            key_num = str(num).split('.')[0] # 1.1 -> 1
            
            if parent_name:
                cluster_key = parent_name
            elif nome.upper().startswith("V"):
                # Extrair prefixo base (V1 de V1a, V1b, etc)
                import re
                match = re.match(r'^([Vv]?\d+[a-zA-Z]*)', nome)
                if match:
                    cluster_key = match.group(1)
                else:
                    # Fallback: usar nome completo
                    cluster_key = nome
            elif key_num:
                cluster_key = f"V{key_num}" # Fallback
            else:
                cluster_key = "Outros"
                
            if cluster_key not in clusters: clusters[cluster_key] = []
            clusters[cluster_key].append((num, dados, m2))


        # Adicionar à Tree
        import re
        def natural_key(text):
            if not text: return []
            return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

        for cluster_name in sorted(clusters.keys(), key=natural_key):
            items = clusters[cluster_name]
            
            # Criar Parent Item
            parent = QTreeWidgetItem([cluster_name, "", "", ""])
            parent.setBackground(0, QBrush(QColor("#333")))
            parent.setExpanded(True)
            self.tree_fundos.addTopLevelItem(parent)
            
            for num, dados, m2 in sorted(items, key=lambda x: natural_key(x[0])): # Sort by num (Natural)
                child = QTreeWidgetItem([str(num), dados.get('nome', ''), dados.get('pavimento', ''), f"{m2:.4f}"])
                # Armazenar número, obra e pavimento para facilitar exclusão
                child.setData(0, Qt.UserRole, str(num))  # Guardar a chave original (numero)
                child.setData(0, Qt.UserRole + 1, current_obra)  # Obra
                # Obter pavimento do dado ou do contexto atual
                pav_item = dados.get('pavimento', current_pav)
                child.setData(0, Qt.UserRole + 2, pav_item)  # Pavimento
                parent.addChild(child)



        self.label_total_m2.setText(f"Total M²: {total_m2:.2f}")

    def open_config(self):
        c = ConfigDialog(self.config_manager, self)
        c.exec()

    def action_excel_export(self):
        QMessageBox.information(self, "Exportar", "Funcionalidade de Exportação em desenvolvimento.")

    def action_excel_import(self):
        QMessageBox.information(self, "Importar", "Funcionalidade de Importação em desenvolvimento.")

    def action_add_obra(self):
        from PySide6.QtWidgets import QInputDialog
        text, ok = QInputDialog.getText(self, "Nova Obra", "Nome da Obra:")
        if ok and text:
            if text not in self.obras_metadata:
                self.obras_metadata[text] = set()
                self.combo_obra.addItem(text)
                self.combo_obra.setCurrentText(text)
            else:
                QMessageBox.warning(self, "Erro", "Obra já existe.")

    def action_rem_obra(self):
        """Remove obra e TODOS os seus pavimentos e itens"""
        curr = self.combo_obra.currentText()
        if not curr: 
            QMessageBox.warning(self, "Aviso", "Nenhuma obra selecionada.")
            return
        
        # Contar itens que serão excluídos
        total_itens = 0
        total_pavimentos = 0
        if curr in self.fundos_salvos:
            for pav, items_dict in self.fundos_salvos[curr].items():
                total_pavimentos += 1
                total_itens += len(items_dict)
        
        # Confirmar com mensagem detalhada
        msg = f"⚠️ ATENÇÃO: Esta ação é IRREVERSÍVEL!\n\n"
        msg += f"Você está prestes a excluir:\n"
        msg += f"• Obra: '{curr}'\n"
        msg += f"• {total_pavimentos} Pavimento(s)\n"
        msg += f"• {total_itens} Item(s) de fundo\n\n"
        msg += f"TODOS os dados serão PERDIDOS PERMANENTEMENTE.\n\n"
        msg += f"Deseja realmente continuar?"
        
        ret = QMessageBox.warning(
            self, 
            "Confirmar Exclusão de Obra", 
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if ret == QMessageBox.Yes:
            # Excluir obra e todos os seus dados
            if curr in self.fundos_salvos:
                del self.fundos_salvos[curr]
            if curr in self.obras_metadata:
                del self.obras_metadata[curr]
            
            # Remover do combo
            idx = self.combo_obra.findText(curr)
            if idx >= 0:
                self.combo_obra.removeItem(idx)
            
            # Salvar dados
            self.save_data()
            
            # Atualizar lista
            self.update_list()
            
            self.log(f"🗑️ Obra '{curr}' e todos os seus dados foram excluídos permanentemente.")
            QMessageBox.information(self, "Sucesso", f"Obra '{curr}' e todos os seus dados foram excluídos.")

    def action_add_pav(self):
        # Apenas adiciona ao metadata, não cria item
        # Mas para aparecer no Combo, deve estar no metadata set()
        curr_obra = self.combo_obra.currentText()
        if not curr_obra: return
        
        from PySide6.QtWidgets import QInputDialog
        text, ok = QInputDialog.getText(self, "Novo Pavimento", "Nome do Pavimento:")
        if ok and text:
            if text in self.obras_metadata[curr_obra]:
                QMessageBox.warning(self, "Erro", "Pavimento já existe nesta obra.")
                return
            self.obras_metadata[curr_obra].add(text)
            self._update_combo_pavimentos()
            _set_cmb_by_data(self.combo_pavimento, text)

    def action_rem_pav(self):
        """Remove pavimento e TODOS os seus itens"""
        curr_obra = self.combo_obra.currentText()
        curr_pav = _cmb_pav_raw(self.combo_pavimento)
        if not curr_obra or not curr_pav:
            QMessageBox.warning(self, "Aviso", "Selecione obra e pavimento.")
            return
        
        # Contar itens que serão excluídos
        total_itens = 0
        if curr_obra in self.fundos_salvos and curr_pav in self.fundos_salvos[curr_obra]:
            total_itens = len(self.fundos_salvos[curr_obra][curr_pav])
        
        # Confirmar com mensagem detalhada
        msg = f"⚠️ ATENÇÃO: Esta ação é IRREVERSÍVEL!\n\n"
        msg += f"Você está prestes a excluir:\n"
        msg += f"• Obra: '{curr_obra}'\n"
        msg += f"• Pavimento: '{curr_pav}'\n"
        msg += f"• {total_itens} Item(s) de fundo\n\n"
        msg += f"TODOS os dados do pavimento serão PERDIDOS PERMANENTEMENTE.\n\n"
        msg += f"Deseja realmente continuar?"
        
        ret = QMessageBox.warning(
            self, 
            "Confirmar Exclusão de Pavimento", 
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if ret == QMessageBox.Yes:
            # Excluir pavimento e todos os seus itens
            if curr_obra in self.fundos_salvos and curr_pav in self.fundos_salvos[curr_obra]:
                del self.fundos_salvos[curr_obra][curr_pav]
            
            # Remover do metadata
            if curr_obra in self.obras_metadata:
                self.obras_metadata[curr_obra].discard(curr_pav)
            
            # Salvar dados
            self.save_data()
            
            # Atualizar combos e lista
            self._update_combo_pavimentos()
            self.update_list()
            
            self.log(f"🗑️ Pavimento '{curr_pav}' da obra '{curr_obra}' e todos os seus dados foram excluídos.")
            QMessageBox.information(self, "Sucesso", f"Pavimento '{curr_pav}' e todos os seus dados foram excluídos.")

    def action_paste_pav(self):
        # Placeholder para edição/colar
        QMessageBox.information(self, "Paste", "Função colar pavimento vindo em breve.")
    
    def action_delete_selected(self):
        """Exclui os itens selecionados na lista (suporta múltiplos)"""
        selected_items = self.tree_fundos.selectedItems()
        
        if not selected_items:
            QMessageBox.warning(self, "Aviso", "Nenhum item selecionado.\n\nUse Shift ou Ctrl para selecionar múltiplos itens.")
            return
        
        # Confirmar exclusão
        count = len(selected_items)
        msg = f"⚠️ ATENÇÃO: Esta ação é IRREVERSÍVEL!\n\n"
        msg += f"Você está prestes a excluir {count} item(s) selecionado(s).\n\n"
        if count > 1:
            msg += f"Os seguintes itens serão excluídos:\n"
            for item in selected_items[:10]:  # Mostrar até 10
                num = item.data(0, Qt.UserRole)
                nome = item.text(1)
                msg += f"  • {num} - {nome}\n"
            if count > 10:
                msg += f"  ... e mais {count - 10} item(s)\n"
        else:
            num = selected_items[0].data(0, Qt.UserRole)
            nome = selected_items[0].text(1)
            msg += f"Item: {num} - {nome}\n"
        msg += f"\nTODOS os dados serão PERDIDOS PERMANENTEMENTE.\n\n"
        msg += f"Deseja realmente continuar?"
        
        ret = QMessageBox.warning(
            self, 
            "Confirmar Exclusão", 
            msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if ret == QMessageBox.Yes:
            # Coletar dados dos itens selecionados
            items_to_delete = []
            for item in selected_items:
                # Ignorar itens parent (clusters) - só processar filhos
                if item.parent() is None and item.childCount() > 0:
                    # É um parent, processar todos os filhos
                    for i in range(item.childCount()):
                        child = item.child(i)
                        num = child.data(0, Qt.UserRole)
                        obra = child.data(0, Qt.UserRole + 1)
                        pav = child.data(0, Qt.UserRole + 2)
                        if num and obra and pav:
                            items_to_delete.append((obra, pav, str(num)))
                else:
                    # É um item filho ou item sem parent
                    num = item.data(0, Qt.UserRole)
                    obra = item.data(0, Qt.UserRole + 1)
                    pav = item.data(0, Qt.UserRole + 2)
                    if num and obra and pav:
                        items_to_delete.append((obra, pav, str(num)))
                    elif not obra or not pav:
                        # Fallback: usar obra e pavimento atuais
                        obra = self.combo_obra.currentText()
                        pav = _cmb_pav_raw(self.combo_pavimento)
                        if num and obra and pav:
                            items_to_delete.append((obra, pav, str(num)))
            
            # Excluir itens
            deleted_count = 0
            for obra, pav, num in items_to_delete:
                if obra in self.fundos_salvos and pav in self.fundos_salvos[obra]:
                    if num in self.fundos_salvos[obra][pav]:
                        del self.fundos_salvos[obra][pav][num]
                        deleted_count += 1
            
            if deleted_count == 0 and len(items_to_delete) > 0:
                # Debug: logar o que foi tentado
                self.log(f"⚠️ Tentativa de exclusão: {len(items_to_delete)} itens coletados, mas nenhum foi encontrado na estrutura de dados.")
                self.log(f"   Estrutura atual: {list(self.fundos_salvos.keys())}")
                for obra, pav, num in items_to_delete[:3]:
                    self.log(f"   Tentado: obra='{obra}', pav='{pav}', num='{num}'")
            
            # Salvar dados
            self.save_data()
            
            # Atualizar lista
            self.update_list()
            
            self.log(f"🗑️ {deleted_count} item(s) excluído(s) com sucesso.")
            QMessageBox.information(self, "Sucesso", f"{deleted_count} item(s) excluído(s) com sucesso.")

    def _realizar_boundary(self):
        if pythoncom is None:
            QMessageBox.critical(self, "Erro", "Bibliotecas COM não disponíveis.")
            return

        try:
            pythoncom.CoInitialize()
            acad = win32com.client.Dispatch("AutoCAD.Application")
            doc = acad.ActiveDocument
            
            self.setWindowState(self.windowState() | Qt.WindowMinimized)
            
            # Usar o novo método de comentário que aguarda
            if not self._mostrar_comentario_flutuante("Clique dentro da área para criar o BOUNDARY.", tempo=None):
                 self.setWindowState(self.windowState() & ~Qt.WindowMinimized)
                 return
            
            # Tentar criar boundary
            max_tentativas = 2
            coords = []
            
            for tentativa in range(max_tentativas):
                try:
                    doc.SendCommand("._BOUNDARY\n")
                    time.sleep(1) 
                    
                    entidade = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
                    if entidade.ObjectName == "AcDbPolyline":
                        coords = list(entidade.Coordinates)
                        entidade.Delete() 
                        break
                except Exception as e:
                    print(f"Tentativa {tentativa+1} falhou: {e}")
                
                if tentativa == 0:
                    doc.SendCommand("._EXTEND\n\n\n") 
            
            self.setWindowState(self.windowState() & ~Qt.WindowMinimized)
            self.show()
            self.activateWindow()
            
            if not coords:
                QMessageBox.warning(self, "Aviso", "Não foi possível capturar o boundary.")
                return

            # Processar coordenadas
            coords_alinhadas = alinhar_boundary_horizontal(coords)
            analise = analisar_coordenadas(coords_alinhadas)
            
            if analise:
                self._preencher_campos(analise)
                # self.learner.add_example(coords_alinhadas, analise) # Learner not fully integrated yet
                QMessageBox.information(self, "Sucesso", "Fundo analisado com sucesso!")
            else:
                QMessageBox.warning(self, "Erro", "Falha na análise geométrica.")

        except Exception as e:
            QMessageBox.critical(self, "Erro no AutoCAD", f"Falha na automação: {e}")
        finally:
            pythoncom.CoUninitialize()

    def _realizar_boundary_duplo(self):
        if pythoncom is None: return
        
        try:
            pythoncom.CoInitialize()
            acad = win32com.client.Dispatch("AutoCAD.Application")
            doc = acad.ActiveDocument
            
            self.setWindowState(self.windowState() | Qt.WindowMinimized)
            
            # Passo 1
            if not self._mostrar_comentario_flutuante("SELECIONE O BOUNDARY PRINCIPAL (1/2)"):
                self.setWindowState(self.windowState() & ~Qt.WindowMinimized); return
            
            doc.SendCommand("._BOUNDARY\n")
            time.sleep(1)
            e1 = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
            coords1 = list(e1.Coordinates)
            e1.Delete()
            
            # Passo 2
            if not self._mostrar_comentario_flutuante("SELECIONE O BOUNDARY DO L (2/2)"):
                self.setWindowState(self.windowState() & ~Qt.WindowMinimized); return

            doc.SendCommand("._BOUNDARY\n")
            time.sleep(1)
            e2 = doc.ModelSpace.Item(doc.ModelSpace.Count - 1)
            coords2 = list(e2.Coordinates)
            e2.Delete()
            
            self.setWindowState(self.windowState() & ~Qt.WindowMinimized)
            self.show()
            self.activateWindow()

            # Processar
            c1_a = alinhar_boundary_horizontal(coords1)
            c2_a = alinhar_boundary_horizontal(coords2)
            a1 = analisar_coordenadas(c1_a)
            a2 = analisar_coordenadas(c2_a)
            
            if not a1 or not a2: raise Exception("Falha na análise de um dos retângulos")
            
            if a1['largura'] < a2['largura']:
                principal, l_panel = a2, a1
            else:
                principal, l_panel = a1, a2

            tipo_l = detectar_tipo_l(principal, l_panel)
            
            analise_dupla = {
                'principal': principal,
                'l_panel': l_panel,
                'tipo_l': tipo_l
            }
            
            self._preencher_campos_duplos(analise_dupla)
            QMessageBox.information(self, "Sucesso L", f"Extensão em L detectada como: {tipo_l}")

        except Exception as e:
            self.setWindowState(self.windowState() & ~Qt.WindowMinimized)
            QMessageBox.critical(self, "Erro Boundary L", str(e))
        finally:
            pythoncom.CoUninitialize()

    def _preencher_campos(self, analise):
        try:
            self.fields['largura'].setText(str(analise.get('largura', 0)))
            self.fields['altura'].setText(str(analise.get('altura', 0)))
            
            for field in self.l_fields.values(): field.setText("0")
            
            paineis = analise.get('paineis', [])
            for i, val in enumerate(paineis):
                if i < len(self.paineis_fields):
                    self.paineis_fields[i].setText(str(val))
                    
            recuos = analise.get('recuos', [0]*4)
            for i, val in enumerate(recuos):
                if i < len(self.chanfros_fields):
                    self.chanfros_fields[i].setText(str(val))
                    
            aberturas = analise.get('aberturas', [[0,0,0]]*4)
            for i, ab in enumerate(aberturas):
                if i < len(self.aberturas_fields):
                    for j, val in enumerate(ab):
                        if j < 3:
                            self.aberturas_fields[i][j].setText(str(val))
            
            self.update_calculations()
        except Exception as e:
            print(f"Erro ao preencher campos: {e}")

    def _preencher_campos_duplos(self, analise_dupla):
        p = analise_dupla['principal']
        l = analise_dupla['l_panel']
        tipo = analise_dupla['tipo_l']
        
        # Principal
        self._preencher_campos(p)
        
        # Painel 2 em L
        self.l_fields['comp2'].setText(str(l.get('largura', 0)))
        self.l_fields['larg2'].setText(str(l.get('altura', 0)))
        
        # Painéis do L (P1-P3)
        paineis_l = l.get('paineis', [])
        for i, val in enumerate(paineis_l):
            key = f"p{i+1}_2"
            if key in self.l_fields:
                self.l_fields[key].setText(str(val))
                
        # Tipo L (Radios)
        for rad in self.l_type_rads:
            if rad.text() == tipo:
                rad.setChecked(True)
                break
                
        # Chanfros do L
        recuos_l = l.get('recuos', [0]*4)
        for i, val in enumerate(recuos_l):
            if i < len(self.l_chanfros_fields):
                self.l_chanfros_fields[i].setText(str(val))
                
        # Aberturas do L
        aberturas_l = l.get('aberturas', [[0,0,0]]*4)
        for i, ab in enumerate(aberturas_l):
            if i < len(self.l_aberturas_fields):
                for j, val in enumerate(ab):
                    if j < 3:
                        self.l_aberturas_fields[i][j].setText(str(val))
        
        self.update_calculations()

    def _mostrar_comentario_flutuante(self, msg):
        """Exibe o comentário flutuante na tela."""
        if self._balcao_flutuante:
            self._balcao_flutuante.accept()
        
        self._balcao_flutuante = FloatingComment(msg, self)
        self._balcao_flutuante.show()
        # Não bloqueante para permitir interação com CAD
        QApplication.processEvents()

    def update_calculations(self):
        # Triggered when width/height changes
        largura = float_safe(self.fields['largura'].text())
        if largura > 0:
            # Mock de cálculo automático de painéis
            paineis = [largura/2, largura/2, 0, 0, 0, 0] # Simplificado
            for i, val in enumerate(paineis):
                if i < len(self.paineis_fields):
                    self.paineis_fields[i].setText(str(round(val, 2)))
        
        self.update_canvas()

    def update_canvas(self):
        dados = self.get_current_data()
        self.canvas.draw_fundo(dados)

    def get_current_data(self):
        print("[DEBUG] Iniciando get_current_data")
        # L-Type string
        l_type = "E/T"
        for i, rad in enumerate(self.l_type_rads):
            if rad.isChecked():
                l_type = ["E/T", "E/F", "D/T", "D/F"][i]
                break
        print(f"[DEBUG] Tipo L determinado: {l_type}")

        data = {
            'numero': self.fields['numero'].text(),
            'nome': self.fields['nome'].text(),
            'obs': self.fields['obs'].text(),
            'pavimento': self.fields['pavimento'].text(),
            'largura': self.fields['largura'].text(),
            'altura': self.fields['altura'].text(),
            'texto_esq': self.fields['texto_esq'].text(),
            'texto_dir': self.fields['texto_dir'].text(),
            'paineis': [f.text() for f in self.paineis_fields],
            'recuos': [f.text() for f in self.chanfros_fields],
            'aberturas': [[f.text() for f in row] for row in self.aberturas_fields],
            'comprimento_2': self.l_fields['comp2'].text(),
            'largura_2': self.l_fields['larg2'].text(),
            'paineis_2': [self.l_fields[f'p{i}_2'].text() for i in range(1, 4)], # TODO: add p4-p6 to UI if needed
            'tipo_painel2': l_type,
            'recuos_2': [f.text() for f in self.l_chanfros_fields],
            'aberturas_2': [[f.text() for f in row] for row in self.l_aberturas_fields],
            'sarrafo_esq': self.chk_sarrafo_esq.isChecked(),
            'sarrafo_dir': self.chk_sarrafo_dir.isChecked()
        }
        
        # IMPORTANTE: Preservar boundary_original se existir (vem do load_selected)
        # Isso garante que a geometria real seja mantida
        if hasattr(self, '_current_boundary_original') and self._current_boundary_original is not None:
            data['boundary_original'] = self._current_boundary_original.copy()
            print(f"[DEBUG] Preservando boundary_original com {len(self._current_boundary_original)} coordenadas")

        print(f"[DEBUG] Dados coletados: numero='{data['numero']}', nome='{data['nome']}', pavimento='{data['pavimento']}', largura='{data['largura']}', altura='{data['altura']}'")
        print(f"[DEBUG] Paineis: {data['paineis']}")
        return data

    def _mostrar_comentario_flutuante(self, msg, tempo=None):
        if self._balcao_flutuante:
            self._balcao_flutuante.close()
        
        self._balcao_flutuante = FloatingComment(msg, self, timeout=tempo)
        return self._balcao_flutuante.exec() == QDialog.Accepted

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FundoMainWindow()
    window.show()
    sys.exit(app.exec())
